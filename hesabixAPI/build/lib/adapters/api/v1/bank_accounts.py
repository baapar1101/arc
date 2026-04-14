from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, Request, Body, HTTPException
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.permissions import require_business_management_dep, require_business_access
from adapters.api.v1.schemas import QueryInfo
from adapters.api.v1.schema_models.bank_account import (
    BankAccountCreateRequest,
    BankAccountUpdateRequest,
)
from app.services.bank_account_service import (
    create_bank_account,
    update_bank_account,
    delete_bank_account,
    get_bank_account_by_id,
    list_bank_accounts,
    bulk_delete_bank_accounts,
)

router = APIRouter(prefix="/bank-accounts", tags=["bank-accounts"])


@router.post(
    "/businesses/{business_id}/bank-accounts",
    summary="لیست حساب‌های بانکی کسب‌وکار",
    description="دریافت لیست حساب‌های بانکی یک کسب و کار با امکان جستجو و فیلتر",
)
@require_business_access("business_id")
async def list_bank_accounts_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    query_dict: Dict[str, Any] = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "search": query_info.search,
        "search_fields": query_info.search_fields,
        "filters": query_info.filters,
    }
    result = list_bank_accounts(db, business_id, query_dict)
    result["items"] = [format_datetime_fields(item, request) for item in result.get("items", [])]
    return success_response(data=result, request=request, message="BANK_ACCOUNTS_LIST_FETCHED")


@router.post(
    "/businesses/{business_id}/bank-accounts/create",
    summary="ایجاد حساب بانکی جدید",
    description="ایجاد حساب بانکی برای یک کسب‌وکار مشخص",
)
@require_business_access("business_id")
async def create_bank_account_endpoint(
    request: Request,
    business_id: int,
    body: BankAccountCreateRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_management_dep),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    created = create_bank_account(db, business_id, payload)
    return success_response(data=format_datetime_fields(created, request), request=request, message="BANK_ACCOUNT_CREATED")


@router.get(
    "/bank-accounts/{account_id}",
    summary="جزئیات حساب بانکی",
    description="دریافت جزئیات حساب بانکی بر اساس شناسه",
)
async def get_bank_account_endpoint(
    request: Request,
    account_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_management_dep),
):
    result = get_bank_account_by_id(db, account_id)
    if not result:
        raise ApiError("BANK_ACCOUNT_NOT_FOUND", "Bank account not found", http_status=404)
    # بررسی دسترسی به کسبوکار مرتبط
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="BANK_ACCOUNT_DETAILS")


@router.put(
    "/bank-accounts/{account_id}",
    summary="ویرایش حساب بانکی",
    description="ویرایش اطلاعات حساب بانکی",
)
async def update_bank_account_endpoint(
    request: Request,
    account_id: int,
    body: BankAccountUpdateRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_management_dep),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    result = update_bank_account(db, account_id, payload)
    if result is None:
        raise ApiError("BANK_ACCOUNT_NOT_FOUND", "Bank account not found", http_status=404)
    # بررسی دسترسی به کسبوکار مرتبط
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="BANK_ACCOUNT_UPDATED")


@router.delete(
    "/bank-accounts/{account_id}",
    summary="حذف حساب بانکی",
    description="حذف یک حساب بانکی",
)
async def delete_bank_account_endpoint(
    request: Request,
    account_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_management_dep),
):
    # ابتدا بررسی دسترسی بر اساس business مربوط به حساب
    result = get_bank_account_by_id(db, account_id)
    if result:
        try:
            biz_id = int(result.get("business_id"))
        except Exception:
            biz_id = None
        if biz_id is not None and not ctx.can_access_business(biz_id):
            raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    ok = delete_bank_account(db, account_id)
    if not ok:
        raise ApiError("BANK_ACCOUNT_NOT_FOUND", "Bank account not found", http_status=404)
    return success_response(data=None, request=request, message="BANK_ACCOUNT_DELETED")


@router.post(
    "/businesses/{business_id}/bank-accounts/bulk-delete",
    summary="حذف گروهی حساب‌های بانکی",
    description="حذف چندین حساب بانکی بر اساس شناسه‌ها",
)
@require_business_access("business_id")
async def bulk_delete_bank_accounts_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_management_dep),
):
    ids = body.get("ids")
    if not isinstance(ids, list):
        ids = []
    try:
        ids = [int(x) for x in ids if isinstance(x, (int, str)) and str(x).isdigit()]
    except Exception:
        ids = []
    
    if not ids:
        return success_response({"deleted": 0, "skipped": 0, "total_requested": 0}, request, message="NO_VALID_IDS_FOR_DELETE")
    
    # فراخوانی تابع حذف گروهی از سرویس
    result = bulk_delete_bank_accounts(db, business_id, ids)
    
    return success_response(result, request, message="BANK_ACCOUNTS_BULK_DELETE_DONE")

@router.post(
    "/businesses/{business_id}/bank-accounts/export/excel",
    summary="خروجی Excel لیست حساب‌های بانکی",
    description="خروجی Excel لیست حساب‌های بانکی با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_bank_accounts_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from app.core.i18n import negotiate_locale

    # دریافت داده‌ها از سرویس
    query_dict = {
        "take": int(body.get("take", 1000)),  # برای export همه داده‌ها
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }
    
    result = list_bank_accounts(db, business_id, query_dict)
    items: List[Dict[str, Any]] = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]
    
    headers: List[str] = [
        "code", "name", "branch", "account_number", "sheba_number", "card_number", "owner_name", "pos_number", "is_active", "is_default"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "BankAccounts"

    # RTL/LTR
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(headers, 1):
            value = item.get(key, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")

    # Auto-width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    content = buffer.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=bank_accounts.xlsx",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/bank-accounts/export/pdf",
    summary="خروجی PDF لیست حساب‌های بانکی",
    description="خروجی PDF لیست حساب‌های بانکی با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_bank_accounts_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale

    # Build query dict similar to persons export
    query_dict = {
        "take": int(body.get("take", 1000)),  # برای export همه داده‌ها
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }

    # دریافت داده‌ها از سرویس
    result = list_bank_accounts(db, business_id, query_dict)
    items: List[Dict[str, Any]] = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Selection handling
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            import json
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare headers/keys from export_columns (order + visibility)
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        if items:
            keys = list(items[0].keys())
            headers = keys
        else:
            keys = [
                "code", "name", "branch", "account_number", "sheba_number",
                "card_number", "owner_name", "pos_number", "is_active", "is_default",
            ]
            headers = keys

    # Load business info
    business_name = ""
    try:
        from adapters.db.models.business import Business
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name or ""
    except Exception:
        business_name = ""

    # Locale and calendar-aware date
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'

    try:
        from app.core.calendar import CalendarConverter
        calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
        formatted_now = CalendarConverter.format_datetime(
            __import__("datetime").datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian",
        )
        now_str = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        from datetime import datetime
        now_str = datetime.now().strftime('%Y/%m/%d %H:%M')

    # Labels
    title_text = "گزارش لیست حساب‌های بانکی" if is_fa else "Bank Accounts List Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    footer_text = "تولید شده توسط Hesabix" if is_fa else "Generated by Hesabix"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "

    def escape_val(v: Any) -> str:
        try:
            return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        except Exception:
            return str(v)

    rows_html: List[str] = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key)
            if value is None:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            tds.append(f"<td>{escape_val(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = ''.join(f"<th>{escape_val(h)}</th>" for h in headers)

    table_html = f"""
    <html lang=\"{html_lang}\" dir=\"{html_dir}\"> 
      <head>
        <meta charset='utf-8'>
        <style>
          @page {{
            size: A4 landscape;
            margin: 12mm;
            @bottom-{ 'left' if is_fa else 'right' } {{
              content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
              font-size: 10px;
              color: #666;
            }}
          }}
          body {{
            font-family: sans-serif;
            font-size: 11px;
            color: #222;
          }}
          .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #444;
            padding-bottom: 6px;
          }}
          .title {{
            font-size: 16px;
            font-weight: 700;
          }}
          .meta {{
            font-size: 11px;
            color: #555;
          }}
          .table-wrapper {{
            width: 100%;
          }}
          table.report-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
          }}
          thead th {{
            background: #f0f3f7;
            border: 1px solid #c7cdd6;
            padding: 6px 4px;
            text-align: center;
            font-weight: 700;
            white-space: nowrap;
          }}
          tbody td {{
            border: 1px solid #d7dde6;
            padding: 5px 4px;
            vertical-align: top;
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: normal;
          }}
          .footer {{
            position: running(footer);
            font-size: 10px;
            color: #666;
            margin-top: 8px;
            text-align: {'left' if is_fa else 'right'};
          }}
        </style>
      </head>
      <body>
        <div class=\"header\"> 
          <div>
            <div class=\"title\">{title_text}</div>
            <div class=\"meta\">{label_biz}: {escape_val(business_name)}</div>
          </div>
          <div class=\"meta\">{label_date}: {escape_val(now_str)}</div>
        </div>
        <div class=\"table-wrapper\"> 
          <table class=\"report-table\"> 
            <thead>
              <tr>{headers_html}</tr>
            </thead>
            <tbody>
              {''.join(rows_html)}
            </tbody>
          </table>
        </div>
        <div class=\"footer\">{footer_text}</div>
      </body>
    </html>
    """

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=table_html).write_pdf(font_config=font_config)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=bank_accounts.pdf",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


