# Removed __future__ annotations to fix OpenAPI schema generation

from typing import Dict, Any
from fastapi import APIRouter, Depends, Request
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.permissions import require_business_access
from app.core.responses import success_response, ApiError, format_datetime_fields
from adapters.api.v1.schemas import QueryInfo
from adapters.api.v1.schema_models.product import (
    ProductCreateRequest,
    ProductUpdateRequest,
    BulkPriceUpdateRequest,
    BulkPriceUpdatePreviewResponse,
)
from app.services.product_service import (
    create_product,
    list_products,
    get_product,
    update_product,
    delete_product,
)
from app.services.bulk_price_update_service import (
    preview_bulk_price_update,
    apply_bulk_price_update,
)
from adapters.db.models.business import Business
from app.core.i18n import negotiate_locale
from fastapi import UploadFile, File, Form


router = APIRouter(prefix="/products", tags=["products"])


@router.post("/business/{business_id}")
@require_business_access("business_id")
def create_product_endpoint(
    request: Request,
    business_id: int,
    payload: ProductCreateRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.has_business_permission("inventory", "write"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.write", http_status=403)
    result = create_product(db, business_id, payload)
    return success_response(data=format_datetime_fields(result["data"], request), request=request, message=result.get("message"))


@router.post("/business/{business_id}/search")
@require_business_access("business_id")
def search_products_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.can_read_section("inventory"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.read", http_status=403)
    result = list_products(db, business_id, {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "search": query_info.search,
        "filters": query_info.filters,
    })
    return success_response(data=format_datetime_fields(result, request), request=request)


@router.get("/business/{business_id}/{product_id}")
@require_business_access("business_id")
def get_product_endpoint(
    request: Request,
    business_id: int,
    product_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.can_read_section("inventory"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.read", http_status=403)
    item = get_product(db, product_id, business_id)
    if not item:
        raise ApiError("NOT_FOUND", "Product not found", http_status=404)
    return success_response(data=format_datetime_fields({"item": item}, request), request=request)


@router.put("/business/{business_id}/{product_id}")
@require_business_access("business_id")
def update_product_endpoint(
    request: Request,
    business_id: int,
    product_id: int,
    payload: ProductUpdateRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.has_business_permission("inventory", "write"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.write", http_status=403)
    result = update_product(db, product_id, business_id, payload)
    if not result:
        raise ApiError("NOT_FOUND", "Product not found", http_status=404)
    return success_response(data=format_datetime_fields(result["data"], request), request=request, message=result.get("message"))


@router.delete("/business/{business_id}/{product_id}")
@require_business_access("business_id")
def delete_product_endpoint(
    request: Request,
    business_id: int,
    product_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.has_business_permission("inventory", "delete"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.delete", http_status=403)
    ok = delete_product(db, product_id, business_id)
    return success_response({"deleted": ok}, request)


@router.post("/business/{business_id}/bulk-delete",
    summary="حذف گروهی محصولات",
    description="حذف چندین آیتم بر اساس شناسه‌ها یا کدها",
)
@require_business_access("business_id")
def bulk_delete_products_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any],
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.has_business_permission("inventory", "delete"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.delete", http_status=403)

    from sqlalchemy import and_ as _and
    from adapters.db.models.product import Product

    ids = body.get("ids")
    codes = body.get("codes")
    deleted = 0
    skipped = 0

    if not ids and not codes:
        return success_response({"deleted": 0, "skipped": 0}, request)

    # Normalize inputs
    if isinstance(ids, list):
        ids = [int(x) for x in ids if isinstance(x, (int, str)) and str(x).isdigit()]
    else:
        ids = []
    if isinstance(codes, list):
        codes = [str(x).strip() for x in codes if str(x).strip()]
    else:
        codes = []

    # Delete by IDs first
    if ids:
        for pid in ids:
            ok = delete_product(db, pid, business_id)
            if ok:
                deleted += 1
            else:
                skipped += 1

    # Delete by codes
    if codes:
        items = db.query(Product).filter(_and(Product.business_id == business_id, Product.code.in_(codes))).all()
        for obj in items:
            try:
                db.delete(obj)
                deleted += 1
            except Exception:
                skipped += 1
        db.commit()

    return success_response({"deleted": deleted, "skipped": skipped}, request)

@router.post("/business/{business_id}/export/excel",
    summary="خروجی Excel لیست محصولات",
    description="خروجی Excel لیست محصولات با قابلیت فیلتر، انتخاب ستون‌ها و ترتیب آن‌ها",
)
@require_business_access("business_id")
async def export_products_excel(
    request: Request,
    business_id: int,
    body: dict,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    import re
    import datetime
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if not ctx.can_read_section("inventory"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.read", http_status=403)

    query_dict = {
        "take": int(body.get("take", 1000)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }
    result = list_products(db, business_id, query_dict)
    items = result.get("items", []) if isinstance(result, dict) else result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Apply selected indices filter if requested
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None and isinstance(items, list):
        indices = None
        if isinstance(selected_indices, str):
            try:
                import json as _json
                indices = _json.loads(selected_indices)
            except Exception:
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        headers = [col.get("label") or col.get("key") for col in export_columns]
        keys = [col.get("key") for col in export_columns]
    else:
        default_cols = [
            ("code", "کد"),
            ("name", "نام"),
            ("item_type", "نوع"),
            ("category_id", "دسته"),
            ("base_sales_price", "قیمت فروش"),
            ("base_purchase_price", "قیمت خرید"),
            ("main_unit_id", "واحد اصلی"),
            ("secondary_unit_id", "واحد فرعی"),
            ("track_inventory", "کنترل موجودی"),
            ("created_at_formatted", "ایجاد"),
        ]
        keys = [k for k, _ in default_cols]
        headers = [v for _, v in default_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Locale and RTL/LTR handling for Excel
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for it in items:
        row = []
        for k in keys:
            row.append(it.get(k))
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            # Align data cells based on locale
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")

    # Auto width columns
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass

    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()

    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    base = "products"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(data)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/business/{business_id}/import/template",
    summary="دانلود تمپلیت ایمپورت محصولات",
    description="فایل Excel تمپلیت برای ایمپورت کالا/خدمت را برمی‌گرداند",
)
@require_business_access("business_id")
async def download_products_import_template(
    request: Request,
    business_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    import datetime
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    if not ctx.has_business_permission("inventory", "write"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.write", http_status=403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    # Localized, user-friendly headers. Import endpoint will map these back to internal keys.
    # For reference fields, provide both ID columns and human-friendly columns.
    columns = [
        ("code", {"fa": "کد", "en": "Code"}),
        ("name", {"fa": "نام", "en": "Name"}),
        ("item_type", {"fa": "نوع", "en": "Type"}),
        ("description", {"fa": "توضیحات", "en": "Description"}),
        ("category_id", {"fa": "شناسه دسته‌بندی", "en": "Category ID"}),
        ("category_path", {"fa": "مسیر دسته‌بندی", "en": "Category Path"}),
        ("main_unit_id", {"fa": "شناسه واحد اصلی", "en": "Main Unit ID"}),
        ("secondary_unit_id", {"fa": "شناسه واحد فرعی", "en": "Secondary Unit ID"}),
        ("unit_conversion_factor", {"fa": "ضریب تبدیل", "en": "Unit Conversion Factor"}),
        ("base_sales_price", {"fa": "قیمت فروش", "en": "Sales Price"}),
        ("base_purchase_price", {"fa": "قیمت خرید", "en": "Purchase Price"}),
        ("track_inventory", {"fa": "کنترل موجودی", "en": "Track Inventory"}),
        ("reorder_point", {"fa": "نقطه سفارش مجدد", "en": "Reorder Point"}),
        ("min_order_qty", {"fa": "حداقل مقدار سفارش", "en": "Min Order Qty"}),
        ("lead_time_days", {"fa": "زمان تامین (روز)", "en": "Lead Time (Days)"}),
        ("is_sales_taxable", {"fa": "مشمول مالیات فروش", "en": "Sales Taxable"}),
        ("is_purchase_taxable", {"fa": "مشمول مالیات خرید", "en": "Purchase Taxable"}),
        ("sales_tax_rate", {"fa": "نرخ مالیات فروش (%)", "en": "Sales Tax Rate (%)"}),
        ("purchase_tax_rate", {"fa": "نرخ مالیات خرید (%)", "en": "Purchase Tax Rate (%)"}),
        ("tax_type_id", {"fa": "شناسه نوع مالیات", "en": "Tax Type ID"}),
        ("tax_type_code", {"fa": "کد نوع مالیات", "en": "Tax Type Code"}),
        ("tax_type_title", {"fa": "عنوان نوع مالیات", "en": "Tax Type Title"}),
        ("tax_code", {"fa": "کد مالیاتی", "en": "Tax Code"}),
        ("tax_unit_id", {"fa": "شناسه واحد مالیاتی", "en": "Tax Unit ID"}),
        ("tax_unit_code", {"fa": "کد واحد مالیاتی", "en": "Tax Unit Code"}),
        ("tax_unit_name", {"fa": "نام واحد مالیاتی", "en": "Tax Unit Name"}),
        ("attribute_ids", {"fa": "شناسه ویژگی‌ها", "en": "Attribute IDs"}),
        ("attribute_titles", {"fa": "نام ویژگی‌ها", "en": "Attribute Titles"}),
    ]

    headers = [labels.get(locale, labels.get("en", key)) for key, labels in columns]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    if locale == 'fa':
        sample = [
            "P1001","نمونه کالا","کالا","توضیح اختیاری", "",
            "مواد اولیه > پلاستیک", "", "", "",
            "150000", "120000", "TRUE",
            "0", "0", "",
            "FALSE", "FALSE", "", "",
            "", "", "", "", "", "", "",
            "1,2,3", "رنگ, سایز",
        ]
    else:
        sample = [
            "P1001","Sample product","product","Optional description", "",
            "Raw materials > Plastics", "", "", "",
            "150000", "120000", "TRUE",
            "0", "0", "",
            "FALSE", "FALSE", "", "",
            "", "", "", "", "", "", "",
            "1,2,3", "Color, Size",
        ]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    # Auto width
    for column in ws.columns:
        try:
            letter = column[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"products_import_template_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/business/{business_id}/import/excel",
    summary="ایمپورت محصولات از فایل Excel",
    description="فایل اکسل را دریافت می‌کند و به‌صورت dry-run یا واقعی پردازش می‌کند",
)
@require_business_access("business_id")
async def import_products_excel(
    request: Request,
    business_id: int,
    file: UploadFile = File(...),
    dry_run: str = Form(default="true"),
    match_by: str = Form(default="code"),
    conflict_policy: str = Form(default="upsert"),
    on_missing_category: str = Form(default="error"),
    on_missing_attributes: str = Form(default="error"),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    import json
    import logging
    import re
    import zipfile
    from decimal import Decimal
    from typing import Optional
    from openpyxl import load_workbook
    from sqlalchemy import and_ as _and
    from adapters.db.models.category import BusinessCategory
    from adapters.db.models.product_attribute import ProductAttribute
    from adapters.db.models.tax_type import TaxType
    from adapters.db.models.tax_unit import TaxUnit

    if not ctx.has_business_permission("inventory", "write"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.write", http_status=403)

    logger = logging.getLogger(__name__)

    def _validate_excel_signature(content: bytes) -> bool:
        try:
            if not content.startswith(b'PK'):
                return False
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                return any(n.startswith('xl/') for n in zf.namelist())
        except Exception:
            return False

    try:
        is_dry_run = str(dry_run).lower() in ("true","1","yes","on")
        on_missing_category = str(on_missing_category or "error").strip().lower()
        on_missing_attributes = str(on_missing_attributes or "error").strip().lower()
        if on_missing_category not in ("error", "create"):
            on_missing_category = "error"
        if on_missing_attributes not in ("error", "create"):
            on_missing_attributes = "error"

        preview_rows: list[dict] = []
        reference_summary: dict[str, Any] = {
            "resolved": {"category": 0, "tax_type": 0, "tax_unit": 0, "attributes": 0},
            "would_create": {"categories": 0, "attributes": 0},
            "created": {"categories": 0, "attributes": 0},
            "policies": {"on_missing_category": on_missing_category, "on_missing_attributes": on_missing_attributes},
        }

        if not file.filename or not file.filename.lower().endswith('.xlsx'):
            raise ApiError("INVALID_FILE", "فرمت فایل معتبر نیست. تنها xlsx پشتیبانی می‌شود", http_status=400)

        content = await file.read()
        if len(content) < 100 or not _validate_excel_signature(content):
            raise ApiError("INVALID_FILE", "فایل Excel معتبر نیست یا خالی است", http_status=400)

        try:
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        except zipfile.BadZipFile:
            raise ApiError("INVALID_FILE", "فایل Excel خراب است یا فرمت آن معتبر نیست", http_status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return success_response(data={"summary": {"total": 0}}, request=request, message="EMPTY_FILE")

        # Headers may be localized (fa/en). Normalize them to internal keys.
        raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        def _normalize_header(v: object) -> str:
            s = "" if v is None else str(v)
            s = s.replace("\u200c", " ")
            s = re.sub(r"\s+", " ", s).strip()
            return s.lower()

        header_aliases: dict[str, str] = {}
        internal_keys = [
            "code","name","item_type","description","category_id",
            "category_path","category",
            "main_unit_id","secondary_unit_id","unit_conversion_factor",
            "base_sales_price","base_purchase_price","track_inventory",
            "reorder_point","min_order_qty","lead_time_days",
            "is_sales_taxable","is_purchase_taxable","sales_tax_rate","purchase_tax_rate",
            "tax_type_id","tax_type_code","tax_type_title","tax_code",
            "tax_unit_id","tax_unit_code","tax_unit_name",
            "attribute_ids","attribute_titles",
        ]
        for k in internal_keys:
            header_aliases[_normalize_header(k)] = k

        header_aliases.update({
            _normalize_header("کد"): "code",
            _normalize_header("نام"): "name",
            _normalize_header("نوع"): "item_type",
            _normalize_header("توضیحات"): "description",
            _normalize_header("شناسه دسته‌بندی"): "category_id",
            _normalize_header("شناسه دسته بندی"): "category_id",
            _normalize_header("مسیر دسته‌بندی"): "category_path",
            _normalize_header("مسیر دسته بندی"): "category_path",
            _normalize_header("دسته‌بندی"): "category",
            _normalize_header("دسته بندی"): "category",
            _normalize_header("شناسه واحد اصلی"): "main_unit_id",
            _normalize_header("شناسه واحد فرعی"): "secondary_unit_id",
            _normalize_header("ضریب تبدیل"): "unit_conversion_factor",
            _normalize_header("قیمت فروش"): "base_sales_price",
            _normalize_header("قیمت خرید"): "base_purchase_price",
            _normalize_header("کنترل موجودی"): "track_inventory",
            _normalize_header("نقطه سفارش مجدد"): "reorder_point",
            _normalize_header("حداقل مقدار سفارش"): "min_order_qty",
            _normalize_header("زمان تامین (روز)"): "lead_time_days",
            _normalize_header("زمان تأمین (روز)"): "lead_time_days",
            _normalize_header("مشمول مالیات فروش"): "is_sales_taxable",
            _normalize_header("مشمول مالیات خرید"): "is_purchase_taxable",
            _normalize_header("نرخ مالیات فروش (%)"): "sales_tax_rate",
            _normalize_header("نرخ مالیات خرید (%)"): "purchase_tax_rate",
            _normalize_header("شناسه نوع مالیات"): "tax_type_id",
            _normalize_header("کد نوع مالیات"): "tax_type_code",
            _normalize_header("عنوان نوع مالیات"): "tax_type_title",
            _normalize_header("کد مالیاتی"): "tax_code",
            _normalize_header("شناسه واحد مالیاتی"): "tax_unit_id",
            _normalize_header("کد واحد مالیاتی"): "tax_unit_code",
            _normalize_header("نام واحد مالیاتی"): "tax_unit_name",
            _normalize_header("شناسه ویژگی‌ها"): "attribute_ids",
            _normalize_header("شناسه ویژگی ها"): "attribute_ids",
            _normalize_header("نام ویژگی‌ها"): "attribute_titles",
            _normalize_header("نام ویژگی ها"): "attribute_titles",
        })

        header_aliases.update({
            _normalize_header("code"): "code",
            _normalize_header("name"): "name",
            _normalize_header("type"): "item_type",
            _normalize_header("description"): "description",
            _normalize_header("category id"): "category_id",
            _normalize_header("category path"): "category_path",
            _normalize_header("category"): "category",
            _normalize_header("main unit id"): "main_unit_id",
            _normalize_header("secondary unit id"): "secondary_unit_id",
            _normalize_header("unit conversion factor"): "unit_conversion_factor",
            _normalize_header("sales price"): "base_sales_price",
            _normalize_header("purchase price"): "base_purchase_price",
            _normalize_header("track inventory"): "track_inventory",
            _normalize_header("reorder point"): "reorder_point",
            _normalize_header("min order qty"): "min_order_qty",
            _normalize_header("lead time (days)"): "lead_time_days",
            _normalize_header("sales taxable"): "is_sales_taxable",
            _normalize_header("purchase taxable"): "is_purchase_taxable",
            _normalize_header("sales tax rate (%)"): "sales_tax_rate",
            _normalize_header("purchase tax rate (%)"): "purchase_tax_rate",
            _normalize_header("tax type id"): "tax_type_id",
            _normalize_header("tax type code"): "tax_type_code",
            _normalize_header("tax type title"): "tax_type_title",
            _normalize_header("tax code"): "tax_code",
            _normalize_header("tax unit id"): "tax_unit_id",
            _normalize_header("tax unit code"): "tax_unit_code",
            _normalize_header("tax unit name"): "tax_unit_name",
            _normalize_header("attribute ids"): "attribute_ids",
            _normalize_header("attribute titles"): "attribute_titles",
        })

        headers = [header_aliases.get(_normalize_header(h), h) for h in raw_headers]
        data_rows = rows[1:]

        def _parse_bool(v: object) -> Optional[bool]:
            if v is None: return None
            s = str(v).strip().lower()
            if s in ("true","1","yes","on","بله","هست"):
                return True
            if s in ("false","0","no","off","خیر","نیست"):
                return False
            return None

        def _normalize_number_text(v: object) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if s == "":
                return ""
            if s.startswith("(") and s.endswith(")"):
                s = "-" + s[1:-1]
            digit_map = str.maketrans({
                "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
                "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
                "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
                "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9",
            })
            s = s.translate(digit_map)
            s = s.replace("\u066b", ".")
            for ch in [",", "\u066c", "\u060c", " ", "\u00a0", "\u202f", "\u2009", "_"]:
                s = s.replace(ch, "")
            return s.strip()

        def _parse_decimal(v: object) -> Optional[Decimal]:
            s = _normalize_number_text(v)
            if s == "":
                return None
            try:
                return Decimal(s)
            except Exception:
                return None

        def _parse_int(v: object) -> Optional[int]:
            s = _normalize_number_text(v)
            if s == "":
                return None
            try:
                return int(s.split(".")[0])
            except Exception:
                return None

        def _normalize_item_type(v: object) -> Optional[str]:
            if v is None: return None
            s = str(v).strip()
            mapping = {"product": "کالا", "service": "خدمت"}
            low = s.lower()
            if low in mapping: return mapping[low]
            if s in ("کالا","خدمت"): return s
            return None

        def _norm_text(v: object) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            s = s.replace("\u200c", " ")
            s = re.sub(r"\s+", " ", s).strip()
            return s

        def _split_list_text(v: object) -> list[str]:
            s = _norm_text(v)
            if not s:
                return []
            parts = re.split(r"[,\u060c;\|]+", s)
            out: list[str] = []
            for p in parts:
                t = _norm_text(p)
                if t:
                    out.append(t)
            return out

        categories_rows: list[BusinessCategory] | None = None
        attrs_rows: list[ProductAttribute] | None = None
        tax_type_by_code: dict[str, int] | None = None
        tax_type_by_title: dict[str, int] | None = None
        tax_unit_by_code: dict[str, int] | None = None
        tax_unit_by_name: dict[str, int] | None = None

        def _ensure_categories_loaded() -> list[BusinessCategory]:
            nonlocal categories_rows
            if categories_rows is None:
                categories_rows = db.query(BusinessCategory).filter(BusinessCategory.business_id == business_id).all()
            return categories_rows

        def _ensure_attributes_loaded() -> list[ProductAttribute]:
            nonlocal attrs_rows
            if attrs_rows is None:
                attrs_rows = db.query(ProductAttribute).filter(ProductAttribute.business_id == business_id).all()
            return attrs_rows

        def _ensure_tax_types_loaded() -> None:
            nonlocal tax_type_by_code, tax_type_by_title
            if tax_type_by_code is not None and tax_type_by_title is not None:
                return
            rows = db.query(TaxType).all()
            tax_type_by_code = {(r.code or "").strip().lower(): r.id for r in rows if r.code}
            tax_type_by_title = {(r.title or "").strip().lower(): r.id for r in rows if r.title}

        def _ensure_tax_units_loaded() -> None:
            nonlocal tax_unit_by_code, tax_unit_by_name
            if tax_unit_by_code is not None and tax_unit_by_name is not None:
                return
            rows = db.query(TaxUnit).all()
            tax_unit_by_code = {(r.code or "").strip().lower(): r.id for r in rows if r.code}
            tax_unit_by_name = {(r.name or "").strip().lower(): r.id for r in rows if r.name}

        def _get_category_titles(cat: BusinessCategory) -> list[str]:
            trans = cat.title_translations or {}
            vals: list[str] = []
            for k in ("fa", "en"):
                t = (trans.get(k) or "").strip()
                if t:
                    vals.append(t)
            for t in trans.values():
                tt = (t or "").strip()
                if tt and tt not in vals:
                    vals.append(tt)
            return vals

        def _resolve_category_by_id(category_id: Optional[int]) -> tuple[Optional[int], Optional[str]]:
            if category_id is None:
                return None, None
            exists = db.query(BusinessCategory.id).filter(_and(BusinessCategory.business_id == business_id, BusinessCategory.id == category_id)).first()
            if not exists:
                return None, f"دسته‌بندی با شناسه {category_id} یافت نشد"
            return category_id, None

        def _resolve_category_by_name_or_path(category_value: object, category_path: object) -> tuple[Optional[int], Optional[str]]:
            path_str = _norm_text(category_path)
            name_str = _norm_text(category_value)
            if not path_str and not name_str:
                return None, None
            path = path_str or name_str
            segments = [s.strip() for s in re.split(r"[>/\u203a\u00bb]+", path) if s and str(s).strip()]
            segments = [_norm_text(s) for s in segments if _norm_text(s)]
            if not segments:
                return None, None

            cats = _ensure_categories_loaded()
            by_parent: dict[int | None, list[BusinessCategory]] = {}
            for c in cats:
                by_parent.setdefault(c.parent_id, []).append(c)

            parent_id: int | None = None
            current_id: int | None = None
            for seg in segments:
                seg_norm = seg.strip().lower()
                candidates: list[BusinessCategory] = []
                for c in by_parent.get(parent_id, []):
                    titles = _get_category_titles(c)
                    if any(t.strip().lower() == seg_norm for t in titles):
                        candidates.append(c)
                if len(candidates) == 1:
                    current_id = candidates[0].id
                    parent_id = current_id
                    continue
                if len(candidates) > 1:
                    return None, f"دسته‌بندی مبهم است: '{seg}'"

                if on_missing_category == "create" and (not is_dry_run):
                    from adapters.db.repositories.category_repository import CategoryRepository
                    repo = CategoryRepository(db)
                    obj = repo.create_category(business_id=business_id, parent_id=parent_id, translations={"fa": seg, "en": seg})
                    reference_summary["created"]["categories"] += 1
                    cats.append(obj)
                    by_parent.setdefault(parent_id, []).append(obj)
                    current_id = obj.id
                    parent_id = current_id
                    continue
                if on_missing_category == "create" and is_dry_run:
                    return None, f"__WOULD_CREATE__:{seg}"
                return None, f"دسته‌بندی یافت نشد: '{seg}'"

            return current_id, None

        def _resolve_tax_type(item: dict, row_errors: list[str]) -> None:
            if item.get("tax_type_id") is not None:
                return
            code = _norm_text(item.get("tax_type_code"))
            if code:
                _ensure_tax_types_loaded()
                tid = (tax_type_by_code or {}).get(code.lower())
                if not tid:
                    row_errors.append(f"نوع مالیات با کد '{code}' یافت نشد")
                else:
                    item["tax_type_id"] = tid
                    return
            title = _norm_text(item.get("tax_type_title"))
            if title:
                _ensure_tax_types_loaded()
                tid = (tax_type_by_title or {}).get(title.lower())
                if not tid:
                    row_errors.append(f"نوع مالیات با عنوان '{title}' یافت نشد")
                else:
                    item["tax_type_id"] = tid

        def _resolve_tax_unit(item: dict, row_errors: list[str]) -> None:
            if item.get("tax_unit_id") is not None:
                return
            code = _norm_text(item.get("tax_unit_code"))
            if code:
                _ensure_tax_units_loaded()
                uid = (tax_unit_by_code or {}).get(code.lower())
                if not uid:
                    row_errors.append(f"واحد مالیاتی با کد '{code}' یافت نشد")
                else:
                    item["tax_unit_id"] = uid
                    return
            name = _norm_text(item.get("tax_unit_name"))
            if name:
                _ensure_tax_units_loaded()
                uid = (tax_unit_by_name or {}).get(name.lower())
                if not uid:
                    row_errors.append(f"واحد مالیاتی با نام '{name}' یافت نشد")
                else:
                    item["tax_unit_id"] = uid

        def _resolve_attributes(item: dict, row_errors: list[str]) -> None:
            if isinstance(item.get("attribute_ids"), list) and item.get("attribute_ids"):
                ids = [i for i in item["attribute_ids"] if isinstance(i, int)]
                existing_ids = set([a.id for a in _ensure_attributes_loaded()])
                missing = [str(i) for i in ids if i not in existing_ids]
                item["attribute_ids"] = [i for i in ids if i in existing_ids]
                if missing:
                    row_errors.append(f"شناسه(های) ویژگی نامعتبر: {', '.join(missing)}")
            titles = _split_list_text(item.get("attribute_titles"))
            if not titles:
                return
            attrs = _ensure_attributes_loaded()
            by_title = {a.title.strip().lower(): a for a in attrs if a.title}
            resolved_ids: list[int] = []
            missing_titles: list[str] = []
            for t in titles:
                key = t.strip().lower()
                found = by_title.get(key)
                if found:
                    resolved_ids.append(found.id)
                    continue
                if on_missing_attributes == "create" and (not is_dry_run):
                    from adapters.db.repositories.product_attribute_repository import ProductAttributeRepository
                    repo = ProductAttributeRepository(db)
                    try:
                        obj = repo.create(business_id=business_id, title=t, description=None, data_type="text", options=None)
                        reference_summary["created"]["attributes"] += 1
                        attrs.append(obj)
                        by_title[obj.title.strip().lower()] = obj
                        resolved_ids.append(obj.id)
                    except Exception:
                        missing_titles.append(t)
                else:
                    missing_titles.append(t)
            if missing_titles:
                if on_missing_attributes == "create" and is_dry_run:
                    item["_would_create_attribute_titles"] = missing_titles
                else:
                    row_errors.append(f"ویژگی(های) یافت نشد: {', '.join(missing_titles)}")
            current = item.get("attribute_ids") if isinstance(item.get("attribute_ids"), list) else []
            merged = list(dict.fromkeys([*(current or []), *resolved_ids]))
            item["attribute_ids"] = merged

        errors: list[dict] = []
        valid_items: list[dict] = []

        for idx, row in enumerate(data_rows, start=2):
            item: dict[str, Any] = {}
            row_errors: list[str] = []
            row_warnings: list[str] = []
            row_preview: dict[str, Any] = {"row": idx, "resolved": {}, "would_create": {}, "warnings": []}

            for ci, key in enumerate(headers):
                if not key:
                    continue
                val = row[ci] if ci < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                item[key] = val

            # normalize & cast
            if 'item_type' in item:
                item['item_type'] = _normalize_item_type(item.get('item_type')) or 'کالا'
            for k in ['base_sales_price','base_purchase_price','sales_tax_rate','purchase_tax_rate','unit_conversion_factor']:
                if k in item:
                    item[k] = _parse_decimal(item.get(k))
            for k in ['reorder_point','min_order_qty','lead_time_days','category_id','main_unit_id','secondary_unit_id','tax_type_id','tax_unit_id']:
                if k in item:
                    item[k] = _parse_int(item.get(k))
            for k in ['track_inventory','is_sales_taxable','is_purchase_taxable']:
                if k in item:
                    item[k] = _parse_bool(item.get(k)) if item.get(k) is not None else None

            # attribute_ids: comma-separated
            if 'attribute_ids' in item and item['attribute_ids']:
                try:
                    parts = [p.strip() for p in str(item['attribute_ids']).split(',') if p and p.strip()]
                    item['attribute_ids'] = [int(p) for p in parts if p.isdigit()]
                except Exception:
                    item['attribute_ids'] = []

            # Resolve category/tax/attributes
            cat_id, cat_err = _resolve_category_by_id(item.get("category_id"))
            if cat_err:
                resolved_id, resolved_err = _resolve_category_by_name_or_path(item.get("category"), item.get("category_path"))
                if resolved_err and str(resolved_err).startswith("__WOULD_CREATE__:"):
                    seg = str(resolved_err).split(":", 1)[1]
                    row_preview["would_create"]["categories"] = [seg]
                    reference_summary["would_create"]["categories"] += 1
                    row_warnings.append("دسته‌بندی در حالت create ساخته خواهد شد")
                elif resolved_err:
                    row_errors.append(cat_err + " / " + resolved_err)
                else:
                    item["category_id"] = resolved_id
                    if resolved_id is not None:
                        row_preview["resolved"]["category"] = {"category_id": resolved_id}
                        reference_summary["resolved"]["category"] += 1
            else:
                if item.get("category_id") is None and (item.get("category") or item.get("category_path")):
                    resolved_id, resolved_err = _resolve_category_by_name_or_path(item.get("category"), item.get("category_path"))
                    if resolved_err and str(resolved_err).startswith("__WOULD_CREATE__:"):
                        seg = str(resolved_err).split(":", 1)[1]
                        row_preview["would_create"]["categories"] = [seg]
                        reference_summary["would_create"]["categories"] += 1
                        row_warnings.append("دسته‌بندی در حالت create ساخته خواهد شد")
                    elif resolved_err:
                        row_errors.append(resolved_err)
                    else:
                        item["category_id"] = resolved_id
                        if resolved_id is not None:
                            row_preview["resolved"]["category"] = {"category_id": resolved_id}
                            reference_summary["resolved"]["category"] += 1

            _resolve_tax_type(item, row_errors)
            _resolve_tax_unit(item, row_errors)
            _resolve_attributes(item, row_errors)
            if "_would_create_attribute_titles" in item:
                titles = item.get("_would_create_attribute_titles") or []
                if isinstance(titles, list) and titles:
                    row_preview["would_create"]["attributes"] = titles
                    reference_summary["would_create"]["attributes"] += len(titles)
                    row_warnings.append("برخی ویژگی‌ها در حالت create ساخته خواهد شد")
                item.pop("_would_create_attribute_titles", None)

            # validations
            name = item.get('name')
            if not name or str(name).strip() == "":
                row_errors.append('name الزامی است')

            # if code is empty, it will be auto-generated in service
            code = item.get('code')
            if code is not None and str(code).strip() == "":
                item['code'] = None

            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
                if is_dry_run:
                    row_preview["warnings"] = row_warnings
                    preview_rows.append(row_preview)
                continue

            valid_items.append(item)
            if is_dry_run:
                row_preview["warnings"] = row_warnings
                preview_rows.append(row_preview)

        inserted = 0
        updated = 0
        skipped = 0

        if not is_dry_run and valid_items:
            from sqlalchemy import and_ as _and
            from adapters.db.models.product import Product
            from adapters.api.v1.schema_models.product import ProductCreateRequest, ProductUpdateRequest
            from app.services.product_service import create_product, update_product

            def _find_existing(session: Session, data: dict) -> Optional[Product]:
                if match_by == 'code' and data.get('code'):
                    return session.query(Product).filter(_and(Product.business_id == business_id, Product.code == str(data['code']).strip())).first()
                if match_by == 'name' and data.get('name'):
                    return session.query(Product).filter(_and(Product.business_id == business_id, Product.name == str(data['name']).strip())).first()
                return None

            for data in valid_items:
                existing = _find_existing(db, data)
                if existing is None:
                    try:
                        create_product(db, business_id, ProductCreateRequest(**data))
                        inserted += 1
                    except Exception as e:
                        logger.error(f"Create product failed: {e}")
                        skipped += 1
                else:
                    if conflict_policy == 'insert':
                        skipped += 1
                    elif conflict_policy in ('update','upsert'):
                        try:
                            update_product(db, existing.id, business_id, ProductUpdateRequest(**data))
                            updated += 1
                        except Exception as e:
                            logger.error(f"Update product failed: {e}")
                            skipped += 1

        summary = {
            "total": len(data_rows),
            "valid": len(valid_items),
            "invalid": len(errors),
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "dry_run": is_dry_run,
        }

        return success_response(
            data={"summary": summary, "errors": errors, "reference_summary": reference_summary, "preview": preview_rows if is_dry_run else None},
            request=request,
            message="PRODUCTS_IMPORT_RESULT",
        )
    except ApiError:
        raise
    except Exception as e:
        logger.error(f"Import error: {e}", exc_info=True)
        raise ApiError("IMPORT_ERROR", f"خطا در پردازش فایل: {e}", http_status=500)
@router.post("/business/{business_id}/export/pdf",
    summary="خروجی PDF لیست محصولات",
    description="خروجی PDF لیست محصولات با قابلیت فیلتر و انتخاب ستون‌ها",
)
@require_business_access("business_id")
async def export_products_pdf(
    request: Request,
    business_id: int,
    body: dict,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration

    if not ctx.can_read_section("inventory"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.read", http_status=403)

    query_dict = {
        "take": int(body.get("take", 100)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }
    result = list_products(db, business_id, query_dict)
    items = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Apply selected indices filter if requested
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        headers = [col.get("label") or col.get("key") for col in export_columns]
        keys = [col.get("key") for col in export_columns]
    else:
        default_cols = [
            ("code", "کد"),
            ("name", "نام"),
            ("item_type", "نوع"),
            ("category_id", "دسته"),
            ("base_sales_price", "قیمت فروش"),
            ("base_purchase_price", "قیمت خرید"),
            ("main_unit_id", "واحد اصلی"),
            ("secondary_unit_id", "واحد فرعی"),
            ("track_inventory", "کنترل موجودی"),
            ("created_at_formatted", "ایجاد"),
        ]
        keys = [k for k, _ in default_cols]
        headers = [v for _, v in default_cols]

    # Locale and direction
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'

    # Load business info for header
    business_name = ""
    try:
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name or ""
    except Exception:
        business_name = ""

    # Escape helper
    def escape(s: Any) -> str:
        try:
            return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        except Exception:
            return str(s)

    # Build rows
    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key)
            if value is None:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            tds.append(f"<td>{escape(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = ''.join(f"<th>{escape(h)}</th>" for h in headers)

    # Format report datetime based on X-Calendar-Type header
    calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian")
        now_str = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        now_str = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')

    title_text = "گزارش فهرست محصولات" if is_fa else "Products List Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    footer_text = "تولید شده توسط Hesabix" if is_fa else "Generated by Hesabix"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "

    table_html = f"""
    <html lang=\"{html_lang}\" dir=\"{html_dir}\">
      <head>
        <meta charset='utf-8'>
        <style>
          @page {{
            size: A4 landscape;
            margin: 12mm;
            @bottom-{ 'left' if is_fa else 'right' } {{
              content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
              font-size: 10px;
              color: #666;
            }}
          }}
          body {{
            font-family: sans-serif;
            font-size: 11px;
            color: #222;
          }}
          .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #444;
            padding-bottom: 6px;
          }}
          .title {{
            font-size: 16px;
            font-weight: 700;
          }}
          .meta {{
            font-size: 11px;
            color: #555;
          }}
          .table-wrapper {{
            width: 100%;
          }}
          table.report-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
          }}
          thead th {{
            background: #f0f3f7;
            border: 1px solid #c7cdd6;
            padding: 6px 4px;
            text-align: center;
            font-weight: 700;
            white-space: nowrap;
          }}
          tbody td {{
            border: 1px solid #d7dde6;
            padding: 5px 4px;
            vertical-align: top;
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: normal;
          }}
          .footer {{
            position: running(footer);
            font-size: 10px;
            color: #666;
            margin-top: 8px;
            text-align: {'left' if is_fa else 'right'};
          }}
        </style>
      </head>
      <body>
        <div class=\"header\">
          <div>
            <div class=\"title\">{title_text}</div>
            <div class=\"meta\">{label_biz}: {escape(business_name)}</div>
          </div>
          <div class=\"meta\">{label_date}: {escape(now_str)}</div>
        </div>
        <div class=\"table-wrapper\">
          <table class=\"report-table\">
            <thead>
              <tr>{headers_html}</tr>
            </thead>
            <tbody>
              {''.join(rows_html)}
            </tbody>
          </table>
        </div>
        <div class=\"footer\">{footer_text}</div>
      </body>
    </html>
    """

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=table_html).write_pdf(font_config=font_config)

    # Build meaningful filename
    biz_name = business_name
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    base = "products"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/business/{business_id}/bulk-price-update/preview",
    summary="پیش‌نمایش تغییر قیمت‌های گروهی",
    description="پیش‌نمایش تغییرات قیمت قبل از اعمال",
)
@require_business_access("business_id")
def preview_bulk_price_update_endpoint(
    request: Request,
    business_id: int,
    payload: BulkPriceUpdateRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.has_business_permission("inventory", "write"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.write", http_status=403)
    
    result = preview_bulk_price_update(db, business_id, payload)
    return success_response(data=result.dict(), request=request)


@router.post("/business/{business_id}/bulk-price-update/apply",
    summary="اعمال تغییر قیمت‌های گروهی",
    description="اعمال تغییرات قیمت بر روی کالاهای انتخاب شده",
)
@require_business_access("business_id")
def apply_bulk_price_update_endpoint(
    request: Request,
    business_id: int,
    payload: BulkPriceUpdateRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if not ctx.has_business_permission("inventory", "write"):
        raise ApiError("FORBIDDEN", "Missing business permission: inventory.write", http_status=403)
    
    result = apply_bulk_price_update(db, business_id, payload)
    return success_response(data=result, request=request)


