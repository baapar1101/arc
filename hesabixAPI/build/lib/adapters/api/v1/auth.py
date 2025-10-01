# Removed __future__ annotations to fix OpenAPI schema generation

import datetime
from fastapi import APIRouter, Depends, Request, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from app.core.responses import success_response, format_datetime_fields
from app.services.captcha_service import create_captcha
from app.services.auth_service import register_user, login_user, create_password_reset, reset_password, change_password, referral_stats
from app.services.pdf import PDFService
from .schemas import (
	RegisterRequest, LoginRequest, ForgotPasswordRequest, ResetPasswordRequest, 
	ChangePasswordRequest, CreateApiKeyRequest, QueryInfo, FilterItem,
	SuccessResponse, CaptchaResponse, LoginResponse, ApiKeyResponse, 
	ReferralStatsResponse, UserResponse
)
from app.core.auth_dependency import get_current_user, AuthContext
from app.services.api_key_service import list_personal_keys, create_personal_key, revoke_key


router = APIRouter(prefix="/auth", tags=["auth"]) 


@router.post("/captcha", 
	summary="تولید کپچای عددی", 
	description="تولید کپچای عددی برای تأیید هویت در عملیات حساس",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کپچا با موفقیت تولید شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کپچا تولید شد",
						"data": {
							"captcha_id": "abc123def456",
							"image_base64": "iVBORw0KGgoAAAANSUhEUgAA...",
							"ttl_seconds": 180
						}
					}
				}
			}
		}
	}
)
def generate_captcha(db: Session = Depends(get_db)) -> dict:
	captcha_id, image_base64, ttl = create_captcha(db)
	return success_response({
		"captcha_id": captcha_id,
		"image_base64": image_base64,
		"ttl_seconds": ttl,
	})


@router.get("/me", 
	summary="دریافت اطلاعات کاربر کنونی", 
	description="دریافت اطلاعات کامل کاربری که در حال حاضر وارد سیستم شده است",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "اطلاعات کاربر با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "اطلاعات کاربر دریافت شد",
						"data": {
							"id": 1,
							"email": "user@example.com",
							"mobile": "09123456789",
							"first_name": "احمد",
							"last_name": "احمدی",
							"is_active": True,
							"referral_code": "ABC123",
							"referred_by_user_id": None,
							"app_permissions": {"admin": True},
							"created_at": "2024-01-01T00:00:00Z",
							"updated_at": "2024-01-01T00:00:00Z"
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "احراز هویت مورد نیاز است",
						"error_code": "UNAUTHORIZED"
					}
				}
			}
		}
	}
)
def get_current_user_info(
    request: Request,
    ctx: AuthContext = Depends(get_current_user)
) -> dict:
    """دریافت اطلاعات کاربر کنونی"""
    return success_response(ctx.to_dict(), request)


@router.post("/register", 
	summary="ثبت‌نام کاربر جدید", 
	description="ثبت‌نام کاربر جدید در سیستم با تأیید کپچا",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کاربر با موفقیت ثبت‌نام شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ثبت‌نام با موفقیت انجام شد",
						"data": {
							"api_key": "sk_1234567890abcdef",
							"expires_at": None,
							"user": {
								"id": 1,
								"first_name": "احمد",
								"last_name": "احمدی",
								"email": "ahmad@example.com",
								"mobile": "09123456789",
								"referral_code": "ABC123",
								"app_permissions": None
							}
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		409: {
			"description": "کاربر با این ایمیل یا موبایل قبلاً ثبت‌نام کرده است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کاربر با این ایمیل قبلاً ثبت‌نام کرده است",
						"error_code": "USER_EXISTS"
					}
				}
			}
		}
	}
)
def register(request: Request, payload: RegisterRequest, db: Session = Depends(get_db)) -> dict:
	user_id = register_user(
		db=db,
		first_name=payload.first_name,
		last_name=payload.last_name,
		email=payload.email,
		mobile=payload.mobile,
		password=payload.password,
		captcha_id=payload.captcha_id,
		captcha_code=payload.captcha_code,
		referrer_code=payload.referrer_code,
	)
	# Create a session api key similar to login
	user_agent = request.headers.get("User-Agent")
	ip = request.client.host if request.client else None
	from app.core.security import generate_api_key
	from adapters.db.repositories.api_key_repo import ApiKeyRepository
	api_key, key_hash = generate_api_key()
	api_repo = ApiKeyRepository(db)
	api_repo.create_session_key(user_id=user_id, key_hash=key_hash, device_id=payload.device_id, user_agent=user_agent, ip=ip, expires_at=None)
	from adapters.db.models.user import User
	user_obj = db.get(User, user_id)
	user = {"id": user_id, "first_name": payload.first_name, "last_name": payload.last_name, "email": payload.email, "mobile": payload.mobile, "referral_code": getattr(user_obj, "referral_code", None), "app_permissions": getattr(user_obj, "app_permissions", None)}
	response_data = {"api_key": api_key, "expires_at": None, "user": user}
	formatted_data = format_datetime_fields(response_data, request)
	return success_response(formatted_data, request)


@router.post("/login", 
	summary="ورود با ایمیل یا موبایل", 
	description="ورود کاربر به سیستم با استفاده از ایمیل یا شماره موبایل و رمز عبور",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "ورود با موفقیت انجام شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ورود با موفقیت انجام شد",
						"data": {
							"api_key": "sk_1234567890abcdef",
							"expires_at": "2024-01-02T00:00:00Z",
							"user": {
								"id": 1,
								"first_name": "احمد",
								"last_name": "احمدی",
								"email": "ahmad@example.com",
								"mobile": "09123456789",
								"referral_code": "ABC123",
								"app_permissions": {"admin": True}
							}
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		401: {
			"description": "اطلاعات ورود نامعتبر است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "ایمیل یا رمز عبور اشتباه است",
						"error_code": "INVALID_CREDENTIALS"
					}
				}
			}
		}
	}
)
def login(request: Request, payload: LoginRequest, db: Session = Depends(get_db)) -> dict:
	user_agent = request.headers.get("User-Agent")
	ip = request.client.host if request.client else None
	api_key, expires_at, user = login_user(
		db=db,
		identifier=payload.identifier,
		password=payload.password,
		captcha_id=payload.captcha_id,
		captcha_code=payload.captcha_code,
		device_id=payload.device_id,
		user_agent=user_agent,
		ip=ip,
	)
	# Ensure referral_code is included
	from adapters.db.repositories.user_repo import UserRepository
	repo = UserRepository(db)
	from adapters.db.models.user import User
	user_obj = None
	if 'id' in user and user['id']:
		user_obj = repo.db.get(User, user['id'])
	if user_obj is not None:
		user["referral_code"] = getattr(user_obj, "referral_code", None)
	response_data = {"api_key": api_key, "expires_at": expires_at, "user": user}
	formatted_data = format_datetime_fields(response_data, request)
	return success_response(formatted_data, request)


@router.post("/forgot-password", 
	summary="ایجاد توکن بازنشانی رمز عبور", 
	description="ایجاد توکن برای بازنشانی رمز عبور کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "توکن بازنشانی با موفقیت ایجاد شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "توکن بازنشانی ارسال شد",
						"data": {
							"ok": True,
							"token": "reset_token_1234567890abcdef"
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		404: {
			"description": "کاربر یافت نشد",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کاربر با این ایمیل یا موبایل یافت نشد",
						"error_code": "USER_NOT_FOUND"
					}
				}
			}
		}
	}
)
def forgot_password(payload: ForgotPasswordRequest, db: Session = Depends(get_db)) -> dict:
	# In production do not return token; send via email/SMS. Here we return for dev/testing.
	token = create_password_reset(db=db, identifier=payload.identifier, captcha_id=payload.captcha_id, captcha_code=payload.captcha_code)
	return success_response({"ok": True, "token": token if token else None})


@router.post("/reset-password", 
	summary="بازنشانی رمز عبور با توکن", 
	description="بازنشانی رمز عبور کاربر با استفاده از توکن دریافتی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "رمز عبور با موفقیت بازنشانی شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "رمز عبور با موفقیت تغییر کرد",
						"data": {
							"ok": True
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		404: {
			"description": "توکن نامعتبر یا منقضی شده است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "توکن نامعتبر یا منقضی شده است",
						"error_code": "INVALID_TOKEN"
					}
				}
			}
		}
	}
)
def reset_password_endpoint(payload: ResetPasswordRequest, db: Session = Depends(get_db)) -> dict:
	reset_password(db=db, token=payload.token, new_password=payload.new_password, captcha_id=payload.captcha_id, captcha_code=payload.captcha_code)
	return success_response({"ok": True})


@router.get("/api-keys", 
	summary="لیست کلیدهای API شخصی", 
	description="دریافت لیست کلیدهای API شخصی کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "لیست کلیدهای API با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "لیست کلیدهای API دریافت شد",
						"data": [
							{
								"id": 1,
								"name": "کلید اصلی",
								"scopes": "read,write",
								"device_id": "device123",
								"user_agent": "Mozilla/5.0...",
								"ip": "192.168.1.1",
								"expires_at": None,
								"last_used_at": "2024-01-01T12:00:00Z",
								"created_at": "2024-01-01T00:00:00Z"
							}
						]
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def list_keys(request: Request, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	items = list_personal_keys(db, ctx.user.id)
	return success_response(items)


@router.post("/api-keys", 
	summary="ایجاد کلید API شخصی", 
	description="ایجاد کلید API جدید برای کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کلید API با موفقیت ایجاد شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کلید API ایجاد شد",
						"data": {
							"id": 1,
							"api_key": "sk_1234567890abcdef"
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def create_key(request: Request, payload: CreateApiKeyRequest, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	id_, api_key = create_personal_key(db, ctx.user.id, payload.name, payload.scopes, None)
	return success_response({"id": id_, "api_key": api_key})


@router.post("/change-password", 
	summary="تغییر رمز عبور", 
	description="تغییر رمز عبور کاربر با تأیید رمز عبور فعلی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "رمز عبور با موفقیت تغییر کرد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "رمز عبور با موفقیت تغییر کرد",
						"data": {
							"ok": True
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "رمز عبور فعلی اشتباه است",
						"error_code": "INVALID_CURRENT_PASSWORD"
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def change_password_endpoint(request: Request, payload: ChangePasswordRequest, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	# دریافت translator از request state
	translator = getattr(request.state, "translator", None)
	
	change_password(
		db=db,
		user_id=ctx.user.id,
		current_password=payload.current_password,
		new_password=payload.new_password,
		confirm_password=payload.confirm_password,
		translator=translator
	)
	return success_response({"ok": True})


@router.delete("/api-keys/{key_id}", 
	summary="حذف کلید API", 
	description="حذف کلید API مشخص شده",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کلید API با موفقیت حذف شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کلید API حذف شد",
						"data": {
							"ok": True
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		},
		404: {
			"description": "کلید API یافت نشد",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کلید API یافت نشد",
						"error_code": "API_KEY_NOT_FOUND"
					}
				}
			}
		}
	}
)
def delete_key(request: Request, key_id: int, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	revoke_key(db, ctx.user.id, key_id)
	return success_response({"ok": True})


@router.get("/referrals/stats", 
	summary="آمار معرفی‌ها", 
	description="دریافت آمار معرفی‌های کاربر فعلی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "آمار معرفی‌ها با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "آمار معرفی‌ها دریافت شد",
						"data": {
							"total_referrals": 25,
							"active_referrals": 20,
							"recent_referrals": 5,
							"referral_rate": 0.8
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def get_referral_stats(request: Request, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db), start: str = Query(None, description="تاریخ شروع (ISO format)"), end: str = Query(None, description="تاریخ پایان (ISO format)")):
	from datetime import datetime
	start_dt = datetime.fromisoformat(start) if start else None
	end_dt = datetime.fromisoformat(end) if end else None
	stats = referral_stats(db=db, user_id=ctx.user.id, start=start_dt, end=end_dt)
	return success_response(stats)


@router.post("/referrals/list", 
	summary="لیست معرفی‌ها با فیلتر پیشرفته", 
	description="دریافت لیست معرفی‌ها با قابلیت فیلتر، جستجو، مرتب‌سازی و صفحه‌بندی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "لیست معرفی‌ها با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "لیست معرفی‌ها دریافت شد",
						"data": {
							"items": [
								{
									"id": 1,
									"first_name": "علی",
									"last_name": "احمدی",
									"email": "ali@example.com",
									"mobile": "09123456789",
									"created_at": "2024-01-01T00:00:00Z"
								}
							],
							"total": 1,
							"page": 1,
							"limit": 10,
							"total_pages": 1,
							"has_next": False,
							"has_prev": False
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def get_referral_list_advanced(
	request: Request,
	query_info: QueryInfo,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""
	دریافت لیست معرفی‌ها با قابلیت فیلتر پیشرفته
	
	پارامترهای QueryInfo:
	- sort_by: فیلد مرتب‌سازی (مثال: created_at, first_name, last_name, email)
	- sort_desc: ترتیب نزولی (true/false)
	- take: تعداد رکورد در هر صفحه (پیش‌فرض: 10)
	- skip: تعداد رکورد صرف‌نظر شده (پیش‌فرض: 0)
	- search: عبارت جستجو
	- search_fields: فیلدهای جستجو (مثال: ["first_name", "last_name", "email"])
	- filters: آرایه فیلترها با ساختار:
	  [
		{
		  "property": "created_at",
		  "operator": ">=",
		  "value": "2024-01-01T00:00:00"
		},
		{
		  "property": "first_name", 
		  "operator": "*",
		  "value": "احمد"
		}
	  ]
	"""
	from adapters.db.repositories.user_repo import UserRepository
	from adapters.db.models.user import User
	from datetime import datetime
	
	# Create a custom query for referrals
	repo = UserRepository(db)
	
	# Add filter for referrals only (users with referred_by_user_id = current user)
	referral_filter = FilterItem(
		property="referred_by_user_id",
		operator="=",
		value=ctx.user.id
	)
	
	# Add referral filter to existing filters
	if query_info.filters is None:
		query_info.filters = [referral_filter]
	else:
		query_info.filters.append(referral_filter)
	
	# Set default search fields for referrals
	if query_info.search_fields is None:
		query_info.search_fields = ["first_name", "last_name", "email"]
	
	# Execute query with filters
	referrals, total = repo.query_with_filters(query_info)
	
	# Convert to dictionary format
	referral_dicts = [repo.to_dict(referral) for referral in referrals]
	
	# Format datetime fields
	formatted_referrals = format_datetime_fields(referral_dicts, request)
	
	# Calculate pagination info
	page = (query_info.skip // query_info.take) + 1
	total_pages = (total + query_info.take - 1) // query_info.take
	
	return success_response({
		"items": formatted_referrals,
		"total": total,
		"page": page,
		"limit": query_info.take,
		"total_pages": total_pages,
		"has_next": page < total_pages,
		"has_prev": page > 1
	}, request)


@router.post("/referrals/export/pdf", 
	summary="خروجی PDF لیست معرفی‌ها", 
	description="خروجی PDF لیست معرفی‌ها با قابلیت فیلتر و انتخاب سطرهای خاص",
	responses={
		200: {
			"description": "فایل PDF با موفقیت تولید شد",
			"content": {
				"application/pdf": {
					"schema": {
						"type": "string",
						"format": "binary"
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def export_referrals_pdf(
	request: Request,
	query_info: QueryInfo,
	selected_only: bool = False,
	selected_indices: str | None = None,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> Response:
	"""
	خروجی PDF لیست معرفی‌ها
	
	پارامترها:
	- selected_only: آیا فقط سطرهای انتخاب شده export شوند
	- selected_indices: لیست ایندکس‌های انتخاب شده (JSON string)
	- سایر پارامترهای QueryInfo برای فیلتر
	"""
	from app.services.pdf import PDFService
	from app.services.auth_service import referral_stats
	import json
	
	# Parse selected indices if provided
	indices = None
	if selected_only and selected_indices:
		try:
			indices = json.loads(selected_indices)
		except (json.JSONDecodeError, TypeError):
			indices = None
	
	# Get stats for the report
	stats = None
	try:
		# Extract date range from filters if available
		start_date = None
		end_date = None
		if query_info.filters:
			for filter_item in query_info.filters:
				if filter_item.property == 'created_at':
					if filter_item.operator == '>=':
						start_date = filter_item.value
					elif filter_item.operator == '<':
						end_date = filter_item.value
		
		stats = referral_stats(
			db=db,
			user_id=ctx.user.id,
			start=start_date,
			end=end_date
		)
	except Exception:
		pass  # Continue without stats
	
	# Get calendar type from request headers
	calendar_header = request.headers.get("X-Calendar-Type", "jalali")
	calendar_type = "jalali" if calendar_header.lower() in ["jalali", "persian", "shamsi"] else "gregorian"
	
	# Generate PDF using new modular service
	pdf_service = PDFService()
	
	# Get locale from request headers
	locale_header = request.headers.get("Accept-Language", "fa")
	locale = "fa" if locale_header.startswith("fa") else "en"
	
	pdf_bytes = pdf_service.generate_pdf(
		module_name='marketing',
		data={},  # Empty data - module will fetch its own data
		calendar_type=calendar_type,
		locale=locale,
		db=db,
		user_id=ctx.user.id,
		query_info=query_info,
		selected_indices=indices,
		stats=stats
	)
	
	# Return PDF response
	from fastapi.responses import Response
	import datetime
	
	filename = f"referrals_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
	
	return Response(
		content=pdf_bytes,
		media_type="application/pdf",
		headers={
			"Content-Disposition": f"attachment; filename={filename}",
			"Content-Length": str(len(pdf_bytes))
		}
	)


@router.post("/referrals/export/excel", 
	summary="خروجی Excel لیست معرفی‌ها", 
	description="خروجی Excel لیست معرفی‌ها با قابلیت فیلتر و انتخاب سطرهای خاص",
	responses={
		200: {
			"description": "فایل Excel با موفقیت تولید شد",
			"content": {
				"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
					"schema": {
						"type": "string",
						"format": "binary"
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def export_referrals_excel(
	request: Request,
	query_info: QueryInfo,
	selected_only: bool = False,
	selected_indices: str | None = None,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> Response:
	"""
	خروجی Excel لیست معرفی‌ها (فایل Excel واقعی برای دانلود)
	
	پارامترها:
	- selected_only: آیا فقط سطرهای انتخاب شده export شوند
	- selected_indices: لیست ایندکس‌های انتخاب شده (JSON string)
	- سایر پارامترهای QueryInfo برای فیلتر
	"""
	from app.services.pdf import PDFService
	import json
	import io
	from openpyxl import Workbook
	from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
	
	# Parse selected indices if provided
	indices = None
	if selected_only and selected_indices:
		try:
			indices = json.loads(selected_indices)
		except (json.JSONDecodeError, TypeError):
			indices = None
	
	# Get calendar type from request headers
	calendar_header = request.headers.get("X-Calendar-Type", "jalali")
	calendar_type = "jalali" if calendar_header.lower() in ["jalali", "persian", "shamsi"] else "gregorian"
	
	# Generate Excel data using new modular service
	pdf_service = PDFService()
	
	# Get locale from request headers
	locale_header = request.headers.get("Accept-Language", "fa")
	locale = "fa" if locale_header.startswith("fa") else "en"
	
	excel_data = pdf_service.generate_excel_data(
		module_name='marketing',
		data={},  # Empty data - module will fetch its own data
		calendar_type=calendar_type,
		locale=locale,
		db=db,
		user_id=ctx.user.id,
		query_info=query_info,
		selected_indices=indices
	)
	
	# Create Excel workbook
	wb = Workbook()
	ws = wb.active
	ws.title = "Referrals"
	
	# Define styles
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	
	# Add headers
	if excel_data:
		headers = list(excel_data[0].keys())
		for col, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col, value=header)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = header_alignment
			cell.border = border
		
		# Add data rows
		for row, data in enumerate(excel_data, 2):
			for col, header in enumerate(headers, 1):
				cell = ws.cell(row=row, column=col, value=data.get(header, ""))
				cell.border = border
				# Center align for numbers and dates
				if header in ["ردیف", "Row", "تاریخ ثبت", "Registration Date"]:
					cell.alignment = Alignment(horizontal="center")
	
	# Auto-adjust column widths
	for column in ws.columns:
		max_length = 0
		column_letter = column[0].column_letter
		for cell in column:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = min(max_length + 2, 50)
		ws.column_dimensions[column_letter].width = adjusted_width
	
	# Save to BytesIO
	excel_buffer = io.BytesIO()
	wb.save(excel_buffer)
	excel_buffer.seek(0)
	
	# Generate filename
	filename = f"referrals_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
	
	# Return Excel file as response
	return Response(
		content=excel_buffer.getvalue(),
		media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		headers={
			"Content-Disposition": f"attachment; filename={filename}",
			"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
		}
	)

