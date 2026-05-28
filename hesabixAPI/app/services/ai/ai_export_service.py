"""
خروجی Excel/PDF برای ابزار export_business_data در AI (همگام، بدون FastAPI).
"""
from __future__ import annotations

import io
from typing import Any, Dict, Tuple

from sqlalchemy.orm import Session

_MAX_ROWS = 10_000


def _excel_from_rows(headers: list[str], rows: list[list[Any]], sheet_title: str = "Export") -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.sheet_view.rightToLeft = True
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_persons_sync(db: Session, business_id: int, query: Dict[str, Any]) -> bytes:
    from app.services.person_service import get_persons_by_business

    q = dict(query)
    q["take"] = min(int(q.get("take") or _MAX_ROWS), _MAX_ROWS)
    q["skip"] = int(q.get("skip") or 0)
    result = get_persons_by_business(db, business_id, q, q.get("fiscal_year_id"))
    items = result.get("items") or []
    if not items:
        return _excel_from_rows(["پیام"], [["داده‌ای یافت نشد"]], "Persons")
    keys = ["id", "code", "name", "person_type", "phone", "mobile", "email"]
    headers = ["شناسه", "کد", "نام", "نوع", "تلفن", "موبایل", "ایمیل"]
    rows = []
    for it in items:
        rows.append([it.get(k) for k in keys])
    return _excel_from_rows(headers, rows, "Persons")


def export_products_sync(db: Session, business_id: int, query: Dict[str, Any]) -> bytes:
    from app.services.product_service import list_products

    q = dict(query)
    q["take"] = min(int(q.get("take") or _MAX_ROWS), _MAX_ROWS)
    q["skip"] = int(q.get("skip") or 0)
    result = list_products(db, business_id, q)
    items = result.get("items") or []
    headers = ["شناسه", "کد", "نام", "نوع", "قیمت فروش", "قیمت خرید"]
    keys = ["id", "code", "name", "item_type", "base_sales_price", "base_purchase_price"]
    rows = [[it.get(k) for k in keys] for it in items]
    if not rows:
        rows = [["—"]]
    return _excel_from_rows(headers, rows, "Products")


def export_business_file(
    db: Session,
    business_id: int,
    export_type: str,
    fmt: str,
    filters: Dict[str, Any],
) -> Tuple[bytes, str, str]:
    """برمی‌گرداند: (bytes, filename, mime_type)."""
    et = export_type.strip().lower()
    f = fmt.strip().lower()
    if f in ("xlsx", "excel"):
        f = "excel"

    if et == "persons":
        if f == "pdf":
            raise ValueError("خروجی PDF اشخاص از AI فعلاً پشتیبانی نمی‌شود؛ format=excel بزنید")
        data = export_persons_sync(db, business_id, filters)
        return data, f"persons_{business_id}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if et == "products":
        if f == "pdf":
            raise ValueError("خروجی PDF محصولات از AI فعلاً پشتیبانی نمی‌شود")
        data = export_products_sync(db, business_id, filters)
        return data, f"products_{business_id}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if et == "expense_income":
        from app.services.expense_income_service import (
            export_expense_income_excel,
            export_expense_income_pdf,
        )

        if f == "pdf":
            data = export_expense_income_pdf(db, business_id, filters)
            return data, f"expense_income_{business_id}.pdf", "application/pdf"
        data = export_expense_income_excel(db, business_id, filters)
        return (
            data,
            f"expense_income_{business_id}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if et == "documents":
        from app.services.document_service import export_documents_excel, export_documents_pdf

        if f == "pdf":
            data = export_documents_pdf(db, business_id, filters)
            return data, f"documents_{business_id}.pdf", "application/pdf"
        data = export_documents_excel(db, business_id, filters)
        return (
            data,
            f"documents_{business_id}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if et == "invoices":
        from app.services.document_service import export_documents_excel

        inv_f = dict(filters)
        if not inv_f.get("document_type"):
            inv_f.setdefault(
                "document_types",
                [
                    "invoice_sales",
                    "invoice_purchase",
                    "invoice_sales_return",
                    "invoice_purchase_return",
                ],
            )
        data = export_documents_excel(db, business_id, inv_f)
        return (
            data,
            f"invoices_{business_id}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    raise ValueError(f"export_type نامعتبر: {export_type}")
