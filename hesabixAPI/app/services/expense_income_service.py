"""
سرویس هزینه و درآمد (Expense & Income)

این سرویس ثبت اسناد «هزینه/درآمد» را با چند سطر حساب و چند سطر طرف‌حساب پشتیبانی می‌کند.
الگوی پیاده‌سازی بر اساس سرویس دریافت/پرداخت است.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional
from datetime import datetime, date
from decimal import Decimal
import logging

from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func, exists
from sqlalchemy.exc import IntegrityError

from adapters.db.models.document import Document
from adapters.db.models.document_line import DocumentLine
from adapters.db.models.account import Account
from adapters.db.models.currency import Currency
from adapters.db.models.fiscal_year import FiscalYear
from adapters.db.models.user import User
from adapters.db.models.check import Check, CheckType, CheckStatus
from app.core.responses import ApiError
from app.core.cache import get_cache
from app.services.document_monetization_service import ensure_document_policy_allows_creation
from app.services.document_numbering_service import generate_document_code


logger = logging.getLogger(__name__)


def invalidate_expense_income_cache(business_id: int, fiscal_year_id: Optional[int] = None, document_id: Optional[int] = None):
	"""
	حذف تمام کش‌های مربوط به لیست هزینه/درآمد یک کسب‌وکار
	
	این تابع از چند روش استفاده می‌کند:
	1. Tag-based invalidation با set ردیس: حذف انتخابی بر اساس business_id و fiscal_year_id (بهینه‌تر)
	2. Pattern-based invalidation: حذف تمام کلیدهای expense_income_list:* (fallback برای اطمینان)
	3. Redis Pub/Sub: انتشار پیام invalidation برای تمام instanceها
	
	Args:
		business_id: شناسه کسب‌وکار
		fiscal_year_id: شناسه سال مالی (اختیاری)
			- اگر None باشد، تمام کش‌های مربوط به business_id حذف می‌شوند
			- اگر مشخص باشد، فقط کش‌های مربوط به آن fiscal_year_id حذف می‌شوند
		document_id: شناسه سند خاص (اختیاری)
	"""
	cache = get_cache()
	if not cache.enabled:
		return
	
	try:
		# روش 1: استفاده از invalidate_expense_income_by_business (بهینه‌ترین روش)
		deleted_count = cache.invalidate_expense_income_by_business(business_id, fiscal_year_id, document_id)
		if deleted_count > 0:
			logger.info(f"Invalidated {deleted_count} cache keys for business_id {business_id}, fiscal_year_id {fiscal_year_id}, document_id {document_id}")
		
		# روش 2: حذف تمام کلیدهای expense_income_list:* (fallback برای اطمینان کامل)
		pattern = "expense_income_list:*"
		deleted_pattern = cache.delete_pattern(pattern)
		if deleted_pattern > 0:
			logger.info(f"Invalidated {deleted_pattern} cache keys using pattern: {pattern}")
		
		# حذف کش سند خاص اگر مشخص شده باشد
		if document_id:
			document_pattern = f"expense_income:{business_id}:{document_id}*"
			deleted_document = cache.delete_pattern(document_pattern)
			if deleted_document > 0:
				logger.info(f"Invalidated {deleted_document} cache keys for document_id {document_id} using pattern: {document_pattern}")
		
		# روش 3: انتشار پیام invalidation از طریق Redis Pub/Sub
		invalidation_message = {
			"type": "expense_income_cache_invalidation",
			"business_id": business_id,
			"fiscal_year_id": fiscal_year_id,
			"document_id": document_id,
			"timestamp": None
		}
		try:
			import time
			invalidation_message["timestamp"] = time.time()
			cache.publish_invalidation("cache_invalidation", invalidation_message)
			logger.info(f"Published invalidation message for business_id {business_id}, fiscal_year_id {fiscal_year_id}, document_id {document_id}")
		except Exception as pub_error:
			logger.warning(f"Error publishing invalidation message: {pub_error}")
	
	except Exception as e:
		# خطا در invalidate نباید مانع عملیات اصلی شود
		logger.warning(f"Error invalidating expense_income cache for business_id {business_id}: {e}")


# نوع‌های سند
DOCUMENT_TYPE_EXPENSE = "expense"
DOCUMENT_TYPE_INCOME = "income"


def _parse_iso_date(dt: str | datetime | date) -> date:
    if isinstance(dt, date) and not isinstance(dt, datetime):
        return dt
    if isinstance(dt, datetime):
        return dt.date()
    try:
        return datetime.fromisoformat(str(dt)).date()
    except Exception:
        return datetime.utcnow().date()


def _get_fixed_account_by_code(db: Session, account_code: str) -> Account:
    account = db.query(Account).filter(Account.code == str(account_code)).first()
    if not account:
        raise ApiError("ACCOUNT_NOT_FOUND", f"Account with code {account_code} not found", http_status=404)
    return account


def _get_business_fiscal_year(db: Session, business_id: int) -> FiscalYear:
    fy = (
        db.query(FiscalYear)
        .filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True,  # noqa: E712
            )
        )
        .order_by(FiscalYear.start_date.desc())
        .first()
    )
    if not fy:
        raise ApiError("FISCAL_YEAR_NOT_FOUND", "Active fiscal year not found", http_status=404)
    return fy


def create_expense_income(
    db: Session,
    business_id: int,
    user_id: int,
    data: Dict[str, Any],
    *,
    commit: bool = True,
    skip_post_commit_hooks: bool = False,
) -> Dict[str, Any]:
    """
    ایجاد سند هزینه/درآمد با چند سطر حساب و چند سطر طرف‌حساب

    data = {
      "document_type": "expense" | "income",
      "document_date": "2025-10-20",
      "currency_id": 1,
      "description": str?,
      "item_lines": [  # سطرهای حساب‌های هزینه/درآمد
        {"account_id": 123, "amount": 100000, "description": str?},
      ],
      "counterparty_lines": [  # سطرهای طرف‌حساب (بانک/صندوق/شخص/چک ...)
        {
          "transaction_type": "bank" | "cash_register" | "petty_cash" | "check" | "person",
          "amount": 100000,
          "transaction_date": "2025-10-20T10:00:00",
          "description": str?,
          "commission": float?,  # اختیاری
          # فیلدهای اختیاری متناسب با نوع
          "bank_id": int?, "bank_name": str?,
          "cash_register_id": int?, "cash_register_name": str?,
          "petty_cash_id": int?, "petty_cash_name": str?,
          "check_id": int?, "check_number": str?,
          "person_id": int?, "person_name": str?,
        }
      ]
    }
    """
    document_type = str(data.get("document_type", "")).lower()
    if document_type not in (DOCUMENT_TYPE_EXPENSE, DOCUMENT_TYPE_INCOME):
        raise ApiError("INVALID_DOCUMENT_TYPE", "document_type must be 'expense' or 'income'", http_status=400)

    is_income = document_type == DOCUMENT_TYPE_INCOME

    # تاریخ
    document_date = _parse_iso_date(data.get("document_date", datetime.utcnow()))

    # ارز
    currency_id = data.get("currency_id")
    if not currency_id:
        raise ApiError("CURRENCY_REQUIRED", "currency_id is required", http_status=400)
    currency = db.query(Currency).filter(Currency.id == int(currency_id)).first()
    if not currency:
        raise ApiError("CURRENCY_NOT_FOUND", "Currency not found", http_status=404)

    # سال مالی فعال
    fiscal_year = _get_business_fiscal_year(db, business_id)

    # اعتبارسنجی خطوط
    item_lines: List[Dict[str, Any]] = list(data.get("item_lines") or [])
    counterparty_lines: List[Dict[str, Any]] = list(data.get("counterparty_lines") or [])
    if not item_lines:
        raise ApiError("LINES_REQUIRED", "item_lines is required", http_status=400)
    if not counterparty_lines:
        raise ApiError("LINES_REQUIRED", "counterparty_lines is required", http_status=400)

    sum_items = Decimal(0)
    for idx, line in enumerate(item_lines):
        if not line.get("account_id"):
            raise ApiError("ACCOUNT_REQUIRED", f"item_lines[{idx}].account_id is required", http_status=400)
        amount = Decimal(str(line.get("amount", 0)))
        if amount <= 0:
            raise ApiError("AMOUNT_INVALID", f"item_lines[{idx}].amount must be > 0", http_status=400)
        sum_items += amount

    sum_counterparties = Decimal(0)
    for idx, line in enumerate(counterparty_lines):
        amount = Decimal(str(line.get("amount", 0)))
        if amount <= 0:
            raise ApiError("AMOUNT_INVALID", f"counterparty_lines[{idx}].amount must be > 0", http_status=400)
        sum_counterparties += amount

    if sum_items != sum_counterparties:
        raise ApiError("LINES_NOT_BALANCED", "Sum of items and counterparties must be equal", http_status=400)

    # ایجاد سند
    user = db.query(User).filter(User.id == int(user_id)).first()
    if not user:
        raise ApiError("USER_NOT_FOUND", "User not found", http_status=404)

    ensure_document_policy_allows_creation(
        db,
        business_id,
        document_type=document_type,
        document_date=document_date,
        amount=sum_items,
    )

    # دریافت project_id (اختیاری)
    project_id = data.get("project_id")
    if project_id:
        # اعتبارسنجی پروژه
        from adapters.db.models.project import Project
        project = db.query(Project).filter(
            and_(Project.id == project_id, Project.business_id == business_id, Project.is_active == True)
        ).first()
        if not project:
            raise ApiError("PROJECT_NOT_FOUND", "پروژه یافت نشد یا غیرفعال است", http_status=404)

    document: Optional[Document] = None
    max_code_attempts = 5
    for _attempt in range(max_code_attempts):
        code = generate_document_code(db, business_id, document_type, document_date)
        candidate = Document(
            code=code,
            business_id=business_id,
            fiscal_year_id=fiscal_year.id,
            currency_id=int(currency_id),
            created_by_user_id=int(user_id),
            document_date=document_date,
            document_type=document_type,
            is_proforma=False,
            description=(data.get("description") or None),
            extra_info=(data.get("extra_info") if isinstance(data.get("extra_info"), dict) else None),
            project_id=project_id,
        )
        try:
            with db.begin_nested():
                db.add(candidate)
                db.flush()
        except IntegrityError as exc:
            msg = str(getattr(exc.orig, "args", exc))
            if "uq_documents_business_code" in msg or "Duplicate entry" in msg:
                continue
            raise
        else:
            document = candidate
            break

    if not document:
        raise ApiError(
            "DOCUMENT_CODE_RACE",
            "تولید شماره سند پس از چند تلاش ناموفق بود. لطفاً دوباره تلاش کنید.",
            http_status=409,
        )

    # سطرهای حساب‌های هزینه/درآمد
    for line in item_lines:
        account = db.query(Account).filter(
            and_(
                Account.id == int(line.get("account_id")),
                or_(Account.business_id == business_id, Account.business_id == None),  # noqa: E711
            )
        ).first()
        if not account:
            raise ApiError("ACCOUNT_NOT_FOUND", "Item account not found", http_status=404)

        amount = Decimal(str(line.get("amount", 0)))
        description = (line.get("description") or "").strip() or None

        debit_amount = amount if not is_income else Decimal(0)
        credit_amount = amount if is_income else Decimal(0)

        db.add(DocumentLine(
            document_id=document.id,
            account_id=account.id,
            debit=debit_amount,
            credit=credit_amount,
            description=description,
        ))

    # سطرهای طرف‌حساب (بانک/صندوق/شخص/چک/تنخواه)
    for line in counterparty_lines:
        amount = Decimal(str(line.get("amount", 0)))
        description = (line.get("description") or "").strip() or None
        commission = Decimal(str(line.get("commission", 0))) if line.get("commission") else Decimal(0)
        transaction_type: Optional[str] = line.get("transaction_type")

        # انتخاب حساب طرف‌حساب
        account: Optional[Account] = None
        # شناسه‌های موجود برای نگاشت به فیلدهای خط سند
        resolved_bank_account_id = None
        cash_register_id_val = line.get("cash_register_id")
        petty_cash_id_val = line.get("petty_cash_id")
        check_id_val = line.get("check_id")
        person_id_val = line.get("person_id")

        if transaction_type == "bank":
            # سازگاری: bank_account_id یا bank_id
            bank_account_id = line.get("bank_account_id") or line.get("bank_id")
            if bank_account_id:
                try:
                    from adapters.db.models.bank_account import BankAccount
                    bank_account = db.query(BankAccount).filter(
                        and_(
                            BankAccount.id == int(bank_account_id),
                            BankAccount.business_id == business_id,
                        )
                    ).first()
                    if bank_account:
                        resolved_bank_account_id = int(bank_account_id)
                except Exception:
                    resolved_bank_account_id = None
            # حساب دفترکل بانک همیشه حساب ثابت بانک است
            account = _get_fixed_account_by_code(db, "10203")
        elif transaction_type == "cash_register":
            if cash_register_id_val:
                try:
                    from adapters.db.models.cash_register import CashRegister
                    cash_register = db.query(CashRegister).filter(
                        and_(
                            CashRegister.id == int(cash_register_id_val),
                            CashRegister.business_id == business_id,
                        )
                    ).first()
                    if not cash_register:
                        cash_register_id_val = None
                except Exception:
                    cash_register_id_val = None
            account = _get_fixed_account_by_code(db, "10202")
        elif transaction_type == "petty_cash":
            if petty_cash_id_val:
                try:
                    from adapters.db.models.petty_cash import PettyCash
                    petty_cash = db.query(PettyCash).filter(
                        and_(
                            PettyCash.id == int(petty_cash_id_val),
                            PettyCash.business_id == business_id,
                        )
                    ).first()
                    if not petty_cash:
                        petty_cash_id_val = None
                except Exception:
                    petty_cash_id_val = None
            account = _get_fixed_account_by_code(db, "10201")
        elif transaction_type in ("check", "check_expense"):
            # برای چک‌ها از کدهای اسناد دریافتنی/پرداختنی استفاده شود
            account = _get_fixed_account_by_code(db, "10403" if is_income else "20202")
        elif transaction_type == "person":
            # حساب شخص بر اساس نوع (دریافتنی در درآمد / پرداختنی در هزینه)
            if person_id_val:
                try:
                    account = _get_person_account(db, business_id, int(person_id_val), is_income)
                except Exception:
                    account = _get_fixed_account_by_code(db, "20201" if not is_income else "1211")
            else:
                account = _get_fixed_account_by_code(db, "20201" if not is_income else "1211")
        elif line.get("account_id"):
            account = db.query(Account).filter(
                and_(
                    Account.id == int(line.get("account_id")),
                    or_(Account.business_id == business_id, Account.business_id == None),  # noqa: E711
                )
            ).first()
        if not account:
            raise ApiError("ACCOUNT_NOT_FOUND", "Account not found for counterparty line", http_status=404)

        extra_info: Dict[str, Any] = {}
        if transaction_type:
            extra_info["transaction_type"] = transaction_type
        if line.get("transaction_date"):
            extra_info["transaction_date"] = line.get("transaction_date")
        if commission and commission > 0:
            extra_info["commission"] = float(commission)
        if transaction_type == "bank":
            # همواره هر دو کلید را برای سازگاری نگه داریم
            bank_id_val = line.get("bank_account_id") or line.get("bank_id") or resolved_bank_account_id
            if bank_id_val:
                extra_info["bank_id"] = bank_id_val
                extra_info["bank_account_id"] = bank_id_val
            if line.get("bank_name"):
                extra_info["bank_name"] = line.get("bank_name")
            if line.get("bank_account_name"):
                extra_info["bank_account_name"] = line.get("bank_account_name")
        elif transaction_type == "cash_register":
            if line.get("cash_register_id"):
                extra_info["cash_register_id"] = line.get("cash_register_id")
            if line.get("cash_register_name"):
                extra_info["cash_register_name"] = line.get("cash_register_name")
        elif transaction_type == "petty_cash":
            if line.get("petty_cash_id"):
                extra_info["petty_cash_id"] = line.get("petty_cash_id")
            if line.get("petty_cash_name"):
                extra_info["petty_cash_name"] = line.get("petty_cash_name")
        elif transaction_type in ("check", "check_expense"):
            if line.get("check_id"):
                extra_info["check_id"] = line.get("check_id")
            if line.get("check_number"):
                extra_info["check_number"] = line.get("check_number")
        elif transaction_type == "person":
            if line.get("person_id"):
                extra_info["person_id"] = line.get("person_id")
            if line.get("person_name"):
                extra_info["person_name"] = line.get("person_name")

        # برای چک: منطق حسابداری متفاوت است
        # در اسناد هزینه با چک دریافتی: چک از 10403 خارج می‌شود → باید بستانکار شود
        # در اسناد هزینه با چک پرداختی: چک از 20202 خارج می‌شود → باید بدهکار شود
        # در اسناد درآمد: اگر چک قبلاً ثبت شده، نباید دوباره حساب 10403 را بدهکار کرد
        if transaction_type in ("check", "check_expense") and check_id_val:
            try:
                check_obj_for_debit_credit = db.query(Check).filter(
                    and_(
                        Check.id == int(check_id_val),
                        Check.business_id == business_id,
                    )
                ).first()
                
                if check_obj_for_debit_credit:
                    if check_obj_for_debit_credit.type == CheckType.RECEIVED:
                        # چک دریافتی
                        if is_income:
                            # در اسناد درآمد: اگر چک قبلاً ثبت شده، نباید دوباره بدهکار شود
                            # برای حال حاضر، فقط حساب درآمد را ثبت می‌کنیم
                            debit_amount = amount
                            credit_amount = Decimal(0)
                        else:
                            # در اسناد هزینه: چک از 10403 خارج می‌شود → بستانکار می‌شود
                            debit_amount = Decimal(0)
                            credit_amount = amount
                    else:
                        # چک پرداختی
                        if is_income:
                            # در اسناد درآمد: چک پرداختی نمی‌تواند استفاده شود
                            debit_amount = amount
                            credit_amount = Decimal(0)
                        else:
                            # در اسناد هزینه: چک از 20202 خارج می‌شود → بدهکار می‌شود
                            debit_amount = amount
                            credit_amount = Decimal(0)
                else:
                    # اگر چک پیدا نشد، از منطق قبلی استفاده می‌کنیم
                    debit_amount = amount if is_income else Decimal(0)
                    credit_amount = amount if not is_income else Decimal(0)
            except Exception:
                # در صورت خطا، از منطق قبلی استفاده می‌کنیم
                debit_amount = amount if is_income else Decimal(0)
                credit_amount = amount if not is_income else Decimal(0)
        else:
            # برای سایر حساب‌ها: منطق عادی
            debit_amount = amount if is_income else Decimal(0)
            credit_amount = amount if not is_income else Decimal(0)

        db.add(DocumentLine(
            document_id=document.id,
            account_id=account.id,
            person_id=(int(person_id_val) if transaction_type == "person" and person_id_val else None),
            bank_account_id=(int(resolved_bank_account_id) if transaction_type == "bank" and resolved_bank_account_id else None),
            cash_register_id=cash_register_id_val,
            petty_cash_id=petty_cash_id_val,
            check_id=check_id_val,
            debit=debit_amount,
            credit=credit_amount,
            description=description,
            extra_info=extra_info or None,
        ))

        # اگر کارمزد وجود دارد، یک خط کارمزد اضافه کن (هماهنگ با update_expense_income)
        if commission > 0:
            commission_account = _get_fixed_account_by_code(db, "5111")  # کارمزد
            db.add(DocumentLine(
                document_id=document.id,
                account_id=commission_account.id,
                debit=commission if is_income else Decimal(0),
                credit=commission if not is_income else Decimal(0),
                description=f"کارمزد {description or ''}".strip(),
                extra_info={"is_commission_line": True}
            ))

    db.flush()
    
    # تغییر وضعیت چک‌های استفاده شده
    logger.info("=== تغییر وضعیت چک‌های استفاده شده ===")
    for line in counterparty_lines:
        check_id = line.get("check_id")
        if not check_id:
            continue
        
        try:
            check_obj = db.query(Check).filter(Check.id == int(check_id)).first()
            if not check_obj:
                logger.warning(f"چک با شناسه {check_id} یافت نشد")
                continue
            
            transaction_type_line = line.get("transaction_type")
            if transaction_type_line == "check" or transaction_type_line == "check_expense":
                logger.info(f"تغییر وضعیت چک {check_obj.check_number} (id={check_obj.id})")
                logger.info(f"وضعیت قبلی: {check_obj.status}, نوع چک: {check_obj.type}, نوع سند: {document_type}")
                
                if check_obj.type == CheckType.RECEIVED:
                    # چک دریافتی
                    if is_income:
                        # در اسناد درآمد: چک دریافتی استفاده نمی‌شود (قبلاً ثبت شده)
                        # وضعیت را تغییر نمی‌دهیم
                        pass
                    else:
                        # در اسناد هزینه: چک خرج می‌شود
                        check_obj.status = CheckStatus.CLEARED
                        logger.info(f"وضعیت جدید: {check_obj.status}")
                
                elif check_obj.type == CheckType.TRANSFERRED:
                    # چک پرداختی
                    if not is_income:
                        # در اسناد هزینه: چک پرداخته می‌شود
                        check_obj.status = CheckStatus.CLEARED
                        logger.info(f"وضعیت جدید: {check_obj.status}")
                
                check_obj.status_at = datetime.utcnow()
                check_obj.last_action_document_id = document.id
                logger.info(f"وضعیت چک {check_obj.check_number} به {check_obj.status} تغییر یافت")
        
        except Exception as e:
            logger.error(f"خطا در تغییر وضعیت چک {check_id}: {e}", exc_info=True)
    
    if commit:
        db.commit()
        db.refresh(document)
    else:
        db.flush()

    result = document_to_dict(db, document)

    if skip_post_commit_hooks or not commit:
        return result

    # Invalidate cache بعد از ایجاد موفق سند هزینه/درآمد
    invalidate_expense_income_cache(
        business_id=business_id,
        fiscal_year_id=document.fiscal_year_id,
        document_id=document.id
    )

    # همچنین اسناد عمومی را هم invalidate کن
    from app.services.document_service import invalidate_documents_cache
    invalidate_documents_cache(
        business_id=business_id,
        fiscal_year_id=document.fiscal_year_id,
        document_id=document.id,
        document_type=document.document_type
    )

    try:
        from app.services.workflow.workflow_trigger_service import trigger_document_created

        trigger_document_created(
            db=db,
            business_id=business_id,
            document_id=document.id,
            document_type=str(document.document_type),
            user_id=user_id,
            extra_fields=None,
        )
    except Exception as e:
        logger.warning(
            "Failed to trigger workflows for expense/income document %s: %s",
            document.id,
            e,
            exc_info=True,
        )

    return result


def document_to_dict(db: Session, document: Document) -> Dict[str, Any]:
    lines = db.query(DocumentLine).filter(DocumentLine.document_id == document.id).all()
    items: List[Dict[str, Any]] = []
    counterparties: List[Dict[str, Any]] = []
    for ln in lines:
        account = db.query(Account).filter(Account.id == ln.account_id).first()
        row = {
            "id": ln.id,
            "account_id": ln.account_id,
            "account_code": getattr(account, "code", None),
            "account_name": account.name if account else None,
            "debit": float(ln.debit or 0),
            "credit": float(ln.credit or 0),
            "description": ln.description,
            "extra_info": ln.extra_info,
            "person_id": ln.person_id,
            "bank_account_id": ln.bank_account_id,
            "cash_register_id": ln.cash_register_id,
            "petty_cash_id": ln.petty_cash_id,
            "check_id": ln.check_id,
        }
        # ساده: بر اساس وجود transaction_type در extra_info، به عنوان طرف‌حساب تلقی می‌شود
        if ln.extra_info and ln.extra_info.get("transaction_type"):
            counterparties.append(row)
        else:
            items.append(row)

    # ساخته‌های نمایشی: نام ایجادکننده و کد ارز
    from adapters.db.models.user import User
    from adapters.db.models.currency import Currency
    created_by = db.query(User).filter(User.id == document.created_by_user_id).first()
    created_by_name = f"{getattr(created_by, 'first_name', '')} {getattr(created_by, 'last_name', '')}".strip() if created_by else None
    currency = db.query(Currency).filter(Currency.id == document.currency_id).first()
    currency_code = getattr(currency, "code", None)
    currency_symbol = getattr(currency, "symbol", None)
    
    # دریافت نام پروژه
    project_name = None
    if document.project_id:
        from adapters.db.models.project import Project
        project = db.query(Project).filter(Project.id == document.project_id).first()
        if project:
            project_name = project.name

    return {
        "id": document.id,
        "code": document.code,
        "business_id": document.business_id,
        "fiscal_year_id": document.fiscal_year_id,
        "currency_id": document.currency_id,
        "currency_code": currency_code,
        "currency_symbol": currency_symbol,
        "document_type": document.document_type,
        "document_date": document.document_date.isoformat(),
        "registered_at": (document.registered_at.isoformat() if getattr(document, "registered_at", None) else document.document_date.isoformat()),
        "created_by_user_id": document.created_by_user_id,
        "created_by_name": created_by_name,
        "description": document.description,
        "project_id": document.project_id,
        "project_name": project_name,
        "items": items,
        "counterparties": counterparties,
    }


def _apply_expense_income_extra_filters(db: Session, q, query: Dict[str, Any]):
    """فیلتر ستونی جدول (مثل چندانتخابی پروژه)."""
    filters_raw = query.get("filters") or []
    for raw in filters_raw:
        if not isinstance(raw, dict):
            continue
        prop = raw.get("property")
        operator = str(raw.get("operator") or "").strip().lower()
        val = raw.get("value")
        if prop == "project_name" and operator == "in" and val:
            ids: List[int] = []
            for x in (val if isinstance(val, list) else [val]):
                try:
                    ids.append(int(x))
                except (TypeError, ValueError):
                    continue
            if ids:
                q = q.filter(Document.project_id.in_(ids))
        elif prop == "project_name" and operator == "=" and val not in (None, ""):
            try:
                q = q.filter(Document.project_id == int(val))
            except (TypeError, ValueError):
                pass
    return q


def _apply_expense_income_search(db: Session, q, query: Dict[str, Any]):
    search = query.get("search")
    if not search:
        return q
    pattern = f"%{search}%"
    search_fields = query.get("search_fields")
    if not search_fields or not isinstance(search_fields, list):
        return q.filter(Document.code.ilike(pattern))
    sf_set = {str(x) for x in search_fields}
    parts = []
    if "code" in sf_set:
        parts.append(Document.code.ilike(pattern))
    if "description" in sf_set:
        parts.append(Document.description.ilike(pattern))
    if "created_by_name" in sf_set:
        uid_rows = db.query(User.id).filter(
            or_(
                func.concat(User.first_name, " ", User.last_name).ilike(pattern),
                User.first_name.ilike(pattern),
                User.last_name.ilike(pattern),
            )
        ).all()
        uid_list = [row[0] for row in uid_rows if row[0] is not None]
        if uid_list:
            parts.append(Document.created_by_user_id.in_(uid_list))
        else:
            parts.append(Document.id == -1)
    if not parts:
        return q.filter(Document.code.ilike(pattern))
    return q.filter(or_(*parts))


def list_expense_income(
    db: Session,
    business_id: int,
    query: Dict[str, Any],
) -> Dict[str, Any]:
    """لیست اسناد هزینه و درآمد با فیلتر، جست‌وجو و صفحه‌بندی"""
    q = db.query(Document).filter(
        and_(
            Document.business_id == business_id,
            Document.document_type.in_([DOCUMENT_TYPE_EXPENSE, DOCUMENT_TYPE_INCOME]),
        )
    )

    # سال مالی
    fiscal_year_id = query.get("fiscal_year_id")
    try:
        fiscal_year_id_int = int(fiscal_year_id) if fiscal_year_id is not None else None
    except Exception:
        fiscal_year_id_int = None
    if fiscal_year_id_int is None:
        try:
            fy = _get_business_fiscal_year(db, business_id)
            fiscal_year_id_int = fy.id
        except Exception:
            fiscal_year_id_int = None
    if fiscal_year_id_int is not None:
        q = q.filter(Document.fiscal_year_id == fiscal_year_id_int)

    # نوع سند
    doc_type = query.get("document_type")
    if doc_type in (DOCUMENT_TYPE_EXPENSE, DOCUMENT_TYPE_INCOME):
        q = q.filter(Document.document_type == doc_type)

    # پروژه
    project_id = query.get("project_id")
    try:
        project_id_int = int(project_id) if project_id is not None else None
    except Exception:
        project_id_int = None
    if project_id_int is not None:
        q = q.filter(Document.project_id == project_id_int)

    acc_line_id = query.get("account_id")
    if acc_line_id is not None:
        try:
            aid = int(acc_line_id)
            q = q.filter(
                exists().where(
                    and_(
                        DocumentLine.document_id == Document.id,
                        DocumentLine.account_id == aid,
                    )
                )
            )
        except (TypeError, ValueError):
            pass

    q = _apply_expense_income_extra_filters(db, q, query)

    # فیلتر تاریخ
    from_date = query.get("from_date")
    to_date = query.get("to_date")
    if from_date:
        try:
            q = q.filter(Document.document_date >= _parse_iso_date(from_date))
        except Exception:
            pass
    if to_date:
        try:
            q = q.filter(Document.document_date <= _parse_iso_date(to_date))
        except Exception:
            pass

    q = _apply_expense_income_search(db, q, query)

    # مرتب‌سازی (sort چندستونه / sort_by)
    from app.services.document_list_sort import apply_document_dynamic_ordering_from_dict

    q = apply_document_dynamic_ordering_from_dict(q, query)

    # صفحه‌بندی
    skip = int(query.get("skip", 0))
    take = int(query.get("take", 20))
    total = q.count()
    docs = q.offset(skip).limit(take).all()

    return {
        "items": [document_to_dict(db, d) for d in docs],
        "pagination": {
            "total": total,
            "page": (skip // take) + 1,
            "per_page": take,
            "total_pages": (total + take - 1) // take,
            "has_next": skip + take < total,
            "has_prev": skip > 0,
        },
        "query_info": query,
    }


def get_expense_income(db: Session, document_id: int) -> Optional[Dict[str, Any]]:
    """دریافت جزئیات یک سند هزینه/درآمد"""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        return None
    
    return document_to_dict(db, document)


def update_expense_income(
    db: Session,
    document_id: int,
    user_id: int,
    data: Dict[str, Any]
) -> Dict[str, Any]:
    """ویرایش سند هزینه/درآمد"""
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise ApiError("DOCUMENT_NOT_FOUND", "Document not found", http_status=404)
    
    # بررسی نوع سند
    if document.document_type not in (DOCUMENT_TYPE_EXPENSE, DOCUMENT_TYPE_INCOME):
        raise ApiError("INVALID_DOCUMENT_TYPE", "Document is not expense/income", http_status=400)
    
    is_income = document.document_type == DOCUMENT_TYPE_INCOME
    
    # تاریخ
    document_date = _parse_iso_date(data.get("document_date", document.document_date))
    
    # ارز
    currency_id = data.get("currency_id")
    if not currency_id:
        raise ApiError("CURRENCY_REQUIRED", "currency_id is required", http_status=400)
    currency = db.query(Currency).filter(Currency.id == int(currency_id)).first()
    if not currency:
        raise ApiError("CURRENCY_NOT_FOUND", "Currency not found", http_status=404)
    
    # سال مالی فعال
    fiscal_year = _get_business_fiscal_year(db, document.business_id)
    
    # اعتبارسنجی خطوط
    item_lines: List[Dict[str, Any]] = list(data.get("item_lines") or [])
    counterparty_lines: List[Dict[str, Any]] = list(data.get("counterparty_lines") or [])
    if not item_lines:
        raise ApiError("LINES_REQUIRED", "item_lines is required", http_status=400)
    if not counterparty_lines:
        raise ApiError("LINES_REQUIRED", "counterparty_lines is required", http_status=400)
    
    sum_items = Decimal(0)
    for idx, line in enumerate(item_lines):
        if not line.get("account_id"):
            raise ApiError("ACCOUNT_REQUIRED", f"item_lines[{idx}].account_id is required", http_status=400)
        amount = Decimal(str(line.get("amount", 0)))
        if amount <= 0:
            raise ApiError("AMOUNT_INVALID", f"item_lines[{idx}].amount must be > 0", http_status=400)
        sum_items += amount
    
    sum_counterparties = Decimal(0)
    for idx, line in enumerate(counterparty_lines):
        amount = Decimal(str(line.get("amount", 0)))
        if amount <= 0:
            raise ApiError("AMOUNT_INVALID", f"counterparty_lines[{idx}].amount must be > 0", http_status=400)
        sum_counterparties += amount
    
    if sum_items != sum_counterparties:
        raise ApiError("LINES_NOT_BALANCED", "Sum of items and counterparties must be equal", http_status=400)
    
    # حذف خطوط قبلی
    db.query(DocumentLine).filter(DocumentLine.document_id == document_id).delete()
    
    # به‌روزرسانی اطلاعات سند
    document.document_date = document_date
    document.currency_id = int(currency_id)
    document.fiscal_year_id = fiscal_year.id
    document.description = (data.get("description") or "").strip() or None
    document.extra_info = data.get("extra_info") if isinstance(data.get("extra_info"), dict) else None
    
    # سطرهای حساب‌های هزینه/درآمد
    for line in item_lines:
        account = db.query(Account).filter(
            and_(
                Account.id == int(line.get("account_id")),
                or_(Account.business_id == document.business_id, Account.business_id == None),  # noqa: E711
            )
        ).first()
        if not account:
            raise ApiError("ACCOUNT_NOT_FOUND", "Item account not found", http_status=404)
        
        amount = Decimal(str(line.get("amount", 0)))
        description = (line.get("description") or "").strip() or None
        
        debit_amount = amount if not is_income else Decimal(0)
        credit_amount = amount if is_income else Decimal(0)
        
        db.add(DocumentLine(
            document_id=document.id,
            account_id=account.id,
            debit=debit_amount,
            credit=credit_amount,
            description=description,
        ))
    
    # سطرهای طرف‌حساب
    for line in counterparty_lines:
        amount = Decimal(str(line.get("amount", 0)))
        description = (line.get("description") or "").strip() or None
        commission = Decimal(str(line.get("commission", 0))) if line.get("commission") else Decimal(0)
        
        # تعیین نوع تراکنش و حساب مربوطه
        transaction_type = line.get("transaction_type", "bank")
        account = None
        # شناسه‌های موجود برای نگاشت به فیلدهای خط سند
        resolved_bank_account_id = None
        cash_register_id_val = line.get("cash_register_id")
        petty_cash_id_val = line.get("petty_cash_id")
        check_id_val = line.get("check_id")
        person_id_val = line.get("person_id")
        
        if transaction_type == "bank":
            # سازگاری: bank_account_id یا bank_id
            bank_account_id = line.get("bank_account_id") or line.get("bank_id")
            if bank_account_id:
                try:
                    from adapters.db.models.bank_account import BankAccount
                    bank_account = db.query(BankAccount).filter(
                        and_(
                            BankAccount.id == int(bank_account_id),
                            BankAccount.business_id == document.business_id,
                        )
                    ).first()
                    if bank_account:
                        resolved_bank_account_id = int(bank_account_id)
                except Exception:
                    resolved_bank_account_id = None
            # حساب دفترکل بانک همیشه حساب ثابت بانک است
            account = _get_fixed_account_by_code(db, "10203")
        elif transaction_type == "cash_register":
            if cash_register_id_val:
                try:
                    from adapters.db.models.cash_register import CashRegister
                    cash_register = db.query(CashRegister).filter(
                        and_(
                            CashRegister.id == int(cash_register_id_val),
                            CashRegister.business_id == document.business_id,
                        )
                    ).first()
                    if not cash_register:
                        cash_register_id_val = None
                except Exception:
                    cash_register_id_val = None
            account = _get_fixed_account_by_code(db, "10202")
        elif transaction_type == "petty_cash":
            if petty_cash_id_val:
                try:
                    from adapters.db.models.petty_cash import PettyCash
                    petty_cash = db.query(PettyCash).filter(
                        and_(
                            PettyCash.id == int(petty_cash_id_val),
                            PettyCash.business_id == document.business_id,
                        )
                    ).first()
                    if not petty_cash:
                        petty_cash_id_val = None
                except Exception:
                    petty_cash_id_val = None
            account = _get_fixed_account_by_code(db, "10201")
        elif transaction_type in ("check", "check_expense"):
            # برای چک‌ها باید نوع چک را بررسی کنیم تا حساب مناسب را انتخاب کنیم
            check_obj = None
            if check_id_val:
                try:
                    check_obj = db.query(Check).filter(
                        and_(
                            Check.id == int(check_id_val),
                            Check.business_id == document.business_id,
                        )
                    ).first()
                except Exception:
                    pass
            
            # بر اساس نوع چک، حساب مناسب را انتخاب می‌کنیم
            if check_obj:
                if check_obj.type == CheckType.RECEIVED:
                    # چک دریافتی: حساب 10403 (اسناد دریافتنی)
                    account = _get_fixed_account_by_code(db, "10403")
                else:
                    # چک پرداختی: حساب 20202 (اسناد پرداختنی)
                    account = _get_fixed_account_by_code(db, "20202")
            else:
                # اگر چک پیدا نشد یا check_id موجود نبود، از منطق قبلی استفاده می‌کنیم
                account = _get_fixed_account_by_code(db, "10403" if is_income else "20202")
        elif transaction_type == "person":
            # حساب شخص بر اساس نوع (دریافتنی در درآمد / پرداختنی در هزینه)
            if person_id_val:
                try:
                    account = _get_person_account(db, document.business_id, int(person_id_val), is_income)
                except Exception:
                    account = _get_fixed_account_by_code(db, "20201" if not is_income else "1211")
            else:
                account = _get_fixed_account_by_code(db, "20201" if not is_income else "1211")
        elif line.get("account_id"):
            account = db.query(Account).filter(
                and_(
                    Account.id == int(line.get("account_id")),
                    or_(Account.business_id == document.business_id, Account.business_id == None),  # noqa: E711
                )
            ).first()
        if not account:
            raise ApiError("ACCOUNT_NOT_FOUND", "Account not found for counterparty line", http_status=404)
        
        extra_info: Dict[str, Any] = {}
        if transaction_type:
            extra_info["transaction_type"] = transaction_type
        if line.get("transaction_date"):
            extra_info["transaction_date"] = line.get("transaction_date")
        if commission and commission > 0:
            extra_info["commission"] = float(commission)
        if transaction_type == "bank":
            # همواره هر دو کلید را برای سازگاری نگه داریم
            bank_id_val = line.get("bank_account_id") or line.get("bank_id") or resolved_bank_account_id
            if bank_id_val:
                extra_info["bank_id"] = bank_id_val
                extra_info["bank_account_id"] = bank_id_val
            if line.get("bank_name"):
                extra_info["bank_name"] = line.get("bank_name")
            if line.get("bank_account_name"):
                extra_info["bank_account_name"] = line.get("bank_account_name")
        elif transaction_type == "cash_register":
            if line.get("cash_register_id"):
                extra_info["cash_register_id"] = line.get("cash_register_id")
            if line.get("cash_register_name"):
                extra_info["cash_register_name"] = line.get("cash_register_name")
        elif transaction_type == "petty_cash":
            if line.get("petty_cash_id"):
                extra_info["petty_cash_id"] = line.get("petty_cash_id")
            if line.get("petty_cash_name"):
                extra_info["petty_cash_name"] = line.get("petty_cash_name")
        elif transaction_type in ("check", "check_expense"):
            if line.get("check_id"):
                extra_info["check_id"] = line.get("check_id")
            if line.get("check_number"):
                extra_info["check_number"] = line.get("check_number")
        elif transaction_type == "person":
            if line.get("person_id"):
                extra_info["person_id"] = line.get("person_id")
            if line.get("person_name"):
                extra_info["person_name"] = line.get("person_name")

        # برای چک: منطق حسابداری متفاوت است
        # در اسناد هزینه با چک دریافتی: چک از 10403 خارج می‌شود → باید بستانکار شود
        # در اسناد هزینه با چک پرداختی: چک از 20202 خارج می‌شود → باید بدهکار شود
        if transaction_type in ("check", "check_expense") and check_id_val:
            try:
                check_obj_for_debit_credit = db.query(Check).filter(
                    and_(
                        Check.id == int(check_id_val),
                        Check.business_id == document.business_id,
                    )
                ).first()
                
                if check_obj_for_debit_credit:
                    if check_obj_for_debit_credit.type == CheckType.RECEIVED:
                        # چک دریافتی
                        if is_income:
                            # در اسناد درآمد: اگر چک قبلاً ثبت شده، نباید دوباره بدهکار شود
                            debit_amount = amount
                            credit_amount = Decimal(0)
                        else:
                            # در اسناد هزینه: چک از 10403 خارج می‌شود → بستانکار می‌شود
                            debit_amount = Decimal(0)
                            credit_amount = amount
                    else:
                        # چک پرداختی
                        if is_income:
                            # در اسناد درآمد: چک پرداختی نمی‌تواند استفاده شود
                            debit_amount = amount
                            credit_amount = Decimal(0)
                        else:
                            # در اسناد هزینه: چک از 20202 خارج می‌شود → بدهکار می‌شود
                            debit_amount = amount
                            credit_amount = Decimal(0)
                else:
                    # اگر چک پیدا نشد، از منطق قبلی استفاده می‌کنیم
                    debit_amount = amount if is_income else Decimal(0)
                    credit_amount = amount if not is_income else Decimal(0)
            except Exception:
                # در صورت خطا، از منطق قبلی استفاده می‌کنیم
                debit_amount = amount if is_income else Decimal(0)
                credit_amount = amount if not is_income else Decimal(0)
        else:
            # برای سایر حساب‌ها: منطق عادی
            debit_amount = amount if is_income else Decimal(0)
            credit_amount = amount if not is_income else Decimal(0)

        db.add(DocumentLine(
            document_id=document.id,
            account_id=account.id,
            person_id=(int(person_id_val) if transaction_type == "person" and person_id_val else None),
            bank_account_id=(int(resolved_bank_account_id) if transaction_type == "bank" and resolved_bank_account_id else None),
            cash_register_id=cash_register_id_val,
            petty_cash_id=petty_cash_id_val,
            check_id=check_id_val,
            debit=debit_amount,
            credit=credit_amount,
            description=description,
            extra_info=extra_info or None,
        ))
        
        # اگر کارمزد وجود دارد، خط کارمزد اضافه کن
        if commission > 0:
            commission_account = _get_fixed_account_by_code(db, "5111")  # کارمزد
            db.add(DocumentLine(
                document_id=document.id,
                account_id=commission_account.id,
                debit=commission if is_income else Decimal(0),
                credit=commission if not is_income else Decimal(0),
                description=f"کارمزد {description or ''}",
                extra_info={"is_commission_line": True}
            ))
    
    db.flush()
    
    # تغییر وضعیت چک‌های استفاده شده (در update)
    logger.info("[UPDATE_EXPENSE_INCOME] === تغییر وضعیت چک‌های استفاده شده ===")
    for line in counterparty_lines:
        check_id = line.get("check_id")
        if not check_id:
            continue
        
        try:
            check_obj = db.query(Check).filter(Check.id == int(check_id)).first()
            if not check_obj:
                logger.warning(f"[UPDATE_EXPENSE_INCOME] چک با شناسه {check_id} یافت نشد")
                continue
            
            transaction_type_line = line.get("transaction_type")
            if transaction_type_line == "check" or transaction_type_line == "check_expense":
                logger.info(f"[UPDATE_EXPENSE_INCOME] تغییر وضعیت چک {check_obj.check_number} (id={check_obj.id})")
                logger.info(f"[UPDATE_EXPENSE_INCOME] وضعیت قبلی: {check_obj.status}, نوع چک: {check_obj.type}, نوع سند: {document.document_type}")
                
                if check_obj.type == CheckType.RECEIVED:
                    # چک دریافتی
                    if is_income:
                        # در اسناد درآمد: چک دریافتی استفاده نمی‌شود (قبلاً ثبت شده)
                        pass
                    else:
                        # در اسناد هزینه: چک خرج می‌شود
                        check_obj.status = CheckStatus.CLEARED
                        logger.info(f"[UPDATE_EXPENSE_INCOME] وضعیت جدید: {check_obj.status}")
                
                elif check_obj.type == CheckType.TRANSFERRED:
                    # چک پرداختی
                    if not is_income:
                        # در اسناد هزینه: چک پرداخته می‌شود
                        check_obj.status = CheckStatus.CLEARED
                        logger.info(f"[UPDATE_EXPENSE_INCOME] وضعیت جدید: {check_obj.status}")
                
                check_obj.status_at = datetime.utcnow()
                check_obj.last_action_document_id = document.id
                logger.info(f"[UPDATE_EXPENSE_INCOME] وضعیت چک {check_obj.check_number} به {check_obj.status} تغییر یافت")
        
        except Exception as e:
            logger.error(f"[UPDATE_EXPENSE_INCOME] خطا در تغییر وضعیت چک {check_id}: {e}", exc_info=True)
    
    db.commit()
    db.refresh(document)
    
    result = document_to_dict(db, document)
    
    # Invalidate cache بعد از به‌روزرسانی موفق سند هزینه/درآمد
    invalidate_expense_income_cache(
        business_id=document.business_id,
        fiscal_year_id=document.fiscal_year_id,
        document_id=document.id
    )
    
    # همچنین اسناد عمومی را هم invalidate کن
    from app.services.document_service import invalidate_documents_cache
    invalidate_documents_cache(
        business_id=document.business_id,
        fiscal_year_id=document.fiscal_year_id,
        document_id=document.id,
        document_type=document.document_type
    )
    
    return result


def delete_expense_income(db: Session, document_id: int, *, commit: bool = True) -> bool:
    """حذف یک سند هزینه/درآمد"""
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            return False
        
        # بررسی نوع سند
        if document.document_type not in (DOCUMENT_TYPE_EXPENSE, DOCUMENT_TYPE_INCOME):
            return False
        
        # بررسی ارتباط با تراکنش‌های کیف پول
        try:
            from app.services.wallet_service import check_document_has_wallet_transactions
            wallet_check = check_document_has_wallet_transactions(db, document_id)
            if wallet_check["has_wallet_transactions"] and wallet_check.get("has_protected_transactions", False):
                raise ApiError(
                    "DOCUMENT_HAS_WALLET_TRANSACTIONS",
                    wallet_check["message"],
                    http_status=409
                )
        except ApiError:
            raise
        except Exception:
            # اگر به هر دلیل نتوانستیم بررسی کنیم، حذف را متوقف نکن (برای backward compatibility)
            pass
        
        # دریافت اطلاعات قبل از حذف برای invalidation
        business_id = document.business_id
        fiscal_year_id = document.fiscal_year_id
        document_type = document.document_type
        
        # حذف خطوط سند
        db.query(DocumentLine).filter(DocumentLine.document_id == document_id).delete()
        
        # حذف سند
        db.delete(document)
        if commit:
            db.commit()
        else:
            db.flush()
        
        if commit:
            invalidate_expense_income_cache(
                business_id=business_id,
                fiscal_year_id=fiscal_year_id,
                document_id=document_id
            )
            
            from app.services.document_service import invalidate_documents_cache
            invalidate_documents_cache(
                business_id=business_id,
                fiscal_year_id=fiscal_year_id,
                document_id=document_id,
                document_type=document_type
            )
        
        return True
    except ApiError:
        raise
    except Exception as e:
        logger.error(f"Error deleting expense/income document {document_id}: {e}")
        if commit:
            db.rollback()
            return False
        raise ApiError("DELETE_FAILED", str(e), http_status=500) from e


def delete_multiple_expense_income(db: Session, document_ids: List[int]) -> bool:
    """حذف چندین سند هزینه/درآمد"""
    try:
        documents = db.query(Document).filter(
            and_(
                Document.id.in_(document_ids),
                Document.document_type.in_([DOCUMENT_TYPE_EXPENSE, DOCUMENT_TYPE_INCOME])
            )
        ).all()
        
        if not documents:
            return False
        
        # حذف خطوط اسناد
        db.query(DocumentLine).filter(DocumentLine.document_id.in_(document_ids)).delete()
        
        # حذف اسناد
        for document in documents:
            db.delete(document)
        
        db.commit()
        return True
    except Exception as e:
        logger.error(f"Error deleting multiple expense/income documents: {e}")
        db.rollback()
        return False


def export_expense_income_excel(db: Session, business_id: int, query: Dict[str, Any]) -> bytes:
    """خروجی Excel اسناد هزینه/درآمد"""
    # این تابع باید پیاده‌سازی شود
    # فعلاً یک فایل Excel خالی برمی‌گرداند
    import io
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "هزینه و درآمد"
    
    # هدرها
    headers = ["کد سند", "نوع", "تاریخ سند", "مبلغ کل", "توضیحات", "ایجادکننده", "تاریخ ثبت"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # دریافت داده‌ها
    result = list_expense_income(db, business_id, query)
    items = result.get("items", [])
    
    # اضافه کردن داده‌ها
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.get("code", ""))
        ws.cell(row=row, column=2, value=item.get("document_type_name", ""))
        ws.cell(row=row, column=3, value=item.get("document_date", ""))
        ws.cell(row=row, column=4, value=item.get("total_amount", 0))
        ws.cell(row=row, column=5, value=item.get("description", ""))
        ws.cell(row=row, column=6, value=item.get("created_by_name", ""))
        ws.cell(row=row, column=7, value=item.get("registered_at", ""))
    
    # ذخیره در بایت
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_expense_income_pdf(db: Session, business_id: int, query: Dict[str, Any]) -> bytes:
    """خروجی PDF اسناد هزینه/درآمد با WeasyPrint (بدون وابستگی به reportlab)."""
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    import datetime
    # دریافت داده‌ها
    result = list_expense_income(db, business_id, query)
    items = result.get("items", [])
    # ساخت جدول ساده HTML
    headers = ["کد", "نوع", "تاریخ", "مبلغ کل"]
    def _get(item, key, default=""):
        return escape(str(item.get(key, default)))
    rows_html = []
    for item in items:
        total_amount = _get(item, 'total_amount', 0)
        currency_symbol = item.get('currency_symbol') or ''
        amount_display = f"{total_amount} {currency_symbol}".strip() if currency_symbol else total_amount
        rows_html.append(
            "<tr>"
            f"<td>{_get(item, 'code')}</td>"
            f"<td>{_get(item, 'document_type_name')}</td>"
            f"<td>{_get(item, 'document_date')}</td>"
            f"<td>{amount_display}</td>"
            "</tr>"
        )
    now_str = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    html = f"""
<!doctype html>
<html lang="fa">
  <head>
    <meta charset="utf-8">
    <style>
      body {{ font-family: DejaVu Sans, sans-serif; font-size: 12px; }}
      h1 {{ font-size: 16px; margin-bottom: 8px; }}
      table {{ width: 100%; border-collapse: collapse; }}
      th, td {{ border: 1px solid #999; padding: 6px; text-align: right; }}
      th {{ background: #f2f2f2; }}
      .footer {{ margin-top: 12px; font-size: 11px; color: #666; }}
    </style>
  </head>
  <body dir="rtl">
    <h1>گزارش هزینه و درآمد</h1>
    <table>
      <thead>
        <tr>{"".join(f"<th>{escape(h)}</th>" for h in headers)}</tr>
      </thead>
      <tbody>
        {"".join(rows_html)}
      </tbody>
    </table>
    <div class="footer">تولید شده در {escape(now_str)}</div>
  </body>
</html>
"""
    return HTML(string=html).write_pdf(font_config=FontConfiguration())


def generate_expense_income_pdf(db: Session, document_id: int) -> bytes:
    """تولید PDF یک سند هزینه/درآمد با WeasyPrint (بدون reportlab)."""
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    # دریافت داکیومنت به‌صورت دیکشنری قابل استفاده
    doc = get_expense_income(db, document_id)
    if not doc:
        raise ApiError("DOCUMENT_NOT_FOUND", "Document not found", http_status=404)
    code = escape(str(doc.get("code") or ""))
    dtype = escape(str(doc.get("document_type_name") or doc.get("document_type") or ""))
    date = escape(str(doc.get("document_date") or ""))
    total = escape(str(doc.get("total_amount") or ""))
    description = escape(str(doc.get("description") or ""))
    currency_symbol = doc.get("currency_symbol") or ""
    # خطوط
    item_lines = doc.get("item_lines") or doc.get("items") or []
    counterparty_lines = doc.get("counterparty_lines") or doc.get("counterparties") or []
    def _esc(x): return escape(str(x)) if x is not None else ""
    def _fmt_amount(amt): 
        amt_str = _esc(amt)
        return f"{amt_str} {currency_symbol}".strip() if currency_symbol else amt_str
    item_rows = []
    for it in item_lines:
        # پشتیبانی از دو شکل: amount یا debit/credit
        amount = it.get("amount")
        if amount is None:
            debit = it.get("debit") or 0
            credit = it.get("credit") or 0
            amount = debit if abs(debit) >= abs(credit) else credit
        item_rows.append(
            "<tr>"
            f"<td>{_esc(it.get('account_code',''))}</td>"
            f"<td>{_esc(it.get('account_name',''))}</td>"
            f"<td>{_fmt_amount(amount)}</td>"
            f"<td>{_esc(it.get('description',''))}</td>"
            "</tr>"
        )
    cp_rows = []
    for cp in counterparty_lines:
        extra = cp.get("extra_info") or {}
        tx_type = extra.get("transaction_type") or cp.get("transaction_type") or ""
        tx_name = extra.get("transaction_type_name") or cp.get("transaction_type_name") or tx_type
        amount = cp.get("amount")
        if amount is None:
            debit = cp.get("debit") or 0
            credit = cp.get("credit") or 0
            amount = debit if abs(debit) >= abs(credit) else credit
        cp_rows.append(
            "<tr>"
            f"<td>{_esc(tx_name)}</td>"
            f"<td>{_esc(cp.get('account_name') or extra.get('person_name') or extra.get('bank_account_name') or '')}</td>"
            f"<td>{_fmt_amount(amount)}</td>"
            f"<td>{_esc(cp.get('description',''))}</td>"
            "</tr>"
        )
    html = f"""
<!doctype html>
<html lang="fa">
  <head>
    <meta charset="utf-8">
    <style>
      body {{ font-family: DejaVu Sans, sans-serif; font-size: 12px; }}
      h1 {{ font-size: 16px; margin-bottom: 8px; }}
      .section-title {{ margin-top: 14px; margin-bottom: 6px; font-weight: bold; }}
      table {{ width: 100%; border-collapse: collapse; }}
      th, td {{ border: 1px solid #999; padding: 6px; text-align: right; }}
      th {{ background: #f2f2f2; }}
    </style>
  </head>
  <body dir="rtl">
    <h1>سند هزینه/درآمد</h1>
    <div>کد سند: {code}</div>
    <div>نوع سند: {dtype}</div>
    <div>تاریخ سند: {date}</div>
    <div>مبلغ کل: {total} {currency_symbol}</div>
    <div>توضیحات: {description}</div>

    <div class="section-title">اقلام</div>
    <table>
      <thead>
        <tr>
          <th>کد حساب</th>
          <th>نام حساب</th>
          <th>مبلغ</th>
          <th>توضیح</th>
        </tr>
      </thead>
      <tbody>
        {"".join(item_rows)}
      </tbody>
    </table>

    <div class="section-title">طرف‌حساب‌ها</div>
    <table>
      <thead>
        <tr>
          <th>نوع تراکنش</th>
          <th>شرح</th>
          <th>مبلغ</th>
          <th>توضیح</th>
        </tr>
      </thead>
      <tbody>
        {"".join(cp_rows)}
      </tbody>
    </table>
  </body>
</html>
"""
    return HTML(string=html).write_pdf(font_config=FontConfiguration())


def _get_person_account(
    db: Session,
    business_id: int,
    person_id: int,
    is_receivable: bool
) -> Account:
    """دریافت حساب شخص (دریافتنی یا پرداختنی)"""
    from adapters.db.models.person import Person
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise ApiError("PERSON_NOT_FOUND", "Person not found", http_status=404)
    
    # تعیین کد حساب بر اساس نوع
    if is_receivable:
        account_code = "1211"  # دریافتنی‌ها
    else:
        account_code = "20201"  # پرداختنی‌ها
    
    return _get_fixed_account_by_code(db, account_code)


