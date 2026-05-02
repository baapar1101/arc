from __future__ import annotations

import datetime as dt
import io
import re
from typing import Any, Dict, List, Optional, Tuple

from decimal import Decimal
from sqlalchemy.orm import Session

from adapters.db.models.product import Product
from adapters.db.models.price_list import PriceItem, PriceList
from adapters.db.models.currency import Currency
from adapters.api.v1.schema_models.product import (
    BulkProductPriceSheetApplyRequest,
    BulkProductPriceSheetItemsRequest,
    BulkProductPriceSheetRow,
)
from app.services.product_service import list_products


def _quantize_non_negative_integer(value: Decimal) -> Decimal:
    quantized = value.quantize(Decimal("1"))
    if quantized < 0:
        return Decimal("0")
    return quantized


def list_bulk_price_sheet_items(
    db: Session, business_id: int, payload: BulkProductPriceSheetItemsRequest
) -> Dict[str, Any]:
    """لیست PriceItemها برای کالاهای داده‌شده و لیست‌های قیمت انتخابی."""
    q = (
        db.query(PriceItem, PriceList.name, Currency.code)
        .join(PriceList, PriceList.id == PriceItem.price_list_id)
        .outerjoin(Currency, Currency.id == PriceItem.currency_id)
        .filter(
            PriceList.business_id == business_id,
            PriceItem.product_id.in_(payload.product_ids),
            PriceItem.price_list_id.in_(payload.price_list_ids),
        )
        .order_by(PriceList.name, Currency.code, PriceItem.tier_name, PriceItem.product_id)
    )
    rows: List[Dict[str, Any]] = []
    for pi, pl_name, cur_code in q.all():
        rows.append(
            {
                "price_item_id": pi.id,
                "product_id": pi.product_id,
                "price_list_id": pi.price_list_id,
                "price_list_name": pl_name or "",
                "currency_id": pi.currency_id,
                "currency_code": cur_code or "",
                "tier_name": pi.tier_name or "",
                "price": pi.price,
                "created_at": pi.created_at,
                "updated_at": pi.updated_at,
            }
        )
    return {"items": rows}


def apply_bulk_product_price_sheet(
    db: Session, business_id: int, payload: BulkProductPriceSheetApplyRequest
) -> Dict[str, Any]:
    """به‌روزرسانی دسته‌ای قیمت پایه و ردیف‌های لیست قیمت."""
    updated_products = 0
    updated_price_items = 0
    errors: List[str] = []

    for row in payload.items:
        try:
            product = db.get(Product, row.product_id)
            if not product or product.business_id != business_id:
                errors.append(f"کالا یافت نشد: {row.product_id}")
                continue

            product_touched = False

            if row.clear_base_sales_price:
                product.base_sales_price = None
                product_touched = True
            elif row.base_sales_price is not None:
                product.base_sales_price = _quantize_non_negative_integer(Decimal(row.base_sales_price))
                product_touched = True

            if row.clear_base_purchase_price:
                product.base_purchase_price = None
                product_touched = True
            elif row.base_purchase_price is not None:
                product.base_purchase_price = _quantize_non_negative_integer(Decimal(row.base_purchase_price))
                product_touched = True

            for piu in row.price_item_updates:
                pi = db.get(PriceItem, piu.price_item_id)
                if not pi:
                    errors.append(f"ردیف لیست قیمت یافت نشد: {piu.price_item_id}")
                    continue
                if pi.product_id != row.product_id:
                    errors.append(f"ردیف {piu.price_item_id} متعلق به کالای دیگری است")
                    continue
                pl = db.get(PriceList, pi.price_list_id)
                if not pl or pl.business_id != business_id:
                    errors.append(f"لیست قیمت نامعتبر برای ردیف {piu.price_item_id}")
                    continue
                pi.price = _quantize_non_negative_integer(Decimal(piu.price))
                updated_price_items += 1
                product_touched = True

            if product_touched:
                updated_products += 1

        except Exception as e:  # noqa: BLE001
            errors.append(f"خطا در کالای {row.product_id}: {e!s}")

    db.commit()

    msg = f"{updated_products} کالا به‌روز شد"
    if updated_price_items:
        msg += f" ({updated_price_items} ردیف لیست قیمت)"

    return {
        "message": msg,
        "updated_count": updated_products,
        "updated_price_items": updated_price_items,
        "errors": errors,
    }


BULK_PRICE_SHEET_EXCEL_SHEET_NAME = "BulkPrices"
_MAX_IMPORT_ROWS = 20000
_PRODUCT_PAGE_SIZE = 1000
_PID_IN_CHUNK = 500

_FA_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")


def _normalize_search(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    t = str(s).strip()
    return t or None


def collect_matching_product_ids(
    db: Session,
    business_id: int,
    *,
    search: Optional[str],
    search_fields: Optional[List[str]],
) -> List[int]:
    """تمام شناسه‌های کالا مطابق همان پارامترهای جستجوی UI."""
    out: List[int] = []
    skip = 0
    q_search = _normalize_search(search)
    while True:
        query: Dict[str, Any] = {
            "take": _PRODUCT_PAGE_SIZE,
            "skip": skip,
            "search": q_search,
            "search_fields": search_fields,
        }
        res = list_products(db, business_id, query)
        items = res.get("items") or []
        if not items:
            break
        for it in items:
            pid = it.get("id")
            if pid is None:
                continue
            try:
                out.append(int(pid))
            except (TypeError, ValueError):
                continue
        if len(items) < _PRODUCT_PAGE_SIZE:
            break
        skip += _PRODUCT_PAGE_SIZE
    return out


def discover_price_item_columns(
    db: Session,
    business_id: int,
    product_ids: List[int],
    price_list_ids: List[int],
) -> List[Tuple[int, str]]:
    """مرتب‌سازی ستون‌ها مشابه UI (برچسب لیست · ارز · پله)."""
    if not price_list_ids or not product_ids:
        return []
    labels_by_id: Dict[int, str] = {}
    for i in range(0, len(product_ids), _PID_IN_CHUNK):
        chunk = product_ids[i : i + _PID_IN_CHUNK]
        rows = (
            db.query(PriceItem.id, PriceList.name, Currency.code, PriceItem.tier_name)
            .join(PriceList, PriceList.id == PriceItem.price_list_id)
            .outerjoin(Currency, Currency.id == PriceItem.currency_id)
            .filter(
                PriceList.business_id == business_id,
                PriceItem.product_id.in_(chunk),
                PriceItem.price_list_id.in_(price_list_ids),
            )
            .distinct()
        )
        for pi_id, pl_name, cur_code, tier in rows.all():
            pid_int = int(pi_id)
            label = f"{pl_name or ''} · {cur_code or ''} · {tier or ''}"
            labels_by_id.setdefault(pid_int, label)
    return sorted(labels_by_id.items(), key=lambda x: x[1])


def _price_map_for_products(
    db: Session,
    business_id: int,
    product_ids: List[int],
    price_list_ids: List[int],
    column_price_item_ids: List[int],
) -> Dict[Tuple[int, int], Decimal]:
    out: Dict[Tuple[int, int], Decimal] = {}
    if not product_ids or not column_price_item_ids:
        return out
    flt = [
        PriceList.business_id == business_id,
        PriceItem.product_id.in_(product_ids),
        PriceItem.id.in_(column_price_item_ids),
    ]
    if price_list_ids:
        flt.append(PriceItem.price_list_id.in_(price_list_ids))
    rows = db.query(PriceItem.product_id, PriceItem.id, PriceItem.price).join(
        PriceList, PriceList.id == PriceItem.price_list_id
    ).filter(*flt)
    for prod_id, pi_id, price in rows.all():
        out[(int(prod_id), int(pi_id))] = price if isinstance(price, Decimal) else Decimal(str(price))
    return out


def _num_or_none(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    return None


def export_bulk_product_price_sheet_excel(
    db: Session,
    business_id: int,
    *,
    search: Optional[str],
    search_fields: Optional[List[str]],
    price_list_ids: List[int],
    rtl_sheet: bool = False,
) -> Tuple[bytes, str]:
    """ساخت فایل xlsx؛ ستون‌های pi_<id> برای ردیف‌های PriceItem."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    product_ids = collect_matching_product_ids(
        db, business_id, search=search, search_fields=search_fields
    )
    col_meta = discover_price_item_columns(db, business_id, product_ids, price_list_ids)
    col_ids = [c[0] for c in col_meta]

    wb = Workbook()
    ws = wb.active
    ws.title = BULK_PRICE_SHEET_EXCEL_SHEET_NAME
    if rtl_sheet:
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["product_id", "code", "name", "base_sales_price", "base_purchase_price"]
    headers.extend([f"pi_{cid}" for cid in col_ids])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    skip = 0
    q_search = _normalize_search(search)
    while True:
        chunk_items = list_products(
            db,
            business_id,
            {
                "take": _PRODUCT_PAGE_SIZE,
                "skip": skip,
                "search": q_search,
                "search_fields": search_fields,
            },
        ).get("items") or []
        if not chunk_items:
            break
        ids_chunk = []
        for it in chunk_items:
            pid = it.get("id")
            if pid is not None:
                try:
                    ids_chunk.append(int(pid))
                except (TypeError, ValueError):
                    pass
        pmap = _price_map_for_products(db, business_id, ids_chunk, price_list_ids, col_ids)

        for it in chunk_items:
            pid = it.get("id")
            if pid is None:
                continue
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                continue
            row_vals: List[Any] = [
                pid_int,
                it.get("code"),
                it.get("name"),
                _num_or_none(it.get("base_sales_price")),
                _num_or_none(it.get("base_purchase_price")),
            ]
            for cid in col_ids:
                key = (pid_int, cid)
                pr = pmap.get(key)
                row_vals.append(float(pr) if pr is not None else None)
            ws.append(row_vals)

        if len(chunk_items) < _PRODUCT_PAGE_SIZE:
            break
        skip += _PRODUCT_PAGE_SIZE

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bulk_prices_sheet_{business_id}_{ts}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename


def _norm_header_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip().replace("\u200c", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _translate_digits(s: str) -> str:
    return s.translate(_FA_DIGITS)


def _parse_excel_price_cell(val: Any) -> Optional[Decimal]:
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, Decimal):
        return _quantize_non_negative_integer(val)
    if isinstance(val, (int, float)):
        return _quantize_non_negative_integer(Decimal(str(val)))
    s = _translate_digits(str(val).strip()).replace(",", "").replace("٬", "").replace(" ", "")
    if not s:
        return None
    try:
        return _quantize_non_negative_integer(Decimal(s))
    except Exception:
        return None


_PI_COL_RE = re.compile(r"^pi_(\d+)$", re.I)


def _resolve_header_key(h: str) -> Optional[str]:
    raw = _norm_header_cell(h)
    if not raw:
        return None
    low = raw.lower()
    aliases = {
        "product_id": "product_id",
        "شناسه کالا": "product_id",
        "کد": "code",
        "code": "code",
        "نام": "name",
        "name": "name",
        "base_sales_price": "base_sales_price",
        "قیمت فروش پایه": "base_sales_price",
        "base_purchase_price": "base_purchase_price",
        "قیمت خرید پایه": "base_purchase_price",
    }
    if low in aliases:
        return aliases[low]
    return aliases.get(raw)


def import_bulk_product_price_sheet_from_excel(
    db: Session,
    business_id: int,
    content: bytes,
) -> Dict[str, Any]:
    """Parse و اعمال دسته‌ای (در بسته‌های ۵۰۰ تایی)."""
    import zipfile
    from openpyxl import load_workbook

    def _validate_sig(buf: bytes) -> bool:
        try:
            if not buf.startswith(b"PK"):
                return False
            with zipfile.ZipFile(io.BytesIO(buf), "r") as zf:
                return any(n.startswith("xl/") for n in zf.namelist())
        except Exception:
            return False

    if len(content) < 100 or not _validate_sig(content):
        return {"ok": False, "message": "فایل Excel معتبر نیست", "errors": []}

    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception:
        return {"ok": False, "message": "خواندن فایل اکسل ناموفق بود", "errors": []}

    ws = wb[BULK_PRICE_SHEET_EXCEL_SHEET_NAME] if BULK_PRICE_SHEET_EXCEL_SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": False, "message": "فایل خالی است", "errors": []}

    raw_headers = list(rows[0])
    col_by_key: Dict[str, int] = {}
    pi_cols: Dict[int, int] = {}
    for idx, h in enumerate(raw_headers):
        if h is None:
            continue
        hs = _norm_header_cell(h)
        sk = _resolve_header_key(hs)
        if sk:
            col_by_key[sk] = idx
            continue
        m = _PI_COL_RE.match(hs.strip())
        if m:
            try:
                pi_cols[int(m.group(1))] = idx
            except ValueError:
                pass

    if "product_id" not in col_by_key:
        return {
            "ok": False,
            "message": "ستون product_id (شناسه کالا) در ردیف اول یافت نشد",
            "errors": [],
        }

    parsed_rows: List[BulkProductPriceSheetRow] = []
    errors: List[str] = []
    seen_pid: Dict[int, int] = {}
    data_row_count = 0

    for r_idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        data_row_count += 1
        if data_row_count > _MAX_IMPORT_ROWS:
            errors.append(f"حداکثر {_MAX_IMPORT_ROWS} ردیف داده مجاز است")
            break

        pid_cell = row[col_by_key["product_id"]] if col_by_key["product_id"] < len(row) else None
        if pid_cell is None or str(pid_cell).strip() == "":
            continue
        try:
            if isinstance(pid_cell, float):
                pid = int(pid_cell)
            elif isinstance(pid_cell, int):
                pid = int(pid_cell)
            elif isinstance(pid_cell, Decimal):
                pid = int(pid_cell)
            else:
                pid = int(_translate_digits(str(pid_cell).strip()).replace(",", "").split(".")[0])
        except ValueError:
            errors.append(f"ردیف {r_idx}: شناسه کالا نامعتبر است")
            continue
        if pid in seen_pid:
            errors.append(f"ردیف {r_idx}: شناسه کالا {pid} تکراری است (اولین ردیف {seen_pid[pid]})")
            continue
        seen_pid[pid] = r_idx

        item_dict: Dict[str, Any] = {"product_id": pid, "price_item_updates": []}
        any_action = False

        def col_val(key: str) -> Any:
            j = col_by_key.get(key)
            if j is None or j >= len(row):
                return None
            return row[j]

        bs = _parse_excel_price_cell(col_val("base_sales_price"))
        if bs is not None:
            item_dict["base_sales_price"] = bs
            any_action = True
        bp = _parse_excel_price_cell(col_val("base_purchase_price"))
        if bp is not None:
            item_dict["base_purchase_price"] = bp
            any_action = True

        updates = []
        for pi_id, cidx in sorted(pi_cols.items()):
            if cidx >= len(row):
                continue
            pv = _parse_excel_price_cell(row[cidx])
            if pv is not None:
                updates.append({"price_item_id": pi_id, "price": pv})
                any_action = True
        item_dict["price_item_updates"] = updates

        if not any_action:
            continue

        try:
            parsed_rows.append(BulkProductPriceSheetRow.model_validate(item_dict))
        except Exception as e:
            errors.append(f"ردیف {r_idx}: {e}")

    if errors and not parsed_rows:
        return {"ok": False, "message": "خطا در ایمپورت", "errors": errors}

    if not parsed_rows:
        return {
            "ok": False,
            "message": "هیچ ردیفی با قیمت قابل اعمال یافت نشد؛ ستون‌ها را با خروجی همین صفحه مقایسه کنید",
            "errors": errors,
        }

    total_updated = 0
    total_pi = 0
    batch_errors: List[str] = list(errors)
    chunk_sz = 500
    for i in range(0, len(parsed_rows), chunk_sz):
        chunk = parsed_rows[i : i + chunk_sz]
        req = BulkProductPriceSheetApplyRequest(items=chunk)
        res = apply_bulk_product_price_sheet(db, business_id, req)
        total_updated += int(res.get("updated_count") or 0)
        total_pi += int(res.get("updated_price_items") or 0)
        batch_errors.extend(str(x) for x in (res.get("errors") or []) if x)

    msg = f"{total_updated} کالا به‌روز شد"
    if total_pi:
        msg += f" ({total_pi} ردیف لیست قیمت)"
    return {
        "ok": True,
        "message": msg,
        "updated_count": total_updated,
        "updated_price_items": total_pi,
        "processed_rows": len(parsed_rows),
        "errors": batch_errors,
    }
