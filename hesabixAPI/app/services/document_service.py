"""
سرویس مدیریت اسناد حسابداری عمومی (General Accounting Documents)

این سرویس برای مدیریت تمام اسناد حسابداری (عمومی و اتوماتیک) استفاده می‌شود.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional
from datetime import datetime, date
import logging

from sqlalchemy.orm import Session

from adapters.db.repositories.document_repository import DocumentRepository
from app.core.responses import ApiError
from app.core.cache import get_cache
from app.services.document_monetization_service import process_document_usage_for_document

logger = logging.getLogger(__name__)

RECEIVED_LOAN_FACILITY_DOC_SOURCE = "received_loan_facility"
RECEIVED_LOAN_FACILITY_DOC_TYPE = "received_loan_facility"


def _is_received_loan_facility_manual_document(document: Any) -> bool:
    ex = getattr(document, "extra_info", None)
    if not isinstance(ex, dict):
        ex = {}
    return (
        ex.get("source") == RECEIVED_LOAN_FACILITY_DOC_SOURCE
        or getattr(document, "document_type", None) == RECEIVED_LOAN_FACILITY_DOC_TYPE
    )


def _assert_not_received_loan_facility_document(document: Any, *, action: str) -> None:
    if not _is_received_loan_facility_manual_document(document):
        return
    verb = "حذف" if action == "delete" else "ویرایش"
    raise ApiError(
        "RECEIVED_LOAN_DOCUMENT_SOURCE_ONLY",
        f"اسناد تسهیلات فقط از بخش تسهیلات و اقساط قابل {verb} هستند",
        http_status=409,
    )


def _assert_received_loan_manual_fiscal_year_editable(db: Session, document: Any) -> None:
    """اسناد تسهیلات دستی مانند اسناد فاکتور فقط در سال مالی آخر صف قابل حذف/ویرایش اند."""
    if not _is_received_loan_facility_manual_document(document):
        return
    from adapters.db.models.fiscal_year import FiscalYear

    fy = db.query(FiscalYear).filter(FiscalYear.id == document.fiscal_year_id).first()
    if fy is None:
        return
    if getattr(fy, "is_last", False) is not True:
        raise ApiError(
            "FISCAL_YEAR_LOCKED",
            "سند تسهیلات متعلق به سال مالی جاری نیست و قابل حذف یا ویرایش نمی‌باشد",
            http_status=409,
        )


def invalidate_documents_cache(business_id: int, fiscal_year_id: Optional[int] = None, document_id: Optional[int] = None, document_type: Optional[str] = None):
	"""
	حذف تمام کش‌های مربوط به لیست اسناد عمومی یک کسب‌وکار
	
	این تابع از چند روش استفاده می‌کند:
	1. Tag-based invalidation با set ردیس: حذف انتخابی بر اساس business_id, fiscal_year_id و document_type (بهینه‌تر)
	2. Pattern-based invalidation: حذف تمام کلیدهای documents_list:* (fallback برای اطمینان)
	3. Redis Pub/Sub: انتشار پیام invalidation برای تمام instanceها
	
	Args:
		business_id: شناسه کسب‌وکار
		fiscal_year_id: شناسه سال مالی (اختیاری)
			- اگر None باشد، تمام کش‌های مربوط به business_id حذف می‌شوند
			- اگر مشخص باشد، فقط کش‌های مربوط به آن fiscal_year_id حذف می‌شوند
		document_id: شناسه سند خاص (اختیاری)
		document_type: نوع سند (expense, income, receipt, payment, ...) (اختیاری)
	"""
	cache = get_cache()
	if not cache.enabled:
		return
	
	try:
		# روش 1: استفاده از invalidate_documents_by_business (بهینه‌ترین روش)
		deleted_count = cache.invalidate_documents_by_business(business_id, fiscal_year_id, document_id, document_type)
		if deleted_count > 0:
			logger.info(f"Invalidated {deleted_count} cache keys for business_id {business_id}, fiscal_year_id {fiscal_year_id}, document_id {document_id}, document_type {document_type}")
		
		# روش 2: حذف تمام کلیدهای documents_list:* و receipts_payments_list:* (fallback برای اطمینان کامل)
		for pattern in ("documents_list:*", "receipts_payments_list:*"):
			deleted_pattern = cache.delete_pattern(pattern)
			if deleted_pattern > 0:
				logger.info(f"Invalidated {deleted_pattern} cache keys using pattern: {pattern}")
		
		# حذف کش سند خاص اگر مشخص شده باشد
		if document_id:
			document_pattern = f"document:{business_id}:{document_id}*"
			deleted_document = cache.delete_pattern(document_pattern)
			if deleted_document > 0:
				logger.info(f"Invalidated {deleted_document} cache keys for document_id {document_id} using pattern: {document_pattern}")
		
		# روش 3: انتشار پیام invalidation از طریق Redis Pub/Sub
		invalidation_message = {
			"type": "documents_cache_invalidation",
			"business_id": business_id,
			"fiscal_year_id": fiscal_year_id,
			"document_id": document_id,
			"document_type": document_type,
			"timestamp": None
		}
		try:
			import time
			invalidation_message["timestamp"] = time.time()
			cache.publish_invalidation("cache_invalidation", invalidation_message)
			logger.info(f"Published invalidation message for business_id {business_id}, fiscal_year_id {fiscal_year_id}, document_id {document_id}")
		except Exception as pub_error:
			logger.warning(f"Error publishing invalidation message: {pub_error}")
	
	except Exception as e:
		# خطا در invalidate نباید مانع عملیات اصلی شود
		logger.warning(f"Error invalidating documents cache for business_id {business_id}: {e}")


def list_documents(
    db: Session,
    business_id: int,
    query: Dict[str, Any],
) -> Dict[str, Any]:
    """
    دریافت لیست اسناد حسابداری با فیلتر و صفحه‌بندی
    
    Args:
        db: جلسه دیتابیس
        business_id: شناسه کسب‌وکار
        query: فیلترها و تنظیمات:
            - document_type: نوع سند (expense, income, receipt, payment, transfer, manual, ...)
            - fiscal_year_id: شناسه سال مالی
            - from_date: از تاریخ
            - to_date: تا تاریخ
            - currency_id: شناسه ارز
            - is_proforma: پیش‌فاکتور یا قطعی
            - search: عبارت جستجو
            - sort_by: فیلد مرتب‌سازی
            - sort_desc: ترتیب نزولی
            - take: تعداد رکورد در هر صفحه
            - skip: تعداد رکورد صرف‌نظر شده
    
    Returns:
        دیکشنری حاوی items و pagination
    """
    repo = DocumentRepository(db)
    
    # دریافت لیست اسناد
    documents, total = repo.list_documents_with_filters(business_id, query)
    
    # محاسبه pagination
    take = query.get("take", 50)
    skip = query.get("skip", 0)
    page = (skip // take) + 1
    total_pages = (total + take - 1) // take
    
    return {
        "items": documents,
        "pagination": {
            "total": total,
            "page": page,
            "per_page": take,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
        },
    }


def get_document(db: Session, document_id: int) -> Optional[Dict[str, Any]]:
    """
    دریافت جزئیات کامل یک سند شامل سطرهای سند
    
    Args:
        db: جلسه دیتابیس
        document_id: شناسه سند
    
    Returns:
        دیکشنری حاوی اطلاعات کامل سند یا None
    """
    repo = DocumentRepository(db)
    return repo.get_document_details(document_id)


def delete_document(db: Session, document_id: int, *, commit: bool = True) -> bool:
    """
    حذف یک سند حسابداری
    
    توجه: فقط اسناد عمومی (manual) قابل حذف هستند
    
    Args:
        db: جلسه دیتابیس
        document_id: شناسه سند
    
    Returns:
        True در صورت موفقیت، False در غیر این صورت
    
    Raises:
        ApiError: در صورت عدم وجود سند یا عدم امکان حذف
    """
    repo = DocumentRepository(db)
    
    # بررسی وجود سند
    document = repo.get_document(document_id)
    if not document:
        raise ApiError(
            "DOCUMENT_NOT_FOUND",
            "Document not found",
            http_status=404
        )

    _assert_not_received_loan_facility_document(document, action="delete")
    
    # بررسی نوع سند - فقط اسناد manual قابل حذف هستند
    if document.document_type != "manual":
        raise ApiError(
            "CANNOT_DELETE_AUTO_DOCUMENT",
            "Cannot delete automatically generated documents. Please delete from the original source.",
            http_status=400
        )
    
    # بررسی ارتباط با تراکنش‌های کیف پول
    try:
        from app.services.wallet_service import check_document_has_wallet_transactions
        wallet_check = check_document_has_wallet_transactions(db, document_id)
        if wallet_check["has_wallet_transactions"] and wallet_check.get("has_protected_transactions", False):
            raise ApiError(
                "DOCUMENT_HAS_WALLET_TRANSACTIONS",
                wallet_check["message"],
                http_status=409
            )
    except ApiError:
        raise
    except Exception:
        # اگر به هر دلیل نتوانستیم بررسی کنیم، حذف را متوقف نکن (برای backward compatibility)
        pass

    try:
        _assert_received_loan_manual_fiscal_year_editable(db, document)
    except ApiError:
        raise
    except Exception:
        pass

    # دریافت اطلاعات قبل از حذف برای invalidation
    business_id = document.business_id
    fiscal_year_id = document.fiscal_year_id
    document_type = document.document_type
    
    # حذف سند
    success = repo.delete_document(document_id, commit=commit)
    if not success:
        raise ApiError(
            "DELETE_FAILED",
            "Failed to delete document",
            http_status=500
        )
    
    # Invalidate cache بعد از حذف موفق سند
    if commit:
        invalidate_documents_cache(
            business_id=business_id,
            fiscal_year_id=fiscal_year_id,
            document_id=document_id,
            document_type=document_type
        )
    
    return True


def delete_multiple_documents(db: Session, document_ids: List[int]) -> Dict[str, Any]:
    """
    حذف گروهی اسناد حسابداری
    
    فقط اسناد عمومی (manual) حذف می‌شوند، اسناد اتوماتیک نادیده گرفته می‌شوند
    
    Args:
        db: جلسه دیتابیس
        document_ids: لیست شناسه‌های سند
    
    Returns:
        دیکشنری حاوی تعداد حذف شده و خطاها
    """
    repo = DocumentRepository(db)
    
    deleted_count = 0
    errors = []
    skipped_auto = []
    
    for doc_id in document_ids:
        try:
            document = repo.get_document(doc_id)
            if not document:
                errors.append({"id": doc_id, "error": "Document not found"})
                continue

            try:
                _assert_not_received_loan_facility_document(document, action="delete")
            except ApiError as exc:
                err = getattr(exc, "detail", None)
                msg = str(err) if err is not None else str(exc)
                errors.append({"id": doc_id, "error": msg})
                continue
            
            # بررسی نوع سند
            if document.document_type != "manual":
                skipped_auto.append({
                    "id": doc_id,
                    "type": document.document_type,
                    "code": document.code
                })
                continue

            try:
                _assert_received_loan_manual_fiscal_year_editable(db, document)
            except ApiError as exc:
                err = getattr(exc, "detail", None)
                msg = str(err) if err is not None else str(exc)
                errors.append({"id": doc_id, "error": msg})
                continue
            except Exception:
                pass
            
            # حذف سند
            if repo.delete_document(doc_id):
                deleted_count += 1
            else:
                errors.append({"id": doc_id, "error": "Delete failed"})
                
        except Exception as e:
            logger.error(f"Error deleting document {doc_id}: {str(e)}")
            errors.append({"id": doc_id, "error": str(e)})
    
    return {
        "deleted_count": deleted_count,
        "total_requested": len(document_ids),
        "errors": errors,
        "skipped_auto_documents": skipped_auto,
    }


def get_document_types_summary(db: Session, business_id: int) -> Dict[str, Any]:
    """
    دریافت خلاصه آماری انواع اسناد
    
    Args:
        db: جلسه دیتابیس
        business_id: شناسه کسب‌وکار
    
    Returns:
        دیکشنری حاوی آمار هر نوع سند
    """
    from adapters.db.models.document import Document
    from sqlalchemy import func
    
    results = (
        db.query(
            Document.document_type,
            func.count(Document.id).label("count")
        )
        .filter(Document.business_id == business_id)
        .group_by(Document.document_type)
        .all()
    )
    
    summary = {}
    for row in results:
        summary[row.document_type] = row.count
    
    return summary


def export_documents_excel(
    db: Session,
    business_id: int,
    filters: Dict[str, Any],
) -> bytes:
    """
    خروجی Excel لیست اسناد
    
    Args:
        db: جلسه دیتابیس
        business_id: شناسه کسب‌وکار
        filters: فیلترها
    
    Returns:
        محتوای فایل Excel به صورت bytes
    """
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        repo = DocumentRepository(db)
        
        # دریافت تمام اسناد (بدون pagination)
        filters_copy = filters.copy()
        filters_copy["take"] = 10000
        filters_copy["skip"] = 0
        
        documents, _ = repo.list_documents_with_filters(business_id, filters_copy)
        
        # ایجاد Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Documents"
        
        # تنظیم راست‌چین برای فارسی
        ws.sheet_view.rightToLeft = True
        
        # هدر
        headers = [
            "شماره سند",
            "نوع سند",
            "تاریخ سند",
            "سال مالی",
            "ارز",
            "بدهکار",
            "بستانکار",
            "وضعیت",
            "توضیحات",
            "ایجاد کننده",
            "تاریخ ثبت",
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # داده‌ها
        for row_num, doc in enumerate(documents, 2):
            ws.cell(row=row_num, column=1, value=doc.get("code"))
            ws.cell(row=row_num, column=2, value=doc.get("document_type"))
            ws.cell(row=row_num, column=3, value=str(doc.get("document_date")))
            ws.cell(row=row_num, column=4, value=doc.get("fiscal_year_title"))
            ws.cell(row=row_num, column=5, value=doc.get("currency_code"))
            ws.cell(row=row_num, column=6, value=doc.get("total_debit", 0))
            ws.cell(row=row_num, column=7, value=doc.get("total_credit", 0))
            ws.cell(row=row_num, column=8, value="پیش‌فاکتور" if doc.get("is_proforma") else "قطعی")
            ws.cell(row=row_num, column=9, value=doc.get("description", ""))
            ws.cell(row=row_num, column=10, value=doc.get("created_by_name") or "")
            ws.cell(row=row_num, column=11, value=str(doc.get("created_at")))
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 30
        ws.column_dimensions["J"].width = 22
        ws.column_dimensions["K"].width = 20
        
        # ذخیره در حافظه
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
        
    except ImportError:
        raise ApiError(
            "OPENPYXL_NOT_INSTALLED",
            "openpyxl library is not installed",
            http_status=500
        )
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        raise ApiError(
            "EXCEL_GENERATION_FAILED",
            f"Failed to generate Excel file: {str(e)}",
            http_status=500
        )


def export_documents_pdf(
    db: Session,
    business_id: int,
    filters: Dict[str, Any],
) -> bytes:
    """
    خروجی PDF لیست اسناد
    
    Args:
        db: جلسه دیتابیس
        business_id: شناسه کسب‌وکار
        filters: فیلترها
    
    Returns:
        محتوای فایل PDF به صورت bytes
    """
    # TODO: پیاده‌سازی export PDF
    # می‌توان از WeasyPrint یا ReportLab استفاده کرد
    raise ApiError(
        "NOT_IMPLEMENTED",
        "PDF export is not implemented yet",
        http_status=501
    )


def create_manual_document(
    db: Session,
    business_id: int,
    fiscal_year_id: int,
    user_id: int,
    data: Dict[str, Any],
) -> Dict[str, Any]:
    """
    ایجاد سند حسابداری دستی جدید
    
    Args:
        db: جلسه دیتابیس
        business_id: شناسه کسب‌وکار
        fiscal_year_id: شناسه سال مالی
        user_id: شناسه کاربر ایجادکننده
        data: اطلاعات سند و سطرها
    
    Returns:
        دیکشنری حاوی اطلاعات کامل سند ایجاد شده
    
    Raises:
        ApiError: در صورت بروز خطا
    """
    repo = DocumentRepository(db)
    
    # اعتبارسنجی سطرهای سند
    lines_data = data.get("lines", [])
    is_valid, error_msg = repo.validate_document_balance(lines_data)
    if not is_valid:
        raise ApiError(
            "INVALID_DOCUMENT",
            error_msg,
            http_status=400
        )
    
    # تولید کد سند اگر وجود نداشت
    code = data.get("code")
    if not code:
        code = repo.generate_document_code(business_id, "manual")
    
    # تبدیل lines به فرمت مناسب repository
    lines_for_db = []
    for line in lines_data:
        line_dict = {
            "account_id": line.get("account_id"),
            "person_id": line.get("person_id"),
            "product_id": line.get("product_id"),
            "bank_account_id": line.get("bank_account_id"),
            "cash_register_id": line.get("cash_register_id"),
            "petty_cash_id": line.get("petty_cash_id"),
            "check_id": line.get("check_id"),
            "quantity": line.get("quantity"),
            "debit": line.get("debit", 0),
            "credit": line.get("credit", 0),
            "description": line.get("description"),
            "extra_info": line.get("extra_info"),
        }
        lines_for_db.append(line_dict)
    
    # اعتبارسنجی پروژه (اگر ارسال شده باشد)
    project_id = data.get("project_id")
    if project_id:
        from adapters.db.models.project import Project
        from sqlalchemy import and_
        project = db.query(Project).filter(
            and_(Project.id == project_id, Project.business_id == business_id, Project.is_active == True)
        ).first()
        if not project:
            raise ApiError("PROJECT_NOT_FOUND", "پروژه یافت نشد یا غیرفعال است", http_status=404)
    
    # آماده‌سازی داده‌های سند
    document_data = {
        "code": code,
        "business_id": business_id,
        "fiscal_year_id": fiscal_year_id,
        "currency_id": data.get("currency_id"),
        "created_by_user_id": user_id,
        "document_date": data.get("document_date"),
        "document_type": "manual",
        "is_proforma": data.get("is_proforma", False),
        "description": data.get("description"),
        "extra_info": data.get("extra_info"),
        "project_id": project_id,
        "lines": lines_for_db,
    }
    
    try:
        # ایجاد سند
        document = repo.create_document(document_data)
        
        # دریافت جزئیات کامل سند (با روابط)
        result = repo.get_document_details(document.id)

        try:
            process_document_usage_for_document(db, document.id)
        except Exception as monetization_error:
            logger.warning(
                "document_monetization_processing_failed",
                extra={"document_id": document.id, "error": str(monetization_error)},
            )
        
        # Invalidate cache بعد از ایجاد موفق سند
        invalidate_documents_cache(
            business_id=business_id,
            fiscal_year_id=fiscal_year_id,
            document_id=document.id,
            document_type="manual"
        )
        
        result_data = result
        
        # فراخوانی workflow triggers
        try:
            from app.services.workflow.workflow_trigger_service import trigger_document_created
            manual_document_type = document_data.get("document_type", "manual")
            _doc_extra = {}
            if document_data.get("description"):
                _doc_extra["description"] = document_data.get("description")
            trigger_document_created(
                db=db,
                business_id=business_id,
                document_id=document.id,
                document_type=manual_document_type,
                user_id=user_id,
                extra_fields=_doc_extra or None,
            )
        except Exception as e:
            # عدم موفقیت در trigger نباید مانع بازگشت سند شود
            logger.warning(f"Failed to trigger workflows for document {document.id}: {e}")
        
        return result_data
        
    except Exception as e:
        logger.error(f"Error creating manual document: {str(e)}")
        db.rollback()
        raise ApiError(
            "CREATE_DOCUMENT_FAILED",
            f"Failed to create document: {str(e)}",
            http_status=500
        )


def update_manual_document(
    db: Session,
    document_id: int,
    data: Dict[str, Any],
) -> Dict[str, Any]:
    """
    ویرایش سند حسابداری دستی
    
    Args:
        db: جلسه دیتابیس
        document_id: شناسه سند
        data: اطلاعات جدید سند
    
    Returns:
        دیکشنری حاوی اطلاعات کامل سند ویرایش شده
    
    Raises:
        ApiError: در صورت بروز خطا
    """
    repo = DocumentRepository(db)
    
    # بررسی وجود سند
    document = repo.get_document(document_id)
    if not document:
        raise ApiError(
            "DOCUMENT_NOT_FOUND",
            "Document not found",
            http_status=404
        )

    _assert_not_received_loan_facility_document(document, action="edit")

    # بررسی اینکه فقط اسناد manual قابل ویرایش هستند
    if document.document_type != "manual":
        raise ApiError(
            "CANNOT_EDIT_AUTO_DOCUMENT",
            "Cannot edit automatically generated documents. Please edit from the original source.",
            http_status=400
        )

    try:
        _assert_received_loan_manual_fiscal_year_editable(db, document)
    except ApiError:
        raise
    except Exception:
        pass
    
    # اعتبارسنجی پروژه (اگر ارسال شده باشد)
    project_id = data.get("project_id")
    if project_id is not None:  # چک می‌کنیم که None نباشد (می‌تواند 0 یا عدد باشد)
        from adapters.db.models.project import Project
        from sqlalchemy import and_
        project = db.query(Project).filter(
            and_(Project.id == project_id, Project.business_id == document.business_id, Project.is_active == True)
        ).first()
        if not project:
            raise ApiError("PROJECT_NOT_FOUND", "پروژه یافت نشد یا غیرفعال است", http_status=404)
    
    # اگر سطرها ارسال شده، اعتبارسنجی کن
    if "lines" in data and data["lines"] is not None:
        lines_data = data["lines"]
        is_valid, error_msg = repo.validate_document_balance(lines_data)
        if not is_valid:
            raise ApiError(
                "INVALID_DOCUMENT",
                error_msg,
                http_status=400
            )
        
        # تبدیل lines به فرمت مناسب repository
        lines_for_db = []
        for line in lines_data:
            line_dict = {
                "account_id": line.get("account_id"),
                "person_id": line.get("person_id"),
                "product_id": line.get("product_id"),
                "bank_account_id": line.get("bank_account_id"),
                "cash_register_id": line.get("cash_register_id"),
                "petty_cash_id": line.get("petty_cash_id"),
                "check_id": line.get("check_id"),
                "quantity": line.get("quantity"),
                "debit": line.get("debit", 0),
                "credit": line.get("credit", 0),
                "description": line.get("description"),
                "extra_info": line.get("extra_info"),
            }
            lines_for_db.append(line_dict)
        
        data["lines"] = lines_for_db
    
    try:
        # ویرایش سند
        updated_document = repo.update_document(document_id, data)
        
        if not updated_document:
            raise ApiError(
                "UPDATE_FAILED",
                "Failed to update document",
                http_status=500
            )
        
        # Invalidate cache بعد از به‌روزرسانی موفق سند
        invalidate_documents_cache(
            business_id=document.business_id,
            fiscal_year_id=document.fiscal_year_id,
            document_id=document_id,
            document_type=document.document_type
        )
        
        # دریافت جزئیات کامل سند
        details = repo.get_document_details(updated_document.id)

        try:
            from app.services.workflow.workflow_trigger_service import trigger_document_updated

            trigger_document_updated(
                db=db,
                business_id=int(document.business_id),
                document_id=int(updated_document.id),
                document_type="manual",
            )
        except Exception as wfe:
            logger.warning("document.updated workflow trigger failed: %s", wfe, exc_info=True)

        return details
        
    except ApiError:
        raise
    except Exception as e:
        logger.error(f"Error updating manual document: {str(e)}")
        db.rollback()
        raise ApiError(
            "UPDATE_DOCUMENT_FAILED",
            f"Failed to update document: {str(e)}",
            http_status=500
        )

