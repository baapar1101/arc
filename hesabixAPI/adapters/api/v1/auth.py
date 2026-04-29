# Removed __future__ annotations to fix OpenAPI schema generation

import datetime
from fastapi import APIRouter, Depends, Request, Query
from fastapi.responses import Response, HTMLResponse
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from adapters.db.repositories.user_repo import UserRepository
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.services.captcha_service import create_captcha, validate_captcha
from app.services.auth_service import register_user, login_user, create_password_reset, reset_password, change_password, referral_stats
from app.services.email_verification_service import verify_email_token, resend_verification_email
from app.services.pdf import PDFService
from .schemas import (
	RegisterRequest, LoginRequest, ForgotPasswordRequest, ResetPasswordRequest, 
	SendLoginOtpRequest, AvailableChannelsRequest, ChangePasswordRequest, UpdateMobileRequest, UpdateEmailRequest,
	SendMobileVerificationRequest,
	CreateApiKeyRequest, UpdateApiKeyRequest, QueryInfo, FilterItem,
	SuccessResponse, CaptchaResponse, LoginResponse, ApiKeyResponse, 
	ReferralStatsResponse, UserResponse
)
from app.core.settings import get_settings
from app.core.auth_dependency import get_current_user, AuthContext
from app.services.api_key_service import list_personal_keys, create_personal_key, revoke_key
from app.services.session_service import list_user_sessions, revoke_session, revoke_other_sessions
from app.core.rate_limiter import get_client_ip, rate_limit
from app.services.auth_dynamic_rate_limit import enforce_auth_rate_limit


router = APIRouter(prefix="/auth", tags=["احراز هویت"])


def _otp_in_response(otp_code: str | None) -> dict:
	"""فقط در حالت debug — در production مقدار OTP در پاسخ API برگردانده نمی‌شود."""
	if get_settings().debug and (otp_code or ""):
		return {"otp_code": otp_code}
	return {}


@router.post("/captcha", 
	summary="تولید کپچای عددی", 
	description="تولید کپچای عددی برای تأیید هویت در عملیات حساس",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کپچا با موفقیت تولید شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کپچا تولید شد",
						"data": {
							"captcha_id": "abc123def456",
							"image_base64": "iVBORw0KGgoAAAANSUhEUgAA...",
							"ttl_seconds": 180
						}
					}
				}
			}
		}
	}
)
async def generate_captcha(request: Request, db: Session = Depends(get_db)) -> dict:
	enforce_auth_rate_limit(
		request,
		db,
		kind="captcha",
		error_message="تعداد درخواست‌های کپچا بیش از حد مجاز است. لطفاً کمی صبر کنید.",
	)
	captcha_id, image_base64, ttl = create_captcha(db, get_client_ip(request))
	from app.services.system_settings_service import get_captcha_auth_security_effective
	_csec = get_captcha_auth_security_effective(db)
	return success_response({
		"captcha_id": captcha_id,
		"image_base64": image_base64,
		"ttl_seconds": ttl,
		"captcha_mode": _csec["captcha_mode"],
		"captcha_length": _csec["captcha_length"],
	})


@router.get("/me", 
	summary="دریافت اطلاعات کاربر کنونی", 
	description="دریافت اطلاعات کامل کاربری که در حال حاضر وارد سیستم شده است",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "اطلاعات کاربر با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "اطلاعات کاربر دریافت شد",
						"data": {
							"id": 1,
							"email": "user@example.com",
							"mobile": "09123456789",
							"first_name": "احمد",
							"last_name": "احمدی",
							"is_active": True,
							"referral_code": "ABC123",
							"referred_by_user_id": None,
							"app_permissions": {"admin": True},
							"created_at": "2024-01-01T00:00:00Z",
							"updated_at": "2024-01-01T00:00:00Z"
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "احراز هویت مورد نیاز است",
						"error_code": "UNAUTHORIZED"
					}
				}
			}
		}
	}
)
def get_current_user_info(
    request: Request,
    ctx: AuthContext = Depends(get_current_user)
) -> dict:
    """دریافت اطلاعات کاربر کنونی"""
    return success_response(ctx.to_dict(), request)


@router.post(
	"/activity",
	summary="ثبت ضربان فعالیت (آخرین فعالیت در اپ)",
	description="""
کلاینت به‌صورت دوره‌ای (مثلاً هر یک دقیقه) این endpoint را با همان کلید ApiKey فراخوانی کند.
سرور فیلد «آخرین فعالیت» را با به‌روزرسانی محدود (throttle پیش‌فرض ~۴۵ ثانیه) در پایگاه داده ذخیره می‌کند.
بدون نیاز به Redis؛ برای نمایش «کاربران اخیراً فعال» در مدیریت سیستم قابل استفاده است.
""",
)
def post_user_activity(
	request: Request,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db),
):
	user_id = ctx.get_user_id()
	if not user_id:
		raise ApiError("UNAUTHORIZED", "احراز هویت الزامی است", http_status=401)
	repo = UserRepository(db)
	ts = repo.touch_last_activity(user_id)
	if ts is None:
		raise ApiError("USER_NOT_FOUND", "کاربر یافت نشد", http_status=404)
	payload = format_datetime_fields({"last_activity_at": ts}, request)
	return success_response(payload, request, message=None)


@router.post("/register", 
	summary="ثبت‌نام کاربر جدید", 
	description="ثبت‌نام کاربر جدید در سیستم با تأیید کپچا",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کاربر با موفقیت ثبت‌نام شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ثبت‌نام با موفقیت انجام شد",
						"data": {
							"api_key": "sk_1234567890abcdef",
							"expires_at": None,
							"user": {
								"id": 1,
								"first_name": "احمد",
								"last_name": "احمدی",
								"email": "ahmad@example.com",
								"mobile": "09123456789",
								"referral_code": "ABC123",
								"app_permissions": None
							}
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		409: {
			"description": "کاربر با این ایمیل یا موبایل قبلاً ثبت‌نام کرده است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کاربر با این ایمیل قبلاً ثبت‌نام کرده است",
						"error_code": "USER_EXISTS"
					}
				}
			}
		}
	}
)
async def register(request: Request, payload: RegisterRequest, db: Session = Depends(get_db)) -> dict:
	enforce_auth_rate_limit(
		request,
		db,
		kind="register",
		error_message="تعداد درخواست‌های ثبت‌نام بیش از حد مجاز است. لطفاً بعداً تلاش کنید.",
	)
	import logging
	logger = logging.getLogger(__name__)
	# ساخت base_url از request برای verification email
	base_url = None
	if request.headers.get("X-Forwarded-Host"):
		proto = request.headers.get("X-Forwarded-Proto", "https")
		host = request.headers.get("X-Forwarded-Host")
		base_url = f"{proto}://{host}"
	elif request.url:
		base_url = str(request.url).replace(request.url.path, "").rstrip("/")
	try:
		user_id = register_user(
			db=db,
			first_name=payload.first_name,
			last_name=payload.last_name,
			email=payload.email,
			mobile=payload.mobile,
			password=payload.password,
			captcha_id=payload.captcha_id,
			captcha_code=payload.captcha_code,
			referrer_code=payload.referrer_code,
			base_url=base_url,
			client_ip=get_client_ip(request),
		)
		# Create a session api key similar to login
		user_agent = request.headers.get("User-Agent")
		ip = request.client.host if request.client else None
		from app.core.security import generate_api_key
		from adapters.db.repositories.api_key_repo import ApiKeyRepository
		api_key, key_hash = generate_api_key()
		api_repo = ApiKeyRepository(db)
		api_repo.create_session_key(user_id=user_id, key_hash=key_hash, device_id=payload.device_id, user_agent=user_agent, ip=ip, expires_at=None)
		from adapters.db.models.user import User
		user_obj = db.get(User, user_id)
		user = {
			"id": user_id,
			"first_name": payload.first_name,
			"last_name": payload.last_name,
			"email": payload.email,
			"mobile": payload.mobile,
			"referral_code": getattr(user_obj, "referral_code", None),
			"app_permissions": getattr(user_obj, "app_permissions", None),
			"email_verified": getattr(user_obj, "email_verified", False)
		}
		response_data = {"api_key": api_key, "expires_at": None, "user": user}
		formatted_data = format_datetime_fields(response_data, request)
		return success_response(formatted_data, request)
	except Exception as exc:
		from app.core.responses import ApiError
		if isinstance(exc, ApiError):
			raise
		logger.error(
			"POST /api/v1/auth/register failed: %s: %s",
			type(exc).__name__,
			str(exc),
			exc_info=True,
			extra={"path": request.url.path, "method": request.method},
		)
		raise


@router.post("/login", 
	summary="ورود با ایمیل یا موبایل", 
	description="ورود کاربر به سیستم با استفاده از ایمیل یا شماره موبایل و رمز عبور",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "ورود با موفقیت انجام شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ورود با موفقیت انجام شد",
						"data": {
							"api_key": "sk_1234567890abcdef",
							"expires_at": "2024-01-02T00:00:00Z",
							"user": {
								"id": 1,
								"first_name": "احمد",
								"last_name": "احمدی",
								"email": "ahmad@example.com",
								"mobile": "09123456789",
								"referral_code": "ABC123",
								"app_permissions": {"admin": True}
							}
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		401: {
			"description": "اطلاعات ورود نامعتبر است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "ایمیل یا رمز عبور اشتباه است",
						"error_code": "INVALID_CREDENTIALS"
					}
				}
			}
		}
	}
)
async def login(request: Request, payload: LoginRequest, db: Session = Depends(get_db)) -> dict:
	enforce_auth_rate_limit(
		request,
		db,
		kind="login_short",
		error_message="تعداد تلاش‌های ورود بیش از حد مجاز است. لطفاً کمی صبر کنید.",
	)
	enforce_auth_rate_limit(
		request,
		db,
		kind="login_long",
		error_message="تعداد درخواست‌های ورود بیش از حد مجاز است. لطفاً کمی صبر کنید.",
	)
	user_agent = request.headers.get("User-Agent")
	ip = get_client_ip(request)
	api_key, expires_at, user = login_user(
		db=db,
		identifier=payload.identifier,
		password=payload.password,
		captcha_id=payload.captcha_id,
		captcha_code=payload.captcha_code,
		device_id=payload.device_id,
		user_agent=user_agent,
		ip=ip,
	)
	# Ensure referral_code is included
	from adapters.db.repositories.user_repo import UserRepository
	repo = UserRepository(db)
	from adapters.db.models.user import User
	user_obj = None
	if 'id' in user and user['id']:
		user_obj = repo.db.get(User, user['id'])
	if user_obj is not None:
		user["referral_code"] = getattr(user_obj, "referral_code", None)
	response_data = {"api_key": api_key, "expires_at": expires_at, "user": user}
	formatted_data = format_datetime_fields(response_data, request)
	return success_response(formatted_data, request)


@router.post("/forgot-password", 
	summary="ایجاد توکن بازنشانی رمز عبور", 
	description="ایجاد توکن برای بازنشانی رمز عبور کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "توکن بازنشانی با موفقیت ایجاد شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "توکن بازنشانی ارسال شد",
						"data": {
							"ok": True,
							"token": "reset_token_1234567890abcdef"
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		404: {
			"description": "کاربر یافت نشد",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کاربر با این ایمیل یا موبایل یافت نشد",
						"error_code": "USER_NOT_FOUND"
					}
				}
			}
		}
	}
)
async def forgot_password(request: Request, payload: ForgotPasswordRequest, db: Session = Depends(get_db)) -> dict:
	enforce_auth_rate_limit(
		request,
		db,
		kind="forgot",
		error_message="تعداد درخواست‌های بازیابی رمز عبور بیش از حد مجاز است. لطفاً بعداً تلاش کنید.",
	)
	# ایجاد token برای reset password
	token = create_password_reset(
		db=db,
		identifier=payload.identifier,
		captcha_id=payload.captcha_id,
		captcha_code=payload.captcha_code,
		client_ip=get_client_ip(request),
	)
	# Send notification via preferred channels
	if token:
		from adapters.db.repositories.user_repo import UserRepository
		from app.services.notification_service import NotificationService
		from app.services.auth_service import _detect_identifier
		import logging
		logger = logging.getLogger(__name__)
		
		# تشخیص نوع identifier و جستجوی صحیح کاربر
		kind, email, mobile = _detect_identifier(payload.identifier)
		if kind != "invalid":
			repo = UserRepository(db)
			user = repo.get_by_email(email) if email else repo.get_by_mobile(mobile)  # type: ignore[arg-type]
			if user:
				try:
					svc = NotificationService(db)
					svc.send(user_id=user.id, event_key="auth.password_reset", context={"token": token})
				except Exception as e:
					# در صورت خطا در ارسال notification، log می‌کنیم اما فرآیند ادامه می‌یابد
					logger.error(f"Failed to send password reset notification for user {user.id}: {e}")
	# همیشه پاسخ موفق برمی‌گردانیم (برای جلوگیری از user enumeration)
	# در production نباید token برگردانده شود
	return success_response({"ok": True})


@router.post("/reset-password", 
	summary="بازنشانی رمز عبور با توکن", 
	description="بازنشانی رمز عبور کاربر با استفاده از توکن دریافتی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "رمز عبور با موفقیت بازنشانی شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "رمز عبور با موفقیت تغییر کرد",
						"data": {
							"ok": True
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کپچا نامعتبر است",
						"error_code": "INVALID_CAPTCHA"
					}
				}
			}
		},
		404: {
			"description": "توکن نامعتبر یا منقضی شده است",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "توکن نامعتبر یا منقضی شده است",
						"error_code": "INVALID_TOKEN"
					}
				}
			}
		}
	}
)
async def reset_password_endpoint(request: Request, payload: ResetPasswordRequest, db: Session = Depends(get_db)) -> dict:
	enforce_auth_rate_limit(
		request,
		db,
		kind="reset",
		error_message="تعداد درخواست‌های بازنشانی رمز عبور بیش از حد مجاز است. لطفاً بعداً تلاش کنید.",
	)
	reset_password(
		db=db,
		token=payload.token,
		new_password=payload.new_password,
		captcha_id=payload.captcha_id,
		captcha_code=payload.captcha_code,
		client_ip=get_client_ip(request),
	)
	return success_response({"ok": True})


@router.post(
	"/password-reset/send-otp",
	summary="ارسال OTP برای بازیابی رمز عبور",
	description="ارسال کد OTP 6 رقمی به شماره موبایل برای بازیابی رمز عبور",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کد OTP با موفقیت ارسال شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کد بازیابی رمز عبور به شماره موبایل ارسال شد"
					}
				}
			}
		},
		400: {
			"description": "شناسه نامعتبر یا کاربر موبایل ندارد",
		},
		429: {
			"description": "تعداد درخواست‌ها بیش از حد مجاز",
		},
		503: {
			"description": "سرویس پیامک پیکربندی نشده است",
		}
	}
)
def send_password_reset_otp(
	request: Request,
	payload: ForgotPasswordRequest,
	db: Session = Depends(get_db)
) -> dict:
	"""ارسال OTP برای بازیابی رمز عبور"""
	enforce_auth_rate_limit(
		request,
		db,
		kind="pr_otp",
		error_message="تعداد درخواست‌های بازیابی رمز عبور بیش از حد مجاز است. لطفاً چند دقیقه بعد دوباره تلاش کنید.",
	)
	from app.services.password_reset_otp_service import PasswordResetOtpService
	from app.services.captcha_service import validate_captcha
	from app.core.responses import ApiError
	
	# تایید کپچا
	if not validate_captcha(db, payload.captcha_id, payload.captcha_code, client_ip=get_client_ip(request)):
		raise ApiError("INVALID_CAPTCHA", "Invalid captcha code")
	
	service = PasswordResetOtpService(db)
	
	try:
		otp_code = service.send_reset_otp(payload.identifier)
		data = {
			"ok": True,
			"message": "کد بازیابی رمز عبور به شماره موبایل ارسال شد",
		}
		data.update(_otp_in_response(otp_code))
		return success_response(data, request)
	except ApiError as e:
		raise e
	except Exception as e:
		raise ApiError("SMS_SEND_FAILED", f"خطا در ارسال پیامک: {str(e)}", http_status=500)


@router.post(
	"/password-reset/verify-otp",
	summary="تایید OTP بازیابی رمز عبور",
	description="تایید کد OTP و دریافت token برای تغییر رمز عبور",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "OTP تایید شد و token دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"data": {
							"reset_token": "token_here"
						}
					}
				}
			}
		},
		400: {
			"description": "کد OTP نامعتبر",
		}
	}
)
def verify_password_reset_otp(
	request: Request,
	identifier: str = Query(..., description="ایمیل یا شماره موبایل"),
	otp_code: str = Query(..., description="کد OTP 6 رقمی"),
	db: Session = Depends(get_db)
) -> dict:
	"""تایید OTP و دریافت reset token"""
	from app.services.password_reset_otp_service import PasswordResetOtpService
	
	service = PasswordResetOtpService(db)
	
	success, reset_token = service.verify_reset_otp(identifier, otp_code)
	
	if not success:
		from app.core.responses import ApiError
		raise ApiError("OTP_VERIFICATION_FAILED", "تایید OTP ناموفق بود", http_status=400)
	
	return success_response({
		"reset_token": reset_token,
		"message": "OTP تایید شد. اکنون می‌توانید رمز عبور جدید تنظیم کنید"
	}, request)


@router.get("/api-keys", 
	summary="لیست کلیدهای API شخصی", 
	description="دریافت لیست کلیدهای API شخصی کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "لیست کلیدهای API با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "لیست کلیدهای API دریافت شد",
						"data": [
							{
								"id": 1,
								"name": "کلید اصلی",
								"scopes": "read,write",
								"device_id": "device123",
								"user_agent": "Mozilla/5.0...",
								"ip": "192.168.1.1",
								"expires_at": None,
								"last_used_at": "2024-01-01T12:00:00Z",
								"created_at": "2024-01-01T00:00:00Z"
							}
						]
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def list_keys(request: Request, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	items = list_personal_keys(db, ctx.user.id)
	return success_response(items)


@router.post("/api-keys", 
	summary="ایجاد کلید API شخصی", 
	description="ایجاد کلید API جدید برای کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کلید API با موفقیت ایجاد شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کلید API ایجاد شد",
						"data": {
							"id": 1,
							"api_key": "sk_1234567890abcdef"
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def create_key(request: Request, payload: CreateApiKeyRequest, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	from datetime import datetime
	
	expires_at = None
	if payload.expires_at:
		try:
			# پارس کردن تاریخ ISO format (مثال: 2024-12-31T23:59:59)
			date_str = payload.expires_at.strip()
			# تبدیل Z به +00:00 برای timezone
			if date_str.endswith('Z'):
				date_str = date_str[:-1] + '+00:00'
			expires_at = datetime.fromisoformat(date_str)
		except (ValueError, AttributeError) as e:
			from app.core.responses import ApiError
			raise ApiError("INVALID_INPUT", f"Invalid expires_at format. Use ISO format: YYYY-MM-DDTHH:MM:SS. Error: {str(e)}", http_status=400)
	
	id_, api_key = create_personal_key(
		db, 
		ctx.user.id, 
		payload.name, 
		payload.scopes, 
		expires_at,
		payload.ip_whitelist
	)
	return success_response({"id": id_, "api_key": api_key, "message": "کلید API با موفقیت ایجاد شد. لطفاً آن را ذخیره کنید زیرا فقط یک بار نمایش داده می‌شود."})


@router.post("/change-password", 
	summary="تغییر رمز عبور", 
	description="تغییر رمز عبور کاربر با تأیید رمز عبور فعلی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "رمز عبور با موفقیت تغییر کرد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "رمز عبور با موفقیت تغییر کرد",
						"data": {
							"ok": True
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی داده‌ها",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "رمز عبور فعلی اشتباه است",
						"error_code": "INVALID_CURRENT_PASSWORD"
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def change_password_endpoint(request: Request, payload: ChangePasswordRequest, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	# دریافت translator از request state
	translator = getattr(request.state, "translator", None)
	
	change_password(
		db=db,
		user_id=ctx.user.id,
		current_password=payload.current_password,
		new_password=payload.new_password,
		confirm_password=payload.confirm_password,
		translator=translator
	)
	return success_response({"ok": True})


@router.delete("/api-keys/{key_id}", 
	summary="حذف کلید API", 
	description="حذف کلید API مشخص شده",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کلید API با موفقیت حذف شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کلید API حذف شد",
						"data": {
							"ok": True
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		},
		404: {
			"description": "کلید API یافت نشد",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"message": "کلید API یافت نشد",
						"error_code": "API_KEY_NOT_FOUND"
					}
				}
			}
		}
	}
)
def delete_key(request: Request, key_id: int, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	revoke_key(db, ctx.user.id, key_id)
	return success_response({"ok": True})


@router.get("/api-keys/{key_id}",
	summary="دریافت جزئیات کلید API",
	description="دریافت اطلاعات کامل یک کلید API",
	response_model=SuccessResponse,
)
def get_key(request: Request, key_id: int, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	from adapters.db.models.api_key import ApiKey
	from app.core.responses import ApiError
	obj = db.get(ApiKey, key_id)
	if not obj or obj.user_id != ctx.user.id or obj.key_type != "personal":
		raise ApiError("NOT_FOUND", "Key not found", http_status=404)
	
	from datetime import datetime
	data = {
		"id": obj.id,
		"name": obj.name,
		"scopes": obj.scopes,
		"ip": obj.ip,
		"user_agent": obj.user_agent,
		"created_at": obj.created_at.isoformat() if obj.created_at else None,
		"expires_at": obj.expires_at.isoformat() if obj.expires_at else None,
		"last_used_at": obj.last_used_at.isoformat() if obj.last_used_at else None,
		"revoked_at": obj.revoked_at.isoformat() if obj.revoked_at else None,
		"is_active": obj.revoked_at is None and (obj.expires_at is None or obj.expires_at > datetime.utcnow()),
	}
	return success_response(data)


@router.put("/api-keys/{key_id}",
	summary="ویرایش کلید API",
	description="ویرایش اطلاعات یک کلید API (نام، محدوده دسترسی، تاریخ انقضا، IP whitelist)",
	response_model=SuccessResponse,
)
def update_key(request: Request, key_id: int, payload: UpdateApiKeyRequest, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
	from datetime import datetime
	from app.services.api_key_service import update_api_key
	
	expires_at = None
	if payload.expires_at is not None:
		if payload.expires_at == "":
			expires_at = None
		else:
			try:
				# پارس کردن تاریخ ISO format
				date_str = payload.expires_at.strip()
				# تبدیل Z به +00:00 برای timezone
				if date_str.endswith('Z'):
					date_str = date_str[:-1] + '+00:00'
				expires_at = datetime.fromisoformat(date_str)
			except (ValueError, AttributeError) as e:
				from app.core.responses import ApiError
				raise ApiError("INVALID_INPUT", f"Invalid expires_at format. Use ISO format: YYYY-MM-DDTHH:MM:SS. Error: {str(e)}", http_status=400)
	
	update_api_key(
		db,
		ctx.user.id,
		key_id,
		payload.name,
		payload.scopes,
		expires_at,
		payload.ip_whitelist
	)
	return success_response({"ok": True, "message": "کلید API با موفقیت به‌روزرسانی شد"})


@router.get("/sessions",
	summary="لیست سشن‌های ورود",
	description="دریافت لیست تمام session های فعال کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "لیست session ها با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"data": [
							{
								"id": 123,
								"device_name": "Chrome on Windows",
								"device_id": "device-uuid-123",
								"user_agent": "Mozilla/5.0...",
								"ip": "192.168.1.100",
								"is_current": True,
								"created_at": "2024-01-15T10:30:00Z",
								"last_used_at": "2024-01-20T14:25:00Z",
								"last_used_relative": "2 ساعت پیش",
								"browser": "Chrome",
								"os": "Windows",
								"device_type": "desktop"
							}
						]
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def list_sessions(
	request: Request,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""لیست تمام session های فعال کاربر"""
	# دریافت hash کلید API فعلی
	from adapters.db.models.api_key import ApiKey
	current_key = db.get(ApiKey, ctx.api_key_id)
	if not current_key:
		from app.core.responses import ApiError
		raise ApiError("INTERNAL_ERROR", "API key not found", http_status=500)
	
	current_key_hash = current_key.key_hash
	sessions = list_user_sessions(db, ctx.user.id, current_key_hash)
	return success_response(sessions, request)


@router.delete("/sessions/others",
	summary="حذف همه سشن‌های دیگر",
	description="حذف تمام session های کاربر به جز session فعلی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "تمام session های دیگر حذف شدند",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "تمام سشن‌های دیگر حذف شدند",
						"data": {"deleted_count": 5}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def revoke_other_sessions_endpoint(
	request: Request,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""حذف تمام session های دیگر (به جز فعلی)"""
	# دریافت hash کلید API فعلی
	from adapters.db.models.api_key import ApiKey
	current_key = db.get(ApiKey, ctx.api_key_id)
	if not current_key:
		from app.core.responses import ApiError
		raise ApiError("INTERNAL_ERROR", "API key not found", http_status=500)
	
	current_key_hash = current_key.key_hash
	deleted_count = revoke_other_sessions(db, ctx.user.id, current_key_hash)
	return success_response(
		{"deleted_count": deleted_count, "message": f"{deleted_count} سشن حذف شد"},
		request
	)


@router.delete("/sessions/{session_id}",
	summary="حذف سشن ورود",
	description="حذف یک session خاص. کاربر نمی‌تواند session فعلی را حذف کند.",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "Session با موفقیت حذف شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "سشن با موفقیت حذف شد",
						"data": {"ok": True}
					}
				}
			}
		},
		400: {
			"description": "نمی‌توانید session فعلی را حذف کنید"
		},
		404: {
			"description": "Session یافت نشد"
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def revoke_session_endpoint(
	session_id: int,
	request: Request,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""حذف یک session"""
	# دریافت hash کلید API فعلی
	from adapters.db.models.api_key import ApiKey
	current_key = db.get(ApiKey, ctx.api_key_id)
	if not current_key:
		from app.core.responses import ApiError
		raise ApiError("INTERNAL_ERROR", "API key not found", http_status=500)
	
	current_key_hash = current_key.key_hash
	revoke_session(db, ctx.user.id, session_id, current_key_hash)
	return success_response({"ok": True, "message": "سشن با موفقیت حذف شد"}, request)


@router.get("/referrals/stats", 
	summary="آمار معرفی‌ها", 
	description="دریافت آمار معرفی‌های کاربر فعلی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "آمار معرفی‌ها با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "آمار معرفی‌ها دریافت شد",
						"data": {
							"total_referrals": 25,
							"active_referrals": 20,
							"recent_referrals": 5,
							"referral_rate": 0.8
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def get_referral_stats(request: Request, ctx: AuthContext = Depends(get_current_user), db: Session = Depends(get_db), start: str = Query(None, description="تاریخ شروع (ISO format)"), end: str = Query(None, description="تاریخ پایان (ISO format)")):
	from datetime import datetime
	start_dt = datetime.fromisoformat(start) if start else None
	end_dt = datetime.fromisoformat(end) if end else None
	stats = referral_stats(db=db, user_id=ctx.user.id, start=start_dt, end=end_dt)
	return success_response(stats)


@router.post("/referrals/list", 
	summary="لیست معرفی‌ها با فیلتر پیشرفته", 
	description="دریافت لیست معرفی‌ها با قابلیت فیلتر، جستجو، مرتب‌سازی و صفحه‌بندی",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "لیست معرفی‌ها با موفقیت دریافت شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "لیست معرفی‌ها دریافت شد",
						"data": {
							"items": [
								{
									"id": 1,
									"first_name": "علی",
									"last_name": "احمدی",
									"email": "ali@example.com",
									"mobile": "09123456789",
									"created_at": "2024-01-01T00:00:00Z"
								}
							],
							"total": 1,
							"page": 1,
							"limit": 10,
							"total_pages": 1,
							"has_next": False,
							"has_prev": False
						}
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def get_referral_list_advanced(
	request: Request,
	query_info: QueryInfo,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""
	دریافت لیست معرفی‌ها با قابلیت فیلتر پیشرفته
	
	پارامترهای QueryInfo:
	- sort_by: فیلد مرتب‌سازی (مثال: created_at, first_name, last_name, email)
	- sort_desc: ترتیب نزولی (true/false)
	- take: تعداد رکورد در هر صفحه (پیش‌فرض: 10)
	- skip: تعداد رکورد صرف‌نظر شده (پیش‌فرض: 0)
	- search: عبارت جستجو
	- search_fields: فیلدهای جستجو (مثال: ["first_name", "last_name", "email"])
	- filters: آرایه فیلترها با ساختار:
	  [
		{
		  "property": "created_at",
		  "operator": ">=",
		  "value": "2024-01-01T00:00:00"
		},
		{
		  "property": "first_name", 
		  "operator": "*",
		  "value": "احمد"
		}
	  ]
	"""
	from adapters.db.repositories.user_repo import UserRepository
	from adapters.db.models.user import User
	from datetime import datetime
	
	# Create a custom query for referrals
	repo = UserRepository(db)
	
	# Add filter for referrals only (users with referred_by_user_id = current user)
	referral_filter = FilterItem(
		property="referred_by_user_id",
		operator="=",
		value=ctx.user.id
	)
	
	# Add referral filter to existing filters
	if query_info.filters is None:
		query_info.filters = [referral_filter]
	else:
		query_info.filters.append(referral_filter)
	
	# Set default search fields for referrals
	if query_info.search_fields is None:
		query_info.search_fields = ["first_name", "last_name", "email"]
	
	# Execute query with filters
	referrals, total = repo.query_with_filters(query_info)
	
	# Convert to dictionary format
	referral_dicts = [repo.to_dict(referral) for referral in referrals]
	
	# Format datetime fields
	formatted_referrals = format_datetime_fields(referral_dicts, request)
	
	# Calculate pagination info
	page = (query_info.skip // query_info.take) + 1
	total_pages = (total + query_info.take - 1) // query_info.take
	
	return success_response({
		"items": formatted_referrals,
		"total": total,
		"page": page,
		"limit": query_info.take,
		"total_pages": total_pages,
		"has_next": page < total_pages,
		"has_prev": page > 1
	}, request)


@router.post("/referrals/export/pdf", 
	summary="خروجی PDF لیست معرفی‌ها", 
	description="خروجی PDF لیست معرفی‌ها با قابلیت فیلتر و انتخاب سطرهای خاص",
	responses={
		200: {
			"description": "فایل PDF با موفقیت تولید شد",
			"content": {
				"application/pdf": {
					"schema": {
						"type": "string",
						"format": "binary"
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def export_referrals_pdf(
	request: Request,
	query_info: QueryInfo,
	selected_only: bool = False,
	selected_indices: str | None = None,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> Response:
	"""
	خروجی PDF لیست معرفی‌ها
	
	پارامترها:
	- selected_only: آیا فقط سطرهای انتخاب شده export شوند
	- selected_indices: لیست ایندکس‌های انتخاب شده (JSON string)
	- سایر پارامترهای QueryInfo برای فیلتر
	"""
	from app.services.pdf import PDFService
	from app.services.auth_service import referral_stats
	import json
	
	# Parse selected indices if provided
	indices = None
	if selected_only and selected_indices:
		try:
			indices = json.loads(selected_indices)
		except (json.JSONDecodeError, TypeError):
			indices = None
	
	# Get stats for the report
	stats = None
	try:
		# Extract date range from filters if available
		start_date = None
		end_date = None
		if query_info.filters:
			for filter_item in query_info.filters:
				if filter_item.property == 'created_at':
					if filter_item.operator == '>=':
						start_date = filter_item.value
					elif filter_item.operator == '<':
						end_date = filter_item.value
		
		stats = referral_stats(
			db=db,
			user_id=ctx.user.id,
			start=start_date,
			end=end_date
		)
	except Exception:
		pass  # Continue without stats
	
	# Get calendar type from request headers
	calendar_header = request.headers.get("X-Calendar-Type", "jalali")
	calendar_type = "jalali" if calendar_header.lower() in ["jalali", "persian", "shamsi"] else "gregorian"
	
	# Generate PDF using new modular service
	pdf_service = PDFService()
	
	# Get locale from request headers
	locale_header = request.headers.get("Accept-Language", "fa")
	locale = "fa" if locale_header.startswith("fa") else "en"
	
	pdf_bytes = pdf_service.generate_pdf(
		module_name='marketing',
		data={},  # Empty data - module will fetch its own data
		calendar_type=calendar_type,
		locale=locale,
		db=db,
		user_id=ctx.user.id,
		query_info=query_info,
		selected_indices=indices,
		stats=stats
	)
	
	# Return PDF response
	from fastapi.responses import Response
	import datetime
	
	filename = f"referrals_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
	
	return Response(
		content=pdf_bytes,
		media_type="application/pdf",
		headers={
			"Content-Disposition": f"attachment; filename={filename}",
			"Content-Length": str(len(pdf_bytes))
		}
	)


@router.post("/referrals/export/excel", 
	summary="خروجی Excel لیست معرفی‌ها", 
	description="خروجی Excel لیست معرفی‌ها با قابلیت فیلتر و انتخاب سطرهای خاص",
	responses={
		200: {
			"description": "فایل Excel با موفقیت تولید شد",
			"content": {
				"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
					"schema": {
						"type": "string",
						"format": "binary"
					}
				}
			}
		},
		401: {
			"description": "کاربر احراز هویت نشده است"
		}
	}
)
def export_referrals_excel(
	request: Request,
	query_info: QueryInfo,
	selected_only: bool = False,
	selected_indices: str | None = None,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> Response:
	"""
	خروجی Excel لیست معرفی‌ها (فایل Excel واقعی برای دانلود)
	
	پارامترها:
	- selected_only: آیا فقط سطرهای انتخاب شده export شوند
	- selected_indices: لیست ایندکس‌های انتخاب شده (JSON string)
	- سایر پارامترهای QueryInfo برای فیلتر
	"""
	from app.services.pdf import PDFService
	import json
	import io
	from openpyxl import Workbook
	from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
	
	# Parse selected indices if provided
	indices = None
	if selected_only and selected_indices:
		try:
			indices = json.loads(selected_indices)
		except (json.JSONDecodeError, TypeError):
			indices = None
	
	# Get calendar type from request headers
	calendar_header = request.headers.get("X-Calendar-Type", "jalali")
	calendar_type = "jalali" if calendar_header.lower() in ["jalali", "persian", "shamsi"] else "gregorian"
	
	# Generate Excel data using new modular service
	pdf_service = PDFService()
	
	# Get locale from request headers
	locale_header = request.headers.get("Accept-Language", "fa")
	locale = "fa" if locale_header.startswith("fa") else "en"
	
	excel_data = pdf_service.generate_excel_data(
		module_name='marketing',
		data={},  # Empty data - module will fetch its own data
		calendar_type=calendar_type,
		locale=locale,
		db=db,
		user_id=ctx.user.id,
		query_info=query_info,
		selected_indices=indices
	)
	
	# Create Excel workbook
	wb = Workbook()
	ws = wb.active
	ws.title = "Referrals"
	
	# Define styles
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	
	# Add headers
	if excel_data:
		headers = list(excel_data[0].keys())
		for col, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col, value=header)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = header_alignment
			cell.border = border
		
		# Add data rows
		for row, data in enumerate(excel_data, 2):
			for col, header in enumerate(headers, 1):
				cell = ws.cell(row=row, column=col, value=data.get(header, ""))
				cell.border = border
				# Center align for numbers and dates
				if header in ["ردیف", "Row", "تاریخ ثبت", "Registration Date"]:
					cell.alignment = Alignment(horizontal="center")
	
	# Auto-adjust column widths
	for column in ws.columns:
		max_length = 0
		column_letter = column[0].column_letter
		for cell in column:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = min(max_length + 2, 50)
		ws.column_dimensions[column_letter].width = adjusted_width
	
	# Save to BytesIO
	excel_buffer = io.BytesIO()
	wb.save(excel_buffer)
	excel_buffer.seek(0)
	
	# Generate filename
	filename = f"referrals_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
	
	# Return Excel file as response
	return Response(
		content=excel_buffer.getvalue(),
		media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		headers={
			"Content-Disposition": f"attachment; filename={filename}",
			"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
		}
	)


@router.get(
	"/verify-email",
	summary="تایید ایمیل کاربر",
	description="تایید ایمیل کاربر با استفاده از token ارسال شده در ایمیل verification",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "ایمیل با موفقیت تایید شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ایمیل با موفقیت تایید شد",
						"data": {
							"user_id": 1,
							"email": "user@example.com",
							"email_verified": True
						}
					}
				}
			}
		},
		400: {
			"description": "Token نامعتبر یا منقضی شده",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"error_code": "INVALID_TOKEN",
						"message": "Token نامعتبر یا استفاده شده است"
					}
				}
			}
		}
	}
)
def verify_email(
	request: Request,
	token: str = Query(..., description="Token verification از ایمیل"),
	db: Session = Depends(get_db)
):
	"""تایید ایمیل کاربر با token. در مرورگر یک صفحهٔ خوانا و در API همان JSON برمی‌گردد."""
	accept = request.headers.get("accept", "") or ""
	wants_html = "text/html" in accept

	try:
		user = verify_email_token(db, token)
	except ApiError as e:
		if wants_html:
			msg = "لینک نامعتبر یا منقضی شده است."
			if isinstance(getattr(e, "detail", None), dict):
				err = e.detail.get("error") if isinstance(e.detail.get("error"), dict) else {}
				msg = err.get("message", msg)
			html = _verify_email_html(success=False, message=msg)
			return HTMLResponse(content=html, status_code=e.status_code)
		raise

	response_data = {
		"user_id": user.id,
		"email": user.email,
		"email_verified": user.email_verified
	}
	if wants_html:
		html = _verify_email_html(success=True, message="ایمیل شما با موفقیت تایید شد.")
		return HTMLResponse(content=html)

	return success_response(response_data, request, message="EMAIL_VERIFIED")


def _verify_email_html(*, success: bool, message: str) -> str:
	"""صفحهٔ HTML ساده برای نمایش نتیجهٔ تایید ایمیل در مرورگر (RTL)."""
	title = "تایید ایمیل" if success else "خطا در تایید ایمیل"
	icon = "✅" if success else "❌"
	color = "#22c55e" if success else "#dc2626"
	return f"""<!DOCTYPE html>
<html dir="rtl" lang="fa">
<head>
	<meta charset="utf-8">
	<meta name="viewport" content="width=device-width, initial-scale=1">
	<title>{title}</title>
	<style>
		body {{ font-family: Tahoma, Arial, sans-serif; background: #f8fafc; margin: 0; padding: 2rem; display: flex; justify-content: center; align-items: center; min-height: 100vh; box-sizing: border-box; }}
		.card {{ background: white; border-radius: 12px; padding: 2rem; max-width: 420px; box-shadow: 0 4px 12px rgba(0,0,0,0.08); text-align: center; }}
		.icon {{ font-size: 3rem; margin-bottom: 1rem; }}
		.message {{ color: #334155; font-size: 1.1rem; line-height: 1.6; }}
	</style>
</head>
<body>
	<div class="card">
		<div class="icon">{icon}</div>
		<h1 style="color: {color}; margin: 0 0 1rem;">{title}</h1>
		<p class="message">{message}</p>
	</div>
</body>
</html>"""


@router.post(
	"/resend-verification",
	summary="ارسال مجدد ایمیل verification",
	description="ارسال مجدد ایمیل verification برای کاربر. نیاز به authentication دارد.",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "ایمیل verification با موفقیت ارسال شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ایمیل verification با موفقیت ارسال شد"
					}
				}
			}
		},
		400: {
			"description": "ایمیل قبلاً تایید شده یا تنظیم نشده",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"error_code": "EMAIL_ALREADY_VERIFIED",
						"message": "ایمیل کاربر قبلاً تایید شده است"
					}
				}
			}
		},
		429: {
			"description": "تعداد درخواست‌ها بیش از حد مجاز",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"error_code": "RATE_LIMIT_EXCEEDED",
						"message": "شما بیش از حد مجاز درخواست ارسال مجدد داده‌اید"
					}
				}
			}
		}
	}
)
def resend_verification(
	request: Request,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""ارسال مجدد ایمیل verification"""
	user_id = ctx.get_user_id()
	
	# ساخت base_url از request
	base_url = None
	if request.headers.get("X-Forwarded-Host"):
		proto = request.headers.get("X-Forwarded-Proto", "https")
		host = request.headers.get("X-Forwarded-Host")
		base_url = f"{proto}://{host}"
	elif request.url:
		base_url = str(request.url).replace(request.url.path, "").rstrip("/")
	
	resend_verification_email(db, user_id, base_url)
	
	return success_response({}, request, message="VERIFICATION_EMAIL_SENT")


@router.post(
	"/send-mobile-verification",
	summary="ارسال کد تایید به شماره موبایل",
	description="ارسال کد OTP 6 رقمی به شماره موبایل کاربر برای تایید",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کد تایید با موفقیت ارسال شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "کد تایید به شماره موبایل ارسال شد"
					}
				}
			}
		},
		400: {
			"description": "شماره موبایل نامعتبر یا SMS Provider پیکربندی نشده",
		},
		404: {
			"description": "کاربر با این شماره موبایل یافت نشد",
		},
		429: {
			"description": "تعداد درخواست‌ها بیش از حد مجاز",
		},
		503: {
			"description": "سرویس پیامک پیکربندی نشده است",
		}
	}
)
@rate_limit(
	max_requests=8,
	window_seconds=3600,
	key_func=lambda req: f"send_mobile_verif:{get_client_ip(req)}",
	error_message="تعداد درخواست‌های ارسال کد تایید موبایل بیش از حد مجاز است. لطفاً بعداً تلاش کنید.",
)
def send_mobile_verification(
	request: Request,
	payload: SendMobileVerificationRequest,
	db: Session = Depends(get_db)
) -> dict:
	"""ارسال کد OTP به شماره موبایل کاربر (همراه کپچا در بدنه)"""
	from app.services.mobile_verification_service import MobileVerificationService
	from app.services.captcha_service import validate_captcha
	from adapters.db.repositories.user_repo import UserRepository
	from app.services.auth_service import _normalize_mobile
	from app.core.responses import ApiError
	
	if not validate_captcha(db, payload.captcha_id, payload.captcha_code, client_ip=get_client_ip(request)):
		raise ApiError("INVALID_CAPTCHA", "کد کپچا نامعتبر است", http_status=400)
	
	# نرمال‌سازی شماره موبایل (همان مسیری که در ثبت‌نام است)
	normalized_mobile = _normalize_mobile(payload.mobile)
	if not normalized_mobile:
		raise ApiError("INVALID_MOBILE", "شماره موبایل نامعتبر است", http_status=400)
	
	# پیدا کردن کاربر از شماره موبایل
	user_repo = UserRepository(db)
	user = user_repo.get_by_mobile(normalized_mobile)
	if not user:
		raise ApiError("USER_NOT_FOUND", "کاربری با این شماره موبایل یافت نشد", http_status=404)
	
	service = MobileVerificationService(db)
	
	try:
		otp_code = service.create_mobile_verification(user.id, payload.mobile)
		data = {"message": "کد تایید به شماره موبایل ارسال شد"}
		data.update(_otp_in_response(otp_code))
		return success_response(data, request)
	except ApiError as e:
		raise e
	except Exception as e:
		raise ApiError("SMS_SEND_FAILED", f"خطا در ارسال پیامک: {str(e)}", http_status=500)


@router.post(
	"/verify-mobile",
	summary="تایید شماره موبایل با کد OTP",
	description="تایید شماره موبایل کاربر با استفاده از کد OTP ارسال شده",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "شماره موبایل با موفقیت تایید شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "شماره موبایل با موفقیت تایید شد",
						"data": {
							"user_id": 1,
							"mobile_verified": True
						}
					}
				}
			}
		},
		400: {
			"description": "کد OTP نامعتبر",
		},
		404: {
			"description": "کد تایید یافت نشد یا منقضی شده است",
		},
		429: {
			"description": "تعداد تلاش‌های مجاز به پایان رسیده است",
		}
	}
)
def verify_mobile(
	request: Request,
	otp_code: str = Query(..., description="کد OTP 6 رقمی"),
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""تایید شماره موبایل با کد OTP"""
	from app.services.mobile_verification_service import MobileVerificationService
	from adapters.db.models.user import User
	
	service = MobileVerificationService(db)
	user_id = ctx.get_user_id()
	
	service.verify_mobile_otp(user_id, otp_code)
	
	# به‌روزرسانی اطلاعات کاربر
	user = db.get(User, user_id)
	
	response_data = {
		"user_id": user_id,
		"mobile_verified": user.mobile_verified if user else False
	}
	
	return success_response(response_data, request, message="MOBILE_VERIFIED")


@router.post(
	"/resend-mobile-verification",
	summary="ارسال مجدد کد تایید موبایل",
	description="ارسال مجدد کد OTP به شماره موبایل کاربر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کد تایید مجدداً ارسال شد",
		},
		400: {
			"description": "کاربر شماره موبایل ثبت نکرده است",
		},
		429: {
			"description": "تعداد درخواست‌ها بیش از حد مجاز",
		}
	}
)
@rate_limit(
	max_requests=10,
	window_seconds=3600,
	key_func=lambda req: f"resend_mobile_verif:{get_client_ip(req)}",
	error_message="تعداد درخواست‌های ارسال مجدد کد بیش از حد مجاز است. لطفاً بعداً تلاش کنید.",
)
def resend_mobile_verification(
	request: Request,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""ارسال مجدد کد تایید موبایل"""
	from app.services.mobile_verification_service import MobileVerificationService
	
	service = MobileVerificationService(db)
	user_id = ctx.get_user_id()
	
	otp_code = service.resend_otp(user_id)
	data = {"message": "کد تایید مجدداً ارسال شد"}
	data.update(_otp_in_response(otp_code))
	return success_response(data, request)


@router.post(
	"/update-mobile",
	summary="تغییر شماره موبایل کاربر",
	description="تغییر شماره موبایل کاربر با تایید کپچا و بررسی تایید شده بودن توسط کاربر دیگر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "شماره موبایل با موفقیت به‌روزرسانی شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "شماره موبایل با موفقیت به‌روزرسانی شد",
						"data": {
							"mobile": "09123456789",
							"mobile_verified": False
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی یا تایید شده بودن توسط کاربر دیگر",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"error_code": "MOBILE_IN_USE_VERIFIED",
						"message": "این شماره موبایل قبلاً توسط کاربر دیگری ثبت و تایید شده است"
					}
				}
			}
		},
		429: {
			"description": "تعداد درخواست‌ها بیش از حد مجاز"
		}
	}
)
def update_mobile(
	request: Request,
	payload: UpdateMobileRequest,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""تغییر شماره موبایل کاربر"""
	from app.services.auth_service import update_user_mobile
	
	user_id = ctx.get_user_id()
	
	try:
		result = update_user_mobile(
			db=db,
			user_id=user_id,
			mobile=payload.mobile,
			captcha_id=payload.captcha_id,
			captcha_code=payload.captcha_code,
			force_unverified=payload.force_unverified,
			send_verification_sms=payload.send_verification_sms,
			client_ip=get_client_ip(request),
		)
		
		return success_response(
			result,
			request,
			message="شماره موبایل با موفقیت به‌روزرسانی شد"
		)
	except Exception as e:
		# خطاها از طریق ApiError برمی‌گردند
		raise


@router.post(
	"/update-email",
	summary="تغییر ایمیل کاربر",
	description="تغییر ایمیل کاربر با تایید کپچا و بررسی تایید شده بودن توسط کاربر دیگر",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "ایمیل با موفقیت به‌روزرسانی شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ایمیل با موفقیت به‌روزرسانی شد",
						"data": {
							"email": "newemail@example.com",
							"email_verified": False
						}
					}
				}
			}
		},
		400: {
			"description": "خطا در اعتبارسنجی یا تایید شده بودن توسط کاربر دیگر",
			"content": {
				"application/json": {
					"example": {
						"success": False,
						"error_code": "EMAIL_IN_USE_VERIFIED",
						"message": "این ایمیل قبلاً توسط کاربر دیگری ثبت و تایید شده است"
					}
				}
			}
		},
		429: {
			"description": "تعداد درخواست‌ها بیش از حد مجاز"
		}
	}
)
def update_email(
	request: Request,
	payload: UpdateEmailRequest,
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db)
) -> dict:
	"""تغییر ایمیل کاربر"""
	from app.services.auth_service import update_user_email
	
	user_id = ctx.get_user_id()
	
	try:
		result = update_user_email(
			db=db,
			user_id=user_id,
			email=payload.email,
			captcha_id=payload.captcha_id,
			captcha_code=payload.captcha_code,
			force_unverified=payload.force_unverified,
			client_ip=get_client_ip(request),
		)
		
		return success_response(
			result,
			request,
			message="ایمیل با موفقیت به‌روزرسانی شد"
		)
	except Exception as e:
		# خطاها از طریق ApiError برمی‌گردند
		raise


@router.get(
	"/login/otp-channel-status",
	summary="وضعیت پیکربندی کانال‌های OTP ورود",
	description="نمایش اینکه هر کانال (پیامک، ایمیل، تلگرام، بله) روی سرور پیکربندی شده یا نه — بدون اطلاع از کاربر.",
	response_model=SuccessResponse,
)
@rate_limit(
	max_requests=40,
	window_seconds=60,
	key_func=lambda req: f"login_otp_channel_status:{get_client_ip(req)}",
	error_message="تعداد درخواست‌ها بیش از حد مجاز است. لطفاً کمی صبر کنید.",
)
def get_otp_channel_status(request: Request, db: Session = Depends(get_db)) -> dict:
	from app.services.otp_login_service import OtpLoginService
	
	service = OtpLoginService(db)
	return success_response(service.get_otp_channel_status_flags(), request)


@router.post(
	"/login/available-channels",
	summary="دریافت کانال‌های در دسترس برای ورود با OTP",
	description=(
		"دریافت فهرست کانال‌های اولیه (فقط SMS یا ایمیل) بر اساس فرمت شناسه، "
		"بدون افشای وجود کاربر. نیاز به کپتچای معتبر دارد."
	),
	response_model=SuccessResponse,
)
@rate_limit(
	max_requests=12,
	window_seconds=300,
	key_func=lambda req: f"login_otp_channels:{get_client_ip(req)}",
	error_message="تعداد درخواست‌های دریافت کانال ورود بیش از حد مجاز است. لطفاً چند دقیقه بعد دوباره تلاش کنید.",
)
def post_login_available_channels(
	request: Request,
	payload: AvailableChannelsRequest,
	db: Session = Depends(get_db)
) -> dict:
	"""کانال‌های قابل انتخاب قبل از ارسال OTP (ضد user enumeration)."""
	from app.services.otp_login_service import OtpLoginService
	
	if not validate_captcha(db, payload.captcha_id, payload.captcha_code, client_ip=get_client_ip(request)):
		raise ApiError("INVALID_CAPTCHA", "کد کپتچا نامعتبر است", http_status=400)
	
	service = OtpLoginService(db)
	channels_info = service.get_public_otp_channel_options(payload.identifier)
	return success_response(channels_info, request)


@router.post(
	"/login/send-otp",
	summary="ارسال OTP برای ورود",
	description="ارسال کد OTP 6 رقمی به ایمیل، شماره موبایل یا تلگرام برای ورود بدون نیاز به رمز عبور",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "کد OTP با موفقیت ارسال شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"data": {
							"session_id": "session_id_here",
							"message": "کد ورود ارسال شد",
							"available_channels": ["sms", "email", "telegram"]
						}
					}
				}
			}
		},
		400: {
			"description": "شناسه، کانال یا کپتچا نامعتبر",
		},
		429: {
			"description": "تعداد درخواست‌ها بیش از حد مجاز",
		},
		503: {
			"description": "سرویس مورد نظر پیکربندی نشده است",
		}
	}
)
@rate_limit(
	max_requests=5,
	window_seconds=300,
	key_func=lambda req: f"login_otp:{get_client_ip(req)}",
	error_message="تعداد درخواست‌های ارسال کد ورود بیش از حد مجاز است. لطفاً چند دقیقه بعد دوباره تلاش کنید."
)
def send_login_otp(
	request: Request,
	payload: SendLoginOtpRequest,
	db: Session = Depends(get_db)
) -> dict:
	"""ارسال OTP برای ورود"""
	from app.services.otp_login_service import OtpLoginService
	from app.services.captcha_service import validate_captcha
	from app.core.responses import ApiError
	
	# تایید کپتچا
	if not validate_captcha(db, payload.captcha_id, payload.captcha_code, client_ip=get_client_ip(request)):
		raise ApiError("INVALID_CAPTCHA", "کد کپتچا نامعتبر است", http_status=400)
	
	service = OtpLoginService(db)
	
	# دریافت IP و User Agent
	ip_address = get_client_ip(request)
	user_agent = request.headers.get("User-Agent")
	
	try:
		from app.services.otp_login_service import OTP_LOGIN_CHANNEL_PUBLIC_MESSAGE
		
		success, new_session_id, channels_info = service.send_login_otp(
			identifier=payload.identifier,
			channel=payload.channel,
			ip_address=ip_address,
			user_agent=user_agent,
			session_id=payload.session_id
		)
		
		if not success:
			raise ApiError("SEND_FAILED", OTP_LOGIN_CHANNEL_PUBLIC_MESSAGE, http_status=400)
		
		channel_messages = {
			"sms": "کد ورود به شماره موبایل ارسال شد",
			"email": "کد ورود به ایمیل ارسال شد",
			"telegram": "کد ورود به تلگرام ارسال شد",
			"bale": "کد ورود به بله ارسال شد"
		}
		
		return success_response({
			"session_id": new_session_id,
			"message": channel_messages.get(payload.channel, "کد ورود ارسال شد"),
			"available_channels": channels_info.get("available_channels", [])
		}, request)
	except ApiError as e:
		raise e
	except Exception as e:
		raise ApiError("OTP_SEND_FAILED", f"خطا در ارسال OTP: {str(e)}", http_status=500)


@router.post(
	"/login/verify-otp",
	summary="تایید OTP و ورود",
	description="تایید کد OTP و ورود کاربر به سیستم",
	response_model=SuccessResponse,
	responses={
		200: {
			"description": "ورود با موفقیت انجام شد",
			"content": {
				"application/json": {
					"example": {
						"success": True,
						"message": "ورود با موفقیت انجام شد",
						"data": {
							"api_key": "sk_1234567890abcdef",
							"expires_at": "2024-01-02T00:00:00Z",
							"user": {
								"id": 1,
								"first_name": "احمد",
								"last_name": "احمدی",
								"email": "ahmad@example.com",
								"mobile": "09123456789"
							}
						}
					}
				}
			}
		},
		400: {
			"description": "کد OTP نامعتبر",
		},
		404: {
			"description": "Session یا کاربر یافت نشد",
		},
		429: {
			"description": "تعداد تلاش‌های مجاز به پایان رسیده است",
		}
	}
)
@rate_limit(
	max_requests=10,
	window_seconds=600,
	key_func=lambda req: f"verify_login_otp:{get_client_ip(req)}",
	error_message="تعداد تلاش‌های تایید کد ورود بیش از حد مجاز است. لطفاً بعداً تلاش کنید."
)
def verify_login_otp(
	request: Request,
	session_id: str = Query(..., description="شناسه session"),
	otp_code: str = Query(..., description="کد OTP 6 رقمی"),
	device_id: str = Query(None, description="شناسه دستگاه (اختیاری)"),
	db: Session = Depends(get_db)
) -> dict:
	"""تایید OTP و ورود"""
	from app.services.otp_login_service import OtpLoginService
	from app.core.responses import ApiError
	
	service = OtpLoginService(db)
	
	# دریافت IP و User Agent
	ip_address = request.client.host if request.client else None
	user_agent = request.headers.get("User-Agent")
	
	try:
		success, user_data, api_key = service.verify_login_otp(
			session_id=session_id,
			otp_code=otp_code,
			device_id=device_id,
			user_agent=user_agent,
			ip=ip_address
		)
		
		if not success or not user_data or not api_key:
			raise ApiError("OTP_VERIFICATION_FAILED", "تایید OTP ناموفق بود", http_status=400)
		
		response_data = {
			"api_key": api_key,
			"expires_at": None,  # اگر expires_at نیاز باشد، باید از service برگردانده شود
			"user": user_data
		}
		
		formatted_data = format_datetime_fields(response_data, request)
		return success_response(formatted_data, request, message="LOGIN_SUCCESS")
	except ApiError as e:
		raise e
	except Exception as e:
		raise ApiError("LOGIN_FAILED", f"خطا در ورود: {str(e)}", http_status=500)

