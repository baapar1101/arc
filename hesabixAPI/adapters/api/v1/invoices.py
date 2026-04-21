from typing import Dict, Any, List, Optional
from fastapi import APIRouter, Depends, Request, Body, UploadFile, File, Form
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, cast, Integer
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.exc import IntegrityError
from decimal import Decimal
import io
import json
import datetime
import re
import base64
from pathlib import Path
import logging

from adapters.db.session import get_db
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.permissions import require_business_access, require_business_management_dep, require_business_permission_dep, require_business_permission_by_entity_dep
from app.core.responses import success_response, format_datetime_fields
from app.core.cache import get_cache
from adapters.api.v1.schemas import QueryInfo
from adapters.db.models.document import Document
from adapters.db.models.document_line import DocumentLine
from adapters.db.models.account import Account
from adapters.db.models.currency import Currency
from adapters.db.models.product import Product
from adapters.db.models.warehouse import Warehouse
from adapters.db.models.fiscal_year import FiscalYear
from adapters.db.models.business import Business
from adapters.db.models.business_print_settings import BusinessPrintSettings
from adapters.db.models.user import User
from app.core.responses import ApiError
from app.services import invoice_service
from app.services.invoice_service import (
    create_invoice,
    update_invoice,
    delete_invoice,
    bulk_delete_invoices,
    invoice_document_to_dict,
    calculate_invoice_remaining,
    SUPPORTED_INVOICE_TYPES,
    get_invoice_installment_plan,
    search_installments,
    export_installments_csv,
    export_installments_xlsx,
)
from app.services.tax_submission_service import send_document_to_tax_system, inquire_tax_status
from app.services.pdf.template_renderer import render_template
from app.core.calendar import CalendarConverter
from adapters.db.models.person import Person
from app.services.receipt_payment_service import get_receipt_payment
from app.services.file_storage_service import FileStorageService
from app.services.person_service import calculate_person_balance
from app.services.document_list_sort import apply_invoice_search_ordering, apply_invoice_search_ordering_from_body
from adapters.db.models.bank_account import BankAccount
from adapters.db.models.cash_register import CashRegister
from adapters.db.models.petty_cash import PettyCash
from sqlalchemy import func


logger = logging.getLogger(__name__)


def _format_line_custom_attributes_for_pdf(extra_info: Any) -> Optional[str]:
    if not isinstance(extra_info, dict):
        return None
    raw = extra_info.get("line_custom_attributes")
    if not raw or not isinstance(raw, dict):
        return None
    parts: List[str] = []
    for k, v in raw.items():
        if v is None:
            continue
        parts.append(f"{k}: {v}")
    if not parts:
        return None
    return "؛ ".join(parts)


def _invoice_line_unit_display_for_pdf(pl: Dict[str, Any]) -> str:
    """
    متن واحد برای چاپ PDF: extra_info.unit می‌تواند main/secondary (ایمپورت) یا نام واحد (UI) باشد.
    """
    info = pl.get("extra_info") or {}
    if not isinstance(info, dict):
        info = {}
    raw = info.get("unit")
    main_u = str(pl.get("product_main_unit") or "").strip()
    sec_u = str(pl.get("product_secondary_unit") or "").strip()
    if isinstance(raw, str):
        key = raw.strip().lower()
        if key == "main":
            return main_u or "-"
        if key == "secondary":
            return sec_u or "-"
    if raw is not None:
        s = str(raw).strip()
        if s:
            return s
    return main_u or "-"


router = APIRouter(prefix="/invoices", tags=["اسناد فروش", "اسناد خرید"])


@router.post("/business/{business_id}")
@require_business_access("business_id")
def create_invoice_endpoint(
    request: Request,
    business_id: int,
    payload: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "add")),
) -> Dict[str, Any]:
    result = create_invoice(
        db=db,
        business_id=business_id,
        user_id=ctx.get_user_id(),
        data=payload,
    )
    return success_response(data=result, request=request, message="INVOICE_CREATED")


@router.get("/business/{business_id}/{invoice_id}/installments")
@require_business_access("business_id")
def get_invoice_installments_endpoint(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_by_entity_dep("invoices", "view", Document, "invoice_id")),
):
    data = get_invoice_installment_plan(db=db, business_id=business_id, invoice_id=invoice_id)
    return success_response(data=data, request=request, message="INSTALLMENT_PLAN_FETCHED")


@router.post("/business/{business_id}/installments/search")
@require_business_access("business_id")
def search_installments_endpoint(
    request: Request,
    business_id: int,
    payload: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "view")),
):
    """
    جستجوی اقساط با فیلترهای:
    {
      "fiscal_year_id": int?,
      "due_from": "YYYY-MM-DD"?,
      "due_to": "YYYY-MM-DD"?,
      "status": "pending|partial|paid|overdue"?,
      "status_in": ["pending","overdue"] | "pending,overdue"?,
      "bucket": "unpaid"|"upcoming"|"overdue_only"?,
      "min_overdue_days": int?,
      "group_by": "invoice"?,
      "person_id": int?,
      "invoice_id": int?,
      "take": 200,
      "skip": 0
    }
    """
    result = search_installments(db=db, business_id=business_id, query=payload or {})
    formatted = format_datetime_fields(result, request)
    return success_response(data=formatted, request=request, message="INSTALLMENTS_LIST_FETCHED")


@router.post("/business/{business_id}/installments/export/excel")
@require_business_access("business_id")
def export_installments_excel_endpoint(
    request: Request,
    business_id: int,
    payload: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "export")),
) -> Response:
    """
    خروجی XLSX از اقساط (در صورت نبودن کتابخانه، CSV بازگردانده می‌شود).
    """
    content, mime, ext = export_installments_xlsx(
        db=db,
        business_id=business_id,
        query=payload or {},
        calendar_type=ctx.get_calendar_type(),
    )
    filename = f"installments_{business_id}.{ext}"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": mime,
        # Flutter Web needs this to read filename/content-type headers
        "Access-Control-Expose-Headers": "Content-Disposition, Content-Type",
    }
    return Response(content=content, media_type=mime, headers=headers)


@router.post("/business/{business_id}/installments/export/pdf")
@require_business_access("business_id")
def export_installments_pdf_endpoint(
    request: Request,
    business_id: int,
    payload: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "export")),
) -> Response:
    """
    خروجی PDF گزارش اقساط بر اساس همان فیلترهای search_installments.
    """
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from app.services.pdf.template_renderer import load_farsi_font_data_uris

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    calendar_type = ctx.get_calendar_type()

    body = dict(payload or {})
    body.pop("group_by", None)
    data = search_installments(db=db, business_id=business_id, query=body, disable_pagination=True)
    items = (data.get("items") or [])

    # Resolve business name
    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b and getattr(b, "name", None):
            business_name = b.name
    except Exception:
        business_name = ""

    # Format dates based on calendar
    def _fmt_date(v: Any) -> str:
        if v is None:
            return ""
        try:
            if calendar_type == "jalali":
                fd = CalendarConverter.format_datetime(v, "jalali")
            else:
                fd = CalendarConverter.format_datetime(v, "gregorian")
            return fd.get("date_only") or fd.get("formatted") or str(v)
        except Exception:
            try:
                # Already string
                return str(v)
            except Exception:
                return ""

    for it in items:
        try:
            it["document_date"] = _fmt_date(it.get("document_date"))
            it["due_date"] = _fmt_date(it.get("due_date"))
        except Exception:
            pass

    # Filters summary (human readable)
    filters_summary: List[Dict[str, str]] = []
    try:
        fy_id = body.get("fiscal_year_id")
        if fy_id:
            filters_summary.append({"label": "سال مالی" if is_fa else "Fiscal year", "value": str(fy_id)})
    except Exception:
        pass
    try:
        due_from = body.get("due_from")
        if isinstance(due_from, str) and due_from:
            filters_summary.append({"label": "از سررسید" if is_fa else "Due from", "value": due_from[:10]})
    except Exception:
        pass
    try:
        due_to = body.get("due_to")
        if isinstance(due_to, str) and due_to:
            filters_summary.append({"label": "تا سررسید" if is_fa else "Due to", "value": due_to[:10]})
    except Exception:
        pass
    try:
        status = body.get("status")
        if isinstance(status, str) and status:
            filters_summary.append({"label": "وضعیت" if is_fa else "Status", "value": status})
    except Exception:
        pass
    try:
        person_id = body.get("person_id")
        if person_id:
            person_name = None
            try:
                p = db.query(Person).filter(Person.id == int(person_id), Person.business_id == business_id).first()
                person_name = getattr(p, "name", None) if p else None
            except Exception:
                person_name = None
            filters_summary.append({
                "label": "شخص" if is_fa else "Person",
                "value": (person_name or str(person_id)),
            })
    except Exception:
        pass
    try:
        invoice_id = body.get("invoice_id")
        if invoice_id:
            inv_code = None
            try:
                doc = db.query(Document).filter(Document.id == int(invoice_id), Document.business_id == business_id).first()
                inv_code = getattr(doc, "code", None) if doc else None
            except Exception:
                inv_code = None
            filters_summary.append({
                "label": "فاکتور" if is_fa else "Invoice",
                "value": (inv_code or str(invoice_id)),
            })
    except Exception:
        pass

    # Summary totals (reuse service stats when available)
    stats = data.get("stats") or {}

    now = datetime.datetime.now()
    footer_text = ""
    try:
        footer_label = "زمان چاپ" if is_fa else "Printed at"
        issuer_label = "صادرکننده" if is_fa else "Issued by"
        issuer_name = ""
        try:
            issuer_name = ctx.get_user_name() or ""
        except Exception:
            issuer_name = ""
        try:
            fd = CalendarConverter.format_datetime(now, "jalali" if calendar_type == "jalali" else "gregorian")
            printed_at_str = fd.get("formatted") or fd.get("date_only", "")
        except Exception:
            printed_at_str = now.strftime("%Y/%m/%d %H:%M")
        footer_text = f"{footer_label}: {printed_at_str}"
        if issuer_name:
            footer_text += f" | {issuer_label}: {issuer_name}"
    except Exception:
        footer_text = ""

    fa_font_url_regular = None
    fa_font_url_bold = None
    if is_fa:
        fa_font_url_regular, fa_font_url_bold = load_farsi_font_data_uris()

    html_content = render_template(
        "pdf/installments/list.html",
        {
            "title_text": "گزارش اقساط" if is_fa else "Installments report",
            "business_name": business_name,
            "generated_at": now,
            "is_fa": is_fa,
            "footer_text": footer_text,
            "paper_size": body.get("paper_size"),
            "orientation": body.get("orientation"),
            "fa_font_url_regular": fa_font_url_regular,
            "fa_font_url_bold": fa_font_url_bold,
            "filters_summary": filters_summary,
            "items": items,
            "stats": stats,
        },
    )

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=html_content).write_pdf(font_config=font_config)

    filename = f"installments_{business_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition, Content-Type",
        },
    )


@router.put("/business/{business_id}/{invoice_id}")
@require_business_access("business_id")
def update_invoice_endpoint(
    request: Request,
    business_id: int,
    invoice_id: int,
    payload: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_by_entity_dep("invoices", "edit", Document, "invoice_id")),
) -> Dict[str, Any]:
    # Optional safety: ensure ownership
    doc = db.query(Document).filter(Document.id == invoice_id).first()
    if not doc or doc.business_id != business_id or doc.document_type not in SUPPORTED_INVOICE_TYPES:
        # Lazy import to avoid circular
        from app.core.responses import ApiError
        raise ApiError("DOCUMENT_NOT_FOUND", "Invoice document not found", http_status=404)
    try:
        result = update_invoice(
            db=db,
            document_id=invoice_id,
            user_id=ctx.get_user_id(),
            data=payload,
        )
    except IntegrityError as e:
        db.rollback()
        err_msg = str(getattr(e, "orig", e) or e)
        err_lower = err_msg.lower()
        if "person_id" in err_lower or "document_lines" in err_lower:
            raise ApiError(
                "INVALID_PERSON",
                "شخص انتخاب‌شده معتبر نیست یا به این کسب‌وکار تعلق ندارد. لطفاً شخص دیگری انتخاب کنید.",
                http_status=400,
            )
        if "uq_documents_business_code" in err_lower:
            raise ApiError(
                "DUPLICATE_DOCUMENT_CODE",
                "این شماره فاکتور برای این کسب‌وکار قبلاً ثبت شده است",
                http_status=400,
            )
        raise ApiError(
            "UPDATE_FAILED",
            "خطای یکتایی یا ارجاع در دیتابیس. داده‌های ارسالی را بررسی کنید.",
            http_status=400,
        )
    return success_response(data=result, request=request, message="INVOICE_UPDATED")


@router.delete(
    "/business/{business_id}/{invoice_id}",
    summary="حذف فاکتور",
    description="حذف یک فاکتور",
)
@require_business_access("business_id")
def delete_invoice_endpoint(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "delete")),
) -> Dict[str, Any]:
    """حذف یک فاکتور"""
    # بررسی مالکیت
    doc = db.query(Document).filter(Document.id == invoice_id).first()
    if not doc or doc.business_id != business_id or doc.document_type not in SUPPORTED_INVOICE_TYPES:
        from app.core.responses import ApiError
        raise ApiError("DOCUMENT_NOT_FOUND", "Invoice document not found", http_status=404)
    
    # حذف فاکتور
    success = delete_invoice(db, invoice_id)
    
    if not success:
        from app.core.responses import ApiError
        raise ApiError("DELETE_FAILED", "Failed to delete invoice", http_status=500)
    
    return success_response(
        data={"deleted": True, "invoice_id": invoice_id},
        request=request,
        message="INVOICE_DELETED"
    )


@router.post(
    "/business/{business_id}/bulk-delete",
    summary="حذف گروهی فاکتورها",
    description="حذف چندین فاکتور به صورت همزمان. فاکتورهایی که به هر دلیل حذف نشوند در skipped با دلیل برگردانده می‌شوند.",
)
@require_business_access("business_id")
def bulk_delete_invoices_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "delete")),
) -> Dict[str, Any]:
    """حذف گروهی فاکتورها"""
    invoice_ids = body.get("invoice_ids") or []
    if not isinstance(invoice_ids, list):
        raise ApiError("INVALID_REQUEST", "invoice_ids must be a list", http_status=400)
    invoice_ids = [int(x) for x in invoice_ids if isinstance(x, (int, str)) and str(x).strip().isdigit()]
    if not invoice_ids:
        return success_response(
            data={"deleted": [], "skipped": []},
            request=request,
            message="INVOICE_BULK_DELETED",
        )
    result = bulk_delete_invoices(db, business_id, invoice_ids)
    return success_response(
        data=result,
        request=request,
        message="INVOICE_BULK_DELETED",
    )


@router.get("/business/{business_id}/{invoice_id}/delete-info")
@require_business_access("business_id")
def get_invoice_delete_info(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """دریافت اطلاعات مرتبط با فاکتور برای نمایش در هشدار حذف"""
    from app.core.responses import ApiError
    
    # بررسی مالکیت
    doc = db.query(Document).filter(Document.id == invoice_id).first()
    if not doc or doc.business_id != business_id or doc.document_type not in SUPPORTED_INVOICE_TYPES:
        raise ApiError("DOCUMENT_NOT_FOUND", "Invoice document not found", http_status=404)
    
    extra_info = doc.extra_info or {}
    links = extra_info.get("links") or {}
    
    # بررسی کارپوشه مودیان
    is_in_tax_workspace = bool(extra_info.get("tax_workspace"))
    tax_status = extra_info.get("tax_status", "")
    
    # بررسی اقساط
    has_installments = bool(extra_info.get("installment_plan"))
    installment_info = None
    if has_installments:
        plan = extra_info.get("installment_plan", {})
        schedule = plan.get("schedule", [])
        installment_info = {
            "count": len(schedule),
            "total_amount": plan.get("principal_total", 0),
        }
    
    # پاک‌سازی لینک‌های مرده قبل از بررسی اسناد دریافت/پرداخت
    logger = logging.getLogger(__name__)
    try:
        from app.services.invoice_service import _cleanup_dead_receipt_payment_links
        _cleanup_dead_receipt_payment_links(db, doc)
        db.commit()
        db.refresh(doc)
        # به‌روزرسانی extra_info و links
        extra_info = doc.extra_info or {}
        links = extra_info.get("links") or {}
    except Exception as e:
        logger.warning(f"خطا در پاک‌سازی لینک‌های مرده در get_invoice_delete_info: {e}")
    
    # بررسی اسناد دریافت/پرداخت
    receipt_payment_document_ids = links.get("receipt_payment_document_ids") or []
    receipt_payment_info = []
    if receipt_payment_document_ids:
        related_docs = db.query(Document).filter(
            Document.id.in_(receipt_payment_document_ids)
        ).all()
        for rp_doc in related_docs:
            total_amount = Decimal(0)
            try:
                from adapters.db.models.document_line import DocumentLine
                lines = db.query(DocumentLine).filter(
                    DocumentLine.document_id == rp_doc.id
                ).all()
                for line in lines:
                    debit = Decimal(str(line.debit or 0))
                    credit = Decimal(str(line.credit or 0))
                    total_amount += max(debit, credit)
            except Exception:
                pass
            
            receipt_payment_info.append({
                "id": rp_doc.id,
                "code": rp_doc.code,
                "type": rp_doc.document_type,
                "amount": float(total_amount),
                "is_zero": total_amount == Decimal(0),
            })
    
    # بررسی حواله‌های انبار
    warehouse_document_ids = links.get("warehouse_document_ids") or []
    warehouse_info = []
    if warehouse_document_ids:
        try:
            from adapters.db.models.warehouse_document import WarehouseDocument
            warehouse_docs = db.query(WarehouseDocument).filter(
                WarehouseDocument.id.in_(warehouse_document_ids)
            ).all()
            for wd in warehouse_docs:
                status = getattr(wd, "status", None)
                warehouse_info.append({
                    "id": wd.id,
                    "code": getattr(wd, "code", ""),
                    "status": status,
                    "is_finalized": status == "finalized",
                })
        except ImportError:
            pass
    
    return success_response(
        data={
            "invoice_id": invoice_id,
            "invoice_code": doc.code,
            "is_in_tax_workspace": is_in_tax_workspace,
            "tax_status": tax_status,
            "has_installments": has_installments,
            "installment_info": installment_info,
            "receipt_payment_documents": receipt_payment_info,
            "warehouse_documents": warehouse_info,
        },
        request=request,
        message="INVOICE_DELETE_INFO"
    )


@router.get("/business/{business_id}/{invoice_id}")
@require_business_access("business_id")
def get_invoice_endpoint(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_by_entity_dep("invoices", "view", Document, "invoice_id")),
) -> Dict[str, Any]:
    doc = db.query(Document).filter(Document.id == invoice_id).first()
    if not doc or doc.business_id != business_id or doc.document_type not in SUPPORTED_INVOICE_TYPES:
        from app.core.responses import ApiError
        raise ApiError("DOCUMENT_NOT_FOUND", "Invoice document not found", http_status=404)
    # پاک‌سازی لینک‌های مرده در invoice_document_to_dict انجام می‌شود
    result = invoice_document_to_dict(db, doc)
    return success_response(data={"item": result}, request=request, message="INVOICE")


@router.post("/business/{business_id}/backfill-profit-ledger")
@require_business_access("business_id")
def backfill_invoice_profit_ledger_endpoint(
    request: Request,
    business_id: int,
    payload: Dict[str, Any] = Body(default_factory=dict),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "edit")),
) -> Dict[str, Any]:
    """
    ذخیرهٔ مقادیر بهای تمام‌شده و سود ناخالص قطعی (دفتر) برای اسناد قبلی،
    مطابق تنظیم «زمان شناسایی قطعی» روی کسب‌وکار.
    """
    from adapters.db.models.business import Business

    business = db.query(Business).filter(Business.id == business_id).first()
    if not business:
        raise ApiError("BUSINESS_NOT_FOUND", "کسب‌وکار یافت نشد", http_status=404)

    if business.invoice_profit_calculation_method == "disabled":
        return success_response(
            data={
                "message": "محاسبه سود برای این کسب و کار غیرفعال است",
                "processed": 0,
                "skipped": 0,
            },
            request=request,
            message="PROFIT_CALCULATION_DISABLED",
        )

    fiscal_year_id = payload.get("fiscal_year_id")
    invoice_ids = payload.get("invoice_ids")
    limit = payload.get("limit")
    use_background = bool(payload.get("use_background", False))

    if use_background:
        try:
            from app.core.queue import get_queue_service, QUEUE_DEFAULT
            from app.services.jobs.invoice_profit_ledger_backfill_job import (
                backfill_invoice_profit_ledger_job,
            )

            qs = get_queue_service()
            if qs and qs.enabled:
                job = qs.enqueue(
                    backfill_invoice_profit_ledger_job,
                    business_id=business_id,
                    user_id=ctx.get_user_id(),
                    fiscal_year_id=fiscal_year_id,
                    invoice_ids=invoice_ids,
                    limit=limit,
                    queue_name=QUEUE_DEFAULT,
                    timeout=7200,
                    result_ttl=7200,
                )
                if job:
                    return success_response(
                        data={
                            "job_id": job.id,
                            "status": "queued",
                            "message": "به‌روزرسانی شناسایی قطعی در پس‌زمینه آغاز شد.",
                        },
                        request=request,
                    )
        except Exception as exc:
            logger.warning("profit ledger queue failed, fallback sync: %s", exc)

    from app.services.invoice_profit_ledger_service import (
        backfill_recognized_profit_for_business,
    )

    result = backfill_recognized_profit_for_business(
        db,
        business_id,
        fiscal_year_id=int(fiscal_year_id) if fiscal_year_id is not None else None,
        invoice_ids=[int(x) for x in invoice_ids] if invoice_ids else None,
        limit=int(limit) if limit is not None else None,
    )
    return success_response(data=result, request=request, message="PROFIT_LEDGER_BACKFILLED")


@router.post("/business/{business_id}/recalculate-all-profits")
@require_business_access("business_id")
def recalculate_all_invoice_profits_endpoint(
    request: Request,
    business_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "view")),
) -> Dict[str, Any]:
    """
    محاسبه مجدد سود تمام فاکتورهای یک کسب و کار
    
    این endpoint برای استفاده بعد از تغییر تنظیمات محاسبه سود طراحی شده است.
    به صورت خودکار از background job استفاده می‌کند.
    """
    from adapters.db.models.business import Business
    
    business = db.query(Business).filter(Business.id == business_id).first()
    if not business:
        from app.core.responses import ApiError
        raise ApiError("BUSINESS_NOT_FOUND", "Business not found", http_status=404)
    
    # بررسی اینکه محاسبه سود فعال است
    if business.invoice_profit_calculation_method == "disabled":
        return success_response(
            data={
                "message": "محاسبه سود برای این کسب و کار غیرفعال است",
                "processed": 0,
                "skipped": 0
            },
            request=request,
            message="PROFIT_CALCULATION_DISABLED"
        )
    
    # استفاده از endpoint اصلی با پارامترهای پیش‌فرض
    return recalculate_invoice_profits_endpoint(
        request=request,
        business_id=business_id,
        payload={},  # بدون فیلتر - همه فاکتورها
        ctx=ctx,
        db=db,
        _=_
    )


@router.post("/business/{business_id}/recalculate-profits")
@require_business_access("business_id")
def recalculate_invoice_profits_endpoint(
    request: Request,
    business_id: int,
    payload: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "view")),
) -> Dict[str, Any]:
    """
    محاسبه مجدد سود فاکتورهای قدیمی
    
    پارامترها:
    - invoice_ids: لیست شناسه فاکتورها (اختیاری - اگر نباشد، تمام فاکتورهای کسب و کار)
    - document_type: نوع فاکتور (اختیاری - برای فیلتر)
    - fiscal_year_id: سال مالی (اختیاری - برای فیلتر)
    - batch_size: تعداد فاکتورها در هر batch (پیش‌فرض: 100)
    - use_background: استفاده از background job برای پردازش (پیش‌فرض: true)
    """
    from adapters.db.models.business import Business
    from adapters.db.models.fiscal_year import FiscalYear
    
    business = db.query(Business).filter(Business.id == business_id).first()
    if not business:
        from app.core.responses import ApiError
        raise ApiError("BUSINESS_NOT_FOUND", "Business not found", http_status=404)
    
    # بررسی اینکه محاسبه سود فعال است
    if business.invoice_profit_calculation_method == "disabled":
        return success_response(
            data={
                "message": "محاسبه سود برای این کسب و کار غیرفعال است",
                "processed": 0,
                "skipped": 0
            },
            request=request,
            message="PROFIT_CALCULATION_DISABLED"
        )
    
    invoice_ids = payload.get("invoice_ids")
    document_type = payload.get("document_type")
    fiscal_year_id = payload.get("fiscal_year_id")
    batch_size = payload.get("batch_size", 100)
    use_background = payload.get("use_background", True)
    
    # ساخت query برای فاکتورها
    query = db.query(Document).filter(Document.business_id == business_id)
    query = query.filter(Document.document_type.in_(SUPPORTED_INVOICE_TYPES))
    
    if invoice_ids:
        query = query.filter(Document.id.in_(invoice_ids))
    if document_type:
        query = query.filter(Document.document_type == document_type)
    if fiscal_year_id:
        query = query.filter(Document.fiscal_year_id == fiscal_year_id)
    
    # فقط فاکتورهای فروش و تولید
    query = query.filter(
        Document.document_type.in_(["invoice_sales", "invoice_sales_return", "invoice_production"])
    )
    
    total_invoices = query.count()
    
    if total_invoices == 0:
        return success_response(
            data={
                "message": "هیچ فاکتوری برای محاسبه سود یافت نشد",
                "processed": 0,
                "skipped": 0,
                "total": 0
            },
            request=request,
            message="NO_INVOICES_FOUND"
        )
    
    # اگر تعداد فاکتورها زیاد است یا use_background فعال است، از background job استفاده کن
    if use_background and total_invoices > batch_size:
        from app.core.queue import get_queue_service, QUEUE_DEFAULT
        from app.services.jobs.invoice_profit_job import recalculate_invoice_profits_job
        
        queue_service = get_queue_service()
        if queue_service and queue_service.enabled:
            # دریافت invoice_ids برای job
            invoice_id_list = [d.id for d in query.limit(10000).all()] if not invoice_ids else invoice_ids
            
            job = queue_service.enqueue(
                recalculate_invoice_profits_job,
                business_id=business_id,
                user_id=ctx.get_user_id(),
                invoice_ids=invoice_id_list if invoice_ids else None,
                document_type=document_type,
                fiscal_year_id=fiscal_year_id,
                batch_size=batch_size,
                queue_name=QUEUE_DEFAULT,
                timeout=3600,  # 1 ساعت timeout
                result_ttl=7200,  # نتیجه را 2 ساعت نگه دار
            )
            
            if job:
                return success_response({
                    "job_id": job.id,
                    "status": "queued",
                    "total_invoices": total_invoices,
                    "message": f"محاسبه سود {total_invoices} فاکتور در پس‌زمینه شروع شد. از GET /api/v1/jobs/{job.id} برای بررسی وضعیت استفاده کنید."
                }, request)
    
    # اجرای sync برای تعداد کم فاکتورها
    from decimal import Decimal
    
    # پردازش همه فاکتورها (نه فقط batch_size)
    invoices = query.all()
    processed = 0
    skipped = 0
    errors = []
    
    logger.info(f"Starting profit recalculation for {len(invoices)} invoices (business_id={business_id})")
    
    for doc in invoices:
        try:
            # محاسبه سود (فقط برای بررسی - نتیجه ذخیره نمی‌شود چون on-demand است)
            from app.services.invoice_service import _calculate_invoice_profit
            profit_data = _calculate_invoice_profit(
                db,
                business_id,
                doc.id,
                business.invoice_profit_calculation_method or "automatic",
                business.invoice_profit_calculation_basis or "purchase_price",
                business.invoice_profit_include_overhead or False,
                business.invoice_profit_overhead_type or "none",
                Decimal(str(business.invoice_profit_overhead_percent or 0)) if business.invoice_profit_overhead_percent else None,
                business.invoice_profit_calculation_type or "gross"
            )
            
            # بررسی اینکه آیا سود محاسبه شده است
            if profit_data and (profit_data.get("gross_profit") is not None or profit_data.get("net_profit") is not None):
                processed += 1
                logger.debug(f"Successfully calculated profit for invoice {doc.id} (code: {doc.code})")
            else:
                skipped += 1
                errors.append({
                    "invoice_id": doc.id,
                    "invoice_code": doc.code,
                    "error": "سود محاسبه نشد (نتیجه خالی)"
                })
                logger.warning(f"Empty profit result for invoice {doc.id} (code: {doc.code})")
        except Exception as e:
            skipped += 1
            error_msg = str(e)
            errors.append({
                "invoice_id": doc.id,
                "invoice_code": doc.code,
                "error": error_msg
            })
            logger.error(f"Error calculating profit for invoice {doc.id} (code: {doc.code}): {e}", exc_info=True)
    
    logger.info(f"Profit recalculation completed: processed={processed}, skipped={skipped}, total={total_invoices}")
    
    return success_response(
        data={
            "message": f"محاسبه سود برای {processed} فاکتور انجام شد",
            "processed": processed,
            "skipped": skipped,
            "total": total_invoices,
            "errors": errors[:20] if errors else []  # 20 خطای اول برای بررسی بهتر
        },
        request=request,
        message="PROFIT_RECALCULATED"
    )

@router.get(
    "/business/{business_id}/{invoice_id}/pdf",
    summary="PDF یک فاکتور",
    description="دریافت فایل PDF یک فاکتور با پشتیبانی از قالب سفارشی (invoices/detail)",
)
@require_business_access("business_id")
async def export_single_invoice_pdf(
    business_id: int,
    invoice_id: int,
    request: Request,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    template_id: int | None = None,
    _: None = Depends(require_business_permission_by_entity_dep("invoices", "view", Document, "invoice_id")),
):
    """
    خروجی PDF تک‌سند فاکتور با پشتیبانی از قالب سفارشی:
    - اگر template_id داده شود و منتشرشده باشد، همان استفاده می‌شود.
    - در غیر این صورت اگر قالب پیش‌فرض منتشرشده برای invoices/detail موجود باشد، استفاده می‌شود.
    - در نبود قالب، خروجی HTML پیش‌فرض تولید می‌شود.
    """
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from html import escape
    import datetime

    # دریافت سند و اعتبارسنجی
    doc = db.query(Document).filter(Document.id == invoice_id).first()
    if not doc or doc.business_id != business_id or doc.document_type not in SUPPORTED_INVOICE_TYPES:
        from app.core.responses import ApiError
        raise ApiError("DOCUMENT_NOT_FOUND", "Invoice document not found", http_status=404)

    # جزئیات کامل فاکتور (به‌صورت دیکشنری قابل ارسال به قالب)
    item = invoice_document_to_dict(db, doc)
    item = dict(item or {})

    # اطلاعات کسب‌وکار (اختیاری) + فایل‌های گرافیکی (لوگو/مهر) و امضای مالک
    business_name = ""
    business_info: Dict[str, Any] = {}
    business_logo_data_uri: Optional[str] = None
    business_stamp_data_uri: Optional[str] = None
    owner_signature_data_uri: Optional[str] = None

    storage = FileStorageService(db)

    async def _load_image_data_uri(file_id_str: Optional[str]) -> Optional[str]:
        """دریافت داده فایل و تبدیل به data URI برای استفاده در HTML/PDF."""
        if not file_id_str:
            return None
        try:
            from uuid import UUID

            try:
                file_data = await storage.download_file(UUID(str(file_id_str)))
            except Exception:
                # در صورت بروز خطا، None برمی‌گردانیم تا قالب بدون تصویر ادامه دهد
                return None
            content: bytes = file_data.get("content") or b""
            if not content:
                return None
            mime = file_data.get("mime_type") or "image/png"
            b64 = base64.b64encode(content).decode("ascii")
            return f"data:{mime};base64,{b64}"
        except Exception:
            return None

    show_stamp_override = None

    # تنظیمات چاپ کسب‌وکار (لوگو، مهر، پرداخت‌ها، اقساط و متن انتهایی)
    # یک کانفیگ پیش‌فرض تعریف می‌کنیم تا در صورت بروز خطا یا نبود کسب‌وکار، همچنان در دسترس باشد
    print_settings: Dict[str, Any] = {
        "show_logo": True,
        "show_stamp": True,
        "show_payments": True,
        "show_installment_plan": True,
        "footer_note": None,
    }
    invoice_footer_note: Optional[str] = None

    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
            # اطلاعات اقتصادی و تماس کسب‌وکار
            economic_id = getattr(b, "economic_id", None)
            economic_code = getattr(b, "economic_code", None)
            business_info = {
                "name": getattr(b, "name", None),
                # برای سازگاری با قالب‌های قدیمی
                "economic_id": economic_id or economic_code,
                "economic_code": economic_code or economic_id,
                "national_id": getattr(b, "national_id", None),
                "registration_number": getattr(b, "registration_number", None),
                "address": getattr(b, "address", None),
                "postal_code": getattr(b, "postal_code", None),
                "phone": getattr(b, "phone", None),
                "mobile": getattr(b, "mobile", None),
            }

            # ابتدا تنظیمات چاپ را (در صورت وجود) برای این کسب‌وکار و نوع سند می‌خوانیم
            try:
                print_rows = (
                    db.query(BusinessPrintSettings)
                    .filter(BusinessPrintSettings.business_id == business_id)
                    .all()
                )
            except Exception:
                print_rows = []

            def _pick_print_settings() -> dict:
                # از print_settings فعلی به‌عنوان مقدار اولیه استفاده می‌کنیم
                default_cfg = dict(print_settings)
                per_type_cfg = None
                for r in print_rows:
                    if r.document_type == "all":
                        default_cfg = {
                            "show_logo": bool(getattr(r, "show_logo", True)),
                            "show_stamp": bool(getattr(r, "show_stamp", True)),
                            "show_payments": bool(getattr(r, "show_payments", True)),
                            "show_installment_plan": bool(
                                getattr(r, "show_installment_plan", True)
                            ),
                            "footer_note": getattr(r, "footer_note", None),
                        }
                    elif r.document_type == doc.document_type:
                        per_type_cfg = {
                            "show_logo": bool(getattr(r, "show_logo", True)),
                            "show_stamp": bool(getattr(r, "show_stamp", True)),
                            "show_payments": bool(getattr(r, "show_payments", True)),
                            "show_installment_plan": bool(
                                getattr(r, "show_installment_plan", True)
                            ),
                            "footer_note": getattr(r, "footer_note", None),
                        }
                if per_type_cfg is None:
                    return default_cfg
                # per_type روی default override می‌شود
                merged = dict(default_cfg)
                merged.update({k: v for k, v in per_type_cfg.items() if v is not None})
                return merged

            print_settings = _pick_print_settings()

            def _normalize_bool(value):
                if isinstance(value, bool):
                    return value
                if value is None:
                    return None
                if isinstance(value, (int, float)):
                    return bool(value)
                if isinstance(value, str):
                    v = value.strip().lower()
                    if v in {"1", "true", "yes", "on"}:
                        return True
                    if v in {"0", "false", "no", "off"}:
                        return False
                return None

            override_value = _normalize_bool(show_stamp_override)
            if override_value is not None:
                print_settings["show_stamp"] = override_value

            # لوگو و مهر کسب‌وکار بر اساس تنظیمات چاپ
            if print_settings.get("show_logo", True):
                business_logo_data_uri = await _load_image_data_uri(
                    getattr(b, "logo_file_id", None)
                )
            else:
                business_logo_data_uri = None

            if print_settings.get("show_stamp", True):
                business_stamp_data_uri = await _load_image_data_uri(
                    getattr(b, "stamp_file_id", None)
                )
            else:
                business_stamp_data_uri = None

            # امضای مالک کسب‌وکار (بر اساس owner_id) فقط اگر show_stamp فعال باشد
            try:
                owner_user = db.query(User).filter(User.id == b.owner_id).first()
            except Exception:
                owner_user = None
            if owner_user is not None and print_settings.get("show_stamp", True):
                owner_signature_data_uri = await _load_image_data_uri(
                    getattr(owner_user, "signature_file_id", None)
                )

            invoice_footer_note = print_settings.get("footer_note")
    except Exception:
        business_name = ""
        business_info = {}
        business_logo_data_uri = None
        business_stamp_data_uri = None
        owner_signature_data_uri = None
        invoice_footer_note = None

    # Locale و نوع تقویم
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    calendar_header = (request.headers.get("X-Calendar-Type") or "").strip().lower()
    calendar_type = calendar_header or ("jalali" if is_fa else "gregorian")

    # تاریخ فاکتور با هر دو فرمت
    invoice_date_raw = item.get("document_date")
    invoice_date_jalali = None
    invoice_date_gregorian = None
    if invoice_date_raw:
        try:
            dt = datetime.datetime.fromisoformat(str(invoice_date_raw).replace("Z", "+00:00"))
            jalali = CalendarConverter.format_datetime(dt, "jalali")
            greg = CalendarConverter.format_datetime(dt, "gregorian")
            # فقط تاریخ بدون زمان را برای نمایش استفاده می‌کنیم
            invoice_date_jalali = jalali.get("date_only") or jalali.get("formatted", "")
            invoice_date_gregorian = greg.get("date_only") or greg.get("formatted", "")
        except Exception:
            invoice_date_gregorian = str(invoice_date_raw)

    if calendar_type == "jalali" and invoice_date_jalali:
        invoice_date_display = invoice_date_jalali
    else:
        invoice_date_display = invoice_date_gregorian or invoice_date_raw

    # نوع فاکتور به‌صورت خوانا
    def _type_name(tp: str) -> str:
        mapping = {
            "invoice_sales": ("فروش" if is_fa else "Sales"),
            "invoice_sales_return": ("برگشت از فروش" if is_fa else "Sales return"),
            "invoice_purchase": ("خرید" if is_fa else "Purchase"),
            "invoice_purchase_return": ("برگشت از خرید" if is_fa else "Purchase return"),
            "invoice_direct_consumption": ("مصرف مستقیم" if is_fa else "Direct consumption"),
            "invoice_production": ("تولید" if is_fa else "Production"),
            "invoice_waste": ("ضایعات" if is_fa else "Waste"),
        }
        return mapping.get(str(tp), str(tp))

    invoice_type_name = _type_name(item.get("document_type"))
    is_proforma = bool(item.get("is_proforma"))

    # اطلاعات طرف حساب (خریدار/فروشنده) بر اساس نوع فاکتور
    extra = item.get("extra_info") or {}
    person_id = extra.get("person_id")
    buyer_info: Dict[str, Any] = {}
    seller_info: Dict[str, Any] = {}

    person_obj = None
    try:
        if person_id is not None:
            person_obj = db.query(Person).filter(Person.id == int(person_id)).first()
    except Exception:
        person_obj = None

    person_info: Dict[str, Any] = {}
    if person_obj is not None:
        # اطلاعات اقتصادی و هویتی شخص (با پشتیبانی از فیلدهای قدیمی و جدید)
        national_id = getattr(person_obj, "national_id", None)
        national_code = getattr(person_obj, "national_code", None)
        registration_number = getattr(person_obj, "registration_number", None)
        economic_id = getattr(person_obj, "economic_id", None)
        economic_code = getattr(person_obj, "economic_code", None)

        # تعیین نام: اول display_name یا name، سپس first_name + last_name، در نهایت alias_name
        display_name = getattr(person_obj, "display_name", None)
        name = getattr(person_obj, "name", None)
        first_name = getattr(person_obj, "first_name", None)
        last_name = getattr(person_obj, "last_name", None)
        alias_name = getattr(person_obj, "alias_name", None)
        
        person_name = display_name or name
        if not person_name:
            # اگر display_name و name خالی بودند، از first_name و last_name استفاده می‌کنیم
            if first_name or last_name:
                name_parts = []
                if first_name:
                    name_parts.append(first_name)
                if last_name:
                    name_parts.append(last_name)
                person_name = " ".join(name_parts) if name_parts else None
        
        # اگر هنوز خالی است (یعنی first_name و last_name هم خالی بودند)، از alias_name استفاده می‌کنیم
        if not person_name:
            person_name = alias_name

        le_type = getattr(person_obj, "legal_entity_type", None) or "natural"
        name_prefix = getattr(person_obj, "name_prefix", None)
        person_info = {
            "id": getattr(person_obj, "id", None),
            "code": getattr(person_obj, "code", None),
            "name": person_name,
            "name_prefix": name_prefix,
            "legal_entity_type": le_type,
            "legal_entity_type_label": ("حقوقی" if is_fa else "Legal entity") if le_type == "legal" else ("حقیقی" if is_fa else "Natural person"),
            # برای سازگاری با قالب‌های قدیمی، هر دو کلید نگه داشته می‌شوند
            "national_id": national_id or national_code,
            "national_code": national_code or national_id,
            "registration_number": registration_number,
            "economic_id": economic_id or economic_code,
            "economic_code": economic_code or economic_id,
            "address": getattr(person_obj, "address", None),
            "postal_code": getattr(person_obj, "postal_code", None),
            "mobile": getattr(person_obj, "mobile", None),
            "phone": getattr(person_obj, "phone", None),
        }

    inv_type = str(item.get("document_type") or "")
    # برای فروش، کسب‌وکار فروشنده و شخص خریدار است
    if inv_type in ("invoice_sales", "invoice_sales_return"):
        seller_info = business_info if business_info else {"name": business_name}
        buyer_info = person_info if person_info else {}
    # برای خرید، شخص فروشنده و کسب‌وکار خریدار است
    elif inv_type in ("invoice_purchase", "invoice_purchase_return"):
        seller_info = person_info if person_info else {}
        buyer_info = business_info if business_info else {"name": business_name}
    else:
        # سایر انواع: فقط کسب‌وکار را به‌عنوان صاحب فاکتور نمایش می‌دهیم
        seller_info = business_info if business_info else {"name": business_name}
        buyer_info = person_info if person_info else {}
    
    # لاگ برای دیباگ آدرس
    logger.info(
        "Invoice PDF addresses: invoice_id=%s, seller.address=%s, buyer.address=%s",
        invoice_id,
        seller_info.get("address"),
        buyer_info.get("address"),
    )

    # خطوط فاکتور (کالا/خدمت)
    normalized_lines: list[dict[str, Any]] = []
    try:
        for pl in item.get("product_lines", []) or []:
            info = (pl.get("extra_info") or {}) if isinstance(pl, dict) else {}
            qty = pl.get("quantity")
            unit_price = info.get("unit_price")
            line_discount = info.get("line_discount") or 0
            tax_amount = info.get("tax_amount") or 0
            line_total = info.get("line_total")
            qty_display = None
            try:
                qf = float(qty or 0)
                upf = float(unit_price or 0)
                discf = float(line_discount or 0)
                taxf = float(tax_amount or 0)
                if line_total is None:
                    line_total = (qf * upf) - discf + taxf
                # نمایش تعداد: بدون اعشار اگر عدد صحیح باشد
                if qf.is_integer():
                    qty_display = f"{int(qf):,}"
                else:
                    qty_display = f"{qf:,.3f}".rstrip("0").rstrip(".")
            except Exception:
                qty_display = qty
            attrs_display = _format_line_custom_attributes_for_pdf(info)
            lc_attrs = info.get("line_custom_attributes") if isinstance(info, dict) else None
            if not isinstance(lc_attrs, dict):
                lc_attrs = {}
            normalized_lines.append(
                {
                    "product_code": pl.get("product_code"),
                    "product_name": pl.get("product_name"),
                    "description": pl.get("description"),
                    "quantity": qty,
                    "quantity_display": qty_display,
                    "unit_display": _invoice_line_unit_display_for_pdf(pl if isinstance(pl, dict) else {}),
                    "unit_price": unit_price,
                    "discount": line_discount,
                    "tax_amount": tax_amount,
                    "line_total": line_total,
                    "line_custom_attributes": lc_attrs,
                    "attributes_display": attrs_display,
                }
            )
    except Exception:
        normalized_lines = []

    # جمع مبالغ فاکتور از totals یا محاسبه مجدد
    totals = (extra.get("totals") or {}) if isinstance(extra, dict) else {}
    subtotal = totals.get("gross")
    discount_total = totals.get("discount")
    tax_total = totals.get("tax")
    payable_total = totals.get("net")

    try:
        if subtotal is None or discount_total is None or tax_total is None or payable_total is None:
            gross = 0.0
            discount_sum = 0.0
            tax_sum = 0.0
            net_sum = 0.0
            for ln in normalized_lines:
                try:
                    qf = float(ln.get("quantity") or 0)
                    upf = float(ln.get("unit_price") or 0)
                    discf = float(ln.get("discount") or 0)
                    taxf = float(ln.get("tax_amount") or 0)
                    line_total = ln.get("line_total")
                    if line_total is None:
                        line_total = (qf * upf) - discf + taxf
                    gross += qf * upf
                    discount_sum += discf
                    tax_sum += taxf
                    net_sum += float(line_total)
                except Exception:
                    continue
            if subtotal is None:
                subtotal = gross
            if discount_total is None:
                discount_total = discount_sum
            if tax_total is None:
                tax_total = tax_sum
            if payable_total is None:
                payable_total = net_sum
    except Exception:
        pass

    # تشخیص وجود تخفیف/مالیات در سطح سطرها برای نمایش ستون‌های جداگانه
    has_line_discount = False
    has_line_tax = False
    try:
        for ln in normalized_lines:
            try:
                if float(ln.get("discount") or 0) != 0:
                    has_line_discount = True
                if float(ln.get("tax_amount") or 0) != 0:
                    has_line_tax = True
                if has_line_discount and has_line_tax:
                    break
            except Exception:
                continue
    except Exception:
        has_line_discount = False
        has_line_tax = False

    # مبالغ تکمیلی: قبل از تخفیف و مالیات، و بدون مالیات
    amount_before_discount_and_tax = subtotal
    amount_without_tax = None
    try:
        base = float(subtotal or 0)
        disc = float(discount_total or 0)
        taxv = float(tax_total or 0)
        # مبلغ بدون مالیات = مبلغ بعد از تخفیف و قبل از مالیات
        amount_without_tax = base - disc
    except Exception:
        try:
            if payable_total is not None and tax_total is not None:
                amount_without_tax = float(payable_total) - float(tax_total or 0)
        except Exception:
            amount_without_tax = None

    # محاسبه وضعیت حساب مشتری (فقط برای فاکتورهای دارای person_id و با همان ارز فاکتور)
    customer_balance_info: Dict[str, Any] = {}
    try:
        if person_id is not None:
            # محاسبه تراز فعلی مشتری (فقط اسناد قطعی و با همان ارز فاکتور)
            invoice_currency_id = item.get("currency_id")
            if invoice_currency_id:
                # محاسبه تراز با فیلتر ارز
                query = db.query(
                    func.coalesce(func.sum(DocumentLine.credit), 0).label('total_credit'),
                    func.coalesce(func.sum(DocumentLine.debit), 0).label('total_debit')
                ).join(
                    Document, DocumentLine.document_id == Document.id
                ).filter(
                    DocumentLine.person_id == int(person_id),
                    Document.is_proforma == False,  # فقط اسناد قطعی
                    Document.currency_id == int(invoice_currency_id)  # فقط همان ارز فاکتور
                )
                result = query.first()
                if result is not None:
                    total_credit = float(result.total_credit or 0)
                    total_debit = float(result.total_debit or 0)
                    current_balance = total_credit - total_debit
                    if total_credit == 0 and total_debit == 0:
                        current_status = "بدون تراکنش"
                    elif current_balance > 0:
                        current_status = "بستانکار"
                    elif current_balance < 0:
                        current_status = "بدهکار"
                    else:
                        current_status = "بالانس"
                else:
                    current_balance = 0.0
                    current_status = "بدون تراکنش"
            else:
                # اگر ارز فاکتور مشخص نبود، از تابع قبلی استفاده می‌کنیم
                current_balance, current_status = calculate_person_balance(db, int(person_id))
            
            # اگر فاکتور پیش‌فاکتور است، تراز احتمالی بعد از قطعی شدن را محاسبه می‌کنیم
            if is_proforma:
                # محاسبه تاثیر این فاکتور بر تراز
                # برای فاکتور فروش: بدهکار می‌شود (debit)
                # برای فاکتور برگشت از فروش: بستانکار می‌شود (credit)
                # برای فاکتور خرید: بستانکار می‌شود (credit)
                # برای فاکتور برگشت از خرید: بدهکار می‌شود (debit)
                invoice_impact = 0.0
                if inv_type in ("invoice_sales", "invoice_purchase_return"):
                    # بدهکار می‌شود
                    invoice_impact = -float(payable_total or 0)
                elif inv_type in ("invoice_sales_return", "invoice_purchase"):
                    # بستانکار می‌شود
                    invoice_impact = float(payable_total or 0)
                
                potential_balance = current_balance + invoice_impact
                
                # تعیین وضعیت احتمالی
                if potential_balance > 0:
                    potential_status = "بستانکار" if is_fa else "Creditor"
                elif potential_balance < 0:
                    potential_status = "بدهکار" if is_fa else "Debtor"
                else:
                    potential_status = "بالانس" if is_fa else "Balanced"
                
                customer_balance_info = {
                    "current_balance": current_balance,
                    "current_status": current_status,
                    "potential_balance": potential_balance,
                    "potential_status": potential_status,
                    "invoice_impact": invoice_impact,
                }
            else:
                # فاکتور قطعی است، تراز فعلی شامل این فاکتور است
                customer_balance_info = {
                    "current_balance": current_balance,
                    "current_status": current_status,
                }
    except Exception:
        logger.exception("Error calculating customer balance for invoice_id=%s", invoice_id)
        customer_balance_info = {}

    # تراکنش‌های پرداخت مرتبط با فاکتور (رسید/پرداخت‌ها)
    payments: list[dict[str, Any]] = []
    try:
        show_payments = bool(print_settings.get("show_payments", True))
        logger.info(
            "Invoice PDF payments: show_payments=%s for invoice_id=%s", show_payments, invoice_id
        )
        if show_payments:
            links = (extra.get("links") or {}) if isinstance(extra, dict) else {}
            receipt_payment_ids = links.get("receipt_payment_document_ids") or []
            logger.info(
                "Invoice PDF payments: receipt_payment_document_ids=%s for invoice_id=%s",
                receipt_payment_ids,
                invoice_id,
            )
            for rid in receipt_payment_ids:
                try:
                    rp = get_receipt_payment(db, int(rid))
                except Exception:
                    rp = None
                if not rp:
                    continue
                # تاریخ پرداخت با تقویم کاربر
                pay_date_raw = rp.get("document_date")
                pay_date_display = None
                try:
                    if pay_date_raw:
                        dt = datetime.datetime.fromisoformat(str(pay_date_raw).replace("Z", "+00:00"))
                        if calendar_type == "jalali":
                            fd = CalendarConverter.format_datetime(dt, "jalali")
                            pay_date_display = fd.get("date_only") or fd.get("formatted", "")
                        else:
                            fd = CalendarConverter.format_datetime(dt, "gregorian")
                            pay_date_display = fd.get("date_only") or fd.get("formatted", "")
                except Exception:
                    pay_date_display = str(pay_date_raw) if pay_date_raw is not None else None
                # استخراج اطلاعات کامل از account_lines (نوع پرداخت، نام حساب، توضیحات)
                account_details: list[dict[str, Any]] = []
                methods: list[str] = []
                
                for ln in (rp.get("account_lines") or []):
                    ttype = (ln.get("transaction_type") or "").strip().lower()
                    if not ttype:
                        continue
                    
                    # نام نوع پرداخت
                    if ttype == "bank":
                        method_label = "بانک" if is_fa else "Bank"
                    elif ttype == "cash_register":
                        method_label = "صندوق" if is_fa else "Cash"
                    elif ttype == "petty_cash":
                        method_label = "تنخواه" if is_fa else "Petty cash"
                    elif ttype == "check":
                        method_label = "چک" if is_fa else "Check"
                    elif ttype == "wallet":
                        method_label = "کیف‌پول" if is_fa else "Wallet"
                    elif ttype == "person":
                        method_label = "شخص" if is_fa else "Person"
                    else:
                        method_label = ttype
                    
                    if method_label not in methods:
                        methods.append(method_label)
                    
                    # استخراج نام حساب (بانک/صندوق/تنخواه)
                    account_name = ln.get("account_name") or ""
                    bank_name = ln.get("bank_name") or ""
                    cash_register_name = ln.get("cash_register_name") or ""
                    petty_cash_name = ln.get("petty_cash_name") or ""
                    check_number = ln.get("check_number") or ""
                    description = ln.get("description") or ""
                    
                    # تعیین نام نمایشی
                    display_name = account_name
                    if ttype == "bank" and bank_name:
                        display_name = bank_name
                    elif ttype == "cash_register" and cash_register_name:
                        display_name = cash_register_name
                    elif ttype == "petty_cash" and petty_cash_name:
                        display_name = petty_cash_name
                    elif ttype == "check" and check_number:
                        display_name = f"چک {check_number}" if is_fa else f"Check {check_number}"
                    
                    account_details.append({
                        "transaction_type": ttype,
                        "method_label": method_label,
                        "display_name": display_name,
                        "amount": ln.get("amount", 0),
                        "description": description,
                    })
                
                payments.append(
                    {
                        "id": rp.get("id"),
                        "code": rp.get("code"),
                        "document_type": rp.get("document_type"),
                        "document_type_name": rp.get("document_type_name"),
                        "date": pay_date_display,
                        "total_amount": rp.get("total_amount"),
                        "methods": ", ".join(methods),
                        "account_details": account_details,
                        "description": rp.get("description") or "",
                    }
                )
        logger.info(
            "Invoice PDF payments: built %d payment rows for invoice_id=%s",
            len(payments),
            invoice_id,
        )
    except Exception:
        logger.exception(
            "Invoice PDF payments: error while building payments list for invoice_id=%s",
            invoice_id,
        )
        payments = []

    # طرح اقساط (در صورت وجود)
    installment_plan: dict[str, Any] | None = None
    try:
        if print_settings.get("show_installment_plan", True):
            extra_info = item.get("extra_info") or {}
            if isinstance(extra_info, dict) and isinstance(extra_info.get("installment_plan"), dict):
                # از سرویس نصب اقساط برای غنی‌سازی برنامه استفاده می‌کنیم
                try:
                    plan_view = get_invoice_installment_plan(db=db, business_id=business_id, invoice_id=invoice_id)
                except Exception:
                    plan_view = None
                if isinstance(plan_view, dict) and isinstance(plan_view.get("plan"), dict):
                    plan = dict(plan_view["plan"])
                    schedule = []
                    for it in plan.get("schedule") or []:
                        due_raw = it.get("due_date")
                        due_display = None
                        try:
                            if due_raw:
                                dt = datetime.datetime.fromisoformat(str(due_raw).replace("Z", "+00:00"))
                                if calendar_type == "jalali":
                                    fd = CalendarConverter.format_datetime(dt, "jalali")
                                    due_display = fd.get("date_only") or fd.get("formatted", "")
                                else:
                                    fd = CalendarConverter.format_datetime(dt, "gregorian")
                                    due_display = fd.get("date_only") or fd.get("formatted", "")
                        except Exception:
                            due_display = str(due_raw) if due_raw is not None else None
                        new_it = dict(it)
                        new_it["due_date_display"] = due_display
                        schedule.append(new_it)
                    plan["schedule"] = schedule
                    installment_plan = {
                        "meta": {
                            "invoice_code": plan_view.get("invoice_code"),
                            "person_id": plan_view.get("person_id"),
                        },
                        "data": plan,
                    }
    except Exception:
        installment_plan = None

    # غنی‌سازی دیکشنری فاکتور برای استفاده راحت‌تر در قالب و لیست‌ها
    item["title"] = item.get("title") or ("فاکتور" if is_fa else "Invoice")
    item["issue_date"] = invoice_date_display
    item["invoice_type_name"] = invoice_type_name
    item["is_proforma"] = is_proforma
    item["subtotal"] = subtotal
    item["discount_total"] = discount_total
    item["tax_total"] = tax_total
    item["payable_total"] = payable_total
    item["amount_before_discount_and_tax"] = amount_before_discount_and_tax
    item["amount_without_tax"] = amount_without_tax
    # فلگ فروش اقساطی: اگر طرح اقساط روی سند وجود داشته باشد
    try:
        extra_info_for_flag = item.get("extra_info") or {}
        item["is_installment_sale"] = bool(
            isinstance(extra_info_for_flag, dict)
            and isinstance(extra_info_for_flag.get("installment_plan"), dict)
        )
    except Exception:
        item["is_installment_sale"] = False

    # نام کاربر صادرکننده فاکتور
    issuer_name: Optional[str] = None
    try:
        issuer = db.query(User).filter(User.id == doc.created_by_user_id).first()
        if issuer is not None:
            first = getattr(issuer, "first_name", None) or ""
            last = getattr(issuer, "last_name", None) or ""
            full = (f"{first} {last}").strip()
            issuer_name = full or (issuer.email or issuer.mobile or str(issuer.id))
    except Exception:
        issuer_name = None

    # آدرس/داده فونت فارسی برای PDF (در صورت وجود و زبان فارسی)
    fa_font_url_regular: Optional[str] = None
    fa_font_url_bold: Optional[str] = None
    try:
        if is_fa:
            project_root = Path(__file__).resolve().parents[4]
            fonts_dir = project_root / "hesabixUI" / "hesabix_ui" / "assets" / "fonts"
            regular_path = fonts_dir / "YekanBakhFaNum-Regular.ttf"
            bold_path = fonts_dir / "YekanBakhFaNum-Bold.ttf"
            logger.info("PDF Font detection: fonts_dir=%s", fonts_dir)
            if regular_path.is_file():
                logger.info("PDF Font detection: Regular font found at %s", regular_path)
                import base64 as _b64
                _data = regular_path.read_bytes()
                _b64_data = _b64.b64encode(_data).decode("ascii")
                fa_font_url_regular = f"data:font/ttf;base64,{_b64_data}"
            else:
                logger.warning("PDF Font detection: Regular font NOT found at %s", regular_path)
            if bold_path.is_file():
                logger.info("PDF Font detection: Bold font found at %s", bold_path)
                import base64 as _b64b
                _data_b = bold_path.read_bytes()
                _b64_data_b = _b64b.b64encode(_data_b).decode("ascii")
                fa_font_url_bold = f"data:font/ttf;base64,{_b64_data_b}"
            else:
                logger.warning("PDF Font detection: Bold font NOT found at %s", bold_path)
    except Exception:
        logger.exception("PDF Font detection: error while loading YekanBakhFaNum fonts")
        fa_font_url_regular = None
        fa_font_url_bold = None

    # کانتکست قالب
    template_context = {
        "business_id": business_id,
        "business_name": business_name,
        "business": business_info,
        "business_logo_data_uri": business_logo_data_uri,
        "business_stamp_data_uri": business_stamp_data_uri,
        "owner_signature_data_uri": owner_signature_data_uri,
        "invoice": item,
        "lines": normalized_lines,
        "buyer": buyer_info,
        "seller": seller_info,
        "has_line_discount": has_line_discount,
        "has_line_tax": has_line_tax,
        "payments": payments,
        "installment_plan": installment_plan,
        "invoice_date_jalali": invoice_date_jalali,
        "invoice_date_gregorian": invoice_date_gregorian,
        "generated_at": datetime.datetime.now(),
        "is_fa": is_fa,
        "issuer_name": issuer_name,
        "fa_font_url_regular": fa_font_url_regular,
        "fa_font_url_bold": fa_font_url_bold,
        "invoice_footer_note": invoice_footer_note,
        "customer_balance_info": customer_balance_info,
    }

    # تلاش برای رندر با قالب سفارشی
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if template_id is not None:
                explicit_template_id = int(template_id)
        except Exception:
            explicit_template_id = None
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="invoices",
            subtype="detail",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # HTML پیش‌فرض در نبود قالب: استفاده از قالب فایل
    # پارامترهای صفحه از کوئری (اختیاری)
    show_stamp_override = None
    try:
        qp = request.query_params
        paper_size = qp.get("paper_size")
        orientation = qp.get("orientation")
        disposition = qp.get("disposition") or "attachment"
        show_stamp_override = qp.get("show_stamp")
    except Exception:
        paper_size = None
        orientation = None
        disposition = "attachment"
        show_stamp_override = None

    # حالت پیش‌فرض صفحه برای فاکتور: افقی (landscape)، مگر این‌که صراحتاً چیز دیگری ارسال شده باشد
    if not orientation:
        orientation = "landscape"
    # متن فوتر با زمان چاپ (بر اساس تقویم انتخاب‌شده کاربر) و نام صادرکننده
    try:
        now = template_context["generated_at"]
        footer_text = ""
        if isinstance(now, datetime.datetime):
            footer_label = "زمان چاپ" if is_fa else "Printed at"
            issuer_label = "صادرکننده" if is_fa else "Issued by"
            try:
                if calendar_type == "jalali":
                    fd = CalendarConverter.format_datetime(now, "jalali")
                else:
                    fd = CalendarConverter.format_datetime(now, "gregorian")
                printed_at_str = fd.get("formatted") or fd.get("date_only", "")
                if printed_at_str:
                    footer_text = f"{footer_label}: {printed_at_str}"
                    if issuer_name:
                        footer_text += f" | {issuer_label}: {issuer_name}"
            except Exception:
                footer_text = f"{footer_label}: {now.strftime('%Y/%m/%d %H:%M')}"
                if issuer_name:
                    footer_text += f" | {issuer_label}: {issuer_name}"
    except Exception:
        footer_text = ""

    default_ctx = {
        **template_context,
        "title_text": item.get("title") or ("فاکتور" if is_fa else "Invoice"),
        "paper_size": paper_size,
        "orientation": orientation,
        "footer_text": footer_text,
    }
    html_content = resolved_html or render_template("pdf/invoices/detail.html", default_ctx)

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=html_content).write_pdf(font_config=font_config)

    # نام فایل
    def _slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", (text or "")).strip("_") or "invoice"
    filename = f"invoice_{_slugify(item.get('code'))}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


def _apply_invoice_list_text_search(
    q,
    *,
    business_id: int,
    search: Optional[str],
    search_fields: Optional[List[Any]] = None,
    extra_info_person_id_expr=None,
):
    """
    جستجوی متنی لیست فاکتورها بر اساس search_fields.
    اگر search_fields ارسال نشود: کد سند، توضیح، و نام طرف‌حساب (شخص).
    فیلد counterparty با join به جدول persons روی extra_info.person_id اعمال می‌شود.
    """
    if not isinstance(search, str) or not search.strip():
        return q
    s = f"%{search.strip()}%"
    if search_fields:
        sf_set = {str(x).lower() for x in search_fields}
    else:
        sf_set = {"code", "description", "counterparty"}

    parts: List[Any] = []
    if "code" in sf_set:
        parts.append(Document.code.ilike(s))
    if "description" in sf_set:
        parts.append(Document.description.ilike(s))

    if "counterparty" in sf_set:
        pid_expr = extra_info_person_id_expr
        if pid_expr is None:
            _jb = cast(Document.extra_info, JSONB)
            pid_expr = cast(_jb["person_id"].astext, Integer)
        q = q.outerjoin(
            Person,
            and_(
                Person.id == pid_expr,
                Person.business_id == business_id,
            ),
        )
        full_name = func.nullif(func.trim(func.concat_ws(" ", Person.first_name, Person.last_name)), "")
        parts.append(
            or_(
                Person.alias_name.ilike(s),
                Person.first_name.ilike(s),
                Person.last_name.ilike(s),
                Person.company_name.ilike(s),
                full_name.ilike(s),
            )
        )

    if not parts:
        parts = [Document.code.ilike(s), Document.description.ilike(s)]
    return q.filter(or_(*parts))


@router.post("/business/{business_id}/search")
@require_business_access("business_id")
async def search_invoices_endpoint(
	request: Request,
	business_id: int,
	body_data: Dict[str, Any] = Body(...),
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db),
) -> Dict[str, Any]:
	"""لیست فاکتورها با فیلتر، جست‌وجو، مرتب‌سازی و صفحه‌بندی استاندارد"""

	# Locale for labels
	from app.core.i18n import negotiate_locale
	locale = negotiate_locale(request.headers.get("Accept-Language"))
	is_fa = locale == "fa"

	# Parse QueryInfo from body_data
	try:
		query_info = QueryInfo(**{k: v for k, v in body_data.items() if k in ['take', 'skip', 'sort_by', 'sort_desc', 'sort', 'search', 'search_fields', 'filters', 'include_inventory', 'inventory_as_of_date']})
	except Exception as e:
		raise ApiError("INVALID_QUERY", "پارامترهای جستجو معتبر نیستند.", http_status=400)

	# Extract additional fields for filtering
	body = body_data
	
	# کش نتایج جستجوی فاکتورها
	cache = get_cache()
	cache_key = None
	fiscal_year_id = None
	document_type = None
	project_id = None
	
	if cache.enabled:
		import json, hashlib
		from decimal import Decimal
		from datetime import datetime, date
		from adapters.api.v1.schemas import FilterItem
		
		# Helper function to convert objects to JSON-serializable format
		def to_serializable(obj):
			"""Convert Pydantic models and other non-serializable types to dict/primitive types"""
			if hasattr(obj, 'model_dump'):  # Pydantic v2
				return obj.model_dump()
			elif hasattr(obj, 'dict'):  # Pydantic v1
				return obj.dict()
			elif isinstance(obj, (datetime, date)):
				return obj.isoformat()
			elif isinstance(obj, Decimal):
				return float(obj)
			elif isinstance(obj, dict):
				return {k: to_serializable(v) for k, v in obj.items()}
			elif isinstance(obj, list):
				return [to_serializable(item) for item in obj]
			else:
				return obj
		
		# استخراج fiscal_year_id از header یا body
		try:
			fy_header = request.headers.get("X-Fiscal-Year-ID")
			if fy_header:
				fiscal_year_id = int(fy_header)
		except Exception:
			pass
		if fiscal_year_id is None:
			try:
				if body.get("fiscal_year_id") is not None:
					fiscal_year_id = int(body.get("fiscal_year_id"))
			except Exception:
				pass
		
		# استخراج document_type
		document_type = body.get("document_type")
		if isinstance(document_type, str) and document_type in SUPPORTED_INVOICE_TYPES:
			pass  # document_type از قبل تنظیم شده
		else:
			document_type = None
		
		# استخراج project_id
		try:
			if body.get("project_id") is not None:
				project_id = int(body.get("project_id"))
		except Exception:
			pass
		
		# Convert filters to serializable format
		serializable_filters = None
		if query_info.filters:
			serializable_filters = [to_serializable(f) for f in query_info.filters]
		
		# Convert body_data to serializable format (might contain non-serializable objects)
		serializable_body = to_serializable(body_data)
		
		key_payload = {
			"business_id": business_id,
			"take": query_info.take,
			"skip": query_info.skip,
			"sort_by": query_info.sort_by,
			"sort_desc": query_info.sort_desc,
			"sort": to_serializable(query_info.sort) if getattr(query_info, "sort", None) else None,
			"search": query_info.search,
			"filters": serializable_filters,
			"body": serializable_body,  # شامل تمام فیلترها
		}
		key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
		key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
		cache_key = f"invoices_search:{business_id}:{ctx.get_user_id()}:{key_hash}"
		cached = cache.get(cache_key)
		if cached is not None:
			return success_response(data=cached, request=request)

	# Base query
	q = db.query(Document).filter(
		and_(
			Document.business_id == business_id,
			Document.document_type.in_(list(SUPPORTED_INVOICE_TYPES)),
		)
	)
	# extra_info نوع JSON عمومی SQLAlchemy است؛ astext فقط روی JSONB است — برای PG ابتدا cast به JSONB
	_extra_info_jb = cast(Document.extra_info, JSONB)
	extra_info_person_id_int = cast(_extra_info_jb["person_id"].astext, Integer)

	search: Optional[str] = getattr(query_info, 'search', None)
	search_fields: Optional[List[Any]] = getattr(query_info, 'search_fields', None)
	q = _apply_invoice_list_text_search(
		q,
		business_id=business_id,
		search=search,
		search_fields=search_fields,
		extra_info_person_id_expr=extra_info_person_id_int,
	)

	# Extra filters
	doc_type = body.get("document_type")
	if isinstance(doc_type, str) and doc_type in SUPPORTED_INVOICE_TYPES:
		q = q.filter(Document.document_type == doc_type)

	is_proforma = body.get("is_proforma")
	if isinstance(is_proforma, bool):
		q = q.filter(Document.is_proforma == is_proforma)

	currency_id = body.get("currency_id")
	try:
		if currency_id is not None:
			q = q.filter(Document.currency_id == int(currency_id))
	except Exception:
		pass

	# Fiscal year from header or body
	fiscal_year_id = None
	try:
		fy_header = request.headers.get("X-Fiscal-Year-ID")
		if fy_header:
			fiscal_year_id = int(fy_header)
	except Exception:
		fiscal_year_id = None
	if fiscal_year_id is None:
		try:
			if body.get("fiscal_year_id") is not None:
				fiscal_year_id = int(body.get("fiscal_year_id"))
		except Exception:
			fiscal_year_id = None
	if fiscal_year_id is not None:
		q = q.filter(Document.fiscal_year_id == fiscal_year_id)

	# Project filter
	project_id = body.get("project_id")
	try:
		if project_id is not None:
			q = q.filter(Document.project_id == int(project_id))
	except Exception:
		pass

	# Date range from filters or flat body
	# 1) From QueryInfo.filters operators
	try:
		filters = getattr(query_info, 'filters', None)
	except Exception:
		filters = None
	if filters and isinstance(filters, (list, tuple)):
		for flt in filters:
			try:
				prop = getattr(flt, 'property', None) if not isinstance(flt, dict) else flt.get('property')
				op = getattr(flt, 'operator', None) if not isinstance(flt, dict) else flt.get('operator')
				val = getattr(flt, 'value', None) if not isinstance(flt, dict) else flt.get('value')
				if prop == 'person_id' and val is not None:
					# فیلتر بر اساس person_id در extra_info
					try:
						person_id_val = int(val)
						# فیلتر person_id از extra_info با ->> (معادل astext روی JSONB)
						q = q.filter(extra_info_person_id_int == person_id_val)
					except (ValueError, TypeError, KeyError):
						pass
				elif prop == 'document_date' and isinstance(val, str) and val:
					from app.services.transfer_service import _parse_iso_date as _p
					dt = _p(val)
					col = getattr(Document, prop)
					if op == ">=":
						q = q.filter(col >= dt)
					elif op == "<=":
						q = q.filter(col <= dt)
			except Exception:
				pass

	# 2) From flat body keys
	if isinstance(body.get("from_date"), str):
		try:
			from app.services.transfer_service import _parse_iso_date as _p
			q = q.filter(Document.document_date >= _p(body.get("from_date")))
		except Exception:
			pass
	if isinstance(body.get("to_date"), str):
		try:
			from app.services.transfer_service import _parse_iso_date as _p
			q = q.filter(Document.document_date <= _p(body.get("to_date")))
		except Exception:
			pass

	# پردازش person_id از body (اگر در filters نبود)
	person_id = body.get("person_id")
	if person_id is not None:
		try:
			person_id_val = int(person_id)
			# فیلتر بر اساس person_id در extra_info
			q = q.filter(extra_info_person_id_int == person_id_val)
		except (ValueError, TypeError, KeyError):
			pass

	# فیلتر فاکتورهای اقساطی (فقط اسنادی که طرح اقساط دارند)
	is_installment_sale = body.get("is_installment_sale")
	if is_installment_sale is True:
		# فقط فاکتورهای فروش اقساطی
		q = q.filter(Document.document_type == "invoice_sales")
		q = q.filter(Document.extra_info["installment_plan"].isnot(None))

	q = apply_invoice_search_ordering(q, query_info)

	# Pagination
	take = int(getattr(query_info, 'take', 20) or 20)
	skip = int(getattr(query_info, 'skip', 0) or 0)

	try:
		total = q.count()
		items = q.offset(skip).limit(take).all()
	except Exception as e:
		db.rollback()
		logging.getLogger(__name__).warning("invoice search query failed: %s", e)
		raise ApiError(
			"INVALID_SEARCH_FILTER",
			"پارامترهای جستجو یا داده‌های فاکتورها معتبر نیستند. در صورت استفاده از فیلتر شخص، آن را حذف کنید و دوباره امتحان کنید.",
			http_status=400,
		)

	# Helpers for display fields
	def _type_name(tp: str) -> str:
		mapping = {
			'invoice_sales': ('فروش' if is_fa else 'Sales'),
			'invoice_sales_return': ('برگشت از فروش' if is_fa else 'Sales return'),
			'invoice_purchase': ('خرید' if is_fa else 'Purchase'),
			'invoice_purchase_return': ('برگشت از خرید' if is_fa else 'Purchase return'),
			'invoice_direct_consumption': ('مصرف مستقیم' if is_fa else 'Direct consumption'),
			'invoice_production': ('تولید' if is_fa else 'Production'),
			'invoice_waste': ('ضایعات' if is_fa else 'Waste'),
		}
		return mapping.get(str(tp), str(tp))

	data_items: List[Dict[str, Any]] = []
	_log = logging.getLogger(__name__)
	for d in items:
		try:
			item = invoice_document_to_dict(db, d)
		except Exception as e:
			_log.exception("invoice_document_to_dict failed for document_id=%s business_id=%s", d.id, business_id)
			raise ApiError(
				"INVOICE_DATA_ERROR",
				f"خطا در بارگذاری فاکتور با شناسه {d.id}. داده‌های سند را بررسی کنید.",
				http_status=500,
			)

		# Tax workspace fields from extra_info
		try:
			extra = item.get("extra_info") or {}
		except Exception:
			extra = {}
		tax_workspace = bool(extra.get("tax_workspace"))
		tax_status = (extra.get("tax_status") or "").strip() if isinstance(extra.get("tax_status"), str) else extra.get("tax_status")
		if not tax_status:
			tax_status = "in_workspace" if tax_workspace else "not_in_workspace"
		item["tax_status"] = tax_status

		# Installment sale flag: اگر طرح اقساط روی سند وجود داشته باشد
		try:
			item["is_installment_sale"] = bool(
				isinstance(extra, dict) and isinstance(extra.get("installment_plan"), dict)
			)
		except Exception:
			item["is_installment_sale"] = False

		# total_amount from extra_info.totals.net if available
		total_amount = None
		try:
			totals = (item.get('extra_info') or {}).get('totals') or {}
			if isinstance(totals, dict) and 'net' in totals:
				total_amount = totals.get('net')
		except Exception:
			total_amount = None
		# Fallback compute from product lines
		if total_amount is None:
			try:
				net_sum = 0.0
				for pl in item.get('product_lines', []) or []:
					info = pl.get('extra_info') or {}
					qty = float(pl.get('quantity') or 0)
					unit_price = float(info.get('unit_price') or 0)
					line_discount = float(info.get('line_discount') or 0)
					tax_amount = float(info.get('tax_amount') or 0)
					line_total = info.get('line_total')
					if line_total is None:
						line_total = (qty * unit_price) - line_discount + tax_amount
					net_sum += float(line_total)
				total_amount = float(net_sum)
			except Exception:
				total_amount = None

		item['document_type_name'] = _type_name(item.get('document_type'))
		if total_amount is not None:
			item['total_amount'] = total_amount

		# مبلغ پرداخت‌شده و مبلغ باقی‌مانده فاکتور
		try:
			remaining_result = calculate_invoice_remaining(db, business_id, d.id)
			item["paid_amount"] = remaining_result.get("paid_amount")
			item["remaining_amount"] = remaining_result.get("remaining")
		except Exception:
			item["paid_amount"] = None
			item["remaining_amount"] = None

		# وضعیت اقساط برای فاکتورهای اقساطی: paid | partial | pending | overdue
		if item.get("is_installment_sale"):
			paid = float(item.get("paid_amount") or 0)
			remaining = float(item.get("remaining_amount") or 0)
			total_amt = float(item.get("total_amount") or 0)
			if total_amt <= 0 or remaining <= 0:
				item["installment_status"] = "paid"
			elif paid > 0:
				item["installment_status"] = "partial"
			else:
				item["installment_status"] = "pending"
			# remaining_total برای سازگاری با UI
			item["remaining_total"] = remaining
		else:
			item["installment_status"] = None
			item["remaining_total"] = None

		# افزودن counterparty
		_add_counterparty_to_invoice_item(db, item)

		data_items.append(format_datetime_fields(item, request))

	# مرتب‌سازی بر اساس مانده: برای فاکتورهای اقساطی اگر sort_by=remaining_amount باشد
	sort_by_val = getattr(query_info, 'sort_by', None) or body.get('sort_by')
	if is_installment_sale is True and sort_by_val == 'remaining_amount':
		data_items.sort(
			key=lambda x: float(x.get('remaining_amount') or 0),
			reverse=bool(getattr(query_info, 'sort_desc', True) or body.get('sort_desc', True)),
		)

	# Build pagination info
	page = (skip // take) + 1 if take > 0 else 1
	total_pages = (total + take - 1) // take if take > 0 else 1

	result = {
		"items": data_items,
		"total": total,
		"take": take,
		"skip": skip,
		# Optional standard pagination shape (supported by UI model)
		"pagination": {
			"page": page,
			"per_page": take,
			"total": total,
			"total_pages": total_pages,
		},
		# Flat shape too, for compatibility
		"page": page,
		"limit": take,
		"total_pages": total_pages,
	}
	
	# ذخیره در cache با tag-based caching
	if cache.enabled and cache_key:
		cache.set_with_invoices_tag(
			key=cache_key,
			value=result,
			business_id=business_id,
			fiscal_year_id=fiscal_year_id,
			document_type=document_type,
			project_id=project_id,
			ttl=30  # TTL کوتاه برای فاکتورها
		)
	
	return success_response(
		data=result,
		request=request,
		message="INVOICE_LIST",
	)

def _add_counterparty_to_invoice_item(db: Session, item: Dict[str, Any]) -> None:
    """افزودن فیلد counterparty به آیتم فاکتور بر اساس person_id در extra_info"""
    try:
        inv_type = str(item.get("document_type") or "")
        extra_info = item.get("extra_info") or {}
        person_id = extra_info.get("person_id")
        person_name = None
        if person_id is not None:
            try:
                p = db.query(Person).filter(Person.id == int(person_id)).first()
                if p is not None:
                    # اولویت: display_name > name > alias_name (که اجباری است)
                    person_name = getattr(p, "display_name", None) or getattr(p, "name", None)
                    if not person_name:
                        # اگر display_name و name خالی باشند، از alias_name استفاده می‌کنیم (که اجباری است)
                        person_name = getattr(p, "alias_name", None)
                    if not person_name and (getattr(p, "first_name", None) or getattr(p, "last_name", None)):
                        # اگر هنوز خالی است، از first_name و last_name استفاده می‌کنیم
                        name_parts = []
                        if getattr(p, "first_name", None):
                            name_parts.append(p.first_name)
                        if getattr(p, "last_name", None):
                            name_parts.append(p.last_name)
                        person_name = " ".join(name_parts) if name_parts else None
                    if not person_name and getattr(p, "company_name", None):
                        # اگر هنوز خالی است، از company_name استفاده می‌کنیم
                        person_name = p.company_name
            except Exception:
                person_name = None
        item["counterparty"] = person_name or ""
    except Exception:
        item["counterparty"] = ""


@router.post("/business/{business_id}/tax-workspace/search")
@require_business_access("business_id")
async def search_tax_workspace_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """لیست فاکتورهای موجود در کارپوشه مودیان با فیلتر و صفحه‌بندی."""
    from app.core.i18n import negotiate_locale

    # Base query: all invoice documents for business
    q = db.query(Document).filter(
        and_(
            Document.business_id == business_id,
            Document.document_type.in_(list(SUPPORTED_INVOICE_TYPES)),
        )
    )

    # Merge flat body extras
    body: Dict[str, Any] = {}
    try:
        body_json = await request.json()
        if isinstance(body_json, dict):
            body = body_json
    except Exception:
        body = {}

    search: Optional[str] = getattr(query_info, "search", None)
    search_fields: Optional[List[Any]] = getattr(query_info, "search_fields", None)
    q = _apply_invoice_list_text_search(
        q,
        business_id=business_id,
        search=search,
        search_fields=search_fields,
        extra_info_person_id_expr=None,
    )

    # Document type filter
    doc_type = body.get("document_type")
    if isinstance(doc_type, str) and doc_type in SUPPORTED_INVOICE_TYPES:
        q = q.filter(Document.document_type == doc_type)

    # Proforma filter
    is_proforma = body.get("is_proforma")
    if isinstance(is_proforma, bool):
        q = q.filter(Document.is_proforma == is_proforma)

    # Currency filter
    currency_id = body.get("currency_id")
    try:
        if currency_id is not None:
            q = q.filter(Document.currency_id == int(currency_id))
    except Exception:
        pass

    # Fiscal year from header or body
    fiscal_year_id = None
    try:
        fy_header = request.headers.get("X-Fiscal-Year-ID")
        if fy_header:
            fiscal_year_id = int(fy_header)
    except Exception:
        fiscal_year_id = None
    if fiscal_year_id is None:
        try:
            if body.get("fiscal_year_id") is not None:
                fiscal_year_id = int(body.get("fiscal_year_id"))
        except Exception:
            fiscal_year_id = None
    if fiscal_year_id is not None:
        q = q.filter(Document.fiscal_year_id == fiscal_year_id)

    # Date range filters from QueryInfo.filters
    try:
        filters = getattr(query_info, "filters", None)
    except Exception:
        filters = None
    if filters and isinstance(filters, (list, tuple)):
        for flt in filters:
            try:
                prop = getattr(flt, "property", None) if not isinstance(flt, dict) else flt.get("property")
                op = getattr(flt, "operator", None) if not isinstance(flt, dict) else flt.get("operator")
                val = getattr(flt, "value", None) if not isinstance(flt, dict) else flt.get("value")
                if prop == "document_date" and isinstance(val, str) and val:
                    from app.services.transfer_service import _parse_iso_date as _p

                    dt = _p(val)
                    col = getattr(Document, prop)
                    if op == ">=":
                        q = q.filter(col >= dt)
                    elif op == "<=":
                        q = q.filter(col <= dt)
            except Exception:
                pass

    # Date range from flat body
    q = apply_invoice_search_ordering(q, query_info)

    # Fetch all candidates and filter by workspace/tax_status in Python
    all_docs: List[Document] = q.all()
    requested_status = body.get("tax_status")
    requested_status = requested_status.strip() if isinstance(requested_status, str) else None

    workspace_docs: List[Document] = []
    for d in all_docs:
        extra = d.extra_info or {}
        in_workspace = bool(extra.get("tax_workspace"))
        if not in_workspace:
            continue
        status = extra.get("tax_status")
        if isinstance(status, str):
            status = status.strip()
        if not status:
            status = "not_sent"
        if requested_status and status != requested_status:
            continue
        workspace_docs.append(d)

    # Pagination (after workspace filter)
    take = int(getattr(query_info, "take", 20) or 20)
    skip = int(getattr(query_info, "skip", 0) or 0)
    total = len(workspace_docs)
    page_docs = workspace_docs[skip : skip + take] if take > 0 else workspace_docs

    # Locale for type names
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"

    def _type_name(tp: str) -> str:
        mapping = {
            "invoice_sales": ("فروش" if is_fa else "Sales"),
            "invoice_sales_return": ("برگشت از فروش" if is_fa else "Sales return"),
            "invoice_purchase": ("خرید" if is_fa else "Purchase"),
            "invoice_purchase_return": ("برگشت از خرید" if is_fa else "Purchase return"),
            "invoice_direct_consumption": ("مصرف مستقیم" if is_fa else "Direct consumption"),
            "invoice_production": ("تولید" if is_fa else "Production"),
            "invoice_waste": ("ضایعات" if is_fa else "Waste"),
        }
        return mapping.get(str(tp), str(tp))

    data_items: List[Dict[str, Any]] = []
    for d in page_docs:
        item = invoice_document_to_dict(db, d)
        extra = item.get("extra_info") or {}
        tax_status = extra.get("tax_status")
        if isinstance(tax_status, str):
            tax_status = tax_status.strip()
        if not tax_status:
            tax_status = "not_sent"
        item["tax_status"] = tax_status
        item["tax_tracking_code"] = extra.get("tax_tracking_code")
        item["tax_last_send_at"] = extra.get("tax_last_send_at")

        # total_amount from totals.net or recomputed
        total_amount = None
        try:
            totals = (item.get("extra_info") or {}).get("totals") or {}
            if isinstance(totals, dict) and "net" in totals:
                total_amount = totals.get("net")
        except Exception:
            total_amount = None
        if total_amount is None:
            try:
                net_sum = 0.0
                for pl in item.get("product_lines", []) or []:
                    info = pl.get("extra_info") or {}
                    qty = float(pl.get("quantity") or 0)
                    unit_price = float(info.get("unit_price") or 0)
                    line_discount = float(info.get("line_discount") or 0)
                    tax_amount = float(info.get("tax_amount") or 0)
                    line_total = info.get("line_total")
                    if line_total is None:
                        line_total = (qty * unit_price) - line_discount + tax_amount
                    net_sum += float(line_total)
                total_amount = float(net_sum)
            except Exception:
                total_amount = None

        item["document_type_name"] = _type_name(item.get("document_type"))
        if total_amount is not None:
            item["total_amount"] = total_amount
        
        # افزودن counterparty
        _add_counterparty_to_invoice_item(db, item)
        
        data_items.append(format_datetime_fields(item, request))

    page = (skip // take) + 1 if take > 0 else 1
    total_pages = (total + take - 1) // take if take > 0 else 1

    return success_response(
        data={
            "items": data_items,
            "total": total,
            "take": take,
            "skip": skip,
            "pagination": {
                "page": page,
                "per_page": take,
                "total": total,
                "total_pages": total_pages,
            },
            "page": page,
            "limit": take,
            "total_pages": total_pages,
        },
        request=request,
        message="INVOICE_TAX_WORKSPACE_LIST",
    )

def _get_invoice_for_business(
    db: Session,
    business_id: int,
    invoice_id: int,
) -> Document:
    doc = db.query(Document).filter(Document.id == invoice_id).first()
    if not doc or doc.business_id != business_id or doc.document_type not in SUPPORTED_INVOICE_TYPES:
        raise ApiError("DOCUMENT_NOT_FOUND", "Invoice document not found", http_status=404)
    return doc


def _ensure_sales_or_return(doc: Document) -> None:
    if doc.document_type not in ("invoice_sales", "invoice_sales_return"):
        raise ApiError(
            "TAX_WORKSPACE_NOT_ALLOWED",
            "Only sales and sales-return invoices can be added to tax workspace",
            http_status=400,
        )
    if doc.is_proforma:
        raise ApiError(
            "TAX_WORKSPACE_NOT_ALLOWED",
            "Proforma invoices cannot be added to tax workspace",
            http_status=400,
        )


@router.post("/business/{business_id}/{invoice_id}/tax-workspace/add")
@require_business_access("business_id")
def add_invoice_to_tax_workspace(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """افزودن فاکتور به کارپوشه مودیان."""
    doc = _get_invoice_for_business(db, business_id, invoice_id)
    _ensure_sales_or_return(doc)

    extra = dict(doc.extra_info or {})
    extra["tax_workspace"] = True
    status = extra.get("tax_status")
    if not isinstance(status, str) or not status.strip():
        extra["tax_status"] = "not_sent"
    doc.extra_info = extra
    db.commit()
    db.refresh(doc)

    return success_response(
        data={"id": doc.id, "tax_status": extra.get("tax_status")},
        request=request,
        message="INVOICE_ADDED_TO_TAX_WORKSPACE",
    )


@router.post("/business/{business_id}/{invoice_id}/tax-workspace/remove")
@require_business_access("business_id")
def remove_invoice_from_tax_workspace(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """حذف فاکتور از کارپوشه مودیان (فقط اگر قطعی نشده باشد)."""
    doc = _get_invoice_for_business(db, business_id, invoice_id)
    extra = dict(doc.extra_info or {})
    status = (extra.get("tax_status") or "").strip() if isinstance(extra.get("tax_status"), str) else extra.get("tax_status")
    if status in ("sent", "finalized"):
        raise ApiError(
            "TAX_WORKSPACE_REMOVE_NOT_ALLOWED",
            "Cannot remove invoice that has been sent to tax system from tax workspace",
            http_status=409,
        )

    extra["tax_workspace"] = False
    # Optional: mark as not in workspace
    extra["tax_status"] = status or "not_in_workspace"
    doc.extra_info = extra
    db.commit()
    db.refresh(doc)

    return success_response(
        data={"id": doc.id, "tax_status": extra.get("tax_status")},
        request=request,
        message="INVOICE_REMOVED_FROM_TAX_WORKSPACE",
    )

@router.post("/business/{business_id}/{invoice_id}/tax-workspace/send-to-system")
@require_business_access("business_id")
def send_invoice_to_tax_system(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    ارسال تکی فاکتور موجود در کارپوشه به سامانه مودیان.
    (نسخه MVP: فقط بروزرسانی وضعیت و کد رهگیری آزمایشی)
    """
    doc = _get_invoice_for_business(db, business_id, invoice_id)
    _ensure_sales_or_return(doc)

    extra = dict(doc.extra_info or {})
    if not bool(extra.get("tax_workspace")):
        raise ApiError(
            "TAX_WORKSPACE_NOT_SET",
            "Invoice is not in tax workspace",
            http_status=400,
        )
    status = (extra.get("tax_status") or "").strip() if isinstance(extra.get("tax_status"), str) else extra.get("tax_status")
    if status in ("sent", "finalized"):
        raise ApiError(
            "TAX_ALREADY_SENT",
            "Invoice has already been sent to tax system",
            http_status=409,
        )

    send_document_to_tax_system(db, doc)
    db.commit()
    db.refresh(doc)

    return success_response(
        data={"id": doc.id, "tax_status": (doc.extra_info or {}).get("tax_status")},
        request=request,
        message="INVOICE_SENT_TO_TAX_SYSTEM",
    )


@router.post("/business/{business_id}/tax-workspace/send-to-system-batch")
@require_business_access("business_id")
def send_invoices_to_tax_system_batch(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    ارسال گروهی فاکتورهای موجود در کارپوشه به سامانه مودیان.
    (MVP: شبیه‌سازی ارسال و بروزرسانی وضعیت)
    """
    ids = body.get("invoice_ids") or []
    if not isinstance(ids, list) or not ids:
        raise ApiError("INVALID_REQUEST", "invoice_ids must be a non-empty list", http_status=400)

    succeeded: List[int] = []
    failed: List[Dict[str, Any]] = []

    for raw_id in ids:
        try:
            invoice_id = int(raw_id)
        except Exception:
            failed.append({"id": raw_id, "error": "INVALID_ID"})
            continue
        try:
            doc = _get_invoice_for_business(db, business_id, invoice_id)
            _ensure_sales_or_return(doc)
            extra = dict(doc.extra_info or {})
            if not bool(extra.get("tax_workspace")):
                raise ApiError("TAX_WORKSPACE_NOT_SET", "Invoice is not in tax workspace", http_status=400)
            status = (extra.get("tax_status") or "").strip() if isinstance(extra.get("tax_status"), str) else extra.get("tax_status")
            if status in ("sent", "finalized"):
                raise ApiError("TAX_ALREADY_SENT", "Invoice has already been sent to tax system", http_status=409)

            send_document_to_tax_system(db, doc)
            succeeded.append(invoice_id)
        except ApiError as e:
            error_detail = getattr(e, "detail", {}) or {}
            error_info = error_detail.get("error") if isinstance(error_detail, dict) else {}
            failed.append(
                {
                    "id": invoice_id,
                    "error": (error_info or {}).get("code") or str(e),
                    "message": (error_info or {}).get("message"),
                    "issues": (error_info or {}).get("details", {}).get("issues")
                    if isinstance((error_info or {}).get("details"), dict)
                    else None,
                }
            )
        except Exception as e:
            failed.append({"id": invoice_id, "error": str(e)})

    db.commit()

    return success_response(
        data={"succeeded": succeeded, "failed": failed},
        request=request,
        message="INVOICE_BATCH_SENT_TO_TAX_SYSTEM",
    )


@router.post("/business/{business_id}/tax-workspace/remove-batch")
@require_business_access("business_id")
def remove_invoices_from_tax_workspace_batch(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """حذف گروهی فاکتورها از کارپوشه مودیان (صرفاً غیرقطعی‌ها)."""
    ids = body.get("invoice_ids") or []
    if not isinstance(ids, list) or not ids:
        raise ApiError("INVALID_REQUEST", "invoice_ids must be a non-empty list", http_status=400)

    removed: List[int] = []
    failed: List[Dict[str, Any]] = []

    for raw_id in ids:
        try:
            invoice_id = int(raw_id)
        except Exception:
            failed.append({"id": raw_id, "error": "INVALID_ID"})
            continue
        try:
            doc = _get_invoice_for_business(db, business_id, invoice_id)
            extra = dict(doc.extra_info or {})
            status = (extra.get("tax_status") or "").strip() if isinstance(extra.get("tax_status"), str) else extra.get("tax_status")
            if status in ("sent", "finalized"):
                raise ApiError("TAX_WORKSPACE_REMOVE_NOT_ALLOWED", "Cannot remove invoice that has been sent to tax system", http_status=409)
            extra["tax_workspace"] = False
            extra["tax_status"] = status or "not_in_workspace"
            doc.extra_info = extra
            db.add(doc)
            removed.append(invoice_id)
        except ApiError as e:
            failed.append({"id": invoice_id, "error": e.detail.get("error", {}).get("code")})
        except Exception as e:
            failed.append({"id": invoice_id, "error": str(e)})

    db.commit()

    return success_response(
        data={"removed": removed, "failed": failed},
        request=request,
        message="INVOICE_BATCH_REMOVED_FROM_TAX_WORKSPACE",
    )


@router.get("/business/{business_id}/tax-workspace/health")
@require_business_access("business_id")
def get_tax_system_health(
    request: Request,
    business_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """بررسی سلامت سامانه مالیاتی"""
    from app.services.tax_health_check import check_tax_system_health
    
    health = check_tax_system_health(db, business_id)
    
    return success_response(
        data=health,
        request=request,
        message="TAX_SYSTEM_HEALTH_CHECK",
    )


@router.get("/business/{business_id}/tax-workspace/failed-invoices")
@require_business_access("business_id")
def get_failed_invoices(
    request: Request,
    business_id: int,
    status: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """دریافت لیست فاکتورهای failed از Dead Letter Queue"""
    from app.services.tax_dead_letter_queue import get_failed_invoices
    
    failed = get_failed_invoices(db, business_id, status, limit, offset)
    
    return success_response(
        data={
            "items": [
                {
                    "id": item.id,
                    "invoice_id": item.invoice_id,
                    "tracking_code": item.tracking_code,
                    "error_code": item.error_code,
                    "error_message": item.error_message,
                    "attempt_count": item.attempt_count,
                    "first_failed_at": item.first_failed_at.isoformat() if item.first_failed_at else None,
                    "last_attempt_at": item.last_attempt_at.isoformat() if item.last_attempt_at else None,
                    "status": item.status,
                }
                for item in failed
            ],
            "total": len(failed),
        },
        request=request,
        message="TAX_FAILED_INVOICES_LIST",
    )


@router.post("/business/{business_id}/tax-workspace/failed-invoices/{failed_id}/retry")
@require_business_access("business_id")
def retry_failed_invoice_endpoint(
    request: Request,
    business_id: int,
    failed_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """تلاش مجدد برای ارسال فاکتور failed"""
    from app.services.tax_dead_letter_queue import retry_failed_invoice
    
    result = retry_failed_invoice(db, failed_id)
    db.commit()
    
    return success_response(
        data=result,
        request=request,
        message="TAX_FAILED_INVOICE_RETRY_COMPLETED",
    )


@router.get("/business/{business_id}/{invoice_id}/tax-timeline")
@require_business_access("business_id")
def get_invoice_tax_timeline(
    request: Request,
    business_id: int,
    invoice_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """دریافت Timeline تغییرات وضعیت مالیاتی فاکتور"""
    doc = _get_invoice_for_business(db, business_id, invoice_id)
    extra = dict(doc.extra_info or {})
    
    timeline = []
    
    # تاریخچه از extra_info
    if extra.get("tax_last_send_at"):
        timeline.append({
            "event": "send_attempt",
            "timestamp": extra.get("tax_last_send_at"),
            "status": extra.get("tax_status"),
            "tracking_code": extra.get("tax_tracking_code"),
        })
    
    if extra.get("tax_last_inquiry_at"):
        timeline.append({
            "event": "status_inquiry",
            "timestamp": extra.get("tax_last_inquiry_at"),
            "status": extra.get("tax_status"),
            "error_message": extra.get("tax_error_message"),
        })
    
    # بررسی Dead Letter Queue
    from app.services.tax_dead_letter_queue import get_failed_invoices
    failed_items = get_failed_invoices(db, business_id, status=None, limit=1000, offset=0)
    for failed in failed_items:
        if failed.invoice_id == invoice_id:
            timeline.append({
                "event": "failed",
                "timestamp": failed.first_failed_at.isoformat() if failed.first_failed_at else None,
                "error_code": failed.error_code,
                "error_message": failed.error_message,
                "attempt_count": failed.attempt_count,
            })
    
    # مرتب‌سازی بر اساس زمان
    timeline.sort(key=lambda x: x.get("timestamp") or "", reverse=True)
    
    return success_response(
        data={
            "invoice_id": invoice_id,
            "current_status": extra.get("tax_status"),
            "timeline": timeline,
        },
        request=request,
        message="TAX_TIMELINE_RETRIEVED",
    )


@router.post("/business/{business_id}/tax-workspace/inquire-status")
@require_business_access("business_id")
def inquire_tax_status_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "view")),
) -> Dict[str, Any]:
    invoice_ids = body.get("invoice_ids")
    tracking_codes = body.get("tracking_codes")
    if not isinstance(invoice_ids, list):
        invoice_ids = []
    if not isinstance(tracking_codes, list):
        tracking_codes = []
    if not invoice_ids and not tracking_codes:
        raise ApiError("INVALID_REQUEST", "لیست فاکتور یا کد رهگیری لازم است.", http_status=400)

    result = inquire_tax_status(
        db,
        business_id,
        invoice_ids=[
            int(i)
            for i in invoice_ids
            if i is not None and str(i).isdigit()
        ],
        tracking_codes=[str(code) for code in tracking_codes if code],
    )
    db.commit()
    return success_response(
        data=result,
        request=request,
        message="TAX_STATUS_INQUIRY_COMPLETED",
    )


@router.post("/business/{business_id}/tax-workspace/validate-batch")
@require_business_access("business_id")
def validate_invoices_for_tax_batch(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """اعتبارسنجی گروهی فاکتورها قبل از ارسال"""
    ids = body.get("invoice_ids") or []
    if not isinstance(ids, list) or not ids:
        raise ApiError("INVALID_REQUEST", "invoice_ids must be a non-empty list", http_status=400)
    
    validated: List[int] = []
    invalid: List[Dict[str, Any]] = []
    
    for raw_id in ids:
        try:
            invoice_id = int(raw_id)
        except Exception:
            invalid.append({"id": raw_id, "error": "INVALID_ID", "message": "شناسه نامعتبر"})
            continue
        
        try:
            doc = _get_invoice_for_business(db, business_id, invoice_id)
            _ensure_sales_or_return(doc)
            
            # اعتبارسنجی
            from app.services.tax_validation_service import validate_document_for_tax
            validation = validate_document_for_tax(db, doc)
            
            if not validation["valid"]:
                invalid.append({
                    "id": invoice_id,
                    "error": "VALIDATION_FAILED",
                    "message": "فاکتور حداقل الزامات سامانه مودیان را ندارد.",
                    "issues": validation["issues"],
                })
            else:
                validated.append(invoice_id)
                
        except ApiError as e:
            error_detail = getattr(e, "detail", {}) or {}
            error_info = error_detail.get("error") if isinstance(error_detail, dict) else {}
            invalid.append({
                "id": invoice_id,
                "error": (error_info or {}).get("code") or str(e),
                "message": (error_info or {}).get("message") or str(e),
            })
        except Exception as e:
            invalid.append({
                "id": invoice_id,
                "error": "UNKNOWN_ERROR",
                "message": str(e),
            })
    
    return success_response(
        data={
            "validated": validated,
            "invalid": invalid,
            "total": len(ids),
            "valid_count": len(validated),
            "invalid_count": len(invalid),
        },
        request=request,
        message="TAX_VALIDATION_BATCH_COMPLETED",
    )


@router.post("/business/{business_id}/tax-workspace/quick-actions")
@require_business_access("business_id")
def tax_workspace_quick_actions(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "add")),
) -> Dict[str, Any]:
    """عملیات سریع برای کارپوشه مالیاتی"""
    action = body.get("action")
    if not action or not isinstance(action, str):
        raise ApiError("INVALID_REQUEST", "action is required", http_status=400)
    
    action = action.lower()
    
    # دریافت فاکتورها بر اساس وضعیت
    from sqlalchemy import func
    
    # برای PostgreSQL از JSON operators استفاده می‌کنیم
    from sqlalchemy import cast, Boolean
    base_query = db.query(Document).filter(
        and_(
            Document.business_id == business_id,
            Document.document_type.in_(list(SUPPORTED_INVOICE_TYPES)),
            cast(Document.extra_info['tax_workspace'], Boolean) == True,
        )
    )
    
    invoice_ids: List[int] = []
    
    if action == "send_all_pending":
        # ارسال همه pending
        docs = base_query.filter(
            Document.extra_info['tax_status'].astext == 'pending'
        ).all()
        invoice_ids = [doc.id for doc in docs]
        
        if not invoice_ids:
            return success_response(
                data={"message": "هیچ فاکتور pending یافت نشد", "invoice_ids": []},
                request=request,
                message="NO_PENDING_INVOICES",
            )
        
        # استفاده از endpoint ارسال گروهی
        return send_invoices_to_tax_system_batch(request, business_id, {"invoice_ids": invoice_ids}, ctx, db)
        
    elif action == "inquire_all_sent":
        # استعلام همه sent
        docs = base_query.filter(
            Document.extra_info['tax_status'].astext == 'sent'
        ).all()
        
        tracking_codes: List[str] = []
        for doc in docs:
            code = (doc.extra_info or {}).get("tax_tracking_code")
            if code and isinstance(code, str) and code.strip():
                tracking_codes.append(code.strip())
        
        if not tracking_codes:
            return success_response(
                data={"message": "هیچ فاکتور sent با کد رهگیری یافت نشد", "tracking_codes": []},
                request=request,
                message="NO_SENT_INVOICES",
            )
        
        # استفاده از endpoint استعلام
        return inquire_tax_status_endpoint(request, business_id, {"tracking_codes": tracking_codes}, ctx, db)
        
    elif action == "retry_all_failed":
        # Retry همه failed
        docs = base_query.filter(
            Document.extra_info['tax_status'].astext == 'failed'
        ).all()
        invoice_ids = [doc.id for doc in docs]
        
        if not invoice_ids:
            return success_response(
                data={"message": "هیچ فاکتور failed یافت نشد", "invoice_ids": []},
                request=request,
                message="NO_FAILED_INVOICES",
            )
        
        # استفاده از endpoint ارسال گروهی
        return send_invoices_to_tax_system_batch(request, business_id, {"invoice_ids": invoice_ids}, ctx, db)
        
    else:
        raise ApiError("INVALID_ACTION", f"Unknown action: {action}", http_status=400)


@router.post(
    "/business/{business_id}/export/excel",
    summary="خروجی Excel لیست فاکتورها",
    description="خروجی Excel لیست فاکتورها با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_invoices_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.core.i18n import negotiate_locale

    # Build base query similar to search endpoint
    take_value = min(int(body.get("take", 1000)), 10000)
    skip_value = int(body.get("skip", 0))

    q = db.query(Document).filter(
        and_(
            Document.business_id == business_id,
            Document.document_type.in_(list(SUPPORTED_INVOICE_TYPES)),
        )
    )

    search = body.get("search")
    search_fields = body.get("search_fields")
    sf_list = search_fields if isinstance(search_fields, list) else None
    q = _apply_invoice_list_text_search(
        q,
        business_id=business_id,
        search=search if isinstance(search, str) else None,
        search_fields=sf_list,
        extra_info_person_id_expr=None,
    )

    # Filters
    doc_type = body.get("document_type")
    if isinstance(doc_type, str) and doc_type in SUPPORTED_INVOICE_TYPES:
        q = q.filter(Document.document_type == doc_type)

    is_proforma = body.get("is_proforma")
    if isinstance(is_proforma, bool):
        q = q.filter(Document.is_proforma == is_proforma)

    currency_id = body.get("currency_id")
    try:
        if currency_id is not None:
            q = q.filter(Document.currency_id == int(currency_id))
    except Exception:
        pass

    # Fiscal year
    try:
        fy_header = request.headers.get("X-Fiscal-Year-ID")
        if fy_header:
            q = q.filter(Document.fiscal_year_id == int(fy_header))
        elif body.get("fiscal_year_id") is not None:
            q = q.filter(Document.fiscal_year_id == int(body.get("fiscal_year_id")))
    except Exception:
        pass

    # Project filter
    project_id = body.get("project_id")
    try:
        if project_id is not None:
            q = q.filter(Document.project_id == int(project_id))
    except Exception:
        pass

    # Date range
    from app.services.transfer_service import _parse_iso_date as _p
    if isinstance(body.get("from_date"), str):
        try:
            q = q.filter(Document.document_date >= _p(body.get("from_date")))
        except Exception:
            pass
    if isinstance(body.get("to_date"), str):
        try:
            q = q.filter(Document.document_date <= _p(body.get("to_date")))
        except Exception:
            pass

    q = apply_invoice_search_ordering_from_body(q, body)

    total = q.count()
    docs: List[Document] = q.offset(skip_value).limit(take_value).all()

    # Build items like list endpoint
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    calendar_type = request.state.calendar_type

    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""

    def _type_name(tp: str) -> str:
        mapping = {
            'invoice_sales': ('فروش' if is_fa else 'Sales'),
            'invoice_sales_return': ('برگشت از فروش' if is_fa else 'Sales return'),
            'invoice_purchase': ('خرید' if is_fa else 'Purchase'),
            'invoice_purchase_return': ('برگشت از خرید' if is_fa else 'Purchase return'),
            'invoice_direct_consumption': ('مصرف مستقیم' if is_fa else 'Direct consumption'),
            'invoice_production': ('تولید' if is_fa else 'Production'),
            'invoice_waste': ('ضایعات' if is_fa else 'Waste'),
        }
        return mapping.get(str(tp), str(tp))

    items: List[Dict[str, Any]] = []
    for d in docs:
        item = invoice_document_to_dict(db, d)
        # total_amount
        total_amount = None
        try:
            totals = (item.get('extra_info') or {}).get('totals') or {}
            if isinstance(totals, dict) and 'net' in totals:
                total_amount = totals.get('net')
        except Exception:
            total_amount = None
        if total_amount is None:
            try:
                net_sum = 0.0
                for pl in item.get('product_lines', []) or []:
                    info = pl.get('extra_info') or {}
                    qty = float(pl.get('quantity') or 0)
                    unit_price = float(info.get('unit_price') or 0)
                    line_discount = float(info.get('line_discount') or 0)
                    tax_amount = float(info.get('tax_amount') or 0)
                    line_total = info.get('line_total')
                    if line_total is None:
                        line_total = (qty * unit_price) - line_discount + tax_amount
                    net_sum += float(line_total)
                total_amount = float(net_sum)
            except Exception:
                total_amount = None

        item['document_type_name'] = _type_name(item.get('document_type'))
        if total_amount is not None:
            item['total_amount'] = total_amount
        
        # افزودن counterparty
        _add_counterparty_to_invoice_item(db, item)
        
        items.append(format_datetime_fields(item, request))

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        # Check if document_type_name exists in export_columns
        has_document_type_name = any(col.get('key') == 'document_type_name' for col in export_columns)
        
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
        
        # If document_type_name is missing, add it at the beginning (after code if exists)
        if not has_document_type_name:
            # Find position after 'code' if it exists
            code_index = keys.index('code') if 'code' in keys else -1
            insert_pos = code_index + 1 if code_index >= 0 else 0
            keys.insert(insert_pos, 'document_type_name')
            headers.insert(insert_pos, 'نوع فاکتور' if is_fa else 'Invoice type')
    else:
        default_columns = [
            ('code', 'کد سند' if is_fa else 'Code'),
            ('document_type_name', 'نوع فاکتور' if is_fa else 'Invoice type'),
            ('counterparty', 'طرف حساب' if is_fa else 'Counterparty'),
            ('document_date', 'تاریخ سند' if is_fa else 'Document date'),
            ('total_amount', 'مبلغ کل' if is_fa else 'Total amount'),
            ('currency_code', 'ارز' if is_fa else 'Currency'),
            ('created_by_name', 'ایجادکننده' if is_fa else 'Created by'),
            ('is_proforma', 'پیش‌فاکتور' if is_fa else 'Proforma'),
            ('registered_at', 'تاریخ ثبت' if is_fa else 'Registered at'),
        ]
        for key, label in default_columns:
            if items and key in items[0]:
                keys.append(key)
                headers.append(label)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            
            # Format date fields based on calendar type
            if key in ('document_date', 'registered_at', 'created_at', 'due_date'):
                value = format_date_for_export(item, key)
            elif isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            ws.cell(row=row_idx, column=col_idx, value=value).border = border

    # Auto-width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "invoices"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/business/{business_id}/export/pdf",
    summary="خروجی PDF لیست فاکتورها",
    description="خروجی PDF لیست فاکتورها با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_invoices_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from html import escape

    # Build same list as excel
    take_value = min(int(body.get("take", 1000)), 10000)
    skip_value = int(body.get("skip", 0))

    q = db.query(Document).filter(
        and_(
            Document.business_id == business_id,
            Document.document_type.in_(list(SUPPORTED_INVOICE_TYPES)),
        )
    )

    search = body.get("search")
    search_fields = body.get("search_fields")
    sf_list = search_fields if isinstance(search_fields, list) else None
    q = _apply_invoice_list_text_search(
        q,
        business_id=business_id,
        search=search if isinstance(search, str) else None,
        search_fields=sf_list,
        extra_info_person_id_expr=None,
    )

    doc_type = body.get("document_type")
    if isinstance(doc_type, str) and doc_type in SUPPORTED_INVOICE_TYPES:
        q = q.filter(Document.document_type == doc_type)

    is_proforma = body.get("is_proforma")
    if isinstance(is_proforma, bool):
        q = q.filter(Document.is_proforma == is_proforma)

    currency_id = body.get("currency_id")
    try:
        if currency_id is not None:
            q = q.filter(Document.currency_id == int(currency_id))
    except Exception:
        pass

    try:
        fy_header = request.headers.get("X-Fiscal-Year-ID")
        if fy_header:
            q = q.filter(Document.fiscal_year_id == int(fy_header))
        elif body.get("fiscal_year_id") is not None:
            q = q.filter(Document.fiscal_year_id == int(body.get("fiscal_year_id")))
    except Exception:
        pass

    # Project filter
    project_id = body.get("project_id")
    try:
        if project_id is not None:
            q = q.filter(Document.project_id == int(project_id))
    except Exception:
        pass

    from app.services.transfer_service import _parse_iso_date as _p
    if isinstance(body.get("from_date"), str):
        try:
            q = q.filter(Document.document_date >= _p(body.get("from_date")))
        except Exception:
            pass
    if isinstance(body.get("to_date"), str):
        try:
            q = q.filter(Document.document_date <= _p(body.get("to_date")))
        except Exception:
            pass

    q = apply_invoice_search_ordering_from_body(q, body)

    docs: List[Document] = q.offset(skip_value).limit(take_value).all()

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'

    def _type_name(tp: str) -> str:
        mapping = {
            'invoice_sales': ('فروش' if is_fa else 'Sales'),
            'invoice_sales_return': ('برگشت از فروش' if is_fa else 'Sales return'),
            'invoice_purchase': ('خرید' if is_fa else 'Purchase'),
            'invoice_purchase_return': ('برگشت از خرید' if is_fa else 'Purchase return'),
            'invoice_direct_consumption': ('مصرف مستقیم' if is_fa else 'Direct consumption'),
            'invoice_production': ('تولید' if is_fa else 'Production'),
            'invoice_waste': ('ضایعات' if is_fa else 'Waste'),
        }
        return mapping.get(str(tp), str(tp))

    items: List[Dict[str, Any]] = []
    # Helper to resolve person display name
    def _get_person_display_name(person_id: int | None) -> str | None:
        if person_id is None:
            return None
        try:
            p = db.query(Person).filter(Person.id == int(person_id)).first()
            if p is None:
                return None
            # اولویت: display_name > name > alias_name (که اجباری است)
            person_name = getattr(p, "display_name", None) or getattr(p, "name", None)
            if not person_name:
                # اگر display_name و name خالی باشند، از alias_name استفاده می‌کنیم (که اجباری است)
                person_name = getattr(p, "alias_name", None)
            if not person_name and (getattr(p, "first_name", None) or getattr(p, "last_name", None)):
                # اگر هنوز خالی است، از first_name و last_name استفاده می‌کنیم
                name_parts = []
                if getattr(p, "first_name", None):
                    name_parts.append(p.first_name)
                if getattr(p, "last_name", None):
                    name_parts.append(p.last_name)
                person_name = " ".join(name_parts) if name_parts else None
            if not person_name and getattr(p, "company_name", None):
                # اگر هنوز خالی است، از company_name استفاده می‌کنیم
                person_name = p.company_name
            return person_name
        except Exception:
            return None
    for d in docs:
        item = invoice_document_to_dict(db, d)
        total_amount = None
        try:
            totals = (item.get('extra_info') or {}).get('totals') or {}
            if isinstance(totals, dict) and 'net' in totals:
                total_amount = totals.get('net')
        except Exception:
            total_amount = None
        if total_amount is None:
            try:
                net_sum = 0.0
                for pl in item.get('product_lines', []) or []:
                    info = pl.get('extra_info') or {}
                    qty = float(pl.get('quantity') or 0)
                    unit_price = float(info.get('unit_price') or 0)
                    line_discount = float(info.get('line_discount') or 0)
                    tax_amount = float(info.get('tax_amount') or 0)
                    line_total = info.get('line_total')
                    if line_total is None:
                        line_total = (qty * unit_price) - line_discount + tax_amount
                    net_sum += float(line_total)
                total_amount = float(net_sum)
            except Exception:
                total_amount = None
        item['document_type_name'] = _type_name(item.get('document_type'))
        if total_amount is not None:
            item['total_amount'] = total_amount
        # Counterparty based on type: sales -> buyer (person), purchase -> seller (person)
        try:
            inv_type = str(item.get("document_type") or "")
            extra = item.get("extra_info") or {}
            person_id = extra.get("person_id")
            person_name = _get_person_display_name(person_id)
            counterparty = ""
            if inv_type in ("invoice_sales", "invoice_sales_return"):
                counterparty = person_name or ""
            elif inv_type in ("invoice_purchase", "invoice_purchase_return"):
                counterparty = person_name or ""
            else:
                counterparty = person_name or ""
            item["counterparty"] = counterparty
        except Exception:
            item["counterparty"] = ""
        items.append(format_datetime_fields(item, request))

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        # Check if document_type_name exists in export_columns
        has_document_type_name = any(col.get('key') == 'document_type_name' for col in export_columns)
        
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
        
        # If document_type_name is missing, add it at the beginning (after code if exists)
        if not has_document_type_name:
            # Find position after 'code' if it exists
            code_index = keys.index('code') if 'code' in keys else -1
            insert_pos = code_index + 1 if code_index >= 0 else 0
            keys.insert(insert_pos, 'document_type_name')
            headers.insert(insert_pos, 'نوع فاکتور' if is_fa else 'Invoice type')
    else:
        default_columns = [
            ('code', 'کد سند' if is_fa else 'Code'),
            ('document_type_name', 'نوع فاکتور' if is_fa else 'Invoice type'),
            ('counterparty', 'طرف حساب' if is_fa else 'Counterparty'),
            ('document_date', 'تاریخ سند' if is_fa else 'Document date'),
            ('total_amount', 'مبلغ کل' if is_fa else 'Total amount'),
            ('currency_code', 'ارز' if is_fa else 'Currency'),
            ('created_by_name', 'ایجادکننده' if is_fa else 'Created by'),
            ('is_proforma', 'پیش‌فاکتور' if is_fa else 'Proforma'),
            ('registered_at', 'تاریخ ثبت' if is_fa else 'Registered at'),
        ]
        for key, label in default_columns:
            if items and key in items[0]:
                keys.append(key)
                headers.append(label)

    # Business name & locale
    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    # respect user's calendar for generated_at
    try:
        cal_header = (request.headers.get("X-Calendar-Type") or "").strip().lower()
        cal_type = cal_header or ("jalali" if is_fa else "gregorian")
    except Exception:
        cal_type = "jalali" if is_fa else "gregorian"
    try:
        _now = datetime.datetime.now()
        _fd = CalendarConverter.format_datetime(_now, cal_type)
        now = _fd.get("formatted") or _fd.get("date_only") or _now.strftime('%Y/%m/%d %H:%M')
    except Exception:
        now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    title_text = "لیست فاکتورها" if is_fa else "Invoices List"
    label_biz = "کسب و کار" if is_fa else "Business"
    label_date = "تاریخ تولید" if is_fa else "Generated Date"
    footer_text = f"تولید شده در {now}" if is_fa else f"Generated at {now}"

    headers_html = ''.join(f'<th>{escape(header)}</th>' for header in headers)

    # Determine calendar type for date formatting in table rows
    try:
        cal_header = (request.headers.get("X-Calendar-Type") or "").strip().lower()
        cal_type_for_rows = cal_header or ("jalali" if is_fa else "gregorian")
    except Exception:
        cal_type_for_rows = "jalali" if is_fa else "gregorian"

    # Helper to format date string using CalendarConverter
    def _format_date_for_calendar(value: str) -> str:
        try:
            dt = datetime.datetime.fromisoformat(str(value).replace("Z", "+00:00"))
            fd = CalendarConverter.format_datetime(dt, cal_type_for_rows)
            return fd.get("date_only") or fd.get("formatted") or str(value)
        except Exception:
            return str(value)

    # Helpers for numeric formatting with thousands separator and trimming .00
    def _format_number_for_display(value: object) -> str:
        try:
            if value is None:
                return ""
            v = float(value)
            s = f"{v:,.2f}"
            # Trim trailing .00 or trailing zeros
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return s
        except Exception:
            return str(value)

    # Build rows with numeric alignment and calendar-aware dates
    amount_keys = {"total_amount", "subtotal", "discount_total", "tax_total", "payable_total"}
    date_keys = {"document_date", "registered_at", "created_at"}
    rows_html = []
    total_sum = 0.0
    discount_sum = 0.0
    tax_sum = 0.0
    for item in items:
        row_cells = []
        for key in keys:
            value = item.get(key, "")
            # Normalize list/dict to string
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            # Calendar-aware date formatting
            if key in date_keys and value:
                value = _format_date_for_calendar(value)
            # Amount cells: format and accumulate totals
            if key in amount_keys:
                try:
                    vnum = float(item.get(key)) if item.get(key) is not None else None
                    if key == "total_amount" and vnum is not None:
                        total_sum += vnum
                    if key == "discount_total" and vnum is not None:
                        discount_sum += vnum
                    if key == "tax_total" and vnum is not None:
                        tax_sum += vnum
                except Exception:
                    pass
                disp = _format_number_for_display(value)
                row_cells.append(f'<td class="amount">{escape(disp)}</td>')
            # Proforma: show checkmark for true, empty otherwise
            elif key == "is_proforma":
                checked = (str(value).lower() in ("true", "1"))
                cell = "✓" if checked else ""
                row_cells.append(f'<td style="text-align:center">{cell}</td>')
            else:
                row_cells.append(f'<td>{escape(str(value))}</td>')
        rows_html.append(f'<tr>{"".join(row_cells)}</tr>')

    # Summary block (only total amount and count for list)
    total_count = len(items)
    label_rows = 'تعداد ردیف' if is_fa else 'Rows'
    label_total = 'جمع مبلغ کل' if is_fa else 'Total of amounts'
    label_discount = 'جمع تخفیف' if is_fa else 'Total discount'
    label_tax = 'جمع مالیات' if is_fa else 'Total tax'
    summary_parts = [
        f'<div><strong>{label_rows}:</strong> {total_count}</div>',
        f'<div><strong>{label_total}:</strong> <span class="amount">{total_sum:.2f}</span></div>',
    ]
    # Only render discount/tax if present in any row (non-zero)
    if discount_sum != 0.0:
        summary_parts.append(f'<div><strong>{label_discount}:</strong> <span class="amount">{discount_sum:.2f}</span></div>')
    if tax_sum != 0.0:
        summary_parts.append(f'<div><strong>{label_tax}:</strong> <span class="amount">{tax_sum:.2f}</span></div>')
    summary_html = f'<div class="summary">{"".join(summary_parts)}</div>'

    # کانتکست مشترک برای قالب‌های سفارشی
    template_context: Dict[str, Any] = {
        "title_text": title_text,
        "business_name": business_name,
        "generated_at": now,
        "is_fa": is_fa,
        "fa_font_url_regular": None,
        "fa_font_url_bold": None,
        "headers": headers,
        "keys": keys,
        "items": items,
        # خروجی‌های HTML آماده برای استفاده سریع در قالب
        "table_headers_html": headers_html,
        "table_rows_html": "".join(rows_html),
        "table_summary_html": summary_html,
    }

    # Embed Farsi fonts like single-invoice PDF (if available)
    try:
        if is_fa:
            project_root = Path(__file__).resolve().parents[4]
            fonts_dir = project_root / "hesabixUI" / "hesabix_ui" / "assets" / "fonts"
            regular_path = fonts_dir / "YekanBakhFaNum-Regular.ttf"
            bold_path = fonts_dir / "YekanBakhFaNum-Bold.ttf"
            if regular_path.is_file():
                import base64 as _b64
                _data = regular_path.read_bytes()
                _b64_data = _b64.b64encode(_data).decode("ascii")
                template_context["fa_font_url_regular"] = f"data:font/ttf;base64,{_b64_data}"
            if bold_path.is_file():
                import base64 as _b64b
                _data_b = bold_path.read_bytes()
                _b64_data_b = _b64b.b64encode(_data_b).decode("ascii")
                template_context["fa_font_url_bold"] = f"data:font/ttf;base64,{_b64_data_b}"
    except Exception:
        pass

    # تلاش برای رندر با قالب سفارشی (explicit یا پیش‌فرض)
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if "template_id" in body and body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="invoices",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # HTML پیش‌فرض در نبود قالب: استفاده از فایل قالب
    disposition = "attachment"
    try:
        disposition = str(body.get("disposition") or "attachment")
    except Exception:
        disposition = "attachment"
    paper_size = None
    orientation = None
    try:
        paper_size = body.get("paper_size")
        orientation = body.get("orientation")
    except Exception:
        pass
    html_content = resolved_html or render_template(
        "pdf/invoices/list.html",
        {
            **template_context,
            "title_text": title_text,
            "paper_size": paper_size,
            "orientation": orientation,
            "footer_text": footer_text,
        },
    )

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=html_content).write_pdf(font_config=font_config)

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "invoices"
    if business_name:
        base += f"_{slugify(business_name)}"
    if selected_only:
        base += "_selected"
    # Add filters to filename when available
    try:
        doc_type = body.get("document_type")
        if isinstance(doc_type, str) and doc_type:
            base += f"_{slugify(doc_type)}"
    except Exception:
        pass
    try:
        fd = body.get("from_date")
        td = body.get("to_date")
        if isinstance(fd, str) and fd:
            base += f"_from_{slugify(fd[:10])}"
        if isinstance(td, str) and td:
            base += f"_to_{slugify(td[:10])}"
    except Exception:
        pass
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/business/{business_id}/import/template",
    summary="دانلود تمپلیت ایمپورت فاکتورها",
    description="فایل Excel تمپلیت برای ایمپورت فاکتورها را برمی‌گرداند",
)
@require_business_access("business_id")
async def download_invoices_import_template(
    request: Request,
    business_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "add")),
):
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Header row
    headers = [
        "invoice_number",  # شناسه یکتا برای گروه‌بندی ردیف‌های یک فاکتور
        "invoice_type",  # نوع فاکتور: sales, purchase, sales_return, purchase_return, direct_consumption, production, waste
        "document_date",  # تاریخ فاکتور: YYYY-MM-DD
        "currency_code",  # کد ارز: IRR, USD, etc.
        "is_proforma",  # پیش‌فاکتور: TRUE/FALSE
        "description",  # توضیحات فاکتور
        "person_code",  # کد مشتری/تامین‌کننده (برای sales/purchase)
        "seller_code",  # کد فروشنده/بازاریاب (اختیاری)
        "due_date",  # تاریخ سررسید (اختیاری)
        "post_inventory",  # ثبت انبار: TRUE/FALSE
        "product_code",  # کد کالا/خدمت
        "quantity",  # تعداد
        "unit",  # واحد: main/secondary (پیش‌فرض: main)
        "unit_price",  # قیمت واحد
        "discount_type",  # نوع تخفیف: percent/amount (پیش‌فرض: amount)
        "discount_value",  # مقدار تخفیف
        "tax_rate",  # نرخ مالیات (درصد)
        "line_description",  # توضیحات ردیف
        "movement",  # جهت حرکت: in/out (برای فاکتور تولید)
        "warehouse_code",  # کد انبار (اختیاری)
    ]
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    # Sample data rows
    # نکته: در ردیف‌های بعدی هر فاکتور، ستون‌های هدر (invoice_type, document_date, person_code, ...) خالی می‌مانند
    # فقط invoice_number و اطلاعات ردیف (product_code, quantity, ...) در همه ردیف‌ها وارد می‌شوند
    samples = [
        [
            "INV-001", "sales", "2024-01-15", "IRR", "FALSE", "فاکتور فروش نمونه",
            "CUST-001", "", "2024-02-15", "TRUE",
            "P1001", "10", "main", "100000", "amount", "5000", "9", "ردیف اول", "", "WH-001"
        ],
        [
            "INV-001", "", "", "", "", "", "", "", "", "",
            "P1002", "5", "main", "200000", "percent", "10", "9", "ردیف دوم", "", "WH-001"
        ],
        [
            "INV-002", "purchase", "2024-01-16", "IRR", "FALSE", "فاکتور خرید نمونه",
            "SUPP-001", "", "", "TRUE",
            "P1003", "20", "main", "50000", "amount", "0", "9", "", "", "WH-001"
        ],
    ]
    
    for row_idx, sample in enumerate(samples, start=2):
        for col, val in enumerate(sample, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # Auto width
    for column in ws.columns:
        try:
            letter = column[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"invoices_import_template_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/business/{business_id}/import/excel",
    summary="ایمپورت فاکتورها از فایل Excel",
    description="فایل اکسل را دریافت می‌کند و به‌صورت dry-run یا واقعی پردازش می‌کند",
)
@require_business_access("business_id")
async def import_invoices_excel(
    request: Request,
    business_id: int,
    file: UploadFile = File(...),
    dry_run: str = Form(default="true"),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "add")),
):
    import zipfile
    from decimal import Decimal
    from collections import defaultdict
    from datetime import datetime as dt
    from openpyxl import load_workbook
    from app.services.invoice_service import (
        INVOICE_SALES, INVOICE_PURCHASE, INVOICE_SALES_RETURN, INVOICE_PURCHASE_RETURN,
        INVOICE_DIRECT_CONSUMPTION, INVOICE_PRODUCTION, INVOICE_WASTE,
    )

    def _validate_excel_signature(content: bytes) -> bool:
        try:
            if not content.startswith(b'PK'):
                return False
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                return any(n.startswith('xl/') for n in zf.namelist())
        except Exception:
            return False

    def _parse_bool(v: object) -> Optional[bool]:
        if v is None: return None
        s = str(v).strip().lower()
        if s in ("true", "1", "yes", "on", "بله", "هست"):
            return True
        if s in ("false", "0", "no", "off", "خیر", "نیست"):
            return False
        return None

    def _parse_decimal(v: object) -> Optional[Decimal]:
        if v is None or str(v).strip() == "":
            return None
        try:
            return Decimal(str(v).replace(",", ""))
        except Exception:
            return None

    def _parse_date(v: object) -> Optional[date]:
        if v is None or str(v).strip() == "":
            return None
        try:
            if isinstance(v, dt):
                return v.date()
            s = str(v).strip()
            # Try different date formats
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"]:
                try:
                    return dt.strptime(s, fmt).date()
                except Exception:
                    continue
            return None
        except Exception:
            return None

    def _normalize_invoice_type(v: object) -> Optional[str]:
        if v is None: return None
        s = str(v).strip().lower()
        mapping = {
            "sales": INVOICE_SALES,
            "purchase": INVOICE_PURCHASE,
            "sales_return": INVOICE_SALES_RETURN,
            "purchase_return": INVOICE_PURCHASE_RETURN,
            "direct_consumption": INVOICE_DIRECT_CONSUMPTION,
            "production": INVOICE_PRODUCTION,
            "waste": INVOICE_WASTE,
        }
        if s in mapping:
            return mapping[s]
        # Check if already in correct format
        if s.startswith("invoice_"):
            return s
        return None

    try:
        is_dry_run = str(dry_run).lower() in ("true", "1", "yes", "on")

        if not file.filename or not file.filename.lower().endswith('.xlsx'):
            raise ApiError("INVALID_FILE", "فرمت فایل معتبر نیست. تنها xlsx پشتیبانی می‌شود", http_status=400)

        content = await file.read()
        if len(content) < 100 or not _validate_excel_signature(content):
            raise ApiError("INVALID_FILE", "فایل Excel معتبر نیست یا خالی است", http_status=400)

        try:
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        except zipfile.BadZipFile:
            raise ApiError("INVALID_FILE", "فایل Excel خراب است یا فرمت آن معتبر نیست", http_status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return success_response(data={"summary": {"total": 0}}, request=request, message="EMPTY_FILE")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = rows[1:]

        # Find column indices
        col_map = {h.lower(): i for i, h in enumerate(headers)}
        
        # Required columns
        required_cols = ["invoice_number", "invoice_type", "document_date", "currency_code", "product_code", "quantity", "unit_price"]
        missing_cols = [c for c in required_cols if c.lower() not in col_map]
        if missing_cols:
            raise ApiError("MISSING_COLUMNS", f"ستون‌های الزامی یافت نشد: {', '.join(missing_cols)}", http_status=400)

        # Group rows by invoice_number
        invoices_data: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        
        for row_idx, row in enumerate(data_rows, start=2):
            invoice_number = str(row[col_map["invoice_number"]]).strip() if col_map["invoice_number"] < len(row) else None
            if not invoice_number:
                continue
            
            row_data = {}
            for col_name, col_idx in col_map.items():
                if col_idx < len(row):
                    row_data[col_name] = row[col_idx]
            
            invoices_data[invoice_number].append({
                "row": row_idx,
                "data": row_data
            })

        # Cache lookups
        product_cache: Dict[str, Optional[Product]] = {}
        person_cache: Dict[str, Optional[Person]] = {}
        currency_cache: Dict[str, Optional[Currency]] = {}
        warehouse_cache: Dict[str, Optional[Warehouse]] = {}

        def _get_product(code: str) -> Optional[Product]:
            if code in product_cache:
                return product_cache[code]
            product = db.query(Product).filter(
                and_(Product.business_id == business_id, Product.code == code)
            ).first()
            product_cache[code] = product
            return product

        def _get_person(code: str) -> Optional[Person]:
            if code in person_cache:
                return person_cache[code]
            try:
                code_int = int(code)
                person = db.query(Person).filter(
                    and_(Person.business_id == business_id, Person.code == code_int)
                ).first()
            except Exception:
                person = None
            person_cache[code] = person
            return person

        def _get_currency(code: str) -> Optional[Currency]:
            if code in currency_cache:
                return currency_cache[code]
            currency = db.query(Currency).filter(Currency.code == code.upper()).first()
            currency_cache[code] = currency
            return currency

        def _get_warehouse(code: str) -> Optional[Warehouse]:
            if code in warehouse_cache:
                return warehouse_cache[code]
            warehouse = db.query(Warehouse).filter(
                and_(Warehouse.business_id == business_id, Warehouse.code == code)
            ).first()
            warehouse_cache[code] = warehouse
            return warehouse

        errors: List[Dict[str, Any]] = []
        valid_invoices: List[Dict[str, Any]] = []
        created_count = 0

        # Process each invoice group
        for invoice_number, rows_list in invoices_data.items():
            if not rows_list:
                continue

            # Extract header data from first row (ردیف اول هر فاکتور باید اطلاعات هدر را داشته باشد)
            # در ردیف‌های بعدی، اگر ستون‌های هدر خالی باشند، از ردیف اول استفاده می‌شود
            first_row = rows_list[0]["data"]
            invoice_errors: List[str] = []
            
            # Helper function to get header value (from first row if empty in current row)
            def _get_header_value(key: str, row_data: Dict[str, Any]) -> Any:
                val = row_data.get(key)
                # اگر خالی است، از ردیف اول استفاده کن
                if val is None or (isinstance(val, str) and str(val).strip() == ""):
                    return first_row.get(key)
                return val
            
            # Parse invoice type
            invoice_type_raw = str(_get_header_value("invoice_type", first_row) or "").strip()
            invoice_type = _normalize_invoice_type(invoice_type_raw)
            if not invoice_type:
                invoice_errors.append(f"نوع فاکتور نامعتبر: {invoice_type_raw}")
            
            # Parse document date
            doc_date = _parse_date(_get_header_value("document_date", first_row))
            if not doc_date:
                invoice_errors.append("تاریخ فاکتور الزامی است")
            
            # Parse currency
            currency_code = str(_get_header_value("currency_code", first_row) or "").strip()
            currency = _get_currency(currency_code) if currency_code else None
            if not currency:
                invoice_errors.append(f"ارز یافت نشد: {currency_code}")
            
            # Parse person (for sales/purchase)
            person_id = None
            person_code = str(_get_header_value("person_code", first_row) or "").strip()
            if person_code:
                person = _get_person(person_code)
                if person:
                    person_id = person.id
                else:
                    invoice_errors.append(f"شخص یافت نشد: {person_code}")
            elif invoice_type in {INVOICE_SALES, INVOICE_SALES_RETURN, INVOICE_PURCHASE, INVOICE_PURCHASE_RETURN}:
                invoice_errors.append("کد مشتری/تامین‌کننده الزامی است")
            
            # Parse other header fields
            is_proforma = _parse_bool(_get_header_value("is_proforma", first_row)) or False
            description = str(_get_header_value("description", first_row) or "").strip() or None
            seller_code = str(_get_header_value("seller_code", first_row) or "").strip()
            seller_id = None
            if seller_code:
                seller = _get_person(seller_code)
                if seller:
                    seller_id = seller.id
                else:
                    invoice_errors.append(f"فروشنده یافت نشد: {seller_code}")
            
            due_date = _parse_date(_get_header_value("due_date", first_row))
            post_inventory = _parse_bool(_get_header_value("post_inventory", first_row))
            if post_inventory is None:
                post_inventory = True

            # Parse line items
            lines: List[Dict[str, Any]] = []
            for row_info in rows_list:
                row_data = row_info["data"]
                row_num = row_info["row"]
                line_errors: List[str] = []
                
                # Product
                product_code = str(row_data.get("product_code", "")).strip()
                if not product_code:
                    line_errors.append("کد محصول الزامی است")
                    continue
                
                product = _get_product(product_code)
                if not product:
                    line_errors.append(f"محصول یافت نشد: {product_code}")
                    continue
                
                # Quantity
                quantity = _parse_decimal(row_data.get("quantity"))
                if not quantity or quantity <= 0:
                    line_errors.append("تعداد باید بزرگتر از صفر باشد")
                    continue
                
                # Unit price
                unit_price = _parse_decimal(row_data.get("unit_price"))
                if not unit_price or unit_price < 0:
                    line_errors.append("قیمت واحد الزامی است")
                    continue
                
                # Unit
                unit = str(row_data.get("unit", "main")).strip().lower()
                if unit not in ("main", "secondary"):
                    unit = "main"
                
                # Discount
                discount_type = str(row_data.get("discount_type", "amount")).strip().lower()
                if discount_type not in ("percent", "amount"):
                    discount_type = "amount"
                discount_value = _parse_decimal(row_data.get("discount_value")) or Decimal(0)
                
                # Tax rate
                tax_rate = _parse_decimal(row_data.get("tax_rate")) or Decimal(0)
                
                # Line description
                line_description = str(row_data.get("line_description", "")).strip() or None
                
                # Movement (for production invoices)
                movement = str(row_data.get("movement", "")).strip().lower()
                if movement not in ("in", "out"):
                    movement = None
                
                # Warehouse
                warehouse_id = None
                warehouse_code = row_data.get("warehouse_code", "")
                if warehouse_code:
                    warehouse = _get_warehouse(str(warehouse_code))
                    if warehouse:
                        warehouse_id = warehouse.id
                    else:
                        line_errors.append(f"انبار یافت نشد: {warehouse_code}")
                
                if line_errors:
                    invoice_errors.extend([f"ردیف {row_num}: {e}" for e in line_errors])
                    continue
                
                # Build line extra_info
                line_extra_info: Dict[str, Any] = {
                    "unit_price": float(unit_price),
                    "discount_type": discount_type,
                    "discount_value": float(discount_value),
                    "tax_rate": float(tax_rate),
                }
                if movement:
                    line_extra_info["movement"] = movement
                if warehouse_id:
                    line_extra_info["warehouse_id"] = warehouse_id
                
                lines.append({
                    "product_id": product.id,
                    "quantity": float(quantity),
                    "description": line_description,
                    "extra_info": line_extra_info,
                })
            
            if not lines:
                invoice_errors.append("حداقل یک ردیف معتبر الزامی است")
            
            if invoice_errors:
                errors.append({
                    "invoice_number": invoice_number,
                    "row": rows_list[0]["row"],
                    "errors": invoice_errors
                })
                continue
            
            # Build invoice payload
            extra_info: Dict[str, Any] = {
                "post_inventory": post_inventory,
            }
            if person_id:
                extra_info["person_id"] = person_id
            if seller_id:
                extra_info["seller_id"] = seller_id
            
            invoice_payload = {
                "invoice_type": invoice_type,
                "document_date": doc_date.isoformat(),
                "currency_id": currency.id,
                "is_proforma": is_proforma,
                "description": description,
                "extra_info": extra_info,
                "lines": lines,
            }
            
            if due_date:
                invoice_payload["due_date"] = due_date.isoformat()
            
            valid_invoices.append({
                "invoice_number": invoice_number,
                "payload": invoice_payload,
            })

        # Create invoices if not dry run
        if not is_dry_run and valid_invoices:
            user_id = ctx.get_user_id()
            for inv_info in valid_invoices:
                try:
                    create_invoice(
                        db=db,
                        business_id=business_id,
                        user_id=user_id,
                        data=inv_info["payload"],
                    )
                    created_count += 1
                except Exception as e:
                    logger.error(f"Failed to create invoice {inv_info['invoice_number']}: {e}", exc_info=True)
                    errors.append({
                        "invoice_number": inv_info["invoice_number"],
                        "row": 0,
                        "errors": [f"خطا در ایجاد فاکتور: {str(e)}"]
                    })

        summary = {
            "total": len(invoices_data),
            "valid": len(valid_invoices),
            "invalid": len(errors),
            "created": created_count,
            "dry_run": is_dry_run,
        }

        return success_response(
            data={"summary": summary, "errors": errors},
            request=request,
            message="INVOICES_IMPORT_RESULT",
        )
    except ApiError:
        raise
    except Exception as e:
        logger.error(f"Import error: {e}", exc_info=True)
        raise ApiError("IMPORT_ERROR", f"خطا در پردازش فایل: {e}", http_status=500)


@router.post("/business/{business_id}/invoices/calculate-remaining")
@require_business_access("business_id")
async def calculate_invoices_remaining_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("invoices", "view")),
) -> Dict[str, Any]:
    """
    محاسبه مانده چند فاکتور در یک درخواست
    
    Body:
        {
            "invoice_ids": [123, 456, 789]  # لیست شناسه فاکتورها
        }
    
    Response:
        {
            "results": {
                123: {
                    "invoice_id": 123,
                    "total_amount": 1000000.0,
                    "paid_amount": 500000.0,
                    "remaining": 500000.0,
                    "is_settled": false
                },
                456: {
                    "invoice_id": 456,
                    "total_amount": 2000000.0,
                    "paid_amount": 2000000.0,
                    "remaining": 0.0,
                    "is_settled": true
                }
            },
            "errors": {
                789: "فاکتور یافت نشد"
            }
        }
    """
    from app.services.invoice_service import calculate_invoice_remaining
    
    invoice_ids = body.get("invoice_ids", [])
    
    if not isinstance(invoice_ids, list):
        raise ApiError("INVALID_INPUT", "invoice_ids باید یک لیست باشد", http_status=400)
    
    if len(invoice_ids) > 100:  # محدودیت برای جلوگیری از overload
        raise ApiError("TOO_MANY_INVOICES", "حداکثر 100 فاکتور در یک درخواست", http_status=400)
    
    results = {}
    errors = {}
    
    logger.info(f"محاسبه مانده برای {len(invoice_ids)} فاکتور - invoice_ids: {invoice_ids}")
    
    for invoice_id in invoice_ids:
        try:
            invoice_id_int = int(invoice_id)
            logger.info(f"محاسبه مانده فاکتور {invoice_id_int}")
            result = calculate_invoice_remaining(db, business_id, invoice_id_int)
            # تبدیل کلید به string برای JSON serialization
            results[str(invoice_id_int)] = result
            logger.info(f"نتیجه برای فاکتور {invoice_id_int}: {result}")
        except ApiError as e:
            logger.warning(f"ApiError برای فاکتور {invoice_id}: {e.message}")
            errors[str(invoice_id_int)] = e.message
        except Exception as e:
            logger.exception(f"خطا در محاسبه مانده فاکتور {invoice_id}")
            errors[str(invoice_id_int)] = "خطا در محاسبه مانده"
    
    logger.info(f"نتایج: {len(results)} موفق، {len(errors)} خطا")
    
    return success_response(
        data={
            "results": results,
            "errors": errors,
        },
        request=request,
        message="INVOICE_REMAINING_CALCULATED",
    )

