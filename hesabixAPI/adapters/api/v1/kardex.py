from typing import Any, Dict

from fastapi import APIRouter, Depends, Request, Body
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.responses import success_response, format_datetime_fields
from app.core.permissions import require_business_access
from adapters.api.v1.schemas import QueryInfo
from app.services.kardex_service import list_kardex_lines
from app.services.pdf.template_renderer import render_template
from app.core.i18n import negotiate_locale
from adapters.db.models.business import Business
from app.core.cache import get_cache


router = APIRouter(prefix="/kardex", tags=["گزارش‌ها", "انبارداری"])


@router.post(
    "/businesses/{business_id}/lines",
    summary="لیست کاردکس (خطوط اسناد)",
    description="دریافت خطوط اسناد مرتبط با انتخاب‌های چندگانه موجودیت‌ها با فیلتر تاریخ",
)
@require_business_access("business_id")
async def list_kardex_lines_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    # Compose query dict from QueryInfo and additional parameters from body
    query_dict: Dict[str, Any] = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by or "document_date",
        "sort_desc": query_info.sort_desc,
        "sort": [s.model_dump() for s in query_info.sort] if query_info.sort else None,
        "search": query_info.search,
        "search_fields": query_info.search_fields,
        "filters": query_info.filters,
    }

    # Additional params from body (DataTable additionalParams)
    try:
        body_json = await request.json()
        if isinstance(body_json, dict):
            for key in (
                "from_date",
                "to_date",
                "fiscal_year_id",
                "person_ids",
                "product_ids",
                "bank_account_ids",
                "cash_register_ids",
                "petty_cash_ids",
                "account_ids",
                "check_ids",
                "warehouse_ids",
                "match_mode",
                "result_scope",
            ):
                if key in body_json and body_json.get(key) is not None:
                    query_dict[key] = body_json.get(key)
    except Exception:
        pass

    # کش نتایج لیست کاردکس (خطوط اسناد)
    cache = get_cache()
    cache_key = None

    if cache.enabled:
        import json, hashlib
        key_payload = {
            "business_id": business_id,
            "query": query_dict,
        }
        key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
        cache_key = f"kardex_lines:{key_hash}"
        cached = cache.get(cache_key)
        if cached is not None:
            return success_response(data=cached, request=request, message="KARDEX_LINES")

    result = list_kardex_lines(db, business_id, query_dict)

    # Format date fields in response items (document_date)
    try:
        items = result.get("items", [])
        for item in items:
            # Use format_datetime_fields for consistency
            item.update(format_datetime_fields({"document_date": item.get("document_date")}, request))
    except Exception:
        pass

    if cache.enabled and cache_key:
        cache.set(cache_key, result, ttl=60)

    return success_response(data=result, request=request, message="KARDEX_LINES")


@router.post(
    "/businesses/{business_id}/lines/export/excel",
    summary="خروجی Excel کاردکس",
    description="خروجی اکسل از لیست خطوط کاردکس با فیلترهای اعمال‌شده",
)
@require_business_access("business_id")
async def export_kardex_excel_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    from fastapi.responses import Response
    import datetime
    try:
        max_export_records = 10000
        take_value = min(int(body.get("take", 1000)), max_export_records)
    except Exception:
        take_value = 1000

    query_dict: Dict[str, Any] = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by") or "document_date",
        "sort_desc": bool(body.get("sort_desc", True)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
        "from_date": body.get("from_date"),
        "to_date": body.get("to_date"),
        "person_ids": body.get("person_ids"),
        "product_ids": body.get("product_ids"),
        "bank_account_ids": body.get("bank_account_ids"),
        "cash_register_ids": body.get("cash_register_ids"),
        "petty_cash_ids": body.get("petty_cash_ids"),
        "account_ids": body.get("account_ids"),
        "check_ids": body.get("check_ids"),
        "warehouse_ids": body.get("warehouse_ids"),
        "match_mode": body.get("match_mode") or "any",
        "result_scope": body.get("result_scope") or "lines_matching",
        "include_running_balance": bool(body.get("include_running_balance", False)),
    }

    result = list_kardex_lines(db, business_id, query_dict)
    items = result.get("items", [])
    items = [format_datetime_fields(it, request) for it in items]

    # Build Excel using openpyxl (RTL, thousand separators, safe strings, configurable columns)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import re
    from io import BytesIO
    import datetime as _dt

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == "fa")

    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    if is_fa:
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    # Helpers
    def _slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    def _safe_text(v: Any) -> str:
        s = "" if v is None else str(v)
        # Prevent Excel formula injection
        if s.startswith(("=", "+", "-", "@")):
            return "'" + s
        return s

    def _num(v: Any) -> float | None:
        if v is None:
            return None
        try:
            return float(v)
        except Exception:
            return None

    # Apply selected-only filtering with stable keys (line_id) if provided
    selected_only = bool(body.get("selected_only", False))
    selected_row_keys = body.get("selected_row_keys")
    if selected_only and isinstance(selected_row_keys, list):
        try:
            wanted_ids = set()
            for k in selected_row_keys:
                if isinstance(k, dict) and k.get("line_id") is not None:
                    try:
                        wanted_ids.add(int(k.get("line_id")))
                    except Exception:
                        pass
            if wanted_ids:
                items = [it for it in items if int(it.get("line_id") or -1) in wanted_ids]
        except Exception:
            pass

    # Backward compatible: selected_indices (indexes) - best-effort
    if selected_only and (not selected_row_keys) and body.get("selected_indices") is not None:
        try:
            selected_indices = body.get("selected_indices")
            indices = None
            if isinstance(selected_indices, str):
                import json as _json
                indices = _json.loads(selected_indices)
            elif isinstance(selected_indices, list):
                indices = selected_indices
            if isinstance(indices, list):
                items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
        except Exception:
            pass

    # Determine export columns
    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        keys = [c.get("key") for c in export_columns if c.get("key")]
        headers = [c.get("label") or c.get("key") for c in export_columns if c.get("key")]
    else:
        # Default columns (labels localized)
        headers = [
            ("document_date", "تاریخ سند" if is_fa else "Date"),
            ("document_code", "کد سند" if is_fa else "Code"),
            ("document_type_name", "نوع سند" if is_fa else "Type"),
            ("warehouse_name", "انبار" if is_fa else "Warehouse"),
            ("movement", "جهت" if is_fa else "Movement"),
            ("description", "شرح" if is_fa else "Description"),
            ("debit", "بدهکار" if is_fa else "Debit"),
            ("credit", "بستانکار" if is_fa else "Credit"),
            ("quantity", "تعداد" if is_fa else "Qty"),
        ]
        if bool(query_dict.get("include_running_balance", False)):
            headers += [
                ("running_amount", "مانده مبلغ" if is_fa else "Running Amount"),
                ("running_quantity", "مانده تعداد" if is_fa else "Running Qty"),
            ]
        keys = [k for k, _ in headers]
        headers = [h for _, h in headers]

    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Freeze header + enable filter row
    try:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

    # Column kind for number formatting
    money_keys = {"debit", "credit", "running_amount"}
    qty_keys = {"quantity", "running_quantity"}

    for it in items:
        row_vals = []
        for k in keys:
            # map friendly values
            if k == "document_type_name":
                v = it.get("document_type_name") or it.get("document_type")
            elif k == "warehouse_name":
                v = it.get("warehouse_name") or it.get("warehouse_id")
            else:
                v = it.get(k)
            if k in money_keys:
                row_vals.append(_num(v))
            elif k in qty_keys:
                row_vals.append(_num(v))
            else:
                row_vals.append(_safe_text(v))
        ws.append(row_vals)
        # style row
        for j, cell in enumerate(ws[ws.max_row], start=1):
            cell.border = thin_border
            if is_fa:
                cell.alignment = Alignment(horizontal="right")

        # Apply number formats for the row cells
        try:
            for idx, k in enumerate(keys, start=1):
                c = ws.cell(row=ws.max_row, column=idx)
                if k in money_keys:
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal="right")
                elif k in qty_keys:
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")
        except Exception:
            pass

    # Auto width columns (cap)
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass

    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    filename = f"kardex_{business_id}_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/lines/export/pdf",
    summary="خروجی PDF کاردکس",
    description="خروجی PDF از لیست خطوط کاردکس با فیلترهای اعمال‌شده",
)
@require_business_access("business_id")
async def export_kardex_pdf_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    from fastapi.responses import Response
    import datetime
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from html import escape

    try:
        max_export_records = 10000
        take_value = min(int(body.get("take", 1000)), max_export_records)
    except Exception:
        take_value = 1000

    query_dict: Dict[str, Any] = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by") or "document_date",
        "sort_desc": bool(body.get("sort_desc", True)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
        "from_date": body.get("from_date"),
        "to_date": body.get("to_date"),
        "person_ids": body.get("person_ids"),
        "product_ids": body.get("product_ids"),
        "bank_account_ids": body.get("bank_account_ids"),
        "cash_register_ids": body.get("cash_register_ids"),
        "petty_cash_ids": body.get("petty_cash_ids"),
        "account_ids": body.get("account_ids"),
        "check_ids": body.get("check_ids"),
        "warehouse_ids": body.get("warehouse_ids"),
        "match_mode": body.get("match_mode") or "any",
        "result_scope": body.get("result_scope") or "lines_matching",
        "include_running_balance": bool(body.get("include_running_balance", False)),
    }

    result = list_kardex_lines(db, business_id, query_dict)
    items = result.get("items", [])
    items = [format_datetime_fields(it, request) for it in items]
    all_totals = result.get("totals") or {}
    all_count = (result.get("pagination") or {}).get("total")

    # Apply selected-only filtering (prefer stable keys, fallback to indices)
    selected_only = bool(body.get("selected_only", False))
    selected_row_keys = body.get("selected_row_keys")
    if selected_only and isinstance(selected_row_keys, list):
        try:
            wanted_ids = set()
            for k in selected_row_keys:
                if isinstance(k, dict) and k.get("line_id") is not None:
                    try:
                        wanted_ids.add(int(k.get("line_id")))
                    except Exception:
                        pass
            if wanted_ids:
                items = [it for it in items if int(it.get("line_id") or -1) in wanted_ids]
        except Exception:
            pass
    if selected_only and (not selected_row_keys) and body.get("selected_indices") is not None:
        try:
            selected_indices = body.get("selected_indices")
            indices = None
            if isinstance(selected_indices, str):
                import json as _json
                indices = _json.loads(selected_indices)
            elif isinstance(selected_indices, list):
                indices = selected_indices
            if isinstance(indices, list):
                items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
        except Exception:
            pass

    # Build a human-readable filters summary for PDF header (used in templates and fallback HTML)
    filters_summary = []
    try:
        # locale/is_fa are set later too; compute early for summary
        _locale0 = negotiate_locale(request.headers.get("Accept-Language"))
        _is_fa0 = (_locale0 == "fa")

        def _add(label_fa: str, label_en: str, value: Any):
            if value is None:
                return
            if isinstance(value, str) and not value.strip():
                return
            if isinstance(value, (list, tuple)) and len(value) == 0:
                return
            label = label_fa if _is_fa0 else label_en
            if isinstance(value, (list, tuple)):
                vals = [str(v) for v in value if v is not None and str(v).strip()]
                if not vals:
                    return
                if len(vals) <= 3:
                    vtxt = "، ".join(vals) if _is_fa0 else ", ".join(vals)
                else:
                    vtxt = (f"{'، '.join(vals[:3])} (+{len(vals)-3})" if _is_fa0 else f"{', '.join(vals[:3])} (+{len(vals)-3})")
            else:
                vtxt = str(value)
            filters_summary.append({"label": label, "value": vtxt})

        _add("از تاریخ", "From", query_dict.get("from_date"))
        _add("تا تاریخ", "To", query_dict.get("to_date"))
        _add("سال مالی", "Fiscal year", query_dict.get("fiscal_year_id"))

        def _as_int_list(v: Any) -> list[int]:
            if not isinstance(v, (list, tuple)):
                return []
            out: list[int] = []
            for x in v:
                try:
                    if x is None:
                        continue
                    out.append(int(x))
                except Exception:
                    continue
            # unique preserve order
            seen = set()
            uniq: list[int] = []
            for i in out:
                if i in seen:
                    continue
                seen.add(i)
                uniq.append(i)
            return uniq

        def _names_or_count(label_fa: str, label_en: str, ids_any: Any, rows: list[Any], pick_name):
            ids = _as_int_list(ids_any)
            if not ids:
                return
            # show names for small lists; otherwise first 3 + (+N)
            try:
                if len(ids) <= 3:
                    names = [pick_name(r) for r in rows if r is not None]
                    names = [n for n in names if n]
                    if names:
                        _add(label_fa, label_en, names)
                        return
                # fallback to count
                _add(label_fa, label_en, [f"{len(ids)} مورد" if _is_fa0 else f"{len(ids)} items"])
            except Exception:
                _add(label_fa, label_en, [f"{len(ids)} مورد" if _is_fa0 else f"{len(ids)} items"])

        # persons (show alias_name)
        person_ids = _as_int_list(query_dict.get("person_ids"))
        if person_ids:
            try:
                from adapters.db.models.person import Person
                people = (
                    db.query(Person)
                    .filter(Person.business_id == business_id, Person.id.in_(person_ids[:3]))
                    .all()
                )
                _names_or_count("اشخاص", "Persons", person_ids, people, lambda p: getattr(p, "alias_name", None) or "")
            except Exception:
                _add("اشخاص", "Persons", [f"{len(person_ids)} مورد" if _is_fa0 else f"{len(person_ids)} items"])

        # products
        product_ids = _as_int_list(query_dict.get("product_ids"))
        if product_ids:
            try:
                from adapters.db.models.product import Product
                prows = (
                    db.query(Product)
                    .filter(Product.business_id == business_id, Product.id.in_(product_ids[:3]))
                    .all()
                )
                def _pname(p):
                    code = getattr(p, "code", None)
                    name = getattr(p, "name", None)
                    return f"{code} - {name}" if code and name else (name or code or "")
                _names_or_count("کالاها", "Products", product_ids, prows, _pname)
            except Exception:
                _add("کالاها", "Products", [f"{len(product_ids)} مورد" if _is_fa0 else f"{len(product_ids)} items"])

        # bank accounts
        bank_ids = _as_int_list(query_dict.get("bank_account_ids"))
        if bank_ids:
            try:
                from adapters.db.models.bank_account import BankAccount
                brows = (
                    db.query(BankAccount)
                    .filter(BankAccount.business_id == business_id, BankAccount.id.in_(bank_ids[:3]))
                    .all()
                )
                def _bname(b):
                    code = getattr(b, "code", None)
                    name = getattr(b, "name", None)
                    return f"{code} - {name}" if code and name else (name or code or "")
                _names_or_count("حساب‌های بانکی", "Bank accounts", bank_ids, brows, _bname)
            except Exception:
                _add("حساب‌های بانکی", "Bank accounts", [f"{len(bank_ids)} مورد" if _is_fa0 else f"{len(bank_ids)} items"])

        # cash registers
        cash_ids = _as_int_list(query_dict.get("cash_register_ids"))
        if cash_ids:
            try:
                from adapters.db.models.cash_register import CashRegister
                crows = (
                    db.query(CashRegister)
                    .filter(CashRegister.business_id == business_id, CashRegister.id.in_(cash_ids[:3]))
                    .all()
                )
                _names_or_count("صندوق‌ها", "Cash registers", cash_ids, crows, lambda c: getattr(c, "name", None) or "")
            except Exception:
                _add("صندوق‌ها", "Cash registers", [f"{len(cash_ids)} مورد" if _is_fa0 else f"{len(cash_ids)} items"])

        # petty cash
        petty_ids = _as_int_list(query_dict.get("petty_cash_ids"))
        if petty_ids:
            try:
                from adapters.db.models.petty_cash import PettyCash
                prows = (
                    db.query(PettyCash)
                    .filter(PettyCash.business_id == business_id, PettyCash.id.in_(petty_ids[:3]))
                    .all()
                )
                _names_or_count("تنخواه‌ها", "Petty cash", petty_ids, prows, lambda p: getattr(p, "name", None) or "")
            except Exception:
                _add("تنخواه‌ها", "Petty cash", [f"{len(petty_ids)} مورد" if _is_fa0 else f"{len(petty_ids)} items"])

        # accounts
        acc_ids = _as_int_list(query_dict.get("account_ids"))
        if acc_ids:
            try:
                from adapters.db.models.account import Account
                arows = (
                    db.query(Account)
                    .filter(Account.business_id == business_id, Account.id.in_(acc_ids[:3]))
                    .all()
                )
                def _aname(a):
                    code = getattr(a, "code", None)
                    name = getattr(a, "name", None)
                    return f"{code} - {name}" if code and name else (name or code or "")
                _names_or_count("حساب‌ها", "Accounts", acc_ids, arows, _aname)
            except Exception:
                _add("حساب‌ها", "Accounts", [f"{len(acc_ids)} مورد" if _is_fa0 else f"{len(acc_ids)} items"])

        # checks
        check_ids = _as_int_list(query_dict.get("check_ids"))
        if check_ids:
            try:
                from adapters.db.models.check import Check
                chrows = (
                    db.query(Check)
                    .filter(Check.business_id == business_id, Check.id.in_(check_ids[:3]))
                    .all()
                )
                def _cname(ch):
                    num = getattr(ch, "check_number", None)
                    sayad = getattr(ch, "sayad_code", None)
                    bank = getattr(ch, "bank_name", None)
                    if num and bank:
                        return f"{bank} #{num}"
                    if num:
                        return f"#{num}"
                    if sayad:
                        return f"صیاد {sayad}" if _is_fa0 else f"Sayad {sayad}"
                    return ""
                _names_or_count("چک‌ها", "Checks", check_ids, chrows, _cname)
            except Exception:
                _add("چک‌ها", "Checks", [f"{len(check_ids)} مورد" if _is_fa0 else f"{len(check_ids)} items"])

        # warehouses
        wh_ids = _as_int_list(query_dict.get("warehouse_ids"))
        if wh_ids:
            try:
                from adapters.db.models.warehouse import Warehouse
                wrows = (
                    db.query(Warehouse)
                    .filter(Warehouse.business_id == business_id, Warehouse.id.in_(wh_ids[:3]))
                    .all()
                )
                _names_or_count("انبارها", "Warehouses", wh_ids, wrows, lambda w: getattr(w, "name", None) or "")
            except Exception:
                _add("انبارها", "Warehouses", [f"{len(wh_ids)} مورد" if _is_fa0 else f"{len(wh_ids)} items"])

        mm = str(query_dict.get("match_mode") or "").lower()
        if mm:
            mm_map_fa = {"any": "هرکدام", "same_line": "همان سطر", "document_and": "همان سند (AND)"}
            mm_map_en = {"any": "Any", "same_line": "Same line", "document_and": "Same document (AND)"}
            _add("حالت تطبیق", "Match mode", mm_map_fa.get(mm, mm) if _is_fa0 else mm_map_en.get(mm, mm))
        rs = str(query_dict.get("result_scope") or "").lower()
        if rs:
            rs_map_fa = {"lines_matching": "فقط سطرهای منطبق", "lines_of_document": "تمام سطرهای سند"}
            rs_map_en = {"lines_matching": "Matching lines", "lines_of_document": "All lines of document"}
            _add("دامنه نتایج", "Result scope", rs_map_fa.get(rs, rs) if _is_fa0 else rs_map_en.get(rs, rs))

        if query_dict.get("search"):
            _add("جستجو", "Search", query_dict.get("search"))
        flt = query_dict.get("filters")
        if isinstance(flt, list) and flt:
            _add("فیلترهای ستونی", "Column filters", f"{len(flt)} مورد" if _is_fa0 else f"{len(flt)} items")
    except Exception:
        filters_summary = []

    # Page totals (for exported rows after selected-only filtering)
    try:
        page_debit = 0.0
        page_credit = 0.0
        page_qty_sum = 0.0
        qty_seen = False
        for it in items or []:
            try:
                page_debit += float(it.get("debit") or 0)
            except Exception:
                pass
            try:
                page_credit += float(it.get("credit") or 0)
            except Exception:
                pass
            qv = it.get("quantity")
            if qv is not None:
                try:
                    page_qty_sum += float(qv or 0)
                    qty_seen = True
                except Exception:
                    pass
        page_totals = {
            "debit": page_debit,
            "credit": page_credit,
            "quantity": page_qty_sum if qty_seen else None,
            "count": len(items or []),
        }
    except Exception:
        page_totals = {"debit": 0.0, "credit": 0.0, "quantity": None, "count": len(items or [])}

    # Build simple HTML table
    def cell(val: Any) -> str:
        # Never show literal "None"
        if val is None:
            return ""
        s = str(val)
        if s.strip().lower() == "none":
            return ""
        return escape(s)

    def cell_num(val: Any, decimals: int = 2) -> str:
        if val is None:
            return ""
        try:
            n = float(val)
            if abs(n - round(n)) < 1e-9:
                return escape(f"{int(round(n)):,}")
            s = f"{n:,.{decimals}f}".rstrip("0").rstrip(".")
            return escape(s)
        except Exception:
            return cell(val)

    rows_html = "".join([
        f"<tr>"
        f"<td style='text-align:right'>{cell_num(i+1, 0)}</td>"
        f"<td>{cell(it.get('document_date'))}</td>"
        f"<td>{cell(it.get('document_code'))}</td>"
        f"<td>{cell(it.get('document_type_name') or it.get('document_type'))}</td>"
        f"<td>{cell(it.get('warehouse_name') or it.get('warehouse_id'))}</td>"
        f"<td>{cell(it.get('movement'))}</td>"
        f"<td>{cell(it.get('description'))}</td>"
        f"<td style='text-align:right'>{cell_num(it.get('debit'), 0)}</td>"
        f"<td style='text-align:right'>{cell_num(it.get('credit'), 0)}</td>"
        f"<td style='text-align:right'>{cell_num(it.get('quantity'), 2)}</td>"
        f"<td style='text-align:right'>{cell_num(it.get('running_amount'), 0)}</td>"
        f"<td style='text-align:right'>{cell_num(it.get('running_quantity'), 2)}</td>"
        f"</tr>"
        for i, it in enumerate(items)
    ])

    # Prepare font URIs for base.html / templates
    fa_font_url_regular = ""
    fa_font_url_bold = ""

    # تلاش برای رندر با قالب سفارشی (kardex/list) و سپس قالب پیش‌فرض فایل
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        # اطلاعات کسب‌وکار
        business_name = ""
        try:
            b = db.query(Business).filter(Business.id == business_id).first()
            if b is not None:
                business_name = b.name or ""
        except Exception:
            business_name = ""
        # Locale و جهت
        locale = negotiate_locale(request.headers.get("Accept-Language"))
        is_fa = (locale == "fa")
        calendar_type = getattr(request.state, "calendar_type", "jalali")
        # Load font data URIs early so templates (base.html) can use them
        try:
            if is_fa:
                from app.services.pdf.template_renderer import load_farsi_font_data_uris
                fa_reg, fa_bold = load_farsi_font_data_uris()
                fa_font_url_regular = fa_reg or ""
                fa_font_url_bold = fa_bold or ""
        except Exception:
            fa_font_url_regular = ""
            fa_font_url_bold = ""
        # زمان تولید گزارش بر اساس تقویم انتخابی
        try:
            from app.core.calendar import CalendarConverter
            _now = datetime.datetime.now()
            _gen = CalendarConverter.format_datetime(_now, calendar_type).get("formatted") or ""
            # Trim seconds for nicer display
            generated_at = " ".join(_gen.split(" ")[:2])
            if generated_at.count(":") >= 2:
                generated_at = generated_at.rsplit(":", 1)[0]
        except Exception:
            generated_at = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
        # پارامترهای صفحه از کوئری (اختیاری)
        try:
            qp = request.query_params
            paper_size = qp.get("paper_size")
            orientation = qp.get("orientation") or "landscape"
            disposition = qp.get("disposition") or "attachment"
        except Exception:
            paper_size = None
            orientation = "landscape"
            disposition = "attachment"
        # کانتکست مشترک
        template_context = {
            "title_text": "گزارش کاردکس" if is_fa else "Kardex Report",
            "business_name": business_name,
            "generated_at": generated_at,
            "is_fa": is_fa,
            "locale": locale,
            "paper_size": paper_size,
            "orientation": orientation,
            "show_running": bool(query_dict.get("include_running_balance", False)),
            "filters_summary": filters_summary,
            "fa_font_url_regular": fa_font_url_regular,
            "fa_font_url_bold": fa_font_url_bold,
            "footer_text": f"{'گزارش کاردکس' if is_fa else 'Kardex Report'} • {generated_at}",
            "page_totals": page_totals,
            "all_totals": all_totals,
            "all_count": all_count,
            "items": items,
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="kardex",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # Inject Persian fonts (YekanBakhFaNum/Vazirmatn) for PDF rendering (CSS fallback)
    font_face_css = ""
    if is_fa and fa_font_url_regular:
        font_face_css += f"@font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_regular}') format('truetype'); font-weight: 400; font-style: normal; }}\n"
    if is_fa and fa_font_url_bold:
        font_face_css += f"@font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_bold}') format('truetype'); font-weight: 700; font-style: normal; }}\n"
    
    body_font_family = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if is_fa else "Arial, sans-serif"
    
    html = f"""
    <html>
      <head>
        <meta charset='utf-8'/>
        <style>
          {font_face_css}
          body {{ font-family: {body_font_family}; }}
          table {{ width: 100%; border-collapse: collapse; }}
          th, td {{ border: 1px solid #ddd; padding: 6px; font-size: 12px; }}
          th {{ background: #f5f5f5; text-align: right; }}
        </style>
      </head>
      <body>
        <h3>{'گزارش کاردکس' if is_fa else 'Kardex Report'}</h3>
        {(
          "<div style='margin:6px 0 10px 0; font-size:11px; color:#555'>" +
          "".join([f"<span style=\"display:inline-block;border:1px solid #e5e7eb;border-radius:999px;padding:3px 8px;margin:2px 4px;background:#fff\"><b>{escape(str(f.get('label')))}:</b> {escape(str(f.get('value')))}</span>" for f in (filters_summary or [])]) +
          "</div>"
        ) if filters_summary else ""}
        <table>
          <thead>
            <tr>
              <th>#</th>
              <th>تاریخ سند</th>
              <th>کد سند</th>
              <th>نوع سند</th>
              <th>انبار</th>
              <th>جهت حرکت</th>
              <th>شرح</th>
              <th>بدهکار</th>
              <th>بستانکار</th>
              <th>تعداد</th>
              <th>مانده مبلغ</th>
              <th>مانده تعداد</th>
            </tr>
          </thead>
          <tbody>
            {rows_html}
          </tbody>
        </table>
      </body>
    </html>
    """

    # در صورت نبود قالب سفارشی، از قالب فایل استفاده کن
    if not resolved_html:
        try:
            final_html = render_template("pdf/kardex/list.html", {
                "title_text": "گزارش کاردکس" if is_fa else "Kardex Report",
                "business_name": business_name,
                "generated_at": generated_at if 'generated_at' in locals() else datetime.datetime.now().strftime("%Y/%m/%d %H:%M"),
                "is_fa": is_fa,
                "locale": locale,
                "paper_size": paper_size or "A4",
                "orientation": orientation or "landscape",
                "show_running": bool(query_dict.get("include_running_balance", False)),
                "filters_summary": filters_summary,
                "fa_font_url_regular": fa_font_url_regular,
                "fa_font_url_bold": fa_font_url_bold,
                "footer_text": f"{'گزارش کاردکس' if is_fa else 'Kardex Report'} • {generated_at if 'generated_at' in locals() else ''}",
                "page_totals": page_totals,
                "all_totals": all_totals,
                "all_count": all_count,
                "items": items,
            })
        except Exception:
            final_html = html
    else:
        final_html = resolved_html
        # If a custom template is used (DB), it may not include totals section yet.
        # Ensure totals are still visible by injecting a compact B/W totals block near the top of <body>.
        try:
            if "data-kardex-totals=\"1\"" not in final_html:
                def _fmt_num(v: Any, decimals: int = 0) -> str:
                    if v is None:
                        return "—"
                    try:
                        n = float(v)
                        if abs(n - round(n)) < 1e-9:
                            return f"{int(round(n)):,}"
                        s = f"{n:,.{decimals}f}".rstrip("0").rstrip(".")
                        return s
                    except Exception:
                        return "—"

                _is_fa1 = bool(is_fa)
                pt = page_totals or {}
                at = all_totals or {}
                totals_html = f"""
                <div data-kardex-totals="1" style="margin:10px 0 10px 0; padding:8px 10px; border:1px solid #000; border-radius:6px; background:#fff; font-size:10.5px; color:#111; page-break-inside:avoid;">
                  <div style="display:flex; gap:10px; flex-wrap:wrap; align-items:flex-start;">
                    <div style="flex:1; min-width:240px; border:1px solid #000; border-radius:6px; padding:8px 10px;">
                      <div style="font-weight:700; margin-bottom:6px;">{('جمع صفحه' if _is_fa1 else 'Page totals')}</div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0; border-bottom:1px dashed #999;"><span>{('تعداد سطر' if _is_fa1 else 'Rows')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(pt.get('count', 0), 0))}</span></div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0; border-bottom:1px dashed #999;"><span>{('جمع بدهکار' if _is_fa1 else 'Total debit')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(pt.get('debit', 0), 0))}</span></div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0; border-bottom:1px dashed #999;"><span>{('جمع بستانکار' if _is_fa1 else 'Total credit')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(pt.get('credit', 0), 0))}</span></div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0;"><span>{('جمع تعداد' if _is_fa1 else 'Total qty')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(pt.get('quantity', None), 2))}</span></div>
                    </div>
                    <div style="flex:1; min-width:240px; border:1px solid #000; border-radius:6px; padding:8px 10px;">
                      <div style="font-weight:700; margin-bottom:6px;">{('جمع کل' if _is_fa1 else 'Overall totals')}</div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0; border-bottom:1px dashed #999;"><span>{('تعداد کل' if _is_fa1 else 'Total rows')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(all_count, 0))}</span></div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0; border-bottom:1px dashed #999;"><span>{('جمع بدهکار' if _is_fa1 else 'Total debit')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(at.get('debit', 0), 0))}</span></div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0; border-bottom:1px dashed #999;"><span>{('جمع بستانکار' if _is_fa1 else 'Total credit')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(at.get('credit', 0), 0))}</span></div>
                      <div style="display:flex; justify-content:space-between; gap:8px; padding:2px 0;"><span>{('جمع تعداد' if _is_fa1 else 'Total qty')}</span><span style="font-weight:700; direction:ltr; unicode-bidi:plaintext;">{escape(_fmt_num(at.get('quantity', None), 2))}</span></div>
                    </div>
                  </div>
                </div>
                """

                idx = final_html.lower().find("<body")
                if idx != -1:
                    gt = final_html.find(">", idx)
                    if gt != -1:
                        final_html = final_html[: gt + 1] + totals_html + final_html[gt + 1 :]
                    else:
                        final_html = totals_html + final_html
                else:
                    final_html = totals_html + final_html
        except Exception:
            pass
    # Respect paper size/orientation params (fallback to A4 landscape)
    effective_paper = paper_size or "A4"
    effective_orientation = orientation or "landscape"
    effective_disposition = disposition or "attachment"

    font_config = FontConfiguration()
    pdf_css = f"""
    {font_face_css}
    @page {{ size: {effective_paper} {effective_orientation}; margin: 12mm; }}
    body {{ font-family: {body_font_family}; }}
    """
    pdf_bytes = HTML(string=final_html).write_pdf(stylesheets=[CSS(string=pdf_css)], font_config=font_config)

    filename = f"kardex_{business_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{effective_disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


