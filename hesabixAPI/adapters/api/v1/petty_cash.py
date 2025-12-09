from typing import Any, Dict, List
from fastapi import APIRouter, Depends, Request, Body
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from adapters.db.models.petty_cash import PettyCash
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.permissions import require_business_management_dep, require_business_access, require_business_permission_dep, require_business_permission_by_entity_dep
from adapters.api.v1.schemas import QueryInfo
from app.services.petty_cash_service import (
    create_petty_cash,
    update_petty_cash,
    delete_petty_cash,
    get_petty_cash_by_id,
    list_petty_cash,
    bulk_delete_petty_cash,
)


router = APIRouter(prefix="/petty-cash", tags=["مدیریت مالی"])


@router.post(
    "/businesses/{business_id}/petty-cash",
    summary="لیست تنخواه گردان‌ها",
    description="دریافت لیست تنخواه گردان‌های یک کسب و کار با امکان جستجو و فیلتر",
)
@require_business_access("business_id")
async def list_petty_cash_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    query_dict: Dict[str, Any] = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "search": query_info.search,
        "search_fields": query_info.search_fields,
        "filters": query_info.filters,
    }
    result = list_petty_cash(db, business_id, query_dict)
    result["items"] = [format_datetime_fields(item, request) for item in result.get("items", [])]
    return success_response(data=result, request=request, message="PETTY_CASH_LIST_FETCHED")


@router.post(
    "/businesses/{business_id}/petty-cash/create",
    summary="ایجاد تنخواه گردان جدید",
    description="ایجاد تنخوان گردان برای یک کسب‌وکار مشخص",
)
@require_business_access("business_id")
async def create_petty_cash_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("petty_cash", "add")),
):
    payload: Dict[str, Any] = dict(body or {})
    created = create_petty_cash(db, business_id, payload)
    return success_response(data=format_datetime_fields(created, request), request=request, message="PETTY_CASH_CREATED")


@router.get(
    "/petty-cash/{petty_cash_id}",
    summary="جزئیات تنخواه گردان",
    description="دریافت جزئیات تنخواه گردان بر اساس شناسه",
)
async def get_petty_cash_endpoint(
    request: Request,
    petty_cash_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("petty_cash", "view", PettyCash, "petty_cash_id")),
):
    result = get_petty_cash_by_id(db, petty_cash_id)
    if not result:
        raise ApiError("PETTY_CASH_NOT_FOUND", "Petty cash not found", http_status=404)
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="PETTY_CASH_DETAILS")


@router.put(
    "/petty-cash/{petty_cash_id}",
    summary="ویرایش تنخواه گردان",
    description="ویرایش اطلاعات تنخواه گردان",
)
async def update_petty_cash_endpoint(
    request: Request,
    petty_cash_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("petty_cash", "edit", PettyCash, "petty_cash_id")),
):
    payload: Dict[str, Any] = dict(body or {})
    result = update_petty_cash(db, petty_cash_id, payload)
    if result is None:
        raise ApiError("PETTY_CASH_NOT_FOUND", "Petty cash not found", http_status=404)
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="PETTY_CASH_UPDATED")


@router.delete(
    "/petty-cash/{petty_cash_id}",
    summary="حذف تنخواه گردان",
    description="حذف یک تنخواه گردان",
)
async def delete_petty_cash_endpoint(
    request: Request,
    petty_cash_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("petty_cash", "delete", PettyCash, "petty_cash_id")),
):
    result = get_petty_cash_by_id(db, petty_cash_id)
    if result:
        try:
            biz_id = int(result.get("business_id"))
        except Exception:
            biz_id = None
        if biz_id is not None and not ctx.can_access_business(biz_id):
            raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    from fastapi import HTTPException
    
    success, error_message = delete_petty_cash(db, petty_cash_id)
    if not success:
        if error_message:
            raise HTTPException(status_code=400, detail=error_message)
        raise ApiError("PETTY_CASH_NOT_FOUND", "Petty cash not found", http_status=404)
    return success_response(data=None, request=request, message="PETTY_CASH_DELETED")


@router.post(
    "/businesses/{business_id}/petty-cash/bulk-delete",
    summary="حذف گروهی تنخواه گردان‌ها",
    description="حذف چندین تنخواه گردان بر اساس شناسه‌ها",
)
@require_business_access("business_id")
async def bulk_delete_petty_cash_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("petty_cash", "delete")),
):
    ids = body.get("ids")
    if not isinstance(ids, list):
        ids = []
    try:
        ids = [int(x) for x in ids if isinstance(x, (int, str)) and str(x).isdigit()]
    except Exception:
        ids = []
    if not ids:
        return success_response({"deleted": 0, "skipped": 0, "total_requested": 0}, request, message="NO_VALID_IDS_FOR_DELETE")
    result = bulk_delete_petty_cash(db, business_id, ids)
    return success_response(result, request, message="PETTY_CASH_BULK_DELETE_DONE")


@router.post(
    "/businesses/{business_id}/petty-cash/export/excel",
    summary="خروجی Excel لیست تنخواه گردان‌ها",
    description="خروجی Excel لیست تنخواه گردان‌ها با قابلیت فیلتر و مرتب‌سازی",
)
@require_business_access("business_id")
async def export_petty_cash_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from app.core.i18n import negotiate_locale

    query_dict = {
        "take": int(body.get("take", 1000)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }
    result = list_petty_cash(db, business_id, query_dict)
    items: List[Dict[str, Any]] = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Map currency_id -> currency title for display
    try:
        from adapters.db.models.currency import Currency
        currency_ids = set()
        for it in items:
            cid = it.get("currency_id")
            try:
                if cid is not None:
                    currency_ids.add(int(cid))
            except Exception:
                pass
        currency_map: Dict[int, str] = {}
        if currency_ids:
            rows = db.query(Currency).filter(Currency.id.in_(list(currency_ids))).all()
            currency_map = {c.id: (c.title or c.code or str(c.id)) for c in rows}
        for it in items:
            cid = it.get("currency_id")
            it["currency"] = currency_map.get(cid, cid)
    except Exception:
        # In case of any issue, fallback without blocking export
        for it in items:
            if "currency" not in it and "currency_id" in it:
                it["currency"] = it.get("currency_id")

    headers: List[str] = [
        "code", "name", "currency", "is_active", "is_default", "description",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "PettyCash"

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(headers, 1):
            value = item.get(key, "")
            if key in ("is_active", "is_default"):
                try:
                    truthy = bool(value) if isinstance(value, bool) else str(value).strip().lower() in ("true", "1", "yes", "y", "t")
                except Exception:
                    truthy = bool(value)
                value = "✓" if truthy else "✗"
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=petty_cash.xlsx",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/petty-cash/export/pdf",
    summary="خروجی PDF لیست تنخواه گردان‌ها",
    description="خروجی PDF لیست تنخواه گردان‌ها با قابلیت فیلتر و مرتب‌سازی",
)
@require_business_access("business_id")
async def export_petty_cash_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from fastapi.responses import Response
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale

    query_dict = {
        "take": int(body.get("take", 1000)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }

    result = list_petty_cash(db, business_id, query_dict)
    items: List[Dict[str, Any]] = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Map currency_id -> currency title for display
    try:
        from adapters.db.models.currency import Currency
        currency_ids = set()
        for it in items:
            cid = it.get("currency_id")
            try:
                if cid is not None:
                    currency_ids.add(int(cid))
            except Exception:
                pass
        currency_map: Dict[int, str] = {}
        if currency_ids:
            rows = db.query(Currency).filter(Currency.id.in_(list(currency_ids))).all()
            currency_map = {c.id: (c.title or c.code or str(c.id)) for c in rows}
        for it in items:
            cid = it.get("currency_id")
            it["currency"] = currency_map.get(cid, cid)
    except Exception:
        for it in items:
            if "currency" not in it and "currency_id" in it:
                it["currency"] = it.get("currency_id")

    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            import json
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                # Replace currency_id key with currency to show human-readable value
                if str(key) == 'currency_id':
                    keys.append('currency')
                else:
                    keys.append(str(key))
                headers.append(str(label))
    else:
        keys = [
            "code", "name", "currency", "is_active", "is_default", "description",
        ]
        headers = keys

    business_name = ""
    try:
        from adapters.db.models.business import Business
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name or ""
    except Exception:
        business_name = ""

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'

    try:
        from app.core.calendar import CalendarConverter
        calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
        formatted_now = CalendarConverter.format_datetime(
            __import__("datetime").datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian",
        )
        now_str = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        from datetime import datetime
        now_str = datetime.now().strftime('%Y/%m/%d %H:%M')

    def esc(v: Any) -> str:
        try:
            return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        except Exception:
            return str(v)

    rows_html: List[str] = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key)
            if value is None:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif key in ("is_active", "is_default"):
                try:
                    truthy = bool(value) if isinstance(value, bool) else str(value).strip().lower() in ("true", "1", "yes", "y", "t")
                except Exception:
                    truthy = bool(value)
                value = "✓" if truthy else "✗"
            tds.append(f"<td>{esc(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = ''.join(f"<th>{esc(h)}</th>" for h in headers)

    # تلاش برای رندر با قالب سفارشی (petty_cash/list)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": 'گزارش تنخواه گردان‌ها' if is_fa else 'Petty Cash Report',
            "business_name": business_name,
            "generated_at": now_str,
            "is_fa": is_fa,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="petty_cash",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    table_html = f"""
    <html lang="{html_lang}" dir="{html_dir}"> 
      <head>
        <meta charset='utf-8'>
        <style>
          @page {{ size: A4 landscape; margin: 12mm; }}
          body {{ font-family: sans-serif; font-size: 11px; color: #222; }}
          .title {{ font-size: 16px; font-weight: 700; margin-bottom: 10px; }}
          table.report-table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
          thead th {{ background: #f0f3f7; border: 1px solid #c7cdd6; padding: 6px 4px; text-align: center; white-space: nowrap; }}
          tbody td {{ border: 1px solid #d7dde6; padding: 5px 4px; vertical-align: top; overflow-wrap: anywhere; }}
        </style>
      </head>
      <body>
        <div class="title">{esc('گزارش تنخواه گردان‌ها' if is_fa else 'Petty Cash Report')}</div>
        <div style="margin-bottom:6px;">{esc('نام کسب‌وکار' if is_fa else 'Business Name')}: {esc(business_name)} | {esc('تاریخ گزارش' if is_fa else 'Report Date')}: {esc(now_str)}</div>
        <table class="report-table"> 
          <thead><tr>{headers_html}</tr></thead>
          <tbody>{''.join(rows_html)}</tbody>
        </table>
      </body>
    </html>
    """

    final_html = resolved_html or table_html
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=petty_cash.pdf",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
