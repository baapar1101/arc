"""
API endpoints برای دریافت و پرداخت (Receipt & Payment)

این ماژول شامل endpoint های مربوط به مدیریت اسناد دریافت و پرداخت است.
اسناد دریافت برای ثبت دریافتی از مشتریان و اسناد پرداخت برای ثبت پرداخت به تامین‌کنندگان استفاده می‌شوند.

### روش‌های پرداخت پشتیبانی شده:
- نقدی (Cash)
- چک (Check)
- کارت بانکی (Card)
- انتقال آنلاین (Online)
- سایر (Other)
"""

from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, Request, Body, Path, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
import io
import json
import datetime
import re
import base64

from adapters.db.session import get_db
from adapters.db.models.document import Document
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.cache import get_cache
from app.core.permissions import require_business_management_dep, require_business_access, require_business_permission_dep, require_business_permission_by_entity_dep
from app.services.receipt_payment_service import DOCUMENT_TYPE_RECEIPT, DOCUMENT_TYPE_PAYMENT
from adapters.api.v1.schemas import QueryInfo, SuccessResponse
from adapters.api.v1.schema_models.receipt_payment import (
    ReceiptPaymentCreateRequest,
    ReceiptPaymentResponse,
    ReceiptPaymentListResponse
)
from app.services.receipt_payment_service import (
    create_receipt_payment,
    get_receipt_payment,
    list_receipts_payments,
    delete_receipt_payment,
    update_receipt_payment,
)
from adapters.db.models.business import Business
from adapters.db.models.user import User
from adapters.db.models.business_print_settings import BusinessPrintSettings
from app.services.file_storage_service import FileStorageService
from app.services.pdf.template_renderer import render_template


router = APIRouter(tags=["دریافت و پرداخت", "مدیریت مالی"])


@router.post(
    "/businesses/{business_id}/receipts-payments",
    summary="لیست اسناد دریافت و پرداخت",
    description="دریافت لیست اسناد دریافت و پرداخت با فیلتر و جستجو",
)
@require_business_access("business_id")
async def list_receipts_payments_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """
    لیست اسناد دریافت و پرداخت
    
    پارامترهای اضافی در body:
    - document_type: "receipt" یا "payment" (اختیاری)
    - from_date: تاریخ شروع (اختیاری)
    - to_date: تاریخ پایان (اختیاری)
    """
    query_dict: Dict[str, Any] = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "sort": [s.model_dump() for s in query_info.sort] if query_info.sort else None,
        "search": query_info.search,
    }
    
    # دریافت پارامترهای اضافی از body
    try:
        body_json = await request.json()
        if isinstance(body_json, dict):
            for key in ["document_type", "from_date", "to_date", "sort", "sort_by", "sort_desc"]:
                if key in body_json:
                    query_dict[key] = body_json[key]
    except Exception:
        pass
    
    # دریافت fiscal_year_id از هدر برای اولویت دادن به انتخاب کاربر
    try:
        fy_header = request.headers.get("X-Fiscal-Year-ID")
        if fy_header:
            query_dict["fiscal_year_id"] = int(fy_header)
    except Exception:
        pass

    # کش نتایج لیست دریافت/پرداخت
    cache = get_cache()
    cache_key = None
    fiscal_year_id = query_dict.get("fiscal_year_id")
    document_type = query_dict.get("document_type")  # "receipt" یا "payment"

    if cache.enabled:
        import json, hashlib
        key_payload = {
            "business_id": business_id,
            "query": query_dict,
        }
        key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
        cache_key = f"receipts_payments_list:{key_hash}"
        cached = cache.get(cache_key)
        if cached is not None:
            return success_response(
                data=cached,
                request=request,
                message="RECEIPTS_PAYMENTS_LIST_FETCHED"
            )

    result = list_receipts_payments(db, business_id, query_dict)
    result["items"] = [format_datetime_fields(item, request) for item in result.get("items", [])]

    # ذخیره در cache با tag-based caching
    if cache.enabled and cache_key:
        cache.set_with_documents_tag(
            key=cache_key,
            value=result,
            business_id=business_id,
            fiscal_year_id=fiscal_year_id,
            document_type=document_type,
            ttl=60
        )
    
    return success_response(
        data=result,
        request=request,
        message="RECEIPTS_PAYMENTS_LIST_FETCHED"
    )


@router.post(
    "/businesses/{business_id}/receipts-payments/create",
    summary="ایجاد سند دریافت یا پرداخت",
    description="""
    ایجاد سند دریافت یا پرداخت جدید
    
    ### انواع سند:
    - **receipt**: دریافت وجه از مشتریان
    - **payment**: پرداخت به تامین‌کنندگان
    
    ### روش‌های پرداخت:
    - `cash` - نقدی
    - `check` - چکی (نیاز به اطلاعات چک)
    - `card` - کارتی
    - `online` - آنلاین
    
    ### نکات:
    - برای چک، اطلاعات چک الزامی است
    - می‌توان به فاکتور مرتبط کرد
    - خودکار در دفتر کل ثبت می‌شود
    """,
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "سند ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "سند دریافت ایجاد شد",
                        "data": {"id": 123, "code": "REC-1001"}
                    }
                }
            }
        }
    }
)
@require_business_access("business_id")
async def create_receipt_payment_endpoint(
    request: Request,
    business_id: int = Path(..., description="شناسه کسب‌وکار", examples={"example": {"value": 1}}, gt=0),
    body: Dict[str, Any] = Body(..., description="اطلاعات سند"),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("people_transactions", "add")),
):
    """
    ایجاد سند دریافت یا پرداخت
    
    Body باید شامل موارد زیر باشد:
    {
        "document_type": "receipt" | "payment",
        "document_date": "2025-01-15T10:30:00",
        "currency_id": 1,
        "description": "توضیحات کلی سند (اختیاری)",
        "person_lines": [
            {
                "person_id": 123,
                "person_name": "علی احمدی",
                "amount": 1000000,
                "description": "توضیحات (اختیاری)"
            }
        ],
        "account_lines": [
            {
                "account_id": 456,
                "amount": 1000000,
                "transaction_type": "bank" | "cash_register" | "petty_cash" | "check",
                "transaction_date": "2025-01-15T10:30:00",
                "commission": 5000,  // اختیاری
                "description": "توضیحات (اختیاری)",
                // اطلاعات اضافی بر اساس نوع تراکنش:
                "bank_id": "123",  // برای نوع bank
                "bank_name": "بانک ملی",
                "cash_register_id": "456",  // برای نوع cash_register
                "cash_register_name": "صندوق اصلی",
                "petty_cash_id": "789",  // برای نوع petty_cash
                "petty_cash_name": "تنخواهگردان فروش",
                "check_id": "101",  // برای نوع check
                "check_number": "123456"
            }
        ],
        "extra_info": {}  // اختیاری
    }
    """
    created = create_receipt_payment(db, business_id, ctx.get_user_id(), body)
    
    return success_response(
        data=format_datetime_fields(created, request),
        request=request,
        message="RECEIPT_PAYMENT_CREATED"
    )


@router.get(
    "/receipts-payments/{document_id}",
    summary="جزئیات سند دریافت/پرداخت",
    description="دریافت جزئیات یک سند دریافت یا پرداخت",
)
async def get_receipt_payment_endpoint(
    request: Request,
    document_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """دریافت جزئیات سند"""
    result = get_receipt_payment(db, document_id)
    
    if not result:
        raise ApiError(
            "DOCUMENT_NOT_FOUND",
            "Receipt/Payment document not found",
            http_status=404
        )
    
    # بررسی دسترسی
    business_id = result.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    return success_response(
        data=format_datetime_fields(result, request),
        request=request,
        message="RECEIPT_PAYMENT_DETAILS"
    )


def _validate_receipt_payment_document(document: Document) -> bool:
    """Validator برای بررسی اینکه سند از نوع receipt یا payment باشد"""
    return document.document_type in (DOCUMENT_TYPE_RECEIPT, DOCUMENT_TYPE_PAYMENT)


@router.delete(
    "/receipts-payments/{document_id}",
    summary="حذف سند دریافت/پرداخت",
    description="حذف یک سند دریافت یا پرداخت",
)
async def delete_receipt_payment_endpoint(
    request: Request,
    document_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people_transactions", "delete", Document, "document_id", entity_validator=_validate_receipt_payment_document)),
):
    """حذف سند"""
    # دریافت سند برای بررسی دسترسی
    result = get_receipt_payment(db, document_id)
    
    if result:
        business_id = result.get("business_id")
        if business_id and not ctx.can_access_business(business_id):
            raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    ok = delete_receipt_payment(db, document_id)
    
    if not ok:
        raise ApiError(
            "DOCUMENT_NOT_FOUND",
            "Receipt/Payment document not found",
            http_status=404
        )
    
    return success_response(
        data=None,
        request=request,
        message="RECEIPT_PAYMENT_DELETED"
    )


@router.put(
    "/receipts-payments/{document_id}",
    summary="ویرایش سند دریافت/پرداخت",
    description="به‌روزرسانی یک سند دریافت یا پرداخت",
)
async def update_receipt_payment_endpoint(
    request: Request,
    document_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people_transactions", "edit", Document, "document_id", entity_validator=_validate_receipt_payment_document)),
):
    """ویرایش سند"""
    # دریافت سند برای بررسی دسترسی
    result = get_receipt_payment(db, document_id)
    if not result:
        raise ApiError("DOCUMENT_NOT_FOUND", "Receipt/Payment document not found", http_status=404)

    business_id = result.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)

    updated = update_receipt_payment(db, document_id, ctx.get_user_id(), body)
    return success_response(
        data=format_datetime_fields(updated, request),
        request=request,
        message="RECEIPT_PAYMENT_UPDATED",
    )


@router.post(
    "/businesses/{business_id}/receipts-payments/export/excel",
    summary="خروجی Excel لیست اسناد دریافت و پرداخت",
    description="خروجی Excel لیست اسناد دریافت و پرداخت با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_receipts_payments_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel لیست اسناد دریافت و پرداخت"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.core.i18n import negotiate_locale
    
    # Build query dict from flat body
    # For export, we limit to reasonable number to prevent memory issues
    max_export_records = 10000
    take_value = min(int(body.get("take", 1000)), max_export_records)
    
    query_dict = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
        "document_type": body.get("document_type"),
        "from_date": body.get("from_date"),
        "to_date": body.get("to_date"),
    }

    result = list_receipts_payments(db, business_id, query_dict)
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        # Add a warning row to indicate data was truncated
        warning_item = {
            "code": "⚠️ هشدار",
            "document_type": "حداکثر ۱۰,۰۰۰ رکورد قابل export است",
            "document_date": "",
            "total_amount": "",
            "person_lines_count": "",
            "account_lines_count": "",
            "created_by_name": "",
            "registered_at": "",
        }
        items.append(warning_item)

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare headers based on export_columns (order + visibility)
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns for receipts/payments
        default_columns = [
            ('code', 'کد سند'),
            ('document_type_name', 'نوع سند'),
            ('document_date', 'تاریخ سند'),
            ('total_amount', 'مبلغ کل'),
            ('person_names', 'اشخاص'),
            ('account_lines_count', 'تعداد حساب‌ها'),
            ('created_by_name', 'ایجادکننده'),
            ('registered_at', 'تاریخ ثبت'),
        ]
        for key, label in default_columns:
            if items and key in items[0]:
                keys.append(key)
                headers.append(label)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts & Payments"

    # Locale and RTL/LTR handling
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # RTL alignment for Persian text
            if locale == 'fa' and isinstance(value, str) and any('\u0600' <= c <= '\u06FF' for c in value):
                cell.alignment = Alignment(horizontal="right")

    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "receipts_payments"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.get(
    "/receipts-payments/{document_id}/pdf",
    summary="خروجی PDF تک سند دریافت/پرداخت",
    description="خروجی PDF یک سند دریافت یا پرداخت",
)
async def export_single_receipt_payment_pdf(
    document_id: int,
    request: Request,
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    template_id: int | None = None,
):
    """خروجی PDF تک سند دریافت/پرداخت"""
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from html import escape
    
    # دریافت سند
    result = get_receipt_payment(db, document_id)
    if not result:
        raise ApiError(
            "DOCUMENT_NOT_FOUND",
            "Receipt/Payment document not found",
            http_status=404
        )
    
    # بررسی دسترسی
    business_id = result.get("business_id")
    if business_id and not auth_context.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    # دریافت اطلاعات کسب‌وکار + فایل‌های گرافیکی (لوگو/مهر/امضا)
    business_name = ""
    business_logo_data_uri: Optional[str] = None
    business_stamp_data_uri: Optional[str] = None
    owner_signature_data_uri: Optional[str] = None
    storage = FileStorageService(db)

    async def _load_image_data_uri(file_id_str: Optional[str]) -> Optional[str]:
        if not file_id_str:
            return None
        try:
            from uuid import UUID
            try:
                file_data = await storage.download_file(UUID(str(file_id_str)))
            except Exception:
                return None
            content: bytes = file_data.get("content") or b""
            if not content:
                return None
            mime = file_data.get("mime_type") or "image/png"
            b64 = base64.b64encode(content).decode("ascii")
            return f"data:{mime};base64,{b64}"
        except Exception:
            return None

    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
            # تنظیمات چاپ (all یا بر اساس نوع سند receipt/payment)
            try:
                rows = (
                    db.query(BusinessPrintSettings)
                    .filter(BusinessPrintSettings.business_id == business_id)
                    .all()
                )
            except Exception:
                rows = []

            def _pick_cfg() -> dict:
                cfg = {"show_logo": True, "show_stamp": True, "footer_note": None}
                per_type = None
                for r in rows:
                    if r.document_type == "all":
                        cfg = {
                            "show_logo": bool(getattr(r, "show_logo", True)),
                            "show_stamp": bool(getattr(r, "show_stamp", True)),
                            "footer_note": getattr(r, "footer_note", None),
                        }
                    elif r.document_type in ("receipt", "payment"):
                        per_type = {
                            "show_logo": bool(getattr(r, "show_logo", True)),
                            "show_stamp": bool(getattr(r, "show_stamp", True)),
                            "footer_note": getattr(r, "footer_note", None),
                        }
                if per_type:
                    merged = dict(cfg)
                    merged.update({k: v for k, v in per_type.items() if v is not None})
                    return merged
                return cfg

            cfg = _pick_cfg()
            if cfg.get("show_logo", True):
                business_logo_data_uri = await _load_image_data_uri(getattr(b, "logo_file_id", None))
            if cfg.get("show_stamp", True):
                business_stamp_data_uri = await _load_image_data_uri(getattr(b, "stamp_file_id", None))
                try:
                    owner_user = db.query(User).filter(User.id == b.owner_id).first()
                except Exception:
                    owner_user = None
                if owner_user is not None:
                    owner_signature_data_uri = await _load_image_data_uri(getattr(owner_user, "signature_file_id", None))
    except Exception:
        business_name = business_name or ""

    # Locale handling
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # آماده‌سازی داده‌ها
    doc_type_name = result.get("document_type_name", "")
    doc_code = result.get("code", "")
    doc_date = result.get("document_date", "")
    total_amount = result.get("total_amount", 0)
    description = result.get("description", "")
    person_lines = result.get("person_lines", [])
    account_lines = result.get("account_lines", [])
    
    # تاریخ تولید
    now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    title_text = f"سند {doc_type_name}" if is_fa else f"{doc_type_name} Document"
    label_biz = "کسب و کار" if is_fa else "Business"
    label_date = "تاریخ تولید" if is_fa else "Generated Date"
    footer_text = f"تولید شده در {now}" if is_fa else f"Generated at {now}"

    # تلاش برای رندر با قالب سفارشی (receipts_payments/detail)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if template_id is not None:
                explicit_template_id = int(template_id)
        except Exception:
            explicit_template_id = None
        template_context = {
            "business_id": business_id,
            "business_name": business_name,
            "document": result,
            "person_lines": person_lines,
            "account_lines": account_lines,
            "code": doc_code,
            "document_date": doc_date,
            "total_amount": total_amount,
            "description": description,
            "title_text": title_text,
            "generated_at": now,
            "is_fa": is_fa,
            "business_logo_data_uri": business_logo_data_uri,
            "business_stamp_data_uri": business_stamp_data_uri,
            "owner_signature_data_uri": owner_signature_data_uri,
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="receipts_payments",
            subtype="detail",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # HTML پیش‌فرض در نبود قالب: فایل قالب + پارامترها
    try:
        qp = request.query_params
        paper_size = qp.get("paper_size")
        orientation = qp.get("orientation")
        disposition = qp.get("disposition") or "attachment"
    except Exception:
        paper_size = None
        orientation = None
        disposition = "attachment"
    html_content = resolved_html or render_template(
        "pdf/receipts_payments/detail.html",
        {
            "business_id": business_id,
            "business_name": business_name,
            "document": result,
            "person_lines": person_lines,
            "account_lines": account_lines,
            "code": doc_code,
            "document_date": doc_date,
            "total_amount": total_amount,
            "description": description,
            "title_text": title_text,
            "generated_at": now,
            "is_fa": is_fa,
            "paper_size": paper_size,
            "orientation": orientation,
            "footer_text": footer_text,
            "business_logo_data_uri": business_logo_data_uri,
            "business_stamp_data_uri": business_stamp_data_uri,
            "owner_signature_data_uri": owner_signature_data_uri,
        },
    )

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=html_content).write_pdf(font_config=font_config)

    # Build filename
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    filename = f"receipt_payment_{slugify(doc_code)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/receipts-payments/export/pdf",
    summary="خروجی PDF لیست اسناد دریافت و پرداخت",
    description="خروجی PDF لیست اسناد دریافت و پرداخت با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_receipts_payments_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی PDF لیست اسناد دریافت و پرداخت"""
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from html import escape
    
    # Build query dict from flat body
    # For export, we limit to reasonable number to prevent memory issues
    max_export_records = 10000
    take_value = min(int(body.get("take", 1000)), max_export_records)
    
    query_dict = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
        "document_type": body.get("document_type"),
        "from_date": body.get("from_date"),
        "to_date": body.get("to_date"),
    }

    result = list_receipts_payments(db, business_id, query_dict)
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare headers and data
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns for receipts/payments
        default_columns = [
            ('code', 'کد سند'),
            ('document_type_name', 'نوع سند'),
            ('document_date', 'تاریخ سند'),
            ('total_amount', 'مبلغ کل'),
            ('person_names', 'اشخاص'),
            ('account_lines_count', 'تعداد حساب‌ها'),
            ('created_by_name', 'ایجادکننده'),
            ('registered_at', 'تاریخ ثبت'),
        ]
        for key, label in default_columns:
            if items and key in items[0]:
                keys.append(key)
                headers.append(label)

    # Get business name
    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    # Locale handling
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Prepare data for HTML
    now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    title_text = "لیست اسناد دریافت و پرداخت" if is_fa else "Receipts & Payments List"
    label_biz = "کسب و کار" if is_fa else "Business"
    label_date = "تاریخ تولید" if is_fa else "Generated Date"
    footer_text = f"تولید شده در {now}" if is_fa else f"Generated at {now}"

    # Create headers HTML
    headers_html = ''.join(f'<th>{escape(header)}</th>' for header in headers)
    
    # Create rows HTML
    rows_html = []
    for item in items:
        row_cells = []
        for key in keys:
            value = item.get(key, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            row_cells.append(f'<td>{escape(str(value))}</td>')
        rows_html.append(f'<tr>{"".join(row_cells)}</tr>')

    # کانتکست برای قالب سفارشی لیست
    template_context: Dict[str, Any] = {
        "title_text": title_text,
        "business_name": business_name,
        "generated_at": now,
        "is_fa": is_fa,
        "headers": headers,
        "keys": keys,
        "items": items,
        "table_headers_html": headers_html,
        "table_rows_html": "".join(rows_html),
    }

    # تلاش برای رندر با قالب سفارشی (receipts_payments/list)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="receipts_payments",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # HTML پیش‌فرض جدول با قالب فایل
    disposition = "attachment"
    try:
        disposition = str(body.get("disposition") or "attachment")
    except Exception:
        disposition = "attachment"
    paper_size = None
    orientation = None
    try:
        paper_size = body.get("paper_size")
        orientation = body.get("orientation")
    except Exception:
        pass
    final_html = resolved_html or render_template(
        "pdf/receipts_payments/list.html",
        {
            **template_context,
            "title_text": title_text,
            "paper_size": paper_size,
            "orientation": orientation,
            "footer_text": footer_text,
        },
    )

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)

    # Build meaningful filename
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "receipts_payments"
    if business_name:
        base += f"_{slugify(business_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )

