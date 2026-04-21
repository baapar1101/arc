"""
API endpoints برای مدیریت اسناد حسابداری (General Accounting Documents)
"""

from typing import Any, Dict, List
from fastapi import APIRouter, Depends, Request, Body, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import and_

from adapters.db.session import get_db
from adapters.db.models.document import Document
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.permissions import require_business_access, require_business_management_dep, require_business_permission_dep, require_business_permission_by_entity_dep
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.cache import get_cache
from app.core.response_cache import cache_response
from app.services.document_service import (
    list_documents,
    get_document,
    delete_document,
    delete_multiple_documents,
    get_document_types_summary,
    export_documents_excel,
    create_manual_document,
    update_manual_document,
)
from app.services.invoice_service import get_daily_sales_report, get_monthly_sales_report, get_top_customers_report, get_daily_purchases_report, get_top_suppliers_report, get_materials_consumption_report, get_production_report
from app.services.trial_balance_service import get_trial_balance_report
from app.services.general_ledger_service import get_general_ledger_report
from app.services.pnl_service import get_pnl_period_report, get_pnl_cumulative_report
from app.services.account_review_service import get_accounts_review_report
from app.services.journal_ledger_service import get_journal_ledger_report
from app.core.cache import get_cache
from app.core.i18n import negotiate_locale
from app.services.pdf.template_renderer import render_template, load_farsi_font_data_uris
from adapters.api.v1.schema_models.document import (
    CreateManualDocumentRequest,
    UpdateManualDocumentRequest,
)


router = APIRouter(tags=["حسابداری"])


@router.post(
    "/businesses/{business_id}/documents",
    summary="لیست اسناد حسابداری",
    description="دریافت لیست تمام اسناد حسابداری (عمومی و اتوماتیک) با فیلتر و صفحه‌بندی",
)
@require_business_access("business_id")
async def list_documents_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """
    لیست اسناد حسابداری
    
    Body parameters:
        - document_type: نوع سند (expense, income, receipt, payment, transfer, manual)
        - fiscal_year_id: شناسه سال مالی
        - from_date: از تاریخ (ISO format)
        - to_date: تا تاریخ (ISO format)
        - currency_id: شناسه ارز
        - is_proforma: پیش‌فاکتور یا قطعی
        - search: جستجو در کد سند و توضیحات
        - sort_by: فیلد مرتب‌سازی (document_date, code, document_type, created_at)
        - sort_desc: ترتیب نزولی (true/false)
        - take: تعداد رکورد (1-1000)
        - skip: تعداد رکورد صرف‌نظر شده
    """
    query_dict: Dict[str, Any] = {
        "take": body.get("take", 50),
        "skip": body.get("skip", 0),
        "sort_by": body.get("sort_by", "document_date"),
        "sort_desc": body.get("sort_desc", True),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
    }

    # فیلترهای اضافی
    for key in [
        "document_type",
        "from_date",
        "to_date",
        "currency_id",
        "is_proforma",
        "project_id",
        "person_id",
        "search_fields",
        "filters",
    ]:
        if key in body:
            query_dict[key] = body[key]

    # سال مالی: بدنهٔ درخواست اولویت دارد، سپس هدر
    try:
        if body.get("fiscal_year_id") is not None:
            query_dict["fiscal_year_id"] = int(body["fiscal_year_id"])
        else:
            fy_header = request.headers.get("X-Fiscal-Year-ID")
            if fy_header:
                query_dict["fiscal_year_id"] = int(fy_header)
    except Exception:
        pass

    # کش نتایج لیست اسناد بر اساس پارامترها
    cache = get_cache()
    cache_key = None
    fiscal_year_id = query_dict.get("fiscal_year_id")
    document_type = query_dict.get("document_type")

    if cache.enabled:
        import json, hashlib
        key_payload = {
            "business_id": business_id,
            "query": query_dict,
        }
        key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
        cache_key = f"documents_list:{key_hash}"
        cached = cache.get(cache_key)
        if cached is not None:
            return success_response(
                data=cached,
                request=request,
                message="DOCUMENTS_LIST_FETCHED"
            )

    result = list_documents(db, business_id, query_dict)
    
    # فرمت کردن تاریخ‌ها
    result["items"] = [
        format_datetime_fields(item, request) for item in result.get("items", [])
    ]

    # ذخیره در cache با tag-based caching
    if cache.enabled and cache_key:
        cache.set_with_documents_tag(
            key=cache_key,
            value=result,
            business_id=business_id,
            fiscal_year_id=fiscal_year_id,
            document_type=document_type,
            ttl=60
        )
    
    return success_response(
        data=result,
        request=request,
        message="DOCUMENTS_LIST_FETCHED"
    )


@router.post(
    "/businesses/{business_id}/documents/export/pdf",
    summary="خروجی PDF لیست اسناد حسابداری",
    description="دریافت فایل PDF لیست اسناد حسابداری با پشتیبانی از قالب سفارشی (documents/list)",
)
@require_business_access("business_id")
async def export_documents_pdf_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """خروجی PDF لیست اسناد حسابداری"""
    from fastapi.responses import Response
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from html import escape
    import datetime, json
    # فیلترهایی مشابه export_documents_excel
    filters = {}
    for key in [
        "document_type",
        "from_date",
        "to_date",
        "currency_id",
        "is_proforma",
        "search",
        "search_fields",
        "filters",
        "project_id",
        "person_id",
    ]:
        if key in body:
            filters[key] = body[key]
    try:
        if body.get("fiscal_year_id") is not None:
            filters["fiscal_year_id"] = int(body["fiscal_year_id"])
        else:
            fy_header = request.headers.get("X-Fiscal-Year-ID")
            if fy_header:
                filters["fiscal_year_id"] = int(fy_header)
    except Exception:
        pass
    # دریافت داده‌ها (شامل مرتب‌سازی چندستونه اگر ارسال شده باشد)
    _sort_extra: Dict[str, Any] = {"take": body.get("take", 1000), "skip": body.get("skip", 0)}
    for _k in ("sort_by", "sort_desc", "sort"):
        if _k in body:
            _sort_extra[_k] = body[_k]
    result = list_documents(db, business_id, {**filters, **_sort_extra})
    items = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]
    # ستون‌ها
    headers: list[str] = []
    keys: list[str] = []
    export_columns = body.get("export_columns")
    if export_columns:
        for col in export_columns:
            key = col.get("key")
            label = col.get("label", key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        default_columns = [
            ("code", "کد سند"),
            ("document_type_name", "نوع سند"),
            ("document_date", "تاریخ سند"),
            ("total_debit", "جمع بدهکار"),
            ("total_credit", "جمع بستانکار"),
            ("created_by_name", "ایجادکننده"),
            ("registered_at", "تاریخ ثبت"),
        ]
        for key, label in default_columns:
            if items and key in items[0]:
                keys.append(key)
                headers.append(label)
    # اطلاعات کسب‌وکار
    business_name = ""
    try:
        from adapters.db.models.business import Business
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""
    # Locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    title_text = "لیست اسناد حسابداری" if is_fa else "Documents List"
    label_biz = "کسب و کار" if is_fa else "Business"
    label_date = "تاریخ تولید" if is_fa else "Generated Date"
    footer_text = f"تولید شده در {now}" if is_fa else f"Generated at {now}"
    headers_html = ''.join(f'<th>{escape(header)}</th>' for header in headers)
    rows_html = []
    for item in items:
        row_cells = []
        for key in keys:
            value = item.get(key, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = json.dumps(value, ensure_ascii=False)
            row_cells.append(f'<td>{escape(str(value))}</td>')
        rows_html.append(f'<tr>{"".join(row_cells)}</tr>')
    # کانتکست قالب
    template_context = {
        "title_text": title_text,
        "business_name": business_name,
        "generated_at": now,
        "is_fa": is_fa,
        "headers": headers,
        "keys": keys,
        "items": items,
        "table_headers_html": headers_html,
        "table_rows_html": "".join(rows_html),
    }
    # تلاش برای رندر با قالب سفارشی
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="documents",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None
    # HTML پیش‌فرض با قالب فایل + پارامترها
    disposition = "attachment"
    try:
        disposition = str(body.get("disposition") or "attachment")
    except Exception:
        disposition = "attachment"
    paper_size = None
    orientation = None
    try:
        paper_size = body.get("paper_size")
        orientation = body.get("orientation")
    except Exception:
        pass
    html_content = resolved_html or render_template(
        "pdf/documents/list.html",
        {
            **template_context,
            "title_text": title_text,
            "paper_size": paper_size,
            "orientation": orientation,
            "footer_text": footer_text,
        },
    )
    pdf_bytes = HTML(string=html_content).write_pdf(font_config=FontConfiguration())
    filename = f"documents_{business_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
@router.get(
    "/documents/{document_id}",
    summary="جزئیات سند حسابداری",
    description="دریافت جزئیات کامل یک سند شامل تمام سطرهای سند",
)
async def get_document_endpoint(
    request: Request,
    document_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """دریافت جزئیات کامل سند"""
    result = get_document(db, document_id)
    
    if not result:
        raise ApiError(
            "DOCUMENT_NOT_FOUND",
            "Document not found",
            http_status=404
        )
    
    # بررسی دسترسی
    business_id = result.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    return success_response(
        data=format_datetime_fields(result, request),
        request=request,
        message="DOCUMENT_DETAILS_FETCHED"
    )


@router.delete(
    "/documents/{document_id}",
    summary="حذف سند حسابداری",
    description="حذف یک سند حسابداری (فقط اسناد عمومی manual قابل حذف هستند)",
)
async def delete_document_endpoint(
    request: Request,
    document_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_management_dep),
):
    """
    حذف سند حسابداری
    
    توجه: فقط اسناد عمومی (manual) قابل حذف هستند.
    اسناد اتوماتیک (expense, income, receipt, payment, ...) باید از منبع اصلی حذف شوند.
    """
    # دریافت سند برای بررسی دسترسی
    doc = get_document(db, document_id)
    if not doc:
        raise ApiError("DOCUMENT_NOT_FOUND", "Document not found", http_status=404)
    
    business_id = doc.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    # حذف سند
    success = delete_document(db, document_id)
    
    return success_response(
        data={"deleted": success, "document_id": document_id},
        request=request,
        message="DOCUMENT_DELETED"
    )


@router.post(
    "/documents/bulk-delete",
    summary="حذف گروهی اسناد",
    description="حذف گروهی اسناد حسابداری (فقط اسناد manual حذف می‌شوند)",
)
async def bulk_delete_documents_endpoint(
    request: Request,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """
    حذف گروهی اسناد
    
    Body:
        document_ids: لیست شناسه‌های سند
    
    توجه: اسناد اتوماتیک نادیده گرفته می‌شوند و باید از منبع اصلی حذف شوند.
    """
    document_ids = body.get("document_ids", [])
    if not document_ids:
        raise ApiError(
            "INVALID_REQUEST",
            "document_ids is required",
            http_status=400
        )
    
    # بررسی دسترسی برای هر document
    # اگر business_id ندارند یا دسترسی ندارند، خطا برمی‌گردانیم
    from adapters.db.models.document import Document as DocumentModel
    from adapters.db.repositories.business_permission_repo import BusinessPermissionRepository
    permission_repo = BusinessPermissionRepository(db)
    
    for doc_id in document_ids:
        try:
            doc = db.get(DocumentModel, doc_id)
            if doc:
                business_id = doc.business_id
                if not ctx.can_access_business(business_id):
                    raise ApiError("FORBIDDEN", f"No access to document {doc_id}", http_status=403)
                
                # بررسی دسترسی جزئی برای business_id مشخص
                if ctx.is_superadmin() or ctx.is_business_owner(business_id):
                    continue  # SuperAdmin و مالک تمام دسترسی‌ها را دارند
                
                permission_obj = permission_repo.get_by_user_and_business(ctx.get_user_id(), business_id)
                if not permission_obj or not permission_obj.business_permissions:
                    raise ApiError("FORBIDDEN", f"Missing permission: accounting_documents.delete for document {doc_id}", http_status=403)
                
                permissions = ctx._normalize_permissions_value(permission_obj.business_permissions)
                if "accounting_documents" not in permissions:
                    raise ApiError("FORBIDDEN", f"Missing permission: accounting_documents.delete for document {doc_id}", http_status=403)
                
                section_perms = permissions.get("accounting_documents", {})
                if not section_perms.get("delete", False):
                    raise ApiError("FORBIDDEN", f"Missing permission: accounting_documents.delete for document {doc_id}", http_status=403)
        except ApiError:
            raise
        except Exception:
            continue
    
    result = delete_multiple_documents(db, document_ids)
    
    return success_response(
        data=result,
        request=request,
        message="DOCUMENTS_BULK_DELETED"
    )


@router.get(
    "/businesses/{business_id}/documents/types-summary",
    summary="خلاصه آماری انواع اسناد",
    description="دریافت خلاصه آماری تعداد هر نوع سند",
)
@require_business_access("business_id")
async def get_document_types_summary_endpoint(
    request: Request,
    business_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """دریافت خلاصه آماری انواع اسناد"""
    summary = get_document_types_summary(db, business_id)
    
    total = sum(summary.values())
    
    return success_response(
        data={"summary": summary, "total": total},
        request=request,
        message="DOCUMENT_TYPES_SUMMARY_FETCHED"
    )


@router.post(
    "/businesses/{business_id}/documents/export/excel",
    summary="خروجی Excel اسناد",
    description="دریافت فایل Excel لیست اسناد حسابداری",
)
@require_business_access("business_id")
async def export_documents_excel_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """
    خروجی Excel لیست اسناد
    
    Body: فیلترهای مشابه لیست اسناد
    """
    filters = {}
    
    # فیلترها
    for key in [
        "document_type",
        "from_date",
        "to_date",
        "currency_id",
        "is_proforma",
        "sort_by",
        "sort_desc",
        "sort",
        "search",
        "search_fields",
        "filters",
        "project_id",
        "person_id",
    ]:
        if key in body:
            filters[key] = body[key]

    try:
        if body.get("fiscal_year_id") is not None:
            filters["fiscal_year_id"] = int(body["fiscal_year_id"])
        else:
            fy_header = request.headers.get("X-Fiscal-Year-ID")
            if fy_header:
                filters["fiscal_year_id"] = int(fy_header)
    except Exception:
        pass
    
    excel_data = export_documents_excel(db, business_id, filters)
    
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=documents_{business_id}.xlsx"
        }
    )


@router.get(
    "/documents/{document_id}/pdf",
    summary="PDF یک سند",
    description="دریافت فایل PDF یک سند حسابداری",
)
async def get_document_pdf_endpoint(
    request: Request,
    document_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    template_id: int | None = None,
):
    """PDF یک سند حسابداری (حداقل اطلاعات ضروری)."""
    # بررسی دسترسی
    doc = get_document(db, document_id)
    if not doc:
        raise ApiError("DOCUMENT_NOT_FOUND", "Document not found", http_status=404)
    
    business_id = doc.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    # رندر با قالب سفارشی (documents/detail) یا خروجی پیش‌فرض
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from html import escape
    import datetime, re

    # اطلاعات کسب‌وکار
    business_name = ""
    try:
        from adapters.db.models.business import Business
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    # Locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    now_dt = datetime.datetime.now()

    # فونت فارسی (data URI) برای خروجی بهتر در PDF
    fa_font_url_regular, fa_font_url_bold = load_farsi_font_data_uris()

    # تقویم انتخابی کاربر (X-Calendar-Type) یا پیش‌فرض بر اساس زبان
    try:
        from app.core.calendar import CalendarConverter, get_calendar_type_from_header
        calendar_type = None
        try:
            if hasattr(request.state, "calendar_type") and request.state.calendar_type:
                calendar_type = request.state.calendar_type
        except Exception:
            calendar_type = None
        if not calendar_type:
            cal_header = request.headers.get("X-Calendar-Type")
            if cal_header:
                calendar_type = get_calendar_type_from_header(cal_header)
            else:
                calendar_type = "jalali" if is_fa else "gregorian"
    except Exception:
        calendar_type = "jalali" if is_fa else "gregorian"

    def _trim_seconds(s: str | None) -> str | None:
        if not s:
            return s
        # YYYY/MM/DD HH:MM:SS -> YYYY/MM/DD HH:MM
        # YYYY-MM-DD HH:MM:SS -> YYYY-MM-DD HH:MM
        if len(s) >= 3 and s[-3] == ":" and s[-2:].isdigit():
            return s[:-3]
        return s

    formatted_now = None
    try:
        formatted_now = CalendarConverter.format_datetime(now_dt, calendar_type)
    except Exception:
        formatted_now = None
    generated_at_display = _trim_seconds(
        (formatted_now or {}).get("formatted") if formatted_now else None
    ) or now_dt.strftime("%Y-%m-%d %H:%M")
    
    # تاریخ سند (شمسی/میلادی بر اساس تقویم انتخابی)
    document_date_jalali = None
    document_date_display = None
    if doc.get("document_date"):
        try:
            dt = datetime.datetime.fromisoformat(str(doc.get("document_date")).replace("Z", "+00:00"))
            # برای سازگاری با خروجی قبلی
            try:
                formatted_j = CalendarConverter.format_datetime(dt, "jalali")
                document_date_jalali = formatted_j.get("date_only") or formatted_j.get("formatted")
            except Exception:
                document_date_jalali = None
            formatted_sel = CalendarConverter.format_datetime(dt, calendar_type)
            document_date_display = formatted_sel.get("date_only") or formatted_sel.get("formatted")
        except Exception:
            document_date_display = str(doc.get("document_date"))

    # زمان ثبت سند (registered_at) بر اساس تقویم انتخابی
    registered_at_display = None
    reg_raw = doc.get("registered_at")
    if reg_raw:
        try:
            if isinstance(reg_raw, datetime.datetime):
                reg_dt = reg_raw
            else:
                reg_dt = datetime.datetime.fromisoformat(str(reg_raw).replace("Z", "+00:00"))
            reg_fmt = CalendarConverter.format_datetime(reg_dt, calendar_type)
            registered_at_display = _trim_seconds(reg_fmt.get("formatted"))
        except Exception:
            registered_at_display = str(reg_raw)
    
    # جمع‌آوری اطلاعات assets (مثلاً لوگو کسب‌وکار)
    business_logo = None
    try:
        from adapters.db.models.business import Business
        b = db.query(Business).filter(Business.id == business_id).first()
        if b and hasattr(b, 'logo_url') and b.logo_url:
            business_logo = b.logo_url
    except Exception:
        pass

    # امضای تهیه‌کننده (کاربری که سند را ایجاد کرده)
    prepared_by_signature_data_uri = None
    try:
        import base64
        from uuid import UUID
        from adapters.db.models.user import User
        from app.services.file_storage_service import FileStorageService

        created_by_user_id = doc.get("created_by_user_id")
        if created_by_user_id:
            u = db.query(User).filter(User.id == int(created_by_user_id)).first()
        else:
            u = None

        sig_file_id = getattr(u, "signature_file_id", None) if u else None
        if sig_file_id:
            storage = FileStorageService(db)
            try:
                file_data = await storage.download_file(UUID(str(sig_file_id)))
                content: bytes = file_data.get("content") or b""
                if content:
                    mime = file_data.get("mime_type") or "image/png"
                    prepared_by_signature_data_uri = (
                        f"data:{mime};base64,{base64.b64encode(content).decode('ascii')}"
                    )
            except Exception:
                prepared_by_signature_data_uri = None
    except Exception:
        prepared_by_signature_data_uri = None

    # غنی‌سازی خطوط سند برای PDF:
    # - ساخت مسیر حساب از ریشه تا حساب فعلی (برای راهنمایی حسابدار)
    try:
        from adapters.db.models.account import Account

        raw_lines = doc.get("lines", []) or []
        account_ids: set[int] = set()
        for ln in raw_lines:
            try:
                aid = ln.get("account_id")
                if aid is not None:
                    account_ids.add(int(aid))
            except Exception:
                continue

        nodes: dict[int, dict[str, Any]] = {}
        missing: set[int] = set(account_ids)
        # Load all needed accounts and their ancestors (iteratively) to avoid N+1
        safety_iters = 0
        while missing and safety_iters < 50:
            safety_iters += 1
            rows = (
                db.query(Account.id, Account.code, Account.name, Account.parent_id)
                .filter(Account.id.in_(list(missing)))
                .all()
            )
            missing = set()
            for r in rows:
                nodes[int(r.id)] = {
                    "id": int(r.id),
                    "code": (r.code or "").strip(),
                    "name": (r.name or "").strip(),
                    "parent_id": int(r.parent_id) if r.parent_id else None,
                }
            # add parents not loaded yet
            for n in list(nodes.values()):
                pid = n.get("parent_id")
                if pid and pid not in nodes:
                    missing.add(int(pid))

        _path_memo: dict[int, list[dict[str, Any]]] = {}

        def _account_chain(aid: int) -> list[dict[str, Any]]:
            if aid in _path_memo:
                return _path_memo[aid]
            chain: list[dict[str, Any]] = []
            seen: set[int] = set()
            cur = aid
            depth = 0
            while cur and cur not in seen and depth < 30:
                depth += 1
                seen.add(cur)
                n = nodes.get(cur)
                if not n:
                    break
                chain.append(n)
                cur = n.get("parent_id")
            chain = list(reversed(chain))
            _path_memo[aid] = chain
            return chain

        def _fmt_node(n: dict[str, Any]) -> str:
            code = (n.get("code") or "").strip()
            name = (n.get("name") or "").strip()
            if code and name:
                return f"{code} {name}"
            return code or name or "-"

        for ln in raw_lines:
            try:
                aid = ln.get("account_id")
                if aid is None:
                    continue
                aid_i = int(aid)
                chain = _account_chain(aid_i)
                if not chain or len(chain) <= 1:
                    ln["account_tree_ancestors"] = ""
                else:
                    ancestors = chain[:-1]
                    ln["account_tree_ancestors"] = " / ".join([_fmt_node(n) for n in ancestors])
            except Exception:
                ln["account_tree_ancestors"] = ""
    except Exception:
        # اگر هرگونه خطایی در ساخت مسیر حساب رخ دهد، PDF همچنان باید تولید شود
        pass

    # کانتکست قالب (حداقل اطلاعات ضروری برای حسابداری)
    template_context = {
        "business_id": business_id,
        "business_name": business_name,
        "document": doc,
        "lines": doc.get("lines", []),
        "code": doc.get("code"),
        "document_type": doc.get("document_type"),
        "document_date": doc.get("document_date"),
        "document_date_jalali": document_date_jalali,
        "document_date_display": document_date_display,
        "description": doc.get("description"),
        "fiscal_year_title": doc.get("fiscal_year_title"),
        "currency_code": doc.get("currency_code"),
        "currency_symbol": doc.get("currency_symbol"),
        "created_by_name": doc.get("created_by_name"),
        "created_by_user_id": doc.get("created_by_user_id"),
        "registered_at": doc.get("registered_at"),
        "registered_at_display": registered_at_display,
        "project_name": doc.get("project_name"),
        "total_debit": doc.get("total_debit"),
        "total_credit": doc.get("total_credit"),
        "lines_count": doc.get("lines_count"),
        "generated_at": generated_at_display,
        "generated_at_display": generated_at_display,
        "is_fa": is_fa,
        "calendar_type": calendar_type,
        "fa_font_url_regular": fa_font_url_regular,
        "fa_font_url_bold": fa_font_url_bold,
        "assets": {
            "images": {
                "logo": business_logo or "",
            }
        },
        "prepared_by_signature_data_uri": prepared_by_signature_data_uri,
    }

    # تلاش برای رندر با قالب سفارشی
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if template_id is not None:
                explicit_template_id = int(template_id)
            # همچنین می‌توان از query parameter استفاده کرد
            elif request.query_params.get("template_id"):
                explicit_template_id = int(request.query_params.get("template_id"))
        except (ValueError, TypeError):
            explicit_template_id = None
        
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="documents",
            subtype="detail",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception as e:
        # اگر خطای قالب باشد، لاگ می‌کنیم اما به قالب پیش‌فرض می‌رویم
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Custom template rendering failed, using default: {e}", exc_info=True)
        resolved_html = None

    # پیش‌فرض: قالب فایل + پارامترها
    try:
        qp = request.query_params
        paper_size = qp.get("paper_size")
        orientation = qp.get("orientation")
        disposition = qp.get("disposition") or "attachment"
    except Exception:
        paper_size = None
        orientation = None
        disposition = "attachment"
    if not orientation:
        # برای سند حسابداری (جدول عریض) پیش‌فرض landscape مناسب‌تر است
        orientation = "landscape"
    html_content = resolved_html or render_template(
        "pdf/documents/detail.html",
        {
            **template_context,
            "title_text": doc.get("document_type_name") or ("سند" if is_fa else "Document"),
            "paper_size": paper_size,
            "orientation": orientation,
            "footer_text": (
                f"{'سند' if is_fa else 'Document'}: {doc.get('code') or '-'}"
                + (f" • {business_name}" if business_name else "")
            ),
        },
    )

    # تولید PDF با پیکربندی فونت
    def get_font_config():
        """پیکربندی فونت برای PDF فارسی"""
        return FontConfiguration()
    
    try:
        pdf_bytes = HTML(string=html_content).write_pdf(font_config=get_font_config())
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"PDF generation failed: {e}", exc_info=True)
        raise ApiError("PDF_GENERATION_ERROR", "خطا در تولید فایل PDF", http_status=500)

    def _slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", (text or "")).strip("_") or "document"
    filename = f"document_{_slugify(doc.get('code'))}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/documents/manual",
    summary="ایجاد سند حسابداری دستی",
    description="ایجاد یک سند حسابداری دستی جدید با سطرهای مورد نظر",
)
@require_business_access("business_id")
async def create_manual_document_endpoint(
    request: Request,
    business_id: int,
    body: CreateManualDocumentRequest,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("accounting_documents", "add")),
):
    """
    ایجاد سند حسابداری دستی
    
    Body:
        - code: کد سند (اختیاری - خودکار تولید می‌شود)
        - document_date: تاریخ سند
        - fiscal_year_id: شناسه سال مالی (اختیاری - اگر نباشد، سال مالی فعال استفاده می‌شود)
        - currency_id: شناسه ارز
        - is_proforma: پیش‌فاکتور یا قطعی
        - description: توضیحات سند
        - lines: سطرهای سند (حداقل 2 سطر)
        - extra_info: اطلاعات اضافی
    
    نکته: اگر fiscal_year_id ارسال نشود، سیستم به ترتیب زیر عمل می‌کند:
        1. از X-Fiscal-Year-ID header می‌خواند
        2. سال مالی فعال (is_last=True) را انتخاب می‌کند
        3. اگر سال مالی فعال نداشت، خطا برمی‌گرداند
    
    اعتبارسنجی‌ها:
        - سند باید متوازن باشد (مجموع بدهکار = مجموع بستانکار)
        - حداقل 2 سطر داشته باشد
        - هر سطر باید یا بدهکار یا بستانکار داشته باشد (نه هر دو صفر)
    """
    # الزام: ایجاد سند دستی فقط در سال مالی جاری مجاز است.
    # سال مالی جاری = FiscalYear.is_last == True
    from adapters.db.models.fiscal_year import FiscalYear

    active_fy = db.query(FiscalYear).filter(
        FiscalYear.business_id == business_id,
        FiscalYear.is_last == True  # noqa: E712
    ).first()

    if not active_fy:
        raise ApiError(
            "FISCAL_YEAR_REQUIRED",
            "سال مالی جاری برای این کسب‌وکار یافت نشد. ابتدا سال مالی ایجاد/فعال کنید.",
            http_status=400,
        )

    # اگر کاربر تلاش کند سال مالی دیگری را از طریق body/header تحمیل کند، خطا بده.
    requested_fy = body.fiscal_year_id
    try:
        fy_header = request.headers.get("X-Fiscal-Year-ID")
        header_fy = int(fy_header) if fy_header else None
    except Exception:
        header_fy = None

    if requested_fy is not None and int(requested_fy) != int(active_fy.id):
        raise ApiError(
            "FISCAL_YEAR_NOT_CURRENT",
            "ثبت سند فقط در سال مالی جاری مجاز است.",
            http_status=400,
        )
    if header_fy is not None and int(header_fy) != int(active_fy.id):
        raise ApiError(
            "FISCAL_YEAR_NOT_CURRENT",
            "ثبت سند فقط در سال مالی جاری مجاز است.",
            http_status=400,
        )

    fiscal_year_id = active_fy.id
    
    # تبدیل Pydantic model به dict
    data = body.model_dump()
    data["lines"] = [line.model_dump() for line in body.lines]
    
    # ایجاد سند
    result = create_manual_document(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        user_id=ctx.get_user_id(),
        data=data,
    )
    
    return success_response(
        data=format_datetime_fields(result, request),
        request=request,
        message="MANUAL_DOCUMENT_CREATED"
    )


@router.put(
    "/documents/{document_id}",
    summary="ویرایش سند حسابداری دستی",
    description="ویرایش یک سند حسابداری دستی (فقط اسناد manual قابل ویرایش هستند)",
)
async def update_manual_document_endpoint(
    request: Request,
    document_id: int,
    body: UpdateManualDocumentRequest,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("accounting_documents", "edit", Document, "document_id")),
):
    """
    ویرایش سند حسابداری دستی
    
    Body:
        - code: کد سند
        - document_date: تاریخ سند
        - currency_id: شناسه ارز
        - is_proforma: پیش‌فاکتور یا قطعی
        - description: توضیحات سند
        - lines: سطرهای سند (اختیاری - اگر ارسال شود جایگزین سطرهای قبلی می‌شود)
        - extra_info: اطلاعات اضافی
    
    توجه:
        - فقط اسناد manual قابل ویرایش هستند
        - اسناد اتوماتیک باید از منبع اصلی ویرایش شوند
    """
    # بررسی دسترسی
    doc = get_document(db, document_id)
    if not doc:
        raise ApiError("DOCUMENT_NOT_FOUND", "Document not found", http_status=404)
    
    business_id = doc.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)

    # الزام: ویرایش سند دستی فقط در سال مالی جاری مجاز است
    from adapters.db.models.fiscal_year import FiscalYear
    active_fy = db.query(FiscalYear).filter(
        FiscalYear.business_id == business_id,
        FiscalYear.is_last == True  # noqa: E712
    ).first()
    if not active_fy:
        raise ApiError(
            "FISCAL_YEAR_REQUIRED",
            "سال مالی جاری برای این کسب‌وکار یافت نشد.",
            http_status=400,
        )
    doc_fy_id = doc.get("fiscal_year_id")
    if doc_fy_id is not None and int(doc_fy_id) != int(active_fy.id):
        raise ApiError(
            "FISCAL_YEAR_NOT_CURRENT",
            "ویرایش سند فقط در سال مالی جاری مجاز است.",
            http_status=400,
        )
    
    # تبدیل Pydantic model به dict (فقط فیلدهای set شده)
    data = body.model_dump(exclude_unset=True)
    if "lines" in data and data["lines"] is not None:
        data["lines"] = [line.model_dump() for line in body.lines]
    
    # ویرایش سند
    result = update_manual_document(
        db=db,
        document_id=document_id,
        data=data,
    )
    
    return success_response(
        data=format_datetime_fields(result, request),
        request=request,
        message="MANUAL_DOCUMENT_UPDATED"
    )


@router.post(
    "/businesses/{business_id}/reports/daily-sales",
    summary="گزارش فروش روزانه",
    description="گزارش فروش روزانه با گروه‌بندی بر اساس تاریخ",
)
@require_business_access("business_id")
async def daily_sales_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش فروش روزانه"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_daily_sales_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Daily sales report retrieved successfully" if locale != 'fa' else "گزارش فروش روزانه با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/monthly-sales",
    summary="گزارش فروش ماهانه",
    description="گزارش فروش ماهانه با گروه‌بندی بر اساس ماه",
)
@require_business_access("business_id")
async def monthly_sales_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش فروش ماهانه"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_monthly_sales_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    # برای گزارش ماهانه، format_datetime_fields اعمال نمی‌شود
    # چون month_key یک رشته است و نیازی به فرمت کردن تاریخ نیست
    # فقط فیلد date را برای استفاده در frontend نگه می‌داریم
    formatted_items = []
    for item in items:
        formatted_item = dict(item)
        # اگر فیلد date وجود دارد و یک datetime یا date object است، به ISO string تبدیل کن
        if 'date' in formatted_item and formatted_item['date']:
            try:
                from datetime import date as date_class, datetime as dt_class
                if isinstance(formatted_item['date'], dt_class):
                    formatted_item['date'] = formatted_item['date'].date().isoformat()
                elif isinstance(formatted_item['date'], date_class):
                    formatted_item['date'] = formatted_item['date'].isoformat()
            except Exception:
                pass
        formatted_items.append(formatted_item)
    
    result['items'] = formatted_items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Monthly sales report retrieved successfully" if locale != 'fa' else "گزارش فروش ماهانه با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/top-customers",
    summary="گزارش برترین مشتریان",
    description="گزارش برترین مشتریان بر اساس مبلغ فروش",
)
@require_business_access("business_id")
async def top_customers_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش برترین مشتریان"""
    
    # Extract filters
    fiscal_year_id = body.get('fiscal_year_id')
    currency_id = body.get('currency_id')
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    limit = body.get('limit')
    
    # Convert fiscal_year_id
    if fiscal_year_id is not None:
        try:
            fiscal_year_id = int(fiscal_year_id)
        except (ValueError, TypeError):
            fiscal_year_id = None
    
    # Convert currency_id
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # Convert limit
    if limit is not None:
        try:
            limit = int(limit)
            if limit < 1:
                limit = None
        except (ValueError, TypeError):
            limit = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_top_customers_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        limit=limit,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    # برای گزارش برترین مشتریان، format_datetime_fields روی فیلد last_sale_date اعمال می‌شود
    formatted_items = []
    for item in items:
        formatted_item = dict(item)
        # اگر فیلد last_sale_date وجود دارد، آن را فرمت کن
        if 'last_sale_date' in formatted_item and formatted_item['last_sale_date']:
            try:
                from datetime import date as date_class
                if isinstance(formatted_item['last_sale_date'], str):
                    date_obj = date_class.fromisoformat(formatted_item['last_sale_date'])
                    formatted_dict = format_datetime_fields({'date': date_obj}, request)
                    formatted_item['last_sale_date'] = formatted_dict.get('date', formatted_item['last_sale_date'])
            except Exception:
                pass
        formatted_items.append(formatted_item)
    
    result['items'] = formatted_items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Top customers report retrieved successfully" if locale != 'fa' else "گزارش برترین مشتریان با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/top-suppliers",
    summary="گزارش برترین تامین‌کنندگان",
    description="گزارش برترین تامین‌کنندگان بر اساس مبلغ خرید",
)
@require_business_access("business_id")
async def top_suppliers_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش برترین تامین‌کنندگان"""
    
    # Extract filters
    fiscal_year_id = body.get('fiscal_year_id')
    currency_id = body.get('currency_id')
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    limit = body.get('limit')
    
    # Convert fiscal_year_id
    if fiscal_year_id is not None:
        try:
            fiscal_year_id = int(fiscal_year_id)
        except (ValueError, TypeError):
            fiscal_year_id = None
    
    # Convert currency_id
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # Get fiscal year from header if not in body
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if not fiscal_year_id and fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    # Convert limit
    if limit is not None:
        try:
            limit = int(limit)
            if limit < 1:
                limit = None
        except (ValueError, TypeError):
            limit = None
    
    result = get_top_suppliers_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        limit=limit,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    # برای گزارش برترین تامین‌کنندگان، format_datetime_fields روی فیلد last_purchase_date اعمال می‌شود
    formatted_items = []
    for item in items:
        formatted_item = dict(item)
        # اگر فیلد last_purchase_date وجود دارد، آن را فرمت کن
        if 'last_purchase_date' in formatted_item and formatted_item['last_purchase_date']:
            try:
                from datetime import date as date_class
                if isinstance(formatted_item['last_purchase_date'], str):
                    date_obj = date_class.fromisoformat(formatted_item['last_purchase_date'])
                    formatted_dict = format_datetime_fields({'date': date_obj}, request)
                    formatted_item['last_purchase_date'] = formatted_dict.get('date', formatted_item['last_purchase_date'])
            except Exception:
                pass
        formatted_items.append(formatted_item)
    
    result['items'] = formatted_items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Top suppliers report retrieved successfully" if locale != 'fa' else "گزارش برترین تامین‌کنندگان با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/top-suppliers/export/excel",
    summary="خروجی Excel گزارش برترین تامین‌کنندگان",
    description="خروجی Excel گزارش برترین تامین‌کنندگان با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_top_suppliers_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel گزارش برترین تامین‌کنندگان"""
    import io
    import json
    import datetime
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response
    from adapters.db.models.business import Business
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # Extract filters
    fiscal_year_id = body.get('fiscal_year_id')
    currency_id = body.get('currency_id')
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    limit = body.get('limit')
    
    # Convert fiscal_year_id
    if fiscal_year_id is not None:
        try:
            fiscal_year_id = int(fiscal_year_id)
        except (ValueError, TypeError):
            fiscal_year_id = None
    
    # Convert currency_id
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # Get fiscal year from header if not in body
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if not fiscal_year_id and fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    # Convert limit
    if limit is not None:
        try:
            limit = int(limit)
            if limit < 1:
                limit = None
        except (ValueError, TypeError):
            limit = None
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_top_suppliers_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        limit=limit,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('person_code', 'کد تامین‌کننده' if is_fa else 'Supplier Code'),
            ('person_name', 'نام تامین‌کننده' if is_fa else 'Supplier Name'),
            ('invoice_count', 'تعداد فاکتور' if is_fa else 'Invoice Count'),
            ('total_purchases', 'جمع خرید' if is_fa else 'Total Purchases'),
            ('last_purchase_date', 'آخرین تاریخ خرید' if is_fa else 'Last Purchase Date'),
        ]
        for key, label in default_columns:
            keys.append(key)
            headers.append(label)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش برترین تامین‌کنندگان" if is_fa else "Top Suppliers Report"
    
    # RTL handling for Persian
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            
            # Format numbers
            if key in ['invoice_count', 'total_purchases'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    value = num_value
                except (ValueError, TypeError):
                    pass
            
            # Format dates
            if key == 'last_purchase_date' and value:
                value = format_date_for_export(item, 'last_purchase_date')
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # RTL alignment for Persian text and numbers
            if locale == 'fa':
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, str) and any('\u0600' <= c <= '\u06FF' for c in str(value)):
                    cell.alignment = Alignment(horizontal="right")
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "top_suppliers_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/daily-purchases",
    summary="گزارش خرید روزانه",
    description="گزارش خرید روزانه با گروه‌بندی بر اساس تاریخ",
)
@require_business_access("business_id")
async def daily_purchases_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش خرید روزانه"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_daily_purchases_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Daily purchases report retrieved successfully" if locale != 'fa' else "گزارش خرید روزانه با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/daily-purchases/export/excel",
    summary="خروجی Excel گزارش خرید روزانه",
    description="خروجی Excel گزارش خرید روزانه با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_daily_purchases_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel گزارش خرید روزانه"""
    import io
    import json
    import datetime
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response
    from adapters.db.models.business import Business
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_daily_purchases_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('date', 'تاریخ' if is_fa else 'Date'),
            ('invoice_count', 'تعداد فاکتور' if is_fa else 'Invoice Count'),
            ('total_gross', 'جمع کل' if is_fa else 'Total Gross'),
            ('total_discount', 'جمع تخفیف' if is_fa else 'Total Discount'),
            ('total_tax', 'جمع مالیات' if is_fa else 'Total Tax'),
            ('total_net', 'جمع خالص' if is_fa else 'Total Net'),
        ]
        for key, label in default_columns:
            keys.append(key)
            headers.append(label)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش خرید روزانه" if is_fa else "Daily Purchases Report"
    
    # RTL handling for Persian
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            
            # Format numbers
            if key in ['invoice_count', 'total_gross', 'total_discount', 'total_tax', 'total_net'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    value = num_value
                except (ValueError, TypeError):
                    pass
            
            # Format dates
            if key == 'date' and value:
                value = format_date_for_export(item, 'date')
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # RTL alignment for Persian text and numbers
            if locale == 'fa':
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, str) and any('\u0600' <= c <= '\u06FF' for c in str(value)):
                    cell.alignment = Alignment(horizontal="right")
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "daily_purchases_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/daily-purchases/export/pdf",
    summary="خروجی PDF گزارش خرید روزانه",
    description="خروجی PDF گزارش خرید روزانه با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_daily_purchases_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی PDF گزارش خرید روزانه"""
    import datetime
    import json
    import re
    from fastapi.responses import Response
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from adapters.db.models.business import Business
    from adapters.db.models.fiscal_year import FiscalYear
    from app.core.calendar import CalendarConverter
    from app.services.pdf.template_renderer import render_template

    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    # دریافت calendar_type
    calendar_type = "gregorian"
    try:
        if hasattr(request.state, "calendar_type") and request.state.calendar_type:
            calendar_type = request.state.calendar_type
    except Exception:
        calendar_type = "gregorian"

    # دریافت سال مالی از header یا body (fallback: آخرین سال مالی)
    fiscal_year_id = None
    fy_header = request.headers.get("X-Fiscal-Year-ID")
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            fiscal_year_id = None
    if body.get("fiscal_year_id"):
        try:
            fiscal_year_id = int(body["fiscal_year_id"])
        except (ValueError, TypeError):
            fiscal_year_id = fiscal_year_id
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True,  # noqa: E712
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id

    # پارامترها
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    currency_id = body.get("currency_id")
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None

    # داده‌ها (بدون pagination)
    max_export_records = 10000
    result = get_daily_purchases_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        skip=0,
        take=max_export_records,
    )
    items = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # تاریخ نمایشی برای PDF
    def _date_text(item_dict: dict) -> str:
        try:
            formatted_key = "date_formatted"
            if formatted_key in item_dict and isinstance(item_dict.get(formatted_key), dict):
                d = item_dict.get(formatted_key) or {}
                return str(d.get("date_only") or d.get("formatted") or "").split(" ")[0].split("T")[0]
        except Exception:
            pass
        value = item_dict.get("date")
        if value is None:
            return ""
        try:
            # CalendarConverter.format_datetime expects datetime
            if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
                value = datetime.datetime.combine(value, datetime.datetime.min.time())
            if isinstance(value, str):
                s = value.strip()
                if not s:
                    return ""
                if "T" in s:
                    value = datetime.datetime.fromisoformat(s.replace("Z", "+00:00"))
                else:
                    d = datetime.date.fromisoformat(s)
                    value = datetime.datetime.combine(d, datetime.datetime.min.time())
            if isinstance(value, datetime.datetime):
                out = CalendarConverter.format_datetime(value, calendar_type)
                return str(out.get("date_only") or out.get("formatted") or "").split(" ")[0]
        except Exception:
            pass
        return str(value).split(" ")[0].split("T")[0]

    for it in items:
        try:
            it["date_text"] = _date_text(it)
        except Exception:
            pass

    # انتخاب سطرها
    selected_only = bool(body.get("selected_only", False))
    selected_indices = body.get("selected_indices")
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # ستون‌ها
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    headers: list[str] = []
    keys: list[str] = []
    export_columns = body.get("export_columns")
    if export_columns:
        for col in export_columns:
            key = col.get("key")
            label = col.get("label", key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        default_columns = [
            ("date", "تاریخ" if is_fa else "Date"),
            ("invoice_count", "تعداد فاکتور" if is_fa else "Invoice Count"),
            ("total_gross", "جمع کل" if is_fa else "Total Gross"),
            ("total_discount", "جمع تخفیف" if is_fa else "Total Discount"),
            ("total_tax", "جمع مالیات" if is_fa else "Total Tax"),
            ("total_net", "جمع خالص" if is_fa else "Total Net"),
        ]
        for k, label in default_columns:
            keys.append(k)
            headers.append(label)

    numeric_keys = {"invoice_count", "total_gross", "total_discount", "total_tax", "total_net"}

    # اطلاعات کسب‌وکار / سال مالی
    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    fiscal_year_name = ""
    try:
        if fiscal_year_id:
            fy = db.query(FiscalYear).filter(
                and_(
                    FiscalYear.id == fiscal_year_id,
                    FiscalYear.business_id == business_id,
                )
            ).first()
            if fy is not None:
                fiscal_year_name = getattr(fy, "name", "") or ""
    except Exception:
        fiscal_year_name = ""

    # فونت فارسی (data-uri)
    fa_font_url_regular = None
    fa_font_url_bold = None
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_font_url_regular, fa_font_url_bold = load_farsi_font_data_uris()
    except Exception:
        fa_font_url_regular = None
        fa_font_url_bold = None

    # disposition / paper settings
    disposition = "attachment"
    try:
        disposition = str(body.get("disposition") or "attachment")
    except Exception:
        disposition = "attachment"
    paper_size = None
    orientation = None
    try:
        paper_size = body.get("paper_size")
        orientation = body.get("orientation")
    except Exception:
        paper_size = None
        orientation = None

    generated_at = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
    if is_fa:
        try:
            out = CalendarConverter.format_datetime(datetime.datetime.now(), calendar_type)
            generated_at = str(out.get("formatted") or generated_at)
        except Exception:
            pass

    title_text = "گزارش خرید روزانه" if is_fa else "Daily Purchases Report"
    footer_text = ""
    final_html = render_template(
        "pdf/daily_purchases/list.html",
        {
            "is_fa": is_fa,
            "title_text": title_text,
            "footer_text": footer_text,
            "paper_size": paper_size,
            "orientation": orientation,
            "fa_font_url_regular": fa_font_url_regular,
            "fa_font_url_bold": fa_font_url_bold,
            "business_name": business_name,
            "fiscal_year_name": fiscal_year_name,
            "generated_at": generated_at,
            "headers": headers,
            "keys": keys,
            "numeric_keys": numeric_keys,
            "items": items,
            "date_from": date_from,
            "date_to": date_to,
        },
    )

    pdf_bytes = HTML(string=final_html).write_pdf(font_config=FontConfiguration())

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "daily_purchases_report"
    if business_name:
        base += f"_{slugify(business_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/materials-consumption",
    summary="گزارش مصرف مواد",
    description="گزارش مصرف مواد از فاکتورهای تولید",
)
@require_business_access("business_id")
async def materials_consumption_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش مصرف مواد از فاکتورهای تولید"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    product_id = body.get('product_id')
    warehouse_id = body.get('warehouse_id')
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    if product_id is not None:
        try:
            product_id = int(product_id)
        except (ValueError, TypeError):
            product_id = None
    
    if warehouse_id is not None:
        try:
            warehouse_id = int(warehouse_id)
        except (ValueError, TypeError):
            warehouse_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_materials_consumption_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        product_id=product_id,
        warehouse_id=warehouse_id,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Materials consumption report retrieved successfully" if locale != 'fa' else "گزارش مصرف مواد با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/production",
    summary="گزارش تولید",
    description="گزارش تولید (کالاهای ساخته شده) از فاکتورهای تولید",
)
@require_business_access("business_id")
async def production_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش تولید (کالاهای ساخته شده) از فاکتورهای تولید"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    product_id = body.get('product_id')
    warehouse_id = body.get('warehouse_id')
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    if product_id is not None:
        try:
            product_id = int(product_id)
        except (ValueError, TypeError):
            product_id = None
    
    if warehouse_id is not None:
        try:
            warehouse_id = int(warehouse_id)
        except (ValueError, TypeError):
            warehouse_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_production_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        product_id=product_id,
        warehouse_id=warehouse_id,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Production report retrieved successfully" if locale != 'fa' else "گزارش تولید با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/trial-balance",
    summary="گزارش تراز آزمایشی",
    description="گزارش تراز آزمایشی برای حساب‌های حسابداری",
)
@require_business_access("business_id")
async def trial_balance_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش تراز آزمایشی برای حساب‌های حسابداری"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    account_type = body.get('account_type')
    account_ids = body.get('account_ids')
    include_zero_balance = body.get('include_zero_balance', False)
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    if account_ids is not None and not isinstance(account_ids, list):
        account_ids = None
    
    if account_ids:
        try:
            account_ids = [int(aid) for aid in account_ids if aid is not None]
        except (ValueError, TypeError):
            account_ids = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_trial_balance_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        account_type=account_type,
        account_ids=account_ids,
        include_zero_balance=include_zero_balance,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Trial balance report retrieved successfully" if locale != 'fa' else "گزارش تراز آزمایشی با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/trial-balance/export/excel",
    summary="خروجی Excel گزارش تراز آزمایشی",
    description="خروجی Excel گزارش تراز آزمایشی با قابلیت انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_trial_balance_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی Excel گزارش تراز آزمایشی"""
    import io
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from adapters.db.models.business import Business

    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass

    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass

    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    account_type = body.get('account_type')
    account_ids = body.get('account_ids')
    include_zero_balance = bool(body.get('include_zero_balance', False))
    project_id = body.get('project_id')

    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None

    if project_id is not None:
        try:
            project_id = int(project_id)
        except (ValueError, TypeError):
            project_id = None

    if account_ids is not None and not isinstance(account_ids, list):
        account_ids = None
    if account_ids:
        try:
            account_ids = [int(aid) for aid in account_ids if aid is not None]
        except (ValueError, TypeError):
            account_ids = None

    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_trial_balance_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        account_type=account_type,
        account_ids=account_ids,
        project_id=project_id,
        include_zero_balance=include_zero_balance,
        skip=0,
        take=max_export_records,
    )

    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare headers based on export_columns
    headers: list[str] = []
    keys: list[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get("key")
            label = col.get("label", key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        default_columns = [
            ("account_code", "کد حساب"),
            ("account_name", "نام حساب"),
            ("account_type", "نوع حساب"),
            ("opening_debit", "مانده ابتدای دوره (بدهکار)"),
            ("opening_credit", "مانده ابتدای دوره (بستانکار)"),
            ("period_debit", "جمع بدهکار دوره"),
            ("period_credit", "جمع بستانکار دوره"),
            ("closing_debit", "مانده انتهای دوره (بدهکار)"),
            ("closing_credit", "مانده انتهای دوره (بستانکار)"),
        ]
        if items:
            for key, label in default_columns:
                if key in items[0]:
                    keys.append(key)
                    headers.append(label)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    
    def map_account_type_label(raw_value: object) -> str:
        """
        Map account_type codes to user-friendly labels (fa/en) for exports.
        Supports both legacy string types and seeded numeric codes stored as strings.
        """
        if raw_value is None:
            return ""
        key = str(raw_value).strip()
        if not key:
            return ""
        
        mapping_fa = {
            # string types (common in services/migrations)
            "accounting_document": "حسابداری",
            "bank": "بانک",
            "cash_register": "صندوق",
            "cashdesk": "صندوق",
            "cash": "نقد",
            "petty_cash": "تنخواه",
            "check": "چک",
            "person": "شخص",
            "product": "کالا",
            # numeric types (seeded chart-of-accounts)
            "0": "سایر",
            "1": "صندوق",
            "2": "تنخواه",
            "3": "بانک",
            "4": "حساب‌های دریافتنی",
            "5": "اسناد دریافتنی",
            "6": "اسناد در جریان وصول",
            "7": "موجودی کالا",
            "8": "مالیات ارزش افزوده خرید",
            "9": "حساب‌های پرداختنی",
            "10": "اسناد پرداختنی",
            "11": "مالیات ارزش افزوده فروش",
            "12": "مالیات بر درآمد پرداختنی",
            "13": "سرمایه اولیه",
            "14": "افزایش/کاهش سرمایه",
            "15": "اندوخته قانونی",
            "16": "برداشت‌ها",
            "17": "سهم سود و زیان",
            "18": "سود و زیان انباشته",
            "19": "بهای تمام‌شده کالای فروش‌رفته",
            "20": "برگشت از خرید",
            "21": "تخفیفات نقدی خرید",
            "22": "فروش کالا",
            "23": "برگشت از فروش",
            "24": "تخفیفات نقدی فروش",
            "25": "درآمد فروش خدمات",
            "26": "برگشت از خرید خدمات",
            "27": "درآمد اضافه کالا",
            "28": "درآمد حمل کالا",
            "29": "برگشت از فروش خدمات",
            "30": "خرید خدمات",
            "31": "هزینه حمل کالا",
            "32": "هزینه کسری و ضایعات کالا",
            "33": "کارمزد خدمات بانکی",
            "34": "کنترل کسری/اضافه کالا",
            "35": "خلاصه سود و زیان",
            "36": "درآمد تسعیر ارز",
            "37": "هزینه تسعیر ارز",
            "38": "سود فروش اقساطی",
            "39": "سود تحقق‌نیافته فروش اقساطی",
            "40": "ذخیره مالیات بر درآمد پرداختنی",
            "41": "موجودی کالای در جریان ساخت",
            "42": "حقوق و دستمزد پرداختنی",
            "43": "سربار تولید پرداختنی",
        }
        mapping_en = {
            # string types
            "accounting_document": "Accounting",
            "bank": "Bank",
            "cash_register": "Cash Register",
            "cashdesk": "Cash Register",
            "cash": "Cash",
            "petty_cash": "Petty Cash",
            "check": "Check",
            "person": "Person",
            "product": "Product",
            # numeric types
            "0": "Other",
            "1": "Cash Register",
            "2": "Petty Cash",
            "3": "Bank",
            "4": "Accounts Receivable",
            "5": "Notes Receivable",
            "6": "Notes in Collection",
            "7": "Inventory",
            "8": "Input VAT",
            "9": "Accounts Payable",
            "10": "Notes Payable",
            "11": "Output VAT",
            "12": "Income Tax Payable",
            "13": "Initial Capital",
            "14": "Capital Increase/Decrease",
            "15": "Legal Reserve",
            "16": "Drawings",
            "17": "Profit/Loss Share",
            "18": "Retained Earnings",
            "19": "Cost of Goods Sold",
            "20": "Purchase Returns",
            "21": "Purchase Discounts",
            "22": "Goods Sales",
            "23": "Sales Returns",
            "24": "Sales Discounts",
            "25": "Service Revenue",
            "26": "Service Purchase Returns",
            "27": "Inventory Surplus Income",
            "28": "Freight Income",
            "29": "Service Sales Returns",
            "30": "Service Purchases",
            "31": "Freight Expense",
            "32": "Inventory Shortage/Waste",
            "33": "Bank Fees",
            "34": "Inventory Variance Control",
            "35": "Income Summary",
            "36": "FX Gain",
            "37": "FX Loss",
            "38": "Installment Sales Profit",
            "39": "Unearned Installment Profit",
            "40": "Provision for Income Tax",
            "41": "Work in Process Inventory",
            "42": "Salaries Payable",
            "43": "Manufacturing Overhead Payable",
        }
        return (mapping_fa if is_fa else mapping_en).get(key, key)
    
    numeric_keys = {
        "opening_debit",
        "opening_credit",
        "period_debit",
        "period_credit",
        "closing_debit",
        "closing_credit",
    }
    
    def format_number_for_export(raw_value: object) -> object:
        """Hide useless .0 for integer-like values (keep real decimals)."""
        if raw_value is None:
            return ""
        if isinstance(raw_value, bool):
            return raw_value
        if isinstance(raw_value, int):
            return raw_value
        if isinstance(raw_value, float):
            try:
                if raw_value.is_integer():
                    return int(raw_value)
            except Exception:
                pass
            return raw_value
        if isinstance(raw_value, str):
            s = raw_value.strip()
            if not s:
                return ""
            # Avoid messing with non-numeric strings
            try:
                f = float(s)
                if f.is_integer():
                    return int(f)
            except Exception:
                return raw_value
            return raw_value
        return raw_value

    # Header row
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for item in items:
        row = []
        for key in keys:
            value = item.get(key, "")
            if key == "account_type":
                value = map_account_type_label(value)
            if key in numeric_keys:
                value = format_number_for_export(value)
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)
        ws.append(row)

    # Basic formatting: borders + alignment
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            # پیش‌فرض: وسط‌چین
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "trial_balance_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/trial-balance/export/pdf",
    summary="خروجی PDF گزارش تراز آزمایشی",
    description="خروجی PDF گزارش تراز آزمایشی با قابلیت انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_trial_balance_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی PDF گزارش تراز آزمایشی"""
    import json
    import datetime
    import re
    import logging
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    from adapters.db.models.business import Business
    from adapters.db.models.fiscal_year import FiscalYear
    from adapters.db.models.currency import Currency

    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    logger = logging.getLogger(__name__)
    debug_font = str(request.headers.get("X-Debug-Pdf-Font", "") or "").strip().lower() in ("1", "true", "yes", "y", "on")

    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass

    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass

    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    account_type = body.get('account_type')
    account_ids = body.get('account_ids')
    include_zero_balance = bool(body.get('include_zero_balance', False))
    project_id = body.get('project_id')

    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None

    if project_id is not None:
        try:
            project_id = int(project_id)
        except (ValueError, TypeError):
            project_id = None

    if account_ids is not None and not isinstance(account_ids, list):
        account_ids = None
    if account_ids:
        try:
            account_ids = [int(aid) for aid in account_ids if aid is not None]
        except (ValueError, TypeError):
            account_ids = None

    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_trial_balance_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        account_type=account_type,
        account_ids=account_ids,
        project_id=project_id,
        include_zero_balance=include_zero_balance,
        skip=0,
        take=max_export_records,
    )

    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare headers based on export_columns
    headers: list[str] = []
    keys: list[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get("key")
            label = col.get("label", key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        default_columns = [
            ("account_code", "کد حساب"),
            ("account_name", "نام حساب"),
            ("account_type", "نوع حساب"),
            ("opening_debit", "مانده ابتدای دوره (بدهکار)"),
            ("opening_credit", "مانده ابتدای دوره (بستانکار)"),
            ("period_debit", "جمع بدهکار دوره"),
            ("period_credit", "جمع بستانکار دوره"),
            ("closing_debit", "مانده انتهای دوره (بدهکار)"),
            ("closing_credit", "مانده انتهای دوره (بستانکار)"),
        ]
        if items:
            for key, label in default_columns:
                if key in items[0]:
                    keys.append(key)
                    headers.append(label)

    # Business name
    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    html_lang = "fa" if is_fa else "en"
    html_dir = "rtl" if is_fa else "ltr"

    # Embed Farsi fonts (shared helper)
    fa_font_url_regular = None
    fa_font_url_bold = None
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_font_url_regular, fa_font_url_bold = load_farsi_font_data_uris()
    except Exception:
        fa_font_url_regular = None
        fa_font_url_bold = None

    if debug_font:
        logger.info(
            "trial_balance_export_pdf: locale=%s is_fa=%s accept_language=%s",
            locale,
            is_fa,
            request.headers.get("Accept-Language"),
        )
        logger.info(
            "trial_balance_export_pdf: font_uris present regular=%s bold=%s (len=%s/%s)",
            bool(fa_font_url_regular),
            bool(fa_font_url_bold),
            len(fa_font_url_regular or ""),
            len(fa_font_url_bold or ""),
        )
    
    def map_account_type_label(raw_value: object) -> str:
        """Map account_type codes to user-friendly labels (fa/en) for exports."""
        if raw_value is None:
            return ""
        key = str(raw_value).strip()
        if not key:
            return ""
        mapping_fa = {
            "accounting_document": "حسابداری",
            "bank": "بانک",
            "cash_register": "صندوق",
            "cashdesk": "صندوق",
            "cash": "نقد",
            "petty_cash": "تنخواه",
            "check": "چک",
            "person": "شخص",
            "product": "کالا",
            "0": "سایر",
            "1": "صندوق",
            "2": "تنخواه",
            "3": "بانک",
            "4": "حساب‌های دریافتنی",
            "5": "اسناد دریافتنی",
            "6": "اسناد در جریان وصول",
            "7": "موجودی کالا",
            "8": "مالیات ارزش افزوده خرید",
            "9": "حساب‌های پرداختنی",
            "10": "اسناد پرداختنی",
            "11": "مالیات ارزش افزوده فروش",
            "12": "مالیات بر درآمد پرداختنی",
            "13": "سرمایه اولیه",
            "14": "افزایش/کاهش سرمایه",
            "15": "اندوخته قانونی",
            "16": "برداشت‌ها",
            "17": "سهم سود و زیان",
            "18": "سود و زیان انباشته",
            "19": "بهای تمام‌شده کالای فروش‌رفته",
            "20": "برگشت از خرید",
            "21": "تخفیفات نقدی خرید",
            "22": "فروش کالا",
            "23": "برگشت از فروش",
            "24": "تخفیفات نقدی فروش",
            "25": "درآمد فروش خدمات",
            "26": "برگشت از خرید خدمات",
            "27": "درآمد اضافه کالا",
            "28": "درآمد حمل کالا",
            "29": "برگشت از فروش خدمات",
            "30": "خرید خدمات",
            "31": "هزینه حمل کالا",
            "32": "هزینه کسری و ضایعات کالا",
            "33": "کارمزد خدمات بانکی",
            "34": "کنترل کسری/اضافه کالا",
            "35": "خلاصه سود و زیان",
            "36": "درآمد تسعیر ارز",
            "37": "هزینه تسعیر ارز",
            "38": "سود فروش اقساطی",
            "39": "سود تحقق‌نیافته فروش اقساطی",
            "40": "ذخیره مالیات بر درآمد پرداختنی",
            "41": "موجودی کالای در جریان ساخت",
            "42": "حقوق و دستمزد پرداختنی",
            "43": "سربار تولید پرداختنی",
        }
        mapping_en = {
            "accounting_document": "Accounting",
            "bank": "Bank",
            "cash_register": "Cash Register",
            "cashdesk": "Cash Register",
            "cash": "Cash",
            "petty_cash": "Petty Cash",
            "check": "Check",
            "person": "Person",
            "product": "Product",
            "0": "Other",
            "1": "Cash Register",
            "2": "Petty Cash",
            "3": "Bank",
            "4": "Accounts Receivable",
            "5": "Notes Receivable",
            "6": "Notes in Collection",
            "7": "Inventory",
            "8": "Input VAT",
            "9": "Accounts Payable",
            "10": "Notes Payable",
            "11": "Output VAT",
            "12": "Income Tax Payable",
            "13": "Initial Capital",
            "14": "Capital Increase/Decrease",
            "15": "Legal Reserve",
            "16": "Drawings",
            "17": "Profit/Loss Share",
            "18": "Retained Earnings",
            "19": "Cost of Goods Sold",
            "20": "Purchase Returns",
            "21": "Purchase Discounts",
            "22": "Goods Sales",
            "23": "Sales Returns",
            "24": "Sales Discounts",
            "25": "Service Revenue",
            "26": "Service Purchase Returns",
            "27": "Inventory Surplus Income",
            "28": "Freight Income",
            "29": "Service Sales Returns",
            "30": "Service Purchases",
            "31": "Freight Expense",
            "32": "Inventory Shortage/Waste",
            "33": "Bank Fees",
            "34": "Inventory Variance Control",
            "35": "Income Summary",
            "36": "FX Gain",
            "37": "FX Loss",
            "38": "Installment Sales Profit",
            "39": "Unearned Installment Profit",
            "40": "Provision for Income Tax",
            "41": "Work in Process Inventory",
            "42": "Salaries Payable",
            "43": "Manufacturing Overhead Payable",
        }
        return (mapping_fa if is_fa else mapping_en).get(key, key)
    
    numeric_keys = {
        "opening_debit",
        "opening_credit",
        "period_debit",
        "period_credit",
        "closing_debit",
        "closing_credit",
    }
    
    def format_number_for_export(raw_value: object) -> str:
        """Hide useless .0 and add thousand separators for readability."""
        if raw_value is None:
            return ""
        if isinstance(raw_value, bool):
            return "1" if raw_value else "0"
        
        def _apply_grouping(num_str: str) -> str:
            # group integer part, keep fractional part as-is
            s = (num_str or "").strip()
            if not s:
                return ""
            sign = ""
            if s.startswith("-"):
                sign = "-"
                s = s[1:]
            if s.startswith("+"):
                s = s[1:]
            if "." in s:
                int_part, frac_part = s.split(".", 1)
            else:
                int_part, frac_part = s, ""
            try:
                grouped = f"{int(int_part):,}" if int_part else "0"
            except Exception:
                # fallback: don't modify if int parsing fails
                grouped = int_part
            if frac_part:
                out = f"{sign}{grouped}.{frac_part}"
            else:
                out = f"{sign}{grouped}"
            # Persian thousands separator if locale is fa
            if is_fa:
                out = out.replace(",", "٬")
            return out
        
        if isinstance(raw_value, int):
            return _apply_grouping(str(raw_value))
        if isinstance(raw_value, float):
            try:
                if raw_value.is_integer():
                    return _apply_grouping(str(int(raw_value)))
            except Exception:
                pass
            # Keep decimals but add grouping to integer part
            return _apply_grouping(str(raw_value))
        if isinstance(raw_value, str):
            s = raw_value.strip()
            if not s:
                return ""
            try:
                f = float(s)
                if f.is_integer():
                    return _apply_grouping(str(int(f)))
            except Exception:
                return s
            return _apply_grouping(s)
        return str(raw_value)

    def format_date_for_report(value: object, calendar: str) -> str:
        """Format a date/datetime/ISO string to selected calendar (date only)."""
        if value is None:
            return ""
        import datetime as _dt
        dt_obj: _dt.datetime | None = None
        if isinstance(value, _dt.datetime):
            dt_obj = value
        elif isinstance(value, _dt.date):
            dt_obj = _dt.datetime.combine(value, _dt.datetime.min.time())
        elif isinstance(value, str):
            s = value.strip()
            if not s:
                return ""
            try:
                if "T" in s:
                    dt_obj = _dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
                else:
                    d = _dt.date.fromisoformat(s.split(" ")[0].split("T")[0])
                    dt_obj = _dt.datetime.combine(d, _dt.datetime.min.time())
            except Exception:
                return s
        else:
            return str(value)
        try:
            from app.core.calendar import CalendarConverter
            out = CalendarConverter.format_datetime(dt_obj, calendar)
            return out.get("date_only") or (out.get("formatted", "").split(" ")[0]) or dt_obj.strftime("%Y/%m/%d")
        except Exception:
            return dt_obj.strftime("%Y/%m/%d")

    def esc(v: object) -> str:
        return escape("" if v is None else str(v))

    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key, "")
            if key == "account_type":
                value = map_account_type_label(value)
            if key in numeric_keys:
                value = format_number_for_export(value)
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = json.dumps(value, ensure_ascii=False)
            tds.append(f"<td>{esc(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = ''.join(f"<th>{esc(h)}</th>" for h in headers)

    # Format report datetime based on selected calendar (X-Calendar-Type or middleware state)
    calendar_type = None
    try:
        if hasattr(request.state, "calendar_type") and request.state.calendar_type:
            calendar_type = request.state.calendar_type
    except Exception:
        calendar_type = None
    
    if not calendar_type:
        cal_header = (request.headers.get("X-Calendar-Type", "jalali") or "jalali").lower()
        calendar_type = "jalali" if cal_header in ["jalali", "persian", "shamsi"] else "gregorian"
    
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(), calendar_type)
        now = formatted_now.get("formatted") or formatted_now.get("date_time") or datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    except Exception:
        now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')

    # Build filters summary (show selected filters in PDF)
    filters: list[tuple[str, str]] = []

    # Fiscal year + (optional) date range
    fy_obj = None
    if fiscal_year_id:
        try:
            fy_obj = db.query(FiscalYear).filter(
                FiscalYear.id == int(fiscal_year_id),
                FiscalYear.business_id == int(business_id),
            ).first()
        except Exception:
            fy_obj = None
        fy_title = (fy_obj.title if fy_obj else str(fiscal_year_id))
        filters.append(("سال مالی" if is_fa else "Fiscal Year", str(fy_title)))

    # Date range: if user provided date_from/to show them; otherwise show FY period when available
    if date_from or date_to:
        if date_from:
            filters.append(("از تاریخ" if is_fa else "From", format_date_for_report(date_from, calendar_type)))
        if date_to:
            filters.append(("تا تاریخ" if is_fa else "To", format_date_for_report(date_to, calendar_type)))
    elif fy_obj is not None:
        try:
            dr = f"{format_date_for_report(fy_obj.start_date, calendar_type)} — {format_date_for_report(fy_obj.end_date, calendar_type)}"
            filters.append(("بازه" if is_fa else "Period", dr))
        except Exception:
            pass

    # Currency
    if currency_id:
        cur_label = ""
        try:
            cur = db.query(Currency).filter(Currency.id == int(currency_id)).first()
            if cur is not None:
                if is_fa:
                    cur_label = (cur.title or cur.name or "")
                else:
                    cur_label = (cur.code or cur.name or cur.title or "")
                if cur.symbol:
                    cur_label = f"{cur_label} ({cur.symbol})" if cur_label else cur.symbol
        except Exception:
            cur_label = ""
        filters.append(("ارز" if is_fa else "Currency", cur_label or str(currency_id)))

    # Account type
    if account_type:
        filters.append(("نوع حساب" if is_fa else "Account Type", map_account_type_label(account_type)))

    # Include zero balance (only if present in request or enabled)
    if ("include_zero_balance" in body) or include_zero_balance:
        filters.append((
            "شامل مانده صفر" if is_fa else "Include Zero Balance",
            ("بله" if include_zero_balance else "خیر") if is_fa else ("Yes" if include_zero_balance else "No"),
        ))

    # Project
    if project_id:
        filters.append(("پروژه" if is_fa else "Project", str(project_id)))

    filters_html = ""
    if filters:
        parts = []
        for k, v in filters:
            if v:
                parts.append(f"<span class='filter-item'><strong>{esc(k)}:</strong> {esc(v)}</span>")
        if parts:
            filters_html = f"<div class='filters'>{''.join(parts)}</div>"

    title_text = "گزارش تراز آزمایشی" if is_fa else "Trial Balance Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "

    # تلاش برای رندر با قالب سفارشی (trial_balance/reports/trial-balance)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": title_text,
            "business_name": business_name,
            "generated_at": now,
            "is_fa": is_fa,
            "calendar_type": calendar_type,
            "filters": [{"label": k, "value": v} for k, v in filters],
            "fa_font_url_regular": fa_font_url_regular,
            "fa_font_url_bold": fa_font_url_bold,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="trial_balance",
            subtype="reports/trial-balance",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    table_html = f"""
    <html lang="{html_lang}" dir="{html_dir}">
      <head>
        <meta charset="utf-8">
      </head>
      <body>
        <div class="header">
          <div class="title">{esc(title_text)}</div>
          <div class="meta">
            <div><strong>{esc(label_biz)}:</strong> {esc(business_name)}</div>
            <div><strong>{esc(label_date)}:</strong> {esc(now)}</div>
          </div>
        </div>
        {filters_html}
        <div class="table-wrapper">
          <table class="report-table">
            <thead><tr>{headers_html}</tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
          </table>
        </div>
      </body>
    </html>
    """

    html_content = resolved_html or table_html

    # Build font-face CSS BEFORE injecting into HTML
    font_face_css = ""
    if fa_font_url_regular:
        font_face_css += f"""
        @font-face {{
          font-family: 'YekanBakhFaNum';
          src: url("{fa_font_url_regular}") format('truetype');
          font-weight: 400;
          font-style: normal;
        }}
        """
    if fa_font_url_bold:
        font_face_css += f"""
        @font-face {{
          font-family: 'YekanBakhFaNum';
          src: url("{fa_font_url_bold}") format('truetype');
          font-weight: 700;
          font-style: normal;
        }}
        """

    # Important: some custom templates include their own CSS (sometimes with !important)
    # To match invoices behavior and guarantee font is applied, inject font CSS at the end of <head>.
    try:
        # If locale isn't fa but document is RTL (custom templates), still prefer Persian font stack.
        force_fa = bool(is_fa) or ('dir="rtl"' in html_content) or ("dir='rtl'" in html_content)
        if force_fa:
            preferred_stack = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif"
        else:
            preferred_stack = "Arial, sans-serif"
        injected_style = (
            "<style id=\"hesabix-font-inject\">"
            + (font_face_css or "")
            + f"\nhtml, body, body * {{ font-family: {preferred_stack} !important; }}\n"
            + "</style>"
        )
        if "</head>" in html_content:
            html_content = html_content.replace("</head>", injected_style + "</head>")
        else:
            html_content = injected_style + html_content
    except Exception:
        # Don't fail PDF generation if injection fails
        pass

    if debug_font:
        logger.info(
            "trial_balance_export_pdf: resolved_html=%s head_tag=%s injected_style=%s rtl_in_html=%s font_face_css_len=%s",
            bool(resolved_html),
            ("</head>" in html_content),
            ("hesabix-font-inject" in html_content),
            ('dir="rtl"' in html_content) or ("dir='rtl'" in html_content),
            len(font_face_css or ""),
        )

    # Keep stylesheet as well (layout + additional rules)
    force_font_css = ""
    if is_fa:
        force_font_css = "\nhtml, body, body * { font-family: YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif !important; }\n"
    else:
        force_font_css = "\nhtml, body, body * { font-family: Arial, sans-serif !important; }\n"

    css = CSS(string=(font_face_css or "") + force_font_css + f"""
      @page {{
        size: A4 landscape;
        margin: 12mm;
        @bottom-{'left' if is_fa else 'right'} {{
          content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
          font-size: 10px;
          color: #666;
          font-family: {'YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif' if is_fa else 'Arial, sans-serif'} !important;
        }}
      }}
      body {{
        font-family: {'YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif' if is_fa else 'Arial, sans-serif'} !important;
        font-size: 11px;
        color: #222;
      }}
      .header {{
        display: flex;
        justify-content: space-between;
        align-items: flex-start;
        margin-bottom: 10px;
        border-bottom: 2px solid #444;
        padding-bottom: 6px;
        gap: 12px;
      }}
      .title {{
        font-size: 16px;
        font-weight: 700;
      }}
      .meta {{
        font-size: 11px;
        color: #555;
        text-align: {'right' if is_fa else 'left'};
      }}
      .filters {{
        margin: 8px 0 10px;
        font-size: 10.5px;
        color: #444;
        display: flex;
        flex-wrap: wrap;
        gap: 6px 10px;
      }}
      .filters .filter-item {{
        background: #f7f9fc;
        border: 1px solid #e2e8f0;
        padding: 3px 6px;
        border-radius: 6px;
        white-space: nowrap;
      }}
      table.report-table {{
        width: 100%;
        border-collapse: collapse;
        table-layout: fixed;
      }}
      thead th {{
        background: #f0f3f7;
        border: 1px solid #c7cdd6;
        padding: 5px 4px;
        text-align: center;
        font-weight: 700;
        white-space: normal;
        overflow-wrap: anywhere;
        word-break: break-word;
        line-height: 1.2;
        font-size: 10px;
      }}
      tbody td {{
        border: 1px solid #d7dde6;
        padding: 5px 4px;
        vertical-align: top;
        overflow-wrap: anywhere;
        word-break: break-word;
        /* پیش‌فرض: وسط‌چین */
        text-align: center;
      }}
    """)

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=html_content).write_pdf(
        stylesheets=[css],
        font_config=font_config,
    )

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "trial_balance_report"
    if business_name:
        base += f"_{slugify(business_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/general-ledger",
    summary="گزارش دفتر کل",
    description="گزارش دفتر کل برای حساب‌های حسابداری",
)
@require_business_access("business_id")
async def general_ledger_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش دفتر کل برای حساب‌های حسابداری"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # account_ids اجباری است
    account_ids = body.get('account_ids')
    if not account_ids or not isinstance(account_ids, list) or len(account_ids) == 0:
        raise ApiError("ACCOUNT_IDS_REQUIRED", "account_ids الزامی است و باید لیستی از شناسه‌های حساب باشد", http_status=400)
    
    try:
        account_ids = [int(aid) for aid in account_ids if aid is not None]
    except (ValueError, TypeError):
        raise ApiError("INVALID_ACCOUNT_IDS", "account_ids باید لیستی از اعداد صحیح باشد", http_status=400)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    person_id = body.get('person_id')
    project_id = body.get('project_id')  # 🆕 فیلتر پروژه
    include_proforma = body.get('include_proforma', False)
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    if person_id is not None:
        try:
            person_id = int(person_id)
        except (ValueError, TypeError):
            person_id = None
    
    # 🆕 تبدیل نوع project_id
    if project_id is not None:
        try:
            project_id = int(project_id)
        except (ValueError, TypeError):
            project_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_general_ledger_report(
        db=db,
        business_id=business_id,
        account_ids=account_ids,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        person_id=person_id,
        project_id=project_id,  # 🆕 پاس دادن به سرویس
        include_proforma=include_proforma,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="General ledger report retrieved successfully" if locale != 'fa' else "گزارش دفتر کل با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/general-ledger/export/excel",
    summary="خروجی Excel گزارش دفتر کل",
    description="خروجی Excel گزارش دفتر کل با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_general_ledger_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی Excel گزارش دفتر کل"""
    import io
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from adapters.db.models.business import Business

    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    # account_ids اجباری است
    account_ids = body.get("account_ids")
    if not account_ids or not isinstance(account_ids, list) or len(account_ids) == 0:
        raise ApiError("ACCOUNT_IDS_REQUIRED", "account_ids الزامی است و باید لیستی از شناسه‌های حساب باشد", http_status=400)
    try:
        account_ids = [int(aid) for aid in account_ids if aid is not None]
    except (ValueError, TypeError):
        raise ApiError("INVALID_ACCOUNT_IDS", "account_ids باید لیستی از اعداد صحیح باشد", http_status=400)

    # Fiscal year from header/body
    fiscal_year_id = None
    fy_header = request.headers.get("X-Fiscal-Year-ID")
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    if body.get("fiscal_year_id"):
        try:
            fiscal_year_id = int(body["fiscal_year_id"])
        except (ValueError, TypeError):
            pass

    # Filters
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    currency_id = body.get("currency_id")
    person_id = body.get("person_id")
    project_id = body.get("project_id")
    include_proforma = bool(body.get("include_proforma", False))

    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    if person_id is not None:
        try:
            person_id = int(person_id)
        except (ValueError, TypeError):
            person_id = None
    if project_id is not None:
        try:
            project_id = int(project_id)
        except (ValueError, TypeError):
            project_id = None

    # Fetch all (limited)
    max_export_records = 10000
    result = get_general_ledger_report(
        db=db,
        business_id=business_id,
        account_ids=account_ids,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        person_id=person_id,
        project_id=project_id,
        include_proforma=include_proforma,
        skip=0,
        take=max_export_records,
    )
    items = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Selected rows
    selected_only = bool(body.get("selected_only", False))
    selected_indices = body.get("selected_indices")
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Columns
    headers: list[str] = []
    keys: list[str] = []
    export_columns = body.get("export_columns")
    if export_columns:
        for col in export_columns:
            key = col.get("key")
            label = col.get("label", key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        default_columns = [
            ("document_date", "تاریخ سند"),
            ("document_code", "کد سند"),
            ("document_type_name", "نوع سند"),
            ("account_code", "کد حساب"),
            ("account_name", "نام حساب"),
            ("counterpart_code", "کد طرف حساب"),
            ("counterpart_name", "نام طرف حساب"),
            ("description", "شرح"),
            ("debit", "بدهکار"),
            ("credit", "بستانکار"),
            ("balance", "مانده"),
        ]
        if items:
            for k, label in default_columns:
                if k in items[0]:
                    keys.append(k)
                    headers.append(label)

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"

    wb = Workbook()
    ws = wb.active
    ws.title = "General Ledger"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numeric_keys = {"debit", "credit", "balance"}

    def _normalize_number(val: object) -> object:
        if val is None or val == "":
            return ""
        if isinstance(val, bool):
            return 1 if val else 0
        if isinstance(val, int):
            return val
        if isinstance(val, float):
            try:
                return int(val) if val.is_integer() else val
            except Exception:
                return val
        if isinstance(val, str):
            s = val.strip()
            if not s:
                return ""
            try:
                f = float(s)
                return int(f) if f.is_integer() else f
            except Exception:
                return val
        return val

    for item in items:
        row_values = []
        for key in keys:
            value = item.get(key, "")
            if key in numeric_keys:
                value = _normalize_number(value)
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = json.dumps(value, ensure_ascii=False)
            row_values.append(value)
        ws.append(row_values)

    # Format cells
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if keys[c - 1] in numeric_keys and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # پیش‌فرض: وسط‌چین
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()

    # Filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "general_ledger_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/general-ledger/export/pdf",
    summary="خروجی PDF گزارش دفتر کل",
    description="خروجی PDF گزارش دفتر کل با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_general_ledger_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی PDF گزارش دفتر کل"""
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    from adapters.db.models.business import Business
    from adapters.db.models.fiscal_year import FiscalYear
    from adapters.db.models.currency import Currency

    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    # account_ids اجباری
    account_ids = body.get("account_ids")
    if not account_ids or not isinstance(account_ids, list) or len(account_ids) == 0:
        raise ApiError("ACCOUNT_IDS_REQUIRED", "account_ids الزامی است و باید لیستی از شناسه‌های حساب باشد", http_status=400)
    try:
        account_ids = [int(aid) for aid in account_ids if aid is not None]
    except (ValueError, TypeError):
        raise ApiError("INVALID_ACCOUNT_IDS", "account_ids باید لیستی از اعداد صحیح باشد", http_status=400)

    # Fiscal year from header/body
    fiscal_year_id = None
    fy_header = request.headers.get("X-Fiscal-Year-ID")
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    if body.get("fiscal_year_id"):
        try:
            fiscal_year_id = int(body["fiscal_year_id"])
        except (ValueError, TypeError):
            pass

    # Filters
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    currency_id = body.get("currency_id")
    person_id = body.get("person_id")
    project_id = body.get("project_id")
    include_proforma = bool(body.get("include_proforma", False))

    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    if person_id is not None:
        try:
            person_id = int(person_id)
        except (ValueError, TypeError):
            person_id = None
    if project_id is not None:
        try:
            project_id = int(project_id)
        except (ValueError, TypeError):
            project_id = None

    # Fetch all (limited)
    max_export_records = 10000
    result = get_general_ledger_report(
        db=db,
        business_id=business_id,
        account_ids=account_ids,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        person_id=person_id,
        project_id=project_id,
        include_proforma=include_proforma,
        skip=0,
        take=max_export_records,
    )
    items = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Selected rows
    selected_only = bool(body.get("selected_only", False))
    selected_indices = body.get("selected_indices")
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Columns
    headers: list[str] = []
    keys: list[str] = []
    export_columns = body.get("export_columns")
    if export_columns:
        for col in export_columns:
            key = col.get("key")
            label = col.get("label", key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        default_columns = [
            ("document_date", "تاریخ سند"),
            ("document_code", "کد سند"),
            ("document_type_name", "نوع سند"),
            ("account_code", "کد حساب"),
            ("account_name", "نام حساب"),
            ("counterpart_code", "کد طرف حساب"),
            ("counterpart_name", "نام طرف حساب"),
            ("description", "شرح"),
            ("debit", "بدهکار"),
            ("credit", "بستانکار"),
            ("balance", "مانده"),
        ]
        if items:
            for k, label in default_columns:
                if k in items[0]:
                    keys.append(k)
                    headers.append(label)

    # Calendar type + locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    calendar_type = None
    try:
        if hasattr(request.state, "calendar_type") and request.state.calendar_type:
            calendar_type = request.state.calendar_type
    except Exception:
        calendar_type = None
    if not calendar_type:
        cal_header = (request.headers.get("X-Calendar-Type", "jalali") or "jalali").lower()
        calendar_type = "jalali" if cal_header in ["jalali", "persian", "shamsi"] else "gregorian"

    def esc(v: object) -> str:
        return escape("" if v is None else str(v))

    def format_date_for_report(value: object) -> str:
        if value is None:
            return ""
        import datetime as _dt
        dt_obj: _dt.datetime | None = None
        if isinstance(value, _dt.datetime):
            dt_obj = value
        elif isinstance(value, _dt.date):
            dt_obj = _dt.datetime.combine(value, _dt.datetime.min.time())
        elif isinstance(value, str):
            s = value.strip()
            if not s:
                return ""
            try:
                if "T" in s:
                    dt_obj = _dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
                else:
                    d = _dt.date.fromisoformat(s.split(" ")[0].split("T")[0])
                    dt_obj = _dt.datetime.combine(d, _dt.datetime.min.time())
            except Exception:
                return s
        else:
            return str(value)
        try:
            from app.core.calendar import CalendarConverter
            out = CalendarConverter.format_datetime(dt_obj, calendar_type)
            return out.get("date_only") or (out.get("formatted", "").split(" ")[0]) or dt_obj.strftime("%Y/%m/%d")
        except Exception:
            return dt_obj.strftime("%Y/%m/%d")

    def format_num(value: object) -> str:
        if value is None or value == "":
            return ""
        try:
            n = float(value)
            if n.is_integer():
                s = f"{int(n):,}"
            else:
                s = f"{n:,.2f}"
            if is_fa:
                s = s.replace(",", "٬")
            return s
        except Exception:
            return str(value)

    # Business name
    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    # Filters summary
    filters: list[tuple[str, str]] = []
    fy_obj = None
    if fiscal_year_id:
        try:
            fy_obj = db.query(FiscalYear).filter(FiscalYear.id == int(fiscal_year_id), FiscalYear.business_id == int(business_id)).first()
        except Exception:
            fy_obj = None
        filters.append(("سال مالی" if is_fa else "Fiscal Year", (fy_obj.title if fy_obj else str(fiscal_year_id))))
    if date_from:
        filters.append(("از تاریخ" if is_fa else "From", format_date_for_report(date_from)))
    if date_to:
        filters.append(("تا تاریخ" if is_fa else "To", format_date_for_report(date_to)))
    if currency_id:
        cur_label = ""
        try:
            cur = db.query(Currency).filter(Currency.id == int(currency_id)).first()
            if cur is not None:
                cur_label = (cur.title or cur.name or "") if is_fa else (cur.code or cur.name or cur.title or "")
                if cur.symbol:
                    cur_label = f"{cur_label} ({cur.symbol})" if cur_label else cur.symbol
        except Exception:
            cur_label = ""
        filters.append(("ارز" if is_fa else "Currency", cur_label or str(currency_id)))
    if person_id:
        filters.append(("شخص" if is_fa else "Person", str(person_id)))
    if project_id:
        filters.append(("پروژه" if is_fa else "Project", str(project_id)))
    filters.append(("تعداد حساب" if is_fa else "Accounts", str(len(account_ids))))
    if ("include_proforma" in body) or include_proforma:
        filters.append((
            "شامل پیش‌نویس" if is_fa else "Include Proforma",
            ("بله" if include_proforma else "خیر") if is_fa else ("Yes" if include_proforma else "No"),
        ))

    filters_html = ""
    if filters:
        parts = []
        for k, v in filters:
            if v:
                parts.append(f"<span class='filter-item'><strong>{esc(k)}:</strong> {esc(v)}</span>")
        if parts:
            filters_html = f"<div class='filters'>{''.join(parts)}</div>"

    # Rows html
    rows_html = []
    numeric_keys = {"debit", "credit", "balance"}
    for item in items:
        tds = []
        for k in keys:
            v = item.get(k, "")
            if k == "document_date":
                v = format_date_for_report(v)
            if k in numeric_keys:
                v = format_num(v)
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            elif isinstance(v, dict):
                v = json.dumps(v, ensure_ascii=False)
            tds.append(f"<td>{esc(v)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")
    headers_html = "".join(f"<th>{esc(h)}</th>" for h in headers)

    # Date report (calendar-aware)
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(), calendar_type)
        now_str = formatted_now.get("formatted") or formatted_now.get("date_time") or datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
    except Exception:
        now_str = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")

    title_text = "گزارش دفتر کل" if is_fa else "General Ledger Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "

    # Embed fonts like invoice PDFs
    fa_reg, fa_bold = None, None
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_reg, fa_bold = load_farsi_font_data_uris()
    except Exception:
        fa_reg, fa_bold = None, None

    font_face_css = ""
    if fa_reg:
        font_face_css += f"""
        @font-face {{
          font-family: 'YekanBakhFaNum';
          src: url("{fa_reg}") format('truetype');
          font-weight: 400;
          font-style: normal;
        }}
        """
    if fa_bold:
        font_face_css += f"""
        @font-face {{
          font-family: 'YekanBakhFaNum';
          src: url("{fa_bold}") format('truetype');
          font-weight: 700;
          font-style: normal;
        }}
        """

    preferred_stack = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if (is_fa or html_dir == "rtl") else "Arial, sans-serif"
    injected_style = (
        "<style id=\"hesabix-font-inject\">"
        + (font_face_css or "")
        + f"\nhtml, body, body * {{ font-family: {preferred_stack} !important; }}\n"
        + "</style>"
    )

    table_html = f"""
    <html lang="{ 'fa' if is_fa else 'en' }" dir="{ 'rtl' if is_fa else 'ltr' }">
      <head>
        <meta charset="utf-8">
        {injected_style}
      </head>
      <body>
        <div class="header">
          <div class="title">{esc(title_text)}</div>
          <div class="meta">
            <div><strong>{esc(label_biz)}:</strong> {esc(business_name)}</div>
            <div><strong>{esc(label_date)}:</strong> {esc(now_str)}</div>
          </div>
        </div>
        {filters_html}
        <div class="table-wrapper">
          <table class="report-table">
            <thead><tr>{headers_html}</tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
          </table>
        </div>
      </body>
    </html>
    """

    css = CSS(string=f"""
      {font_face_css}
      html, body, body * {{ font-family: {preferred_stack} !important; }}
      @page {{
        size: A4 landscape;
        margin: 12mm;
        @bottom-{'left' if is_fa else 'right'} {{
          content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
          font-size: 10px;
          color: #666;
          font-family: {preferred_stack} !important;
        }}
      }}
      body {{
        font-size: 11px;
        color: #222;
      }}
      .header {{
        display: flex;
        justify-content: space-between;
        align-items: flex-start;
        margin-bottom: 10px;
        border-bottom: 2px solid #444;
        padding-bottom: 6px;
        gap: 12px;
      }}
      .title {{
        font-size: 16px;
        font-weight: 700;
      }}
      .meta {{
        font-size: 11px;
        color: #555;
        text-align: {'right' if is_fa else 'left'};
      }}
      .filters {{
        margin: 8px 0 10px;
        font-size: 10.5px;
        color: #444;
        display: flex;
        flex-wrap: wrap;
        gap: 6px 10px;
      }}
      .filters .filter-item {{
        background: #f7f9fc;
        border: 1px solid #e2e8f0;
        padding: 3px 6px;
        border-radius: 6px;
        white-space: nowrap;
      }}
      table.report-table {{
        width: 100%;
        border-collapse: collapse;
        table-layout: fixed;
      }}
      thead th {{
        background: #f0f3f7;
        border: 1px solid #c7cdd6;
        padding: 5px 4px;
        text-align: center;
        font-weight: 700;
        white-space: normal;
        overflow-wrap: anywhere;
        word-break: break-word;
        line-height: 1.2;
        font-size: 10px;
      }}
      tbody td {{
        border: 1px solid #d7dde6;
        padding: 5px 4px;
        vertical-align: top;
        overflow-wrap: anywhere;
        word-break: break-word;
        /* پیش‌فرض: وسط‌چین */
        text-align: center;
      }}
    """)

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=table_html).write_pdf(stylesheets=[css], font_config=font_config)

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", (text or "")).strip("_")

    base = "general_ledger_report"
    if business_name:
        base += f"_{slugify(business_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/pnl-period",
    summary="گزارش سود و زیان دوره‌ای",
    description="گزارش سود و زیان دوره‌ای (Profit & Loss Period)",
)
@require_business_access("business_id")
async def pnl_period_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش سود و زیان دوره‌ای"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    project_id = body.get('project_id')  # 🆕 فیلتر پروژه
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # 🆕 تبدیل نوع project_id
    if project_id is not None:
        try:
            project_id = int(project_id)
        except (ValueError, TypeError):
            project_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 100)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 100
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 100
    
    result = get_pnl_period_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        project_id=project_id,  # 🆕 پاس دادن به سرویس
        skip=skip,
        take=take,
    )
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="PnL Period report retrieved successfully" if locale != 'fa' else "گزارش سود و زیان دوره‌ای با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/pnl-cumulative",
    summary="گزارش سود و زیان تجمعی",
    description="گزارش سود و زیان تجمعی از ابتدای سال مالی (Profit & Loss Cumulative)",
)
@require_business_access("business_id")
async def pnl_cumulative_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش سود و زیان تجمعی از ابتدای سال مالی"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_to = body.get('date_to')  # فقط date_to (date_from همیشه ابتدای سال مالی است)
    currency_id = body.get('currency_id')
    project_id = body.get('project_id')  # 🆕 فیلتر پروژه
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # 🆕 تبدیل نوع project_id
    if project_id is not None:
        try:
            project_id = int(project_id)
        except (ValueError, TypeError):
            project_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 100)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 100
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 100
    
    result = get_pnl_cumulative_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_to=date_to,
        project_id=project_id,  # 🆕 پاس دادن به سرویس
        skip=skip,
        take=take,
    )
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="PnL Cumulative report retrieved successfully" if locale != 'fa' else "گزارش سود و زیان تجمعی با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/accounts-review",
    summary="گزارش مرور حساب‌ها",
    description="گزارش مرور حساب‌ها با ساختار درختی و مانده‌ها (Account Review Report)",
)
@require_business_access("business_id")
async def accounts_review_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش مرور حساب‌ها با ساختار درختی"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "دسترسی مجاز نیست", http_status=403)
    # بررسی دسترسی - گزارشات معمولاً نیاز به دسترسی خواندن دارند
    # اما اگر این endpoint خاص نیاز به دسترسی خاصی ندارد، می‌توانیم این بررسی را حذف کنیم
    # یا دسترسی را به صورت optional در نظر بگیریم
    # برای اکنون، بررسی می‌کنیم که آیا کاربر دسترسی خواندن گزارشات دارد یا نه
    try:
        if ctx and hasattr(ctx, 'can_read_section'):
            if not ctx.can_read_section("reports"):
                raise ApiError("FORBIDDEN", "دسترسی مجاز نیست", http_status=403)
    except AttributeError:
        # اگر متد can_read_section وجود نداشت، به کاربر اجازه می‌دهیم
        pass
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    account_type = body.get('account_type')
    include_zero_balance = body.get('include_zero_balance', False)
    account_id = body.get('account_id')  # برای دریافت جزئیات یک حساب خاص
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    if account_id is not None:
        try:
            account_id = int(account_id)
        except (ValueError, TypeError):
            account_id = None
    
    # Pagination (فقط برای جزئیات)
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_accounts_review_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        account_type=account_type,
        include_zero_balance=include_zero_balance,
        account_id=account_id,
        skip=skip,
        take=take,
    )
    
    # Format datetime fields
    def _format_accounts(accounts: list[Dict[str, Any]]) -> list[Dict[str, Any]]:
        formatted = []
        for acc in accounts:
            formatted_acc = dict(acc)
            if formatted_acc.get('children'):
                formatted_acc['children'] = _format_accounts(formatted_acc['children'])
            formatted.append(formatted_acc)
        return formatted
    
    result['accounts'] = _format_accounts(result.get('accounts', []))
    
    if result.get('account_details'):
        result['account_details'] = [format_datetime_fields(item, request) for item in result['account_details']]
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Accounts review report retrieved successfully" if locale != 'fa' else "گزارش مرور حساب‌ها با موفقیت دریافت شد",
        request=request
    )


def _flatten_accounts_review_tree(accounts: list) -> list:
    """تبدیل درخت حساب‌های گزارش مرور به لیست تخت برای export."""
    rows = []
    for acc in accounts:
        row = {k: v for k, v in acc.items() if k != "children"}
        rows.append(row)
        if acc.get("children"):
            rows.extend(_flatten_accounts_review_tree(acc["children"]))
    return rows


@router.post(
    "/businesses/{business_id}/reports/accounts-review/export/excel",
    summary="خروجی Excel گزارش مرور حساب‌ها",
    description="خروجی Excel گزارش مرور حساب‌ها با ساختار درختی و مانده‌ها",
)
@require_business_access("business_id")
async def export_accounts_review_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی Excel گزارش مرور حساب‌ها"""
    import io
    import re
    import datetime
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from adapters.db.models.business import Business

    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    fiscal_year_id = None
    fy_header = request.headers.get("X-Fiscal-Year-ID")
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    if body.get("fiscal_year_id"):
        try:
            fiscal_year_id = int(body["fiscal_year_id"])
        except (ValueError, TypeError):
            pass

    date_from = body.get("date_from")
    date_to = body.get("date_to")
    currency_id = body.get("currency_id")
    account_type = body.get("account_type")
    include_zero_balance = bool(body.get("include_zero_balance", False))

    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None

    result = get_accounts_review_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        account_type=account_type,
        include_zero_balance=include_zero_balance,
        account_id=None,
        skip=0,
        take=50,
    )
    accounts_tree = result.get("accounts", [])
    items = _flatten_accounts_review_tree(accounts_tree)
    items = [format_datetime_fields(item, request) for item in items]

    keys = [
        "account_code",
        "account_name",
        "account_type",
        "opening_debit",
        "opening_credit",
        "period_debit",
        "period_credit",
        "closing_debit",
        "closing_credit",
    ]
    headers_fa = [
        "کد حساب",
        "نام حساب",
        "نوع حساب",
        "مانده ابتدای دوره (بدهکار)",
        "مانده ابتدای دوره (بستانکار)",
        "جمع بدهکار دوره",
        "جمع بستانکار دوره",
        "مانده انتهای دوره (بدهکار)",
        "مانده انتهای دوره (بستانکار)",
    ]
    headers_en = [
        "Account Code",
        "Account Name",
        "Account Type",
        "Opening Debit",
        "Opening Credit",
        "Period Debit",
        "Period Credit",
        "Closing Debit",
        "Closing Credit",
    ]
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    headers = headers_fa if is_fa else headers_en

    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts Review"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    mapping_fa = {
        "accounting_document": "حسابداری",
        "bank": "بانک",
        "cash_register": "صندوق",
        "cashdesk": "صندوق",
        "cash": "نقد",
        "petty_cash": "تنخواه",
        "check": "چک",
        "person": "شخص",
        "product": "کالا",
    }
    mapping_en = {
        "accounting_document": "Accounting",
        "bank": "Bank",
        "cash_register": "Cash Register",
        "cashdesk": "Cash Register",
        "cash": "Cash",
        "petty_cash": "Petty Cash",
        "check": "Check",
        "person": "Person",
        "product": "Product",
    }
    for k in [str(i) for i in range(44)]:
        mapping_fa.setdefault(k, k)
        mapping_en.setdefault(k, k)

    def map_account_type(v):
        if v is None:
            return ""
        key = str(v).strip()
        return (mapping_fa if is_fa else mapping_en).get(key, key)

    numeric_keys = {"opening_debit", "opening_credit", "period_debit", "period_credit", "closing_debit", "closing_credit"}

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in items:
        row = []
        for key in keys:
            value = item.get(key, "")
            if key == "account_type":
                value = map_account_type(value)
            if key in numeric_keys and value != "":
                try:
                    if isinstance(value, float) and value == int(value):
                        value = int(value)
                except Exception:
                    pass
            row.append(value)
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", str(text)).strip("_")

    filename = f"accounts_review_{slugify(biz_name) or business_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/accounts-review/export/pdf",
    summary="خروجی PDF گزارش مرور حساب‌ها",
    description="خروجی PDF گزارش مرور حساب‌ها با ساختار درختی و مانده‌ها",
)
@require_business_access("business_id")
async def export_accounts_review_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی PDF گزارش مرور حساب‌ها"""
    import datetime
    import re
    from html import escape
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from adapters.db.models.business import Business
    from adapters.db.models.fiscal_year import FiscalYear
    from adapters.db.models.currency import Currency

    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    fiscal_year_id = None
    fy_header = request.headers.get("X-Fiscal-Year-ID")
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    if body.get("fiscal_year_id"):
        try:
            fiscal_year_id = int(body["fiscal_year_id"])
        except (ValueError, TypeError):
            pass

    date_from = body.get("date_from")
    date_to = body.get("date_to")
    currency_id = body.get("currency_id")
    account_type = body.get("account_type")
    include_zero_balance = bool(body.get("include_zero_balance", False))

    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None

    result = get_accounts_review_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        account_type=account_type,
        include_zero_balance=include_zero_balance,
        account_id=None,
        skip=0,
        take=50,
    )
    accounts_tree = result.get("accounts", [])
    items = _flatten_accounts_review_tree(accounts_tree)
    items = [format_datetime_fields(item, request) for item in items]

    keys = [
        "account_code",
        "account_name",
        "account_type",
        "opening_debit",
        "opening_credit",
        "period_debit",
        "period_credit",
        "closing_debit",
        "closing_credit",
    ]
    headers_fa = [
        "کد حساب",
        "نام حساب",
        "نوع حساب",
        "مانده ابتدای دوره (بدهکار)",
        "مانده ابتدای دوره (بستانکار)",
        "جمع بدهکار دوره",
        "جمع بستانکار دوره",
        "مانده انتهای دوره (بدهکار)",
        "مانده انتهای دوره (بستانکار)",
    ]
    headers_en = [
        "Account Code",
        "Account Name",
        "Account Type",
        "Opening Debit",
        "Opening Credit",
        "Period Debit",
        "Period Credit",
        "Closing Debit",
        "Closing Credit",
    ]
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == "fa"
    headers = headers_fa if is_fa else headers_en

    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    mapping_fa = {
        "accounting_document": "حسابداری",
        "bank": "بانک",
        "cash_register": "صندوق",
        "cashdesk": "صندوق",
        "cash": "نقد",
        "petty_cash": "تنخواه",
        "check": "چک",
        "person": "شخص",
        "product": "کالا",
    }
    mapping_en = {
        "accounting_document": "Accounting",
        "bank": "Bank",
        "cash_register": "Cash Register",
        "cashdesk": "Cash Register",
        "cash": "Cash",
        "petty_cash": "Petty Cash",
        "check": "Check",
        "person": "Person",
        "product": "Product",
    }
    for k in [str(i) for i in range(44)]:
        mapping_fa.setdefault(k, k)
        mapping_en.setdefault(k, k)

    def map_account_type_label(raw_value):
        if raw_value is None:
            return ""
        key = str(raw_value).strip()
        return (mapping_fa if is_fa else mapping_en).get(key, key)

    numeric_keys = {"opening_debit", "opening_credit", "period_debit", "period_credit", "closing_debit", "closing_credit"}

    def format_number_for_export(raw_value):
        if raw_value is None:
            return ""
        if isinstance(raw_value, bool):
            return "1" if raw_value else "0"
        s = (str(raw_value) or "").strip()
        if not s:
            return ""
        try:
            f = float(s)
            if f == int(f):
                s = str(int(f))
        except Exception:
            pass
        if is_fa and "," in s:
            s = s.replace(",", "٬")
        return s

    def esc(v):
        return escape("" if v is None else str(v))

    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key, "")
            if key == "account_type":
                value = map_account_type_label(value)
            if key in numeric_keys:
                value = format_number_for_export(value)
            tds.append(f"<td>{esc(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = "".join(f"<th>{esc(h)}</th>" for h in headers)

    calendar_type = None
    try:
        if hasattr(request.state, "calendar_type") and request.state.calendar_type:
            calendar_type = request.state.calendar_type
    except Exception:
        calendar_type = None
    if not calendar_type:
        cal_header = (request.headers.get("X-Calendar-Type", "jalali") or "jalali").lower()
        calendar_type = "jalali" if cal_header in ["jalali", "persian", "shamsi"] else "gregorian"
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(), calendar_type)
        now = formatted_now.get("formatted") or formatted_now.get("date_time") or datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
    except Exception:
        now = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")

    filters = []
    fy_obj = None
    if fiscal_year_id:
        try:
            fy_obj = db.query(FiscalYear).filter(
                FiscalYear.id == int(fiscal_year_id),
                FiscalYear.business_id == int(business_id),
            ).first()
        except Exception:
            fy_obj = None
        fy_title = (fy_obj.title if fy_obj else str(fiscal_year_id))
        filters.append(("سال مالی" if is_fa else "Fiscal Year", str(fy_title)))
    if date_from or date_to:
        if date_from:
            filters.append(("از تاریخ" if is_fa else "From", date_from))
        if date_to:
            filters.append(("تا تاریخ" if is_fa else "To", date_to))
    if currency_id:
        cur_label = ""
        try:
            cur = db.query(Currency).filter(Currency.id == int(currency_id)).first()
            if cur is not None:
                cur_label = (cur.title or cur.name or "") if is_fa else (cur.code or cur.name or cur.title or "")
                if cur.symbol:
                    cur_label = f"{cur_label} ({cur.symbol})" if cur_label else cur.symbol
        except Exception:
            cur_label = ""
        filters.append(("ارز" if is_fa else "Currency", cur_label or str(currency_id)))
    if account_type:
        filters.append(("نوع حساب" if is_fa else "Account Type", map_account_type_label(account_type)))
    filters.append((
        "شامل مانده صفر" if is_fa else "Include Zero Balance",
        ("بله" if include_zero_balance else "خیر") if is_fa else ("Yes" if include_zero_balance else "No"),
    ))

    filters_html = ""
    if filters:
        parts = [f"<span class='filter-item'><strong>{esc(k)}:</strong> {esc(v)}</span>" for k, v in filters if v]
        if parts:
            filters_html = f"<div class='filters'>{''.join(parts)}</div>"

    title_text = "گزارش مرور حساب‌ها" if is_fa else "Accounts Review Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    html_lang = "fa" if is_fa else "en"
    html_dir = "rtl" if is_fa else "ltr"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "

    fa_font_url_regular = None
    fa_font_url_bold = None
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_font_url_regular, fa_font_url_bold = load_farsi_font_data_uris()
    except Exception:
        pass

    table_html = f"""
    <html lang="{html_lang}" dir="{html_dir}">
      <head><meta charset="utf-8"></head>
      <body>
        <div class="header">
          <div class="title">{esc(title_text)}</div>
          <div class="meta">
            <div><strong>{esc(label_biz)}:</strong> {esc(business_name)}</div>
            <div><strong>{esc(label_date)}:</strong> {esc(now)}</div>
          </div>
        </div>
        {filters_html}
        <div class="table-wrapper">
          <table class="report-table">
            <thead><tr>{headers_html}</tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
          </table>
        </div>
      </body>
    </html>
    """

    font_face_css = ""
    if fa_font_url_regular:
        font_face_css += f'@font-face {{ font-family: \'YekanBakhFaNum\'; src: url("{fa_font_url_regular}") format(\'truetype\'); font-weight: 400; font-style: normal; }}\n'
    if fa_font_url_bold:
        font_face_css += f'@font-face {{ font-family: \'YekanBakhFaNum\'; src: url("{fa_font_url_bold}") format(\'truetype\'); font-weight: 700; font-style: normal; }}\n'
    try:
        force_fa = is_fa or ('dir="rtl"' in table_html) or ("dir='rtl'" in table_html)
        preferred_stack = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if force_fa else "Arial, sans-serif"
        injected = "<style id=\"hesabix-font-inject\">" + (font_face_css or "") + f"\nhtml, body, body * {{ font-family: {preferred_stack} !important; }}\n</style>"
        if "</head>" in table_html:
            table_html = table_html.replace("</head>", injected + "</head>")
        else:
            table_html = injected + table_html
    except Exception:
        pass

    force_font_css = "\nhtml, body, body * { font-family: YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif !important; }\n" if is_fa else "\nhtml, body, body * { font-family: Arial, sans-serif !important; }\n"
    css = CSS(string=(font_face_css or "") + force_font_css + f"""
      @page {{ size: A4 landscape; margin: 12mm;
        @bottom-{'left' if is_fa else 'right'} {{ content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages); font-size: 10px; color: #666; font-family: {'YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif' if is_fa else 'Arial, sans-serif'} !important; }}
      }}
      body {{ font-size: 11px; color: #222; }}
      .header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 10px; border-bottom: 2px solid #444; padding-bottom: 6px; gap: 12px; }}
      .title {{ font-size: 16px; font-weight: 700; }}
      .meta {{ font-size: 11px; color: #555; text-align: {'right' if is_fa else 'left'}; }}
      .filters {{ margin: 8px 0 10px; font-size: 10.5px; color: #444; display: flex; flex-wrap: wrap; gap: 6px 10px; }}
      .filters .filter-item {{ background: #f7f9fc; border: 1px solid #e2e8f0; padding: 3px 6px; border-radius: 6px; white-space: nowrap; }}
      table.report-table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
      thead th {{ background: #f0f3f7; border: 1px solid #c7cdd6; padding: 5px 4px; text-align: center; font-weight: 700; white-space: normal; font-size: 10px; }}
      tbody td {{ border: 1px solid #d7dde6; padding: 5px 4px; vertical-align: top; overflow-wrap: anywhere; word-break: break-word; text-align: center; }}
    """)

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=table_html).write_pdf(stylesheets=[css], font_config=font_config)

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", str(text)).strip("_")

    filename = f"accounts_review_{slugify(business_name) or business_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/journal-ledger",
    summary="گزارش دفتر روزنامه",
    description="گزارش دفتر روزنامه - تمام تراکنش‌های مالی به ترتیب تاریخ (Journal Ledger)",
)
@require_business_access("business_id")
async def journal_ledger_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش دفتر روزنامه - تمام تراکنش‌های مالی به ترتیب تاریخ"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    document_type = body.get('document_type')
    include_proforma = body.get('include_proforma', False)
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50

    # کش نتایج گزارش دفتر روزنامه
    cache = get_cache()
    cache_key = None

    if cache.enabled:
        import json, hashlib
        key_payload = {
            "business_id": business_id,
            "fiscal_year_id": fiscal_year_id,
            "currency_id": currency_id,
            "date_from": date_from,
            "date_to": date_to,
            "document_type": document_type,
            "include_proforma": include_proforma,
            "skip": skip,
            "take": take,
        }
        key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
        cache_key = f"journal_ledger:{key_hash}"
        cached = cache.get(cache_key)
        if cached is not None:
            locale = negotiate_locale(request.headers.get("Accept-Language"))
            return success_response(
                data=cached,
                message="Journal ledger report retrieved successfully" if locale != 'fa' else "گزارش دفتر روزنامه با موفقیت دریافت شد",
                request=request
            )

    result = get_journal_ledger_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        document_type=document_type,
        include_proforma=include_proforma,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items

    if cache.enabled and cache_key:
        cache.set(cache_key, result, ttl=60)
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Journal ledger report retrieved successfully" if locale != 'fa' else "گزارش دفتر روزنامه با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/journal-ledger/export/excel",
    summary="خروجی Excel گزارش دفتر روزنامه",
    description="خروجی Excel گزارش دفتر روزنامه با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_journal_ledger_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel گزارش دفتر روزنامه"""
    import io
    import json
    import datetime
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response
    from adapters.db.models.business import Business
    from app.core.calendar import CalendarConverter, CalendarType
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت calendar_type از request state
    calendar_type: CalendarType = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    document_type = body.get('document_type')
    include_proforma = body.get('include_proforma', False)
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_journal_ledger_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        document_type=document_type,
        include_proforma=include_proforma,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains Persian or has / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and Persian numbers or has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Helper function to get document type name in Persian
    def get_document_type_name_fa(doc_type: str | None) -> str:
        """تبدیل نوع سند به نام فارسی"""
        if not doc_type:
            return ''
        
        mapping = {
            'invoice_sales': 'فاکتور فروش',
            'invoice_sales_return': 'برگشت از فروش',
            'invoice_purchase': 'فاکتور خرید',
            'invoice_purchase_return': 'برگشت از خرید',
            'invoice_production': 'فاکتور تولید',
            'invoice_direct_consumption': 'مصرف مستقیم',
            'invoice_waste': 'ضایعات',
            'receipt': 'دریافت',
            'payment': 'پرداخت',
            'transfer': 'انتقال',
            'expense': 'هزینه',
            'income': 'درآمد',
            'expense_income': 'درآمد/هزینه',
            'opening_balance': 'تراز افتتاحیه',
            'manual': 'سند دستی',
            'manual_document': 'سند دستی',
            'check_endorse': 'پاسخگویی چک',
            'check_clear': 'وصول چک',
            'check_pay': 'پرداخت چک',
            'check_return': 'برگشت چک',
            'check_bounce': 'برگشت خوردن چک',
            'check_deposit': 'واریز به حساب',
            'check_delete': 'حذف چک',
        }
        return mapping.get(doc_type, doc_type)
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            'document_date': '',
            'document_code': '⚠️',
            'document_type_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'description': '',
            'general_account_code': '',
            'general_account_name': '',
            'subsidiary_account_code': '',
            'subsidiary_account_name': '',
            'debit_account_code': '',
            'debit_account_name': '',
            'debit_amount': '',
            'credit_account_code': '',
            'credit_account_name': '',
            'credit_amount': '',
            'person_name': '',
        }
        items.append(warning_item)
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('document_date', 'تاریخ سند' if is_fa else 'Document Date'),
            ('document_type_name', 'نوع سند' if is_fa else 'Document Type'),
            ('document_code', 'شماره سند' if is_fa else 'Document Code'),
            ('description', 'شرح' if is_fa else 'Description'),
            ('general_account_code', 'کد حساب کل' if is_fa else 'General Account Code'),
            ('general_account_name', 'عنوان حساب کل' if is_fa else 'General Account Name'),
            ('subsidiary_account_code', 'کد حساب معین' if is_fa else 'Subsidiary Account Code'),
            ('subsidiary_account_name', 'عنوان حساب معین' if is_fa else 'Subsidiary Account Name'),
            ('debit_account_code', 'کد حساب بدهکار' if is_fa else 'Debit Account Code'),
            ('debit_account_name', 'نام حساب بدهکار' if is_fa else 'Debit Account Name'),
            ('debit_amount', 'مبلغ بدهکار' if is_fa else 'Debit Amount'),
            ('credit_account_code', 'کد حساب بستانکار' if is_fa else 'Credit Account Code'),
            ('credit_account_name', 'نام حساب بستانکار' if is_fa else 'Credit Account Name'),
            ('credit_amount', 'مبلغ بستانکار' if is_fa else 'Credit Amount'),
            ('person_name', 'طرف حساب' if is_fa else 'Counterpart'),
        ]
        for key, label in default_columns:
            keys.append(key)
            headers.append(label)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "دفتر روزنامه" if is_fa else "Journal Ledger"
    
    # RTL handling for Persian
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            
            # Format dates based on calendar type
            if key == 'document_date' and value:
                value = format_date_for_export(item, 'document_date')
            
            # Format document type name (ensure it's in Persian)
            if key == 'document_type_name':
                # Always use document_type to get the correct translation
                doc_type = item.get('document_type', '')
                if doc_type:
                    value = get_document_type_name_fa(doc_type)
                elif value:
                    # If document_type is not available, check if value is English and translate
                    # Check if value contains Persian characters
                    has_persian = any('\u0600' <= c <= '\u06FF' for c in str(value))
                    if not has_persian:
                        # Value is likely in English, try to translate it
                        value = get_document_type_name_fa(value)
                else:
                    value = ""
            
            # Format numbers
            if key in ['debit_amount', 'credit_amount'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    value = num_value
                except (ValueError, TypeError):
                    pass
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # RTL alignment for Persian text and numbers
            if locale == 'fa':
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, str) and any('\u0600' <= c <= '\u06FF' for c in str(value)):
                    cell.alignment = Alignment(horizontal="right")
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if max_length < cell_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Add summary row at the end
    summary_row = len(items) + 3
    ws.cell(row=summary_row, column=1, value="جمع" if is_fa else "Total").font = Font(bold=True)
    ws.cell(row=summary_row, column=1).border = border
    
    # Find debit_amount and credit_amount column indices
    debit_col_idx = None
    credit_col_idx = None
    for idx, key in enumerate(keys, 1):
        if key == 'debit_amount':
            debit_col_idx = idx
        elif key == 'credit_amount':
            credit_col_idx = idx
    
    if debit_col_idx:
        total_debit = result.get('summary', {}).get('total_debit', 0)
        ws.cell(row=summary_row, column=debit_col_idx, value=float(total_debit)).font = Font(bold=True)
        ws.cell(row=summary_row, column=debit_col_idx).border = border
        if locale == 'fa':
            ws.cell(row=summary_row, column=debit_col_idx).alignment = Alignment(horizontal="right")
    
    if credit_col_idx:
        total_credit = result.get('summary', {}).get('total_credit', 0)
        ws.cell(row=summary_row, column=credit_col_idx, value=float(total_credit)).font = Font(bold=True)
        ws.cell(row=summary_row, column=credit_col_idx).border = border
        if locale == 'fa':
            ws.cell(row=summary_row, column=credit_col_idx).alignment = Alignment(horizontal="right")
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "journal_ledger_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/journal-ledger/export/pdf",
    summary="خروجی PDF گزارش دفتر روزنامه",
    description="خروجی PDF گزارش دفتر روزنامه با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_journal_ledger_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی PDF گزارش دفتر روزنامه"""
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    import datetime
    import re
    import json
    from adapters.db.models.business import Business
    from app.core.calendar import CalendarConverter, CalendarType
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت calendar_type از request state
    calendar_type: CalendarType = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    document_type = body.get('document_type')
    include_proforma = body.get('include_proforma', False)
    
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_journal_ledger_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        document_type=document_type,
        include_proforma=include_proforma,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        if isinstance(value, str):
            if '/' in value and (len(value.split('/')) == 3):
                if '-' in value:
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        return str(value) if value else ""
    
    # Helper function to get document type name in Persian
    def get_document_type_name_fa(doc_type: str | None) -> str:
        """تبدیل نوع سند به نام فارسی"""
        if not doc_type:
            return ''
        
        mapping = {
            'invoice_sales': 'فاکتور فروش',
            'invoice_sales_return': 'برگشت از فروش',
            'invoice_purchase': 'فاکتور خرید',
            'invoice_purchase_return': 'برگشت از خرید',
            'invoice_production': 'فاکتور تولید',
            'invoice_direct_consumption': 'مصرف مستقیم',
            'invoice_waste': 'ضایعات',
            'receipt': 'دریافت',
            'payment': 'پرداخت',
            'transfer': 'انتقال',
            'expense': 'هزینه',
            'income': 'درآمد',
            'expense_income': 'درآمد/هزینه',
            'opening_balance': 'تراز افتتاحیه',
            'manual': 'سند دستی',
            'manual_document': 'سند دستی',
            'check_endorse': 'پاسخگویی چک',
            'check_clear': 'وصول چک',
            'check_pay': 'پرداخت چک',
            'check_return': 'برگشت چک',
            'check_bounce': 'برگشت خوردن چک',
            'check_deposit': 'واریز به حساب',
            'check_delete': 'حذف چک',
        }
        return mapping.get(doc_type, doc_type)
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            'document_date': '',
            'document_code': '⚠️',
            'document_type_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'description': '',
            'general_account_code': '',
            'general_account_name': '',
            'subsidiary_account_code': '',
            'subsidiary_account_name': '',
            'debit_account_code': '',
            'debit_account_name': '',
            'debit_amount': '',
            'credit_account_code': '',
            'credit_account_name': '',
            'credit_amount': '',
            'person_name': '',
        }
        items.append(warning_item)
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('document_date', 'تاریخ سند' if is_fa else 'Document Date'),
            ('document_type_name', 'نوع سند' if is_fa else 'Document Type'),
            ('document_code', 'شماره سند' if is_fa else 'Document Code'),
            ('description', 'شرح' if is_fa else 'Description'),
            ('general_account_code', 'کد حساب کل' if is_fa else 'General Account Code'),
            ('general_account_name', 'عنوان حساب کل' if is_fa else 'General Account Name'),
            ('subsidiary_account_code', 'کد حساب معین' if is_fa else 'Subsidiary Account Code'),
            ('subsidiary_account_name', 'عنوان حساب معین' if is_fa else 'Subsidiary Account Name'),
            ('debit_account', 'حساب بدهکار' if is_fa else 'Debit Account'),
            ('debit_amount', 'مبلغ بدهکار' if is_fa else 'Debit Amount'),
            ('credit_account', 'حساب بستانکار' if is_fa else 'Credit Account'),
            ('credit_amount', 'مبلغ بستانکار' if is_fa else 'Credit Amount'),
            ('person_name', 'طرف حساب' if is_fa else 'Counterpart'),
        ]
        for key, label in default_columns:
            keys.append(key)
            headers.append(label)
    
    # Get business name
    business_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""
    
    # Prepare data for HTML
    now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    title_text = "گزارش دفتر روزنامه" if is_fa else "Journal Ledger Report"
    label_biz = "کسب و کار" if is_fa else "Business"
    label_date = "تاریخ تولید" if is_fa else "Generated Date"
    footer_text = f"تولید شده در {now}" if is_fa else f"Generated at {now}"
    
    # Create headers HTML
    headers_html = ''.join(f'<th>{escape(header)}</th>' for header in headers)
    
    # Create rows HTML
    rows_html = []
    for item in items:
        row_cells = []
        for key in keys:
            value = item.get(key, "")
            
            # Format dates based on calendar type
            if key == 'document_date' and value:
                value = format_date_for_export(item, 'document_date')
            
            # Format document type name (ensure it's in Persian)
            if key == 'document_type_name':
                doc_type = item.get('document_type', '')
                if doc_type:
                    value = get_document_type_name_fa(doc_type)
                elif value:
                    has_persian = any('\u0600' <= c <= '\u06FF' for c in str(value))
                    if not has_persian:
                        value = get_document_type_name_fa(value)
                else:
                    value = ""
            
            # Format numbers
            if key in ['debit_amount', 'credit_amount'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    value = f"{num_value:,.0f}" if is_fa else f"{num_value:,.2f}"
                except (ValueError, TypeError):
                    pass
            
            # Handle debit_account and credit_account (they might be formatted as "code - name")
            if key in ['debit_account', 'credit_account']:
                debit_code = item.get('debit_account_code', '')
                debit_name = item.get('debit_account_name', '')
                credit_code = item.get('credit_account_code', '')
                credit_name = item.get('credit_account_name', '')
                if key == 'debit_account':
                    if debit_code and debit_name:
                        value = f"{debit_code} - {debit_name}"
                    elif debit_code:
                        value = debit_code
                    elif debit_name:
                        value = debit_name
                elif key == 'credit_account':
                    if credit_code and credit_name:
                        value = f"{credit_code} - {credit_name}"
                    elif credit_code:
                        value = credit_code
                    elif credit_name:
                        value = credit_name
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            row_cells.append(f'<td>{escape(str(value))}</td>')
        rows_html.append(f'<tr>{"".join(row_cells)}</tr>')
    
    # Add summary row
    summary = result.get('summary', {})
    total_debit = summary.get('total_debit', 0)
    total_credit = summary.get('total_credit', 0)
    
    # Find debit_amount and credit_amount column indices
    debit_col_idx = None
    credit_col_idx = None
    for idx, key in enumerate(keys):
        if key == 'debit_amount':
            debit_col_idx = idx
        elif key == 'credit_amount':
            credit_col_idx = idx
    
    # ساخت summary row (جدا از rows_html تا در آخرین صفحه قرار بگیرد)
    summary_row_cells = ['<td></td>'] * len(keys)
    summary_row_cells[0] = f'<td><strong>{"جمع" if is_fa else "Total"}</strong></td>'
    if debit_col_idx is not None:
        debit_format = f"{total_debit:,.0f}" if is_fa else f"{total_debit:,.2f}"
        summary_row_cells[debit_col_idx] = f'<td><strong>{debit_format}</strong></td>'
    if credit_col_idx is not None:
        credit_format = f"{total_credit:,.0f}" if is_fa else f"{total_credit:,.2f}"
        summary_row_cells[credit_col_idx] = f'<td><strong>{credit_format}</strong></td>'
    summary_row_html = f'<tr class="summary-row">{"".join(summary_row_cells)}</tr>'
    
    # کانتکست برای قالب سفارشی لیست
    template_context: Dict[str, Any] = {
        "title_text": title_text,
        "business_name": business_name,
        "generated_at": now,
        "is_fa": is_fa,
        "headers": headers,
        "keys": keys,
        "items": items,
        "table_headers_html": headers_html,
        "table_rows_html": "".join(rows_html),
        "summary": summary,
    }
    
    # تلاش برای رندر با قالب سفارشی (journal_ledger/list)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="journal_ledger",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None
    
    # HTML پیش‌فرض جدول با قالب فایل
    disposition = "attachment"
    try:
        disposition = str(body.get("disposition") or "attachment")
    except Exception:
        disposition = "attachment"
    paper_size = body.get("paper_size") or "A4"
    orientation = body.get("orientation") or "landscape"  # پیش‌فرض افقی
    
    # تقسیم سطرها به صفحات (حداکثر 25 سطر در هر صفحه برای A4 landscape)
    rows_per_page = 25
    total_rows = len(rows_html)
    pages_data = []
    
    for i in range(0, total_rows, rows_per_page):
        page_rows = rows_html[i:i + rows_per_page]
        pages_data.append(page_rows)
    
    # ساخت HTML برای هر صفحه با header در هر صفحه
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "
    
    pages_html = []
    for page_num, page_rows in enumerate(pages_data, 1):
        is_last_page = page_num == len(pages_data)
        # در آخرین صفحه، summary row را اضافه می‌کنیم
        tbody_content = "".join(page_rows)
        if is_last_page:
            tbody_content += summary_row_html
        
        page_html = f"""
        <div class="page-container">
            <div class="header">
                <div>
                    <div class="title">{title_text}</div>
                    {f'<div class="meta"><strong>{label_biz}:</strong> {escape(business_name)}</div>' if business_name else ''}
                </div>
                <div class="meta">{label_date}: {now}</div>
            </div>
            <table class="report-table">
                <thead>
                    <tr>{headers_html}</tr>
                </thead>
                <tbody>
                    {tbody_content}
                </tbody>
            </table>
            {f'<div class="footer">{footer_text}</div>' if is_last_page else ''}
        </div>
        """
        pages_html.append(page_html)
    
    # Use default template if custom template not found
    final_html = resolved_html or f"""
    <!DOCTYPE html>
    <html lang="{'fa' if is_fa else 'en'}" dir="{'rtl' if is_fa else 'ltr'}">
    <head>
        <meta charset="UTF-8">
        <title>{title_text}</title>
        <style>
            @page {{
                size: {paper_size} {orientation};
                margin: 12mm;
                @bottom-{'left' if is_fa else 'right'} {{
                    content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
                    font-size: 10px;
                    color: #666;
                }}
            }}
            body {{
                font-family: {'Tahoma, Arial' if is_fa else 'Arial, sans-serif'};
                direction: {'rtl' if is_fa else 'ltr'};
                font-size: 11px;
                color: #222;
                margin: 0;
                padding: 0;
            }}
            .page-container {{
                page-break-after: always;
                min-height: 100%;
            }}
            .page-container:last-child {{
                page-break-after: auto;
            }}
            .header {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 10px;
                border-bottom: 2px solid #444;
                padding-bottom: 6px;
            }}
            .title {{
                font-size: 16px;
                font-weight: 700;
                color: #366092;
                margin-bottom: 4px;
            }}
            .meta {{
                font-size: 11px;
                color: #555;
            }}
            table.report-table {{
                width: 100%;
                border-collapse: collapse;
                table-layout: fixed;
                margin-top: 10px;
            }}
            thead {{
                display: table-header-group;
            }}
            thead th {{
                background: #366092;
                color: white;
                border: 1px solid #2a4a6e;
                padding: 8px 6px;
                text-align: {'right' if is_fa else 'left'};
                font-weight: 700;
                white-space: nowrap;
                font-size: 10px;
            }}
            tbody tr {{
                page-break-inside: avoid;
                break-inside: avoid;
            }}
            tbody tr:nth-child(even) {{
                background-color: #f2f2f2;
            }}
            tbody tr.summary-row {{
                background-color: #e8f4f8;
                font-weight: bold;
                page-break-inside: avoid;
            }}
            tbody tr.summary-row td {{
                border-top: 2px solid #366092;
            }}
            tbody td {{
                border: 1px solid #d7dde6;
                padding: 6px 5px;
                text-align: {'right' if is_fa else 'left'};
                vertical-align: top;
                overflow-wrap: anywhere;
                word-break: break-word;
                white-space: normal;
                font-size: 10px;
            }}
            .footer {{
                margin-top: 10px;
                font-size: 10px;
                color: #666;
                text-align: center;
                padding-top: 10px;
                border-top: 1px solid #ddd;
            }}
        </style>
    </head>
    <body>
        {"".join(pages_html)}
    </body>
    </html>
    """
    
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)
    
    # Build meaningful filename
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "journal_ledger_report"
    if business_name:
        base += f"_{slugify(business_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )

