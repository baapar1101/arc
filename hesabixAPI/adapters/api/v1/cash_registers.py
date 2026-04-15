from typing import Any, Dict, List
from fastapi import APIRouter, Depends, Request, Body
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from adapters.db.models.cash_register import CashRegister
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.permissions import require_business_management_dep, require_business_access, require_business_permission_dep, require_business_permission_by_entity_dep
from adapters.api.v1.schemas import QueryInfo
from app.services.cash_register_service import (
    create_cash_register,
    update_cash_register,
    delete_cash_register,
    get_cash_register_by_id,
    list_cash_registers,
    bulk_delete_cash_registers,
    get_cash_petty_turnover_report,
)
from app.core.cache import get_cache


router = APIRouter(prefix="/cash-registers", tags=["مدیریت مالی"])


@router.post(
    "/businesses/{business_id}/cash-registers",
    summary="لیست صندوق‌ها",
    description="دریافت لیست صندوق‌های یک کسب و کار با امکان جستجو و فیلتر",
)
@require_business_access("business_id")
async def list_cash_registers_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    # دریافت سال مالی از header یا query
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    query_dict: Dict[str, Any] = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "sort": [s.model_dump() for s in query_info.sort] if query_info.sort else None,
        "search": query_info.search,
        "search_fields": query_info.search_fields,
        "filters": query_info.filters,
    }
    
    if fiscal_year_id:
        query_dict["fiscal_year_id"] = fiscal_year_id

    # کش لیست صندوق‌ها
    cache = get_cache()
    cache_key = None

    if cache.enabled:
        import json, hashlib
        key_payload = {
            "business_id": business_id,
            "query": query_dict,
        }
        key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
        cache_key = f"cash_registers_list:{key_hash}"
        cached = cache.get(cache_key)
        if cached is not None:
            return success_response(data=cached, request=request, message="CASH_REGISTERS_LIST_FETCHED")

    result = list_cash_registers(db, business_id, query_dict)
    result["items"] = [format_datetime_fields(item, request) for item in result.get("items", [])]

    if cache.enabled and cache_key:
        cache.set(cache_key, result, ttl=60)

    return success_response(data=result, request=request, message="CASH_REGISTERS_LIST_FETCHED")


@router.post(
    "/businesses/{business_id}/cash-registers/create",
    summary="ایجاد صندوق جدید",
    description="ایجاد صندوق برای یک کسب‌وکار مشخص",
)
@require_business_access("business_id")
async def create_cash_register_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("cash", "add")),
):
    payload: Dict[str, Any] = dict(body or {})
    created = create_cash_register(db, business_id, payload)
    return success_response(data=format_datetime_fields(created, request), request=request, message="CASH_REGISTER_CREATED")


@router.get(
    "/cash-registers/{cash_id}",
    summary="جزئیات صندوق",
    description="دریافت جزئیات صندوق بر اساس شناسه",
)
async def get_cash_register_endpoint(
    request: Request,
    cash_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("cash", "view", CashRegister, "cash_id")),
):
    result = get_cash_register_by_id(db, cash_id)
    if not result:
        raise ApiError("CASH_REGISTER_NOT_FOUND", "Cash register not found", http_status=404)
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CASH_REGISTER_DETAILS")


@router.put(
    "/cash-registers/{cash_id}",
    summary="ویرایش صندوق",
    description="ویرایش اطلاعات صندوق",
)
async def update_cash_register_endpoint(
    request: Request,
    cash_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("cash", "edit", CashRegister, "cash_id")),
):
    payload: Dict[str, Any] = dict(body or {})
    result = update_cash_register(db, cash_id, payload)
    if result is None:
        raise ApiError("CASH_REGISTER_NOT_FOUND", "Cash register not found", http_status=404)
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CASH_REGISTER_UPDATED")


@router.delete(
    "/cash-registers/{cash_id}",
    summary="حذف صندوق",
    description="حذف یک صندوق",
)
async def delete_cash_register_endpoint(
    request: Request,
    cash_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("cash", "delete", CashRegister, "cash_id")),
):
    result = get_cash_register_by_id(db, cash_id)
    if result:
        try:
            biz_id = int(result.get("business_id"))
        except Exception:
            biz_id = None
        if biz_id is not None and not ctx.can_access_business(biz_id):
            raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    from fastapi import HTTPException
    
    success, error_message = delete_cash_register(db, cash_id)
    if not success:
        if error_message:
            raise HTTPException(status_code=400, detail=error_message)
        raise ApiError("CASH_REGISTER_NOT_FOUND", "Cash register not found", http_status=404)
    return success_response(data=None, request=request, message="CASH_REGISTER_DELETED")


@router.post(
    "/businesses/{business_id}/cash-registers/bulk-delete",
    summary="حذف گروهی صندوق‌ها",
    description="حذف چندین صندوق بر اساس شناسه‌ها",
)
@require_business_access("business_id")
async def bulk_delete_cash_registers_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("cash", "delete")),
):
    ids = body.get("ids")
    if not isinstance(ids, list):
        ids = []
    try:
        ids = [int(x) for x in ids if isinstance(x, (int, str)) and str(x).isdigit()]
    except Exception:
        ids = []
    if not ids:
        return success_response({"deleted": 0, "skipped": 0, "total_requested": 0}, request, message="NO_VALID_IDS_FOR_DELETE")
    result = bulk_delete_cash_registers(db, business_id, ids)
    return success_response(result, request, message="CASH_REGISTERS_BULK_DELETE_DONE")


@router.post(
    "/businesses/{business_id}/cash-registers/export/excel",
    summary="خروجی Excel لیست صندوق‌ها",
    description="خروجی Excel لیست صندوق‌ها با قابلیت فیلتر و مرتب‌سازی",
)
@require_business_access("business_id")
async def export_cash_registers_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from app.core.i18n import negotiate_locale

    query_dict = {
        "take": int(body.get("take", 1000)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }
    result = list_cash_registers(db, business_id, query_dict)
    items: List[Dict[str, Any]] = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Map currency_id -> currency title for display
    try:
        from adapters.db.models.currency import Currency
        currency_ids = set()
        for it in items:
            cid = it.get("currency_id")
            try:
                if cid is not None:
                    currency_ids.add(int(cid))
            except Exception:
                pass
        currency_map: Dict[int, str] = {}
        if currency_ids:
            rows = db.query(Currency).filter(Currency.id.in_(list(currency_ids))).all()
            currency_map = {c.id: (c.title or c.code or str(c.id)) for c in rows}
        for it in items:
            cid = it.get("currency_id")
            it["currency"] = currency_map.get(cid, cid)
    except Exception:
        # In case of any issue, fallback without blocking export
        for it in items:
            if "currency" not in it and "currency_id" in it:
                it["currency"] = it.get("currency_id")

    headers: List[str] = [
        "code", "name", "currency", "is_active", "is_default",
        "payment_switch_number", "payment_terminal_number", "merchant_id",
        "description",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "CashRegisters"

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(headers, 1):
            value = item.get(key, "")
            if key in ("is_active", "is_default"):
                try:
                    truthy = bool(value) if isinstance(value, bool) else str(value).strip().lower() in ("true", "1", "yes", "y", "t")
                except Exception:
                    truthy = bool(value)
                value = "✓" if truthy else "✗"
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=cash_registers.xlsx",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/cash-registers/export/pdf",
    summary="خروجی PDF لیست صندوق‌ها",
    description="خروجی PDF لیست صندوق‌ها با قابلیت فیلتر و مرتب‌سازی",
)
@require_business_access("business_id")
async def export_cash_registers_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from fastapi.responses import Response
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale

    query_dict = {
        "take": int(body.get("take", 1000)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }

    result = list_cash_registers(db, business_id, query_dict)
    items: List[Dict[str, Any]] = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Map currency_id -> currency title for display
    try:
        from adapters.db.models.currency import Currency
        currency_ids = set()
        for it in items:
            cid = it.get("currency_id")
            try:
                if cid is not None:
                    currency_ids.add(int(cid))
            except Exception:
                pass
        currency_map: Dict[int, str] = {}
        if currency_ids:
            rows = db.query(Currency).filter(Currency.id.in_(list(currency_ids))).all()
            currency_map = {c.id: (c.title or c.code or str(c.id)) for c in rows}
        for it in items:
            cid = it.get("currency_id")
            it["currency"] = currency_map.get(cid, cid)
    except Exception:
        for it in items:
            if "currency" not in it and "currency_id" in it:
                it["currency"] = it.get("currency_id")

    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            import json
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                # Replace currency_id key with currency to show human-readable value
                if str(key) == 'currency_id':
                    keys.append('currency')
                else:
                    keys.append(str(key))
                headers.append(str(label))
    else:
        keys = [
            "code", "name", "currency", "is_active", "is_default",
            "payment_switch_number", "payment_terminal_number", "merchant_id",
            "description",
        ]
        headers = keys

    business_name = ""
    try:
        from adapters.db.models.business import Business
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name or ""
    except Exception:
        business_name = ""

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'

    try:
        from app.core.calendar import CalendarConverter
        calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
        formatted_now = CalendarConverter.format_datetime(
            __import__("datetime").datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian",
        )
        now_str = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        from datetime import datetime
        now_str = datetime.now().strftime('%Y/%m/%d %H:%M')

    def esc(v: Any) -> str:
        try:
            return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        except Exception:
            return str(v)

    rows_html: List[str] = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key)
            if value is None:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif key in ("is_active", "is_default"):
                try:
                    truthy = bool(value) if isinstance(value, bool) else str(value).strip().lower() in ("true", "1", "yes", "y", "t")
                except Exception:
                    truthy = bool(value)
                value = "✓" if truthy else "✗"
            tds.append(f"<td>{esc(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = ''.join(f"<th>{esc(h)}</th>" for h in headers)

    # تلاش برای رندر با قالب سفارشی (cash_registers/list)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": 'گزارش صندوق‌ها' if is_fa else 'Cash Registers Report',
            "business_name": business_name,
            "generated_at": now_str,
            "is_fa": is_fa,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="cash_registers",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # Inject Persian fonts (YekanBakhFaNum/Vazirmatn) for PDF rendering
    fa_font_url_regular = ""
    fa_font_url_bold = ""
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_reg, fa_bold = load_farsi_font_data_uris()
            fa_font_url_regular = fa_reg or ""
            fa_font_url_bold = fa_bold or ""
    except Exception:
        fa_font_url_regular = ""
        fa_font_url_bold = ""
    
    font_face_css = ""
    if is_fa and fa_font_url_regular:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_regular}') format('truetype'); font-weight: 400; font-style: normal; }}
        """
    if is_fa and fa_font_url_bold:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_bold}') format('truetype'); font-weight: 700; font-style: normal; }}
        """
    
    body_font_family = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if is_fa else "Arial, sans-serif"
    
    table_html = f"""
    <html lang="{html_lang}" dir="{html_dir}"> 
      <head>
        <meta charset='utf-8'>
        <style>
          @page {{ size: A4 landscape; margin: 12mm; }}
          {font_face_css}
          body {{ font-family: {body_font_family}; font-size: 11px; color: #222; }}
          .title {{ font-size: 16px; font-weight: 700; margin-bottom: 10px; }}
          table.report-table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
          thead th {{ background: #f0f3f7; border: 1px solid #c7cdd6; padding: 6px 4px; text-align: center; white-space: nowrap; }}
          tbody td {{ border: 1px solid #d7dde6; padding: 5px 4px; vertical-align: top; overflow-wrap: anywhere; }}
        </style>
      </head>
      <body>
        <div class="title">{esc('گزارش صندوق‌ها' if is_fa else 'Cash Registers Report')}</div>
        <div style="margin-bottom:6px;">{esc('نام کسب‌وکار' if is_fa else 'Business Name')}: {esc(business_name)} | {esc('تاریخ گزارش' if is_fa else 'Report Date')}: {esc(now_str)}</div>
        <table class="report-table"> 
          <thead><tr>{headers_html}</tr></thead>
          <tbody>{''.join(rows_html)}</tbody>
        </table>
      </body>
    </html>
    """

    final_html = resolved_html or table_html
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=cash_registers.pdf",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/reports/cash-petty-turnover",
    summary="گزارش گردش صندوق و تنخواه",
    description="گزارش برداشت‌ها و واریزهای هر صندوق و تنخواه در یک بازه زمانی",
)
@require_business_access("business_id")
async def cash_petty_turnover_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش گردش صندوق و تنخواه"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        from sqlalchemy import and_
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    cash_register_ids = body.get('cash_register_ids')
    if cash_register_ids is not None and not isinstance(cash_register_ids, list):
        cash_register_ids = None
    
    petty_cash_ids = body.get('petty_cash_ids')
    if petty_cash_ids is not None and not isinstance(petty_cash_ids, list):
        petty_cash_ids = None
    
    search = body.get('search')
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_cash_petty_turnover_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        cash_register_ids=cash_register_ids,
        petty_cash_ids=petty_cash_ids,
        search=search,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    from app.core.i18n import negotiate_locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Cash and petty cash turnover report retrieved successfully" if locale != 'fa' else "گزارش گردش صندوق و تنخواه با موفقیت دریافت شد",
        request=request
    )


@router.post("/businesses/{business_id}/reports/cash-petty-turnover/export/excel",
    summary="خروجی Excel گزارش گردش صندوق و تنخواه",
    description="خروجی Excel گزارش گردش صندوق و تنخواه با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_cash_petty_turnover_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel گزارش گردش صندوق و تنخواه"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    from app.core.i18n import negotiate_locale
    
    max_export_records = 10000
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        from sqlalchemy import and_
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    cash_register_ids = body.get('cash_register_ids')
    if cash_register_ids is not None and not isinstance(cash_register_ids, list):
        cash_register_ids = None
    
    petty_cash_ids = body.get('petty_cash_ids')
    if petty_cash_ids is not None and not isinstance(petty_cash_ids, list):
        petty_cash_ids = None
    
    search = body.get('search')
    
    # دریافت همه رکوردها برای export
    result = get_cash_petty_turnover_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        cash_register_ids=cash_register_ids,
        petty_cash_ids=petty_cash_ids,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        import datetime as dt_module
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, dt_module.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, dt_module.date):
            try:
                dt_value = dt_module.datetime.combine(value, dt_module.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = dt_module.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = dt_module.date.fromisoformat(value)
                            dt_value = dt_module.datetime.combine(date_value, dt_module.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = dt_module.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = dt_module.date.fromisoformat(value)
                        dt_value = dt_module.datetime.combine(date_value, dt_module.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    # تعریف ستون‌ها
    columns = [
        ('document_date', 'تاریخ' if is_fa else 'Date'),
        ('document_type_name', 'نوع سند' if is_fa else 'Document Type'),
        ('document_code', 'شماره سند' if is_fa else 'Document Code'),
        ('source_type', 'نوع' if is_fa else 'Type'),
        ('source_code', 'کد' if is_fa else 'Code'),
        ('source_name', 'نام' if is_fa else 'Name'),
        ('deposit', 'واریز' if is_fa else 'Deposit'),
        ('withdrawal', 'برداشت' if is_fa else 'Withdrawal'),
        ('balance', 'مانده' if is_fa else 'Balance'),
        ('description', 'توضیحات' if is_fa else 'Description'),
    ]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CashPettyTurnover"
    
    # RTL برای فارسی
    if is_fa:
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Header
    for col_idx, (key, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            value = item.get(key, '')
            
            # Format dates
            if key == 'document_date' and value:
                value = format_date_for_export(item, 'document_date')
            
            if key == 'source_type':
                # تبدیل source_type به نام فارسی
                if value == 'cash_register':
                    value = 'صندوق' if is_fa else 'Cash Register'
                elif value == 'petty_cash':
                    value = 'تنخواه' if is_fa else 'Petty Cash'
            elif isinstance(value, (int, float)):
                if key in ('deposit', 'withdrawal', 'balance'):
                    # Format numbers with thousand separators
                    value = f"{value:,.2f}"
            elif value is None:
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if is_fa:
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(value, str) and value.replace(',', '').replace('.', '').isdigit():
                cell.alignment = Alignment(horizontal="right")
    
    # Auto-width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    content = buffer.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=cash_petty_turnover.xlsx",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
