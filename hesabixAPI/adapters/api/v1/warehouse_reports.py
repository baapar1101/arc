"""
API endpoints for warehouse reports
"""
from typing import Dict, Any, List, Optional
from fastapi import APIRouter, Depends, Request, Body, Response
from sqlalchemy.orm import Session
import io
import datetime
import re

from adapters.db.session import get_db
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.permissions import require_business_access, require_business_permission_dep
from app.core.responses import success_response, format_datetime_fields
from app.core.i18n import negotiate_locale
from app.services.warehouse_reports_service import (
    get_warehouse_documents_summary_report,
    get_slow_moving_items_report,
    get_critical_stock_report,
    get_inter_warehouse_transfers_report,
    get_adjustment_documents_report,
    get_warehouse_performance_report,
    get_product_movement_history_report,
    get_inventory_valuation_report,
    get_pending_documents_report,
    get_inventory_turnover_report,
)

router = APIRouter(prefix="/warehouse-reports", tags=["warehouse_reports"])


def _create_excel_export(items: List[Dict[str, Any]], headers: List[tuple], filename_base: str, request: Request, db: Session, business_id: int) -> Response:
    """تابع کمکی برای ساخت فایل Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from adapters.db.models.business import Business
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    wb = Workbook()
    ws = wb.active
    ws.title = filename_base[:31]  # Excel sheet name limit
    
    # RTL for Persian
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    header_labels = [h[1] for h in headers]
    ws.append(header_labels)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Add data rows
    keys = [h[0] for h in headers]
    for it in items:
        row = []
        for k in keys:
            value = it.get(k)
            if value is None:
                row.append('')
            elif isinstance(value, (int, float)):
                row.append(float(value))
            elif isinstance(value, (dict, list)):
                row.append(str(value))
            else:
                row.append(str(value))
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")
    
    # Auto width columns
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass
    
    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()
    
    # Build filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = filename_base
    if biz_name:
        base += f"_{slugify(biz_name)}"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(data)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/documents-summary",
    summary="گزارش خلاصه حواله‌های انبار",
    description="خلاصه حواله‌های انبار به تفکیک نوع با آمار ورود و خروج",
)
@require_business_access("business_id")
async def warehouse_documents_summary_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش خلاصه حواله‌های انبار"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    doc_types = body.get("doc_types")
    warehouse_ids = body.get("warehouse_ids")
    status = body.get("status")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_warehouse_documents_summary_report(
        db=db,
        business_id=business_id,
        date_from=date_from,
        date_to=date_to,
        doc_types=doc_types,
        warehouse_ids=warehouse_ids,
        status=status,
        skip=skip,
        take=take,
    )
    
    result["items"] = [format_datetime_fields(item, request) for item in result["items"]]
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/documents-summary/export/excel",
    summary="خروجی Excel گزارش خلاصه حواله‌های انبار",
)
@require_business_access("business_id")
async def export_warehouse_documents_summary_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش خلاصه حواله‌های انبار"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_warehouse_documents_summary_report(
        db=db,
        business_id=business_id,
        date_from=body.get("date_from"),
        date_to=body.get("date_to"),
        doc_types=body.get("doc_types"),
        warehouse_ids=body.get("warehouse_ids"),
        status=body.get("status"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('code', 'کد حواله' if is_fa else 'Document Code'),
        ('document_date', 'تاریخ' if is_fa else 'Date'),
        ('doc_type', 'نوع حواله' if is_fa else 'Document Type'),
        ('items_count', 'تعداد اقلام' if is_fa else 'Items Count'),
        ('total_quantity', 'مقدار کل' if is_fa else 'Total Quantity'),
    ]
    
    return _create_excel_export(items, headers, "warehouse_documents_summary", request, db, business_id)


@router.post("/businesses/{business_id}/slow-moving-items",
    summary="گزارش کالاهای کم‌گردش",
    description="کالاهایی که در بازه زمانی مشخص شده هیچ حرکتی نداشته‌اند",
)
@require_business_access("business_id")
async def slow_moving_items_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش کالاهای کم‌گردش"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    days_without_movement = int(body.get("days_without_movement", 90))
    warehouse_ids = body.get("warehouse_ids")
    category_ids = body.get("category_ids")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_slow_moving_items_report(
        db=db,
        business_id=business_id,
        days_without_movement=days_without_movement,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        skip=skip,
        take=take,
    )
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/slow-moving-items/export/excel",
    summary="خروجی Excel گزارش کالاهای کم‌گردش",
)
@require_business_access("business_id")
async def export_slow_moving_items_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش کالاهای کم‌گردش"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_slow_moving_items_report(
        db=db,
        business_id=business_id,
        days_without_movement=int(body.get("days_without_movement", 90)),
        warehouse_ids=body.get("warehouse_ids"),
        category_ids=body.get("category_ids"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('product_code', 'کد محصول' if is_fa else 'Product Code'),
        ('product_name', 'نام محصول' if is_fa else 'Product Name'),
        ('current_stock', 'موجودی فعلی' if is_fa else 'Current Stock'),
        ('unit', 'واحد' if is_fa else 'Unit'),
        ('days_without_movement', 'روز بدون حرکت' if is_fa else 'Days Without Movement'),
    ]
    
    return _create_excel_export(items, headers, "slow_moving_items", request, db, business_id)


@router.post("/businesses/{business_id}/critical-stock",
    summary="گزارش کالاهای با موجودی بحرانی",
    description="کالاهایی که موجودی آن‌ها کمتر از حد تعیین شده است",
)
@require_business_access("business_id")
async def critical_stock_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش کالاهای با موجودی بحرانی"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    warehouse_ids = body.get("warehouse_ids")
    category_ids = body.get("category_ids")
    as_of_date = body.get("as_of_date")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_critical_stock_report(
        db=db,
        business_id=business_id,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        as_of_date=as_of_date,
        skip=skip,
        take=take,
    )
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/critical-stock/export/excel",
    summary="خروجی Excel گزارش کالاهای با موجودی بحرانی",
)
@require_business_access("business_id")
async def export_critical_stock_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش کالاهای با موجودی بحرانی"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_critical_stock_report(
        db=db,
        business_id=business_id,
        warehouse_ids=body.get("warehouse_ids"),
        category_ids=body.get("category_ids"),
        as_of_date=body.get("as_of_date"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('product_code', 'کد محصول' if is_fa else 'Product Code'),
        ('product_name', 'نام محصول' if is_fa else 'Product Name'),
        ('warehouse_name', 'انبار' if is_fa else 'Warehouse'),
        ('current_stock', 'موجودی فعلی' if is_fa else 'Current Stock'),
        ('min_stock', 'حداقل موجودی' if is_fa else 'Min Stock'),
        ('difference', 'تفاوت' if is_fa else 'Difference'),
        ('unit', 'واحد' if is_fa else 'Unit'),
    ]
    
    return _create_excel_export(items, headers, "critical_stock", request, db, business_id)


@router.post("/businesses/{business_id}/inter-warehouse-transfers",
    summary="گزارش انتقالات بین انبارها",
    description="جزئیات انتقالات بین انبارها",
)
@require_business_access("business_id")
async def inter_warehouse_transfers_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش انتقالات بین انبارها"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    warehouse_from_ids = body.get("warehouse_from_ids")
    warehouse_to_ids = body.get("warehouse_to_ids")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_inter_warehouse_transfers_report(
        db=db,
        business_id=business_id,
        date_from=date_from,
        date_to=date_to,
        warehouse_from_ids=warehouse_from_ids,
        warehouse_to_ids=warehouse_to_ids,
        skip=skip,
        take=take,
    )
    
    result["items"] = [format_datetime_fields(item, request) for item in result["items"]]
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/inter-warehouse-transfers/export/excel",
    summary="خروجی Excel گزارش انتقالات بین انبارها",
)
@require_business_access("business_id")
async def export_inter_warehouse_transfers_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش انتقالات بین انبارها"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_inter_warehouse_transfers_report(
        db=db,
        business_id=business_id,
        date_from=body.get("date_from"),
        date_to=body.get("date_to"),
        warehouse_from_ids=body.get("warehouse_from_ids"),
        warehouse_to_ids=body.get("warehouse_to_ids"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('code', 'کد حواله' if is_fa else 'Document Code'),
        ('document_date', 'تاریخ' if is_fa else 'Date'),
        ('warehouse_from_name', 'انبار مبدا' if is_fa else 'From Warehouse'),
        ('warehouse_to_name', 'انبار مقصد' if is_fa else 'To Warehouse'),
        ('items_count', 'تعداد اقلام' if is_fa else 'Items Count'),
        ('total_quantity', 'مقدار کل' if is_fa else 'Total Quantity'),
    ]
    
    return _create_excel_export(items, headers, "inter_warehouse_transfers", request, db, business_id)


@router.post("/businesses/{business_id}/adjustment-documents",
    summary="گزارش حواله‌های تعدیل",
    description="تحلیل حواله‌های تعدیل و تفاوت‌های موجودی",
)
@require_business_access("business_id")
async def adjustment_documents_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش حواله‌های تعدیل"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    warehouse_ids = body.get("warehouse_ids")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_adjustment_documents_report(
        db=db,
        business_id=business_id,
        date_from=date_from,
        date_to=date_to,
        warehouse_ids=warehouse_ids,
        skip=skip,
        take=take,
    )
    
    result["items"] = [format_datetime_fields(item, request) for item in result["items"]]
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/adjustment-documents/export/excel",
    summary="خروجی Excel گزارش حواله‌های تعدیل",
)
@require_business_access("business_id")
async def export_adjustment_documents_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش حواله‌های تعدیل"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_adjustment_documents_report(
        db=db,
        business_id=business_id,
        date_from=body.get("date_from"),
        date_to=body.get("date_to"),
        warehouse_ids=body.get("warehouse_ids"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('code', 'کد حواله' if is_fa else 'Document Code'),
        ('document_date', 'تاریخ' if is_fa else 'Date'),
        ('items_count', 'تعداد اقلام' if is_fa else 'Items Count'),
        ('quantity_increase', 'افزایش' if is_fa else 'Increase'),
        ('quantity_decrease', 'کاهش' if is_fa else 'Decrease'),
        ('net_adjustment', 'خالص تعدیل' if is_fa else 'Net Adjustment'),
    ]
    
    return _create_excel_export(items, headers, "adjustment_documents", request, db, business_id)


@router.post("/businesses/{business_id}/warehouse-performance",
    summary="گزارش عملکرد انبارها",
    description="مقایسه عملکرد انبارها",
)
@require_business_access("business_id")
async def warehouse_performance_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش عملکرد انبارها"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    warehouse_ids = body.get("warehouse_ids")
    
    result = get_warehouse_performance_report(
        db=db,
        business_id=business_id,
        date_from=date_from,
        date_to=date_to,
        warehouse_ids=warehouse_ids,
    )
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/warehouse-performance/export/excel",
    summary="خروجی Excel گزارش عملکرد انبارها",
)
@require_business_access("business_id")
async def export_warehouse_performance_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش عملکرد انبارها"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_warehouse_performance_report(
        db=db,
        business_id=business_id,
        date_from=body.get("date_from"),
        date_to=body.get("date_to"),
        warehouse_ids=body.get("warehouse_ids"),
    )
    
    items = result.get("items", [])
    headers = [
        ('warehouse_code', 'کد انبار' if is_fa else 'Warehouse Code'),
        ('warehouse_name', 'نام انبار' if is_fa else 'Warehouse Name'),
        ('total_documents', 'تعداد حواله‌ها' if is_fa else 'Total Documents'),
        ('total_items', 'تعداد اقلام' if is_fa else 'Total Items'),
        ('total_quantity_in', 'کل ورود' if is_fa else 'Total In'),
        ('total_quantity_out', 'کل خروج' if is_fa else 'Total Out'),
        ('net_quantity', 'خالص' if is_fa else 'Net'),
    ]
    
    return _create_excel_export(items, headers, "warehouse_performance", request, db, business_id)


@router.post("/businesses/{business_id}/product-movement-history",
    summary="گزارش تاریخچه حرکات یک کالا",
    description="تاریخچه کامل حرکات یک کالا در تمام انبارها",
)
@require_business_access("business_id")
async def product_movement_history_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش تاریخچه حرکات یک کالا"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    product_id = body.get("product_id")
    if not product_id:
        from app.core.responses import ApiError
        raise ApiError("PRODUCT_ID_REQUIRED", "product_id is required", http_status=400)
    
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    warehouse_ids = body.get("warehouse_ids")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_product_movement_history_report(
        db=db,
        business_id=business_id,
        product_id=int(product_id),
        date_from=date_from,
        date_to=date_to,
        warehouse_ids=warehouse_ids,
        skip=skip,
        take=take,
    )
    
    result["items"] = [format_datetime_fields(item, request) for item in result["items"]]
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/product-movement-history/export/excel",
    summary="خروجی Excel گزارش تاریخچه حرکات یک کالا",
)
@require_business_access("business_id")
async def export_product_movement_history_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش تاریخچه حرکات یک کالا"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    product_id = body.get("product_id")
    if not product_id:
        from app.core.responses import ApiError
        raise ApiError("PRODUCT_ID_REQUIRED", "product_id is required", http_status=400)
    
    result = get_product_movement_history_report(
        db=db,
        business_id=business_id,
        product_id=int(product_id),
        date_from=body.get("date_from"),
        date_to=body.get("date_to"),
        warehouse_ids=body.get("warehouse_ids"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('document_code', 'کد حواله' if is_fa else 'Document Code'),
        ('document_date', 'تاریخ' if is_fa else 'Date'),
        ('doc_type', 'نوع حواله' if is_fa else 'Document Type'),
        ('warehouse_name', 'انبار' if is_fa else 'Warehouse'),
        ('movement', 'نوع حرکت' if is_fa else 'Movement'),
        ('quantity', 'مقدار' if is_fa else 'Quantity'),
    ]
    
    return _create_excel_export(items, headers, "product_movement_history", request, db, business_id)


@router.post("/businesses/{business_id}/inventory-valuation",
    summary="گزارش ارزش موجودی انبار",
    description="ارزش ریالی موجودی انبارها",
)
@require_business_access("business_id")
async def inventory_valuation_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش ارزش موجودی انبار"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    as_of_date = body.get("as_of_date")
    warehouse_ids = body.get("warehouse_ids")
    category_ids = body.get("category_ids")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_inventory_valuation_report(
        db=db,
        business_id=business_id,
        as_of_date=as_of_date,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        skip=skip,
        take=take,
    )
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/inventory-valuation/export/excel",
    summary="خروجی Excel گزارش ارزش موجودی انبار",
)
@require_business_access("business_id")
async def export_inventory_valuation_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش ارزش موجودی انبار"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_inventory_valuation_report(
        db=db,
        business_id=business_id,
        as_of_date=body.get("as_of_date"),
        warehouse_ids=body.get("warehouse_ids"),
        category_ids=body.get("category_ids"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('product_code', 'کد محصول' if is_fa else 'Product Code'),
        ('product_name', 'نام محصول' if is_fa else 'Product Name'),
        ('warehouse_name', 'انبار' if is_fa else 'Warehouse'),
        ('quantity', 'موجودی' if is_fa else 'Quantity'),
        ('unit', 'واحد' if is_fa else 'Unit'),
        ('cost_price', 'قیمت تمام شده' if is_fa else 'Cost Price'),
        ('value', 'ارزش' if is_fa else 'Value'),
    ]
    
    return _create_excel_export(items, headers, "inventory_valuation", request, db, business_id)


@router.post("/businesses/{business_id}/pending-documents",
    summary="گزارش حواله‌های در انتظار تایید",
    description="حواله‌های draft یا در انتظار تایید",
)
@require_business_access("business_id")
async def pending_documents_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش حواله‌های در انتظار تایید"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    warehouse_ids = body.get("warehouse_ids")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_pending_documents_report(
        db=db,
        business_id=business_id,
        warehouse_ids=warehouse_ids,
        skip=skip,
        take=take,
    )
    
    result["items"] = [format_datetime_fields(item, request) for item in result["items"]]
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/pending-documents/export/excel",
    summary="خروجی Excel گزارش حواله‌های در انتظار تایید",
)
@require_business_access("business_id")
async def export_pending_documents_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش حواله‌های در انتظار تایید"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_pending_documents_report(
        db=db,
        business_id=business_id,
        warehouse_ids=body.get("warehouse_ids"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('code', 'کد حواله' if is_fa else 'Document Code'),
        ('document_date', 'تاریخ' if is_fa else 'Date'),
        ('doc_type', 'نوع حواله' if is_fa else 'Document Type'),
        ('items_count', 'تعداد اقلام' if is_fa else 'Items Count'),
        ('days_pending', 'روز انتظار' if is_fa else 'Days Pending'),
    ]
    
    return _create_excel_export(items, headers, "pending_documents", request, db, business_id)


@router.post("/businesses/{business_id}/inventory-turnover",
    summary="گزارش گردش موجودی",
    description="نرخ گردش موجودی کالاها",
)
@require_business_access("business_id")
async def inventory_turnover_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش گردش موجودی"""
    if not ctx.can_read_section("reports"):
        from app.core.responses import ApiError
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    warehouse_ids = body.get("warehouse_ids")
    category_ids = body.get("category_ids")
    
    skip = int(body.get("skip", 0))
    take = int(body.get("take", 50))
    
    result = get_inventory_turnover_report(
        db=db,
        business_id=business_id,
        date_from=date_from,
        date_to=date_to,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        skip=skip,
        take=take,
    )
    
    return success_response(data=result, request=request)


@router.post("/businesses/{business_id}/inventory-turnover/export/excel",
    summary="خروجی Excel گزارش گردش موجودی",
)
@require_business_access("business_id")
async def export_inventory_turnover_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
) -> Response:
    """خروجی Excel گزارش گردش موجودی"""
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    
    result = get_inventory_turnover_report(
        db=db,
        business_id=business_id,
        date_from=body.get("date_from"),
        date_to=body.get("date_to"),
        warehouse_ids=body.get("warehouse_ids"),
        category_ids=body.get("category_ids"),
        skip=0,
        take=10000,
    )
    
    items = result.get("items", [])
    headers = [
        ('product_code', 'کد محصول' if is_fa else 'Product Code'),
        ('product_name', 'نام محصول' if is_fa else 'Product Name'),
        ('average_stock', 'موجودی متوسط' if is_fa else 'Average Stock'),
        ('total_out', 'کل خروج' if is_fa else 'Total Out'),
        ('turnover_rate', 'نرخ گردش' if is_fa else 'Turnover Rate'),
        ('unit', 'واحد' if is_fa else 'Unit'),
    ]
    
    return _create_excel_export(items, headers, "inventory_turnover", request, db, business_id)

