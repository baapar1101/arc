from typing import Any, Dict
from fastapi import APIRouter, Depends, Request, Body
from sqlalchemy.orm import Session

from adapters.db.session import get_db
from adapters.db.models.check import Check
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.permissions import require_business_management_dep, require_business_access, require_business_permission_dep, require_business_permission_by_entity_dep
from adapters.api.v1.schemas import QueryInfo
from adapters.api.v1.list_query_common import DocumentListQueryBody, document_list_query_to_dict
from adapters.api.v1.schema_models.check import (
    CheckCreateRequest,
    CheckUpdateRequest,
    CheckEndorseRequest,
    CheckClearRequest,
    CheckReturnRequest,
    CheckBounceRequest,
    CheckPayRequest,
    CheckDepositRequest,
    CheckReconciliationCalculateRequest,
    CheckReconciliationCreateRequest,
)
from app.services.check_service import (
    create_check,
    update_check,
    delete_check,
    get_check_by_id,
    list_checks,
    endorse_check,
    clear_check,
    return_check,
    bounce_check,
    pay_check,
    deposit_check,
    get_check_history_and_documents,
)
from app.services.check_reconciliation_service import (
    calculate_checks_reconciliation,
    create_reconciliation,
    get_reconciliation_by_id,
    list_reconciliations,
    delete_reconciliation,
)


router = APIRouter(prefix="/checks", tags=["مدیریت مالی", "دریافت و پرداخت"])


@router.post(
    "/businesses/{business_id}/checks",
    summary="لیست چک‌های کسب‌وکار",
    description="دریافت لیست چک‌ها با جستجو/فیلتر",
)
@require_business_access("business_id")
async def list_checks_endpoint(
    request: Request,
    business_id: int,
    body: DocumentListQueryBody,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    query_dict = document_list_query_to_dict(body, request=request, fiscal_year_from_header=False)
    if query_dict.get("person_id") is None and request.query_params.get("person_id"):
        try:
            query_dict["person_id"] = int(request.query_params.get("person_id"))
        except (TypeError, ValueError):
            pass
    result = list_checks(db, business_id, query_dict)
    result["items"] = [format_datetime_fields(item, request) for item in result.get("items", [])]
    return success_response(data=result, request=request, message="CHECKS_LIST_FETCHED")


@router.post(
    "/businesses/{business_id}/checks/create",
    summary="ایجاد چک",
    description="ایجاد چک جدید برای کسب‌وکار",
)
@require_business_access("business_id")
async def create_check_endpoint(
    request: Request,
    business_id: int,
    body: CheckCreateRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("checks", "add")),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    # ثبت سند حسابداری الزامی است - بررسی دسترسی نوشتن حسابداری
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check creation)", http_status=403)
    created = create_check(db, business_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(created, request), request=request, message="CHECK_CREATED")
@router.post(
    "/checks/{check_id}/actions/endorse",
    summary="واگذاری چک دریافتی به شخص",
    description="واگذاری چک دریافتی به شخص دیگر",
)
async def endorse_check_endpoint(
    request: Request,
    check_id: int,
    body: CheckEndorseRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    # access check
    before = get_check_by_id(db, check_id)
    if not before:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(before.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    # ثبت سند حسابداری الزامی است - بررسی دسترسی نوشتن حسابداری
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check actions)", http_status=403)
    result = endorse_check(db, check_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_ENDORSED")


@router.post(
    "/checks/{check_id}/actions/clear",
    summary="وصول/پاس چک",
    description="انتقال حساب چک به بانک در زمان پاس/وصول",
)
async def clear_check_endpoint(
    request: Request,
    check_id: int,
    body: CheckClearRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    before = get_check_by_id(db, check_id)
    if not before:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(before.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    # ثبت سند حسابداری الزامی است - بررسی دسترسی نوشتن حسابداری
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check actions)", http_status=403)
    result = clear_check(db, check_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_CLEARED")


@router.post(
    "/checks/{check_id}/actions/return",
    summary="عودت چک",
    description="عودت چک: from_endorsee (برگشت از واگذارشونده) یا to_drawer (عودت به صادرکننده)",
)
async def return_check_endpoint(
    request: Request,
    check_id: int,
    body: CheckReturnRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    before = get_check_by_id(db, check_id)
    if not before:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(before.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    # ثبت سند حسابداری الزامی است - بررسی دسترسی نوشتن حسابداری
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check actions)", http_status=403)
    result = return_check(db, check_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_RETURNED")


@router.post(
    "/checks/{check_id}/actions/bounce",
    summary="برگشت چک",
    description="برگشت چک و ثبت هزینه احتمالی",
)
async def bounce_check_endpoint(
    request: Request,
    check_id: int,
    body: CheckBounceRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    before = get_check_by_id(db, check_id)
    if not before:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(before.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    # ثبت سند حسابداری الزامی است - بررسی دسترسی نوشتن حسابداری
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check actions)", http_status=403)
    result = bounce_check(db, check_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_BOUNCED")


@router.post(
    "/checks/{check_id}/actions/pay",
    summary="پرداخت چک پرداختنی",
    description="پاس چک پرداختنی از بانک",
)
async def pay_check_endpoint(
    request: Request,
    check_id: int,
    body: CheckPayRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    before = get_check_by_id(db, check_id)
    if not before:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(before.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    # ثبت سند حسابداری الزامی است - بررسی دسترسی نوشتن حسابداری
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check actions)", http_status=403)
    result = pay_check(db, check_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_PAID")


@router.post(
    "/checks/{check_id}/actions/deposit",
    summary="سپرده چک به بانک (اختیاری)",
    description="انتقال به اسناد در جریان وصول",
)
async def deposit_check_endpoint(
    request: Request,
    check_id: int,
    body: CheckDepositRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    before = get_check_by_id(db, check_id)
    if not before:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(before.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    # ثبت سند حسابداری الزامی است - بررسی دسترسی نوشتن حسابداری
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check actions)", http_status=403)
    result = deposit_check(db, check_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_DEPOSITED")



@router.get(
    "/checks/{check_id}",
    summary="جزئیات چک",
    description="دریافت جزئیات چک",
)
async def get_check_endpoint(
    request: Request,
    check_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    result = get_check_by_id(db, check_id)
    if not result:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_DETAILS")


@router.get(
    "/checks/{check_id}/history",
    summary="سوابق چک و اسناد حسابداری",
    description="دریافت سوابق چک و اسناد حسابداری مرتبط",
)
async def get_check_history_endpoint(
    request: Request,
    check_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    # بررسی دسترسی
    check = get_check_by_id(db, check_id)
    if not check:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(check.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    result = get_check_history_and_documents(db, check_id)
    
    # فرمت کردن تاریخ‌ها
    if result.get("history"):
        result["history"] = [format_datetime_fields(item, request) for item in result["history"]]
    if result.get("documents"):
        result["documents"] = [format_datetime_fields(item, request) for item in result["documents"]]
    
    return success_response(data=result, request=request, message="CHECK_HISTORY_FETCHED")


@router.put(
    "/checks/{check_id}",
    summary="ویرایش چک",
    description="ویرایش اطلاعات چک",
)
async def update_check_endpoint(
    request: Request,
    check_id: int,
    body: CheckUpdateRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("checks", "edit", Check, "check_id")),
):
    # بررسی دسترسی قبل از update
    before = get_check_by_id(db, check_id)
    if not before:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    try:
        biz_id = int(before.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    result = update_check(db, check_id, payload)
    if result is None:
        raise ApiError("CHECK_NOT_FOUND", "Check not found", http_status=404)
    return success_response(data=format_datetime_fields(result, request), request=request, message="CHECK_UPDATED")


@router.delete(
    "/checks/{check_id}",
    summary="حذف چک",
    description="حذف یک چک",
)
async def delete_check_endpoint(
    request: Request,
    check_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("checks", "delete", Check, "check_id")),
):
    result = get_check_by_id(db, check_id)
    if result:
        try:
            biz_id = int(result.get("business_id"))
        except Exception:
            biz_id = None
        if biz_id is not None and not ctx.can_access_business(biz_id):
            raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    # بررسی دسترسی حسابداری برای حذف اسناد مرتبط
    if not ctx.has_any_permission("accounting", "write"):
        raise ApiError("FORBIDDEN", "Missing permission: accounting.write (required for check deletion)", http_status=403)
    delete_check(db, check_id, user_id=ctx.get_user_id())
    return success_response(data=None, request=request, message="CHECK_DELETED")


# =====================
# Reconciliation Endpoints
# =====================

@router.post(
    "/businesses/{business_id}/checks/reconciliations/calculate",
    summary="محاسبه راس چک‌ها",
    description="محاسبه راس چک‌ها بدون ذخیره",
)
@require_business_access("business_id")
async def calculate_reconciliation_endpoint(
    request: Request,
    business_id: int,
    body: CheckReconciliationCalculateRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    result = calculate_checks_reconciliation(
        db,
        business_id,
        payload.get("check_ids", []),
        payload.get("base_date"),
        payload.get("currency_id"),
    )
    return success_response(data=format_datetime_fields(result, request), request=request, message="RECONCILIATION_CALCULATED")


@router.post(
    "/businesses/{business_id}/checks/reconciliations",
    summary="ایجاد جلسه راس‌گیری",
    description="ایجاد و ذخیره جلسه راس‌گیری چک‌ها",
)
@require_business_access("business_id")
async def create_reconciliation_endpoint(
    request: Request,
    business_id: int,
    body: CheckReconciliationCreateRequest = Body(...),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    payload: Dict[str, Any] = body.model_dump(exclude_unset=True)
    result = create_reconciliation(db, business_id, ctx.get_user_id(), payload)
    return success_response(data=format_datetime_fields(result, request), request=request, message="RECONCILIATION_CREATED")


@router.post(
    "/businesses/{business_id}/checks/reconciliations/list",
    summary="لیست جلسات راس‌گیری",
    description="دریافت لیست جلسات راس‌گیری با جستجو/فیلتر",
)
@require_business_access("business_id")
async def list_reconciliations_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    query_dict: Dict[str, Any] = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "sort": [s.model_dump() for s in query_info.sort] if query_info.sort else None,
        "search": query_info.search,
        "search_fields": query_info.search_fields,
        "filters": query_info.filters,
    }
    result = list_reconciliations(db, business_id, query_dict)
    result["items"] = [format_datetime_fields(item, request) for item in result.get("items", [])]
    return success_response(data=result, request=request, message="RECONCILIATIONS_LIST_FETCHED")


@router.get(
    "/reconciliations/{reconciliation_id}",
    summary="جزئیات جلسه راس‌گیری",
    description="دریافت جزئیات یک جلسه راس‌گیری",
)
async def get_reconciliation_endpoint(
    request: Request,
    reconciliation_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    result = get_reconciliation_by_id(db, reconciliation_id)
    if not result:
        raise ApiError("RECONCILIATION_NOT_FOUND", "جلسه راس‌گیری پیدا نشد", http_status=404)
    try:
        biz_id = int(result.get("business_id"))
    except Exception:
        biz_id = None
    if biz_id is not None and not ctx.can_access_business(biz_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="RECONCILIATION_DETAILS")


@router.delete(
    "/reconciliations/{reconciliation_id}",
    summary="حذف جلسه راس‌گیری",
    description="حذف یک جلسه راس‌گیری",
)
async def delete_reconciliation_endpoint(
    request: Request,
    reconciliation_id: int,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    result = get_reconciliation_by_id(db, reconciliation_id)
    if result:
        try:
            biz_id = int(result.get("business_id"))
        except Exception:
            biz_id = None
        if biz_id is not None and not ctx.can_access_business(biz_id):
            raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    delete_reconciliation(db, reconciliation_id)
    return success_response(data=None, request=request, message="RECONCILIATION_DELETED")


# =====================
# Export Endpoints
# =====================

@router.post(
    "/businesses/{business_id}/checks/export/excel",
    summary="خروجی Excel لیست چک‌ها",
    description="خروجی Excel لیست چک‌ها با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_checks_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel لیست چک‌ها"""
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import Response
    from app.core.i18n import negotiate_locale, locale_dependency, Translator
    from app.core.calendar import CalendarConverter, get_calendar_type_from_header
    import datetime as dt_module
    
    # Get locale and translator
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    translator = await locale_dependency(request)
    is_fa = locale == 'fa'
    
    # Get calendar type
    calendar_type = get_calendar_type_from_header(request.headers.get("X-Calendar-Type"))
    if not calendar_type:
        calendar_type = "jalali" if is_fa else "gregorian"
    
    # Build query dict from flat body
    max_export_records = 10000
    take_value = min(int(body.get("take", 1000)), max_export_records)
    
    query_dict = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }
    
    # Handle person_id from body
    if body.get("person_id") is not None:
        try:
            query_dict["person_id"] = int(body.get("person_id"))
        except Exception:
            pass

    result = list_checks(db, business_id, query_dict)
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            "row_number": "⚠️",
            "type": translator.t("warning", "هشدار"),
            "check_number": translator.t("max_export_limit", "حداکثر ۱۰,۰۰۰ رکورد قابل export است"),
            "person_name": "",
            "amount": "",
            "due_date": "",
            "status": "",
        }
        items.append(warning_item)

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Helper functions
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, dt_module.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, dt_module.date):
            try:
                dt_value = dt_module.datetime.combine(value, dt_module.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = dt_module.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = dt_module.date.fromisoformat(value)
                            dt_value = dt_module.datetime.combine(date_value, dt_module.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = dt_module.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = dt_module.date.fromisoformat(value)
                        dt_value = dt_module.datetime.combine(date_value, dt_module.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    def format_number_for_display(value) -> str:
        """Format number with thousands separator and remove .0"""
        try:
            if value is None:
                return ""
            v = float(value)
            s = f"{v:,.2f}"
            # Trim trailing .00 or trailing zeros
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return s
        except Exception:
            return str(value) if value is not None else ""
    
    def translate_status(status: str) -> str:
        """Translate check status"""
        if not status:
            return ""
        status_map = {
            'RECEIVED_ON_HAND': translator.t('check_status_received_on_hand', 'در دست (دریافتی)'),
            'TRANSFERRED_ISSUED': translator.t('check_status_transferred_issued', 'صادر شده (پرداختنی)'),
            'DEPOSITED': translator.t('check_status_deposited', 'سپرده به بانک'),
            'CLEARED': translator.t('check_status_cleared', 'پاس/وصول شده'),
            'ENDORSED': translator.t('check_status_endorsed', 'واگذار شده'),
            'RETURNED': translator.t('check_status_returned', 'عودت شده'),
            'BOUNCED': translator.t('check_status_bounced', 'برگشت خورده'),
            'CANCELLED': translator.t('check_status_cancelled', 'ابطال'),
        }
        return status_map.get(status.upper(), status)
    
    def translate_type(check_type: str) -> str:
        """Translate check type"""
        if not check_type:
            return ""
        type_map = {
            'received': translator.t('check_type_received', 'دریافتی'),
            'transferred': translator.t('check_type_transferred', 'واگذار شده'),
        }
        return type_map.get(check_type.lower(), check_type)

    # Prepare headers based on export_columns (add row number at the beginning)
    headers: list = []
    keys: list = []
    
    # Always add row number as first column
    headers.append(translator.t('row_number', 'ردیف'))
    keys.append('row_number')
    
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key and key != 'row_number':  # Skip row_number if already added
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns for checks
        default_columns = [
            ('type', translator.t('check_type', 'نوع')),
            ('person_name', translator.t('person', 'شخص')),
            ('issue_date', translator.t('issue_date', 'تاریخ صدور')),
            ('due_date', translator.t('due_date', 'تاریخ سررسید')),
            ('check_number', translator.t('check_number', 'شماره چک')),
            ('sayad_code', translator.t('sayad_code', 'شناسه صیاد')),
            ('bank_name', translator.t('bank', 'بانک')),
            ('branch_name', translator.t('branch', 'شعبه')),
            ('amount', translator.t('amount', 'مبلغ')),
            ('currency', translator.t('currency', 'ارز')),
            ('status', translator.t('status', 'وضعیت')),
        ]
        for key, label in default_columns:
            if items and key in items[0]:
                keys.append(key)
                headers.append(label)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "چک‌ها"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data with formatting
    date_keys = {'issue_date', 'due_date'}
    amount_keys = {'amount'}
    for row_idx, item in enumerate(items, start=2):
        for col_idx, key in enumerate(keys, start=1):
            if key == 'row_number':
                value = row_idx - 1  # Row number (starting from 1)
            else:
                value = item.get(key, '')
                # Format dates (date only, no time)
                if key in date_keys:
                    value = format_date_for_export(item, key)
                # Format numbers with thousands separator
                elif key in amount_keys:
                    value = format_number_for_display(value)
                # Translate status and type
                elif key == 'status':
                    value = translate_status(str(value) if value else '')
                elif key == 'type':
                    value = translate_type(str(value) if value else '')
                # Format None values
                elif value is None:
                    value = ''
                else:
                    value = str(value) if value is not None else ''
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(header)
        for row_idx in range(2, len(items) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=checks_{business_id}.xlsx"}
    )


@router.post(
    "/businesses/{business_id}/checks/export/pdf",
    summary="خروجی PDF لیست چک‌ها",
    description="خروجی PDF لیست چک‌ها با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_checks_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی PDF لیست چک‌ها"""
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale, locale_dependency, Translator
    from app.core.calendar import CalendarConverter, get_calendar_type_from_header
    from html import escape
    from fastapi.responses import Response
    import json
    import datetime as dt_module
    
    # Get locale and translator
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    translator = await locale_dependency(request)
    is_rtl = locale.startswith('fa')
    is_fa = locale == 'fa'
    
    # Get calendar type
    calendar_type = get_calendar_type_from_header(request.headers.get("X-Calendar-Type"))
    if not calendar_type:
        calendar_type = "jalali" if is_fa else "gregorian"
    
    # Build query dict from flat body
    max_export_records = 10000
    take_value = min(int(body.get("take", 1000)), max_export_records)
    
    query_dict = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }
    
    # Handle person_id from body
    if body.get("person_id") is not None:
        try:
            query_dict["person_id"] = int(body.get("person_id"))
        except Exception:
            pass

    result = list_checks(db, business_id, query_dict)
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            "row_number": "⚠️",
            "type": translator.t("warning", "هشدار"),
            "check_number": translator.t("max_export_limit", "حداکثر ۱۰,۰۰۰ رکورد قابل export است"),
            "person_name": "",
            "amount": "",
            "due_date": "",
            "status": "",
        }
        items.append(warning_item)

    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Helper functions
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, dt_module.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, dt_module.date):
            try:
                dt_value = dt_module.datetime.combine(value, dt_module.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = dt_module.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = dt_module.date.fromisoformat(value)
                            dt_value = dt_module.datetime.combine(date_value, dt_module.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = dt_module.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = dt_module.date.fromisoformat(value)
                        dt_value = dt_module.datetime.combine(date_value, dt_module.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    def format_number_for_display(value) -> str:
        """Format number with thousands separator and remove .0"""
        try:
            if value is None:
                return ""
            v = float(value)
            s = f"{v:,.2f}"
            # Trim trailing .00 or trailing zeros
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return s
        except Exception:
            return str(value) if value is not None else ""
    
    def translate_status(status: str) -> str:
        """Translate check status"""
        if not status:
            return ""
        status_map = {
            'RECEIVED_ON_HAND': translator.t('check_status_received_on_hand', 'در دست (دریافتی)'),
            'TRANSFERRED_ISSUED': translator.t('check_status_transferred_issued', 'صادر شده (پرداختنی)'),
            'DEPOSITED': translator.t('check_status_deposited', 'سپرده به بانک'),
            'CLEARED': translator.t('check_status_cleared', 'پاس/وصول شده'),
            'ENDORSED': translator.t('check_status_endorsed', 'واگذار شده'),
            'RETURNED': translator.t('check_status_returned', 'عودت شده'),
            'BOUNCED': translator.t('check_status_bounced', 'برگشت خورده'),
            'CANCELLED': translator.t('check_status_cancelled', 'ابطال'),
        }
        return status_map.get(status.upper(), status)
    
    def translate_type(check_type: str) -> str:
        """Translate check type"""
        if not check_type:
            return ""
        type_map = {
            'received': translator.t('check_type_received', 'دریافتی'),
            'transferred': translator.t('check_type_transferred', 'واگذار شده'),
        }
        return type_map.get(check_type.lower(), check_type)

    # Prepare headers based on export_columns (add row number at the beginning)
    headers: list = []
    keys: list = []
    
    # Always add row number as first column
    headers.append(translator.t('row_number', 'ردیف'))
    keys.append('row_number')
    
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key and key != 'row_number':  # Skip row_number if already added
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns for checks
        default_columns = [
            ('type', translator.t('check_type', 'نوع')),
            ('person_name', translator.t('person', 'شخص')),
            ('issue_date', translator.t('issue_date', 'تاریخ صدور')),
            ('due_date', translator.t('due_date', 'تاریخ سررسید')),
            ('check_number', translator.t('check_number', 'شماره چک')),
            ('sayad_code', translator.t('sayad_code', 'شناسه صیاد')),
            ('bank_name', translator.t('bank', 'بانک')),
            ('branch_name', translator.t('branch', 'شعبه')),
            ('amount', translator.t('amount', 'مبلغ')),
            ('currency', translator.t('currency', 'ارز')),
            ('status', translator.t('status', 'وضعیت')),
        ]
        for key, label in default_columns:
            if items and key in items[0]:
                keys.append(key)
                headers.append(label)
    
    # Format generated date based on calendar
    try:
        _now = dt_module.datetime.now()
        _fd = CalendarConverter.format_datetime(_now, calendar_type)
        generated_at = _fd.get("formatted") or _fd.get("date_only") or _now.strftime('%Y/%m/%d %H:%M')
    except Exception:
        generated_at = dt_module.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    # Build HTML table
    title_text = translator.t('checks_list', 'لیست چک‌ها')
    date_keys = {'issue_date', 'due_date'}
    amount_keys = {'amount'}
    
    html_content = f"""
    <!DOCTYPE html>
    <html dir="{'rtl' if is_rtl else 'ltr'}" lang="{locale}">
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 1cm;
            }}
            body {{
                font-family: {'Arial, sans-serif' if not is_rtl else 'Tahoma, Arial, sans-serif'};
                font-size: 9pt;
                direction: {'rtl' if is_rtl else 'ltr'};
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            th {{
                background-color: #366092;
                color: white;
                padding: 8px;
                text-align: center;
                font-weight: bold;
                border: 1px solid #ddd;
            }}
            td {{
                padding: 6px;
                text-align: center;
                border: 1px solid #ddd;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
            .header {{
                text-align: center;
                margin-bottom: 20px;
            }}
            .header h1 {{
                margin: 0;
                font-size: 18pt;
            }}
            .amount {{
                text-align: left;
                direction: ltr;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>{escape(title_text)}</h1>
            <p>{translator.t('generated_at', 'تاریخ گزارش')}: {escape(generated_at)}</p>
        </div>
        <table>
            <thead>
                <tr>
    """
    
    for header in headers:
        html_content += f'<th>{escape(str(header))}</th>'
    
    html_content += """
                </tr>
            </thead>
            <tbody>
    """
    
    for idx, item in enumerate(items, 1):
        html_content += '<tr>'
        for key in keys:
            if key == 'row_number':
                value = str(idx)
            else:
                value = item.get(key, '')
                # Format dates (date only, no time)
                if key in date_keys:
                    value = format_date_for_export(item, key)
                # Format numbers with thousands separator
                elif key in amount_keys:
                    value = format_number_for_display(value)
                # Translate status and type
                elif key == 'status':
                    value = translate_status(str(value) if value else '')
                elif key == 'type':
                    value = translate_type(str(value) if value else '')
                # Format None values
                elif value is None:
                    value = ''
                else:
                    value = str(value) if value is not None else ''
            
            # Use amount class for numeric columns
            cell_class = ' class="amount"' if key in amount_keys else ''
            html_content += f'<td{cell_class}>{escape(str(value))}</td>'
        html_content += '</tr>'
    
    html_content += """
            </tbody>
        </table>
    </body>
    </html>
    """
    
    # Generate PDF
    font_config = FontConfiguration()
    html_doc = HTML(string=html_content)
    pdf_bytes = html_doc.write_pdf(font_config=font_config)
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=checks_{business_id}.pdf"}
    )


