from fastapi import APIRouter, Depends, HTTPException, Query, Request, Body, Form
from fastapi import UploadFile, File
from fastapi.encoders import jsonable_encoder
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import Dict, Any, List, Optional

from adapters.db.session import get_db
from adapters.api.v1.schema_models.person import (
    PersonCreateRequest,
    PersonUpdateRequest,
    PersonResponse,
    PersonListResponse,
    PersonSummaryResponse,
    PersonBankAccountCreateRequest,
    PersonShareLinkCreateRequest,
)
from adapters.api.v1.schemas import QueryInfo, SuccessResponse
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.cache import get_cache
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.permissions import require_business_management_dep, require_business_access, require_business_permission_dep, require_business_permission_by_entity_dep, require_business_access_dep
from app.core.i18n import negotiate_locale
from app.services.person_service import (
    invalidate_persons_cache,
    create_person,
    get_person_by_id,
    get_persons_by_business,
    update_person,
    delete_person,
    get_person_summary,
    get_debtors_report,
    get_creditors_report,
    get_people_transactions_report,
)
from app.services.person_share_link_service import (
    create_share_link as create_person_share_link_service,
    get_active_share_link_for_person,
    revoke_share_link as revoke_person_share_link_service,
    serialize_share_link,
)
from adapters.db.models.person import Person
from adapters.db.models.business import Business
from adapters.db.models.fiscal_year import FiscalYear

router = APIRouter(prefix="/persons", tags=["اشخاص و مشتریان"])


@router.post("/businesses/{business_id}/persons/bulk-delete",
    summary="حذف گروهی اشخاص",
    description="حذف چندین شخص بر اساس شناسه‌ها یا کدها",
)
async def bulk_delete_persons_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("people", "delete")),
):
    """حذف گروهی اشخاص برای یک کسب‌وکار مشخص

    ورودی:
      - ids: لیست شناسه‌های اشخاص
      - codes: لیست کدهای اشخاص در همان کسب‌وکار
    """
    from sqlalchemy import and_ as _and
    from adapters.db.models.person import Person
    from app.services.person_service import delete_person

    ids = body.get("ids")
    codes = body.get("codes")
    deleted = 0
    skipped = 0
    errors = []

    if not ids and not codes:
        return success_response({"deleted": 0, "skipped": 0, "errors": []}, request)

    # Normalize inputs
    if isinstance(ids, list):
        try:
            ids = [int(x) for x in ids if isinstance(x, (int, str)) and str(x).isdigit()]
        except Exception:
            ids = []
    else:
        ids = []

    if isinstance(codes, list):
        try:
            codes = [int(str(x).strip()) for x in codes if str(x).strip().isdigit()]
        except Exception:
            codes = []
    else:
        codes = []

    # Delete by IDs first
    if ids:
        for pid in ids:
            try:
                person = db.query(Person).filter(_and(Person.id == pid, Person.business_id == business_id)).first()
                if person is None:
                    skipped += 1
                    errors.append(f"شخص با شناسه {pid} یافت نشد")
                    continue
                
                success, error_message = delete_person(db, pid, business_id)
                if success:
                    deleted += 1
                else:
                    skipped += 1
                    if error_message:
                        errors.append(f"شخص {person.alias_name or person.code or pid}: {error_message}")
                    else:
                        errors.append(f"شخص {person.alias_name or person.code or pid}: امکان حذف وجود ندارد")
            except Exception as e:
                skipped += 1
                errors.append(f"خطا در حذف شخص با شناسه {pid}: {str(e)}")

    # Delete by codes
    if codes:
        try:
            items = db.query(Person).filter(_and(Person.business_id == business_id, Person.code.in_(codes))).all()
            for obj in items:
                try:
                    success, error_message = delete_person(db, obj.id, business_id)
                    if success:
                        deleted += 1
                    else:
                        skipped += 1
                        if error_message:
                            errors.append(f"شخص {obj.alias_name or obj.code or obj.id}: {error_message}")
                        else:
                            errors.append(f"شخص {obj.alias_name or obj.code or obj.id}: امکان حذف وجود ندارد")
                except Exception as e:
                    skipped += 1
                    errors.append(f"خطا در حذف شخص {obj.alias_name or obj.code or obj.id}: {str(e)}")
        except Exception as e:
            # In case of query issues, treat all as skipped
            skipped += len(codes)
            errors.append(f"خطا در جستجوی اشخاص: {str(e)}")

    # Invalidate کش لیست اشخاص یک بار در انتها (بهینه‌سازی برای bulk delete)
    # اگرچه delete_person خودش invalidate می‌کند، اما این کار برای اطمینان از حذف کامل کش انجام می‌شود
    if deleted > 0:
        invalidate_persons_cache(business_id, fiscal_year_id=None)

    return success_response({
        "deleted": deleted, 
        "skipped": skipped,
        "errors": errors
    }, request)


@router.post("/businesses/{business_id}/persons/create", 
    summary="ایجاد شخص جدید", 
    description="ایجاد شخص جدید برای کسب و کار مشخص",
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "شخص با موفقیت ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "شخص با موفقیت ایجاد شد",
                        "data": {
                            "id": 1,
                            "business_id": 1,
                            "alias_name": "علی احمدی",
                            "person_type": "مشتری",
                            "created_at": "2024-01-01T00:00:00Z"
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی داده‌ها"
        },
        401: {
            "description": "عدم احراز هویت"
        },
        403: {
            "description": "عدم دسترسی به کسب و کار"
        }
    }
)
async def create_person_endpoint(
    request: Request,
    business_id: int,
    person_data: PersonCreateRequest,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_access_dep),
):
    """ایجاد شخص جدید برای کسب و کار"""
    result = create_person(db, business_id, person_data)
    return success_response(
        data=format_datetime_fields(result['data'], request),
        request=request,
        message=result['message'],
    )


@router.post("/businesses/{business_id}/persons",
    summary="لیست اشخاص کسب و کار",
    description="دریافت لیست اشخاص یک کسب و کار با امکان جستجو و فیلتر",
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "لیست اشخاص با موفقیت دریافت شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "لیست اشخاص با موفقیت دریافت شد",
                        "data": {
                            "items": [],
                            "pagination": {
                                "total": 0,
                                "page": 1,
                                "per_page": 20,
                                "total_pages": 0,
                                "has_next": False,
                                "has_prev": False
                            },
                            "query_info": {}
                        }
                    }
                }
            }
        }
    }
)
async def get_persons_endpoint(
    request: Request,
    business_id: int,
    query_info: QueryInfo,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
):
    """دریافت لیست اشخاص کسب و کار"""
    # دریافت سال مالی از header
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    # اگر سال مالی مشخص نشده، از سال مالی جاری business استفاده می‌کنیم
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    query_dict = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "sort": [s.model_dump() for s in query_info.sort] if query_info.sort else None,
        "search": query_info.search,
        "search_fields": query_info.search_fields,
        "filters": query_info.filters,
    }

    # کش نتایج لیست اشخاص بر اساس پارامترها (با بهینه‌سازی tag-based)
    cache = get_cache()
    cache_key = None

    if cache.enabled:
        import json, hashlib
        # تبدیل query_info به dict برای serialize کردن (تبدیل FilterItem objects به dict)
        # استفاده از jsonable_encoder برای تبدیل Pydantic models به dict قابل serialize
        query_dict_for_cache = jsonable_encoder(query_dict)
        key_payload = {
            "business_id": business_id,
            "fiscal_year_id": fiscal_year_id,
            "query": query_dict_for_cache,
        }
        key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
        cache_key = f"persons_list:{key_hash}"
        cached = cache.get(cache_key)
        if cached is not None:
            return success_response(
                data=cached,
                request=request,
                message="لیست اشخاص با موفقیت دریافت شد",
            )

    result = get_persons_by_business(db, business_id, query_dict, fiscal_year_id)
    
    # فرمت کردن تاریخ‌ها
    result['items'] = [
        format_datetime_fields(item, request) for item in result['items']
    ]

    if cache.enabled and cache_key:
        # استفاده از set_with_business_tag برای مدیریت بهتر با set ردیس
        # این متد کلید را در set های مربوط به business_id و fiscal_year_id ذخیره می‌کند
        cache.set_with_business_tag(cache_key, result, business_id=business_id, fiscal_year_id=fiscal_year_id, ttl=60)
    
    return success_response(
        data=result,
        request=request,
        message="لیست اشخاص با موفقیت دریافت شد",
    )


@router.post("/businesses/{business_id}/persons/export/excel",
    summary="خروجی Excel لیست اشخاص",
    description="خروجی Excel لیست اشخاص با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
async def export_persons_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    import json
    import datetime
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response

    # دریافت سال مالی از header
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    # اگر سال مالی مشخص نشده، از سال مالی جاری business استفاده می‌کنیم
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # Build query dict similar to list endpoint from flat body
    query_dict = {
        "take": int(body.get("take", 20)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }

    result = get_persons_by_business(db, business_id, query_dict, fiscal_year_id)

    items = result.get('items', [])
    # Format date/time fields using existing helper
    items = [format_datetime_fields(item, request) for item in items]

    # Apply selected indices filter if requested
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    # Prepare headers based on export_columns (order + visibility)
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Fallback to item keys if no columns provided
        if items:
            keys = list(items[0].keys())
            headers = keys

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Persons"

    # Locale and RTL/LTR handling
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")

    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    base = "persons"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/persons/export/pdf",
    summary="خروجی PDF لیست اشخاص",
    description="خروجی PDF لیست اشخاص با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
async def export_persons_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration

    # دریافت سال مالی از header
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    # اگر سال مالی مشخص نشده، از سال مالی جاری business استفاده می‌کنیم
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # Build query dict from flat body
    query_dict = {
        "take": int(body.get("take", 20)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields"),
        "filters": body.get("filters"),
    }

    result = get_persons_by_business(db, business_id, query_dict, fiscal_year_id)
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]

    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        if items:
            keys = list(items[0].keys())
            headers = keys

    # Load business info for header
    business_name = ""
    try:
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name
    except Exception:
        business_name = ""

    # Styled HTML with dynamic direction/locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'

    def escape(s: Any) -> str:
        try:
            return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        except Exception:
            return str(s)

    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key)
            if value is None:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            tds.append(f"<td>{escape(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = ''.join(f"<th>{escape(h)}</th>" for h in headers)
    # Format report datetime based on X-Calendar-Type header
    calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian")
        now = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    title_text = "گزارش لیست اشخاص" if is_fa else "Persons List Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    footer_text = "تولید شده توسط Hesabix" if is_fa else "Generated by Hesabix"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "

    # تلاش برای رندر با قالب سفارشی (persons/list)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": title_text,
            "business_name": business_name,
            "generated_at": now,
            "is_fa": is_fa,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="persons",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # Inject Persian fonts (YekanBakhFaNum/Vazirmatn) for PDF rendering
    fa_font_url_regular = ""
    fa_font_url_bold = ""
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_reg, fa_bold = load_farsi_font_data_uris()
            fa_font_url_regular = fa_reg or ""
            fa_font_url_bold = fa_bold or ""
    except Exception:
        fa_font_url_regular = ""
        fa_font_url_bold = ""
    
    font_face_css = ""
    if is_fa and fa_font_url_regular:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_regular}') format('truetype'); font-weight: 400; font-style: normal; }}
        """
    if is_fa and fa_font_url_bold:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_bold}') format('truetype'); font-weight: 700; font-style: normal; }}
        """
    
    body_font_family = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if is_fa else "Arial, sans-serif"
    
    table_html = f"""
    <html lang=\"{html_lang}\" dir=\"{html_dir}\"> 
      <head>
        <meta charset='utf-8'>
        <style>
          {font_face_css}
          @page {{
            size: A4 landscape;
            margin: 12mm;
            @bottom-{ 'left' if is_fa else 'right' } {{
              content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
              font-size: 10px;
              color: #666;
              font-family: {body_font_family};
            }}
          }}
          body {{
            font-family: {body_font_family};
            font-size: 11px;
            color: #222;
          }}
          .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #444;
            padding-bottom: 6px;
          }}
          .title {{
            font-size: 16px;
            font-weight: 700;
          }}
          .meta {{
            font-size: 11px;
            color: #555;
          }}
          .table-wrapper {{
            width: 100%;
          }}
          table.report-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
          }}
          thead th {{
            background: #f0f3f7;
            border: 1px solid #c7cdd6;
            padding: 6px 4px;
            text-align: center;
            font-weight: 700;
            white-space: nowrap;
          }}
          tbody td {{
            border: 1px solid #d7dde6;
            padding: 5px 4px;
            vertical-align: top;
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: normal;
          }}
          .footer {{
            position: running(footer);
            font-size: 10px;
            color: #666;
            margin-top: 8px;
            text-align: {'left' if is_fa else 'right'};
          }}
        </style>
      </head>
      <body>
        <div class=\"header\">
          <div>
            <div class=\"title\">{title_text}</div>
            <div class=\"meta\">{label_biz}: {escape(business_name)}</div>
          </div>
          <div class=\"meta\">{label_date}: {escape(now)}</div>
        </div>
        <div class=\"table-wrapper\">
          <table class=\"report-table\">
            <thead>
              <tr>{headers_html}</tr>
            </thead>
            <tbody>
              {''.join(rows_html)}
            </tbody>
          </table>
        </div>
        <div class=\"footer\">{footer_text}</div>
      </body>
    </html>
    """

    final_html = resolved_html or table_html
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)

    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    base = "persons"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/persons/import/template",
    summary="دانلود تمپلیت ایمپورت اشخاص",
    description="فایل Excel تمپلیت برای ایمپورت اشخاص را برمی‌گرداند",
)
async def download_persons_import_template(
    business_id: int,
    request: Request,
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    import datetime
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = [
        'code','alias_name','first_name','last_name','person_type','person_types','company_name','payment_id',
        'national_id','registration_number','economic_id','country','province','city','address','postal_code',
        'phone','mobile','fax','email','website','share_count','commission_sale_percent','commission_sales_return_percent',
        'commission_sales_amount','commission_sales_return_amount'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Sample row
    sample = [
        '', 'نمونه نام مستعار', 'علی', 'احمدی', 'مشتری', 'مشتری, فروشنده', 'نمونه شرکت', 'PID123',
        '0012345678', '12345', 'ECO-1', 'ایران', 'تهران', 'تهران', 'خیابان مثال ۱', '1234567890',
        '02112345678', '09120000000', '', 'test@example.com', 'example.com', '', '5', '0', '0', '0'
    ]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    # Auto width
    for column in ws.columns:
        try:
            letter = column[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"persons_import_template_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.get("/persons/{person_id}",
    summary="جزئیات شخص",
    description="دریافت جزئیات یک شخص",
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "جزئیات شخص با موفقیت دریافت شد"
        },
        404: {
            "description": "شخص یافت نشد"
        }
    }
)
async def get_person_endpoint(
    request: Request,
    person_id: int,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people", "view", Person, "person_id")),
):
    """دریافت جزئیات شخص"""
    # ابتدا باید business_id را از person دریافت کنیم
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    
    result = get_person_by_id(db, person_id, person.business_id)
    if not result:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    
    return success_response(
        data=format_datetime_fields(result, request),
        request=request,
        message="جزئیات شخص با موفقیت دریافت شد",
    )


@router.put("/persons/{person_id}",
    summary="ویرایش شخص",
    description="ویرایش اطلاعات یک شخص",
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "شخص با موفقیت ویرایش شد"
        },
        404: {
            "description": "شخص یافت نشد"
        }
    }
)
async def update_person_endpoint(
    request: Request,
    person_id: int,
    person_data: PersonUpdateRequest,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people", "edit", Person, "person_id")),
):
    """ویرایش شخص"""
    # ابتدا باید business_id را از person دریافت کنیم
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    
    result = update_person(db, person_id, person.business_id, person_data)
    if not result:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    
    return success_response(
        data=format_datetime_fields(result['data'], request),
        request=request,
        message=result['message'],
    )


@router.delete("/persons/{person_id}",
    summary="حذف شخص",
    description="حذف یک شخص",
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "شخص با موفقیت حذف شد"
        },
        404: {
            "description": "شخص یافت نشد"
        }
    }
)
async def delete_person_endpoint(
    request: Request,
    person_id: int,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people", "delete", Person, "person_id")),
):
    """حذف شخص"""
    # ابتدا باید business_id را از person دریافت کنیم
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    
    success, error_message = delete_person(db, person_id, person.business_id)
    if not success:
        if error_message:
            raise HTTPException(status_code=400, detail=error_message)
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    
    return success_response(data=None, message="شخص با موفقیت حذف شد", request=request)


@router.get(
    "/persons/{person_id}/share-link",
    summary="وضعیت لینک اشتراک شخص",
    response_model=SuccessResponse,
)
async def get_person_share_link_endpoint(
    request: Request,
    person_id: int,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people", "view", Person, "person_id")),
):
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    link = get_active_share_link_for_person(db, person.business_id, person.id)
    return success_response(
        data={"link": serialize_share_link(link, request_base_url=str(request.base_url))},
        request=request,
        message="وضعیت لینک اشتراک",
    )


@router.post(
    "/persons/{person_id}/share-link",
    summary="ایجاد یا بروزرسانی لینک اشتراک",
    response_model=SuccessResponse,
)
async def create_person_share_link_endpoint(
    request: Request,
    person_id: int,
    payload: PersonShareLinkCreateRequest,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people", "edit", Person, "person_id")),
):
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")

    link = create_person_share_link_service(
        db,
        business_id=person.business_id,
        person_id=person.id,
        user_id=auth_context.get_user_id(),
        expires_in_hours=payload.expires_in_hours,
        max_view_count=payload.max_view_count,
        options=payload.options.model_dump(),
        replace_existing=payload.replace_existing,
    )
    return success_response(
        data=serialize_share_link(link, request_base_url=str(request.base_url)),
        request=request,
        message="لینک اشتراک ایجاد شد",
    )


@router.delete(
    "/persons/{person_id}/share-link",
    summary="لغو لینک اشتراک شخص",
    response_model=SuccessResponse,
)
async def revoke_person_share_link_endpoint(
    request: Request,
    person_id: int,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("people", "edit", Person, "person_id")),
):
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="شخص یافت نشد")
    revoked = revoke_person_share_link_service(
        db,
        business_id=person.business_id,
        person_id=person.id,
        user_id=auth_context.get_user_id(),
    )
    if not revoked:
        raise HTTPException(status_code=404, detail="لینک فعالی برای لغو وجود ندارد")
    return success_response(
        data=None,
        message="لینک اشتراک لغو شد",
        request=request,
    )


@router.get("/businesses/{business_id}/persons/summary",
    summary="خلاصه اشخاص کسب و کار",
    description="دریافت خلاصه آماری اشخاص یک کسب و کار",
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "خلاصه اشخاص با موفقیت دریافت شد"
        }
    }
)
async def get_persons_summary_endpoint(
    request: Request,
    business_id: int,
    db: Session = Depends(get_db),
    auth_context: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("people", "view")),
):
    """دریافت خلاصه اشخاص کسب و کار"""
    result = get_person_summary(db, business_id)
    
    return success_response(
        data=result,
        request=request,
        message="خلاصه اشخاص با موفقیت دریافت شد",
    )


@router.post("/businesses/{business_id}/persons/import/excel",
    summary="ایمپورت اشخاص از فایل Excel",
    description="فایل اکسل را دریافت می‌کند و به‌صورت dry-run یا واقعی پردازش می‌کند",
)
async def import_persons_excel(
    business_id: int,
    request: Request,
    file: UploadFile = File(...),
    dry_run: str = Form(default="true"),
    match_by: str = Form(default="code"),
    conflict_policy: str = Form(default="upsert"),
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import io
    import json
    import re
    from openpyxl import load_workbook
    from fastapi import HTTPException
    import logging
    import zipfile

    logger = logging.getLogger(__name__)
    
    def validate_excel_file(content: bytes) -> bool:
        """
        Validate if the content is a valid Excel file
        """
        try:
            # Check if it starts with PK signature (zip file)
            if not content.startswith(b'PK'):
                return False
            
            # Try to open as zip file
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zip_file:
                file_list = zip_file.namelist()
                # Check for Excel structure (xl/ folder for .xlsx files)
                excel_structure = any(f.startswith('xl/') for f in file_list)
                if excel_structure:
                    return True
                
                # Check for older Excel format (.xls) - this would be a different structure
                # But since we only support .xlsx, we'll return False for .xls
                return False
        except zipfile.BadZipFile:
            logger.error("File is not a valid zip file")
            return False
        except Exception as e:
            logger.error(f"Error validating Excel file: {str(e)}")
            return False
    
    try:
        # Convert dry_run string to boolean
        dry_run_bool = dry_run.lower() in ('true', '1', 'yes', 'on')
        
        logger.info(f"Import request: business_id={business_id}, dry_run={dry_run_bool}, match_by={match_by}, conflict_policy={conflict_policy}")
        logger.info(f"File info: filename={file.filename}, content_type={file.content_type}")

        if not file.filename or not file.filename.lower().endswith('.xlsx'):
            logger.error(f"Invalid file format: {file.filename}")
            raise HTTPException(status_code=400, detail="فرمت فایل معتبر نیست. تنها xlsx پشتیبانی می‌شود")

        content = await file.read()
        logger.info(f"File content size: {len(content)} bytes")
        
        # Log first few bytes for debugging
        logger.info(f"File header (first 20 bytes): {content[:20].hex()}")
        logger.info(f"File header (first 20 bytes as text): {content[:20]}")
        
        # Check if content is empty or too small
        if len(content) < 100:
            logger.error(f"File too small: {len(content)} bytes")
            raise HTTPException(status_code=400, detail="فایل خیلی کوچک است یا خالی است")
        
        # Validate Excel file format
        if not validate_excel_file(content):
            logger.error("File is not a valid Excel file")
            raise HTTPException(status_code=400, detail="فرمت فایل معتبر نیست. فایل Excel معتبر نیست")
        
        try:
            # Try to load the workbook with additional error handling
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
            logger.info(f"Successfully loaded workbook with {len(wb.worksheets)} worksheets")
        except zipfile.BadZipFile as e:
            logger.error(f"Bad zip file error: {str(e)}")
            raise HTTPException(status_code=400, detail="فایل Excel خراب است یا فرمت آن معتبر نیست")
        except Exception as e:
            logger.error(f"Error loading workbook: {str(e)}")
            raise HTTPException(status_code=400, detail=f"امکان خواندن فایل وجود ندارد: {str(e)}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return success_response(data={"summary": {"total": 0}}, request=request, message="فایل خالی است")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = rows[1:]

        # helper to map enum strings (fa/en) to internal value
        def normalize_person_type(value: str) -> Optional[str]:
            if not value:
                return None
            value = str(value).strip()
            mapping = {
                'customer': 'مشتری', 'marketer': 'بازاریاب', 'employee': 'کارمند', 'supplier': 'تامین‌کننده',
                'partner': 'همکار', 'seller': 'فروشنده', 'shareholder': 'سهامدار'
            }
            for en, fa in mapping.items():
                if value.lower() == en or value == fa:
                    return fa
            return value  # assume already fa

        errors: list[dict] = []
        valid_items: list[dict] = []

        for idx, row in enumerate(data_rows, start=2):
            item: dict[str, Any] = {}
            row_errors: list[str] = []
            for ci, key in enumerate(headers):
                if not key:
                    continue
                val = row[ci] if ci < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                item[key] = val
            # normalize types
            if 'person_type' in item and item['person_type']:
                item['person_type'] = normalize_person_type(item['person_type'])
            if 'person_types' in item and item['person_types']:
                # split by comma
                parts = [normalize_person_type(p.strip()) for p in str(item['person_types']).split(',') if str(p).strip()]
                item['person_types'] = parts

            # alias_name required
            if not item.get('alias_name'):
                row_errors.append('alias_name الزامی است')

            # shareholder rule
            if (item.get('person_type') == 'سهامدار') or (isinstance(item.get('person_types'), list) and 'سهامدار' in item.get('person_types', [])):
                sc = item.get('share_count')
                try:
                    sc_val = int(sc) if sc is not None and str(sc).strip() != '' else None
                except Exception:
                    sc_val = None
                if sc_val is None or sc_val <= 0:
                    row_errors.append('برای سهامدار share_count باید > 0 باشد')
                else:
                    item['share_count'] = sc_val

            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
                continue

            valid_items.append(item)

        inserted = 0
        updated = 0
        skipped = 0

        if not dry_run_bool and valid_items:
            # apply import with conflict policy
            from adapters.db.models.person import Person
            from sqlalchemy import and_

            def find_existing(session: Session, data: dict) -> Optional[Person]:
                if match_by == 'national_id' and data.get('national_id'):
                    return session.query(Person).filter(and_(Person.business_id == business_id, Person.national_id == data['national_id'])).first()
                if match_by == 'email' and data.get('email'):
                    return session.query(Person).filter(and_(Person.business_id == business_id, Person.email == data['email'])).first()
                if match_by == 'code' and data.get('code'):
                    try:
                        code_int = int(data['code'])
                        return session.query(Person).filter(and_(Person.business_id == business_id, Person.code == code_int)).first()
                    except Exception:
                        return None
                return None

            for data in valid_items:
                existing = find_existing(db, data)
                match_value = None
                try:
                    match_value = data.get(match_by)
                except Exception:
                    match_value = None
                if existing is None:
                    # create
                    try:
                        create_person(db, business_id, PersonCreateRequest(**data))
                        inserted += 1
                    except Exception as e:
                        logger.error(f"Create person failed for data={data}: {str(e)}")
                        skipped += 1
                else:
                    if conflict_policy == 'insert':
                        logger.info(f"Skipping existing person (match_by={match_by}, value={match_value}) due to conflict_policy=insert")
                        skipped += 1
                    elif conflict_policy in ('update', 'upsert'):
                        try:
                            update_person(db, existing.id, business_id, PersonUpdateRequest(**data))
                            updated += 1
                        except Exception as e:
                            logger.error(f"Update person failed for id={existing.id}, data={data}: {str(e)}")
                            skipped += 1

        summary = {
            "total": len(data_rows),
            "valid": len(valid_items),
            "invalid": len(errors),
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "dry_run": dry_run_bool,
        }

        # Invalidate کش لیست اشخاص یک بار در انتها (فقط اگر dry_run نباشد و تغییراتی انجام شده باشد)
        if not dry_run_bool and (inserted > 0 or updated > 0):
            invalidate_persons_cache(business_id, fiscal_year_id=None)

        return success_response(
            data={
                "summary": summary,
                "errors": errors,
            },
            request=request,
            message="نتیجه ایمپورت اشخاص",
        )
    except Exception as e:
        logger.error(f"Import error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"خطا در پردازش فایل: {str(e)}")


@router.post("/businesses/{business_id}/reports/debtors",
    summary="گزارش بدهکاران",
    description="گزارش لیست بدهکاران با امکان فیلتر بر اساس سال مالی، تاریخ، حداقل بدهی و جستجو",
)
async def debtors_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش بدهکاران"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    # اگر سال مالی در body مشخص شده، اولویت با body است
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # اگر سال مالی مشخص نشده، از سال مالی جاری business استفاده می‌کنیم
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    min_balance = body.get('min_balance')
    if min_balance is not None:
        try:
            min_balance = float(min_balance)
        except (ValueError, TypeError):
            min_balance = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    search = body.get('search')
    skip = int(body.get('skip', 0))
    take = int(body.get('take', 50))
    
    # دریافت گزارش
    result = get_debtors_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        min_balance=min_balance,
        person_ids=person_ids,
        search=search,
        skip=skip,
        take=take,
    )
    
    # فرمت کردن تاریخ‌ها
    result['items'] = [
        format_datetime_fields(item, request) for item in result['items']
    ]
    
    return success_response(
        data=result,
        request=request,
        message="گزارش بدهکاران با موفقیت دریافت شد",
    )


@router.post("/businesses/{business_id}/reports/creditors",
    summary="گزارش بستانکاران",
    description="گزارش لیست بستانکاران با امکان فیلتر بر اساس سال مالی، تاریخ، حداقل بستانکاری و جستجو",
)
async def creditors_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """گزارش بستانکاران"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    # اگر سال مالی در body مشخص شده، اولویت با body است
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # اگر سال مالی مشخص نشده، از سال مالی جاری business استفاده می‌کنیم
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    min_balance = body.get('min_balance')
    if min_balance is not None:
        try:
            min_balance = float(min_balance)
        except (ValueError, TypeError):
            min_balance = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    search = body.get('search')
    skip = int(body.get('skip', 0))
    take = int(body.get('take', 50))
    
    # دریافت گزارش
    result = get_creditors_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        min_balance=min_balance,
        person_ids=person_ids,
        search=search,
        skip=skip,
        take=take,
    )
    
    # فرمت کردن تاریخ‌ها
    result['items'] = [
        format_datetime_fields(item, request) for item in result['items']
    ]
    
    return success_response(
        data=result,
        request=request,
        message="گزارش بستانکاران با موفقیت دریافت شد",
    )


@router.post("/businesses/{business_id}/reports/debtors/export/excel",
    summary="خروجی Excel گزارش بدهکاران",
    description="خروجی Excel گزارش بدهکاران با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_debtors_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel گزارش بدهکاران"""
    import io
    import json
    import datetime
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    min_balance = body.get('min_balance')
    if min_balance is not None:
        try:
            min_balance = float(min_balance)
        except (ValueError, TypeError):
            min_balance = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    search = body.get('search')
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_debtors_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        min_balance=min_balance,
        person_ids=person_ids,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            'code': '⚠️',
            'display_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'balance': '',
            'total_debit': '',
            'total_credit': '',
            'last_transaction_date': '',
            'status': '',
        }
        items.append(warning_item)
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('code', 'کد' if is_fa else 'Code'),
            ('display_name', 'نام' if is_fa else 'Name'),
            ('balance', 'تراز' if is_fa else 'Balance'),
            ('total_debit', 'بدهکار' if is_fa else 'Debit'),
            ('total_credit', 'بستانکار' if is_fa else 'Credit'),
            ('last_transaction_date', 'تاریخ آخرین تراکنش' if is_fa else 'Last Transaction Date'),
            ('status', 'وضعیت' if is_fa else 'Status'),
        ]
        for key, label in default_columns:
            if items and (key in items[0] or key == 'display_name'):
                keys.append(key)
                headers.append(label)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش بدهکاران" if is_fa else "Debtors Report"
    
    # RTL handling for Persian
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            
            # Handle display_name specially
            if key == 'display_name':
                value = (
                    item.get('display_name') or
                    item.get('alias_name') or
                    f"{item.get('first_name', '')} {item.get('last_name', '')}".strip()
                )
            
            # Format numbers
            if key in ['balance', 'total_debit', 'total_credit'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    value = num_value
                except (ValueError, TypeError):
                    pass
            
            # Format dates
            if key == 'last_transaction_date' and value:
                value = format_date_for_export(item, 'last_transaction_date')
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # RTL alignment for Persian text and numbers
            if locale == 'fa':
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, str) and any('\u0600' <= c <= '\u06FF' for c in value):
                    cell.alignment = Alignment(horizontal="right")
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "debtors_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/reports/debtors/export/pdf",
    summary="خروجی PDF گزارش بدهکاران",
    description="خروجی PDF گزارش بدهکاران با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_debtors_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی PDF گزارش بدهکاران"""
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    min_balance = body.get('min_balance')
    if min_balance is not None:
        try:
            min_balance = float(min_balance)
        except (ValueError, TypeError):
            min_balance = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    search = body.get('search')
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_debtors_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        min_balance=min_balance,
        person_ids=person_ids,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted
        if isinstance(value, str):
            if '/' in value and (len(value.split('/')) == 3):
                if '-' in value:
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            'code': '⚠️',
            'display_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'balance': '',
            'total_debit': '',
            'total_credit': '',
            'last_transaction_date': '',
            'status': '',
        }
        items.append(warning_item)
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('code', 'کد' if is_fa else 'Code'),
            ('display_name', 'نام' if is_fa else 'Name'),
            ('balance', 'تراز' if is_fa else 'Balance'),
            ('total_debit', 'بدهکار' if is_fa else 'Debit'),
            ('total_credit', 'بستانکار' if is_fa else 'Credit'),
            ('last_transaction_date', 'تاریخ آخرین تراکنش' if is_fa else 'Last Transaction Date'),
            ('status', 'وضعیت' if is_fa else 'Status'),
        ]
        for key, label in default_columns:
            if items and (key in items[0] or key == 'display_name'):
                keys.append(key)
                headers.append(label)
    
    # Load business info for header
    business_name = ""
    try:
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name
    except Exception:
        business_name = ""
    
    def esc(s: Any) -> str:
        try:
            return escape(str(s))
        except Exception:
            return str(s)
    
    # Build table rows
    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key, "")
            
            # Handle display_name specially
            if key == 'display_name':
                value = (
                    item.get('display_name') or
                    item.get('alias_name') or
                    f"{item.get('first_name', '')} {item.get('last_name', '')}".strip()
                )
            
            # Format numbers
            if key in ['balance', 'total_debit', 'total_credit'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    # Format with thousand separators
                    if is_fa:
                        value = f"{num_value:,.0f}".replace(',', '٬')
                    else:
                        value = f"{num_value:,.2f}"
                except (ValueError, TypeError):
                    pass
            
            # Format dates
            if key == 'last_transaction_date' and value:
                value = format_date_for_export(item, 'last_transaction_date')
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            tds.append(f"<td>{esc(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")
    
    headers_html = ''.join(f"<th>{esc(h)}</th>" for h in headers)
    
    # Format report datetime based on X-Calendar-Type header
    calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian")
        now = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    title_text = "گزارش بدهکاران" if is_fa else "Debtors Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    footer_text = "تولید شده توسط Hesabix" if is_fa else "Generated by Hesabix"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "
    
    # تلاش برای رندر با قالب سفارشی (persons/reports/debtors)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": title_text,
            "business_name": business_name,
            "generated_at": now,
            "is_fa": is_fa,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="persons",
            subtype="reports/debtors",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None
    
    # Inject Persian fonts (YekanBakhFaNum/Vazirmatn) for PDF rendering
    fa_font_url_regular = ""
    fa_font_url_bold = ""
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_reg, fa_bold = load_farsi_font_data_uris()
            fa_font_url_regular = fa_reg or ""
            fa_font_url_bold = fa_bold or ""
    except Exception:
        fa_font_url_regular = ""
        fa_font_url_bold = ""
    
    font_face_css = ""
    if is_fa and fa_font_url_regular:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_regular}') format('truetype'); font-weight: 400; font-style: normal; }}
        """
    if is_fa and fa_font_url_bold:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_bold}') format('truetype'); font-weight: 700; font-style: normal; }}
        """
    
    body_font_family = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if is_fa else "Arial, sans-serif"
    
    table_html = f"""
    <html lang="{html_lang}" dir="{html_dir}"> 
      <head>
        <meta charset='utf-8'>
        <style>
          {font_face_css}
          @page {{
            size: A4 landscape;
            margin: 12mm;
            @bottom-{'left' if is_fa else 'right'} {{
              content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
              font-size: 10px;
              color: #666;
              font-family: {body_font_family};
            }}
          }}
          body {{
            font-family: {body_font_family};
            font-size: 11px;
            color: #222;
          }}
          .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #444;
            padding-bottom: 6px;
          }}
          .title {{
            font-size: 16px;
            font-weight: 700;
          }}
          .meta {{
            font-size: 11px;
            color: #555;
          }}
          .table-wrapper {{
            width: 100%;
          }}
          table.report-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
          }}
          thead th {{
            background: #f0f3f7;
            border: 1px solid #c7cdd6;
            padding: 6px 4px;
            text-align: center;
            font-weight: 700;
            white-space: nowrap;
          }}
          tbody td {{
            border: 1px solid #d7dde6;
            padding: 5px 4px;
            vertical-align: top;
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: normal;
            text-align: {'right' if is_fa else 'left'};
          }}
          tbody td:has-text(number) {{
            text-align: right;
          }}
          .footer {{
            position: running(footer);
            font-size: 10px;
            color: #666;
            margin-top: 8px;
            text-align: {'left' if is_fa else 'right'};
          }}
        </style>
      </head>
      <body>
        <div class="header">
          <div>
            <div class="title">{title_text}</div>
            <div class="meta">{label_biz}: {esc(business_name)}</div>
          </div>
          <div class="meta">{label_date}: {esc(now)}</div>
        </div>
        <div class="table-wrapper">
          <table class="report-table">
            <thead>
              <tr>{headers_html}</tr>
            </thead>
            <tbody>
              {''.join(rows_html)}
            </tbody>
          </table>
        </div>
        <div class="footer">{footer_text}</div>
      </body>
    </html>
    """
    
    final_html = resolved_html or table_html
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "debtors_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/reports/creditors/export/excel",
    summary="خروجی Excel گزارش بستانکاران",
    description="خروجی Excel گزارش بستانکاران با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_creditors_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel گزارش بستانکاران"""
    import io
    import json
    import datetime
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    min_balance = body.get('min_balance')
    if min_balance is not None:
        try:
            min_balance = float(min_balance)
        except (ValueError, TypeError):
            min_balance = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    search = body.get('search')
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_creditors_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        min_balance=min_balance,
        person_ids=person_ids,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            'code': '⚠️',
            'display_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'balance': '',
            'total_debit': '',
            'total_credit': '',
            'last_transaction_date': '',
            'status': '',
        }
        items.append(warning_item)
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('code', 'کد' if is_fa else 'Code'),
            ('display_name', 'نام' if is_fa else 'Name'),
            ('balance', 'تراز' if is_fa else 'Balance'),
            ('total_debit', 'بدهکار' if is_fa else 'Debit'),
            ('total_credit', 'بستانکار' if is_fa else 'Credit'),
            ('last_transaction_date', 'تاریخ آخرین تراکنش' if is_fa else 'Last Transaction Date'),
            ('status', 'وضعیت' if is_fa else 'Status'),
        ]
        for key, label in default_columns:
            if items and (key in items[0] or key == 'display_name'):
                keys.append(key)
                headers.append(label)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش بستانکاران" if is_fa else "Creditors Report"
    
    # RTL handling for Persian
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")
            
            # Handle display_name specially
            if key == 'display_name':
                value = (
                    item.get('display_name') or
                    item.get('alias_name') or
                    f"{item.get('first_name', '')} {item.get('last_name', '')}".strip()
                )
            
            # Format numbers
            if key in ['balance', 'total_debit', 'total_credit'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    value = num_value
                except (ValueError, TypeError):
                    pass
            
            # Format dates
            if key == 'last_transaction_date' and value:
                value = format_date_for_export(item, 'last_transaction_date')
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # RTL alignment for Persian text and numbers
            if locale == 'fa':
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, str) and any('\u0600' <= c <= '\u06FF' for c in value):
                    cell.alignment = Alignment(horizontal="right")
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "creditors_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/reports/creditors/export/pdf",
    summary="خروجی PDF گزارش بستانکاران",
    description="خروجی PDF گزارش بستانکاران با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_creditors_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی PDF گزارش بستانکاران"""
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    min_balance = body.get('min_balance')
    if min_balance is not None:
        try:
            min_balance = float(min_balance)
        except (ValueError, TypeError):
            min_balance = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    search = body.get('search')
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_creditors_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        min_balance=min_balance,
        person_ids=person_ids,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted
        if isinstance(value, str):
            if '/' in value and (len(value.split('/')) == 3):
                if '-' in value:
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            'code': '⚠️',
            'display_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'balance': '',
            'total_debit': '',
            'total_credit': '',
            'last_transaction_date': '',
            'status': '',
        }
        items.append(warning_item)
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('code', 'کد' if is_fa else 'Code'),
            ('display_name', 'نام' if is_fa else 'Name'),
            ('balance', 'تراز' if is_fa else 'Balance'),
            ('total_debit', 'بدهکار' if is_fa else 'Debit'),
            ('total_credit', 'بستانکار' if is_fa else 'Credit'),
            ('last_transaction_date', 'تاریخ آخرین تراکنش' if is_fa else 'Last Transaction Date'),
            ('status', 'وضعیت' if is_fa else 'Status'),
        ]
        for key, label in default_columns:
            if items and (key in items[0] or key == 'display_name'):
                keys.append(key)
                headers.append(label)
    
    # Load business info for header
    business_name = ""
    try:
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name
    except Exception:
        business_name = ""
    
    def esc(s: Any) -> str:
        try:
            return escape(str(s))
        except Exception:
            return str(s)
    
    # Build table rows
    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key, "")
            
            # Handle display_name specially
            if key == 'display_name':
                value = (
                    item.get('display_name') or
                    item.get('alias_name') or
                    f"{item.get('first_name', '')} {item.get('last_name', '')}".strip()
                )
            
            # Format numbers
            if key in ['balance', 'total_debit', 'total_credit'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    # Format with thousand separators
                    if is_fa:
                        value = f"{num_value:,.0f}".replace(',', '٬')
                    else:
                        value = f"{num_value:,.2f}"
                except (ValueError, TypeError):
                    pass
            
            # Format dates
            if key == 'last_transaction_date' and value:
                value = format_date_for_export(item, 'last_transaction_date')
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            tds.append(f"<td>{esc(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")
    
    headers_html = ''.join(f"<th>{esc(h)}</th>" for h in headers)
    
    # Format report datetime based on X-Calendar-Type header
    calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian")
        now = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    title_text = "گزارش بستانکاران" if is_fa else "Creditors Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    footer_text = "تولید شده توسط Hesabix" if is_fa else "Generated by Hesabix"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "
    
    # تلاش برای رندر با قالب سفارشی (persons/reports/creditors)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": title_text,
            "business_name": business_name,
            "generated_at": now,
            "is_fa": is_fa,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="persons",
            subtype="reports/creditors",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None
    
    # Inject Persian fonts (YekanBakhFaNum/Vazirmatn) for PDF rendering
    fa_font_url_regular = ""
    fa_font_url_bold = ""
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_reg, fa_bold = load_farsi_font_data_uris()
            fa_font_url_regular = fa_reg or ""
            fa_font_url_bold = fa_bold or ""
    except Exception:
        fa_font_url_regular = ""
        fa_font_url_bold = ""
    
    font_face_css = ""
    if is_fa and fa_font_url_regular:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_regular}') format('truetype'); font-weight: 400; font-style: normal; }}
        """
    if is_fa and fa_font_url_bold:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_bold}') format('truetype'); font-weight: 700; font-style: normal; }}
        """
    
    body_font_family = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if is_fa else "Arial, sans-serif"
    
    table_html = f"""
    <html lang="{html_lang}" dir="{html_dir}"> 
      <head>
        <meta charset='utf-8'>
        <style>
          {font_face_css}
          @page {{
            size: A4 landscape;
            margin: 12mm;
            @bottom-{'left' if is_fa else 'right'} {{
              content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
              font-size: 10px;
              color: #666;
              font-family: {body_font_family};
            }}
          }}
          body {{
            font-family: {body_font_family};
            font-size: 11px;
            color: #222;
          }}
          .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #444;
            padding-bottom: 6px;
          }}
          .title {{
            font-size: 16px;
            font-weight: 700;
          }}
          .meta {{
            font-size: 11px;
            color: #555;
          }}
          .table-wrapper {{
            width: 100%;
          }}
          table.report-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
          }}
          thead th {{
            background: #f0f3f7;
            border: 1px solid #c7cdd6;
            padding: 6px 4px;
            text-align: center;
            font-weight: 700;
            white-space: nowrap;
          }}
          tbody td {{
            border: 1px solid #d7dde6;
            padding: 5px 4px;
            vertical-align: top;
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: normal;
            text-align: {'right' if is_fa else 'left'};
          }}
          tbody td:has-text(number) {{
            text-align: right;
          }}
          .footer {{
            position: running(footer);
            font-size: 10px;
            color: #666;
            margin-top: 8px;
            text-align: {'left' if is_fa else 'right'};
          }}
        </style>
      </head>
      <body>
        <div class="header">
          <div>
            <div class="title">{title_text}</div>
            <div class="meta">{label_biz}: {esc(business_name)}</div>
          </div>
          <div class="meta">{label_date}: {esc(now)}</div>
        </div>
        <div class="table-wrapper">
          <table class="report-table">
            <thead>
              <tr>{headers_html}</tr>
            </thead>
            <tbody>
              {''.join(rows_html)}
            </tbody>
          </table>
        </div>
        <div class="footer">{footer_text}</div>
      </body>
    </html>
    """
    
    final_html = resolved_html or table_html
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "creditors_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/reports/people-transactions",
    summary="گزارش تراکنش‌های اشخاص",
    description="گزارش ریز دریافت‌ها و پرداخت‌ها به تفکیک شخص",
)
@require_business_access("business_id")
async def people_transactions_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """گزارش تراکنش‌های اشخاص"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    document_type = body.get('document_type')  # هر نوع سند یا None
    # پشتیبانی از همه انواع اسناد
    if document_type is not None and not isinstance(document_type, str):
        document_type = None
    
    search = body.get('search')
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_people_transactions_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        person_ids=person_ids,
        document_type=document_type,
        search=search,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="People transactions report retrieved successfully" if locale != 'fa' else "گزارش تراکنش‌های اشخاص با موفقیت دریافت شد"
    )


@router.post("/businesses/{business_id}/reports/people-transactions/export/excel",
    summary="خروجی Excel گزارش تراکنش‌های اشخاص",
    description="خروجی Excel گزارش تراکنش‌های اشخاص با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_people_transactions_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی Excel گزارش تراکنش‌های اشخاص"""
    import io
    import json
    import datetime
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response
    from app.core.i18n import negotiate_locale
    from adapters.db.models.business import Business
    from adapters.db.models.fiscal_year import FiscalYear
    from sqlalchemy import and_

    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)

    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass

    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass

    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id

    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None

    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None

    document_type = body.get('document_type')
    # پشتیبانی از همه انواع اسناد
    if document_type is not None and not isinstance(document_type, str):
        document_type = None

    search = body.get('search')

    max_export_records = 10000
    result = get_people_transactions_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        person_ids=person_ids,
        document_type=document_type,
        search=search,
        skip=0,
        take=max_export_records,
    )

    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""

    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    if len(items) >= max_export_records:
        warning_item = {
            'document_code': '⚠️',
            'person_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'document_type_name': '',
            'debit': '',
            'credit': '',
            'running_balance': '',
            'description': '',
        }
        items.append(warning_item)

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'

    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        default_columns = [
            ('document_date', 'تاریخ سند' if is_fa else 'Date'),
            ('document_code', 'کد سند' if is_fa else 'Document Code'),
            ('person_name', 'نام شخص' if is_fa else 'Person Name'),
            ('document_type_name', 'نوع سند' if is_fa else 'Document Type'),
            ('debit', 'بدهکار' if is_fa else 'Debit'),
            ('credit', 'بستانکار' if is_fa else 'Credit'),
            ('running_balance', 'تراز متحرک' if is_fa else 'Running Balance'),
            ('description', 'توضیحات' if is_fa else 'Description'),
        ]
        for key, label in default_columns:
            if items and (key in items[0] or key == 'person_name'):
                keys.append(key)
                headers.append(label)

    wb = Workbook()
    ws = wb.active
    ws.title = "People Transactions Report" if locale != 'fa' else "گزارش تراکنش‌های اشخاص"

    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            value = item.get(key, "")

            # Handle None values - convert to empty string
            if value is None:
                value = ""

            if key == 'person_name':
                value = (
                    item.get('person_name') or
                    item.get('display_name') or
                    item.get('alias_name') or
                    f"{item.get('first_name', '')} {item.get('last_name', '')}".strip()
                )

            if key in ['debit', 'credit', 'running_balance'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    value = num_value
                except (ValueError, TypeError):
                    pass

            if key == 'document_date' and value:
                value = format_date_for_export(item, 'document_date')

            # Handle description - ensure None/empty values are displayed as empty
            if key == 'description':
                if value is None or value == "None" or str(value).strip() == "":
                    value = ""

            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)

            # Final check: if value is None or "None" string, convert to empty string
            if value is None or str(value) == "None":
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            if locale == 'fa':
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, str) and any('\u0600' <= c <= '\u06FF' for c in value):
                    cell.alignment = Alignment(horizontal="right")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "people_transactions_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post("/businesses/{business_id}/reports/people-transactions/export/pdf",
    summary="خروجی PDF گزارش تراکنش‌های اشخاص",
    description="خروجی PDF گزارش تراکنش‌های اشخاص با قابلیت فیلتر، انتخاب سطرها و رعایت ترتیب/نمایش ستون‌ها",
)
@require_business_access("business_id")
async def export_people_transactions_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """خروجی PDF گزارش تراکنش‌های اشخاص"""
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    person_ids = body.get('person_ids')
    if person_ids is not None and not isinstance(person_ids, list):
        person_ids = None
    
    document_type = body.get('document_type')
    if document_type is not None and not isinstance(document_type, str):
        document_type = None
    
    search = body.get('search')
    
    # برای export، همه رکوردها را بدون pagination می‌گیریم
    max_export_records = 10000
    result = get_people_transactions_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        person_ids=person_ids,
        document_type=document_type,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted
        if isinstance(value, str):
            if '/' in value and (len(value.split('/')) == 3):
                if '-' in value:
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # Handle selected rows
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # Check if we hit the limit
    if len(items) >= max_export_records:
        warning_item = {
            'document_code': '⚠️',
            'person_name': 'حداکثر ۱۰,۰۰۰ رکورد قابل export است',
            'document_type_name': '',
            'debit': '',
            'credit': '',
            'running_balance': '',
            'description': '',
        }
        items.append(warning_item)
    
    # Get locale
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'
    
    # Prepare headers based on export_columns
    headers: List[str] = []
    keys: List[str] = []
    export_columns = body.get('export_columns')
    if export_columns:
        for col in export_columns:
            key = col.get('key')
            label = col.get('label', key)
            if key:
                keys.append(str(key))
                headers.append(str(label))
    else:
        # Default columns
        default_columns = [
            ('document_date', 'تاریخ سند' if is_fa else 'Date'),
            ('document_code', 'کد سند' if is_fa else 'Document Code'),
            ('person_name', 'نام شخص' if is_fa else 'Person Name'),
            ('document_type_name', 'نوع سند' if is_fa else 'Document Type'),
            ('debit', 'بدهکار' if is_fa else 'Debit'),
            ('credit', 'بستانکار' if is_fa else 'Credit'),
            ('running_balance', 'تراز متحرک' if is_fa else 'Running Balance'),
            ('description', 'توضیحات' if is_fa else 'Description'),
        ]
        for key, label in default_columns:
            if items and (key in items[0] or key == 'person_name'):
                keys.append(key)
                headers.append(label)
    
    # Load business info for header
    business_name = ""
    try:
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name
    except Exception:
        business_name = ""
    
    def esc(s: Any) -> str:
        try:
            return escape(str(s))
        except Exception:
            return str(s)
    
    # Build table rows
    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key, "")
            
            # Handle None values - convert to empty string
            if value is None:
                value = ""
            
            # Handle person_name specially
            if key == 'person_name':
                value = (
                    item.get('person_name') or
                    item.get('display_name') or
                    item.get('alias_name') or
                    f"{item.get('first_name', '')} {item.get('last_name', '')}".strip()
                )
            
            # Format numbers
            if key in ['debit', 'credit', 'running_balance'] and value:
                try:
                    num_value = float(value) if not isinstance(value, (int, float)) else value
                    # Format with thousand separators
                    if is_fa:
                        value = f"{num_value:,.0f}".replace(',', '٬')
                    else:
                        value = f"{num_value:,.2f}"
                except (ValueError, TypeError):
                    pass
            
            # Format dates
            if key == 'document_date' and value:
                value = format_date_for_export(item, 'document_date')
            
            # Handle description - ensure None/empty values are displayed as empty
            if key == 'description':
                if value is None or value == "None" or str(value).strip() == "":
                    value = ""
            
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = str(value)
            
            # Final check: if value is None or "None" string, convert to empty string
            if value is None or str(value) == "None":
                value = ""
            
            tds.append(f"<td>{esc(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")
    
    headers_html = ''.join(f"<th>{esc(h)}</th>" for h in headers)
    
    # Format report datetime based on X-Calendar-Type header
    calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian")
        now = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    title_text = "گزارش تراکنش‌های اشخاص" if is_fa else "People Transactions Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    footer_text = "تولید شده توسط Hesabix" if is_fa else "Generated by Hesabix"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "
    
    # تلاش برای رندر با قالب سفارشی (persons/reports/people-transactions)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": title_text,
            "business_name": business_name,
            "generated_at": now,
            "is_fa": is_fa,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="persons",
            subtype="reports/people-transactions",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None
    
    # Inject Persian fonts (YekanBakhFaNum/Vazirmatn) for PDF rendering
    fa_font_url_regular = ""
    fa_font_url_bold = ""
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_reg, fa_bold = load_farsi_font_data_uris()
            fa_font_url_regular = fa_reg or ""
            fa_font_url_bold = fa_bold or ""
    except Exception:
        fa_font_url_regular = ""
        fa_font_url_bold = ""
    
    font_face_css = ""
    if is_fa and fa_font_url_regular:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_regular}') format('truetype'); font-weight: 400; font-style: normal; }}
        """
    if is_fa and fa_font_url_bold:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_bold}') format('truetype'); font-weight: 700; font-style: normal; }}
        """
    
    body_font_family = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if is_fa else "Arial, sans-serif"
    
    table_html = f"""
    <html lang="{html_lang}" dir="{html_dir}"> 
      <head>
        <meta charset='utf-8'>
        <style>
          {font_face_css}
          @page {{
            size: A4 landscape;
            margin: 12mm;
            @bottom-{'left' if is_fa else 'right'} {{
              content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
              font-size: 10px;
              color: #666;
              font-family: {body_font_family};
            }}
          }}
          body {{
            font-family: {body_font_family};
            font-size: 11px;
            color: #222;
          }}
          .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #444;
            padding-bottom: 6px;
          }}
          .title {{
            font-size: 16px;
            font-weight: 700;
          }}
          .meta {{
            font-size: 11px;
            color: #555;
          }}
          .table-wrapper {{
            width: 100%;
          }}
          table.report-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
          }}
          thead th {{
            background: #f0f3f7;
            border: 1px solid #c7cdd6;
            padding: 6px 4px;
            text-align: center;
            font-weight: 700;
            white-space: nowrap;
          }}
          tbody td {{
            border: 1px solid #d7dde6;
            padding: 5px 4px;
            vertical-align: top;
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: normal;
            text-align: {'right' if is_fa else 'left'};
          }}
          tbody td:has-text(number) {{
            text-align: right;
          }}
          .footer {{
            position: running(footer);
            font-size: 10px;
            color: #666;
            margin-top: 8px;
            text-align: {'left' if is_fa else 'right'};
          }}
        </style>
      </head>
      <body>
        <div class="header">
          <div>
            <div class="title">{title_text}</div>
            <div class="meta">{label_biz}: {esc(business_name)}</div>
          </div>
          <div class="meta">{label_date}: {esc(now)}</div>
        </div>
        <div class="table-wrapper">
          <table class="report-table">
            <thead>
              <tr>{headers_html}</tr>
            </thead>
            <tbody>
              {''.join(rows_html)}
            </tbody>
          </table>
        </div>
        <div class="footer">{footer_text}</div>
      </body>
    </html>
    """
    
    final_html = resolved_html or table_html
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "people_transactions_report"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
