# Removed __future__ annotations to fix OpenAPI schema generation

from typing import Annotated, Dict, Any
from fastapi import APIRouter, Depends, Request, Body
from sqlalchemy.orm import Session
from sqlalchemy import and_

from adapters.db.session import get_db
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.permissions import require_business_access, require_business_permission_dep, require_business_permission_by_entity_dep
from app.core.responses import success_response, ApiError, format_datetime_fields
from app.core.cache import get_cache
from adapters.api.v1.schemas import QueryInfo
from adapters.api.v1.schema_models.product import (
    ProductCreateRequest,
    ProductUpdateRequest,
    ProductResponse,
    ProductListResponse,
    BulkPriceUpdateRequest,
    BulkPriceUpdatePreviewResponse,
    BulkDefaultWarehouseRequest,
    BulkDefaultWarehousePreviewResponse,
    BulkDefaultWarehouseApplyResponse,
)
from adapters.api.v1.schema_models.common import SuccessResponse, ErrorResponse
from app.services.product_service import (
    create_product,
    list_products,
    get_product,
    update_product,
    delete_product,
    preview_bulk_default_warehouse_update,
    apply_bulk_default_warehouse_update,
    get_item_movements_report,
    get_sales_by_product_report,
    get_inventory_kardex_report,
    get_inventory_stock_report,
)
from app.services.bulk_price_update_service import (
    preview_bulk_price_update,
    apply_bulk_price_update,
)
from adapters.db.models.business import Business
from adapters.db.models.product import Product
from app.core.i18n import negotiate_locale
from fastapi import UploadFile, File, Form, HTTPException
from adapters.api.v1.helpers.product_request_helper import process_product_request
import os


router = APIRouter(prefix="/products", tags=["محصولات و کالاها", "انبارداری"])


async def _get_products_search_query_info(request: Request) -> QueryInfo:
	"""
	خواندن پارامترهای جستجوی محصولات از body (JSON) یا در صورت خالی بودن از query string.
	برای سازگاری با پروکسی/کلاینت‌هایی که body را در query می‌فرستند.
	"""
	import json
	body_bytes = await request.body()
	if body_bytes and body_bytes.strip():
		try:
			data = json.loads(body_bytes)
			if isinstance(data, dict) and data:
				_allowed = (
					"take", "skip", "sort_by", "sort_desc", "sort", "search",
					"search_fields", "searchFields",
					"category_ids", "categoryIds",
					"filters", "include_inventory", "inventory_as_of_date",
				)
				return QueryInfo(**{k: v for k, v in data.items() if k in _allowed})
		except (json.JSONDecodeError, TypeError, ValueError):
			pass
	q = request.query_params
	def _int(name: str, default: int) -> int:
		v = q.get(name)
		if v is None:
			return default
		try:
			return int(v)
		except (ValueError, TypeError):
			return default
	def _bool(name: str, default: bool) -> bool:
		v = q.get(name)
		if v is None:
			return default
		return str(v).lower() in ("1", "true", "yes", "on")
	filters_val = q.get("filters")
	filters_parsed = None
	if filters_val:
		try:
			filters_parsed = json.loads(filters_val)
			if not isinstance(filters_parsed, list):
				filters_parsed = None
		except (json.JSONDecodeError, TypeError):
			pass
	sort_param = q.get("sort")
	sort_parsed = None
	if sort_param:
		try:
			sp = json.loads(sort_param)
			if isinstance(sp, list):
				sort_parsed = sp
		except (json.JSONDecodeError, TypeError, ValueError):
			pass
	return QueryInfo(
		take=max(1, min(1000, _int("take", 20))),
		skip=max(0, _int("skip", 0)),
		sort_by=q.get("sort_by") or None,
		sort_desc=_bool("sort_desc", False),
		sort=sort_parsed,
		search=q.get("search") or None,
		search_fields=None,
		filters=filters_parsed,
		include_inventory=_bool("include_inventory", False),
		inventory_as_of_date=q.get("inventory_as_of_date") or None,
	)


def _ensure_products_pagination(result: Dict[str, Any], take: int, skip: int) -> None:
	"""
	اطمینان از وجود فیلدهای pagination (total_count, has_more) در پاسخ.
	برای پاسخ تازه از list_products معمولاً وجود دارند؛ برای کش قدیمی ممکن است نباشند.
	"""
	if not isinstance(result, dict):
		return
	items = result.get("items") or []
	if "total_count" not in result:
		result["total_count"] = len(items)
	if "has_more" not in result:
		total = result.get("total_count", len(items))
		result["has_more"] = (skip + take) < total


@router.post(
    "/business/{business_id}",
    summary="ایجاد محصول جدید",
    description="""
    ایجاد یک محصول یا خدمت جدید برای کسب‌وکار
    
    ### ویژگی‌ها:
    - پشتیبانی از آپلود تصویر (multipart/form-data یا JSON)
    - پشتیبانی از 60+ فیلد شامل: قیمت‌گذاری، موجودی، مالیات، ویژگی‌ها
    - اگر کد محصول ارسال نشود، به صورت خودکار تولید می‌شود
    
    ### نکات مهم:
    - برای آپلود تصویر از multipart/form-data استفاده کنید
    - فیلد `item_type` می‌تواند "کالا" یا "خدمت" باشد
    - برای کالاهای با کنترل موجودی، `track_inventory` را true قرار دهید
    """,
    response_model=SuccessResponse[ProductResponse],
    responses={
        200: {
            "description": "محصول با موفقیت ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "محصول با موفقیت ایجاد شد",
                        "data": {
                            "id": 123,
                            "code": "P1001",
                            "name": "لپ‌تاپ Asus Vivobook 15",
                            "item_type": "کالا",
                            "description": "لپ‌تاپ 15.6 اینچی با پردازنده Core i5",
                            "category_id": 1,
                            "category_name": "لپ‌تاپ",
                            "main_unit": "عدد",
                            "base_sales_price": 15000000,
                            "base_purchase_price": 12000000,
                            "track_inventory": True,
                            "default_warehouse_id": 1,
                            "is_sales_taxable": True,
                            "sales_tax_rate": 9,
                            "is_active": True,
                            "created_at": "2024-01-15T10:30:00Z"
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی داده‌ها یا محدودیت ذخیره‌سازی",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "VALIDATION_ERROR",
                        "message": "نام محصول الزامی است"
                    }
                }
            }
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.add"
        },
        404: {
            "description": "کسب‌وکار یافت نشد"
        }
    }
)
@require_business_access("business_id")
async def create_product_endpoint(
    request: Request,
    business_id: int,
    payload: ProductCreateRequest = None,
    file: UploadFile = File(None),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "add")),
) -> Dict[str, Any]:
    
    # استفاده از helper function برای پردازش درخواست
    processed_payload, processed_file, image_file_id = await process_product_request(
        request=request,
        business_id=business_id,
        ctx=ctx,
        db=db,
        is_update=False,
    )
    
    # اگر payload از helper آمده، از آن استفاده می‌کنیم
    if processed_payload:
        payload = processed_payload
    
    if not payload:
        raise ApiError("INVALID_PAYLOAD", "داده‌های محصول ارسال نشده است", http_status=400)
    
    result = create_product(db, business_id, payload)
    
    # به‌روزرسانی context_id فایل با product_id
    if image_file_id and result.get("data", {}).get("id"):
        from adapters.db.models.file_storage import FileStorage
        file_storage = db.query(FileStorage).filter(FileStorage.id == image_file_id).first()
        if file_storage:
            file_storage.context_id = str(result["data"]["id"])
            db.commit()
    
    return success_response(data=format_datetime_fields(result["data"], request), request=request, message=result.get("message"))


@router.post(
    "/business/{business_id}/search",
    summary="جستجو و فیلتر محصولات",
    description="""
    جستجو، فیلتر و لیست‌بندی محصولات با قابلیت‌های پیشرفته
    
    ### قابلیت‌ها:
    - جستجو در چندین فیلد (نام، کد، توضیحات)
    - فیلتر بر اساس دسته‌بندی، نوع محصول، وضعیت فعال/غیرفعال
    - مرتب‌سازی بر اساس فیلدهای مختلف
    - صفحه‌بندی نتایج
    - فیلتر بر اساس دسته‌بندی (category_ids)
    
    ### نکات:
    - حداکثر تعداد رکورد در هر درخواست: 1000 (پارامتر `take`)
    - نتایج جستجو برای مدت کوتاهی cache می‌شوند
    - برای جستجوی دقیق‌تر از فیلترها استفاده کنید
    """,
    response_model=SuccessResponse[ProductListResponse],
    responses={
        200: {
            "description": "لیست محصولات با موفقیت دریافت شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "لیست محصولات دریافت شد",
                        "data": {
                            "items": [
                                {
                                    "id": 1,
                                    "code": "P1001",
                                    "name": "لپ‌تاپ Asus Vivobook 15",
                                    "item_type": "کالا",
                                    "category_name": "لپ‌تاپ",
                                    "base_sales_price": 15000000,
                                    "is_active": True
                                }
                            ],
                            "total_count": 100,
                            "has_more": True
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی پارامترها"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.view"
        },
        404: {
            "description": "کسب‌وکار یافت نشد"
        },
        500: {
            "description": "خطای داخلی سرور"
        }
    }
)
@require_business_access("business_id")
def search_products_endpoint(
	request: Request,
	business_id: int,
	query_info: Annotated[QueryInfo, Depends(_get_products_search_query_info)],
	ctx: AuthContext = Depends(get_current_user),
	db: Session = Depends(get_db),
	_: None = Depends(require_business_permission_dep("products", "view")),
) -> Dict[str, Any]:
	import logging
	from app.core.responses import ApiError
	from sqlalchemy.exc import SQLAlchemyError, OperationalError, TimeoutError as SQLTimeoutError
	
	logger = logging.getLogger(__name__)
	
	# کش نتایج جستجوی محصولات بر اساس پارامترها
	cache = get_cache()
	cache_key = None
	category_id = None

	if cache.enabled:
		import json, hashlib
		from decimal import Decimal
		from datetime import datetime, date
		from adapters.api.v1.schemas import FilterItem
		
		# Helper function to convert objects to JSON-serializable format
		def to_serializable(obj):
			"""Convert Pydantic models and other non-serializable types to dict/primitive types"""
			if hasattr(obj, 'model_dump'):  # Pydantic v2
				return obj.model_dump()
			elif hasattr(obj, 'dict'):  # Pydantic v1
				return obj.dict()
			elif isinstance(obj, (datetime, date)):
				return obj.isoformat()
			elif isinstance(obj, Decimal):
				return float(obj)
			elif isinstance(obj, dict):
				return {k: to_serializable(v) for k, v in obj.items()}
			elif isinstance(obj, list):
				return [to_serializable(item) for item in obj]
			else:
				return obj
		
		# استخراج category_id از filters اگر موجود باشد
		if query_info.filters:
			for filter_item in query_info.filters:
				# Convert FilterItem to dict if needed
				if isinstance(filter_item, FilterItem):
					filter_dict = to_serializable(filter_item)
				else:
					filter_dict = filter_item
				
				if isinstance(filter_dict, dict):
					field = filter_dict.get("property") or filter_dict.get("field")
					if field == "category_id":
						value = filter_dict.get("value")
						operator = filter_dict.get("operator", "=")
						if operator == "=" and value:
							try:
								category_id = int(value)
							except (ValueError, TypeError):
								pass
						# اگر operator "in" است، اولین category_id را می‌گیریم
						# (برای سادگی، تمام کش‌های مربوط به این category ها invalidate می‌شوند)
						elif operator == "in" and isinstance(value, list) and value:
							try:
								category_id = int(value[0])
							except (ValueError, TypeError):
								pass
						break
		
		if category_id is None and getattr(query_info, "category_ids", None):
			for cid in query_info.category_ids:
				try:
					category_id = int(cid)
					break
				except (ValueError, TypeError):
					pass
		
		# Convert filters to serializable format
		serializable_filters = None
		if query_info.filters:
			serializable_filters = [to_serializable(f) for f in query_info.filters]
		
		key_payload = {
			"business_id": business_id,
			"take": query_info.take,
			"skip": query_info.skip,
			"sort_by": query_info.sort_by,
			"sort_desc": query_info.sort_desc,
			"sort": to_serializable(query_info.sort) if getattr(query_info, "sort", None) else None,
			"search": query_info.search,
			"filters": serializable_filters,
			"category_ids": to_serializable(query_info.category_ids) if query_info.category_ids else None,
			"include_inventory": query_info.include_inventory,
			"inventory_as_of_date": query_info.inventory_as_of_date,
		}
		key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
		key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
		cache_key = f"products_search:{business_id}:{ctx.get_user_id()}:{key_hash}"
		cached = cache.get(cache_key)
		if cached is not None:
			# پاسخ کش‌شده ممکن است فاقد pagination باشد (کش قدیمی)؛ برای فرانت همیشه pagination لازم است
			_ensure_products_pagination(cached, query_info.take, query_info.skip)
			return success_response(data=cached, request=request)

	try:
		result = list_products(db, business_id, {
			"take": query_info.take,
			"skip": query_info.skip,
			"sort_by": query_info.sort_by,
			"sort_desc": query_info.sort_desc,
			"sort": [s.model_dump() for s in query_info.sort] if query_info.sort else None,
			"search": query_info.search,
			"filters": query_info.filters,
			"category_ids": query_info.category_ids,
			"include_inventory": query_info.include_inventory,
			"inventory_as_of_date": query_info.inventory_as_of_date,
		})
		_ensure_products_pagination(result, query_info.take, query_info.skip)
		formatted = format_datetime_fields(result, request)

		if cache.enabled and cache_key:
			# استفاده از tag-based caching برای مدیریت بهتر invalidation
			# چون موجودی و قیمت ممکن است سریع تغییر کند، TTL کوتاه
			cache.set_with_products_tag(
				key=cache_key,
				value=formatted,
				business_id=business_id,
				category_id=category_id,
				ttl=30
			)

		return success_response(data=formatted, request=request)
	except (OperationalError, SQLTimeoutError) as e:
		# Timeout یا خطای اتصال - rollback و بستن session
		logger.error(
			f"Database timeout/connection error in search_products for business_id={business_id}: {str(e)}",
			exc_info=True,
			extra={
				"business_id": business_id,
				"user_id": ctx.get_user_id() if ctx else None,
				"path": request.url.path,
			}
		)
		try:
			db.rollback()
		except Exception:
			pass
		raise ApiError(
			"DATABASE_TIMEOUT",
			"زمان اجرای درخواست به پایان رسید. لطفاً فیلترهای جستجو را محدودتر کنید یا بعداً تلاش کنید.",
			http_status=504
		)
	except SQLAlchemyError as e:
		logger.error(
			f"Database error in search_products for business_id={business_id}",
			exc_info=True,
			extra={
				"business_id": business_id,
				"user_id": ctx.get_user_id() if ctx else None,
				"path": request.url.path,
				"query": {
					"take": query_info.take,
					"skip": query_info.skip,
					"search": query_info.search,
				}
			}
		)
		try:
			db.rollback()
		except Exception:
			pass
		raise ApiError(
			"DATABASE_ERROR",
			"خطا در ارتباط با پایگاه داده. لطفاً بعداً تلاش کنید.",
			http_status=500
		)
	except Exception as e:
		logger.error(
			f"Unexpected error in search_products for business_id={business_id}: {str(e)}",
			exc_info=True,
			extra={
				"business_id": business_id,
				"user_id": ctx.get_user_id() if ctx else None,
				"path": request.url.path,
			}
		)
		try:
			db.rollback()
		except Exception:
			pass
		raise ApiError(
			"INTERNAL_SERVER_ERROR",
			"خطای داخلی سرور رخ داد. لطفاً با پشتیبانی تماس بگیرید.",
			http_status=500
		)


@router.get(
    "/business/{business_id}/{product_id}",
    summary="دریافت اطلاعات محصول",
    description="""
    دریافت اطلاعات کامل یک محصول یا خدمت
    
    ### اطلاعات برگشتی:
    - اطلاعات پایه محصول (کد، نام، نوع، توضیحات)
    - اطلاعات دسته‌بندی
    - قیمت‌گذاری (فروش و خرید)
    - تنظیمات موجودی و انبار
    - اطلاعات مالیات
    - تصویر محصول (در صورت وجود)
    - ویژگی‌های محصول
    - موجودی در انبارها (اختیاری)
    """,
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "اطلاعات محصول با موفقیت دریافت شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "item": {
                                "id": 123,
                                "code": "P1001",
                                "name": "لپ‌تاپ Asus Vivobook 15",
                                "item_type": "کالا",
                                "description": "لپ‌تاپ 15.6 اینچی با پردازنده Core i5",
                                "category_id": 1,
                                "category_name": "لپ‌تاپ",
                                "main_unit": "عدد",
                                "base_sales_price": 15000000,
                                "base_purchase_price": 12000000,
                                "track_inventory": True,
                                "default_warehouse_id": 1,
                                "default_warehouse_name": "انبار اصلی",
                                "total_quantity": 25,
                                "total_value": 300000000,
                                "is_sales_taxable": True,
                                "sales_tax_rate": 9,
                                "is_active": True,
                                "created_at": "2024-01-15T10:30:00Z",
                                "updated_at": "2024-01-20T14:20:00Z"
                            }
                        }
                    }
                }
            }
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.view"
        },
        404: {
            "description": "محصول یافت نشد",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "NOT_FOUND",
                        "message": "Product not found"
                    }
                }
            }
        }
    }
)
@require_business_access("business_id")
def get_product_endpoint(
    request: Request,
    business_id: int,
    product_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_by_entity_dep("products", "view", Product, "product_id")),
) -> Dict[str, Any]:
    item = get_product(db, product_id, business_id)
    if not item:
        raise ApiError("NOT_FOUND", "Product not found", http_status=404)
    return success_response(data=format_datetime_fields({"item": item}, request), request=request)


@router.put(
    "/business/{business_id}/{product_id}",
    summary="ویرایش محصول",
    description="""
    ویرایش اطلاعات یک محصول یا خدمت موجود
    
    ### ویژگی‌ها:
    - پشتیبانی از آپلود تصویر جدید (multipart/form-data یا JSON)
    - تمام فیلدها اختیاری هستند (فقط فیلدهای ارسال شده به‌روزرسانی می‌شوند)
    - تصویر قبلی در صورت آپلود تصویر جدید حذف می‌شود
    
    ### نکات مهم:
    - برای آپلود تصویر از multipart/form-data استفاده کنید
    - تغییر کد محصول ممکن است روی فاکتورهای موجود تأثیر بگذارد
    - تغییر `track_inventory` ممکن است نیاز به بررسی موجودی داشته باشد
    """,
    response_model=SuccessResponse[ProductResponse],
    responses={
        200: {
            "description": "محصول با موفقیت به‌روزرسانی شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "محصول با موفقیت به‌روزرسانی شد",
                        "data": {
                            "id": 123,
                            "code": "P1001",
                            "name": "لپ‌تاپ Asus Vivobook 15 (ویرایش شده)",
                            "item_type": "کالا",
                            "base_sales_price": 15500000,
                            "updated_at": "2024-01-20T14:20:00Z"
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی داده‌ها یا محدودیت ذخیره‌سازی"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.edit"
        },
        404: {
            "description": "محصول یافت نشد",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "NOT_FOUND",
                        "message": "Product not found"
                    }
                }
            }
        }
    }
)
@require_business_access("business_id")
async def update_product_endpoint(
    request: Request,
    business_id: int,
    product_id: int,
    file: UploadFile | None = File(None),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_by_entity_dep("products", "edit", Product, "product_id")),
) -> Dict[str, Any]:
    
    # بررسی وجود محصول
    from adapters.db.models.product import Product
    product = db.get(Product, product_id)
    if not product or product.business_id != business_id:
        raise ApiError("NOT_FOUND", "Product not found", http_status=404)
    
    # بررسی اینکه آیا multipart/form-data است یا JSON
    content_type = request.headers.get("content-type", "")
    is_multipart = "multipart/form-data" in content_type
    
    image_file_id = None
    old_image_file_id = product.image_file_id
    payload: ProductUpdateRequest | None = None
    
    # اگر multipart/form-data است، فایل و داده‌ها را از form می‌خوانیم
    if is_multipart:
        form_data = await request.form()
        
        # آپلود فایل اگر وجود دارد
        if "file" in form_data:
            file = form_data["file"]
            if hasattr(file, 'filename') and file.filename:
                # بررسی فرمت فایل
                allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}
                file_ext = os.path.splitext(file.filename)[1].lower()
                if file_ext not in allowed_extensions:
                    raise ApiError("INVALID_FILE_FORMAT", "فرمت فایل معتبر نیست. فقط فرمت‌های JPG, PNG, GIF, WebP و BMP پشتیبانی می‌شوند", http_status=400)
                
                # آپلود فایل جدید
                from app.services.file_storage_service import FileStorageService
                storage_service = FileStorageService(db)
                try:
                    upload_result = await storage_service.upload_file(
                        file=file,
                        user_id=ctx.get_user_id(),  # user_id به صورت int ارسال می‌شود
                        module_context="products",
                        context_id=str(product_id),
                        developer_data={"business_id": business_id, "product_id": product_id},
                        is_temporary=False,
                        expires_in_days=3650,
                        business_id=business_id,
                        check_storage_limit=True,
                    )
                    image_file_id = upload_result.get("file_id")
                except HTTPException as e:
                    # اگر خطای محدودیت ذخیره‌سازی باشد، جزئیات را برمی‌گردانیم
                    if e.status_code == 400 and isinstance(e.detail, dict) and e.detail.get("error") == "STORAGE_LIMIT_EXCEEDED":
                        error_detail = {
                            "success": False,
                            "error": {
                                "code": "STORAGE_LIMIT_EXCEEDED",
                                "message": e.detail.get("message", "حجم فایل از محدودیت ذخیره‌سازی تجاوز می‌کند"),
                                "total_limit_gb": e.detail.get("total_limit_gb"),
                                "current_usage_gb": e.detail.get("current_usage_gb"),
                                "available_gb": e.detail.get("available_gb"),
                                "required_gb": e.detail.get("required_gb"),
                                "over_usage_gb": e.detail.get("over_usage_gb"),
                            }
                        }
                        raise HTTPException(status_code=400, detail=error_detail)
                    raise ApiError("FILE_UPLOAD_ERROR", f"خطا در آپلود فایل: {str(e.detail)}", http_status=400)
                except Exception as e:
                    raise ApiError("FILE_UPLOAD_ERROR", f"خطا در آپلود فایل: {str(e)}", http_status=400)
        
        # ساخت payload از form data
        import json
        # فیلدهای string که نباید به int تبدیل شوند
        string_fields = {
            'code', 'name', 'description', 'main_unit', 'secondary_unit',
            'base_sales_note', 'base_purchase_note', 'tax_code', 'image_file_id'
        }
        product_data = {}
        for key, value in form_data.items():
            if key != "file":
                if isinstance(value, str):
                    # سعی می‌کنیم به عنوان JSON parse کنیم
                    try:
                        parsed = json.loads(value)
                        product_data[key] = parsed
                    except (json.JSONDecodeError, ValueError):
                        # اگر JSON نیست، بررسی می‌کنیم که آیا boolean یا number است
                        value_lower = value.strip().lower()
                        if value_lower in ('true', 'false'):
                            product_data[key] = value_lower == 'true'
                        elif value_lower == 'null' or value_lower == '':
                            product_data[key] = None
                        elif key in string_fields:
                            # فیلدهای string را به صورت string نگه می‌داریم
                            product_data[key] = value
                        elif value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                            product_data[key] = int(value)
                        elif value.replace('.', '', 1).replace('-', '', 1).isdigit():
                            product_data[key] = float(value)
                        else:
                            product_data[key] = value
                else:
                    product_data[key] = value
        
        try:
            payload = ProductUpdateRequest(**product_data)
        except Exception as e:
            raise ApiError("INVALID_PAYLOAD", f"خطا در پردازش داده‌ها: {str(e)}", http_status=400)
    else:
        # اگر JSON است، payload را از body می‌خوانیم
        try:
            body_data = await request.json()
            if not isinstance(body_data, dict):
                raise ApiError("INVALID_PAYLOAD", "داده‌های ارسالی باید یک object JSON باشد", http_status=400)
            # اگر default_warehouse_id در body_data وجود دارد (حتی اگر null باشد)، آن را به صورت صریح set می‌کنیم
            # تا Pydantic آن را در fields_set قرار دهد
            default_warehouse_id_value = body_data.get('default_warehouse_id')
            if 'default_warehouse_id' in body_data:
                # اگر null است، آن را به صورت صریح None set می‌کنیم
                body_data['default_warehouse_id'] = default_warehouse_id_value
            payload = ProductUpdateRequest(**body_data)
            # اضافه کردن به fields_set برای Pydantic v2 (حتی اگر null باشد)
            if 'default_warehouse_id' in body_data:
                if hasattr(payload, 'model_fields_set'):
                    payload.model_fields_set.add('default_warehouse_id')
                elif hasattr(payload, '__fields_set__'):
                    payload.__fields_set__.add('default_warehouse_id')
        except ValueError as e:
            raise ApiError("INVALID_PAYLOAD", f"خطا در parse کردن JSON: {str(e)}", http_status=400)
        except Exception as e:
            raise ApiError("INVALID_PAYLOAD", f"خطا در پردازش داده‌های JSON: {str(e)}", http_status=400)
        
        # اگر فایل هم ارسال شده، آن را پردازش می‌کنیم
        if file and file.filename:
            # بررسی فرمت فایل
            allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}
            file_ext = os.path.splitext(file.filename)[1].lower()
            if file_ext not in allowed_extensions:
                raise ApiError("INVALID_FILE_FORMAT", "فرمت فایل معتبر نیست. فقط فرمت‌های JPG, PNG, GIF, WebP و BMP پشتیبانی می‌شوند", http_status=400)
            
            # آپلود فایل جدید
            from app.services.file_storage_service import FileStorageService
            storage_service = FileStorageService(db)
            try:
                upload_result = await storage_service.upload_file(
                    file=file,
                    user_id=ctx.get_user_id(),  # user_id به صورت int ارسال می‌شود
                    module_context="products",
                    context_id=str(product_id),
                    developer_data={"business_id": business_id, "product_id": product_id},
                    is_temporary=False,
                    expires_in_days=3650,
                    business_id=business_id,
                    check_storage_limit=True,
                )
                image_file_id = upload_result.get("file_id")
            except HTTPException as e:
                # اگر خطای محدودیت ذخیره‌سازی باشد، جزئیات را برمی‌گردانیم
                if e.status_code == 400 and isinstance(e.detail, dict) and e.detail.get("error") == "STORAGE_LIMIT_EXCEEDED":
                    error_detail = {
                        "success": False,
                        "error": {
                            "code": "STORAGE_LIMIT_EXCEEDED",
                            "message": e.detail.get("message", "حجم فایل از محدودیت ذخیره‌سازی تجاوز می‌کند"),
                            "total_limit_gb": e.detail.get("total_limit_gb"),
                            "current_usage_gb": e.detail.get("current_usage_gb"),
                            "available_gb": e.detail.get("available_gb"),
                            "required_gb": e.detail.get("required_gb"),
                            "over_usage_gb": e.detail.get("over_usage_gb"),
                        }
                    }
                    raise HTTPException(status_code=400, detail=error_detail)
                raise ApiError("FILE_UPLOAD_ERROR", f"خطا در آپلود فایل: {str(e.detail)}", http_status=400)
            except Exception as e:
                raise ApiError("FILE_UPLOAD_ERROR", f"خطا در آپلود فایل: {str(e)}", http_status=400)
    
    # تنظیم image_file_id در payload
    if image_file_id and payload:
        payload.image_file_id = image_file_id
    
    if not payload:
        raise ApiError("INVALID_PAYLOAD", "داده‌های محصول ارسال نشده است", http_status=400)
    
    result = update_product(db, product_id, business_id, payload)
    if not result:
        raise ApiError("NOT_FOUND", "Product not found", http_status=404)
    
    # حذف عکس قبلی اگر عکس جدید آپلود شده
    if image_file_id and old_image_file_id and old_image_file_id != image_file_id:
        from app.services.file_storage_service import FileStorageService
        from uuid import UUID
        storage_service = FileStorageService(db)
        try:
            await storage_service.delete_file(UUID(old_image_file_id))
        except Exception:
            pass  # اگر حذف فایل با خطا مواجه شد، ادامه می‌دهیم
    
    return success_response(data=format_datetime_fields(result["data"], request), request=request, message=result.get("message"))


@router.delete(
    "/business/{business_id}/{product_id}",
    summary="حذف محصول",
    description="""
    حذف یک محصول یا خدمت از سیستم
    
    ### نکات مهم:
    - محصولات که در فاکتورها استفاده شده‌اند ممکن است قابل حذف نباشند
    - در صورت وجود وابستگی‌ها، خطا برگردانده می‌شود
    - تصویر محصول نیز همراه با محصول حذف می‌شود
    - این عملیات غیرقابل بازگشت است
    """,
    response_model=SuccessResponse[Dict[str, bool]],
    responses={
        200: {
            "description": "محصول با موفقیت حذف شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "کالا با موفقیت حذف شد",
                        "data": {
                            "deleted": True
                        }
                    }
                }
            }
        },
        400: {
            "description": "محصول قابل حذف نیست (به دلیل وابستگی‌ها)",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "HAS_DEPENDENCIES",
                        "message": "این محصول در فاکتورها استفاده شده و قابل حذف نیست"
                    }
                }
            }
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.delete"
        },
        404: {
            "description": "محصول یافت نشد",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "NOT_FOUND",
                        "message": "کالا یافت نشد"
                    }
                }
            }
        }
    }
)
@require_business_access("business_id")
def delete_product_endpoint(
    request: Request,
    business_id: int,
    product_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_by_entity_dep("products", "delete", Product, "product_id")),
) -> Dict[str, Any]:
    from fastapi import HTTPException
    
    success, error_message = delete_product(db, product_id, business_id)
    if not success:
        if error_message:
            raise HTTPException(status_code=400, detail=error_message)
        raise HTTPException(status_code=404, detail="کالا یافت نشد")
    
    return success_response({"deleted": True}, request, message="کالا با موفقیت حذف شد")


@router.post(
    "/business/{business_id}/bulk-delete",
    summary="حذف گروهی محصولات",
    description="""
    حذف چندین محصول به صورت همزمان بر اساس شناسه‌ها یا کدها
    
    ### ویژگی‌ها:
    - حذف بر اساس شناسه‌ها (ids) یا کدها (codes) یا هر دو
    - در صورت وجود وابستگی، محصول نادیده گرفته می‌شود (skipped)
    - نتیجه شامل تعداد حذف شده، نادیده گرفته شده و خطاها است
    
    ### نکات:
    - محصولاتی که در فاکتورها استفاده شده‌اند قابل حذف نیستند
    - تصاویر محصولات حذف شده نیز حذف می‌شوند
    - این عملیات غیرقابل بازگشت است
    """,
    response_model=SuccessResponse[Dict[str, Any]],
    responses={
        200: {
            "description": "عملیات حذف گروهی انجام شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "deleted": 5,
                            "skipped": 2,
                            "errors": [
                                {
                                    "product_id": 123,
                                    "code": "P1001",
                                    "message": "این محصول در فاکتورها استفاده شده است"
                                }
                            ]
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی داده‌ها - ids و codes نباید هر دو خالی باشند"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.delete"
        }
    }
)
@require_business_access("business_id")
def bulk_delete_products_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any],
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "delete")),
) -> Dict[str, Any]:

    from sqlalchemy import and_ as _and
    from adapters.db.models.product import Product

    ids = body.get("ids")
    codes = body.get("codes")
    deleted = 0
    skipped = 0
    errors = []

    if not ids and not codes:
        return success_response({"deleted": 0, "skipped": 0, "errors": []}, request)

    # Normalize inputs
    if isinstance(ids, list):
        ids = [int(x) for x in ids if isinstance(x, (int, str)) and str(x).isdigit()]
    else:
        ids = []
    if isinstance(codes, list):
        codes = [str(x).strip() for x in codes if str(x).strip()]
    else:
        codes = []

    # Delete by IDs first
    if ids:
        for pid in ids:
            try:
                product = db.get(Product, pid)
                if not product or product.business_id != business_id:
                    skipped += 1
                    errors.append(f"کالا با شناسه {pid} یافت نشد")
                    continue
                
                success, error_message = delete_product(db, pid, business_id)
                if success:
                    deleted += 1
                else:
                    skipped += 1
                    if error_message:
                        errors.append(f"کالا {product.name or product.code or pid}: {error_message}")
                    else:
                        errors.append(f"کالا {product.name or product.code or pid}: امکان حذف وجود ندارد")
            except Exception as e:
                skipped += 1
                errors.append(f"خطا در حذف کالا با شناسه {pid}: {str(e)}")

    # Delete by codes
    if codes:
        try:
            items = db.query(Product).filter(_and(Product.business_id == business_id, Product.code.in_(codes))).all()
            for obj in items:
                try:
                    success, error_message = delete_product(db, obj.id, business_id)
                    if success:
                        deleted += 1
                    else:
                        skipped += 1
                        if error_message:
                            errors.append(f"کالا {obj.name or obj.code or obj.id}: {error_message}")
                        else:
                            errors.append(f"کالا {obj.name or obj.code or obj.id}: امکان حذف وجود ندارد")
                except Exception as e:
                    skipped += 1
                    errors.append(f"خطا در حذف کالا {obj.name or obj.code or obj.id}: {str(e)}")
        except Exception as e:
            # In case of query issues, treat all as skipped
            skipped += len(codes)
            errors.append(f"خطا در جستجوی کالاها: {str(e)}")

    return success_response({
        "deleted": deleted, 
        "skipped": skipped,
        "errors": errors
    }, request)

@router.post(
    "/business/{business_id}/export/excel",
    summary="خروجی Excel لیست محصولات",
    description="""
    دریافت لیست محصولات به صورت فایل Excel
    
    ### قابلیت‌ها:
    - فیلتر محصولات بر اساس معیارهای مختلف
    - انتخاب ستون‌های دلخواه برای خروجی
    - مرتب‌سازی داده‌ها
    - انتخاب محصولات خاص (selected_indices)
    - پشتیبانی از RTL برای زبان فارسی
    
    ### فرمت فایل:
    - فرمت: `.xlsx` (Excel 2007+)
    - نام فایل شامل نام کسب‌وکار و timestamp است
    """,
    responses={
        200: {
            "description": "فایل Excel با موفقیت ایجاد شد",
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                    "schema": {
                        "type": "string",
                        "format": "binary"
                    }
                }
            },
            "headers": {
                "Content-Disposition": {
                    "description": "نام فایل خروجی",
                    "schema": {
                        "type": "string",
                        "example": "attachment; filename=products_business_20240115_103000.xlsx"
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.export"
        }
    }
)
@require_business_access("business_id")
async def export_products_excel(
    request: Request,
    business_id: int,
    body: dict,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "export")),
):
    import io
    import re
    import datetime
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query_dict = {
        "take": int(body.get("take", 1000)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields") or body.get("searchFields"),
        "filters": body.get("filters"),
        "category_ids": body.get("category_ids") or body.get("categoryIds"),
    }
    result = list_products(db, business_id, query_dict)
    items = result.get("items", []) if isinstance(result, dict) else result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Apply selected indices filter if requested
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None and isinstance(items, list):
        indices = None
        if isinstance(selected_indices, str):
            try:
                import json as _json
                indices = _json.loads(selected_indices)
            except Exception:
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        headers = [col.get("label") or col.get("key") for col in export_columns]
        keys = [col.get("key") for col in export_columns]
    else:
        default_cols = [
            ("code", "کد"),
            ("name", "نام"),
            ("item_type", "نوع"),
            ("category_id", "دسته"),
            ("base_sales_price", "قیمت فروش"),
            ("base_purchase_price", "قیمت خرید"),
            ("main_unit", "واحد اصلی"),
            ("secondary_unit", "واحد فرعی"),
            ("track_inventory", "کنترل موجودی"),
            ("created_at_formatted", "ایجاد"),
        ]
        keys = [k for k, _ in default_cols]
        headers = [v for _, v in default_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Locale and RTL/LTR handling for Excel
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for it in items:
        row = []
        for k in keys:
            row.append(it.get(k))
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            # Align data cells based on locale
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")

    # Auto width columns
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass

    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()

    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    base = "products"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(data)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/business/{business_id}/import/template",
    summary="دانلود تمپلیت ایمپورت محصولات",
    description="""
    دریافت فایل Excel تمپلیت برای ایمپورت محصولات
    
    ### ویژگی‌ها:
    - شامل تمام ستون‌های مورد نیاز برای ایمپورت
    - هدرهای فارسی و انگلیسی
    - نمونه داده برای راهنمایی
    - پشتیبانی از RTL برای زبان فارسی
    
    ### استفاده:
    1. این endpoint را فراخوانی کنید تا تمپلیت را دریافت کنید
    2. فایل را با اطلاعات محصولات پر کنید
    3. فایل را از طریق endpoint `/import/excel` آپلود کنید
    
    ### فرمت فایل:
    - فرمت: `.xlsx` (Excel 2007+)
    """,
    responses={
        200: {
            "description": "فایل تمپلیت با موفقیت ایجاد شد",
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                    "schema": {
                        "type": "string",
                        "format": "binary"
                    }
                }
            },
            "headers": {
                "Content-Disposition": {
                    "description": "نام فایل تمپلیت",
                    "schema": {
                        "type": "string",
                        "example": "attachment; filename=products_import_template.xlsx"
                    }
                }
            }
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.edit"
        }
    }
)
@require_business_access("business_id")
async def download_products_import_template(
    request: Request,
    business_id: int,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "edit")),
):
    import io
    import datetime
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    # Template headers should be user-friendly and localized.
    # Import endpoint will map these localized headers back to internal keys.
    #
    # NOTE: For reference fields (category, tax, attributes), we provide both ID columns
    # and human-friendly columns (name/code). Users can fill either; import will resolve.
    columns = [
        ("code", {"fa": "کد", "en": "Code"}),
        ("name", {"fa": "نام", "en": "Name"}),
        ("item_type", {"fa": "نوع", "en": "Type"}),
        ("description", {"fa": "توضیحات", "en": "Description"}),
        ("category_id", {"fa": "شناسه دسته‌بندی", "en": "Category ID"}),
        ("category_path", {"fa": "مسیر دسته‌بندی", "en": "Category Path"}),
        ("main_unit", {"fa": "واحد اصلی", "en": "Main Unit"}),
        ("secondary_unit", {"fa": "واحد فرعی", "en": "Secondary Unit"}),
        ("unit_conversion_factor", {"fa": "ضریب تبدیل", "en": "Unit Conversion Factor"}),
        ("base_sales_price", {"fa": "قیمت فروش", "en": "Sales Price"}),
        ("base_purchase_price", {"fa": "قیمت خرید", "en": "Purchase Price"}),
        ("track_inventory", {"fa": "کنترل موجودی", "en": "Track Inventory"}),
        ("reorder_point", {"fa": "نقطه سفارش مجدد", "en": "Reorder Point"}),
        ("min_order_qty", {"fa": "حداقل مقدار سفارش", "en": "Min Order Qty"}),
        ("lead_time_days", {"fa": "زمان تامین (روز)", "en": "Lead Time (Days)"}),
        ("is_sales_taxable", {"fa": "مشمول مالیات فروش", "en": "Sales Taxable"}),
        ("is_purchase_taxable", {"fa": "مشمول مالیات خرید", "en": "Purchase Taxable"}),
        ("sales_tax_rate", {"fa": "نرخ مالیات فروش (%)", "en": "Sales Tax Rate (%)"}),
        ("purchase_tax_rate", {"fa": "نرخ مالیات خرید (%)", "en": "Purchase Tax Rate (%)"}),
        ("tax_type_id", {"fa": "شناسه نوع مالیات", "en": "Tax Type ID"}),
        ("tax_type_code", {"fa": "کد نوع مالیات", "en": "Tax Type Code"}),
        ("tax_type_title", {"fa": "عنوان نوع مالیات", "en": "Tax Type Title"}),
        ("tax_code", {"fa": "کد مالیاتی", "en": "Tax Code"}),
        ("tax_unit_id", {"fa": "شناسه واحد مالیاتی", "en": "Tax Unit ID"}),
        ("tax_unit_code", {"fa": "کد واحد مالیاتی", "en": "Tax Unit Code"}),
        ("tax_unit_name", {"fa": "نام واحد مالیاتی", "en": "Tax Unit Name"}),
        ("attribute_ids", {"fa": "شناسه ویژگی‌ها", "en": "Attribute IDs"}),
        ("attribute_titles", {"fa": "نام ویژگی‌ها", "en": "Attribute Titles"}),
    ]

    headers = [labels.get(locale, labels.get("en", key)) for key, labels in columns]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Sample data row (row 2)
    if locale == 'fa':
        sample = [
            "P1001", "نمونه کالا", "کالا", "توضیح اختیاری", "",
            "مواد اولیه > پلاستیک", "", "", "",
            "150000", "120000", "TRUE",
            "0", "0", "",
            "FALSE", "FALSE", "", "",
            "", "", "", "", "", "", "",
            "1,2,3", "رنگ, سایز",
        ]
    else:
        sample = [
            "P1001", "Sample product", "product", "Optional description", "",
            "Raw materials > Plastics", "", "", "",
            "150000", "120000", "TRUE",
            "0", "0", "",
            "FALSE", "FALSE", "", "",
            "", "", "", "", "", "", "",
            "1,2,3", "Color, Size",
        ]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    # Auto width
    for column in ws.columns:
        try:
            letter = column[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"products_import_template_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/business/{business_id}/import/excel",
    summary="ایمپورت محصولات از فایل Excel",
    description="""
    آپلود و پردازش فایل Excel برای ایمپورت محصولات
    
    ### قابلیت‌ها:
    - پشتیبانی از فایل Excel (.xlsx)
    - پردازش به صورت dry-run (پیش‌نمایش) یا واقعی
    - ایجاد یا به‌روزرسانی محصولات بر اساس کد
    - پشتیبانی از دسته‌بندی، مالیات، ویژگی‌ها و سایر فیلدها
    - گزارش خطاها و هشدارها
    
    ### نکات مهم:
    - برای دریافت تمپلیت از endpoint `/import/template` استفاده کنید
    - در حالت dry-run تغییری در داده‌ها ایجاد نمی‌شود
    - محصولات با کد یکسان به‌روزرسانی می‌شوند
    - خطاهای اعتبارسنجی در پاسخ برگردانده می‌شوند
    """,
    response_model=SuccessResponse[Dict[str, Any]],
    responses={
        200: {
            "description": "فایل با موفقیت پردازش شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "ایمپورت با موفقیت انجام شد",
                        "data": {
                            "total_rows": 100,
                            "processed": 95,
                            "created": 80,
                            "updated": 15,
                            "skipped": 5,
                            "errors": [
                                {
                                    "row": 10,
                                    "code": "P-010",
                                    "message": "نام محصول الزامی است"
                                }
                            ],
                            "warnings": []
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی فایل یا داده‌ها"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.edit"
        },
        413: {
            "description": "حجم فایل بیش از حد مجاز است"
        }
    }
)
@require_business_access("business_id")
async def import_products_excel(
    request: Request,
    business_id: int,
    file: UploadFile = File(...),
    dry_run: str = Form(default="true"),
    match_by: str = Form(default="code"),
    conflict_policy: str = Form(default="upsert"),
    on_missing_category: str = Form(default="error"),
    on_missing_attributes: str = Form(default="error"),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "edit")),
):
    import io
    import json
    import logging
    import re
    import zipfile
    from decimal import Decimal
    from typing import Optional
    from openpyxl import load_workbook
    from sqlalchemy import and_ as _and
    from adapters.db.models.category import BusinessCategory
    from adapters.db.models.product_attribute import ProductAttribute
    from adapters.db.models.tax_type import TaxType
    from adapters.db.models.tax_unit import TaxUnit

    logger = logging.getLogger(__name__)

    def _validate_excel_signature(content: bytes) -> bool:
        try:
            if not content.startswith(b'PK'):
                return False
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                return any(n.startswith('xl/') for n in zf.namelist())
        except Exception:
            return False

    try:
        is_dry_run = str(dry_run).lower() in ("true","1","yes","on")
        logger.info(f"[IMPORT] Starting Excel import - business_id={business_id}, dry_run={is_dry_run}, match_by={match_by}, conflict_policy={conflict_policy}")
        on_missing_category = str(on_missing_category or "error").strip().lower()
        on_missing_attributes = str(on_missing_attributes or "error").strip().lower()
        if on_missing_category not in ("error", "create"):
            on_missing_category = "error"
        if on_missing_attributes not in ("error", "create"):
            on_missing_attributes = "error"

        preview_rows: list[dict] = []
        reference_summary: dict[str, Any] = {
            "resolved": {
                "category": 0,
                "tax_type": 0,
                "tax_unit": 0,
                "attributes": 0,
            },
            "would_create": {
                "categories": 0,
                "attributes": 0,
            },
            "created": {
                "categories": 0,
                "attributes": 0,
            },
            "policies": {
                "on_missing_category": on_missing_category,
                "on_missing_attributes": on_missing_attributes,
            },
        }

        if not file.filename or not file.filename.lower().endswith('.xlsx'):
            raise ApiError("INVALID_FILE", "فرمت فایل معتبر نیست. تنها xlsx پشتیبانی می‌شود", http_status=400)

        content = await file.read()
        logger.info(f"[IMPORT] File received - filename={file.filename}, size={len(content)} bytes")
        if len(content) < 100 or not _validate_excel_signature(content):
            raise ApiError("INVALID_FILE", "فایل Excel معتبر نیست یا خالی است", http_status=400)

        try:
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        except zipfile.BadZipFile:
            raise ApiError("INVALID_FILE", "فایل Excel خراب است یا فرمت آن معتبر نیست", http_status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        logger.info(f"[IMPORT] Excel file loaded - total rows={len(rows)}")
        if not rows:
            return success_response(data={"summary": {"total": 0}}, request=request, message="EMPTY_FILE")

        # Headers may be localized (fa/en). Normalize them to internal keys.
        raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        def _normalize_header(v: object) -> str:
            s = "" if v is None else str(v)
            s = s.replace("\u200c", " ")  # ZWNJ -> space
            s = re.sub(r"\s+", " ", s).strip()
            return s.lower()

        # Aliases for headers (localized labels -> internal keys)
        header_aliases: dict[str, str] = {}
        internal_keys = [
            "code","name","item_type","description","category_id",
            "category_path","category",
            "main_unit","secondary_unit","unit_conversion_factor",
            "base_sales_price","base_purchase_price","track_inventory",
            "reorder_point","min_order_qty","lead_time_days",
            "is_sales_taxable","is_purchase_taxable","sales_tax_rate","purchase_tax_rate",
            "tax_type_id","tax_type_code","tax_type_title","tax_code",
            "tax_unit_id","tax_unit_code","tax_unit_name",
            "attribute_ids","attribute_titles",
        ]
        for k in internal_keys:
            header_aliases[_normalize_header(k)] = k

        # Persian labels
        header_aliases.update({
            _normalize_header("کد"): "code",
            _normalize_header("نام"): "name",
            _normalize_header("نوع"): "item_type",
            _normalize_header("توضیحات"): "description",
            _normalize_header("شناسه دسته‌بندی"): "category_id",
            _normalize_header("شناسه دسته بندی"): "category_id",
            _normalize_header("مسیر دسته‌بندی"): "category_path",
            _normalize_header("مسیر دسته بندی"): "category_path",
            _normalize_header("دسته‌بندی"): "category",
            _normalize_header("دسته بندی"): "category",
            _normalize_header("واحد اصلی"): "main_unit",
            _normalize_header("واحد فرعی"): "secondary_unit",
            _normalize_header("ضریب تبدیل"): "unit_conversion_factor",
            _normalize_header("قیمت فروش"): "base_sales_price",
            _normalize_header("قیمت خرید"): "base_purchase_price",
            _normalize_header("کنترل موجودی"): "track_inventory",
            _normalize_header("نقطه سفارش مجدد"): "reorder_point",
            _normalize_header("حداقل مقدار سفارش"): "min_order_qty",
            _normalize_header("زمان تامین (روز)"): "lead_time_days",
            _normalize_header("زمان تأمین (روز)"): "lead_time_days",
            _normalize_header("مشمول مالیات فروش"): "is_sales_taxable",
            _normalize_header("مشمول مالیات خرید"): "is_purchase_taxable",
            _normalize_header("نرخ مالیات فروش (%)"): "sales_tax_rate",
            _normalize_header("نرخ مالیات خرید (%)"): "purchase_tax_rate",
            _normalize_header("شناسه نوع مالیات"): "tax_type_id",
            _normalize_header("کد نوع مالیات"): "tax_type_code",
            _normalize_header("عنوان نوع مالیات"): "tax_type_title",
            _normalize_header("کد مالیاتی"): "tax_code",
            _normalize_header("شناسه واحد مالیاتی"): "tax_unit_id",
            _normalize_header("کد واحد مالیاتی"): "tax_unit_code",
            _normalize_header("نام واحد مالیاتی"): "tax_unit_name",
            _normalize_header("شناسه ویژگی‌ها"): "attribute_ids",
            _normalize_header("شناسه ویژگی ها"): "attribute_ids",
            _normalize_header("نام ویژگی‌ها"): "attribute_titles",
            _normalize_header("نام ویژگی ها"): "attribute_titles",
        })

        # English labels
        header_aliases.update({
            _normalize_header("code"): "code",
            _normalize_header("name"): "name",
            _normalize_header("type"): "item_type",
            _normalize_header("description"): "description",
            _normalize_header("category id"): "category_id",
            _normalize_header("category path"): "category_path",
            _normalize_header("category"): "category",
            _normalize_header("main unit"): "main_unit",
            _normalize_header("secondary unit"): "secondary_unit",
            _normalize_header("unit conversion factor"): "unit_conversion_factor",
            _normalize_header("sales price"): "base_sales_price",
            _normalize_header("purchase price"): "base_purchase_price",
            _normalize_header("track inventory"): "track_inventory",
            _normalize_header("reorder point"): "reorder_point",
            _normalize_header("min order qty"): "min_order_qty",
            _normalize_header("lead time (days)"): "lead_time_days",
            _normalize_header("sales taxable"): "is_sales_taxable",
            _normalize_header("purchase taxable"): "is_purchase_taxable",
            _normalize_header("sales tax rate (%)"): "sales_tax_rate",
            _normalize_header("purchase tax rate (%)"): "purchase_tax_rate",
            _normalize_header("tax type id"): "tax_type_id",
            _normalize_header("tax type code"): "tax_type_code",
            _normalize_header("tax type title"): "tax_type_title",
            _normalize_header("tax code"): "tax_code",
            _normalize_header("tax unit id"): "tax_unit_id",
            _normalize_header("tax unit code"): "tax_unit_code",
            _normalize_header("tax unit name"): "tax_unit_name",
            _normalize_header("attribute ids"): "attribute_ids",
            _normalize_header("attribute titles"): "attribute_titles",
        })

        headers = [header_aliases.get(_normalize_header(h), h) for h in raw_headers]
        data_rows = rows[1:]
        logger.info(f"[IMPORT] Headers parsed: {headers}, data rows count: {len(data_rows)}")

        def _parse_bool(v: object) -> Optional[bool]:
            if v is None: return None
            s = str(v).strip().lower()
            if s in ("true","1","yes","on","بله","هست"):
                return True
            if s in ("false","0","no","off","خیر","نیست"):
                return False
            return None

        def _normalize_number_text(v: object) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if s == "":
                return ""
            # Handle negative numbers in parentheses: (123) => -123
            if s.startswith("(") and s.endswith(")"):
                s = "-" + s[1:-1]
            # Convert Persian/Arabic digits to English
            digit_map = str.maketrans({
                "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
                "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
                "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
                "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9",
            })
            s = s.translate(digit_map)
            # Normalize separators:
            # - Thousands separators: "," "٬" "،" spaces
            # - Decimal separator: "٫" -> "."
            s = s.replace("\u066b", ".")  # Arabic decimal separator
            for ch in [",", "\u066c", "\u060c", " ", "\u00a0", "\u202f", "\u2009", "_"]:
                s = s.replace(ch, "")
            s = s.strip()
            return s

        def _parse_decimal(v: object) -> Optional[Decimal]:
            s = _normalize_number_text(v)
            if s == "":
                return None
            try:
                return Decimal(s)
            except Exception:
                return None

        def _parse_int(v: object) -> Optional[int]:
            s = _normalize_number_text(v)
            if s == "":
                return None
            try:
                return int(s.split(".")[0])
            except Exception:
                return None

        def _normalize_item_type(v: object) -> Optional[str]:
            if v is None: return None
            s = str(v).strip()
            mapping = {"product": "کالا", "service": "خدمت"}
            low = s.lower()
            if low in mapping: return mapping[low]
            if s in ("کالا","خدمت"): return s
            return None

        def _norm_text(v: object) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            s = s.replace("\u200c", " ")
            s = re.sub(r"\s+", " ", s).strip()
            return s

        def _norm_key(v: object) -> str:
            return _norm_text(v).lower()

        def _split_list_text(v: object) -> list[str]:
            s = _norm_text(v)
            if not s:
                return []
            # Support comma / Persian comma / semicolon / pipe
            parts = re.split(r"[,\u060c;\|]+", s)
            out: list[str] = []
            for p in parts:
                t = _norm_text(p)
                if t:
                    out.append(t)
            return out

        # Preload reference data for faster resolve (only when needed).
        categories_rows: list[BusinessCategory] | None = None
        attrs_rows: list[ProductAttribute] | None = None
        tax_type_by_code: dict[str, int] | None = None
        tax_type_by_title: dict[str, int] | None = None
        tax_unit_by_code: dict[str, int] | None = None
        tax_unit_by_name: dict[str, int] | None = None

        def _ensure_categories_loaded() -> list[BusinessCategory]:
            nonlocal categories_rows
            if categories_rows is None:
                categories_rows = db.query(BusinessCategory).filter(BusinessCategory.business_id == business_id).all()
            return categories_rows

        def _ensure_attributes_loaded() -> list[ProductAttribute]:
            nonlocal attrs_rows
            if attrs_rows is None:
                attrs_rows = db.query(ProductAttribute).filter(ProductAttribute.business_id == business_id).all()
            return attrs_rows

        def _ensure_tax_types_loaded() -> None:
            nonlocal tax_type_by_code, tax_type_by_title
            if tax_type_by_code is not None and tax_type_by_title is not None:
                return
            rows = db.query(TaxType).all()
            tax_type_by_code = {(r.code or "").strip().lower(): r.id for r in rows if r.code}
            tax_type_by_title = {(r.title or "").strip().lower(): r.id for r in rows if r.title}

        def _ensure_tax_units_loaded() -> None:
            nonlocal tax_unit_by_code, tax_unit_by_name
            if tax_unit_by_code is not None and tax_unit_by_name is not None:
                return
            rows = db.query(TaxUnit).all()
            tax_unit_by_code = {(r.code or "").strip().lower(): r.id for r in rows if r.code}
            tax_unit_by_name = {(r.name or "").strip().lower(): r.id for r in rows if r.name}

        def _get_category_titles(cat: BusinessCategory) -> list[str]:
            trans = cat.title_translations or {}
            vals = []
            for k in ("fa", "en"):
                t = (trans.get(k) or "").strip()
                if t:
                    vals.append(t)
            # include any other translations
            for t in trans.values():
                tt = (t or "").strip()
                if tt and tt not in vals:
                    vals.append(tt)
            return vals

        def _resolve_category_by_id(category_id: Optional[int]) -> tuple[Optional[int], Optional[str]]:
            if category_id is None:
                return None, None
            exists = db.query(BusinessCategory.id).filter(_and(BusinessCategory.business_id == business_id, BusinessCategory.id == category_id)).first()
            if not exists:
                return None, f"دسته‌بندی با شناسه {category_id} یافت نشد"
            return category_id, None

        def _resolve_category_by_name_or_path(category_value: object, category_path: object) -> tuple[Optional[int], Optional[str], list[str]]:
            """
            Returns: (category_id, error, created_paths)
            created_paths is for reporting (best-effort).
            """
            created_paths: list[str] = []
            would_create_paths: list[str] = []
            # If path provided, prefer it.
            path_str = _norm_text(category_path)
            name_str = _norm_text(category_value)
            if not path_str and not name_str:
                return None, None, created_paths

            # Parse path segments
            path = path_str or name_str
            # split on common separators
            segments = [s.strip() for s in re.split(r"[>/\u203a\u00bb]+", path) if s and str(s).strip()]
            segments = [_norm_text(s) for s in segments if _norm_text(s)]
            if not segments:
                return None, None, created_paths

            cats = _ensure_categories_loaded()
            # Build index by parent_id and normalized title
            by_parent: dict[int | None, list[BusinessCategory]] = {}
            for c in cats:
                by_parent.setdefault(c.parent_id, []).append(c)

            parent_id: int | None = None
            current_id: int | None = None
            for seg in segments:
                seg_norm = seg.strip().lower()
                candidates: list[BusinessCategory] = []
                for c in by_parent.get(parent_id, []):
                    titles = _get_category_titles(c)
                    if any(t.strip().lower() == seg_norm for t in titles):
                        candidates.append(c)
                if len(candidates) == 1:
                    current_id = candidates[0].id
                    parent_id = current_id
                    continue
                if len(candidates) > 1:
                    opts = [f"{c.id}:{(_get_category_titles(c)[0] if _get_category_titles(c) else '')}" for c in candidates[:5]]
                    return None, f"دسته‌بندی مبهم است: '{seg}' (گزینه‌ها: {', '.join(opts)})", created_paths

                # no match
                if on_missing_category == "create" and is_dry_run:
                    # In dry-run, don't fail the row; just report what would be created.
                    # We can't know the IDs yet, so keep category_id as None.
                    would_create_paths.append(seg if not would_create_paths else f"{would_create_paths[-1]} > {seg}")
                    # Also include remaining segments (all would be created under the last known parent).
                    # Continue collecting for reporting.
                    continue
                if on_missing_category == "create" and (not is_dry_run):
                    # Create the category under current parent
                    from adapters.db.repositories.category_repository import CategoryRepository
                    repo = CategoryRepository(db)
                    obj = repo.create_category(business_id=business_id, parent_id=parent_id, translations={"fa": seg, "en": seg})
                    reference_summary["created"]["categories"] += 1
                    created_paths.append(seg if not created_paths else f"{created_paths[-1]} > {seg}")
                    # update local caches
                    cats.append(obj)
                    by_parent.setdefault(parent_id, []).append(obj)
                    current_id = obj.id
                    parent_id = current_id
                    continue
                return None, f"دسته‌بندی یافت نشد: '{seg}'", created_paths

            # If we collected would_create paths during dry-run, return None (unknown id) but no error.
            if would_create_paths:
                # Encode as a special marker in created_paths for reporting (caller will read helper key).
                # We don't want to change return signature widely; keep created_paths for real creates.
                return None, None, created_paths + [f"__WOULD_CREATE__:{p}" for p in would_create_paths]

            return current_id, None, created_paths

        def _resolve_tax_type(item: dict, row_errors: list[str]) -> None:
            # If id provided, accept as-is (no FK), but validate if possible
            if item.get("tax_type_id") is not None:
                # basic int already parsed
                _ensure_tax_types_loaded()
                tid = item.get("tax_type_id")
                if isinstance(tid, int) and tax_type_by_code is not None and tax_type_by_title is not None:
                    if tid not in set(tax_type_by_code.values()) and tid not in set(tax_type_by_title.values()):
                        row_errors.append(f"شناسه نوع مالیات نامعتبر است: {tid}")
                return
            code = _norm_text(item.get("tax_type_code"))
            if code:
                _ensure_tax_types_loaded()
                tid = (tax_type_by_code or {}).get(code.lower())
                if not tid:
                    row_errors.append(f"نوع مالیات با کد '{code}' یافت نشد")
                else:
                    item["tax_type_id"] = tid
                    return
            title = _norm_text(item.get("tax_type_title"))
            if title:
                _ensure_tax_types_loaded()
                tid = (tax_type_by_title or {}).get(title.lower())
                if not tid:
                    row_errors.append(f"نوع مالیات با عنوان '{title}' یافت نشد")
                else:
                    item["tax_type_id"] = tid

        def _resolve_tax_unit(item: dict, row_errors: list[str]) -> None:
            if item.get("tax_unit_id") is not None:
                _ensure_tax_units_loaded()
                uid = item.get("tax_unit_id")
                if isinstance(uid, int) and tax_unit_by_code is not None and tax_unit_by_name is not None:
                    if uid not in set(tax_unit_by_code.values()) and uid not in set(tax_unit_by_name.values()):
                        row_errors.append(f"شناسه واحد مالیاتی نامعتبر است: {uid}")
                return
            code = _norm_text(item.get("tax_unit_code"))
            if code:
                _ensure_tax_units_loaded()
                uid = (tax_unit_by_code or {}).get(code.lower())
                if not uid:
                    row_errors.append(f"واحد مالیاتی با کد '{code}' یافت نشد")
                else:
                    item["tax_unit_id"] = uid
                    return
            name = _norm_text(item.get("tax_unit_name"))
            if name:
                _ensure_tax_units_loaded()
                uid = (tax_unit_by_name or {}).get(name.lower())
                if not uid:
                    row_errors.append(f"واحد مالیاتی با نام '{name}' یافت نشد")
                else:
                    item["tax_unit_id"] = uid

        def _resolve_attributes(item: dict, row_errors: list[str]) -> None:
            # If attribute_ids already parsed (list[int]), validate and keep only valid; but report missing.
            if isinstance(item.get("attribute_ids"), list) and item.get("attribute_ids"):
                ids = [i for i in item["attribute_ids"] if isinstance(i, int)]
                if not ids:
                    item["attribute_ids"] = []
                    return
                # validate against business attributes
                existing_ids = set([a.id for a in _ensure_attributes_loaded()])
                missing = [str(i) for i in ids if i not in existing_ids]
                item["attribute_ids"] = [i for i in ids if i in existing_ids]
                if missing:
                    row_errors.append(f"شناسه(های) ویژگی نامعتبر: {', '.join(missing)}")
            # If attribute_titles provided, resolve and optionally create
            titles = _split_list_text(item.get("attribute_titles"))
            if not titles:
                return
            attrs = _ensure_attributes_loaded()
            by_title = {a.title.strip().lower(): a for a in attrs if a.title}
            resolved_ids: list[int] = []
            missing_titles: list[str] = []
            created: list[str] = []
            for t in titles:
                key = t.strip().lower()
                found = by_title.get(key)
                if found:
                    resolved_ids.append(found.id)
                    continue
                if on_missing_attributes == "create" and (not is_dry_run):
                    from adapters.db.repositories.product_attribute_repository import ProductAttributeRepository
                    repo = ProductAttributeRepository(db)
                    try:
                        obj = repo.create(business_id=business_id, title=t, description=None, data_type="text", options=None)
                        reference_summary["created"]["attributes"] += 1
                        attrs.append(obj)
                        by_title[obj.title.strip().lower()] = obj
                        resolved_ids.append(obj.id)
                        created.append(t)
                    except Exception:
                        missing_titles.append(t)
                else:
                    missing_titles.append(t)
            if created:
                # store for later summary if needed
                item["_created_attribute_titles"] = created
            if missing_titles:
                if on_missing_attributes == "create" and is_dry_run:
                    # In dry-run, just report what would be created; don't fail the row.
                    item["_would_create_attribute_titles"] = missing_titles
                else:
                    row_errors.append(f"ویژگی(های) یافت نشد: {', '.join(missing_titles)}")
            # merge with any existing attribute_ids already present
            current = item.get("attribute_ids") if isinstance(item.get("attribute_ids"), list) else []
            merged = list(dict.fromkeys([*(current or []), *resolved_ids]))
            item["attribute_ids"] = merged

        def _extract_would_create_paths(paths: list[str] | None) -> list[str]:
            if not paths:
                return []
            out: list[str] = []
            for p in paths:
                if isinstance(p, str) and p.startswith("__WOULD_CREATE__:"):
                    out.append(p.split(":", 1)[1])
            return out

        errors: list[dict] = []
        valid_items: list[dict] = []

        for idx, row in enumerate(data_rows, start=2):
            item: dict[str, Any] = {}
            row_errors: list[str] = []
            row_warnings: list[str] = []
            row_preview: dict[str, Any] = {"row": idx, "resolved": {}, "would_create": {}, "warnings": []}

            for ci, key in enumerate(headers):
                if not key:
                    continue
                val = row[ci] if ci < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                item[key] = val

            # normalize & cast
            if 'item_type' in item:
                item['item_type'] = _normalize_item_type(item.get('item_type')) or 'کالا'
            for k in ['base_sales_price','base_purchase_price','sales_tax_rate','purchase_tax_rate','unit_conversion_factor']:
                if k in item:
                    item[k] = _parse_decimal(item.get(k))
            for k in ['reorder_point','min_order_qty','lead_time_days','category_id','tax_type_id','tax_unit_id']:
                if k in item:
                    item[k] = _parse_int(item.get(k))
            # Handle boolean fields - always set them, default to False if not provided or invalid
            for k in ['track_inventory','is_sales_taxable','is_purchase_taxable']:
                if k in item:
                    parsed = _parse_bool(item.get(k))
                    # For boolean fields, if None or invalid, use False as default
                    item[k] = parsed if parsed is not None else False
                else:
                    # If field doesn't exist in item, set default to False
                    item[k] = False

            # attribute_ids: comma-separated
            if 'attribute_ids' in item and item['attribute_ids']:
                try:
                    parts = [p.strip() for p in str(item['attribute_ids']).split(',') if p and p.strip()]
                    item['attribute_ids'] = [int(p) for p in parts if p.isdigit()]
                except Exception:
                    item['attribute_ids'] = []

            # Resolve references: category, tax, attributes
            # Category: accept category_id, or resolve from category_path/category (name), optionally create
            cat_id, cat_err = _resolve_category_by_id(item.get("category_id"))
            if cat_err:
                # allow resolve by path/name as fallback
                resolved_id, resolved_err, _created_paths = _resolve_category_by_name_or_path(item.get("category"), item.get("category_path"))
                if resolved_err:
                    row_errors.append(cat_err + " / " + resolved_err)
                else:
                    item["category_id"] = resolved_id
                    would_create = _extract_would_create_paths(_created_paths)
                    if would_create:
                        row_preview["would_create"]["categories"] = would_create
                        reference_summary["would_create"]["categories"] += len(would_create)
                        row_warnings.append("دسته‌بندی پیدا نشد و در حالت create ساخته خواهد شد")
                    elif resolved_id is not None:
                        row_preview["resolved"]["category"] = {"by": "path_or_name", "category_id": resolved_id}
                        reference_summary["resolved"]["category"] += 1
            else:
                if cat_id is not None:
                    row_preview["resolved"]["category"] = {"by": "id", "category_id": cat_id}
                    reference_summary["resolved"]["category"] += 1
                # if category_id not provided but path/name provided, resolve it
                if item.get("category_id") is None and (item.get("category") or item.get("category_path")):
                    resolved_id, resolved_err, _created_paths = _resolve_category_by_name_or_path(item.get("category"), item.get("category_path"))
                    if resolved_err:
                        row_errors.append(resolved_err)
                    else:
                        item["category_id"] = resolved_id
                        # detect dry-run would create marker
                        would_create = _extract_would_create_paths(_created_paths)
                        if would_create:
                            row_preview["would_create"]["categories"] = would_create
                            reference_summary["would_create"]["categories"] += len(would_create)
                            row_warnings.append("دسته‌بندی پیدا نشد و در حالت create ساخته خواهد شد")
                        elif resolved_id is not None:
                            row_preview["resolved"]["category"] = {"by": "path_or_name", "category_id": resolved_id}
                            reference_summary["resolved"]["category"] += 1

            before_tax_type = item.get("tax_type_id")
            _resolve_tax_type(item, row_errors)
            if before_tax_type is None and item.get("tax_type_id") is not None:
                row_preview["resolved"]["tax_type"] = {"tax_type_id": item.get("tax_type_id")}
                reference_summary["resolved"]["tax_type"] += 1
            before_tax_unit = item.get("tax_unit_id")
            _resolve_tax_unit(item, row_errors)
            if before_tax_unit is None and item.get("tax_unit_id") is not None:
                row_preview["resolved"]["tax_unit"] = {"tax_unit_id": item.get("tax_unit_id")}
                reference_summary["resolved"]["tax_unit"] += 1

            before_attrs = list(item.get("attribute_ids") or []) if isinstance(item.get("attribute_ids"), list) else []
            _resolve_attributes(item, row_errors)
            after_attrs = list(item.get("attribute_ids") or []) if isinstance(item.get("attribute_ids"), list) else []
            if len(after_attrs) > len(before_attrs):
                row_preview["resolved"]["attributes"] = {"attribute_ids": after_attrs}
                reference_summary["resolved"]["attributes"] += (len(after_attrs) - len(before_attrs))
            if "_would_create_attribute_titles" in item:
                would_titles = item.get("_would_create_attribute_titles") or []
                if isinstance(would_titles, list) and would_titles:
                    row_preview["would_create"]["attributes"] = would_titles
                    reference_summary["would_create"]["attributes"] += len(would_titles)
                    row_warnings.append("برخی ویژگی‌ها وجود ندارند و در حالت create ساخته خواهند شد")

            # validations
            name = item.get('name')
            if not name or str(name).strip() == "":
                row_errors.append('name الزامی است')

            # if code is empty, it will be auto-generated in service
            code = item.get('code')
            if code is not None:
                code_str = str(code).strip()
                # Handle string "None" or empty string
                if code_str == "" or code_str.lower() == "none":
                    item['code'] = None
                else:
                    item['code'] = code_str

            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
                logger.debug(f"[IMPORT] Row {idx} validation failed: {row_errors}")
                if is_dry_run:
                    row_preview["warnings"] = row_warnings
                    preview_rows.append(row_preview)
                continue

            # Remove helper keys not part of schema
            if "_created_attribute_titles" in item:
                item.pop("_created_attribute_titles", None)
            if "_would_create_attribute_titles" in item:
                item.pop("_would_create_attribute_titles", None)

            valid_items.append(item)
            logger.debug(f"[IMPORT] Row {idx} validated successfully - name={item.get('name')}, code={item.get('code')}")
            if is_dry_run:
                row_preview["warnings"] = row_warnings
                preview_rows.append(row_preview)

        inserted = 0
        updated = 0
        skipped = 0

        logger.info(f"[IMPORT] Processing summary - total_rows={len(data_rows)}, valid_items={len(valid_items)}, errors={len(errors)}, is_dry_run={is_dry_run}")

        if not is_dry_run and valid_items:
            logger.info(f"[IMPORT] Starting REAL import (not dry-run) for {len(valid_items)} items")
            from sqlalchemy import and_ as _and
            from adapters.db.models.product import Product
            from adapters.api.v1.schema_models.product import ProductCreateRequest, ProductUpdateRequest
            from app.services.product_service import create_product, update_product

            def _find_existing(session: Session, data: dict) -> Optional[Product]:
                if match_by == 'code' and data.get('code'):
                    result = session.query(Product).filter(_and(Product.business_id == business_id, Product.code == str(data['code']).strip())).first()
                    logger.debug(f"[IMPORT] Searching by code='{data.get('code')}' - found: {result is not None}")
                    return result
                if match_by == 'name' and data.get('name'):
                    result = session.query(Product).filter(_and(Product.business_id == business_id, Product.name == str(data['name']).strip())).first()
                    logger.debug(f"[IMPORT] Searching by name='{data.get('name')}' - found: {result is not None}")
                    return result
                logger.debug(f"[IMPORT] No match criteria - match_by={match_by}, code={data.get('code')}, name={data.get('name')}")
                return None

            for idx, data in enumerate(valid_items, start=1):
                item_name = data.get('name', 'N/A')
                item_code = data.get('code', 'N/A')
                logger.info(f"[IMPORT] Processing item {idx}/{len(valid_items)}: name='{item_name}', code='{item_code}'")
                logger.debug(f"[IMPORT] Full item data: {data}")
                
                existing = _find_existing(db, data)
                if existing is None:
                    logger.info(f"[IMPORT] Item '{item_name}' not found - will CREATE new product")
                    try:
                        logger.debug(f"[IMPORT] Calling create_product with business_id={business_id}, data keys: {list(data.keys())}")
                        # Log data before creating ProductCreateRequest to see what's being passed
                        logger.debug(f"[IMPORT] Data to create ProductCreateRequest: {json.dumps({k: str(v) for k, v in data.items()}, ensure_ascii=False, default=str)}")
                        try:
                            product_request = ProductCreateRequest(**data)
                            logger.debug(f"[IMPORT] ProductCreateRequest created successfully")
                        except Exception as validation_error:
                            logger.error(f"[IMPORT] ❌ ValidationError creating ProductCreateRequest for '{item_name}': {validation_error}")
                            # Try to get detailed validation errors
                            try:
                                if hasattr(validation_error, 'errors'):
                                    errors_list = validation_error.errors()
                                    logger.error(f"[IMPORT] Validation errors details: {json.dumps(errors_list, ensure_ascii=False, indent=2)}")
                                elif hasattr(validation_error, 'error_dict'):
                                    logger.error(f"[IMPORT] Validation error_dict: {json.dumps(validation_error.error_dict(), ensure_ascii=False, indent=2)}")
                                # Log the string representation as fallback
                                logger.error(f"[IMPORT] Full validation error: {str(validation_error)}")
                            except Exception as log_error:
                                logger.error(f"[IMPORT] Could not serialize validation error: {log_error}")
                            raise
                        result = create_product(db, business_id, product_request)
                        logger.info(f"[IMPORT] ✅ Successfully CREATED product '{item_name}' - result: {result.get('message', 'N/A')}")
                        if result.get('data', {}).get('id'):
                            logger.info(f"[IMPORT] Created product ID: {result['data']['id']}")
                        inserted += 1
                        logger.info(f"[IMPORT] Insert counter: {inserted}")
                    except Exception as e:
                        logger.error(f"[IMPORT] ❌ Create product failed for '{item_name}': {e}", exc_info=True)
                        logger.error(f"[IMPORT] Exception type: {type(e).__name__}, args: {e.args}")
                        if hasattr(e, 'errors'):
                            logger.error(f"[IMPORT] Validation errors: {json.dumps(e.errors(), ensure_ascii=False, indent=2)}")
                        skipped += 1
                else:
                    logger.info(f"[IMPORT] Item '{item_name}' EXISTS (id={existing.id}) - conflict_policy={conflict_policy}")
                    if conflict_policy == 'insert':
                        logger.info(f"[IMPORT] Skipping existing item due to conflict_policy='insert'")
                        skipped += 1
                    elif conflict_policy in ('update','upsert'):
                        logger.info(f"[IMPORT] Will UPDATE existing product id={existing.id}")
                        try:
                            logger.debug(f"[IMPORT] Calling update_product with id={existing.id}, business_id={business_id}")
                            result = update_product(db, existing.id, business_id, ProductUpdateRequest(**data))
                            logger.info(f"[IMPORT] ✅ Successfully UPDATED product '{item_name}' (id={existing.id})")
                            updated += 1
                            logger.info(f"[IMPORT] Update counter: {updated}")
                        except Exception as e:
                            logger.error(f"[IMPORT] ❌ Update product failed for '{item_name}' (id={existing.id}): {e}", exc_info=True)
                            logger.error(f"[IMPORT] Exception type: {type(e).__name__}, args: {e.args}")
                            skipped += 1
        else:
            if is_dry_run:
                logger.info(f"[IMPORT] DRY-RUN mode - skipping actual database operations")
            else:
                logger.warning(f"[IMPORT] No valid items to process (valid_items is empty)")

        summary = {
            "total": len(data_rows),
            "valid": len(valid_items),
            "invalid": len(errors),
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "dry_run": is_dry_run,
        }

        logger.info(f"[IMPORT] Final summary: {summary}")
        logger.info(f"[IMPORT] Import completed - inserted={inserted}, updated={updated}, skipped={skipped}")

        return success_response(
            data={
                "summary": summary,
                "errors": errors,
                "reference_summary": reference_summary,
                "preview": preview_rows if is_dry_run else None,
            },
            request=request,
            message="PRODUCTS_IMPORT_RESULT",
        )
    except ApiError:
        raise
    except Exception as e:
        logger.error(f"Import error: {e}", exc_info=True)
        raise ApiError("IMPORT_ERROR", f"خطا در پردازش فایل: {e}", http_status=500)
@router.post(
    "/business/{business_id}/export/pdf",
    summary="خروجی PDF لیست محصولات",
    description="""
    دریافت لیست محصولات به صورت فایل PDF
    
    ### قابلیت‌ها:
    - فیلتر محصولات بر اساس معیارهای مختلف
    - انتخاب ستون‌های دلخواه برای خروجی
    - مرتب‌سازی داده‌ها
    - پشتیبانی از تقویم شمسی و میلادی (بر اساس header X-Calendar-Type)
    - پشتیبانی از RTL برای زبان فارسی
    
    ### فرمت فایل:
    - فرمت: `.pdf`
    - نام فایل شامل نام کسب‌وکار و timestamp است
    """,
    responses={
        200: {
            "description": "فایل PDF با موفقیت ایجاد شد",
            "content": {
                "application/pdf": {
                    "schema": {
                        "type": "string",
                        "format": "binary"
                    }
                }
            },
            "headers": {
                "Content-Disposition": {
                    "description": "نام فایل خروجی",
                    "schema": {
                        "type": "string",
                        "example": "attachment; filename=products_business_20240115_103000.pdf"
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.export"
        }
    }
)
@require_business_access("business_id")
async def export_products_pdf(
    request: Request,
    business_id: int,
    body: dict,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "export")),
):
    import json
    import datetime
    import re
    from fastapi.responses import Response
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration

    query_dict = {
        "take": int(body.get("take", 100)),
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "search_fields": body.get("search_fields") or body.get("searchFields"),
        "filters": body.get("filters"),
        "category_ids": body.get("category_ids") or body.get("categoryIds"),
    }
    result = list_products(db, business_id, query_dict)
    items = result.get("items", [])
    items = [format_datetime_fields(item, request) for item in items]

    # Apply selected indices filter if requested
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                indices = json.loads(selected_indices)
            except (json.JSONDecodeError, TypeError):
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]

    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        headers = [col.get("label") or col.get("key") for col in export_columns]
        keys = [col.get("key") for col in export_columns]
    else:
        default_cols = [
            ("code", "کد"),
            ("name", "نام"),
            ("item_type", "نوع"),
            ("category_id", "دسته"),
            ("base_sales_price", "قیمت فروش"),
            ("base_purchase_price", "قیمت خرید"),
            ("main_unit", "واحد اصلی"),
            ("secondary_unit", "واحد فرعی"),
            ("track_inventory", "کنترل موجودی"),
            ("created_at_formatted", "ایجاد"),
        ]
        keys = [k for k, _ in default_cols]
        headers = [v for _, v in default_cols]

    # Locale and direction
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == 'fa')
    html_lang = 'fa' if is_fa else 'en'
    html_dir = 'rtl' if is_fa else 'ltr'

    # Load business info for header
    business_name = ""
    try:
        biz = db.query(Business).filter(Business.id == business_id).first()
        if biz is not None:
            business_name = biz.name or ""
    except Exception:
        business_name = ""

    # Escape helper
    def escape(s: Any) -> str:
        try:
            return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        except Exception:
            return str(s)

    # Build rows
    rows_html = []
    for item in items:
        tds = []
        for key in keys:
            value = item.get(key)
            if value is None:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            tds.append(f"<td>{escape(value)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    headers_html = ''.join(f"<th>{escape(h)}</th>" for h in headers)

    # Format report datetime based on X-Calendar-Type header
    calendar_header = request.headers.get("X-Calendar-Type", "jalali").lower()
    try:
        from app.core.calendar import CalendarConverter
        formatted_now = CalendarConverter.format_datetime(datetime.datetime.now(),
            "jalali" if calendar_header in ["jalali", "persian", "shamsi"] else "gregorian")
        now_str = formatted_now.get('formatted', formatted_now.get('date_time', ''))
    except Exception:
        now_str = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')

    title_text = "گزارش فهرست محصولات" if is_fa else "Products List Report"
    label_biz = "نام کسب‌وکار" if is_fa else "Business Name"
    label_date = "تاریخ گزارش" if is_fa else "Report Date"
    footer_text = "تولید شده توسط Hesabix" if is_fa else "Generated by Hesabix"
    page_label_left = "صفحه " if is_fa else "Page "
    page_label_of = " از " if is_fa else " of "

    # تلاش برای رندر با قالب سفارشی (products/list)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        template_context = {
            "title_text": title_text,
            "business_name": business_name,
            "generated_at": now_str,
            "is_fa": is_fa,
            "headers": headers,
            "keys": keys,
            "items": items,
            "table_headers_html": headers_html,
            "table_rows_html": "".join(rows_html),
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="products",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # Inject Persian fonts (YekanBakhFaNum/Vazirmatn) for PDF rendering
    fa_font_url_regular = ""
    fa_font_url_bold = ""
    try:
        if is_fa:
            from app.services.pdf.template_renderer import load_farsi_font_data_uris
            fa_reg, fa_bold = load_farsi_font_data_uris()
            fa_font_url_regular = fa_reg or ""
            fa_font_url_bold = fa_bold or ""
    except Exception:
        fa_font_url_regular = ""
        fa_font_url_bold = ""
    
    font_face_css = ""
    if is_fa and fa_font_url_regular:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_regular}') format('truetype'); font-weight: 400; font-style: normal; }}
        """
    if is_fa and fa_font_url_bold:
        font_face_css += f"""
          @font-face {{ font-family: 'YekanBakhFaNum'; src: url('{fa_font_url_bold}') format('truetype'); font-weight: 700; font-style: normal; }}
        """
    
    body_font_family = "YekanBakhFaNum, Vazirmatn, Tahoma, Arial, sans-serif" if is_fa else "Arial, sans-serif"
    
    table_html = f"""
    <html lang=\"{html_lang}\" dir=\"{html_dir}\">
      <head>
        <meta charset='utf-8'>
        <style>
          {font_face_css}
          @page {{
            size: A4 landscape;
            margin: 12mm;
            @bottom-{ 'left' if is_fa else 'right' } {{
              content: "{page_label_left}" counter(page) "{page_label_of}" counter(pages);
              font-size: 10px;
              color: #666;
              font-family: {body_font_family};
            }}
          }}
          body {{
            font-family: {body_font_family};
            font-size: 11px;
            color: #222;
          }}
          .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #444;
            padding-bottom: 6px;
          }}
          .title {{
            font-size: 16px;
            font-weight: 700;
          }}
          .meta {{
            font-size: 11px;
            color: #555;
          }}
          .table-wrapper {{
            width: 100%;
          }}
          table.report-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
          }}
          thead th {{
            background: #f0f3f7;
            border: 1px solid #c7cdd6;
            padding: 6px 4px;
            text-align: center;
            font-weight: 700;
            white-space: nowrap;
          }}
          tbody td {{
            border: 1px solid #d7dde6;
            padding: 5px 4px;
            vertical-align: top;
            overflow-wrap: anywhere;
            word-break: break-word;
            white-space: normal;
          }}
          .footer {{
            position: running(footer);
            font-size: 10px;
            color: #666;
            margin-top: 8px;
            text-align: {'left' if is_fa else 'right'};
          }}
        </style>
      </head>
      <body>
        <div class=\"header\">
          <div>
            <div class=\"title\">{title_text}</div>
            <div class=\"meta\">{label_biz}: {escape(business_name)}</div>
          </div>
          <div class=\"meta\">{label_date}: {escape(now_str)}</div>
        </div>
        <div class=\"table-wrapper\">
          <table class=\"report-table\">
            <thead>
              <tr>{headers_html}</tr>
            </thead>
            <tbody>
              {''.join(rows_html)}
            </tbody>
          </table>
        </div>
        <div class=\"footer\">{footer_text}</div>
      </body>
    </html>
    """

    final_html = resolved_html or table_html
    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(font_config=font_config)

    # Build meaningful filename
    biz_name = business_name
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    base = "products"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/business/{business_id}/bulk-price-update/preview",
    summary="پیش‌نمایش تغییر قیمت‌های گروهی",
    description="""
    پیش‌نمایش تغییرات قیمت قبل از اعمال تغییرات واقعی
    
    ### قابلیت‌ها:
    - تغییر قیمت بر اساس درصد یا مبلغ ثابت
    - تغییر قیمت فروش، خرید یا هر دو
    - فیلتر بر اساس دسته‌بندی، ارز، لیست قیمت و نوع محصول
    - نمایش تغییرات پیش از اعمال
    
    ### نکات:
    - این endpoint فقط پیش‌نمایش است و تغییری در قیمت‌ها ایجاد نمی‌کند
    - برای اعمال تغییرات از endpoint `/bulk-price-update/apply` استفاده کنید
    - نتایج شامل خلاصه آماری تغییرات است
    """,
    response_model=SuccessResponse[BulkPriceUpdatePreviewResponse],
    responses={
        200: {
            "description": "پیش‌نمایش تغییرات با موفقیت محاسبه شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "total_products": 100,
                            "affected_products": [
                                {
                                    "product_id": 1,
                                    "product_name": "محصول A",
                                    "product_code": "P-001",
                                    "category_name": "دسته ۱",
                                    "current_sales_price": 100000,
                                    "current_purchase_price": 80000,
                                    "new_sales_price": 110000,
                                    "new_purchase_price": 88000,
                                    "sales_price_change": 10000,
                                    "purchase_price_change": 8000
                                }
                            ],
                            "summary": {
                                "total_products": 100,
                                "affected_products": 95,
                                "products_with_sales_change": 95,
                                "products_with_purchase_change": 95,
                                "total_sales_change": 950000,
                                "total_purchase_change": 760000
                            }
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.edit"
        }
    }
)
@require_business_access("business_id")
def preview_bulk_price_update_endpoint(
    request: Request,
    business_id: int,
    payload: BulkPriceUpdateRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "edit")),
) -> Dict[str, Any]:
    
    result = preview_bulk_price_update(db, business_id, payload)
    # Pydantic v2: prefer model_dump(); keep response JSON-friendly
    return success_response(data=result.model_dump(), request=request)


@router.post(
    "/business/{business_id}/bulk-price-update/apply",
    summary="اعمال تغییر قیمت‌های گروهی",
    description="""
    اعمال تغییرات قیمت بر روی کالاهای انتخاب شده
    
    ### قابلیت‌ها:
    - تغییر قیمت بر اساس درصد یا مبلغ ثابت
    - تغییر قیمت فروش، خرید یا هر دو
    - فیلتر بر اساس دسته‌بندی، ارز، لیست قیمت و نوع محصول
    
    ### نکات مهم:
    - این عملیات تغییرات را به صورت واقعی اعمال می‌کند
    - قبل از اعمال، از endpoint `/bulk-price-update/preview` برای پیش‌نمایش استفاده کنید
    - تغییرات در قیمت پایه محصولات اعمال می‌شود
    - این عملیات قابل بازگشت نیست (مگر با تغییر مجدد)
    """,
    response_model=SuccessResponse[Dict[str, Any]],
    responses={
        200: {
            "description": "تغییرات قیمت با موفقیت اعمال شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "تغییرات قیمت با موفقیت اعمال شد",
                        "data": {
                            "total_products": 100,
                            "updated_products": 95,
                            "skipped_products": 5
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.edit"
        }
    }
)
@require_business_access("business_id")
def apply_bulk_price_update_endpoint(
    request: Request,
    business_id: int,
    payload: BulkPriceUpdateRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "edit")),
) -> Dict[str, Any]:
    
    result = apply_bulk_price_update(db, business_id, payload)
    return success_response(data=result, request=request)


@router.post(
    "/business/{business_id}/bulk-default-warehouse/preview",
    summary="پیش‌نمایش تغییر گروهی انبار پیش‌فرض کالاها",
    description="""
    پیش‌نمایش تغییر انبار پیش‌فرض برای چندین کالا به صورت همزمان
    
    ### قابلیت‌ها:
    - تغییر انبار پیش‌فرض برای لیستی از کالاها
    - اعمال بر روی کالاهای انبارداری، غیرانبارداری یا همه
    - نمایش کالاهایی که تغییر می‌یابند و کالاهایی که نادیده گرفته می‌شوند
    
    ### نکات:
    - این endpoint فقط پیش‌نمایش است و تغییری ایجاد نمی‌کند
    - برای اعمال تغییرات از endpoint `/bulk-default-warehouse/apply` استفاده کنید
    - کالاهای خدمتی (service) انبار پیش‌فرض ندارند
    """,
    response_model=SuccessResponse[BulkDefaultWarehousePreviewResponse],
    responses={
        200: {
            "description": "پیش‌نمایش تغییرات با موفقیت محاسبه شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "total_requested": 50,
                            "found_count": 48,
                            "will_update_count": 45,
                            "forced_service_null_count": 3,
                            "skipped": [
                                {
                                    "id": 100,
                                    "code": "S-001",
                                    "name": "خدمت A",
                                    "reason": "کالاهای خدمتی انبار پیش‌فرض ندارند"
                                }
                            ],
                            "notes": []
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.edit"
        }
    }
)
@require_business_access("business_id")
def preview_bulk_default_warehouse_endpoint(
    request: Request,
    business_id: int,
    payload: BulkDefaultWarehouseRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "edit")),
) -> Dict[str, Any]:
    result = preview_bulk_default_warehouse_update(db, business_id, payload)
    return success_response(data=result, request=request)


@router.post(
    "/business/{business_id}/bulk-default-warehouse/apply",
    summary="اعمال تغییر گروهی انبار پیش‌فرض کالاها",
    description="""
    اعمال تغییر انبار پیش‌فرض برای چندین کالا به صورت واقعی
    
    ### قابلیت‌ها:
    - تغییر انبار پیش‌فرض برای لیستی از کالاها
    - اعمال بر روی کالاهای انبارداری، غیرانبارداری یا همه
    - حذف انبار پیش‌فرض (با ارسال null)
    
    ### نکات مهم:
    - این عملیات تغییرات را به صورت واقعی اعمال می‌کند
    - قبل از اعمال، از endpoint `/bulk-default-warehouse/preview` برای پیش‌نمایش استفاده کنید
    - کالاهای خدمتی (service) نادیده گرفته می‌شوند
    - تغییر انبار پیش‌فرض بر روی موجودی تأثیری ندارد
    """,
    response_model=SuccessResponse[BulkDefaultWarehouseApplyResponse],
    responses={
        200: {
            "description": "تغییرات با موفقیت اعمال شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "انبار پیش‌فرض با موفقیت به‌روزرسانی شد",
                        "data": {
                            "total_requested": 50,
                            "found_count": 48,
                            "updated_count": 45,
                            "forced_service_null_count": 3,
                            "skipped": [],
                            "notes": []
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز - نیاز به مجوز products.edit"
        }
    }
)
@require_business_access("business_id")
def apply_bulk_default_warehouse_endpoint(
    request: Request,
    business_id: int,
    payload: BulkDefaultWarehouseRequest,
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("products", "edit")),
) -> Dict[str, Any]:
    result = apply_bulk_default_warehouse_update(db, business_id, ctx.get_user_id(), payload)
    db.commit()
    return success_response(data=result, request=request)


@router.post(
    "/businesses/{business_id}/reports/item-movements",
    summary="گزارش گردش کالا",
    description="""
    دریافت گزارش گردش کالاها در بازه زمانی مشخص
    
    ### اطلاعات گزارش:
    - ورود و خروج کالاها در بازه زمانی
    - مانده ابتدا و انتهای دوره
    - تفکیک بر اساس انبارها
    - فیلتر بر اساس محصولات، انبارها و تاریخ
    
    ### کاربردها:
    - بررسی حرکت کالاها
    - تحلیل ورود و خروج
    - کنترل موجودی در بازه زمانی
    """,
    response_model=SuccessResponse[Dict[str, Any]],
    responses={
        200: {
            "description": "گزارش با موفقیت ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "items": [
                                {
                                    "product_id": 1,
                                    "product_name": "کالا A",
                                    "warehouse_id": 1,
                                    "warehouse_name": "انبار اصلی",
                                    "opening_balance": 100,
                                    "incoming": 50,
                                    "outgoing": 30,
                                    "closing_balance": 120
                                }
                            ],
                            "total_count": 100,
                            "summary": {}
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی پارامترها"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def item_movements_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش گردش کالا"""
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    
    include_zero_balance = bool(body.get('include_zero_balance', False))
    search = body.get('search')
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_item_movements_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        product_ids=product_ids,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        include_zero_balance=include_zero_balance,
        search=search,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Item movements report retrieved successfully" if locale != 'fa' else "گزارش گردش کالا با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/item-movements/export/excel",
    summary="خروجی Excel گزارش گردش کالا",
    description="""
    دریافت گزارش گردش کالا به صورت فایل Excel
    
    ### قابلیت‌ها:
    - فیلتر بر اساس محصولات، انبارها و تاریخ
    - انتخاب سطرهای خاص برای خروجی
    - انتخاب و مرتب‌سازی ستون‌ها
    - پشتیبانی از RTL برای زبان فارسی
    
    ### فرمت فایل:
    - فرمت: `.xlsx` (Excel 2007+)
    """,
    responses={
        200: {
            "description": "فایل Excel با موفقیت ایجاد شد",
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                    "schema": {
                        "type": "string",
                        "format": "binary"
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def export_item_movements_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی Excel گزارش گردش کالا"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    import datetime
    import re
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    
    include_zero_balance = bool(body.get('include_zero_balance', False))
    search = body.get('search')
    
    # اعمال فیلتر سطرهای انتخاب شده
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    
    # دریافت گزارش با take بزرگ برای export
    max_export_records = 10000
    result = get_item_movements_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        product_ids=product_ids,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        include_zero_balance=include_zero_balance,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    
    # فیلتر سطرهای انتخاب شده
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                import json as _json
                indices = _json.loads(selected_indices)
            except Exception:
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # ستون‌های export
    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        headers = [col.get("label") or col.get("key") for col in export_columns]
        keys = [col.get("key") for col in export_columns]
    else:
        # ستون‌های پیش‌فرض
        locale = negotiate_locale(request.headers.get("Accept-Language"))
        is_fa = (locale == 'fa')
        default_columns = [
            ('product_code', 'کد کالا' if is_fa else 'Product Code'),
            ('product_name', 'نام کالا' if is_fa else 'Product Name'),
            ('unit', 'واحد' if is_fa else 'Unit'),
            ('category_name', 'دسته‌بندی' if is_fa else 'Category'),
            ('opening_balance', 'مانده ابتدای دوره' if is_fa else 'Opening Balance'),
            ('total_in', 'ورود' if is_fa else 'Total In'),
            ('total_out', 'خروج' if is_fa else 'Total Out'),
            ('closing_balance', 'مانده انتهای دوره' if is_fa else 'Closing Balance'),
        ]
        keys = [k for k, _ in default_columns]
        headers = [h for _, h in default_columns]
    
    # ساخت Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Movements"
    
    # Locale and RTL/LTR handling for Excel
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # افزودن ردیف‌های داده
    for it in items:
        row = []
        for k in keys:
            value = it.get(k)
            if value is None:
                row.append('')
            elif isinstance(value, (int, float)):
                row.append(float(value))
            else:
                row.append(str(value))
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            # Align data cells based on locale
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")
    
    # Auto width columns
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass
    
    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "item_movements"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(data)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/sales-by-product",
    summary="گزارش فروش به تفکیک کالا",
    description="""
    دریافت گزارش عملکرد فروش هر کالا در بازه زمانی
    
    ### اطلاعات گزارش:
    - تعداد و مقدار فروش هر کالا
    - مجموع مبلغ فروش
    - میانگین قیمت فروش
    - تعداد فاکتورها
    - فیلتر بر اساس محصولات، تاریخ و سایر معیارها
    
    ### کاربردها:
    - تحلیل عملکرد فروش محصولات
    - شناسایی پرفروش‌ترین کالاها
    - برنامه‌ریزی فروش و موجودی
    """,
    response_model=SuccessResponse[Dict[str, Any]],
    responses={
        200: {
            "description": "گزارش با موفقیت ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "items": [
                                {
                                    "product_id": 1,
                                    "product_name": "کالا A",
                                    "quantity": 100,
                                    "total_amount": 15000000,
                                    "average_price": 150000,
                                    "invoice_count": 10
                                }
                            ],
                            "total_count": 50,
                            "summary": {
                                "total_amount": 750000000,
                                "total_quantity": 5000
                            }
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی پارامترها"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def sales_by_product_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش فروش به تفکیک کالا"""
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    
    include_zero_sales = bool(body.get('include_zero_sales', False))
    search = body.get('search')
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_sales_by_product_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        product_ids=product_ids,
        category_ids=category_ids,
        warehouse_ids=warehouse_ids,
        include_zero_sales=include_zero_sales,
        search=search,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Sales by product report retrieved successfully" if locale != 'fa' else "گزارش فروش به تفکیک کالا با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/inventory-kardex",
    summary="گزارش کاردکس موجودی",
    description="""
    دریافت گزارش کاردکس (کاردکس) موجودی کالاها
    
    ### اطلاعات گزارش:
    - جزئیات تمام حرکات موجودی (ورود، خروج، انتقال)
    - مانده تجمعی پس از هر حرکت
    - تاریخ و زمان هر حرکت
    - نوع سند و شماره سند
    - فیلتر بر اساس محصولات، انبارها و تاریخ
    
    ### کاربردها:
    - ردیابی دقیق حرکات موجودی
    - بررسی تاریخچه موجودی
    - کنترل و ممیزی موجودی
    - تجزیه و تحلیل الگوهای ورود و خروج
    """,
    response_model=SuccessResponse[Dict[str, Any]],
    responses={
        200: {
            "description": "گزارش با موفقیت ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "items": [
                                {
                                    "date": "2024-01-15",
                                    "document_type": "فاکتور خرید",
                                    "document_number": "INV-001",
                                    "product_id": 1,
                                    "product_name": "کالا A",
                                    "warehouse_id": 1,
                                    "warehouse_name": "انبار اصلی",
                                    "incoming": 100,
                                    "outgoing": 0,
                                    "balance": 100
                                }
                            ],
                            "total_count": 500,
                            "summary": {}
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی پارامترها"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def inventory_kardex_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش کاردکس موجودی"""
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    
    search = body.get('search')
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_inventory_kardex_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        date_from=date_from,
        date_to=date_to,
        product_ids=product_ids,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        search=search,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Inventory kardex report retrieved successfully" if locale != 'fa' else "گزارش کاردکس موجودی با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/inventory-stock",
    summary="گزارش موجودی انبار",
    description="""
    دریافت گزارش موجودی فعلی محصولات در انبارها
    
    ### اطلاعات گزارش:
    - موجودی فعلی هر محصول در هر انبار
    - ارزش موجودی
    - موجودی رزرو شده
    - موجودی قابل استفاده
    - فیلترهای مختلف:
      - فقط کالاهای با کنترل موجودی
      - موجودی منفی
      - فاقد حواله (موجودی صفر)
      - فیلتر بر اساس محصولات و انبارها
    
    ### کاربردها:
    - کنترل موجودی فعلی
    - شناسایی موجودی منفی
    - مدیریت سفارشات
    - ارزیابی ارزش موجودی
    """,
    response_model=SuccessResponse[Dict[str, Any]],
    responses={
        200: {
            "description": "گزارش با موفقیت ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "data": {
                            "items": [
                                {
                                    "product_id": 1,
                                    "product_name": "کالا A",
                                    "product_code": "P-001",
                                    "warehouse_id": 1,
                                    "warehouse_name": "انبار اصلی",
                                    "quantity": 100,
                                    "reserved_quantity": 20,
                                    "available_quantity": 80,
                                    "unit_cost": 150000,
                                    "total_value": 15000000
                                }
                            ],
                            "total_count": 200,
                            "summary": {
                                "total_value": 3000000000
                            }
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی پارامترها"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def inventory_stock_report_endpoint(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "view")),
) -> Dict[str, Any]:
    """گزارش موجودی انبار"""
    # بررسی دسترسی
    if not ctx.can_read_section("reports"):
        raise ApiError("FORBIDDEN", "Missing business permission: reports.read", http_status=403)
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    # استخراج پارامترها از body
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    
    as_of_date = body.get('as_of_date')
    track_inventory = body.get('track_inventory')
    if track_inventory is not None:
        try:
            track_inventory = bool(track_inventory)
        except (ValueError, TypeError):
            track_inventory = None
    
    only_negative_stock = bool(body.get('only_negative_stock', False))
    only_without_movements = bool(body.get('only_without_movements', False))
    include_zero = bool(body.get('include_zero', False))
    search = body.get('search')
    
    # Pagination
    skip = body.get('skip', 0)
    take = body.get('take', 50)
    try:
        skip = int(skip)
        take = int(take)
        if take > 500:
            take = 500
        if take < 1:
            take = 50
        if skip < 0:
            skip = 0
    except (ValueError, TypeError):
        skip = 0
        take = 50
    
    result = get_inventory_stock_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        product_ids=product_ids,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        as_of_date=as_of_date,
        track_inventory=track_inventory,
        only_negative_stock=only_negative_stock,
        only_without_movements=only_without_movements,
        include_zero=include_zero,
        search=search,
        skip=skip,
        take=take,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    result['items'] = items
    
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    return success_response(
        data=result,
        message="Inventory stock report retrieved successfully" if locale != 'fa' else "گزارش موجودی انبار با موفقیت دریافت شد",
        request=request
    )


@router.post(
    "/businesses/{business_id}/reports/inventory-kardex/export/excel",
    summary="خروجی Excel گزارش کاردکس موجودی",
    description="""
    دریافت گزارش کاردکس موجودی به صورت فایل Excel
    
    ### قابلیت‌ها:
    - فیلتر بر اساس محصولات، انبارها و تاریخ
    - انتخاب سطرهای خاص برای خروجی
    - انتخاب و مرتب‌سازی ستون‌ها
    - نمایش تمام حرکات موجودی با مانده تجمعی
    - پشتیبانی از RTL برای زبان فارسی
    
    ### فرمت فایل:
    - فرمت: `.xlsx` (Excel 2007+)
    """,
    responses={
        200: {
            "description": "فایل Excel با موفقیت ایجاد شد",
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                    "schema": {
                        "type": "string",
                        "format": "binary"
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def export_inventory_kardex_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی Excel گزارش کاردکس موجودی"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    import datetime
    import re
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    
    search = body.get('search')
    
    # اعمال فیلتر سطرهای انتخاب شده
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    selected_row_keys = body.get('selected_row_keys')
    
    # دریافت گزارش با take بزرگ برای export
    max_export_records = 10000
    result = get_inventory_kardex_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        date_from=date_from,
        date_to=date_to,
        product_ids=product_ids,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # فیلتر سطرهای انتخاب شده (ترجیحاً بر اساس کلیدهای پایدار)
    if selected_only and selected_row_keys is not None and isinstance(selected_row_keys, list):
        try:
            wanted = []
            for k in selected_row_keys:
                if isinstance(k, dict):
                    wanted.append(k)
            if wanted:
                def _match_key(it: dict, key: dict) -> bool:
                    # Match by strongest key first
                    if key.get("document_id") is not None and it.get("document_id") is not None:
                        try:
                            if int(it.get("document_id")) != int(key.get("document_id")):
                                return False
                        except Exception:
                            return False
                    # Match by product/warehouse if present
                    for fld in ("product_id", "warehouse_id"):
                        if key.get(fld) is not None and it.get(fld) is not None:
                            try:
                                if int(it.get(fld)) != int(key.get(fld)):
                                    return False
                            except Exception:
                                return False
                    # Match movement
                    if key.get("movement") is not None and it.get("movement") is not None:
                        if str(it.get("movement")).lower() != str(key.get("movement")).lower():
                            return False
                    # Match document_code/date if provided
                    if key.get("document_code") is not None and it.get("document_code") is not None:
                        if str(it.get("document_code")) != str(key.get("document_code")):
                            return False
                    if key.get("document_date") is not None and it.get("document_date") is not None:
                        # Compare date-only portion
                        if str(it.get("document_date")).split("T")[0].split(" ")[0] != str(key.get("document_date")).split("T")[0].split(" ")[0]:
                            return False
                    return True
                filtered = []
                for it in items:
                    for k in wanted:
                        if _match_key(it, k):
                            filtered.append(it)
                            break
                items = filtered
        except Exception:
            pass

    # فیلتر سطرهای انتخاب شده (سازگاری عقب‌رو با اندیس‌ها)
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                import json as _json
                indices = _json.loads(selected_indices)
            except Exception:
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # ستون‌های export
    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        headers = [col.get("label") or col.get("key") for col in export_columns]
        keys = [col.get("key") for col in export_columns]
    else:
        # ستون‌های پیش‌فرض
        locale = negotiate_locale(request.headers.get("Accept-Language"))
        is_fa = (locale == 'fa')
        default_columns = [
            ('document_date', 'تاریخ' if is_fa else 'Document Date'),
            ('document_type_name', 'نوع سند' if is_fa else 'Document Type'),
            ('document_code', 'شماره سند' if is_fa else 'Document Code'),
            ('product_code', 'کد کالا' if is_fa else 'Product Code'),
            ('product_name', 'نام کالا' if is_fa else 'Product Name'),
            ('quantity_in', 'ورود' if is_fa else 'Quantity In'),
            ('quantity_out', 'خروج' if is_fa else 'Quantity Out'),
            ('balance', 'مانده' if is_fa else 'Balance'),
            ('unit_price', 'قیمت واحد' if is_fa else 'Unit Price'),
            ('total_amount', 'مبلغ کل' if is_fa else 'Total Amount'),
            ('warehouse_name', 'انبار' if is_fa else 'Warehouse'),
            ('description', 'توضیحات' if is_fa else 'Description'),
        ]
        keys = [k for k, _ in default_columns]
        headers = [h for _, h in default_columns]
    
    # ساخت Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Kardex"
    
    # Locale and RTL/LTR handling for Excel
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # افزودن ردیف‌های داده
    for it in items:
        row = []
        for k in keys:
            value = it.get(k)
            if value is None:
                row.append('')
            elif k == 'document_date' and value:
                # Format date based on calendar type
                formatted_date = format_date_for_export(it, 'document_date')
                row.append(formatted_date)
            elif isinstance(value, (int, float)):
                row.append(float(value))
            else:
                row.append(str(value))
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            # Align data cells based on locale
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")
    
    # Auto width columns
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass
    
    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()
    
    # Build meaningful filename
    biz_name = ""
    try:
        from adapters.db.models.business import Business
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "inventory_kardex"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(data)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/inventory-kardex/export/pdf",
    summary="خروجی PDF گزارش کاردکس موجودی",
    description="""
    دریافت گزارش کاردکس موجودی به صورت فایل PDF
    
    ### قابلیت‌ها:
    - فیلتر بر اساس محصولات، انبارها و تاریخ
    - نمایش تمام حرکات موجودی با مانده تجمعی
    - پشتیبانی از تقویم شمسی و میلادی (بر اساس header X-Calendar-Type)
    - پشتیبانی از RTL برای زبان فارسی
    - صفحه‌بندی خودکار
    
    ### فرمت فایل:
    - فرمت: `.pdf`
    """,
    responses={
        200: {
            "description": "فایل PDF با موفقیت ایجاد شد",
            "content": {
                "application/pdf": {
                    "schema": {
                        "type": "string",
                        "format": "binary"
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def export_inventory_kardex_report_pdf(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی PDF گزارش کاردکس موجودی"""
    from fastapi.responses import Response
    import datetime
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    from app.core.calendar import CalendarConverter
    from app.services.pdf.template_renderer import render_template, load_farsi_font_data_uris

    # Params
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id

    date_from = body.get('date_from')
    date_to = body.get('date_to')
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    search = body.get('search')

    # Build a compact, human-readable filters summary for PDF (printer friendly)
    filters_summary: list[dict] = []
    try:
        if date_from:
            filters_summary.append({"label": "از" if getattr(request.state, "locale", "fa") == "fa" else "From", "value": str(date_from)})
        if date_to:
            filters_summary.append({"label": "تا" if getattr(request.state, "locale", "fa") == "fa" else "To", "value": str(date_to)})
        if search:
            filters_summary.append({"label": "جستجو" if getattr(request.state, "locale", "fa") == "fa" else "Search", "value": str(search)})

        # Products / Warehouses / Categories: show names for short lists, otherwise show count
        locale_for_labels = getattr(request.state, "locale", "fa")
        is_fa_label = (locale_for_labels == "fa")

        if product_ids:
            try:
                if len(product_ids) <= 3:
                    rows = (
                        db.query(Product)
                        .filter(Product.business_id == business_id, Product.id.in_(product_ids))
                        .all()
                    )
                    names = []
                    for p in rows:
                        code = getattr(p, "code", None)
                        name = getattr(p, "name", None)
                        names.append((f"{code} - {name}" if code and name else (name or code or str(p.id))))
                    vtxt = "، ".join(names) if is_fa_label else ", ".join(names)
                else:
                    vtxt = f"{len(product_ids)} مورد" if is_fa_label else f"{len(product_ids)} items"
                filters_summary.append({"label": "کالا" if is_fa_label else "Products", "value": vtxt})
            except Exception:
                filters_summary.append({"label": "کالا" if is_fa_label else "Products", "value": f"{len(product_ids)}"})

        if warehouse_ids:
            try:
                from adapters.db.models.warehouse import Warehouse
                if len(warehouse_ids) <= 3:
                    rows = (
                        db.query(Warehouse)
                        .filter(Warehouse.business_id == business_id, Warehouse.id.in_(warehouse_ids))
                        .all()
                    )
                    names = [getattr(w, "name", None) or str(getattr(w, "id", "")) for w in rows]
                    vtxt = "، ".join(names) if is_fa_label else ", ".join(names)
                else:
                    vtxt = f"{len(warehouse_ids)} مورد" if is_fa_label else f"{len(warehouse_ids)} items"
                filters_summary.append({"label": "انبار" if is_fa_label else "Warehouses", "value": vtxt})
            except Exception:
                filters_summary.append({"label": "انبار" if is_fa_label else "Warehouses", "value": f"{len(warehouse_ids)}"})

        if category_ids:
            try:
                from adapters.db.models.support.category import Category
                if len(category_ids) <= 3:
                    rows = db.query(Category).filter(Category.id.in_(category_ids)).all()
                    names = [getattr(c, "name", None) or str(getattr(c, "id", "")) for c in rows]
                    vtxt = "، ".join(names) if is_fa_label else ", ".join(names)
                else:
                    vtxt = f"{len(category_ids)} مورد" if is_fa_label else f"{len(category_ids)} items"
                filters_summary.append({"label": "دسته" if is_fa_label else "Categories", "value": vtxt})
            except Exception:
                filters_summary.append({"label": "دسته" if is_fa_label else "Categories", "value": f"{len(category_ids)}"})
    except Exception:
        filters_summary = []

    selected_only = bool(body.get('selected_only', False))
    selected_row_keys = body.get('selected_row_keys')
    selected_indices = body.get('selected_indices')

    max_export_records = 10000
    result = get_inventory_kardex_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        date_from=date_from,
        date_to=date_to,
        product_ids=product_ids,
        warehouse_ids=warehouse_ids,
        category_ids=category_ids,
        search=search,
        skip=0,
        take=max_export_records,
    )

    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]

    # Determine calendar/locale
    calendar_type = getattr(request.state, "calendar_type", "jalali")
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = (locale == "fa")
    # زمان تولید گزارش بر اساس تقویم انتخابی
    try:
        _now = datetime.datetime.now()
        _gen = CalendarConverter.format_datetime(_now, calendar_type).get("formatted") or ""
        generated_at = " ".join(_gen.split(" ")[:2])
        if generated_at.count(":") >= 2:
            generated_at = generated_at.rsplit(":", 1)[0]
    except Exception:
        generated_at = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")

    # Apply selected rows filtering (prefer stable keys)
    if selected_only and selected_row_keys is not None and isinstance(selected_row_keys, list):
        try:
            wanted = [k for k in selected_row_keys if isinstance(k, dict)]
            if wanted:
                def _match_key(it: dict, key: dict) -> bool:
                    if key.get("document_id") is not None and it.get("document_id") is not None:
                        try:
                            if int(it.get("document_id")) != int(key.get("document_id")):
                                return False
                        except Exception:
                            return False
                    for fld in ("product_id", "warehouse_id"):
                        if key.get(fld) is not None and it.get(fld) is not None:
                            try:
                                if int(it.get(fld)) != int(key.get(fld)):
                                    return False
                            except Exception:
                                return False
                    if key.get("movement") is not None and it.get("movement") is not None:
                        if str(it.get("movement")).lower() != str(key.get("movement")).lower():
                            return False
                    if key.get("document_code") is not None and it.get("document_code") is not None:
                        if str(it.get("document_code")) != str(key.get("document_code")):
                            return False
                    if key.get("document_date") is not None and it.get("document_date") is not None:
                        if str(it.get("document_date")).split("T")[0].split(" ")[0] != str(key.get("document_date")).split("T")[0].split(" ")[0]:
                            return False
                    return True
                filtered = []
                for it in items:
                    for k in wanted:
                        if _match_key(it, k):
                            filtered.append(it)
                            break
                items = filtered
        except Exception:
            pass

    # Backward compatible selected_indices (may be mismatched if backend ordering differs)
    if selected_only and (not items) and selected_indices is not None:
        try:
            indices = None
            if isinstance(selected_indices, str):
                import json as _json
                indices = _json.loads(selected_indices)
            elif isinstance(selected_indices, list):
                indices = selected_indices
            if isinstance(indices, list):
                items = [result.get('items', [])[i] for i in indices if isinstance(i, int)]
        except Exception:
            pass

    # Note: get_inventory_kardex_report now returns document_date as a date object;
    # format_datetime_fields will convert it to the selected calendar string already.

    # Business name
    business_name = ""
    try:
        from adapters.db.models.business import Business
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    # Page params via query params (optional)
    try:
        qp = request.query_params
        paper_size = qp.get("paper_size")
        orientation = qp.get("orientation") or "landscape"
        disposition = qp.get("disposition") or "attachment"
    except Exception:
        paper_size = None
        orientation = "landscape"
        disposition = "attachment"

    # Provide font data URIs for base.html/templates
    fa_font_url_regular = ""
    fa_font_url_bold = ""
    try:
        if is_fa:
            fa_reg, fa_bold = load_farsi_font_data_uris()
            fa_font_url_regular = fa_reg or ""
            fa_font_url_bold = fa_bold or ""
    except Exception:
        fa_font_url_regular = ""
        fa_font_url_bold = ""

    # Template selection (custom templates supported)
    resolved_html = None
    explicit_template_id = None
    try:
        if body.get("template_id") is not None:
            explicit_template_id = int(body.get("template_id"))
    except Exception:
        explicit_template_id = None
    try:
        from app.services.report_template_service import ReportTemplateService
        template_context = {
            "title_text": "گزارش کاردکس موجودی" if is_fa else "Inventory Kardex Report",
            "business_name": business_name,
            "generated_at": generated_at,
            "is_fa": is_fa,
            "locale": locale,
            "paper_size": paper_size,
            "orientation": orientation,
            "fa_font_url_regular": fa_font_url_regular,
            "fa_font_url_bold": fa_font_url_bold,
            "footer_text": f"{'گزارش کاردکس موجودی' if is_fa else 'Inventory Kardex Report'} • {generated_at}",
            "filters_summary": filters_summary,
            "items": items,
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="inventory_kardex",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    if not resolved_html:
        final_html = render_template("pdf/inventory_kardex/list.html", {
            "title_text": "گزارش کاردکس موجودی" if is_fa else "Inventory Kardex Report",
            "business_name": business_name,
            "generated_at": generated_at,
            "is_fa": is_fa,
            "locale": locale,
            "paper_size": paper_size or "A4",
            "orientation": orientation or "landscape",
            "fa_font_url_regular": fa_font_url_regular,
            "fa_font_url_bold": fa_font_url_bold,
            "footer_text": f"{'گزارش کاردکس موجودی' if is_fa else 'Inventory Kardex Report'} • {generated_at}",
            "filters_summary": filters_summary,
            "items": items,
        })
    else:
        final_html = resolved_html

    page_css = f"@page {{ size: {(paper_size or 'A4')} {(orientation or 'landscape')}; margin: 1cm; }}"

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=final_html).write_pdf(
        stylesheets=[CSS(string=page_css)],
        font_config=font_config,
    )

    filename = f"inventory_kardex_{business_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/reports/sales-by-product/export/excel",
    summary="خروجی Excel گزارش فروش به تفکیک کالا",
    description="""
    دریافت گزارش فروش به تفکیک کالا به صورت فایل Excel
    
    ### قابلیت‌ها:
    - فیلتر بر اساس محصولات و تاریخ
    - انتخاب سطرهای خاص برای خروجی
    - انتخاب و مرتب‌سازی ستون‌ها
    - نمایش آمار فروش هر محصول
    - پشتیبانی از RTL برای زبان فارسی
    
    ### فرمت فایل:
    - فرمت: `.xlsx` (Excel 2007+)
    """,
    responses={
        200: {
            "description": "فایل Excel با موفقیت ایجاد شد",
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                    "schema": {
                        "type": "string",
                        "format": "binary"
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی درخواست"
        },
        401: {
            "description": "کاربر احراز هویت نشده است"
        },
        403: {
            "description": "دسترسی غیرمجاز"
        }
    }
)
@require_business_access("business_id")
async def export_sales_by_product_report_excel(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(default={}),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("reports", "export")),
):
    """خروجی Excel گزارش فروش به تفکیک کالا"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    import datetime
    import re
    
    # دریافت سال مالی از header یا body
    fiscal_year_id = None
    fy_header = request.headers.get('X-Fiscal-Year-ID')
    if fy_header:
        try:
            fiscal_year_id = int(fy_header)
        except (ValueError, TypeError):
            pass
    
    if body.get('fiscal_year_id'):
        try:
            fiscal_year_id = int(body['fiscal_year_id'])
        except (ValueError, TypeError):
            pass
    
    if not fiscal_year_id:
        from adapters.db.models.fiscal_year import FiscalYear
        fiscal_year = db.query(FiscalYear).filter(
            and_(
                FiscalYear.business_id == business_id,
                FiscalYear.is_last == True
            )
        ).first()
        if fiscal_year:
            fiscal_year_id = fiscal_year.id
    
    # استخراج پارامترها از body
    date_from = body.get('date_from')
    date_to = body.get('date_to')
    currency_id = body.get('currency_id')
    if currency_id is not None:
        try:
            currency_id = int(currency_id)
        except (ValueError, TypeError):
            currency_id = None
    
    product_ids = body.get('product_ids')
    if product_ids is not None and not isinstance(product_ids, list):
        product_ids = None
    
    category_ids = body.get('category_ids')
    if category_ids is not None and not isinstance(category_ids, list):
        category_ids = None
    
    warehouse_ids = body.get('warehouse_ids')
    if warehouse_ids is not None and not isinstance(warehouse_ids, list):
        warehouse_ids = None
    
    include_zero_sales = bool(body.get('include_zero_sales', False))
    search = body.get('search')
    
    # اعمال فیلتر سطرهای انتخاب شده
    selected_only = bool(body.get('selected_only', False))
    selected_indices = body.get('selected_indices')
    
    # دریافت گزارش با take بزرگ برای export
    max_export_records = 10000
    result = get_sales_by_product_report(
        db=db,
        business_id=business_id,
        fiscal_year_id=fiscal_year_id,
        currency_id=currency_id,
        date_from=date_from,
        date_to=date_to,
        product_ids=product_ids,
        category_ids=category_ids,
        warehouse_ids=warehouse_ids,
        include_zero_sales=include_zero_sales,
        search=search,
        skip=0,
        take=max_export_records,
    )
    
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]
    
    # Get calendar type
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_export(item_dict: dict, date_key: str) -> str:
        """Format date based on calendar type (date only, no time)"""
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item_dict:
            formatted_value = item_dict.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item_dict.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""
    
    # فیلتر سطرهای انتخاب شده
    if selected_only and selected_indices is not None:
        indices = None
        if isinstance(selected_indices, str):
            try:
                import json as _json
                indices = _json.loads(selected_indices)
            except Exception:
                indices = None
        elif isinstance(selected_indices, list):
            indices = selected_indices
        if isinstance(indices, list):
            items = [items[i] for i in indices if isinstance(i, int) and 0 <= i < len(items)]
    
    # ستون‌های export
    export_columns = body.get("export_columns")
    if export_columns and isinstance(export_columns, list):
        headers = [col.get("label") or col.get("key") for col in export_columns]
        keys = [col.get("key") for col in export_columns]
    else:
        # ستون‌های پیش‌فرض
        locale = negotiate_locale(request.headers.get("Accept-Language"))
        is_fa = (locale == 'fa')
        default_columns = [
            ('product_code', 'کد کالا' if is_fa else 'Product Code'),
            ('product_name', 'نام کالا' if is_fa else 'Product Name'),
            ('unit', 'واحد' if is_fa else 'Unit'),
            ('category_name', 'دسته‌بندی' if is_fa else 'Category'),
            ('total_quantity', 'تعداد فروش' if is_fa else 'Total Quantity'),
            ('total_amount', 'مبلغ کل فروش' if is_fa else 'Total Amount'),
            ('average_price', 'میانگین قیمت' if is_fa else 'Average Price'),
            ('last_sale_date', 'آخرین تاریخ فروش' if is_fa else 'Last Sale Date'),
        ]
        keys = [k for k, _ in default_columns]
        headers = [h for _, h in default_columns]
    
    # ساخت Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales by Product"
    
    # Locale and RTL/LTR handling for Excel
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    if locale == 'fa':
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # افزودن ردیف‌های داده
    for it in items:
        row = []
        for k in keys:
            value = it.get(k)
            if value is None:
                row.append('')
            elif k == 'last_sale_date' and value:
                # Format date based on calendar type
                formatted_date = format_date_for_export(it, 'last_sale_date')
                row.append(formatted_date)
            elif isinstance(value, (int, float)):
                row.append(float(value))
            else:
                row.append(str(value))
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            # Align data cells based on locale
            if locale == 'fa':
                cell.alignment = Alignment(horizontal="right")
    
    # Auto width columns
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass
    
    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()
    
    # Build meaningful filename
    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""
    
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    base = "sales_by_product"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    if selected_only:
        base += "_selected"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(data)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


