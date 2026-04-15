"""
API endpoints برای انتقال وجه (Transfers)

این ماژول شامل endpoint های مربوط به مدیریت اسناد انتقال وجه بین حساب‌های مختلف است.
اسناد انتقال برای جابجایی وجه بین حساب‌های بانکی، صندوق و تنخواه استفاده می‌شوند.
"""

from typing import Any, Dict, Optional
from fastapi import APIRouter, Depends, Request, Body, Path, Query
from sqlalchemy.orm import Session
from fastapi.responses import Response
import io, datetime, re, base64

from adapters.db.session import get_db
from adapters.db.models.document import Document
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.responses import success_response, format_datetime_fields, ApiError
from app.core.permissions import require_business_management_dep, require_business_access, require_business_permission_dep, require_business_permission_by_entity_dep
from adapters.api.v1.schemas import QueryInfo, SuccessResponse
from adapters.api.v1.schema_models.transfer import (
    TransferCreateRequest,
    TransferUpdateRequest,
    TransferResponse,
    TransferListResponse,
    TransferExportRequest
)
from app.services.transfer_service import (
    create_transfer,
    get_transfer,
    list_transfers,
    delete_transfer,
    update_transfer,
)
from adapters.db.models.business import Business
from adapters.db.models.user import User
from adapters.db.models.business_print_settings import BusinessPrintSettings
from app.services.file_storage_service import FileStorageService
from app.core.cache import get_cache


router = APIRouter(tags=["اسناد انتقال", "مدیریت مالی"])


@router.post(
    "/businesses/{business_id}/transfers",
    summary="لیست اسناد انتقال",
    description="""
    دریافت لیست اسناد انتقال با امکانات پیشرفته فیلتر، جستجو و صفحه‌بندی
    
    ### قابلیت‌های جستجو:
    - جستجو در کد سند، نام مبدا، نام مقصد و توضیحات
    - جستجو در مبلغ و تاریخ
    
    ### فیلترهای موجود:
    - **محدوده تاریخ**: `from_date`, `to_date`
    - **نوع مبدا**: `source_type` (bank_account, cash_register, petty_cash)
    - **نوع مقصد**: `destination_type`
    - **محدوده مبلغ**: فیلتر با operator های `>=` و `<=` روی `total_amount`
    - **سال مالی**: از طریق header `X-Fiscal-Year-ID`
    
    ### مرتب‌سازی:
    می‌توانید بر اساس فیلدهای زیر مرتب کنید:
    - `document_date`: تاریخ سند (پیش‌فرض)
    - `total_amount`: مبلغ کل
    - `created_at`: تاریخ ایجاد
    - `code`: کد سند
    - `source_name`: نام مبدا
    - `destination_name`: نام مقصد
    
    ### نکات:
    - نتایج به صورت صفحه‌بندی شده برمی‌گردند
    - حداکثر 1000 رکورد در هر درخواست
    - برای صادرات تعداد زیاد، از endpoint های export استفاده کنید
    """,
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "لیست اسناد با موفقیت دریافت شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "لیست اسناد انتقال دریافت شد",
                        "data": {
                            "items": [
                                {
                                    "id": 123,
                                    "code": "T-1001",
                                    "document_date": "1403/10/15",
                                    "source_type_name": "حساب بانکی",
                                    "source_name": "بانک ملت - 1234567890",
                                    "destination_type_name": "صندوق",
                                    "destination_name": "صندوق اصلی",
                                    "total_amount": 1000000,
                                    "commission": 5000,
                                    "created_by_name": "احمد احمدی"
                                }
                            ],
                            "total_count": 45,
                            "has_more": True
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در پارامترهای درخواست",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "INVALID_PARAMETERS",
                        "message": "تاریخ شروع باید قبل از تاریخ پایان باشد"
                    }
                }
            }
        },
        403: {
            "description": "عدم دسترسی به کسب‌وکار",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "FORBIDDEN",
                        "message": "شما به این کسب‌وکار دسترسی ندارید"
                    }
                }
            }
        }
    }
)
@require_business_access("business_id")
async def list_transfers_endpoint(
    request: Request,
    business_id: int = Path(
        ..., 
        description="شناسه کسب‌وکار",
        examples={"example": {"value": 1}},
        gt=0
    ),
    query_info: QueryInfo = Body(
        ...,
        description="پارامترهای جستجو، فیلتر و صفحه‌بندی",
        examples={
            "example": {
                "take": 20,
                "skip": 0,
                "sort_by": "document_date",
                "sort_desc": True,
                "search": "بانک ملت",
                "filters": [
                    {
                        "property": "total_amount",
                        "operator": ">=",
                        "value": 1000000
                    }
                ]
            }
        }
    ),
    x_fiscal_year_id: Optional[int] = Query(
        None,
        alias="X-Fiscal-Year-ID",
        description="شناسه سال مالی (اگر ارسال نشود، سال مالی فعال استفاده می‌شود)",
        examples={"example": {"value": 1}}
    ),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """
    دریافت لیست اسناد انتقال
    
    این endpoint برای نمایش لیست اسناد انتقال با قابلیت‌های پیشرفته استفاده می‌شود.
    شامل جستجو، فیلتر، مرتب‌سازی و صفحه‌بندی.
    """
    query_dict: Dict[str, Any] = {
        "take": query_info.take,
        "skip": query_info.skip,
        "sort_by": query_info.sort_by,
        "sort_desc": query_info.sort_desc,
        "sort": [s.model_dump() for s in query_info.sort] if query_info.sort else None,
        "search": query_info.search,
    }
    try:
        body_json = await request.json()
        if isinstance(body_json, dict):
            # Forward simple date range params
            for key in ["from_date", "to_date", "sort", "sort_by", "sort_desc"]:
                if key in body_json:
                    query_dict[key] = body_json[key]
            # Forward advanced filters from DataTable (e.g., document_date range)
            if "filters" in body_json:
                query_dict["filters"] = body_json.get("filters")
    except Exception:
        pass

    try:
        fy_header = request.headers.get("X-Fiscal-Year-ID")
        if fy_header:
            query_dict["fiscal_year_id"] = int(fy_header)
    except Exception:
        pass

    # کش نتایج لیست انتقال‌ها
    cache = get_cache()
    cache_key = None
    fiscal_year_id = query_dict.get("fiscal_year_id")

    if cache.enabled:
        import json, hashlib
        key_payload = {
            "business_id": business_id,
            "query": query_dict,
        }
        key_str = json.dumps(key_payload, sort_keys=True, ensure_ascii=False)
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()[:16]
        cache_key = f"transfers_list:{key_hash}"
        cached = cache.get(cache_key)
        if cached is not None:
            return success_response(data=cached, request=request, message="TRANSFERS_LIST_FETCHED")

    result = list_transfers(db, business_id, query_dict)
    result["items"] = [format_datetime_fields(item, request) for item in result.get("items", [])]

    # ذخیره در cache با tag-based caching
    if cache.enabled and cache_key:
        cache.set_with_documents_tag(
            key=cache_key,
            value=result,
            business_id=business_id,
            fiscal_year_id=fiscal_year_id,
            document_type="transfer",
            ttl=60
        )

    return success_response(data=result, request=request, message="TRANSFERS_LIST_FETCHED")


@router.post(
    "/businesses/{business_id}/transfers/create",
    summary="ایجاد سند انتقال",
    description="""
    ایجاد سند انتقال جدید برای جابجایی وجه بین حساب‌های مختلف
    
    ### انواع حساب مبدا/مقصد:
    - `bank_account`: حساب بانکی - برای انتقال از/به حساب‌های بانکی
    - `cash_register`: صندوق - برای انتقال از/به صندوق‌های نقدی
    - `petty_cash`: تنخواه - برای انتقال از/به تنخواه‌گردان‌ها
    
    ### قوانین و محدودیت‌ها:
    - ✅ مبلغ انتقال باید بزرگتر از صفر باشد
    - ✅ مبدا و مقصد نمی‌توانند یکسان باشند
    - ✅ تاریخ سند باید در بازه سال مالی فعال باشد
    - ✅ کارمزد (اختیاری) باید صفر یا مثبت باشد
    - ✅ کاربر باید مجوز "ایجاد سند انتقال" داشته باشد
    
    ### فرآیند ثبت:
    1. اعتبارسنجی مبدا و مقصد
    2. بررسی موجودی کافی در مبدا (در صورت نیاز)
    3. ایجاد سند انتقال
    4. ثبت اتوماتیک در دفتر کل حسابداری
    5. به‌روزرسانی موجودی حساب‌ها
    
    ### نکات:
    - سند انتقال به صورت خودکار در دفتر کل ثبت می‌شود
    - می‌توانید کارمزد انتقال را جداگانه ثبت کنید
    - تاریخ را می‌توانید به فرمت ISO (YYYY-MM-DD) یا جلالی (YYYY/MM/DD) ارسال کنید
    """,
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "سند انتقال با موفقیت ایجاد شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "سند انتقال با موفقیت ایجاد شد",
                        "data": {
                            "id": 123,
                            "code": "T-1001",
                            "business_id": 1,
                            "document_type_name": "انتقال",
                            "source_type": "bank_account",
                            "source_name": "بانک ملت - 1234567890",
                            "destination_type": "cash_register",
                            "destination_name": "صندوق اصلی",
                            "total_amount": 1000000,
                            "commission": 5000,
                            "document_date": "1403/10/15",
                            "created_by_name": "احمد احمدی",
                            "created_at": "1403/10/15 14:30:00"
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی داده‌ها",
            "content": {
                "application/json": {
                    "examples": {
                        "invalid_amount": {
                            "summary": "مبلغ نامعتبر",
                            "value": {
                                "success": False,
                                "error_code": "INVALID_AMOUNT",
                                "message": "مبلغ باید بزرگتر از صفر باشد"
                            }
                        },
                        "same_source_destination": {
                            "summary": "مبدا و مقصد یکسان",
                            "value": {
                                "success": False,
                                "error_code": "SAME_SOURCE_DESTINATION",
                                "message": "مبدا و مقصد نمی‌توانند یکسان باشند"
                            }
                        },
                        "invalid_date": {
                            "summary": "تاریخ خارج از سال مالی",
                            "value": {
                                "success": False,
                                "error_code": "DATE_OUT_OF_FISCAL_YEAR",
                                "message": "تاریخ سند باید در بازه سال مالی فعال باشد"
                            }
                        },
                        "insufficient_balance": {
                            "summary": "موجودی ناکافی",
                            "value": {
                                "success": False,
                                "error_code": "INSUFFICIENT_BALANCE",
                                "message": "موجودی حساب مبدا کافی نیست"
                            }
                        }
                    }
                }
            }
        },
        403: {
            "description": "عدم دسترسی - کاربر مجوز ایجاد سند انتقال را ندارد",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "FORBIDDEN",
                        "message": "شما مجوز ایجاد سند انتقال را ندارید"
                    }
                }
            }
        },
        404: {
            "description": "مبدا یا مقصد یافت نشد",
            "content": {
                "application/json": {
                    "examples": {
                        "source_not_found": {
                            "summary": "مبدا یافت نشد",
                            "value": {
                                "success": False,
                                "error_code": "SOURCE_NOT_FOUND",
                                "message": "حساب مبدا یافت نشد"
                            }
                        },
                        "destination_not_found": {
                            "summary": "مقصد یافت نشد",
                            "value": {
                                "success": False,
                                "error_code": "DESTINATION_NOT_FOUND",
                                "message": "حساب مقصد یافت نشد"
                            }
                        }
                    }
                }
            }
        }
    }
)
@require_business_access("business_id")
async def create_transfer_endpoint(
    request: Request,
    business_id: int = Path(
        ..., 
        description="شناسه کسب‌وکار",
        examples={"example": {"value": 1}},
        gt=0
    ),
    body: TransferCreateRequest = Body(
        ...,
        description="اطلاعات سند انتقال",
        examples={
            "example": {
                "source_type": "bank_account",
                "source_id": 1,
                "destination_type": "cash_register",
                "destination_id": 2,
                "total_amount": 1000000,
                "commission": 5000,
                "document_date": "2024-01-15",
                "currency_id": 1,
                "description": "انتقال وجه بابت خرید مواد اولیه"
            }
        }
    ),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_dep("transfers", "add")),
):
    """
    ایجاد سند انتقال جدید
    
    این endpoint برای ثبت انتقال وجه بین حساب‌های مختلف استفاده می‌شود.
    سند به صورت خودکار در دفتر کل ثبت شده و موجودی حساب‌ها به‌روزرسانی می‌شود.
    """
    # تبدیل داده‌های flat به nested برای سازگاری با سرویس
    # تبدیل bank_account به bank (سرویس انتظار دارد bank باشد)
    source_type = body.source_type
    if source_type == "bank_account":
        source_type = "bank"
    
    destination_type = body.destination_type
    if destination_type == "bank_account":
        destination_type = "bank"
    
    # ساخت ساختار nested برای سرویس
    # استفاده از dict برای اطمینان از اینکه همه فیلدها به درستی منتقل می‌شوند
    body_dict = body.dict(exclude_none=False)
    
    # اطمینان از اینکه currency_id وجود دارد
    currency_id_value = body_dict.get("currency_id")
    if not currency_id_value:
        raise ApiError("CURRENCY_REQUIRED", "currency_id is required", http_status=400)
    
    service_data = {
        "document_date": body_dict.get("document_date"),
        "currency_id": int(currency_id_value),
        "source": {
            "type": source_type,
            "id": body_dict.get("source_id")
        },
        "destination": {
            "type": destination_type,
            "id": body_dict.get("destination_id")
        },
        "amount": body_dict.get("total_amount"),  # مبلغ انتقال (بدون commission)
    }
    
    # اضافه کردن فیلدهای اختیاری
    if body_dict.get("commission") is not None:
        service_data["commission"] = body_dict.get("commission")
    if body_dict.get("description") is not None:
        service_data["description"] = body_dict.get("description")
    if body_dict.get("fiscal_year_id") is not None:
        service_data["fiscal_year_id"] = body_dict.get("fiscal_year_id")
    
    created = create_transfer(db, business_id, ctx.get_user_id(), service_data)
    return success_response(data=format_datetime_fields(created, request), request=request, message="TRANSFER_CREATED")


@router.get(
    "/transfers/{document_id}",
    summary="جزئیات سند انتقال",
    description="""
    دریافت جزئیات کامل یک سند انتقال
    
    ### اطلاعات برگشتی:
    - اطلاعات کامل مبدا و مقصد
    - آیتم‌های حسابداری سند
    - اطلاعات ایجادکننده و ویرایش‌کننده
    - تاریخ‌های ثبت و ویرایش
    - وضعیت سند
    
    ### کاربردها:
    - نمایش جزئیات سند
    - چاپ و دانلود PDF
    - ویرایش سند (دریافت اطلاعات فعلی)
    - بررسی سابقه تراکنش‌ها
    """,
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "جزئیات سند با موفقیت دریافت شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "جزئیات سند انتقال دریافت شد",
                        "data": {
                            "id": 123,
                            "code": "T-1001",
                            "business_id": 1,
                            "source_type": "bank_account",
                            "source_name": "بانک ملت - 1234567890",
                            "destination_type": "cash_register",
                            "destination_name": "صندوق اصلی",
                            "total_amount": 1000000,
                            "commission": 5000,
                            "document_date": "1403/10/15",
                            "description": "انتقال وجه بابت خرید مواد اولیه",
                            "account_lines": [
                                {
                                    "account_code": "1201",
                                    "account_name": "بانک ملت",
                                    "debit": 0,
                                    "credit": 1005000
                                },
                                {
                                    "account_code": "1101",
                                    "account_name": "صندوق",
                                    "debit": 1000000,
                                    "credit": 0
                                }
                            ],
                            "created_by_name": "احمد احمدی",
                            "created_at": "1403/10/15 14:30:00"
                        }
                    }
                }
            }
        },
        403: {
            "description": "عدم دسترسی به سند",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "FORBIDDEN",
                        "message": "شما به این سند دسترسی ندارید"
                    }
                }
            }
        },
        404: {
            "description": "سند یافت نشد",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "DOCUMENT_NOT_FOUND",
                        "message": "سند انتقال یافت نشد"
                    }
                }
            }
        }
    }
)
async def get_transfer_endpoint(
    request: Request,
    document_id: int = Path(
        ..., 
        description="شناسه سند انتقال",
        examples={"example": {"value": 123}},
        gt=0
    ),
    include_lines: bool = Query(
        False,
        description="شامل کردن آیتم‌های حسابداری سند"
    ),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
):
    """
    دریافت جزئیات کامل سند انتقال
    
    این endpoint تمام اطلاعات مربوط به یک سند انتقال را برمی‌گرداند.
    """
    result = get_transfer(db, document_id)
    if not result:
        raise ApiError("DOCUMENT_NOT_FOUND", "سند انتقال یافت نشد", http_status=404)
    business_id = result.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "شما به این سند دسترسی ندارید", http_status=403)
    return success_response(data=format_datetime_fields(result, request), request=request, message="TRANSFER_DETAILS")


@router.get(
    "/transfers/{document_id}/pdf",
    summary="خروجی PDF تک سند انتقال",
    description="خروجی PDF یک سند انتقال",
)
async def export_single_transfer_pdf(
    document_id: int,
    request: Request,
    auth_context: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    template_id: int | None = None,
):
    """خروجی PDF تک سند انتقال"""
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from app.core.i18n import negotiate_locale
    from app.core.calendar import CalendarConverter, get_calendar_type_from_header
    from html import escape
    from app.services.pdf.template_renderer import render_template
    
    # دریافت سند
    result = get_transfer(db, document_id)
    if not result:
        raise ApiError(
            "DOCUMENT_NOT_FOUND",
            "Transfer document not found",
            http_status=404
        )
    
    # بررسی دسترسی
    business_id = result.get("business_id")
    if business_id and not auth_context.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "Access denied", http_status=403)
    
    # دریافت اطلاعات کسب‌وکار + فایل‌های گرافیکی (لوگو/مهر/امضا)
    business_name = ""
    business_logo_data_uri: Optional[str] = None
    business_stamp_data_uri: Optional[str] = None
    owner_signature_data_uri: Optional[str] = None
    storage = FileStorageService(db)

    async def _load_image_data_uri(file_id_str: Optional[str]) -> Optional[str]:
        if not file_id_str:
            return None
        try:
            from uuid import UUID
            try:
                file_data = await storage.download_file(UUID(str(file_id_str)))
            except Exception:
                return None
            content: bytes = file_data.get("content") or b""
            if not content:
                return None
            mime = file_data.get("mime_type") or "image/png"
            b64 = base64.b64encode(content).decode("ascii")
            return f"data:{mime};base64,{b64}"
        except Exception:
            return None

    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
            # تنظیمات چاپ
            try:
                rows = (
                    db.query(BusinessPrintSettings)
                    .filter(BusinessPrintSettings.business_id == business_id)
                    .all()
                )
            except Exception:
                rows = []

            def _pick_cfg() -> dict:
                cfg = {"show_logo": True, "show_stamp": True, "footer_note": None}
                per_type = None
                for r in rows:
                    if r.document_type == "all":
                        cfg = {
                            "show_logo": bool(getattr(r, "show_logo", True)),
                            "show_stamp": bool(getattr(r, "show_stamp", True)),
                            "footer_note": getattr(r, "footer_note", None),
                        }
                    elif r.document_type == "transfer":
                        per_type = {
                            "show_logo": bool(getattr(r, "show_logo", True)),
                            "show_stamp": bool(getattr(r, "show_stamp", True)),
                            "footer_note": getattr(r, "footer_note", None),
                        }
                if per_type:
                    merged = dict(cfg)
                    merged.update({k: v for k, v in per_type.items() if v is not None})
                    return merged
                return cfg

            cfg = _pick_cfg()
            if cfg.get("show_logo", True):
                business_logo_data_uri = await _load_image_data_uri(getattr(b, "logo_file_id", None))
            if cfg.get("show_stamp", True):
                business_stamp_data_uri = await _load_image_data_uri(getattr(b, "stamp_file_id", None))
                try:
                    owner_user = db.query(User).filter(User.id == b.owner_id).first()
                except Exception:
                    owner_user = None
                if owner_user is not None:
                    owner_signature_data_uri = await _load_image_data_uri(getattr(owner_user, "signature_file_id", None))
    except Exception:
        business_name = business_name or ""

    # Locale handling
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Get calendar type for date formatting
    calendar_type = get_calendar_type_from_header(request.headers.get("X-Calendar-Type"))
    if not calendar_type:
        calendar_type = "jalali" if is_fa else "gregorian"
    
    # آماده‌سازی داده‌ها
    doc_type_name = result.get("document_type_name", "انتقال")
    doc_code = result.get("code", "")
    doc_date_dt = result.get("document_date")
    doc_date = ""
    if doc_date_dt:
        try:
            if isinstance(doc_date_dt, datetime.datetime):
                formatted = CalendarConverter.format_datetime(doc_date_dt, calendar_type)
                doc_date = formatted.get("date_only", "") or formatted.get("formatted", "")
            else:
                doc_date = str(doc_date_dt)
        except Exception:
            doc_date = str(doc_date_dt) if doc_date_dt else ""
    
    total_amount = result.get("total_amount", 0)
    commission = result.get("commission", 0)
    description = result.get("description", "")
    account_lines = result.get("account_lines", [])
    source_type = result.get("source_type", "")
    source_type_name = result.get("source_type_name", "")
    source_name = result.get("source_name", "")
    destination_type = result.get("destination_type", "")
    destination_type_name = result.get("destination_type_name", "")
    destination_name = result.get("destination_name", "")
    
    # تاریخ تولید
    try:
        _now = datetime.datetime.now()
        _fd = CalendarConverter.format_datetime(_now, calendar_type)
        generated_at = _fd.get("formatted") or _fd.get("date_only") or _now.strftime('%Y/%m/%d %H:%M')
    except Exception:
        generated_at = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    title_text = f"سند {doc_type_name}" if is_fa else f"{doc_type_name} Document"
    footer_text = f"تولید شده در {generated_at}" if is_fa else f"Generated at {generated_at}"

    # تلاش برای رندر با قالب سفارشی (transfers/detail)
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if template_id is not None:
                explicit_template_id = int(template_id)
        except Exception:
            explicit_template_id = None
        template_context = {
            "business_id": business_id,
            "business_name": business_name,
            "document": result,
            "account_lines": account_lines,
            "code": doc_code,
            "document_date": doc_date,
            "total_amount": total_amount,
            "commission": commission,
            "description": description,
            "source_type": source_type,
            "source_type_name": source_type_name,
            "source_name": source_name,
            "destination_type": destination_type,
            "destination_type_name": destination_type_name,
            "destination_name": destination_name,
            "title_text": title_text,
            "generated_at": generated_at,
            "is_fa": is_fa,
            "business_logo_data_uri": business_logo_data_uri,
            "business_stamp_data_uri": business_stamp_data_uri,
            "owner_signature_data_uri": owner_signature_data_uri,
        }
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="transfers",
            subtype="detail",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # HTML پیش‌فرض در نبود قالب: فایل قالب + پارامترها
    try:
        qp = request.query_params
        paper_size = qp.get("paper_size")
        orientation = qp.get("orientation")
        disposition = qp.get("disposition") or "attachment"
    except Exception:
        paper_size = None
        orientation = None
        disposition = "attachment"
    html_content = resolved_html or render_template(
        "pdf/transfers/detail.html",
        {
            "business_id": business_id,
            "business_name": business_name,
            "document": result,
            "account_lines": account_lines,
            "code": doc_code,
            "document_date": doc_date,
            "total_amount": total_amount,
            "commission": commission,
            "description": description,
            "source_type": source_type,
            "source_type_name": source_type_name,
            "source_name": source_name,
            "destination_type": destination_type,
            "destination_type_name": destination_type_name,
            "destination_name": destination_name,
            "title_text": title_text,
            "generated_at": generated_at,
            "is_fa": is_fa,
            "paper_size": paper_size,
            "orientation": orientation,
            "footer_text": footer_text,
            "business_logo_data_uri": business_logo_data_uri,
            "business_stamp_data_uri": business_stamp_data_uri,
            "owner_signature_data_uri": owner_signature_data_uri,
        },
    )

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=html_content).write_pdf(font_config=font_config)

    # Build filename
    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    
    filename = f"transfer_{slugify(doc_code)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"{disposition}; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.delete(
    "/transfers/{document_id}",
    summary="حذف سند انتقال",
    description="""
    حذف یک سند انتقال موجود
    
    ### عملیات حذف:
    - سند از سیستم حذف می‌شود
    - آیتم‌های حسابداری سند حذف می‌شوند
    - موجودی حساب‌ها به حالت قبل برمی‌گردد
    - عملیات قابل بازگشت نیست
    
    ### محدودیت‌ها:
    - نمی‌توانید سند قفل شده را حذف کنید
    - نمی‌توانید سند تایید شده را حذف کنید (ابتدا باید لغو تایید شود)
    - نمی‌توانید سند مرتبط با سایر اسناد را حذف کنید
    
    ### هشدار:
    ⚠️ این عملیات غیرقابل بازگشت است! سند به صورت کامل از سیستم حذف می‌شود.
    
    ### نکات امنیتی:
    - نیاز به مجوز "حذف سند انتقال" دارید
    - حذف در لاگ سیستم ثبت می‌شود
    - اطلاعات کاربر حذف‌کننده ذخیره می‌شود
    """,
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "سند با موفقیت حذف شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "سند انتقال با موفقیت حذف شد",
                        "data": {
                            "deleted": True,
                            "document_id": 123
                        }
                    }
                }
            }
        },
        400: {
            "description": "سند قابل حذف نیست",
            "content": {
                "application/json": {
                    "examples": {
                        "locked_document": {
                            "summary": "سند قفل شده",
                            "value": {
                                "success": False,
                                "error_code": "DOCUMENT_LOCKED",
                                "message": "سند قفل شده و قابل حذف نیست"
                            }
                        },
                        "confirmed_document": {
                            "summary": "سند تایید شده",
                            "value": {
                                "success": False,
                                "error_code": "DOCUMENT_CONFIRMED",
                                "message": "سند تایید شده است. ابتدا باید لغو تایید شود"
                            }
                        },
                        "has_dependencies": {
                            "summary": "سند دارای وابستگی",
                            "value": {
                                "success": False,
                                "error_code": "HAS_DEPENDENCIES",
                                "message": "این سند با سایر اسناد مرتبط است و قابل حذف نیست"
                            }
                        }
                    }
                }
            }
        },
        403: {
            "description": "عدم مجوز حذف",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "FORBIDDEN",
                        "message": "شما مجوز حذف این سند را ندارید"
                    }
                }
            }
        },
        404: {
            "description": "سند یافت نشد",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "DOCUMENT_NOT_FOUND",
                        "message": "سند انتقال یافت نشد"
                    }
                }
            }
        }
    }
)
async def delete_transfer_endpoint(
    request: Request,
    document_id: int = Path(
        ..., 
        description="شناسه سند انتقال برای حذف",
        examples={"example": {"value": 123}},
        gt=0
    ),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("transfers", "delete", Document, "document_id")),
):
    """
    حذف سند انتقال
    
    ⚠️ هشدار: این عملیات غیرقابل بازگشت است!
    """
    result = get_transfer(db, document_id)
    if result:
        business_id = result.get("business_id")
        if business_id and not ctx.can_access_business(business_id):
            raise ApiError("FORBIDDEN", "شما به این سند دسترسی ندارید", http_status=403)
    ok = delete_transfer(db, document_id)
    if not ok:
        raise ApiError("DOCUMENT_NOT_FOUND", "سند انتقال یافت نشد", http_status=404)
    return success_response(data={"deleted": True, "document_id": document_id}, request=request, message="TRANSFER_DELETED")


@router.put(
    "/transfers/{document_id}",
    summary="ویرایش سند انتقال",
    description="""
    ویرایش و به‌روزرسانی یک سند انتقال موجود
    
    ### فیلدهای قابل ویرایش:
    - مبدا و مقصد (نوع و شناسه)
    - مبلغ کل
    - کارمزد
    - تاریخ سند
    - توضیحات
    
    ### محدودیت‌ها:
    - نمی‌توانید سند قفل شده را ویرایش کنید
    - نمی‌توانید سند تایید شده در دفتر کل را ویرایش کنید (باید ابتدا لغو تایید شود)
    - تاریخ جدید باید در بازه سال مالی باشد
    - مبدا و مقصد نمی‌توانند یکسان باشند
    
    ### نکات:
    - فقط فیلدهایی که ارسال می‌کنید تغییر می‌کنند
    - سایر فیلدها بدون تغییر باقی می‌مانند
    - پس از ویرایش، موجودی حساب‌ها خودکار به‌روزرسانی می‌شود
    - سابقه ویرایش ثبت می‌شود
    """,
    response_model=SuccessResponse,
    responses={
        200: {
            "description": "سند با موفقیت ویرایش شد",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "سند انتقال با موفقیت ویرایش شد",
                        "data": {
                            "id": 123,
                            "code": "T-1001",
                            "total_amount": 1200000,
                            "description": "انتقال وجه بابت خرید مواد اولیه - ویرایش شده",
                            "updated_at": "1403/10/16 09:15:00"
                        }
                    }
                }
            }
        },
        400: {
            "description": "خطا در اعتبارسنجی",
            "content": {
                "application/json": {
                    "examples": {
                        "locked_document": {
                            "summary": "سند قفل شده",
                            "value": {
                                "success": False,
                                "error_code": "DOCUMENT_LOCKED",
                                "message": "این سند قفل شده و قابل ویرایش نیست"
                            }
                        },
                        "confirmed_document": {
                            "summary": "سند تایید شده",
                            "value": {
                                "success": False,
                                "error_code": "DOCUMENT_CONFIRMED",
                                "message": "سند تایید شده است. ابتدا باید لغو تایید شود"
                            }
                        }
                    }
                }
            }
        },
        403: {
            "description": "عدم مجوز ویرایش",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "FORBIDDEN",
                        "message": "شما مجوز ویرایش این سند را ندارید"
                    }
                }
            }
        },
        404: {
            "description": "سند یافت نشد",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "error_code": "DOCUMENT_NOT_FOUND",
                        "message": "سند انتقال یافت نشد"
                    }
                }
            }
        }
    }
)
async def update_transfer_endpoint(
    request: Request,
    document_id: int = Path(
        ..., 
        description="شناسه سند انتقال",
        examples={"example": {"value": 123}},
        gt=0
    ),
    body: TransferUpdateRequest = Body(
        ...,
        description="فیلدهای جدید برای ویرایش (فقط فیلدهای ارسال شده تغییر می‌کنند)",
        examples={
            "example": {
                "total_amount": 1200000,
                "description": "انتقال وجه بابت خرید مواد اولیه - ویرایش شده"
            }
        }
    ),
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(get_current_user),
    _: None = Depends(require_business_permission_by_entity_dep("transfers", "edit", Document, "document_id")),
):
    """
    ویرایش سند انتقال موجود
    
    فقط فیلدهایی که در body ارسال می‌شوند تغییر می‌کنند.
    """
    result = get_transfer(db, document_id)
    if not result:
        raise ApiError("DOCUMENT_NOT_FOUND", "سند انتقال یافت نشد", http_status=404)
    business_id = result.get("business_id")
    if business_id and not ctx.can_access_business(business_id):
        raise ApiError("FORBIDDEN", "شما به این سند دسترسی ندارید", http_status=403)
    
    # تبدیل Pydantic model به dict (فقط فیلدهای set شده)
    body_dict = body.dict(exclude_unset=True)
    
    # تبدیل داده‌های flat به nested برای سازگاری با سرویس
    service_data = {}
    
    # کپی فیلدهای مستقیم
    if "document_date" in body_dict:
        service_data["document_date"] = body_dict["document_date"]
    if "currency_id" in body_dict:
        service_data["currency_id"] = body_dict["currency_id"]
    if "description" in body_dict:
        service_data["description"] = body_dict["description"]
    if "fiscal_year_id" in body_dict:
        service_data["fiscal_year_id"] = body_dict["fiscal_year_id"]
    if "extra_info" in body_dict:
        service_data["extra_info"] = body_dict["extra_info"]
    
    # تبدیل source و destination
    # اعتبارسنجی: هر دو type و id باید ارائه شوند
    if "source_type" in body_dict or "source_id" in body_dict:
        source_type = body_dict.get("source_type")
        source_id = body_dict.get("source_id")
        
        if source_type is None:
            raise ApiError(
                "VALIDATION_ERROR",
                "برای تغییر source، هر دو فیلد source_type و source_id باید ارائه شوند",
                http_status=400
            )
        if source_id is None:
            raise ApiError(
                "VALIDATION_ERROR",
                "برای تغییر source، هر دو فیلد source_type و source_id باید ارائه شوند",
                http_status=400
            )
        
        if source_type == "bank_account":
            source_type = "bank"
        service_data["source"] = {
            "type": source_type,
            "id": source_id
        }
    
    if "destination_type" in body_dict or "destination_id" in body_dict:
        destination_type = body_dict.get("destination_type")
        destination_id = body_dict.get("destination_id")
        
        if destination_type is None:
            raise ApiError(
                "VALIDATION_ERROR",
                "برای تغییر destination، هر دو فیلد destination_type و destination_id باید ارائه شوند",
                http_status=400
            )
        if destination_id is None:
            raise ApiError(
                "VALIDATION_ERROR",
                "برای تغییر destination، هر دو فیلد destination_type و destination_id باید ارائه شوند",
                http_status=400
            )
        
        if destination_type == "bank_account":
            destination_type = "bank"
        service_data["destination"] = {
            "type": destination_type,
            "id": destination_id
        }
    
    # تبدیل total_amount به amount
    if "total_amount" in body_dict:
        service_data["amount"] = body_dict["total_amount"]
    
    if "commission" in body_dict:
        service_data["commission"] = body_dict["commission"]
    
    updated = update_transfer(db, document_id, ctx.get_user_id(), service_data)
    return success_response(data=format_datetime_fields(updated, request), request=request, message="TRANSFER_UPDATED")


@router.post(
    "/businesses/{business_id}/transfers/export/excel",
    summary="خروجی Excel لیست اسناد انتقال",
    description="خروجی Excel لیست اسناد انتقال با فیلتر و جستجو",
)
@require_business_access("business_id")
async def export_transfers_excel(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    max_export_records = 10000
    take_value = min(int(body.get("take", 1000)), max_export_records)
    query_dict = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "from_date": body.get("from_date"),
        "to_date": body.get("to_date"),
    }

    result = list_transfers(db, business_id, query_dict)
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]

    wb = Workbook()
    ws = wb.active
    ws.title = "Transfers"

    # Get calendar type for date formatting
    calendar_type = "gregorian"
    if hasattr(request.state, 'calendar_type'):
        calendar_type = request.state.calendar_type
    
    # Helper function to format date based on calendar type
    def format_date_for_excel(item, date_key):
        """Format date based on calendar type (date only, no time)"""
        import datetime as dt_module
        from app.core.calendar import CalendarConverter
        
        # First check if there's a _formatted field (from format_datetime_fields)
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item:
            formatted_value = item.get(formatted_key)
            if isinstance(formatted_value, dict):
                date_only = formatted_value.get("date_only")
                if date_only:
                    return str(date_only)
                formatted = formatted_value.get("formatted", "")
                if formatted:
                    # Extract date part only (remove time)
                    date_part = str(formatted).split(' ')[0].split('T')[0]
                    return date_part
        
        # Get the main field value
        value = item.get(date_key)
        if value is None:
            return ""
        
        # If it's a dict (from _formatted field), use date_only
        if isinstance(value, dict):
            date_only = value.get("date_only")
            if date_only:
                return str(date_only)
            formatted = value.get("formatted", "")
            if formatted:
                date_part = str(formatted).split(' ')[0].split('T')[0]
                return date_part
        
        # If it's a datetime object, format it based on calendar type
        if isinstance(value, datetime.datetime):
            try:
                formatted = CalendarConverter.format_datetime(value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a date object, format it based on calendar type
        if isinstance(value, datetime.date):
            try:
                dt_value = datetime.datetime.combine(value, datetime.datetime.min.time())
                formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
            except Exception:
                pass
        
        # If it's a string, check if it's already formatted (contains / separator for Jalali)
        if isinstance(value, str):
            # Check if it looks like a Jalali date (contains / and has YYYY/MM/DD format)
            if '/' in value and (len(value.split('/')) == 3):
                # Might be already formatted, but check if it's ISO format (YYYY-MM-DD) or Jalali (YYYY/MM/DD)
                if '-' in value:
                    # ISO format (YYYY-MM-DD), parse and format
                    try:
                        if 'T' in value:
                            dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                        else:
                            date_value = datetime.date.fromisoformat(value)
                            dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                        formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                        return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                    except Exception:
                        pass
                else:
                    # Might be Jalali format (YYYY/MM/DD), return as is but remove time if exists
                    if ' ' in value:
                        return value.split(' ')[0]
                    return value
            else:
                # Try to parse as ISO format
                try:
                    if 'T' in value:
                        dt_value = datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                    else:
                        date_value = datetime.date.fromisoformat(value)
                        dt_value = datetime.datetime.combine(date_value, datetime.datetime.min.time())
                    formatted = CalendarConverter.format_datetime(dt_value, calendar_type)
                    return formatted.get("date_only", "") or formatted.get("formatted", "").split(' ')[0]
                except Exception:
                    # If parsing fails, return as is (might already be formatted)
                    if ' ' in value or 'T' in value:
                        date_part = value.split(' ')[0].split('T')[0]
                        return date_part
                    return value
        
        # Fallback
        return str(value) if value else ""

    headers = [
        "کد سند",
        "تاریخ سند",
        "تاریخ ثبت",
        "نوع مبدا",
        "نام مبدا",
        "نوع مقصد",
        "نام مقصد",
        "مبلغ کل",
        "کارمزد",
        "توضیحات",
        "ایجادکننده"
    ]
    keys = [
        "code",
        "document_date",
        "registered_at",
        "source_type_name",
        "source_name",
        "destination_type_name",
        "destination_name",
        "total_amount",
        "commission",
        "description",
        "created_by_name"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            if key in ["document_date", "registered_at"]:
                # Format dates based on calendar type
                val = format_date_for_excel(item, key)
            else:
                val = item.get(key, "")
                # Handle None values
                if val is None:
                    val = ""
            ws.cell(row=row_idx, column=col_idx, value=val).border = border

    # Auto width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    biz_name = ""
    try:
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            biz_name = b.name or ""
    except Exception:
        biz_name = ""

    def slugify(text: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")

    base = "transfers"
    if biz_name:
        base += f"_{slugify(biz_name)}"
    filename = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content = buffer.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.post(
    "/businesses/{business_id}/transfers/export/pdf",
    summary="خروجی PDF لیست اسناد انتقال",
    description="خروجی PDF لیست اسناد انتقال با فیلتر و جستجو",
)
@require_business_access("business_id")
async def export_transfers_pdf(
    business_id: int,
    request: Request,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
    from html import escape
    from app.core.i18n import negotiate_locale
    from app.core.calendar import CalendarConverter, get_calendar_type_from_header
    from pathlib import Path

    max_export_records = 10000
    take_value = min(int(body.get("take", 1000)), max_export_records)
    query_dict = {
        "take": take_value,
        "skip": int(body.get("skip", 0)),
        "sort_by": body.get("sort_by"),
        "sort_desc": bool(body.get("sort_desc", False)),
        "sort": body.get("sort") if isinstance(body.get("sort"), list) else None,
        "search": body.get("search"),
        "from_date": body.get("from_date"),
        "to_date": body.get("to_date"),
    }
    result = list_transfers(db, business_id, query_dict)
    items = result.get('items', [])
    items = [format_datetime_fields(item, request) for item in items]

    # Locale and calendar
    locale = negotiate_locale(request.headers.get("Accept-Language"))
    is_fa = locale == 'fa'
    
    # Get calendar type for date formatting
    calendar_type = get_calendar_type_from_header(request.headers.get("X-Calendar-Type"))
    if not calendar_type:
        calendar_type = "jalali" if is_fa else "gregorian"
    
    # Format generated date based on calendar
    try:
        _now = datetime.datetime.now()
        _fd = CalendarConverter.format_datetime(_now, calendar_type)
        generated_at = _fd.get("formatted") or _fd.get("date_only") or _now.strftime('%Y/%m/%d %H:%M')
    except Exception:
        generated_at = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    # Helper function to format date based on calendar type
    def format_date_for_pdf(item, date_key):
        # If already formatted by format_datetime_fields, use it
        formatted_key = f"{date_key}_formatted"
        if formatted_key in item:
            formatted = item[formatted_key]
            if isinstance(formatted, dict):
                return formatted.get("date_only", "")
            return str(formatted) if formatted else ""
        # Fallback to original value
        date_value = item.get(date_key, "")
        if isinstance(date_value, datetime.datetime):
            # Format based on calendar type
            if calendar_type == "jalali":
                jalali = CalendarConverter.to_jalali(date_value)
                return jalali.get("date_only", "") if jalali else ""
            else:
                return date_value.strftime("%Y-%m-%d")
        return str(date_value) if date_value else ""

    # Helper for numeric formatting with thousands separator and trimming .00
    def format_number_for_display(value):
        try:
            if value is None:
                return ""
            v = float(value)
            s = f"{v:,.2f}"
            # Trim trailing .00 or trailing zeros
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return s
        except Exception:
            return str(value) if value is not None else ""

    # Add row number column
    headers = [
        "ردیف",
        "کد سند",
        "تاریخ سند",
        "تاریخ ثبت",
        "نوع مبدا",
        "نام مبدا",
        "نوع مقصد",
        "نام مقصد",
        "مبلغ کل",
        "کارمزد",
        "توضیحات",
        "ایجادکننده"
    ]
    keys = [
        "row_number",  # Will be generated
        "code",
        "document_date",
        "registered_at",
        "source_type_name",
        "source_name",
        "destination_type_name",
        "destination_name",
        "total_amount",
        "commission",
        "description",
        "created_by_name"
    ]

    header_html = ''.join(f'<th>{escape(h)}</th>' for h in headers)
    rows_html = []
    amount_keys = {"total_amount", "commission"}
    date_keys = {"document_date", "registered_at"}
    
    for idx, it in enumerate(items, 1):
        row_cells = []
        for k in keys:
            if k == "row_number":
                row_cells.append(f'<td style="text-align:center">{idx}</td>')
            elif k in date_keys:
                # Format dates based on calendar type
                v = format_date_for_pdf(it, k)
                row_cells.append(f'<td>{escape(str(v))}</td>')
            elif k in amount_keys:
                # Format amounts
                v = it.get(k, 0)
                disp = format_number_for_display(v)
                row_cells.append(f'<td class="amount">{escape(disp)}</td>')
            else:
                v = it.get(k, "")
                # Handle None values
                if v is None:
                    v = ""
                row_cells.append(f'<td>{escape(str(v))}</td>')
        rows_html.append(f'<tr>{"".join(row_cells)}</tr>')

    # Business name
    business_name = ""
    try:
        from adapters.db.models.business import Business
        b = db.query(Business).filter(Business.id == business_id).first()
        if b is not None:
            business_name = b.name or ""
    except Exception:
        business_name = ""

    title_text = "لیست انتقال‌ها" if is_fa else "Transfers List"
    footer_text = f"تولید شده در {generated_at}" if is_fa else f"Generated at {generated_at}"

    # Template context
    template_context: Dict[str, Any] = {
        "title_text": title_text,
        "business_name": business_name,
        "generated_at": generated_at,
        "is_fa": is_fa,
        "fa_font_url_regular": None,
        "fa_font_url_bold": None,
        "headers": headers,
        "keys": keys,
        "items": items,
        "table_headers_html": header_html,
        "table_rows_html": "".join(rows_html),
        "paper_size": "A4",
        "orientation": "landscape",  # افقی
        "footer_text": footer_text,
    }

    # Embed Farsi fonts (YekanBakhFaNum)
    try:
        if is_fa:
            project_root = Path(__file__).resolve().parents[4]
            fonts_dir = project_root / "hesabixUI" / "hesabix_ui" / "assets" / "fonts"
            regular_path = fonts_dir / "YekanBakhFaNum-Regular.ttf"
            bold_path = fonts_dir / "YekanBakhFaNum-Bold.ttf"
            if regular_path.is_file():
                import base64 as _b64
                _data = regular_path.read_bytes()
                _b64_data = _b64.b64encode(_data).decode("ascii")
                template_context["fa_font_url_regular"] = f"data:font/ttf;base64,{_b64_data}"
            if bold_path.is_file():
                import base64 as _b64b
                _data_b = bold_path.read_bytes()
                _b64_data_b = _b64b.b64encode(_data_b).decode("ascii")
                template_context["fa_font_url_bold"] = f"data:font/ttf;base64,{_b64_data_b}"
    except Exception:
        pass

    # Try to render with custom template
    resolved_html = None
    try:
        from app.services.report_template_service import ReportTemplateService
        explicit_template_id = None
        try:
            if body.get("template_id") is not None:
                explicit_template_id = int(body.get("template_id"))
        except Exception:
            explicit_template_id = None
        resolved_html = ReportTemplateService.try_render_resolved(
            db=db,
            business_id=business_id,
            module_key="transfers",
            subtype="list",
            context=template_context,
            explicit_template_id=explicit_template_id,
        )
    except Exception:
        resolved_html = None

    # Default HTML template using base.html
    if not resolved_html:
        try:
            from app.services.pdf.template_renderer import render_template
            resolved_html = render_template(
                "pdf/transfers/list.html",
                **template_context
            )
        except Exception:
            # Fallback simple HTML
            resolved_html = f"""
            <!DOCTYPE html>
            <html dir='rtl' lang='fa'>
              <head>
                <meta charset='utf-8'>
                <title>{title_text}</title>
                <style>
                  @page {{ margin: 1cm; size: A4 landscape; }}
                  body {{ font-family: 'YekanBakhFaNum', Tahoma, Arial; font-size: 12px; color: #333; }}
                  .header {{ display: flex; justify-content: space-between; margin-bottom: 16px; border-bottom: 2px solid #366092; padding-bottom: 8px; }}
                  .title {{ font-weight: bold; color: #366092; font-size: 18px; }}
                  table {{ width: 100%; border-collapse: collapse; }}
                  th, td {{ border: 1px solid #ddd; padding: 6px; text-align: right; }}
                  thead th {{ background-color: #f0f0f0; }}
                  .amount {{ text-align: left; font-weight: bold; }}
                </style>
              </head>
              <body>
                <div class='header'>
                  <div class='title'>{title_text}</div>
                  <div>تاریخ تولید: {escape(generated_at)}</div>
                </div>
                <table>
                  <thead><tr>{header_html}</tr></thead>
                  <tbody>
                    {"".join(rows_html)}
                  </tbody>
                </table>
              </body>
            </html>
            """

    font_config = FontConfiguration()
    pdf_bytes = HTML(string=resolved_html).write_pdf(font_config=font_config)
    filename = f"transfers_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(pdf_bytes)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


