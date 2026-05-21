"""
API endpoints برای گزارش‌های مالیاتی
"""

from __future__ import annotations

from fastapi import APIRouter, Request, Depends, Body
from typing import Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, and_

from app.core.responses import success_response, ApiError
from app.core.auth_dependency import get_current_user, AuthContext
from app.core.permissions import require_business_access, require_business_permission_dep
from app.core.moadian_plugin_dependency import ensure_moadian_plugin_active
from adapters.db.session import get_db
from adapters.db.models.document import Document
from adapters.api.v1.invoices import SUPPORTED_INVOICE_TYPES

router = APIRouter(prefix="/tax-reports", tags=["tax-reports"])


@router.post("/business/{business_id}/export")
@require_business_access("business_id")
def export_tax_report(
    request: Request,
    business_id: int,
    body: Dict[str, Any] = Body(...),
    ctx: AuthContext = Depends(get_current_user),
    db: Session = Depends(get_db),
    _: None = Depends(require_business_permission_dep("moadian", "export_reports")),
) -> Dict[str, Any]:
    """
    Export گزارش ارسال‌های مالیاتی به Excel
    
    Body:
        - from_date: تاریخ شروع (اختیاری)
        - to_date: تاریخ پایان (اختیاری)
        - status: وضعیت (اختیاری)
        - format: excel یا pdf (پیش‌فرض: excel)
    """
    ensure_moadian_plugin_active(db, business_id)
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    from_date = body.get("from_date")
    to_date = body.get("to_date")
    status = body.get("status")
    format_type = body.get("format", "excel")
    
    from sqlalchemy import cast, Boolean
    from sqlalchemy.dialects.postgresql import JSONB

    _extra_info_jb = cast(Document.extra_info, JSONB)
    query = db.query(Document).filter(
        and_(
            Document.business_id == business_id,
            Document.document_type.in_(list(SUPPORTED_INVOICE_TYPES)),
            cast(_extra_info_jb['tax_workspace'], Boolean) == True,
        )
    )
    
    # فیلتر تاریخ
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date.replace('Z', '+00:00'))
            query = query.filter(Document.date >= from_dt.date())
        except Exception:
            pass
    
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date.replace('Z', '+00:00'))
            query = query.filter(Document.date <= to_dt.date())
        except Exception:
            pass
    
    # فیلتر وضعیت
    if status:
        query = query.filter(
            _extra_info_jb['tax_status'].astext == status
        )
    
    docs = query.order_by(Document.date.desc()).limit(10000).all()
    
    if format_type == "excel":
        # ایجاد Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "گزارش مالیاتی"
        
        # Header
        headers = [
            "کد فاکتور",
            "تاریخ",
            "وضعیت",
            "کد رهگیری",
            "تاریخ ارسال",
            "تاریخ آخرین استعلام",
            "پیام خطا",
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # داده‌ها
        for row, doc in enumerate(docs, 2):
            extra = dict(doc.extra_info or {})
            ws.cell(row=row, column=1, value=doc.code or "")
            ws.cell(row=row, column=2, value=doc.date.isoformat() if doc.date else "")
            ws.cell(row=row, column=3, value=extra.get("tax_status", ""))
            ws.cell(row=row, column=4, value=extra.get("tax_tracking_code", ""))
            ws.cell(row=row, column=5, value=extra.get("tax_last_send_at", ""))
            ws.cell(row=row, column=6, value=extra.get("tax_last_inquiry_at", ""))
            ws.cell(row=row, column=7, value=extra.get("tax_error_message", ""))
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        
        # ذخیره در memory
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from fastapi.responses import Response
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="tax_report_{business_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            }
        )
    
    else:
        raise ApiError("UNSUPPORTED_FORMAT", f"Format {format_type} is not supported", http_status=400)

